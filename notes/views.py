from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Avg, Count, Sum
from django.core.paginator import Paginator
from django.core.cache import cache
from django.utils import timezone
from django.template.loader import render_to_string
from django.db import IntegrityError
from django.views.decorators.http import require_POST
from decimal import Decimal, InvalidOperation
import json
import os
import io
from datetime import datetime
from eleves.models import Classe as ClasseEleve, Eleve
from utilisateurs.utils import filter_by_user_school, user_school
from ecole_moderne.security_decorators import admin_required, require_school_object
from utilisateurs.permissions import any_permission_required, can_manage_notes
from .forms import ClasseNoteForm, MatiereNoteForm, EvaluationForm, NoteEleveForm
from .models import ClasseNote, MatiereNote, Evaluation, NoteEleve, NoteMensuelle, CompositionNote
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.units import cm
from .calculs_intelligent import (
    obtenir_mention_intelligente,
    obtenir_appreciation_intelligente,
    formater_rang_intelligent,
)

# Import centralisé du filigrane - chargé une seule fois
try:
    from ecole_moderne.pdf_utils import draw_logo_watermark as _draw_watermark_base
    WATERMARK_AVAILABLE = True
except ImportError:
    WATERMARK_AVAILABLE = False
    _draw_watermark_base = None

def _apply_watermark(c, width, height, ecole=None):
    """Applique le filigrane si disponible - fonction utilitaire centralisée"""
    if WATERMARK_AVAILABLE and _draw_watermark_base:
        try:
            _draw_watermark_base(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=ecole)
        except Exception:
            pass

# Groupes de niveaux pour l'affichage
PRIMAIRE = {
    'PRIMAIRE_1', 'PRIMAIRE_2', 'PRIMAIRE_3', 'PRIMAIRE_4', 'PRIMAIRE_5', 'PRIMAIRE_6'
}
COLLEGE = {
    'COLLEGE_7', 'COLLEGE_8', 'COLLEGE_9', 'COLLEGE_10'
}
LYCEE = {
    'LYCEE_11', 'LYCEE_12', 'TERMINALE'
}

def _draw_school_header(c, ecole, *, y_start, margin, page_width):
    """Dessine un en-tête officiel (centré) avec logo, nom en MAJUSCULES, coordonnées et encadré.
    Retourne la nouvelle coordonnée y après dessin."""
    from reportlab.lib import colors
    y = y_start
    # En-tête national
    center_x = page_width / 2
    c.setFont('Helvetica-Bold', 18)
    c.drawCentredString(center_x, y, "République de Guinée")
    y -= 12
    c.setFont('Helvetica-Oblique', 10)
    # Dessiner la devise avec couleurs par mot: Travail (rouge), Justice (jaune), Solidarité (vert)
    from reportlab.pdfbase import pdfmetrics
    from reportlab.lib import colors
    parts = [
        ("Travail", colors.red),
        (" - ", colors.black),
        ("Justice", colors.yellow),
        (" - ", colors.black),
        ("Solidarité", colors.green),
    ]
    total_w = sum(pdfmetrics.stringWidth(t, 'Helvetica-Oblique', 10) for t, _ in parts)
    start_x = center_x - (total_w / 2)
    x = start_x
    for text, col in parts:
        c.setFillColor(col)
        c.drawString(x, y, text)
        x += pdfmetrics.stringWidth(text, 'Helvetica-Oblique', 10)
    c.setFillColor(colors.black)
    y -= 12
    c.setFont('Helvetica', 10)
    c.drawCentredString(center_x, y, "Ministère de l’Enseignement Pré-Universitaire et de l’Alphabétisation")
    y -= 12
    # Abréviations sur 3 lignes (centrées)
    c.setFont('Helvetica-Bold', 10)
    y -= 6
    ire = getattr(ecole, 'ire', None) or ''
    dpe = getattr(ecole, 'dpe', None) or ''
    desee = getattr(ecole, 'desee', None) or ''
    c.drawCentredString(center_x, y, f"IRE: {ire}")
    y -= 12
    c.drawCentredString(center_x, y, f"DPE: {dpe}")
    y -= 12
    c.drawCentredString(center_x, y, f"DESEE: {desee}")
    y -= 16
    # Espace supplémentaire pour descendre le premier cadre du bulletin
    y -= 30

    # Mémoriser la position du haut du cadre pour le dessiner après le contenu
    frame_top = y
    box_height = 60

    # Logo (gauche) si disponible
    logo_path = None
    try:
        if hasattr(ecole, 'logo') and getattr(ecole.logo, 'path', None) and os.path.exists(ecole.logo.path):
            logo_path = ecole.logo.path
    except Exception:
        logo_path = None
    if logo_path:
        try:
            c.drawImage(logo_path, margin + 8, y - 62, width=54, height=54, preserveAspectRatio=True, mask='auto')
        except Exception:
            pass

    # Texte centré
    top_line_y = y + 12
    school_name = (getattr(ecole, 'nom', '') or 'ÉCOLE').upper()
    c.setFont('Helvetica-Bold', 16)
    c.drawCentredString(center_x, top_line_y, school_name)

    c.setFont('Helvetica', 10)
    adresse = getattr(ecole, 'adresse', None) or ''
    telephone = getattr(ecole, 'telephone', None) or ''
    email = getattr(ecole, 'email', None) or ''
    directeur = getattr(ecole, 'directeur', None) or ''
    censeur = getattr(ecole, 'censeur', None) or ''

    # Helper: wrap centered text within available width
    from reportlab.pdfbase import pdfmetrics
    def draw_wrapped_centered(text, y_pos, max_width, line_height=12):
        words = text.split()
        lines = []
        cur = ''
        for w in words:
            test = (cur + ' ' + w).strip()
            if pdfmetrics.stringWidth(test, 'Helvetica', 10) <= max_width:
                cur = test
            else:
                if cur:
                    lines.append(cur)
                cur = w
        if cur:
            lines.append(cur)
        for ln in lines:
            c.drawCentredString(center_x, y_pos, ln)
            y_pos -= line_height
        return y_pos

    # Centrer les informations (adresse/contacts/directeur) au milieu du cadre
    line_y = top_line_y - 30
    # Contacts en gris léger pour hiérarchie visuelle
    try:
        c.setFillGray(0.3)
    except Exception:
        pass
    if adresse:
        # keep inside box: reduce available width a bit
        avail_w = page_width - 2*margin - 20
        line_y = draw_wrapped_centered(f"Adresse: {adresse}", line_y, avail_w)
    # Afficher téléphone et email sur des lignes séparées pour éviter le débordement
    if telephone:
        avail_w = page_width - 2*margin - 20
        line_y = draw_wrapped_centered(f"Tél: {telephone}", line_y, avail_w)
    if email:
        avail_w = page_width - 2*margin - 20
        line_y = draw_wrapped_centered(f"Email: {email}", line_y, avail_w)
    if directeur:
        c.drawCentredString(center_x, line_y, f"Directeur: {directeur}")
        line_y -= 12
    if censeur:
        c.drawCentredString(center_x, line_y, f"Censeur: {censeur}")
    # Rétablir la couleur par défaut
    try:
        c.setFillGray(0.0)
    except Exception:
        pass

    # Dessiner le cadre maintenant que le contenu est placé
    # Le cadre commence au-dessus du nom de l'école pour avoir plus d'espace en haut
    frame_start_y = top_line_y - 8  # En dessous du nom de l'école
    adjusted_box_height = box_height  # Utiliser la hauteur fixe définie
    c.setLineWidth(1)
    c.setStrokeColor(colors.black)
    c.roundRect(margin, frame_start_y - adjusted_box_height, page_width - 2*margin, adjusted_box_height, 6, stroke=1, fill=0)

    # Retourner y en dessous du cadre
    y = y - box_height - 8
    # Ligne séparatrice légère
    c.setFillColor(colors.grey)
    c.rect(margin, y, page_width-2*margin, 1, fill=1, stroke=0)
    c.setFillColor(colors.black)
    y -= 8
    return y

@login_required
def tableau_bord(request):
    """Tableau de bord des notes: liste les classes par groupe de niveaux.
    Filtré par l'école de l'utilisateur (sauf admin) et par année active.
    """
    from eleves.utils_annee import get_annee_active
    classes_qs = filter_by_user_school(
        ClasseEleve.objects.select_related('ecole').order_by('niveau', 'nom'),
        request.user, 'ecole'
    )
    # Filtrer par année scolaire active
    ecole_user = user_school(request.user)
    if ecole_user:
        annee_active = get_annee_active(request, ecole_user)
        if annee_active:
            classes_qs = classes_qs.filter(annee_scolaire=annee_active)

    def group_classes(qs):
        primaire, college, lycee = [], [], []
        for c in qs:
            if c.niveau in PRIMAIRE:
                primaire.append(c)
            elif c.niveau in COLLEGE:
                college.append(c)
            elif c.niveau in LYCEE:
                lycee.append(c)
        return primaire, college, lycee

    primaire, college, lycee = group_classes(classes_qs)

    context = {
        'classes_primaire': primaire,
        'classes_college': college,
        'classes_lycee': lycee,
    }
    return render(request, 'notes/tableau_bord.html', context)

@admin_required
def creer_classe(request, niveau):
    """Créer une classe pour un niveau donné dans l'école de l'utilisateur."""
    if request.method == 'POST':
        form = ClasseNotesForm(request.POST, niveau_initial=niveau)
        if form.is_valid():
            classe = form.save(commit=False)
            classe.ecole = user_school(request.user)
            if classe.ecole is None:
                messages.error(request, "Aucune école associée à votre compte.")
                return redirect('notes:tableau_bord')
            classe.save()
            messages.success(request, f"Classe '{classe.nom}' créée avec succès.")
            return redirect('notes:tableau_bord')
    else:
        form = ClasseNotesForm(niveau_initial=niveau)
    return render(request, 'notes/classe_form.html', {'form': form, 'niveau': niveau})

@admin_required
def supprimer_classe(request, classe_id):
    """Supprimer une classe si elle appartient à l'école de l'utilisateur et qu'elle est vide."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.all(), request.user, 'ecole'), pk=classe_id)
    if request.method == 'POST':
        if hasattr(classe, 'eleves') and classe.eleves.exists():
            messages.error(request, "Impossible de supprimer une classe qui contient des élèves.")
            return redirect('notes:tableau_bord')
        classe.delete()
        messages.success(request, "Classe supprimée avec succès.")
        return redirect('notes:tableau_bord')
    return render(request, 'notes/confirm_delete.html', {
        'objet': classe,
        'message': "Confirmez-vous la suppression de cette classe ? (Cette action est irréversible)",
        'action_url': reverse('notes:supprimer_classe', args=[classe.id])
    })

@can_manage_notes
def matieres_classe(request, classe_id):
    """Liste optimisée et gestion des matières d'une classe avec cache."""
    # Cache de la classe
    classe_cache_key = f'classe_{classe_id}_{request.user.id}'
    classe = cache.get(classe_cache_key)
    
    if classe is None:
        classe = get_object_or_404(
            filter_by_user_school(ClasseEleve.objects.select_related('ecole'), request.user, 'ecole'), 
            pk=classe_id
        )
        cache.set(classe_cache_key, classe, 300)  # 5 minutes
    
    # Cache des matières
    matieres_cache_key = f'matieres_classe_{classe_id}'
    matieres = cache.get(matieres_cache_key)
    
    if matieres is None:
        matieres = list(
            MatiereClasse.objects
            .filter(classe=classe, ecole=classe.ecole)
            .select_related('classe', 'ecole')
            .prefetch_related('evaluations')
            .order_by('nom')
        )
        cache.set(matieres_cache_key, matieres, 180)  # 3 minutes
    
    return render(request, 'notes/matieres_classe.html', {
        'classe': classe,
        'matieres': matieres,
    })

@admin_required
def creer_matiere(request, classe_id):
    """Créer une matière pour une classe donnée."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.all(), request.user, 'ecole'), pk=classe_id)
    if request.method == 'POST':
        form = MatiereClasseForm(request.POST)
        if form.is_valid():
            mat = form.save(commit=False)
            mat.classe = classe
            mat.ecole = classe.ecole
            try:
                mat.save()
                messages.success(request, f"Matière '{mat.nom}' ajoutée.")
                return redirect('notes:matieres_classe', classe.id)
            except Exception as e:
                messages.error(request, f"Erreur lors de la création: {e}")
    else:
        form = MatiereClasseForm()
    return render(request, 'notes/matiere_form.html', {'form': form, 'classe': classe})

@admin_required
def supprimer_matiere(request, pk):
    """Supprimer une matière de classe."""
    matiere = get_object_or_404(filter_by_user_school(MatiereClasse.objects.select_related('classe', 'ecole'), request.user, 'ecole'), pk=pk)
    if request.method == 'POST':
        classe_id = matiere.classe_id
        matiere.delete()
        messages.success(request, "Matière supprimée.")
        return redirect('notes:matieres_classe', classe_id)
    return render(request, 'notes/confirm_delete.html', {
        'objet': matiere,
        'message': "Confirmez-vous la suppression de cette matière ?",
        'action_url': reverse('notes:supprimer_matiere', args=[matiere.id])
    })

@admin_required
def creer_evaluation(request, classe_id, matiere_id):
    """Créer une évaluation pour une classe/matière donnée."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.all(), request.user, 'ecole'), pk=classe_id)
    matiere = get_object_or_404(MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole), pk=matiere_id)
    if request.method == 'POST':
        form = EvaluationForm(request.POST)
        if form.is_valid():
            ev = form.save(commit=False)
            ev.ecole = classe.ecole
            ev.classe = classe
            ev.matiere = matiere
            ev.annee_scolaire = getattr(classe, 'annee_scolaire', None)
            ev.cree_par = request.user
            ev.save()
            messages.success(request, f"Évaluation '{ev.titre}' créée pour {classe.nom} — {matiere.nom}.")
            return redirect('notes:saisie_notes', evaluation_id=ev.id)
    else:
        form = EvaluationForm()
    return render(request, 'notes/evaluation_form.html', {
        'form': form,
        'classe': classe,
        'matiere': matiere,
    })

@can_manage_notes
def saisie_notes(request, evaluation_id):
    """Saisie en masse des notes par matricule pour une évaluation.
    Format par ligne: MATRICULE;NOTE
    """
    # Récupération évaluation dans le périmètre école
    evaluation = get_object_or_404(
        filter_by_user_school(Evaluation.objects.select_related('classe', 'matiere', 'ecole'), request.user, 'ecole'),
        pk=evaluation_id
    )
    if request.method == 'POST':
        form = NotesBulkForm(request.POST)
        if form.is_valid():
            donnees = form.cleaned_data['donnees']
            lignes = [l.strip() for l in donnees.splitlines() if l.strip()]
            ok, erreurs, maj, crees = 0, [], 0, 0
            # Restreindre aux élèves de la classe + école de l'évaluation
            eleves_qs = Eleve.objects.select_related('classe', 'classe__ecole').filter(classe=evaluation.classe)
            eleves_qs = filter_by_user_school(eleves_qs, request.user, 'classe__ecole')
            # Index par matricule (upper)
            index_mat = { (e.matricule or '').strip().upper(): e for e in eleves_qs }
            for i, ligne in enumerate(lignes, start=1):
                parts = [p.strip() for p in ligne.split(';')]
                if len(parts) < 2:
                    erreurs.append(f"Ligne {i}: format invalide (attendu MATRICULE;NOTE)")
                    continue
                matricule, note_txt = parts[0].upper(), parts[1].replace(',', '.').strip()
                obs = ''  # Pas d'observation requise
                if matricule not in index_mat:
                    erreurs.append(f"Ligne {i}: matricule inconnu pour la classe ({matricule})")
                    continue
                # Parse note 0..20
                try:
                    val = Decimal(note_txt)
                except Exception:
                    erreurs.append(f"Ligne {i}: note invalide '{note_txt}'")
                    continue
                if val < 0 or val > 20:
                    erreurs.append(f"Ligne {i}: la note doit être entre 0 et 20 (reçu {val})")
                    continue
                eleve = index_mat[matricule]
                # Créer/mettre à jour la note
                obj, created = Note.objects.update_or_create(
                    evaluation=evaluation,
                    eleve=eleve,
                    defaults={
                        'ecole': evaluation.ecole,
                        'classe': evaluation.classe,
                        'matiere': evaluation.matiere,
                        'matricule': eleve.matricule or matricule,
                        'note': val,
                        'observation': obs or None,
                        'saisie_par': request.user,
                    }
                )
                if created:
                    crees += 1
                else:
                    maj += 1
                ok += 1
            if ok:
                messages.success(request, f"{ok} note(s) traitée(s) — {crees} créée(s), {maj} mise(s) à jour.")
            if erreurs:
                messages.warning(request, "\n".join(erreurs[:10]) + ("\n…" if len(erreurs) > 10 else ''))
            return redirect('notes:saisie_notes', evaluation_id=evaluation.id)
    else:
        form = NotesBulkForm()
    # Préparer un export des élèves de la classe avec matricules pour aide à la saisie
    eleves = Eleve.objects.select_related('classe').filter(classe=evaluation.classe).order_by('prenom', 'nom')
    eleves = filter_by_user_school(eleves, request.user, 'classe__ecole')
    # Notes existantes pour cette évaluation (map JSON par élève)
    notes_qs = evaluation.notes.select_related('eleve')
    try:
        import json as _json
        notes_existantes_map = {
            str(n.eleve_id): {
                'note': float(n.note) if getattr(n, 'note', None) is not None else None,
                'absent': bool(getattr(n, 'absent', False)),
            }
            for n in notes_qs
        }
        notes_existantes_json = _json.dumps(notes_existantes_map)
    except Exception:
        notes_existantes_json = '{}'

    return render(request, 'notes/saisir_notes.html', {
        'evaluation': evaluation,
        'form': form,
        'eleves': eleves,
        'notes_existantes_json': notes_existantes_json,
    })

@can_manage_notes
def saisie_notes_individuelle(request, evaluation_id: int):
    """Affiche la saisie individuelle des notes pour une évaluation donnée."""
    evaluation = get_object_or_404(
        filter_by_user_school(
            Evaluation.objects.select_related('classe', 'matiere', 'ecole'), request.user, 'ecole'
        ),
        pk=evaluation_id,
    )

    # Élèves de la classe (dans le périmètre école)
    eleves_qs = Eleve.objects.select_related('classe').filter(classe=evaluation.classe)
    eleves_qs = filter_by_user_school(eleves_qs, request.user, 'classe__ecole')

    # Notes existantes pour cette évaluation
    notes_map = {
        n.eleve_id: n for n in Note.objects.filter(evaluation=evaluation).select_related('eleve')
    }

    eleves_context = []
    notes_saisies = 0
    for e in eleves_qs.order_by('prenom', 'nom'):
        n = notes_map.get(e.id)
        note_val = n.note if n else None
        if n:
            notes_saisies += 1
        eleves_context.append({
            'id': e.id,
            'nom': e.nom,
            'prenom': e.prenom,
            'matricule': e.matricule,
            'date_naissance': e.date_naissance,
            'photo': getattr(e, 'photo', None),
            'note_actuelle': note_val,
            'appreciation_actuelle': n.appreciation_finale if n else None,
        })

    return render(request, 'notes/saisie_notes_individuelle.html', {
        'evaluation': evaluation,
        'eleves': eleves_context,
        'notes_saisies': notes_saisies,
    })

@can_manage_notes
def saisie_notes_simple(request, evaluation_id):
    """Interface simplifiée et intuitive pour la saisie des notes avec tableau interactif."""
    # Récupération évaluation dans le périmètre école
    evaluation = get_object_or_404(
        filter_by_user_school(Evaluation.objects.select_related('classe', 'matiere', 'ecole'), request.user, 'ecole'),
        pk=evaluation_id
    )
    
    if request.method == 'POST':
        # Récupérer toutes les notes soumises
        ok, erreurs, maj, crees = 0, [], 0, 0
        
        # Restreindre aux élèves de la classe + école de l'évaluation
        eleves_qs = Eleve.objects.select_related('classe', 'classe__ecole').filter(classe=evaluation.classe)
        eleves_qs = filter_by_user_school(eleves_qs, request.user, 'classe__ecole')
        
        for eleve in eleves_qs:
            note_key = f'note_{eleve.id}'
            note_value = request.POST.get(note_key, '').strip()
            
            # Si pas de note saisie, passer au suivant
            if not note_value:
                continue
            
            try:
                val = Decimal(note_value.replace(',', '.'))
            except Exception:
                erreurs.append(f"{eleve.nom} {eleve.prenom}: note invalide '{note_value}'")
                continue
            
            if val < 0 or val > 20:
                erreurs.append(f"{eleve.nom} {eleve.prenom}: la note doit être entre 0 et 20 (reçu {val})")
                continue
            
            # Créer/mettre à jour la note
            obj, created = Note.objects.update_or_create(
                evaluation=evaluation,
                eleve=eleve,
                defaults={
                    'ecole': evaluation.ecole,
                    'classe': evaluation.classe,
                    'matiere': evaluation.matiere,
                    'matricule': eleve.matricule or '',
                    'note': val,
                    'observation': None,
                    'saisie_par': request.user,
                }
            )
            if created:
                crees += 1
            else:
                maj += 1
            ok += 1
        
        if ok:
            messages.success(request, f"✅ {ok} note(s) enregistrée(s) avec succès — {crees} créée(s), {maj} mise(s) à jour.")
        if erreurs:
            messages.warning(request, "⚠️ Erreurs:\n" + "\n".join(erreurs[:10]) + ("\n…" if len(erreurs) > 10 else ''))
        
        return redirect('notes:saisie_notes_simple', evaluation_id=evaluation.id)
    
    # GET: Afficher le formulaire
    eleves = Eleve.objects.select_related('classe').filter(classe=evaluation.classe).order_by('prenom', 'nom')
    eleves = filter_by_user_school(eleves, request.user, 'classe__ecole')
    
    # Notes existantes pour cette évaluation
    notes_existantes = evaluation.notes.select_related('eleve').order_by('eleve__nom', 'eleve__prenom')
    
    return render(request, 'notes/saisie_notes_simple.html', {
        'evaluation': evaluation,
        'eleves': eleves,
        'notes_existantes': notes_existantes,
    })

@require_POST
@can_manage_notes
def ajax_sauvegarder_note(request):
    """Enregistre ou met à jour une note individuelle (JSON)."""
    try:
        data = json.loads(request.body.decode('utf-8'))
        evaluation_id = int(data.get('evaluation_id'))
        eleve_id = int(data.get('eleve_id'))
        note_val = Decimal(str(data.get('note')))
    except Exception:
        return JsonResponse({'success': False, 'error': "Payload invalide"}, status=400)

    if note_val < 0 or note_val > 20:
        return JsonResponse({'success': False, 'error': "La note doit être entre 0 et 20"}, status=400)

    evaluation = get_object_or_404(
        filter_by_user_school(
            Evaluation.objects.select_related('classe', 'matiere', 'ecole'), request.user, 'ecole'
        ),
        pk=evaluation_id,
    )
    eleve = get_object_or_404(
        filter_by_user_school(Eleve.objects.select_related('classe'), request.user, 'classe__ecole'),
        pk=eleve_id,
        classe=evaluation.classe,
    )

    obj, created = Note.objects.update_or_create(
        evaluation=evaluation,
        eleve=eleve,
        defaults={
            'ecole': evaluation.ecole,
            'classe': evaluation.classe,
            'matiere': evaluation.matiere,
            'matricule': eleve.matricule or '',
            'note': note_val,
            'saisie_par': request.user,
        }
    )

    total_notes = Note.objects.filter(evaluation=evaluation).count()
    return JsonResponse({'success': True, 'created': created, 'total_notes': total_notes})

@require_POST
@can_manage_notes
def ajax_sauvegarder_notes_masse(request):
    """Enregistre plusieurs notes d'un coup (JSON)."""
    try:
        data = json.loads(request.body.decode('utf-8'))
        evaluation_id = int(data.get('evaluation_id'))
        notes = data.get('notes') or []
    except Exception:
        return JsonResponse({'success': False, 'error': "Payload invalide"}, status=400)

    evaluation = get_object_or_404(
        filter_by_user_school(
            Evaluation.objects.select_related('classe', 'matiere', 'ecole'), request.user, 'ecole'
        ),
        pk=evaluation_id,
    )

    # Préparer les élèves de la classe
    eleves_qs = Eleve.objects.filter(classe=evaluation.classe)
    eleves_qs = filter_by_user_school(eleves_qs, request.user, 'classe__ecole')
    index_eleves = {e.id: e for e in eleves_qs}

    saved = 0
    for item in notes:
        try:
            eleve_id = int(item.get('eleve_id'))
            note_val = Decimal(str(item.get('note')))
        except Exception:
            continue
        if note_val < 0 or note_val > 20:
            continue
        eleve = index_eleves.get(eleve_id)
        if not eleve:
            continue
        Note.objects.update_or_create(
            evaluation=evaluation,
            eleve=eleve,
            defaults={
                'ecole': evaluation.ecole,
                'classe': evaluation.classe,
                'matiere': evaluation.matiere,
                'matricule': eleve.matricule or '',
                'note': note_val,
                'saisie_par': request.user,
            }
        )
        saved += 1

    total_notes = Note.objects.filter(evaluation=evaluation).count()
    return JsonResponse({'success': True, 'saved_count': saved, 'total_notes': total_notes})

@require_POST
@can_manage_notes
def ajax_supprimer_note(request):
    """Supprime la note d'un élève pour une évaluation (JSON)."""
    try:
        data = json.loads(request.body.decode('utf-8'))
        evaluation_id = int(data.get('evaluation_id'))
        eleve_id = int(data.get('eleve_id'))
    except Exception:
        return JsonResponse({'success': False, 'error': "Payload invalide"}, status=400)

    evaluation = get_object_or_404(
        filter_by_user_school(
            Evaluation.objects.select_related('classe', 'matiere', 'ecole'), request.user, 'ecole'
        ),
        pk=evaluation_id,
    )
    eleve = get_object_or_404(
        filter_by_user_school(Eleve.objects.select_related('classe'), request.user, 'classe__ecole'),
        pk=eleve_id,
        classe=evaluation.classe,
    )

    Note.objects.filter(evaluation=evaluation, eleve=eleve).delete()
    total_notes = Note.objects.filter(evaluation=evaluation).count()
    return JsonResponse({'success': True, 'total_notes': total_notes})

@can_manage_notes
def evaluations_matiere(request, classe_id, matiere_id):
    """Liste des évaluations d'une matière pour une classe, avec accès rapide à la saisie et à l'affichage des notes."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.all(), request.user, 'ecole'), pk=classe_id)
    matiere = get_object_or_404(MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole), pk=matiere_id)
    evaluations = (
        Evaluation.objects.filter(classe=classe, matiere=matiere)
        .order_by('-date', '-id')
    )
    return render(request, 'notes/evaluations_matiere.html', {
        'classe': classe,
        'matiere': matiere,
        'evaluations': evaluations,
    })

@can_manage_notes
def evaluation_detail(request, evaluation_id):
    """Affiche un tableau des élèves de la classe avec leurs notes (ou vide si non saisie)."""
    evaluation = get_object_or_404(
        filter_by_user_school(Evaluation.objects.select_related('classe', 'matiere', 'ecole'), request.user, 'ecole'),
        pk=evaluation_id
    )
    # Élèves de la classe
    eleves = Eleve.objects.select_related('classe').filter(classe=evaluation.classe).order_by('prenom', 'nom')
    eleves = filter_by_user_school(eleves, request.user, 'classe__ecole')
    # Index des notes
    notes_map = {n.eleve_id: n for n in evaluation.notes.select_related('eleve')}
    rows = []
    for e in eleves:
        n = notes_map.get(e.id)
        rows.append({
            'eleve': e,
            'matricule': e.matricule,
            'note': getattr(n, 'note', None),
            'observation': getattr(n, 'observation', ''),
        })
    return render(request, 'notes/evaluation_detail.html', {
        'evaluation': evaluation,
        'rows': rows,
    })

@login_required
@require_school_object(model=Eleve, pk_kwarg='eleve_id', field_path='classe__ecole')
def bulletin_pdf(request, classe_id: int, eleve_id: int, trimestre: str = "T1"):
    """Génère un bulletin de notes PDF en utilisant le système intelligent"""
    # Sécuriser l'accès à la classe / élève
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.select_related('ecole'), request.user, 'ecole'), pk=classe_id)
    eleve = get_object_or_404(filter_by_user_school(Eleve.objects.select_related('classe', 'classe__ecole'), request.user, 'classe__ecole'), pk=eleve_id, classe=classe)

    # Trouver la ClasseNote correspondante
    from notes.models import ClasseNote
    
    # Mapping spécial pour certaines classes (ClasseEleve ID → ClasseNote ID)
    mapping_inverse = {
        8: 59,   # ClasseEleve '11ème série littéraire' → ClasseNote '11ème Série littéraire'
        56: 61,  # ClasseEleve '12ÈME ANNÉE' → ClasseNote '12ème Année'
    }
    
    try:
        # Essayer le mapping spécial d'abord
        if classe.id in mapping_inverse:
            classe_note = ClasseNote.objects.filter(id=mapping_inverse[classe.id]).first()
        else:
            # Sinon chercher par nom
            classe_note = ClasseNote.objects.filter(
                nom=classe.nom,
                annee_scolaire=classe.annee_scolaire,
                ecole=classe.ecole
            ).first()
    except Exception:
        classe_note = None

    if not classe_note:
        # Fallback: chercher un ClasseNote avec l'école
        classe_note = ClasseNote.objects.filter(
            ecole=classe.ecole
        ).order_by('nom').first()

    if not classe_note:
        return HttpResponse("Classe de notes non trouvée", status=404)

    # Convertir trimestre format court vers format attendu par bulletin_intelligent_pdf
    # Le système intelligent attend: 'TRIMESTRE_1', 'SEMESTRE_1', ou 'ANNUEL'
    trimestre_mapping = {
        'T1': 'TRIMESTRE_1',
        'T2': 'TRIMESTRE_2',
        'T3': 'TRIMESTRE_3',
        'S1': 'SEMESTRE_1',
        'S2': 'SEMESTRE_2'
    }
    periode = trimestre_mapping.get(trimestre, trimestre)

    # Appeler la fonction intelligente qui gère tout (photos, totaux, etc)
    from .bulletin_intelligent import bulletin_intelligent_pdf as gen_bulletin_intelligent
    return gen_bulletin_intelligent(request, eleve_id, classe_note.id, periode)

    # Récupérer les notes de l'élève pour ces évaluations
    # Filtrer par matières de la période (evaluation__classe n'existe pas)
    notes_by_eval = {n.evaluation_id: n for n in NoteEleve.objects.filter(eleve=eleve, evaluation__matiere__in=[m.id for m in matieres], evaluation__periode=periode_longue).select_related('evaluation', 'evaluation__matiere')}

    # Calculs des moyennes par matière
    lignes = []
    somme_moyennes_coef = Decimal('0')
    somme_coef_matieres = Decimal('0')

    for mat in matieres:
        evals = evals_by_matiere.get(mat.id, [])
        if not evals:
            moy_mat = None
        else:
            num = Decimal('0')
            den = Decimal('0')
            for ev in evals:
                n = notes_by_eval.get(ev.id)
                c = Decimal(ev.coefficient or 1)
                if n is None or n.note is None:
                    # Absence ou note manquante = 0
                    num += Decimal('0') * c
                else:
                    num += Decimal(n.note) * c
                den += c
            moy_mat = (num / den).quantize(Decimal('0.01')) if den > 0 else None

        # RÈGLE PÉDAGOGIQUE: Toutes les matières comptent (sans notes = 0)
        if moy_mat is None:
            moy_mat = Decimal('0')
        somme_moyennes_coef += moy_mat * Decimal(mat.coefficient or 1)
        somme_coef_matieres += Decimal(mat.coefficient or 1)

        lignes.append({
            'matiere': mat.nom,
            'coef_matiere': mat.coefficient,
            'moyenne': moy_mat,
        })

    moyenne_generale = None
    if somme_coef_matieres > 0:
        moyenne_generale = (somme_moyennes_coef / somme_coef_matieres).quantize(Decimal('0.01'))

    # Moyennes de classe par matière (pondérées par coeffs d'évaluations)
    moyennes_classe_par_matiere: dict[int, Decimal | None] = {}
    for mat in matieres:
        evals = evals_by_matiere.get(mat.id, [])
        total_num = Decimal('0'); total_den = Decimal('0')
        for ev in evals:
            # toutes les notes de l'évaluation pour la classe
            for n in Note.objects.filter(evaluation=ev).only('note'):
                if n.note is None:
                    continue
                cc = Decimal(ev.coefficient or 1)
                total_num += Decimal(n.note) * cc
                total_den += cc
        moyennes_classe_par_matiere[mat.id] = (total_num / total_den).quantize(Decimal('0.01')) if total_den > 0 else None

    # IMPORTANT: Utiliser la source centralisée pour garantir la cohérence
    # entre bulletins PDF, bulletins affichés et classements
    from .utils_rangs import calculer_rangs_classe_periode
    
    rang = None
    total_eleves_ayant_moyenne = 0
    
    if classe_note:
        rangs_dict = calculer_rangs_classe_periode(classe_note, periode_longue, use_cache=False)
        total_eleves_ayant_moyenne = len(rangs_dict)
        
        rang_info = rangs_dict.get(eleve.id)
        if rang_info:
            rang = rang_info.get('rang_num')
            # Utiliser la moyenne de la source centralisée pour cohérence
            moyenne_generale = rang_info.get('moyenne', moyenne_generale)

    # Mention selon barème adapté au niveau
    def mention_for(avg: Decimal | None) -> str:
        if avg is None:
            return ""
        if avg >= Decimal('18'):
            return "Excellent"
        if avg >= Decimal('16'):
            return "Très Bien"
        if avg >= Decimal('14'):
            return "Bien"
        if avg >= Decimal('12'):
            return "Assez Bien"
        if avg >= Decimal('10'):
            return "Passable"
        if avg >= Decimal('8'):
            return "Insuffisant"
        if avg >= Decimal('6'):
            return "Faible"
        return "Très faible"
    mention = mention_for(moyenne_generale)

    # Génération PDF
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab requis (pip install reportlab)", status=500)

    response = HttpResponse(content_type='application/pdf')
    filename = f"bulletin_{eleve.matricule}_{trimestre}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'

    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Filigrane
    _apply_watermark(c, width, height, ecole=getattr(classe, 'ecole', None))

    margin = 2 * cm
    y = height - margin

    # En-tête avec logo et coordonnées
    if getattr(classe, 'ecole', None):
        y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
    y -= 20; c.setFont('Helvetica-Bold', 14); c.drawCentredString(width/2, y, f"Bulletin de notes — {trimestre}"); y -= 40
    c.setFont('Helvetica', 12)
    c.drawString(margin, y, f"Élève: {eleve.prenom} {eleve.nom}  (Matricule: {eleve.matricule or '-'} )")
    y -= 16
    c.drawString(margin, y, f"Classe: {classe.nom} — Année: {getattr(classe, 'annee_scolaire', '')}")
    y -= 12
    c.setFillColor(colors.grey)
    c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
    c.setFillColor(colors.black)
    y -= 16

    # Détecter le niveau scolaire pour masquer les coefficients en primaire/maternelle
    from .calculs_moyennes import detecter_niveau_scolaire
    niveau_scolaire = detecter_niveau_scolaire(classe.nom)
    est_primaire_ou_maternelle = niveau_scolaire in ['PRIMAIRE', 'MATERNELLE']
    
    # Tableau entêtes
    c.setFont('Helvetica-Bold', 14)
    if est_primaire_ou_maternelle:
        headers = ["Matière", "Moyenne /20", "Moy. classe"]
        colw = [9*cm, 4*cm, 4*cm]
    else:
        headers = ["Matière", "Coef.", "Moyenne /20", "Moy. classe"]
        colw = [8*cm, 2.2*cm, 3.2*cm, 3.2*cm]
    x = margin
    for i, htxt in enumerate(headers):
        c.drawString(x, y, htxt)
        x += colw[i]
    y -= 14
    c.setFillColor(colors.lightgrey)
    c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
    c.setFillColor(colors.black)
    y -= 10

    c.setFont('Helvetica', 15)
    for row in lignes:
        if y < margin + 60:
            c.showPage()
            _apply_watermark(c, width, height, ecole=getattr(classe, 'ecole', None))
            y = height - margin
        x = margin
        c.drawString(x, y, str(row['matiere'])); x += colw[0]
        if not est_primaire_ou_maternelle:
            c.drawString(x, y, str(row['coef_matiere'])); x += colw[1]
        moy_txt = '-' if row['moyenne'] is None else f"{row['moyenne']}"
        c.drawString(x, y, moy_txt); x += colw[1] if est_primaire_ou_maternelle else colw[2]
        # moyenne de classe
        mc = moyennes_classe_par_matiere.get(next((m.id for m in matieres if m.nom == row['matiere']), None), None)
        mc_txt = '-' if mc is None else f"{mc}"
        c.drawString(x, y, mc_txt)
        y -= 14

    # Séparateur
    y -= 6
    c.setFillColor(colors.grey)
    c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
    c.setFillColor(colors.black)
    y -= 16

    # Moyenne générale + Rang + Mention
    c.setFont('Helvetica-Bold', 13)
    c.drawString(margin, y, f"Moyenne générale: {moyenne_generale if moyenne_generale is not None else '-'} / 20")
    y -= 16
    c.setFont('Helvetica', 12)
    if rang is not None:
        # Format intelligent du rang avec accord grammatical
        from .calculs_intelligent import formater_rang_intelligent
        sexe = getattr(eleve, 'sexe', 'M') or 'M'
        rang_str = formater_rang_intelligent(rang, sexe, total_eleves_ayant_moyenne)
        c.drawString(margin, y, f"Rang: {rang_str}")
        y -= 14
    if mention:
        c.drawString(margin, y, f"Mention: {mention}")
        y -= 18

    # Pied de page
    # Signatures
    c.setFont('Helvetica', 11)
    sig_y = margin + 50
    c.drawString(margin, sig_y, "Prof. principal:")
    c.line(margin + 120, sig_y-2, margin + 250, sig_y-2)
    c.drawString(margin + 280, sig_y, "Chef d'établ.:")
    c.line(margin + 380, sig_y-2, margin + 510, sig_y-2)
    c.drawString(margin, sig_y - 28, "Parent/Tuteur:")
    c.line(margin + 150, sig_y-30, margin + 320, sig_y-30)

    c.setFont('Helvetica-Oblique', 10)
    c.setFillColor(colors.darkgrey)
    from datetime import datetime
    c.drawString(margin, margin/2, f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    c.showPage(); c.save()
    return response

@login_required
def imprimer_tableau_notes_pdf(request):
    """Imprimer le tableau des notes avec ajustement des colonnes sur A4 landscape"""
    from django.template.loader import render_to_string
    from weasyprint import HTML, CSS
    from django.http import HttpResponse
    
    # Récupérer les paramètres
    classe_id = request.GET.get('classe_id')
    periode = request.GET.get('periode')
    
    if not classe_id or not periode:
        return HttpResponse("Paramètres manquants", status=400)
    
    try:
        # Récupérer la classe et les données
        classe_note = get_object_or_404(ClasseNote, pk=classe_id)
        classe = classe_note.classe
        
        # Récupérer les matières
        matieres = MatiereNote.objects.filter(classe=classe_note).order_by('matiere__nom')
        
        # Calculer le classement
        from .calculs_moyennes import calculer_classement_classe
        from .calculs_intelligent import calculer_rang_intelligent
        
        # Récupérer les élèves
        from eleves.models import Eleve
        eleves = Eleve.objects.filter(classe=classe, est_actif=True).order_by('matricule')
        
        # Calculer les moyennes et rangs
        classement_resultat = calculer_classement_classe(eleves, matieres, periode, 'mensuel')
        
        # Préparer les données pour le template
        classement_data = []
        for eleve in eleves:
            # Récupérer les détails des notes par matière
            details_matieres = {}
            for matiere in matieres:
                from .calculs_moyennes import calculer_moyenne_matiere
                result = calculer_moyenne_matiere(eleve, matiere, periode, 'mensuel')
                details_matieres[matiere.id] = result
            
            # Récupérer le rang et la moyenne
            rang_num = classement_resultat['rang_map'].get(eleve.id)
            rang_str = str(rang_num) if rang_num else "-"
            moyenne = classement_resultat['moyennes_par_eleve'].get(eleve.id)
            
            # Formatter le rang avec ex-æquo si nécessaire
            if rang_num:
                from .calculs_intelligent import formater_rang_intelligent
                sexe = getattr(eleve, 'sexe', 'M') or 'M'
                rang_str = formater_rang_intelligent(rang_num, sexe)
            
            classement_data.append({
                'matricule': eleve.matricule,
                'nom_complet': eleve.nom_complet,
                'rang': rang_str,
                'moyenne': moyenne,
                'details_matieres': details_matieres,
                'sexe': getattr(eleve, 'sexe', 'M') or 'M'
            })
        
        # Trier par rang
        classement_data.sort(key=lambda x: x['rang'] if x['rang'] != '-' else '999')
        
        # Contexte pour le template
        context = {
            'classe_selectionnee': classe,
            'periode_selectionnee': periode,
            'matieres': matieres,
            'classement_data': classement_data,
        }
        
        # Générer le HTML
        html_content = render_to_string('notes/impression_tableau_notes.html', context, request=request)
        
        # Créer le PDF avec WeasyPrint
        html = HTML(string=html_content)
        css = CSS(string='''
            @page {
                size: A4 landscape;
                margin: 10mm;
            }
        ''')
        
        pdf = html.write_pdf(stylesheets=[css])
        
        # Retourner le PDF
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="tableau_notes_{classe.nom}_{periode}.pdf"'
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'impression du tableau: {str(e)}")
        return HttpResponse(f"Erreur: {str(e)}", status=500)

@admin_required
def bulletins_mensuels_classe_pdf(request, classe_id: int, mois: int):
    """Génère en un seul PDF les bulletins mensuels de tous les élèves d'une classe (Collège/Lycée)."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.select_related('ecole'), request.user, 'ecole'), pk=classe_id)
    eleves = Eleve.objects.select_related('classe').filter(classe=classe).order_by('prenom', 'nom')
    eleves = filter_by_user_school(eleves, request.user, 'classe__ecole')
    matieres = MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole, actif=True).order_by('nom')
    annee_scolaire = getattr(classe, 'annee_scolaire', None)
    
    # Détecter le niveau scolaire pour masquer les coefficients en primaire/maternelle
    from .calculs_moyennes import detecter_niveau_scolaire
    niveau_scolaire = detecter_niveau_scolaire(classe.nom)
    est_primaire_ou_maternelle = niveau_scolaire in ['PRIMAIRE', 'MATERNELLE']

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab requis (pip install reportlab)", status=500)

    response = HttpResponse(content_type='application/pdf')
    filename = f"bulletins_classe_mois_{mois:02d}_{classe.nom}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    mois_label = [None,'Janvier','Février','Mars','Avril','Mai','Juin','Juillet','Août','Septembre','Octobre','Novembre','Décembre'][mois] if 1 <= mois <= 12 else f"Mois {mois}"

    # Pré-calcul du classement mensuel (moy. générale mensuelle pondérée par coef matière)
    classement_list = []  # list of (eleve, moy_mensuelle)
    for e in eleves:
        s_num = Decimal('0'); s_den = Decimal('0')
        for mat in matieres:
            mm = monthly_avg(e, mat, annee_scolaire, mois, mode='weighted')
            if mm is not None:
                s_num += mm * Decimal(mat.coefficient or 1)
                s_den += Decimal(mat.coefficient or 1)
        avg = (s_num / s_den).quantize(Decimal('0.01')) if s_den > 0 else None
        if avg is not None:
            classement_list.append((e, avg))
    classement_list.sort(key=lambda t: t[1], reverse=True)
    # Trier les élèves par classement mensuel (1er au dernier)
    rang_map_mensuel = {e.id: idx + 1 for idx, (e, _) in enumerate(classement_list)}
    eleves = sorted(eleves, key=lambda e: rang_map_mensuel.get(e.id, 9999))

    # Page 1: Couverture
    margin = 2 * cm
    _apply_watermark(c, width, height, ecole=getattr(classe, 'ecole', None))
    y = height - margin
    if getattr(classe, 'ecole', None):
        y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
    c.setFont('Helvetica-Bold', 20); c.drawCentredString(width/2, y-10, f"Bulletins mensuels — {mois_label}"); y -= 50
    c.setFont('Helvetica', 15); c.drawCentredString(width/2, y, f"Classe: {classe.nom}"); y -= 20
    c.drawCentredString(width/2, y, f"Année scolaire: {annee_scolaire or ''}"); y -= 20
    c.setFont('Helvetica-Oblique', 11); c.setFillColorRGB(0.3,0.3,0.3)
    c.drawCentredString(width/2, y, f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    try:
        c.setFillColorRGB(0,0,0)
    except Exception:
        pass
    c.showPage()

    # Page 2: Table des matières
    _apply_watermark(c, width, height, ecole=getattr(classe, 'ecole', None))
    y = height - margin
    if getattr(classe, 'ecole', None):
        y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
    c.setFont('Helvetica-Bold', 16); c.drawCentredString(width/2, y-10, f"Table des matières — {mois_label}"); y -= 40
    c.setFont('Helvetica', 12); c.drawString(margin, y, f"Classe: {classe.nom} — Année: {annee_scolaire or ''}"); y -= 20
    c.setFont('Helvetica-Bold', 13); c.drawString(margin, y, "Table des matières"); y -= 18
    c.setFont('Helvetica', 11)
    start_page = 3  # 1 page de couverture + 1 page de TOC, puis 1 page par élève
    page_no = start_page
    for e in eleves:
        c.drawString(margin, y, f"{e.prenom} {e.nom}")
        c.drawRightString(width - margin, y, f"p. {page_no}")
        y -= 14
        page_no += 1
        if y < margin + 40:
            c.showPage()
            _apply_watermark(c, width, height, ecole=getattr(classe, 'ecole', None))
            y = height - margin
    c.showPage()

    for eleve in eleves:
        # Filigrane
        _apply_watermark(c, width, height, ecole=getattr(classe, 'ecole', None))
        y = height - margin
        if getattr(classe, 'ecole', None):
            y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
        y -= 20; c.setFont('Helvetica-Bold', 14); c.drawCentredString(width/2, y, f"Bulletin mensuel — {mois_label}"); y -= 40
        c.setFont('Helvetica', 12)
        c.drawString(margin, y, f"Élève: {eleve.prenom} {eleve.nom}  (Matricule: {eleve.matricule or '-'} )"); y -= 16
        c.drawString(margin, y, f"Classe: {classe.nom} — Année: {annee_scolaire or ''}"); y -= 12
        c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16

        # Calculs
        lignes = []
        somme_moyennes_coef = Decimal('0'); somme_coef_matieres = Decimal('0')
        for mat in matieres:
            moy_cours = course_month_avg(eleve, mat, annee_scolaire, mois)
            moy_compo = compo_month_avg(eleve, mat, annee_scolaire, mois)
            moy_mois = monthly_avg(eleve, mat, annee_scolaire, mois, mode='weighted')
            if moy_mois is not None:
                somme_moyennes_coef += moy_mois * Decimal(mat.coefficient or 1)
                somme_coef_matieres += Decimal(mat.coefficient or 1)
            lignes.append({'matiere': mat.nom,'coef_matiere': mat.coefficient,'moy_cours': moy_cours,'moy_compo': moy_compo,'moy_mois': moy_mois})
        moyenne_generale = (somme_moyennes_coef / somme_coef_matieres).quantize(Decimal('0.01')) if somme_coef_matieres > 0 else None

        # Entêtes
        c.setFont('Helvetica-Bold', 12)
        if est_primaire_ou_maternelle:
            headers = ["Matière", "Moy. cours", "Moy. compo", "Moy. mois"]
            colw = [8*cm, 3.5*cm, 3.5*cm, 3.5*cm]
        else:
            headers = ["Matière", "Coef.", "Moy. cours", "Moy. compo", "Moy. mois"]
            colw = [7.5*cm, 2.0*cm, 3.0*cm, 3.0*cm, 3.0*cm]
        x = margin
        for i, htxt in enumerate(headers):
            c.drawString(x, y, htxt); x += colw[i]
        y -= 14
        c.setFillColor(colors.lightgrey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 10
        c.setFont('Helvetica', 15)
        for row in lignes:
            if y < margin + 60:
                c.showPage(); y = height - margin
                _apply_watermark(c, width, height, ecole=getattr(classe, 'ecole', None))
            x = margin
            c.drawString(x, y, str(row['matiere'])); x += colw[0]
            if not est_primaire_ou_maternelle:
                c.drawString(x, y, str(row['coef_matiere'])); x += colw[1]
            c.drawString(x, y, '-' if row['moy_cours'] is None else f"{row['moy_cours']}"); x += colw[1] if est_primaire_ou_maternelle else colw[2]
            c.drawString(x, y, '-' if row['moy_compo'] is None else f"{row['moy_compo']}"); x += colw[2] if est_primaire_ou_maternelle else colw[3]
            c.drawString(x, y, '-' if row['moy_mois'] is None else f"{row['moy_mois']}")
            y -= 14
        y -= 6; c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16
        c.setFont('Helvetica-Bold', 13)
        c.drawString(margin, y, f"Moyenne générale mensuelle: {moyenne_generale if moyenne_generale is not None else '-'} / 20");
        c.showPage()
    c.save();
    return response


@login_required
def imprimer_tableau_notes_pdf(request):
    """Imprimer le tableau des notes avec ajustement des colonnes sur A4 landscape"""
    from django.template.loader import render_to_string
    from weasyprint import HTML, CSS
    from django.http import HttpResponse
    
    # Récupérer les paramètres
    classe_id = request.GET.get('classe_id')
    periode = request.GET.get('periode')
    
    if not classe_id or not periode:
        return HttpResponse("Paramètres manquants", status=400)
    
    try:
        # Récupérer la classe et les données
        classe_note = get_object_or_404(ClasseNote, pk=classe_id)
        classe = classe_note.classe
        
        # Récupérer les matières
        matieres = MatiereNote.objects.filter(classe=classe_note).order_by('matiere__nom')
        
        # Calculer le classement
        from .calculs_moyennes import calculer_classement_classe
        from .calculs_intelligent import calculer_rang_intelligent
        
        # Récupérer les élèves
        from eleves.models import Eleve
        eleves = Eleve.objects.filter(classe=classe, est_actif=True).order_by('matricule')
        
        # Calculer les moyennes et rangs
        classement_resultat = calculer_classement_classe(eleves, matieres, periode, 'mensuel')
        
        # Préparer les données pour le template
        classement_data = []
        for eleve in eleves:
            # Récupérer les détails des notes par matière
            details_matieres = {}
            for matiere in matieres:
                from .calculs_moyennes import calculer_moyenne_matiere
                result = calculer_moyenne_matiere(eleve, matiere, periode, 'mensuel')
                details_matieres[matiere.id] = result
            
            # Récupérer le rang et la moyenne
            rang_num = classement_resultat['rang_map'].get(eleve.id)
            rang_str = str(rang_num) if rang_num else "-"
            moyenne = classement_resultat['moyennes_par_eleve'].get(eleve.id)
            
            # Formatter le rang avec ex-æquo si nécessaire
            if rang_num:
                from .calculs_intelligent import formater_rang_intelligent
                sexe = getattr(eleve, 'sexe', 'M') or 'M'
                rang_str = formater_rang_intelligent(rang_num, sexe)
            
            classement_data.append({
                'matricule': eleve.matricule,
                'nom_complet': eleve.nom_complet,
                'rang': rang_str,
                'moyenne': moyenne,
                'details_matieres': details_matieres,
                'sexe': getattr(eleve, 'sexe', 'M') or 'M'
            })
        
        # Trier par rang
        classement_data.sort(key=lambda x: x['rang'] if x['rang'] != '-' else '999')
        
        # Contexte pour le template
        context = {
            'classe_selectionnee': classe,
            'periode_selectionnee': periode,
            'matieres': matieres,
            'classement_data': classement_data,
        }
        
        # Générer le HTML
        html_content = render_to_string('notes/impression_tableau_notes.html', context, request=request)
        
        # Créer le PDF avec WeasyPrint
        html = HTML(string=html_content)
        css = CSS(string='''
            @page {
                size: A4 landscape;
                margin: 10mm;
            }
        ''')
        
        pdf = html.write_pdf(stylesheets=[css])
        
        # Retourner le PDF
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="tableau_notes_{classe.nom}_{periode}.pdf"'
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'impression du tableau: {str(e)}")
        return HttpResponse(f"Erreur: {str(e)}", status=500)

@admin_required
def bulletins_semestre_classe_pdf(request, classe_id: int, semestre: int = 1):
    """Génère en un seul PDF les bulletins semestriels de tous les élèves d'une classe (Collège/Lycée)."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.select_related('ecole'), request.user, 'ecole'), pk=classe_id)
    eleves = Eleve.objects.select_related('classe').filter(classe=classe).order_by('prenom', 'nom')
    eleves = filter_by_user_school(eleves, request.user, 'classe__ecole')
    matieres = MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole, actif=True).order_by('nom')
    annee_scolaire = getattr(classe, 'annee_scolaire', None)
    
    # Détecter le niveau scolaire pour masquer les coefficients en primaire/maternelle
    from .calculs_moyennes import detecter_niveau_scolaire
    niveau_scolaire = detecter_niveau_scolaire(classe.nom)
    est_primaire_ou_maternelle = niveau_scolaire in ['PRIMAIRE', 'MATERNELLE']

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab requis (pip install reportlab)", status=500)

    response = HttpResponse(content_type='application/pdf')
    filename = f"bulletins_classe_semestre{semestre}_{classe.nom}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Pré-calcul des moyennes générales pour classement
    moyenne_map = {}
    for e in eleves:
        s_num = Decimal('0'); s_den = Decimal('0')
        for mat in matieres:
            moy_sem = semester_avg(e, mat, annee_scolaire, semestre, mode='weighted')
            if moy_sem is not None:
                s_num += moy_sem * Decimal(mat.coefficient or 1)
                s_den += Decimal(mat.coefficient or 1)
        moyenne_map[e.id] = (s_num / s_den).quantize(Decimal('0.01')) if s_den > 0 else None
    classement_list = [(e, moyenne_map.get(e.id)) for e in eleves]
    classement_list = [(e, m) for (e, m) in classement_list if m is not None]
    classement_list.sort(key=lambda t: t[1], reverse=True)
    rang_map = {e.id: idx for idx, (e, _) in enumerate(classement_list, start=1)}
    # Trier les élèves par classement semestriel (1er au dernier)
    eleves = sorted(eleves, key=lambda e: rang_map.get(e.id, 9999))

    # Page 1: Couverture
    margin = 2 * cm
    _apply_watermark(c, width, height, ecole=getattr(classe, 'ecole', None))
    y = height - margin
    if getattr(classe, 'ecole', None):
        y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
    c.setFont('Helvetica-Bold', 20); c.drawCentredString(width/2, y-10, f"Bulletins semestriels — S{semestre}"); y -= 50
    c.setFont('Helvetica', 15); c.drawCentredString(width/2, y, f"Classe: {classe.nom}"); y -= 20
    c.drawCentredString(width/2, y, f"Année scolaire: {annee_scolaire or ''}"); y -= 20
    c.setFont('Helvetica-Oblique', 11); c.setFillColorRGB(0.3,0.3,0.3)
    c.drawCentredString(width/2, y, f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    try:
        c.setFillColorRGB(0,0,0)
    except Exception:
        pass
    c.showPage()

    # Page 2: Table des matières
    _apply_watermark(c, width, height, ecole=getattr(classe, 'ecole', None))
    y = height - margin
    if getattr(classe, 'ecole', None):
        y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
    c.setFont('Helvetica-Bold', 16); c.drawCentredString(width/2, y-10, f"Table des matières — S{semestre}"); y -= 40
    c.setFont('Helvetica', 12); c.drawString(margin, y, f"Classe: {classe.nom} — Année: {annee_scolaire or ''}"); y -= 20
    c.setFont('Helvetica-Bold', 13); c.drawString(margin, y, "Table des matières"); y -= 18
    c.setFont('Helvetica', 11)
    start_page = 3
    page_no = start_page
    for e in eleves:
        c.drawString(margin, y, f"{e.prenom} {e.nom}")
        c.drawRightString(width - margin, y, f"p. {page_no}")
        y -= 14
        page_no += 1
        if y < margin + 40:
            c.showPage()
            _apply_watermark(c, width, height, ecole=getattr(classe, 'ecole', None))
            y = height - margin
    c.showPage()

    for eleve in eleves:
        _apply_watermark(c, width, height, ecole=getattr(classe, 'ecole', None))
        y = height - margin
        if getattr(classe, 'ecole', None):
            y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
        y -= 20; c.setFont('Helvetica-Bold', 14); c.drawCentredString(width/2, y, f"Bulletin semestriel — S{semestre}"); y -= 40
        c.setFont('Helvetica', 12)
        c.drawString(margin, y, f"Élève: {eleve.prenom} {eleve.nom}  (Matricule: {eleve.matricule or '-'} )"); y -= 16
        c.drawString(margin, y, f"Classe: {classe.nom} — Année: {annee_scolaire or ''}"); y -= 12
        c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16

        # Calculs
        lignes = []
        somme_moyennes_coef = Decimal('0'); somme_coef_matieres = Decimal('0')
        for mat in matieres:
            moy_cours = semester_course_avg(eleve, mat, annee_scolaire, semestre)
            moy_compo = semester_compo_avg(eleve, mat, annee_scolaire, semestre)
            moy_sem = semester_avg(eleve, mat, annee_scolaire, semestre, mode='weighted')
            if moy_sem is not None:
                somme_moyennes_coef += moy_sem * Decimal(mat.coefficient or 1)
                somme_coef_matieres += Decimal(mat.coefficient or 1)
            lignes.append({'matiere': mat.nom,'coef_matiere': mat.coefficient,'moy_cours': moy_cours,'moy_compo': moy_compo,'moy_sem': moy_sem})
        moyenne_generale = (somme_moyennes_coef / somme_coef_matieres).quantize(Decimal('0.01')) if somme_coef_matieres > 0 else None

        # Entêtes
        c.setFont('Helvetica-Bold', 12)
        if est_primaire_ou_maternelle:
            headers = ["Matière", "Moy. cours", "Moy. compo", f"Moy. S{semestre}"]
            colw = [8*cm, 3.5*cm, 3.5*cm, 3.5*cm]
        else:
            headers = ["Matière", "Coef.", "Moy. cours", "Moy. compo", f"Moy. S{semestre}"]
            colw = [7.5*cm, 2.0*cm, 3.0*cm, 3.0*cm, 3.0*cm]
        x = margin
        for i, htxt in enumerate(headers):
            c.drawString(x, y, htxt); x += colw[i]
        y -= 14
        c.setFillColor(colors.lightgrey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 10
        c.setFont('Helvetica', 15)
        for row in lignes:
            if y < margin + 60:
                c.showPage(); y = height - margin
                _apply_watermark(c, width, height, ecole=getattr(classe, 'ecole', None))
            x = margin
            c.drawString(x, y, str(row['matiere'])); x += colw[0]
            if not est_primaire_ou_maternelle:
                c.drawString(x, y, str(row['coef_matiere'])); x += colw[1]
            c.drawString(x, y, '-' if row['moy_cours'] is None else f"{row['moy_cours']}"); x += colw[1] if est_primaire_ou_maternelle else colw[2]
            c.drawString(x, y, '-' if row['moy_compo'] is None else f"{row['moy_compo']}"); x += colw[2] if est_primaire_ou_maternelle else colw[3]
            c.drawString(x, y, '-' if row['moy_sem'] is None else f"{row['moy_sem']}")
            y -= 14
        y -= 6; c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16
        c.setFont('Helvetica-Bold', 13)
        c.drawString(margin, y, f"Moyenne générale semestrielle: {moyenne_generale if moyenne_generale is not None else '-'} / 20");
        c.showPage()
    c.save();
    return response


@login_required
def imprimer_tableau_notes_pdf(request):
    """Imprimer le tableau des notes avec ajustement des colonnes sur A4 landscape"""
    from django.template.loader import render_to_string
    from weasyprint import HTML, CSS
    from django.http import HttpResponse
    
    # Récupérer les paramètres
    classe_id = request.GET.get('classe_id')
    periode = request.GET.get('periode')
    
    if not classe_id or not periode:
        return HttpResponse("Paramètres manquants", status=400)
    
    try:
        # Récupérer la classe et les données
        classe_note = get_object_or_404(ClasseNote, pk=classe_id)
        classe = classe_note.classe
        
        # Récupérer les matières
        matieres = MatiereNote.objects.filter(classe=classe_note).order_by('matiere__nom')
        
        # Calculer le classement
        from .calculs_moyennes import calculer_classement_classe
        from .calculs_intelligent import calculer_rang_intelligent
        
        # Récupérer les élèves
        from eleves.models import Eleve
        eleves = Eleve.objects.filter(classe=classe, est_actif=True).order_by('matricule')
        
        # Calculer les moyennes et rangs
        classement_resultat = calculer_classement_classe(eleves, matieres, periode, 'mensuel')
        
        # Préparer les données pour le template
        classement_data = []
        for eleve in eleves:
            # Récupérer les détails des notes par matière
            details_matieres = {}
            for matiere in matieres:
                from .calculs_moyennes import calculer_moyenne_matiere
                result = calculer_moyenne_matiere(eleve, matiere, periode, 'mensuel')
                details_matieres[matiere.id] = result
            
            # Récupérer le rang et la moyenne
            rang_num = classement_resultat['rang_map'].get(eleve.id)
            rang_str = str(rang_num) if rang_num else "-"
            moyenne = classement_resultat['moyennes_par_eleve'].get(eleve.id)
            
            # Formatter le rang avec ex-æquo si nécessaire
            if rang_num:
                from .calculs_intelligent import formater_rang_intelligent
                sexe = getattr(eleve, 'sexe', 'M') or 'M'
                rang_str = formater_rang_intelligent(rang_num, sexe)
            
            classement_data.append({
                'matricule': eleve.matricule,
                'nom_complet': eleve.nom_complet,
                'rang': rang_str,
                'moyenne': moyenne,
                'details_matieres': details_matieres,
                'sexe': getattr(eleve, 'sexe', 'M') or 'M'
            })
        
        # Trier par rang
        classement_data.sort(key=lambda x: x['rang'] if x['rang'] != '-' else '999')
        
        # Contexte pour le template
        context = {
            'classe_selectionnee': classe,
            'periode_selectionnee': periode,
            'matieres': matieres,
            'classement_data': classement_data,
        }
        
        # Générer le HTML
        html_content = render_to_string('notes/impression_tableau_notes.html', context, request=request)
        
        # Créer le PDF avec WeasyPrint
        html = HTML(string=html_content)
        css = CSS(string='''
            @page {
                size: A4 landscape;
                margin: 10mm;
            }
        ''')
        
        pdf = html.write_pdf(stylesheets=[css])
        
        # Retourner le PDF
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="tableau_notes_{classe.nom}_{periode}.pdf"'
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'impression du tableau: {str(e)}")
        return HttpResponse(f"Erreur: {str(e)}", status=500)

def course_month_avg(eleve, matiere, annee_scolaire, mois):
    """Calcule la moyenne des devoirs pour un mois donné"""
    from django.db.models import Q
    
    periode_str = {
        1: 'JANVIER', 2: 'FEVRIER', 3: 'MARS', 4: 'AVRIL',
        5: 'MAI', 6: 'JUIN', 7: 'JUILLET', 8: 'AOUT',
        9: 'SEPTEMBRE', 10: 'OCTOBRE', 11: 'NOVEMBRE', 12: 'DECEMBRE'
    }.get(mois, 'OCTOBRE')
    
    # Récupérer les évaluations de type DEVOIR pour ce mois
    evals = Evaluation.objects.filter(
        matiere=matiere,
        periode=periode_str,
        type_evaluation__in=['DEVOIR', 'CONTROLE', 'INTERROGATION']
    )
    
    total = Decimal('0')
    count = 0
    
    for ev in evals:
        try:
            n = NoteEleve.objects.get(eleve=eleve, evaluation=ev)
            if n.absent or n.note is None:
                total += Decimal('0')  # Absence = 0
                count += 1
            else:
                total += Decimal(str(n.note))
                count += 1
        except NoteEleve.DoesNotExist:
            total += Decimal('0')  # Pas de note = 0
            count += 1
    
    return (total / count).quantize(Decimal('0.01')) if count > 0 else None

def compo_month_avg(eleve, matiere, annee_scolaire, mois):
    """Calcule la moyenne de composition pour un mois donné"""
    from django.db.models import Q
    
    periode_str = {
        1: 'JANVIER', 2: 'FEVRIER', 3: 'MARS', 4: 'AVRIL',
        5: 'MAI', 6: 'JUIN', 7: 'JUILLET', 8: 'AOUT',
        9: 'SEPTEMBRE', 10: 'OCTOBRE', 11: 'NOVEMBRE', 12: 'DECEMBRE'
    }.get(mois, 'OCTOBRE')
    
    # Récupérer les évaluations de type COMPOSITION pour ce mois
    evals = Evaluation.objects.filter(
        matiere=matiere,
        periode=periode_str,
        type_evaluation__in=['COMPOSITION', 'EXAMEN']
    )
    
    total = Decimal('0')
    count = 0
    
    for ev in evals:
        try:
            n = NoteEleve.objects.get(eleve=eleve, evaluation=ev)
            if n.absent or n.note is None:
                total += Decimal('0')  # Absence = 0
                count += 1
            else:
                total += Decimal(str(n.note))
                count += 1
        except NoteEleve.DoesNotExist:
            total += Decimal('0')  # Pas de note = 0
            count += 1
    
    return (total / count).quantize(Decimal('0.01')) if count > 0 else None

def monthly_avg(eleve, matiere, annee_scolaire, mois, mode='weighted'):
    """Calcule la moyenne mensuelle (pondérée ou simple)"""
    moy_cours = course_month_avg(eleve, matiere, annee_scolaire, mois)
    moy_compo = compo_month_avg(eleve, matiere, annee_scolaire, mois)
    
    if mode == 'weighted':
        # Formule pondérée : (moy_cours + 2 * moy_compo) / 3
        if moy_cours is not None and moy_compo is not None:
            return ((moy_cours + moy_compo * 2) / 3).quantize(Decimal('0.01'))
        elif moy_compo is not None:
            return moy_compo
        elif moy_cours is not None:
            return moy_cours
    else:
        # Moyenne simple
        if moy_cours is not None and moy_compo is not None:
            return ((moy_cours + moy_compo) / 2).quantize(Decimal('0.01'))
        elif moy_compo is not None:
            return moy_compo
        elif moy_cours is not None:
            return moy_cours
    
    return None

@login_required
@require_school_object(model=Eleve, pk_kwarg='eleve_id', field_path='classe__ecole')
def bulletin_mensuel_pdf(request, classe_id: int, eleve_id: int, mois: int):
    """Bulletin mensuel Collège/Lycée pour un élève.
    Colonnes: Matière, Coef., Moy. cours (mois), Moy. compo (mois), Moy. mensuelle (pondérée 2:1 si compo présente).
    """
    # Sécuriser classe/élève
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.select_related('ecole'), request.user, 'ecole'), pk=classe_id)
    eleve = get_object_or_404(filter_by_user_school(Eleve.objects.select_related('classe', 'classe__ecole'), request.user, 'classe__ecole'), pk=eleve_id, classe=classe)

    # Matières actives
    matieres = MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole, actif=True).order_by('nom')
    annee_scolaire = getattr(classe, 'annee_scolaire', None)

    lignes = []
    somme_moyennes_coef = Decimal('0'); somme_coef_matieres = Decimal('0')
    for mat in matieres:
        moy_cours = course_month_avg(eleve, mat, annee_scolaire, mois)
        moy_compo = compo_month_avg(eleve, mat, annee_scolaire, mois)
        moy_mois = monthly_avg(eleve, mat, annee_scolaire, mois, mode='weighted')
        
        # RÈGLE PÉDAGOGIQUE: Toutes les matières comptent (sans notes = 0)
        moy_mois_calcul = moy_mois if moy_mois is not None else Decimal('0')
        somme_moyennes_coef += moy_mois_calcul * Decimal(mat.coefficient or 1)
        somme_coef_matieres += Decimal(mat.coefficient or 1)
        
        lignes.append({
            'matiere': mat.nom,
            'coef_matiere': mat.coefficient,
            'moy_cours': moy_cours,
            'moy_compo': moy_compo,
            'moy_mois': moy_mois,  # Garder None pour affichage
        })

    moyenne_generale = (somme_moyennes_coef / somme_coef_matieres).quantize(Decimal('0.01')) if somme_coef_matieres > 0 else None

    # IMPORTANT: Utiliser la source centralisée pour garantir la cohérence
    # entre bulletins PDF, bulletins affichés et classements
    from .utils_rangs import calculer_rangs_classe_periode
    from .models import ClasseNote
    
    # Trouver la ClasseNote correspondante
    classe_note = ClasseNote.objects.filter(
        nom=classe.nom,
        annee_scolaire=classe.annee_scolaire,
        ecole=classe.ecole
    ).first()
    
    # Convertir le mois en période
    mois_mapping = {
        1: 'JANVIER', 2: 'FEVRIER', 3: 'MARS', 4: 'AVRIL', 5: 'MAI', 6: 'JUIN',
        7: 'JUILLET', 8: 'AOUT', 9: 'SEPTEMBRE', 10: 'OCTOBRE', 11: 'NOVEMBRE', 12: 'DECEMBRE'
    }
    periode = mois_mapping.get(mois, 'OCTOBRE')
    
    rang_str = "-"
    total_eleves = 0
    
    if classe_note:
        rangs_dict = calculer_rangs_classe_periode(classe_note, periode, use_cache=False)
        total_eleves = len(rangs_dict)
        
        rang_info = rangs_dict.get(eleve.id)
        if rang_info:
            rang_str = rang_info.get('rang', '-')
            # Utiliser la moyenne de la source centralisée pour cohérence
            moyenne_generale = rang_info.get('moyenne', moyenne_generale)

    # PDF
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab requis (pip install reportlab)", status=500)

    response = HttpResponse(content_type='application/pdf')
    filename = f"bulletin_mensuel_{eleve.matricule}_{mois:02d}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'

    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Filigrane
    _apply_watermark(c, width, height, ecole=getattr(classe, 'ecole', None))

    margin = 2 * cm
    y = height - margin

    # En-tête
    if getattr(classe, 'ecole', None):
        y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
    mois_label = [
        None, 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre'
    ][mois] if 1 <= mois <= 12 else f"Mois {mois}"
    y -= 20; c.setFont('Helvetica-Bold', 14); c.drawCentredString(width/2, y, f"Bulletin mensuel — {mois_label}"); y -= 40
    c.setFont('Helvetica', 12)
    c.drawString(margin, y, f"Élève: {eleve.prenom} {eleve.nom}  (Matricule: {eleve.matricule or '-'} )"); y -= 16
    c.drawString(margin, y, f"Classe: {classe.nom} — Année: {annee_scolaire or ''}"); y -= 12
    c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16

    # Détecter le niveau scolaire pour masquer les coefficients en primaire/maternelle
    from .calculs_moyennes import detecter_niveau_scolaire
    niveau_scolaire = detecter_niveau_scolaire(classe.nom)
    est_primaire_ou_maternelle = niveau_scolaire in ['PRIMAIRE', 'MATERNELLE']
    
    # Entêtes colonnes
    c.setFont('Helvetica-Bold', 12)
    if est_primaire_ou_maternelle:
        headers = ["Matière", "Moy. cours", "Moy. compo", "Moy. mois"]
        colw = [8*cm, 3.5*cm, 3.5*cm, 3.5*cm]
    else:
        headers = ["Matière", "Coef.", "Moy. cours", "Moy. compo", "Moy. mois"]
        colw = [7.5*cm, 2.0*cm, 3.0*cm, 3.0*cm, 3.0*cm]
    x = margin
    for i, htxt in enumerate(headers):
        c.drawString(x, y, htxt); x += colw[i]
    y -= 14
    c.setFillColor(colors.lightgrey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 10

    c.setFont('Helvetica', 15)
    for row in lignes:
        if y < margin + 60:
            c.showPage()
            _apply_watermark(c, width, height, ecole=getattr(classe, 'ecole', None))
            y = height - margin
        x = margin
        c.drawString(x, y, str(row['matiere'])); x += colw[0]
        if not est_primaire_ou_maternelle:
            c.drawString(x, y, str(row['coef_matiere'])); x += colw[1]
        c.drawString(x, y, '-' if row['moy_cours'] is None else f"{row['moy_cours']}"); x += colw[1] if est_primaire_ou_maternelle else colw[2]
        c.drawString(x, y, '-' if row['moy_compo'] is None else f"{row['moy_compo']}"); x += colw[2] if est_primaire_ou_maternelle else colw[3]
        c.drawString(x, y, '-' if row['moy_mois'] is None else f"{row['moy_mois']}")
        y -= 14

    # Séparateur
    y -= 6; c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16

    # Moyenne générale et rang
    c.setFont('Helvetica-Bold', 13)
    c.drawString(margin, y, f"Moyenne générale mensuelle: {moyenne_generale if moyenne_generale is not None else '-'} / 20"); y -= 18
    c.drawString(margin, y, f"Rang: {rang_str}"); y -= 18
    if total_eleves > 0:
        c.setFont('Helvetica', 11)
        c.drawString(margin, y, f"Effectif: {total_eleves} élèves"); y -= 18

    # Pied
    c.setFont('Helvetica-Oblique', 10); c.setFillColor(colors.darkgrey)
    c.drawString(margin, margin/2, f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    c.showPage(); c.save()
    return response


@login_required
def imprimer_tableau_notes_pdf(request):
    """Imprimer le tableau des notes avec ajustement des colonnes sur A4 landscape"""
    from django.template.loader import render_to_string
    from weasyprint import HTML, CSS
    from django.http import HttpResponse
    
    # Récupérer les paramètres
    classe_id = request.GET.get('classe_id')
    periode = request.GET.get('periode')
    
    if not classe_id or not periode:
        return HttpResponse("Paramètres manquants", status=400)
    
    try:
        # Récupérer la classe et les données
        classe_note = get_object_or_404(ClasseNote, pk=classe_id)
        classe = classe_note.classe
        
        # Récupérer les matières
        matieres = MatiereNote.objects.filter(classe=classe_note).order_by('matiere__nom')
        
        # Calculer le classement
        from .calculs_moyennes import calculer_classement_classe
        from .calculs_intelligent import calculer_rang_intelligent
        
        # Récupérer les élèves
        from eleves.models import Eleve
        eleves = Eleve.objects.filter(classe=classe, est_actif=True).order_by('matricule')
        
        # Calculer les moyennes et rangs
        classement_resultat = calculer_classement_classe(eleves, matieres, periode, 'mensuel')
        
        # Préparer les données pour le template
        classement_data = []
        for eleve in eleves:
            # Récupérer les détails des notes par matière
            details_matieres = {}
            for matiere in matieres:
                from .calculs_moyennes import calculer_moyenne_matiere
                result = calculer_moyenne_matiere(eleve, matiere, periode, 'mensuel')
                details_matieres[matiere.id] = result
            
            # Récupérer le rang et la moyenne
            rang_num = classement_resultat['rang_map'].get(eleve.id)
            rang_str = str(rang_num) if rang_num else "-"
            moyenne = classement_resultat['moyennes_par_eleve'].get(eleve.id)
            
            # Formatter le rang avec ex-æquo si nécessaire
            if rang_num:
                from .calculs_intelligent import formater_rang_intelligent
                sexe = getattr(eleve, 'sexe', 'M') or 'M'
                rang_str = formater_rang_intelligent(rang_num, sexe)
            
            classement_data.append({
                'matricule': eleve.matricule,
                'nom_complet': eleve.nom_complet,
                'rang': rang_str,
                'moyenne': moyenne,
                'details_matieres': details_matieres,
                'sexe': getattr(eleve, 'sexe', 'M') or 'M'
            })
        
        # Trier par rang
        classement_data.sort(key=lambda x: x['rang'] if x['rang'] != '-' else '999')
        
        # Contexte pour le template
        context = {
            'classe_selectionnee': classe,
            'periode_selectionnee': periode,
            'matieres': matieres,
            'classement_data': classement_data,
        }
        
        # Générer le HTML
        html_content = render_to_string('notes/impression_tableau_notes.html', context, request=request)
        
        # Créer le PDF avec WeasyPrint
        html = HTML(string=html_content)
        css = CSS(string='''
            @page {
                size: A4 landscape;
                margin: 10mm;
            }
        ''')
        
        pdf = html.write_pdf(stylesheets=[css])
        
        # Retourner le PDF
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="tableau_notes_{classe.nom}_{periode}.pdf"'
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'impression du tableau: {str(e)}")
        return HttpResponse(f"Erreur: {str(e)}", status=500)

@login_required
@require_school_object(model=Eleve, pk_kwarg='eleve_id', field_path='classe__ecole')
def bulletin_semestre_pdf(request, classe_id: int, eleve_id: int, semestre: int = 1):
    """Génère un bulletin semestriel (S1 ou S2) pour Collège/Lycée.
    Règle de pondération: ((Moy. composition * 2) + Moy. cours) / 3.
    """
    # Sécuriser classe/élève
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.select_related('ecole'), request.user, 'ecole'), pk=classe_id)
    eleve = get_object_or_404(filter_by_user_school(Eleve.objects.select_related('classe', 'classe__ecole'), request.user, 'classe__ecole'), pk=eleve_id, classe=classe)

    # Matières actives
    matieres = MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole, actif=True).order_by('nom')
    annee_scolaire = getattr(classe, 'annee_scolaire', None)

    lignes = []
    somme_moyennes_coef = Decimal('0'); somme_coef_matieres = Decimal('0')
    for mat in matieres:
        moy_cours = semester_course_avg(eleve, mat, annee_scolaire, semestre)
        moy_compo = semester_compo_avg(eleve, mat, annee_scolaire, semestre)
        moy_sem = semester_avg(eleve, mat, annee_scolaire, semestre, mode='weighted')
        
        # RÈGLE PÉDAGOGIQUE: Toutes les matières comptent (sans notes = 0)
        moy_sem_calcul = moy_sem if moy_sem is not None else Decimal('0')
        somme_moyennes_coef += moy_sem_calcul * Decimal(mat.coefficient or 1)
        somme_coef_matieres += Decimal(mat.coefficient or 1)
        
        lignes.append({
            'matiere': mat.nom,
            'coef_matiere': mat.coefficient,
            'moy_cours': moy_cours,
            'moy_compo': moy_compo,
            'moy_sem': moy_sem,  # Garder None pour affichage
        })

    moyenne_generale = (somme_moyennes_coef / somme_coef_matieres).quantize(Decimal('0.01')) if somme_coef_matieres > 0 else None

    # PDF
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab requis (pip install reportlab)", status=500)

    response = HttpResponse(content_type='application/pdf')
    filename = f"bulletin_semestre{semestre}_{eleve.matricule}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'

    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Filigrane
    _apply_watermark(c, width, height, ecole=getattr(classe, 'ecole', None))

    margin = 2 * cm
    y = height - margin

    # En-tête
    if getattr(classe, 'ecole', None):
        y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
    y -= 20; c.setFont('Helvetica-Bold', 14); c.drawCentredString(width/2, y, f"Bulletin semestriel — S{semestre}"); y -= 40
    c.setFont('Helvetica', 12)
    c.drawString(margin, y, f"Élève: {eleve.prenom} {eleve.nom}  (Matricule: {eleve.matricule or '-'} )"); y -= 16
    c.drawString(margin, y, f"Classe: {classe.nom} — Année: {annee_scolaire or ''}"); y -= 12
    c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16

    # Détecter le niveau scolaire pour masquer les coefficients en primaire/maternelle
    from .calculs_moyennes import detecter_niveau_scolaire
    niveau_scolaire = detecter_niveau_scolaire(classe.nom)
    est_primaire_ou_maternelle = niveau_scolaire in ['PRIMAIRE', 'MATERNELLE']
    
    # Entêtes colonnes
    c.setFont('Helvetica-Bold', 12)
    if est_primaire_ou_maternelle:
        headers = ["Matière", "Moy. cours", "Moy. compo", "Moy. S{}".format(semestre)]
        colw = [8*cm, 3.5*cm, 3.5*cm, 3.5*cm]
    else:
        headers = ["Matière", "Coef.", "Moy. cours", "Moy. compo", "Moy. S{}".format(semestre)]
        colw = [7.5*cm, 2.0*cm, 3.0*cm, 3.0*cm, 3.0*cm]
    x = margin
    for i, htxt in enumerate(headers):
        c.drawString(x, y, htxt); x += colw[i]
    y -= 14
    c.setFillColor(colors.lightgrey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 10

    c.setFont('Helvetica', 15)
    for row in lignes:
        if y < margin + 60:
            c.showPage()
            _apply_watermark(c, width, height, ecole=getattr(classe, 'ecole', None))
            y = height - margin
        x = margin
        c.drawString(x, y, str(row['matiere'])); x += colw[0]
        if not est_primaire_ou_maternelle:
            c.drawString(x, y, str(row['coef_matiere'])); x += colw[1]
        c.drawString(x, y, '-' if row['moy_cours'] is None else f"{row['moy_cours']}"); x += colw[1] if est_primaire_ou_maternelle else colw[2]
        c.drawString(x, y, '-' if row['moy_compo'] is None else f"{row['moy_compo']}"); x += colw[2] if est_primaire_ou_maternelle else colw[3]
        c.drawString(x, y, '-' if row['moy_sem'] is None else f"{row['moy_sem']}")
        y -= 14

    # Séparateur
    y -= 6; c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16

    # Moyenne générale
    c.setFont('Helvetica-Bold', 13)
    c.drawString(margin, y, f"Moyenne générale semestrielle: {moyenne_generale if moyenne_generale is not None else '-'} / 20"); y -= 18

    # Pied
    c.setFont('Helvetica-Oblique', 10); c.setFillColor(colors.darkgrey)
    c.drawString(margin, margin/2, f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    c.showPage(); c.save()
    return response


@login_required
def imprimer_tableau_notes_pdf(request):
    """Imprimer le tableau des notes avec ajustement des colonnes sur A4 landscape"""
    from django.template.loader import render_to_string
    from weasyprint import HTML, CSS
    from django.http import HttpResponse
    
    # Récupérer les paramètres
    classe_id = request.GET.get('classe_id')
    periode = request.GET.get('periode')
    
    if not classe_id or not periode:
        return HttpResponse("Paramètres manquants", status=400)
    
    try:
        # Récupérer la classe et les données
        classe_note = get_object_or_404(ClasseNote, pk=classe_id)
        classe = classe_note.classe
        
        # Récupérer les matières
        matieres = MatiereNote.objects.filter(classe=classe_note).order_by('matiere__nom')
        
        # Calculer le classement
        from .calculs_moyennes import calculer_classement_classe
        from .calculs_intelligent import calculer_rang_intelligent
        
        # Récupérer les élèves
        from eleves.models import Eleve
        eleves = Eleve.objects.filter(classe=classe, est_actif=True).order_by('matricule')
        
        # Calculer les moyennes et rangs
        classement_resultat = calculer_classement_classe(eleves, matieres, periode, 'mensuel')
        
        # Préparer les données pour le template
        classement_data = []
        for eleve in eleves:
            # Récupérer les détails des notes par matière
            details_matieres = {}
            for matiere in matieres:
                from .calculs_moyennes import calculer_moyenne_matiere
                result = calculer_moyenne_matiere(eleve, matiere, periode, 'mensuel')
                details_matieres[matiere.id] = result
            
            # Récupérer le rang et la moyenne
            rang_num = classement_resultat['rang_map'].get(eleve.id)
            rang_str = str(rang_num) if rang_num else "-"
            moyenne = classement_resultat['moyennes_par_eleve'].get(eleve.id)
            
            # Formatter le rang avec ex-æquo si nécessaire
            if rang_num:
                from .calculs_intelligent import formater_rang_intelligent
                sexe = getattr(eleve, 'sexe', 'M') or 'M'
                rang_str = formater_rang_intelligent(rang_num, sexe)
            
            classement_data.append({
                'matricule': eleve.matricule,
                'nom_complet': eleve.nom_complet,
                'rang': rang_str,
                'moyenne': moyenne,
                'details_matieres': details_matieres,
                'sexe': getattr(eleve, 'sexe', 'M') or 'M'
            })
        
        # Trier par rang
        classement_data.sort(key=lambda x: x['rang'] if x['rang'] != '-' else '999')
        
        # Contexte pour le template
        context = {
            'classe_selectionnee': classe,
            'periode_selectionnee': periode,
            'matieres': matieres,
            'classement_data': classement_data,
        }
        
        # Générer le HTML
        html_content = render_to_string('notes/impression_tableau_notes.html', context, request=request)
        
        # Créer le PDF avec WeasyPrint
        html = HTML(string=html_content)
        css = CSS(string='''
            @page {
                size: A4 landscape;
                margin: 10mm;
            }
        ''')
        
        pdf = html.write_pdf(stylesheets=[css])
        
        # Retourner le PDF
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="tableau_notes_{classe.nom}_{periode}.pdf"'
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'impression du tableau: {str(e)}")
        return HttpResponse(f"Erreur: {str(e)}", status=500)

@admin_required
def bulletins_classe_pdf(request, classe_id: int, trimestre: str = "T1"):
    """Génère en un seul PDF les bulletins de tous les élèves d'une classe pour un trimestre."""
    # Sécuriser la classe
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.select_related('ecole'), request.user, 'ecole'), pk=classe_id)
    
    # Trouver la ClasseNote correspondante
    from notes.models import ClasseNote
    try:
        # Essayer de trouver la ClasseNote correspondante
        classe_note = ClasseNote.objects.filter(
            nom=classe.nom,
            annee_scolaire=classe.annee_scolaire,
            ecole=classe.ecole
        ).first()
    except Exception:
        classe_note = None

    if not classe_note:
        classe_note = ClasseNote.objects.filter(ecole=classe.ecole).order_by('nom').first()

    if not classe_note:
        return HttpResponse("Classe de notes non trouvée", status=404)

    # Convertir trimestre format court vers format attendu par bulletins_classe_pdf du système intelligent
    trimestre_mapping = {
        'T1': 'TRIMESTRE_1',
        'T2': 'TRIMESTRE_2',
        'T3': 'TRIMESTRE_3',
        'S1': 'SEMESTRE_1',
        'S2': 'SEMESTRE_2'
    }
    periode = trimestre_mapping.get(trimestre, trimestre)

    # Appeler la fonction intelligente qui gère tout (photos, totaux, etc)
    from .bulletin_intelligent import bulletins_classe_pdf as gen_bulletins_classe_intelligent
    return gen_bulletins_classe_intelligent(request, classe_note.id, periode)


@admin_required
def export_notes_excel(request, classe_id: int, matiere_id: int, trimestre: str = "T1"):
    """Export Excel des notes d'une classe pour une matière et un trimestre.
    Colonnes: Matricule, Élève, [colonnes de chaque évaluation], Moyenne matière.
    """
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.all(), request.user, 'ecole'), pk=classe_id)
    matiere = get_object_or_404(MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole), pk=matiere_id)

    # Évaluations du trimestre pour cette matière
    evaluations = list(Evaluation.objects.filter(classe=classe, matiere=matiere, trimestre=trimestre).order_by('date', 'id'))
    # Élèves
    eleves = Eleve.objects.filter(classe=classe).order_by('prenom', 'nom')
    eleves = filter_by_user_school(eleves, request.user, 'classe__ecole')

    # Notes indexées par (eleve_id, evaluation_id)
    notes = Note.objects.filter(evaluation__in=evaluations, eleve__in=eleves)
    notes_map = {(n.eleve_id, n.evaluation_id): n for n in notes}

    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except Exception:
        return HttpResponse("openpyxl requis (pip install openpyxl)", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{matiere.nom} {trimestre}"

    # En-tête
    headers = ["Matricule", "Élève"] + [ev.titre or f"Eval {i+1}" for i, ev in enumerate(evaluations)] + ["Moyenne /20"]
    ws.append(headers)

    # Lignes
    from decimal import Decimal as D
    for e in eleves:
        row = [e.matricule, f"{e.prenom} {e.nom}"]
        num = D('0'); den = D('0')
        for ev in evaluations:
            n = notes_map.get((e.id, ev.id))
            if n and n.note is not None:
                row.append(float(n.note))
                c = D(ev.coefficient or 1)
                num += D(n.note) * c
                den += c
            else:
                row.append(None)
        moy = float((num/den)) if den > 0 else None
        row.append(moy)
        ws.append(row)

    # Styles simples: largeur colonnes
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18 if col <= 2 else 12

    # Réponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"notes_{classe.nom}_{matiere.nom}_{trimestre}.xlsx".replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def imprimer_tableau_notes_pdf(request):
    """Imprimer le tableau des notes avec ajustement des colonnes sur A4 landscape"""
    from django.template.loader import render_to_string
    from weasyprint import HTML, CSS
    from django.http import HttpResponse
    
    # Récupérer les paramètres
    classe_id = request.GET.get('classe_id')
    periode = request.GET.get('periode')
    
    if not classe_id or not periode:
        return HttpResponse("Paramètres manquants", status=400)
    
    try:
        # Récupérer la classe et les données
        classe_note = get_object_or_404(ClasseNote, pk=classe_id)
        classe = classe_note.classe
        
        # Récupérer les matières
        matieres = MatiereNote.objects.filter(classe=classe_note).order_by('matiere__nom')
        
        # Calculer le classement
        from .calculs_moyennes import calculer_classement_classe
        from .calculs_intelligent import calculer_rang_intelligent
        
        # Récupérer les élèves
        from eleves.models import Eleve
        eleves = Eleve.objects.filter(classe=classe, est_actif=True).order_by('matricule')
        
        # Calculer les moyennes et rangs
        classement_resultat = calculer_classement_classe(eleves, matieres, periode, 'mensuel')
        
        # Préparer les données pour le template
        classement_data = []
        for eleve in eleves:
            # Récupérer les détails des notes par matière
            details_matieres = {}
            for matiere in matieres:
                from .calculs_moyennes import calculer_moyenne_matiere
                result = calculer_moyenne_matiere(eleve, matiere, periode, 'mensuel')
                details_matieres[matiere.id] = result
            
            # Récupérer le rang et la moyenne
            rang_num = classement_resultat['rang_map'].get(eleve.id)
            rang_str = str(rang_num) if rang_num else "-"
            moyenne = classement_resultat['moyennes_par_eleve'].get(eleve.id)
            
            # Formatter le rang avec ex-æquo si nécessaire
            if rang_num:
                from .calculs_intelligent import formater_rang_intelligent
                sexe = getattr(eleve, 'sexe', 'M') or 'M'
                rang_str = formater_rang_intelligent(rang_num, sexe)
            
            classement_data.append({
                'matricule': eleve.matricule,
                'nom_complet': eleve.nom_complet,
                'rang': rang_str,
                'moyenne': moyenne,
                'details_matieres': details_matieres,
                'sexe': getattr(eleve, 'sexe', 'M') or 'M'
            })
        
        # Trier par rang
        classement_data.sort(key=lambda x: x['rang'] if x['rang'] != '-' else '999')
        
        # Contexte pour le template
        context = {
            'classe_selectionnee': classe,
            'periode_selectionnee': periode,
            'matieres': matieres,
            'classement_data': classement_data,
        }
        
        # Générer le HTML
        html_content = render_to_string('notes/impression_tableau_notes.html', context, request=request)
        
        # Créer le PDF avec WeasyPrint
        html = HTML(string=html_content)
        css = CSS(string='''
            @page {
                size: A4 landscape;
                margin: 10mm;
            }
        ''')
        
        pdf = html.write_pdf(stylesheets=[css])
        
        # Retourner le PDF
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="tableau_notes_{classe.nom}_{periode}.pdf"'
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'impression du tableau: {str(e)}")
        return HttpResponse(f"Erreur: {str(e)}", status=500)

def _moyenne_generale_semestrielle(eleve, matieres, annee_scolaire, semestre: int) -> Decimal | None:
    s_num = Decimal('0'); s_den = Decimal('0')
    for mat in matieres:
        m = semester_avg(eleve, mat, annee_scolaire, semestre, mode='weighted')
        if m is not None:
            s_num += m * Decimal(mat.coefficient or 1)
            s_den += Decimal(mat.coefficient or 1)
    return (s_num / s_den).quantize(Decimal('0.01')) if s_den > 0 else None

@admin_required
def export_admis_semestre_excel(request, classe_id: int, semestre: int = 1):
    """Export Excel de la liste des admis (moyenne générale semestrielle >= 10) pour une classe."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.all(), request.user, 'ecole'), pk=classe_id)
    eleves = filter_by_user_school(Eleve.objects.filter(classe=classe).order_by('prenom', 'nom'), request.user, 'classe__ecole')
    matieres = list(MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole, actif=True))
    annee_scolaire = getattr(classe, 'annee_scolaire', None)

    results = []  # (eleve, moyenne)
    for e in eleves:
        mg = _moyenne_generale_semestrielle(e, matieres, annee_scolaire, semestre)
        if mg is not None and mg >= Decimal('10'):
            results.append((e, mg))
    # Trier par moyenne desc, puis nom
    results.sort(key=lambda t: (-(t[1]), t[0].nom, t[0].prenom))

    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except Exception:
        return HttpResponse("openpyxl requis (pip install openpyxl)", status=500)

    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = f"Admis_S{semestre}"
    ws.append(["Rang", "Matricule", "Élève", f"Moyenne S{semestre}"])
    for idx, (e, avg) in enumerate(results, start=1):
        ws.append([idx, e.matricule, f"{e.prenom} {e.nom}", float(avg)])
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 18 if col != 4 else 14

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"admis_S{semestre}_{classe.nom}.xlsx".replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def imprimer_tableau_notes_pdf(request):
    """Imprimer le tableau des notes avec ajustement des colonnes sur A4 landscape"""
    from django.template.loader import render_to_string
    from weasyprint import HTML, CSS
    from django.http import HttpResponse
    
    # Récupérer les paramètres
    classe_id = request.GET.get('classe_id')
    periode = request.GET.get('periode')
    
    if not classe_id or not periode:
        return HttpResponse("Paramètres manquants", status=400)
    
    try:
        # Récupérer la classe et les données
        classe_note = get_object_or_404(ClasseNote, pk=classe_id)
        classe = classe_note.classe
        
        # Récupérer les matières
        matieres = MatiereNote.objects.filter(classe=classe_note).order_by('matiere__nom')
        
        # Calculer le classement
        from .calculs_moyennes import calculer_classement_classe
        from .calculs_intelligent import calculer_rang_intelligent
        
        # Récupérer les élèves
        from eleves.models import Eleve
        eleves = Eleve.objects.filter(classe=classe, est_actif=True).order_by('matricule')
        
        # Calculer les moyennes et rangs
        classement_resultat = calculer_classement_classe(eleves, matieres, periode, 'mensuel')
        
        # Préparer les données pour le template
        classement_data = []
        for eleve in eleves:
            # Récupérer les détails des notes par matière
            details_matieres = {}
            for matiere in matieres:
                from .calculs_moyennes import calculer_moyenne_matiere
                result = calculer_moyenne_matiere(eleve, matiere, periode, 'mensuel')
                details_matieres[matiere.id] = result
            
            # Récupérer le rang et la moyenne
            rang_num = classement_resultat['rang_map'].get(eleve.id)
            rang_str = str(rang_num) if rang_num else "-"
            moyenne = classement_resultat['moyennes_par_eleve'].get(eleve.id)
            
            # Formatter le rang avec ex-æquo si nécessaire
            if rang_num:
                from .calculs_intelligent import formater_rang_intelligent
                sexe = getattr(eleve, 'sexe', 'M') or 'M'
                rang_str = formater_rang_intelligent(rang_num, sexe)
            
            classement_data.append({
                'matricule': eleve.matricule,
                'nom_complet': eleve.nom_complet,
                'rang': rang_str,
                'moyenne': moyenne,
                'details_matieres': details_matieres,
                'sexe': getattr(eleve, 'sexe', 'M') or 'M'
            })
        
        # Trier par rang
        classement_data.sort(key=lambda x: x['rang'] if x['rang'] != '-' else '999')
        
        # Contexte pour le template
        context = {
            'classe_selectionnee': classe,
            'periode_selectionnee': periode,
            'matieres': matieres,
            'classement_data': classement_data,
        }
        
        # Générer le HTML
        html_content = render_to_string('notes/impression_tableau_notes.html', context, request=request)
        
        # Créer le PDF avec WeasyPrint
        html = HTML(string=html_content)
        css = CSS(string='''
            @page {
                size: A4 landscape;
                margin: 10mm;
            }
        ''')
        
        pdf = html.write_pdf(stylesheets=[css])
        
        # Retourner le PDF
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="tableau_notes_{classe.nom}_{periode}.pdf"'
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'impression du tableau: {str(e)}")
        return HttpResponse(f"Erreur: {str(e)}", status=500)

@admin_required
def export_admis_semestre_pdf(request, classe_id: int, semestre: int = 1):
    """Export PDF de la liste des admis (moyenne générale semestrielle >= 10) pour une classe."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.select_related('ecole'), request.user, 'ecole'), pk=classe_id)
    eleves = filter_by_user_school(Eleve.objects.filter(classe=classe).order_by('prenom', 'nom'), request.user, 'classe__ecole')
    matieres = list(MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole, actif=True))
    annee_scolaire = getattr(classe, 'annee_scolaire', None)

    rows = []  # (eleve, moyenne)
    for e in eleves:
        mg = _moyenne_generale_semestrielle(e, matieres, annee_scolaire, semestre)
        if mg is not None and mg >= Decimal('10'):
            rows.append((e, mg))
    rows.sort(key=lambda t: (-(t[1]), t[0].nom, t[0].prenom))

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab requis (pip install reportlab)", status=500)

    response = HttpResponse(content_type='application/pdf')
    filename = f"admis_S{semestre}_{classe.nom}.pdf".replace(' ', '_')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    margin = 2*cm
    # En-tête standard + titre
    _apply_watermark(c, width, height, ecole=getattr(classe, 'ecole', None))
    y = height - margin
    if getattr(classe, 'ecole', None):
        y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
    c.setFont('Helvetica-Bold', 16)
    c.drawCentredString(width/2, y-10, f"Liste des admis — S{semestre}"); y -= 40
    c.setFont('Helvetica', 12)
    c.drawString(margin, y, f"Classe: {classe.nom} — Année: {annee_scolaire or ''}"); y -= 16
    c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 12

    # En-têtes de colonnes
    c.setFont('Helvetica-Bold', 12)
    c.drawString(margin, y, "Rang");
    c.drawString(margin + 60, y, "Matricule");
    c.drawString(margin + 170, y, "Élève");
    c.drawRightString(width - margin, y, f"Moy. S{semestre}"); y -= 14
    c.setFillColor(colors.lightgrey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 8

    c.setFont('Helvetica', 11)
    for idx, (e, avg) in enumerate(rows, start=1):
        if y < margin + 40:
            c.showPage(); y = height - margin
            _apply_watermark(c, width, height, ecole=getattr(classe, 'ecole', None))
            if getattr(classe, 'ecole', None):
                y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
            c.setFont('Helvetica-Bold', 12)
            c.drawString(margin, y, "Rang");
            c.drawString(margin + 60, y, "Matricule");
            c.drawString(margin + 170, y, "Élève");
            c.drawRightString(width - margin, y, f"Moy. S{semestre}"); y -= 14
            c.setFillColor(colors.lightgrey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 8
            c.setFont('Helvetica', 11)
        c.drawString(margin, y, str(idx))
        c.drawString(margin + 60, y, e.matricule or '-')
        c.drawString(margin + 170, y, f"{e.prenom} {e.nom}")
        c.drawRightString(width - margin, y, f"{avg}")
        y -= 14

    c.setFont('Helvetica-Oblique', 10); c.setFillColor(colors.darkgrey)
    c.drawString(margin, margin/2, f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c.showPage(); c.save()
    return response


@login_required
def imprimer_tableau_notes_pdf(request):
    """Imprimer le tableau des notes avec ajustement des colonnes sur A4 landscape"""
    from django.template.loader import render_to_string
    from weasyprint import HTML, CSS
    from django.http import HttpResponse
    
    # Récupérer les paramètres
    classe_id = request.GET.get('classe_id')
    periode = request.GET.get('periode')
    
    if not classe_id or not periode:
        return HttpResponse("Paramètres manquants", status=400)
    
    try:
        # Récupérer la classe et les données
        classe_note = get_object_or_404(ClasseNote, pk=classe_id)
        classe = classe_note.classe
        
        # Récupérer les matières
        matieres = MatiereNote.objects.filter(classe=classe_note).order_by('matiere__nom')
        
        # Calculer le classement
        from .calculs_moyennes import calculer_classement_classe
        from .calculs_intelligent import calculer_rang_intelligent
        
        # Récupérer les élèves
        from eleves.models import Eleve
        eleves = Eleve.objects.filter(classe=classe, est_actif=True).order_by('matricule')
        
        # Calculer les moyennes et rangs
        classement_resultat = calculer_classement_classe(eleves, matieres, periode, 'mensuel')
        
        # Préparer les données pour le template
        classement_data = []
        for eleve in eleves:
            # Récupérer les détails des notes par matière
            details_matieres = {}
            for matiere in matieres:
                from .calculs_moyennes import calculer_moyenne_matiere
                result = calculer_moyenne_matiere(eleve, matiere, periode, 'mensuel')
                details_matieres[matiere.id] = result
            
            # Récupérer le rang et la moyenne
            rang_num = classement_resultat['rang_map'].get(eleve.id)
            rang_str = str(rang_num) if rang_num else "-"
            moyenne = classement_resultat['moyennes_par_eleve'].get(eleve.id)
            
            # Formatter le rang avec ex-æquo si nécessaire
            if rang_num:
                from .calculs_intelligent import formater_rang_intelligent
                sexe = getattr(eleve, 'sexe', 'M') or 'M'
                rang_str = formater_rang_intelligent(rang_num, sexe)
            
            classement_data.append({
                'matricule': eleve.matricule,
                'nom_complet': eleve.nom_complet,
                'rang': rang_str,
                'moyenne': moyenne,
                'details_matieres': details_matieres,
                'sexe': getattr(eleve, 'sexe', 'M') or 'M'
            })
        
        # Trier par rang
        classement_data.sort(key=lambda x: x['rang'] if x['rang'] != '-' else '999')
        
        # Contexte pour le template
        context = {
            'classe_selectionnee': classe,
            'periode_selectionnee': periode,
            'matieres': matieres,
            'classement_data': classement_data,
        }
        
        # Générer le HTML
        html_content = render_to_string('notes/impression_tableau_notes.html', context, request=request)
        
        # Créer le PDF avec WeasyPrint
        html = HTML(string=html_content)
        css = CSS(string='''
            @page {
                size: A4 landscape;
                margin: 10mm;
            }
        ''')
        
        pdf = html.write_pdf(stylesheets=[css])
        
        # Retourner le PDF
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="tableau_notes_{classe.nom}_{periode}.pdf"'
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'impression du tableau: {str(e)}")
        return HttpResponse(f"Erreur: {str(e)}", status=500)
def _collect_evals_all_trimestres(classe, matieres):
    """Retourne un dict {matiere_id: [evaluations sur T1+T2+T3]} triées par date."""
    evals_by_matiere = {}
    for mat in matieres:
        evals_by_matiere[mat.id] = list(Evaluation.objects.filter(classe=classe, matiere=mat, trimestre__in=["T1", "T2", "T3"]).order_by('date', 'id'))
    return evals_by_matiere

@login_required
@require_school_object(model=Eleve, pk_kwarg='eleve_id', field_path='classe__ecole')
def bulletin_annuel_pdf(request, classe_id: int, eleve_id: int):
    """Bulletin annuel PDF (T1+T2+T3 cumulés) avec moyennes par matière, moyenne générale, rang, mention, signatures."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.select_related('ecole'), request.user, 'ecole'), pk=classe_id)
    eleve = get_object_or_404(filter_by_user_school(Eleve.objects.select_related('classe', 'classe__ecole'), request.user, 'classe__ecole'), pk=eleve_id, classe=classe)

    # Trouver la ClasseNote correspondante pour MatiereNote
    from notes.models import ClasseNote
    
    # Mapping spécial pour certaines classes (ClasseEleve ID → ClasseNote ID)
    mapping_inverse = {
        8: 59,   # ClasseEleve '11ème série littéraire' → ClasseNote '11ème Série littéraire'
        56: 61,  # ClasseEleve '12ÈME ANNÉE' → ClasseNote '12ème Année'
    }
    
    try:
        # Essayer le mapping spécial d'abord
        if classe.id in mapping_inverse:
            classe_note = ClasseNote.objects.filter(id=mapping_inverse[classe.id]).first()
        else:
            # Sinon chercher par nom
            classe_note = ClasseNote.objects.filter(
                nom=classe.nom,
                annee_scolaire=classe.annee_scolaire,
                ecole=classe.ecole
            ).first()
        
        if classe_note:
            matieres = list(MatiereNote.objects.filter(classe=classe_note, actif=True).order_by('nom'))
        else:
            matieres = []
    except Exception:
        matieres = []
    
    evals_by_matiere = _collect_evals_all_trimestres(classe, matieres)

    # Calculs élève
    lignes = []
    somme_moyennes_coef = Decimal('0'); somme_coef_matieres = Decimal('0')
    # Filtrer les notes par matières et par toutes les périodes trimestrielles
    notes_by_eval = {n.evaluation_id: n for n in NoteEleve.objects.filter(eleve=eleve, evaluation__matiere__in=[m.id for m in matieres], evaluation__periode__in=['TRIMESTRE_1', 'TRIMESTRE_2', 'TRIMESTRE_3'])}
    for mat in matieres:
        evals = evals_by_matiere.get(mat.id, [])
        num = Decimal('0'); den = Decimal('0')
        for ev in evals:
            n = notes_by_eval.get(ev.id)
            if not n or n.note is None:
                continue
            cc = Decimal(ev.coefficient or 1)
            num += Decimal(n.note) * cc
            den += cc
        moy_mat = (num/den).quantize(Decimal('0.01')) if den > 0 else None
        # RÈGLE PÉDAGOGIQUE: Toutes les matières comptent (sans notes = 0)
        if moy_mat is None:
            moy_mat = Decimal('0')
        somme_moyennes_coef += moy_mat * Decimal(mat.coefficient or 1)
        somme_coef_matieres += Decimal(mat.coefficient or 1)
        lignes.append({'matiere': mat.nom, 'coef_matiere': mat.coefficient, 'moyenne': moy_mat})
    moyenne_generale = (somme_moyennes_coef / somme_coef_matieres).quantize(Decimal('0.01')) if somme_coef_matieres > 0 else None

    # Moyennes de classe par matière
    moyennes_classe_par_matiere: dict[int, Decimal | None] = {}
    for mat in matieres:
        evals = evals_by_matiere.get(mat.id, [])
        tnum = Decimal('0'); tden = Decimal('0')
        for ev in evals:
            for n in Note.objects.filter(evaluation=ev).only('note'):
                if n.note is None:
                    continue
                cc = Decimal(ev.coefficient or 1)
                tnum += Decimal(n.note) * cc
                tden += cc
        moyennes_classe_par_matiere[mat.id] = (tnum/tden).quantize(Decimal('0.01')) if tden > 0 else None

    # IMPORTANT: Utiliser la source centralisée pour garantir la cohérence
    # entre bulletins PDF, bulletins affichés et classements
    from .utils_rangs import calculer_rangs_classe_periode
    
    rang = None
    classement = []
    
    if classe_note:
        # Utiliser ANNUEL_TRIM pour le bulletin annuel basé sur les trimestres
        rangs_dict = calculer_rangs_classe_periode(classe_note, 'ANNUEL_TRIM', use_cache=False)
        classement = list(rangs_dict.keys())
        
        rang_info = rangs_dict.get(eleve.id)
        if rang_info:
            rang = rang_info.get('rang_num')
            # Utiliser la moyenne de la source centralisée pour cohérence
            moyenne_generale = rang_info.get('moyenne', moyenne_generale)

    def mention_for(avg: Decimal | None) -> str:
        if avg is None:
            return ""
        if avg >= Decimal('18'): return "Excellent"
        if avg >= Decimal('16'): return "Très Bien"
        if avg >= Decimal('14'): return "Bien"
        if avg >= Decimal('12'): return "Assez Bien"
        if avg >= Decimal('10'): return "Passable"
        if avg >= Decimal('8'): return "Insuffisant"
        if avg >= Decimal('6'): return "Faible"
        return "Très faible"
    mention = mention_for(moyenne_generale)

    # PDF
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab requis (pip install reportlab)", status=500)

    response = HttpResponse(content_type='application/pdf')
    filename = f"bulletin_annuel_{eleve.matricule}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    _apply_watermark(c, width, height)
    margin = 2*cm; y = height - margin
    if getattr(classe, 'ecole', None):
        y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
    y -= 20; c.setFont('Helvetica-Bold', 14); c.drawCentredString(width/2, y, "Bulletin de notes — Annuel"); y -= 40
    c.setFont('Helvetica', 12); c.drawString(margin, y, f"Élève: {eleve.prenom} {eleve.nom} (Matricule: {eleve.matricule or '-'})"); y -= 16
    c.drawString(margin, y, f"Classe: {classe.nom} — Année: {getattr(classe, 'annee_scolaire', '')}"); y -= 12
    c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16

    # Détecter le niveau scolaire pour masquer les coefficients en primaire/maternelle
    from .calculs_moyennes import detecter_niveau_scolaire
    niveau_scolaire = detecter_niveau_scolaire(classe.nom)
    est_primaire_ou_maternelle = niveau_scolaire in ['PRIMAIRE', 'MATERNELLE']
    
    c.setFont('Helvetica-Bold', 12)
    if est_primaire_ou_maternelle:
        headers = ["Matière", "Moy. annuelle", "Moy. classe"]
        colw = [9*cm, 4*cm, 4*cm]
    else:
        headers = ["Matière", "Coef.", "Moy. annuelle", "Moy. classe"]
        colw = [8*cm, 2.2*cm, 3.2*cm, 3.2*cm]
    x = margin
    for i, h in enumerate(headers): c.drawString(x, y, h); x += colw[i]
    y -= 14; c.setFillColor(colors.lightgrey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 10
    c.setFont('Helvetica', 15)
    for row in lignes:
        if y < margin + 60:
            c.showPage();
            _apply_watermark(c, width, height)
            y = height - margin
        x = margin
        c.drawString(x, y, row['matiere']); x += colw[0]
        if not est_primaire_ou_maternelle:
            c.drawString(x, y, str(row['coef_matiere'])); x += colw[1]
        c.drawString(x, y, '-' if row['moyenne'] is None else f"{row['moyenne']}"); x += colw[1] if est_primaire_ou_maternelle else colw[2]
        mc = moyennes_classe_par_matiere.get(next((m.id for m in matieres if m.nom == row['matiere']), None), None)
        c.drawString(x, y, '-' if mc is None else f"{mc}")
        y -= 14

    y -= 6
    c.setFillColor(colors.grey)
    c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
    c.setFillColor(colors.black)
    y -= 16
    c.setFont('Helvetica-Bold', 13)
    c.drawString(margin, y, f"Moyenne générale annuelle: {moyenne_generale if moyenne_generale is not None else '-'} / 20")
    y -= 16
    if rang is not None:
        c.setFont('Helvetica', 12)
        c.drawString(margin, y, f"Rang annuel: {rang} / {len(classement)}")
        y -= 14
    men = mention
    if men:
        c.setFont('Helvetica', 12)
        c.drawString(margin, y, f"Mention: {men}")
        y -= 16
    # Signatures
    c.setFont('Helvetica', 11); sig_y = margin + 50
    c.drawString(margin, sig_y, "Professeur principal:"); c.line(margin + 150, sig_y-2, margin + 320, sig_y-2)
    c.drawString(margin + 350, sig_y, "Chef d’établissement:"); c.line(margin + 520, sig_y-2, margin + 700, sig_y-2)
    c.drawString(margin, sig_y - 28, "Parent/Tuteur:"); c.line(margin + 150, sig_y-30, margin + 320, sig_y-30)
    from datetime import datetime
    c.setFont('Helvetica-Oblique', 10); c.setFillColor(colors.darkgrey); c.drawString(margin, margin/2, f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    c.showPage(); c.save(); return response


@login_required
def imprimer_tableau_notes_pdf(request):
    """Imprimer le tableau des notes avec ajustement des colonnes sur A4 landscape"""
    from django.template.loader import render_to_string
    from weasyprint import HTML, CSS
    from django.http import HttpResponse
    
    # Récupérer les paramètres
    classe_id = request.GET.get('classe_id')
    periode = request.GET.get('periode')
    
    if not classe_id or not periode:
        return HttpResponse("Paramètres manquants", status=400)
    
    try:
        # Récupérer la classe et les données
        classe_note = get_object_or_404(ClasseNote, pk=classe_id)
        classe = classe_note.classe
        
        # Récupérer les matières
        matieres = MatiereNote.objects.filter(classe=classe_note).order_by('matiere__nom')
        
        # Calculer le classement
        from .calculs_moyennes import calculer_classement_classe
        from .calculs_intelligent import calculer_rang_intelligent
        
        # Récupérer les élèves
        from eleves.models import Eleve
        eleves = Eleve.objects.filter(classe=classe, est_actif=True).order_by('matricule')
        
        # Calculer les moyennes et rangs
        classement_resultat = calculer_classement_classe(eleves, matieres, periode, 'mensuel')
        
        # Préparer les données pour le template
        classement_data = []
        for eleve in eleves:
            # Récupérer les détails des notes par matière
            details_matieres = {}
            for matiere in matieres:
                from .calculs_moyennes import calculer_moyenne_matiere
                result = calculer_moyenne_matiere(eleve, matiere, periode, 'mensuel')
                details_matieres[matiere.id] = result
            
            # Récupérer le rang et la moyenne
            rang_num = classement_resultat['rang_map'].get(eleve.id)
            rang_str = str(rang_num) if rang_num else "-"
            moyenne = classement_resultat['moyennes_par_eleve'].get(eleve.id)
            
            # Formatter le rang avec ex-æquo si nécessaire
            if rang_num:
                from .calculs_intelligent import formater_rang_intelligent
                sexe = getattr(eleve, 'sexe', 'M') or 'M'
                rang_str = formater_rang_intelligent(rang_num, sexe)
            
            classement_data.append({
                'matricule': eleve.matricule,
                'nom_complet': eleve.nom_complet,
                'rang': rang_str,
                'moyenne': moyenne,
                'details_matieres': details_matieres,
                'sexe': getattr(eleve, 'sexe', 'M') or 'M'
            })
        
        # Trier par rang
        classement_data.sort(key=lambda x: x['rang'] if x['rang'] != '-' else '999')
        
        # Contexte pour le template
        context = {
            'classe_selectionnee': classe,
            'periode_selectionnee': periode,
            'matieres': matieres,
            'classement_data': classement_data,
        }
        
        # Générer le HTML
        html_content = render_to_string('notes/impression_tableau_notes.html', context, request=request)
        
        # Créer le PDF avec WeasyPrint
        html = HTML(string=html_content)
        css = CSS(string='''
            @page {
                size: A4 landscape;
                margin: 10mm;
            }
        ''')
        
        pdf = html.write_pdf(stylesheets=[css])
        
        # Retourner le PDF
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="tableau_notes_{classe.nom}_{periode}.pdf"'
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'impression du tableau: {str(e)}")
        return HttpResponse(f"Erreur: {str(e)}", status=500)

@admin_required
def bulletins_annuels_classe_pdf(request, classe_id: int):
    """Bulletins annuels (T1+T2+T3) pour tous les élèves d'une classe en un seul PDF."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.select_related('ecole'), request.user, 'ecole'), pk=classe_id)
    eleves = filter_by_user_school(Eleve.objects.filter(classe=classe).order_by('nom','prenom'), request.user, 'classe__ecole')
    matieres = list(MatiereClasse.objects.filter(classe=classe, ecole=classe.ecole, actif=True).order_by('nom'))
    evals_by_matiere = _collect_evals_all_trimestres(classe, matieres)
    
    # Détecter le niveau scolaire pour masquer les coefficients en primaire/maternelle
    from .calculs_moyennes import detecter_niveau_scolaire
    niveau_scolaire = detecter_niveau_scolaire(classe.nom)
    est_primaire_ou_maternelle = niveau_scolaire in ['PRIMAIRE', 'MATERNELLE']

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab requis (pip install reportlab)", status=500)

    response = HttpResponse(content_type='application/pdf')
    filename = f"bulletins_annuels_{classe.nom}.pdf".replace(' ','_')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Pré-calcul moyennes classe par matière
    moyennes_classe_par_matiere: dict[int, Decimal | None] = {}
    for mat in matieres:
        evals = evals_by_matiere.get(mat.id, [])
        tnum = Decimal('0'); tden = Decimal('0')
        for ev in evals:
            for n in Note.objects.filter(evaluation=ev).only('note'):
                if n.note is None:
                    continue
                cc = Decimal(ev.coefficient or 1)
                tnum += Decimal(n.note) * cc
                tden += cc
        moyennes_classe_par_matiere[mat.id] = (tnum/tden).quantize(Decimal('0.01')) if tden > 0 else None

    # Classement annuel
    moyenne_generale_map: dict[int, Decimal] = {}
    for e in eleves:
        # Filtrer les notes par matières et par toutes les périodes trimestrielles
        notes_e = {n.evaluation_id: n for n in NoteEleve.objects.filter(eleve=e, evaluation__matiere__in=[m.id for m in matieres], evaluation__periode__in=['TRIMESTRE_1', 'TRIMESTRE_2', 'TRIMESTRE_3'])}
        s_num = Decimal('0'); s_den = Decimal('0')
        for mat in matieres:
            evals = evals_by_matiere.get(mat.id, [])
            num = Decimal('0'); den = Decimal('0')
            for ev in evals:
                nn = notes_e.get(ev.id)
                if not nn or nn.note is None:
                    continue
                cc = Decimal(ev.coefficient or 1)
                num += Decimal(nn.note) * cc
                den += cc
            if den > 0:
                moy_mat_e = (num/den)
                s_num += moy_mat_e * Decimal(mat.coefficient or 1)
                s_den += Decimal(mat.coefficient or 1)
        if s_den > 0:
            moyenne_generale_map[e.id] = (s_num / s_den).quantize(Decimal('0.01'))
    # Classement (tri par moyenne puis par matricule pour stabiliser les ex-æquo)
    # Récupérer les matricules pour le tri secondaire
    matricules_map = {e.id: e.matricule for e in eleves}
    classement = sorted(
        moyenne_generale_map.items(), 
        key=lambda t: (-float(t[1]), matricules_map.get(t[0], ""))
    )
    rang_map: dict[int, int] = {eid: idx for idx, (eid, _) in enumerate(classement, start=1)}
    # Trier les élèves par classement annuel (1er au dernier)
    eleves = sorted(eleves, key=lambda e: rang_map.get(e.id, 9999))

    def mention_for(avg: Decimal | None) -> str:
        if avg is None: return ""
        if avg >= Decimal('18'): return "Excellent"
        if avg >= Decimal('16'): return "Très Bien"
        if avg >= Decimal('14'): return "Bien"
        if avg >= Decimal('12'): return "Assez Bien"
        if avg >= Decimal('10'): return "Passable"
        if avg >= Decimal('8'): return "Insuffisant"
        if avg >= Decimal('6'): return "Faible"
        return "Très faible"

    def draw_for_student(eleve):
        _apply_watermark(c, width, height)
        margin = 2*cm; y = height - margin
        if getattr(classe, 'ecole', None):
            y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
        y -= 20; c.setFont('Helvetica-Bold', 14); c.drawCentredString(width/2, y, "Bulletin de notes — Annuel"); y -= 40
        c.setFont('Helvetica', 12); c.drawString(margin, y, f"Élève: {eleve.prenom} {eleve.nom} (Matricule: {eleve.matricule or '-'})"); y -= 16
        c.drawString(margin, y, f"Classe: {classe.nom} — Année: {getattr(classe, 'annee_scolaire', '')}"); y -= 12
        c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16

        c.setFont('Helvetica-Bold', 12)
        if est_primaire_ou_maternelle:
            headers = ["Matière", "Moy. annuelle", "Moy. classe"]
            colw = [9*cm, 4*cm, 4*cm]
        else:
            headers = ["Matière", "Coef.", "Moy. annuelle", "Moy. classe"]
            colw = [8*cm, 2.2*cm, 3.2*cm, 3.2*cm]
        x = margin
        for i, h in enumerate(headers): c.drawString(x, y, h); x += colw[i]
        y -= 14; c.setFillColor(colors.lightgrey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 10
        c.setFont('Helvetica', 15)

        # Lignes
        lignes = []
        somme_moy_coef = Decimal('0'); somme_coef = Decimal('0')
        # Filtrer les notes par matières et par toutes les périodes trimestrielles
        notes_by_eval = {n.evaluation_id: n for n in NoteEleve.objects.filter(eleve=eleve, evaluation__matiere__in=[m.id for m in matieres], evaluation__periode__in=['TRIMESTRE_1', 'TRIMESTRE_2', 'TRIMESTRE_3'])}
        for mat in matieres:
            evals = evals_by_matiere.get(mat.id, [])
            num = Decimal('0'); den = Decimal('0')
            for ev in evals:
                nn = notes_by_eval.get(ev.id)
                if not nn or nn.note is None:
                    continue
                cc = Decimal(ev.coefficient or 1)
                num += Decimal(nn.note) * cc
                den += cc
            moy_mat = (num/den).quantize(Decimal('0.01')) if den > 0 else None
            if moy_mat is not None:
                somme_moy_coef += moy_mat * Decimal(mat.coefficient or 1)
                somme_coef += Decimal(mat.coefficient or 1)
            lignes.append({'matiere': mat.nom, 'coef_matiere': mat.coefficient, 'moyenne': moy_mat})

        for row in lignes:
            if y < margin + 60:
                c.showPage();
                _apply_watermark(c, width, height)
                y = height - margin
            x = margin
            c.drawString(x, y, row['matiere']); x += colw[0]
            if not est_primaire_ou_maternelle:
                c.drawString(x, y, str(row['coef_matiere'])); x += colw[1]
            c.drawString(x, y, '-' if row['moyenne'] is None else f"{row['moyenne']}"); x += colw[1] if est_primaire_ou_maternelle else colw[2]
            mc = moyennes_classe_par_matiere.get(next((m.id for m in matieres if m.nom == row['matiere']), None), None)
            c.drawString(x, y, '-' if mc is None else f"{mc}")
            y -= 14

        y -= 6; c.setFillColor(colors.grey); c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0); c.setFillColor(colors.black); y -= 16
        mg = (somme_moy_coef / somme_coef).quantize(Decimal('0.01')) if somme_coef > 0 else None
        c.setFont('Helvetica-Bold', 13); c.drawString(margin, y, f"Moyenne générale annuelle: {mg if mg is not None else '-'} / 20"); y -= 16
        rg = rang_map.get(eleve.id)
        if rg is not None: c.setFont('Helvetica', 12); c.drawString(margin, y, f"Rang annuel: {rg} / {len(classement)}"); y -= 14
        men = mention_for(mg)
        if men: c.setFont('Helvetica', 12); c.drawString(margin, y, f"Mention: {men}"); y -= 16
        # Signatures
        c.setFont('Helvetica', 11); sig_y = margin + 50
        c.drawString(margin, sig_y, "Professeur principal:"); c.line(margin + 150, sig_y-2, margin + 320, sig_y-2)
        c.drawString(margin + 350, sig_y, "Chef d’établissement:"); c.line(margin + 520, sig_y-2, margin + 700, sig_y-2)
        c.drawString(margin, sig_y - 28, "Parent/Tuteur:"); c.line(margin + 150, sig_y-30, margin + 320, sig_y-30)
        c.showPage()

    for e in eleves:
        draw_for_student(e)

    c.save(); return response


@login_required
def imprimer_tableau_notes_pdf(request):
    """Imprimer le tableau des notes avec ajustement des colonnes sur A4 landscape"""
    from django.template.loader import render_to_string
    from weasyprint import HTML, CSS
    from django.http import HttpResponse
    
    # Récupérer les paramètres
    classe_id = request.GET.get('classe_id')
    periode = request.GET.get('periode')
    
    if not classe_id or not periode:
        return HttpResponse("Paramètres manquants", status=400)
    
    try:
        # Récupérer la classe et les données
        classe_note = get_object_or_404(ClasseNote, pk=classe_id)
        classe = classe_note.classe
        
        # Récupérer les matières
        matieres = MatiereNote.objects.filter(classe=classe_note).order_by('matiere__nom')
        
        # Calculer le classement
        from .calculs_moyennes import calculer_classement_classe
        from .calculs_intelligent import calculer_rang_intelligent
        
        # Récupérer les élèves
        from eleves.models import Eleve
        eleves = Eleve.objects.filter(classe=classe, est_actif=True).order_by('matricule')
        
        # Calculer les moyennes et rangs
        classement_resultat = calculer_classement_classe(eleves, matieres, periode, 'mensuel')
        
        # Préparer les données pour le template
        classement_data = []
        for eleve in eleves:
            # Récupérer les détails des notes par matière
            details_matieres = {}
            for matiere in matieres:
                from .calculs_moyennes import calculer_moyenne_matiere
                result = calculer_moyenne_matiere(eleve, matiere, periode, 'mensuel')
                details_matieres[matiere.id] = result
            
            # Récupérer le rang et la moyenne
            rang_num = classement_resultat['rang_map'].get(eleve.id)
            rang_str = str(rang_num) if rang_num else "-"
            moyenne = classement_resultat['moyennes_par_eleve'].get(eleve.id)
            
            # Formatter le rang avec ex-æquo si nécessaire
            if rang_num:
                from .calculs_intelligent import formater_rang_intelligent
                sexe = getattr(eleve, 'sexe', 'M') or 'M'
                rang_str = formater_rang_intelligent(rang_num, sexe)
            
            classement_data.append({
                'matricule': eleve.matricule,
                'nom_complet': eleve.nom_complet,
                'rang': rang_str,
                'moyenne': moyenne,
                'details_matieres': details_matieres,
                'sexe': getattr(eleve, 'sexe', 'M') or 'M'
            })
        
        # Trier par rang
        classement_data.sort(key=lambda x: x['rang'] if x['rang'] != '-' else '999')
        
        # Contexte pour le template
        context = {
            'classe_selectionnee': classe,
            'periode_selectionnee': periode,
            'matieres': matieres,
            'classement_data': classement_data,
        }
        
        # Générer le HTML
        html_content = render_to_string('notes/impression_tableau_notes.html', context, request=request)
        
        # Créer le PDF avec WeasyPrint
        html = HTML(string=html_content)
        css = CSS(string='''
            @page {
                size: A4 landscape;
                margin: 10mm;
            }
        ''')
        
        pdf = html.write_pdf(stylesheets=[css])
        
        # Retourner le PDF
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="tableau_notes_{classe.nom}_{periode}.pdf"'
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'impression du tableau: {str(e)}")
        return HttpResponse(f"Erreur: {str(e)}", status=500)

@login_required
@require_school_object(model=ClasseEleve, pk_kwarg='classe_id', field_path='ecole')
def classement_classe(request, classe_id: int, trimestre: str = "T1"):
    """Affiche le classement des élèves d'une classe pour un trimestre donné."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.all(), request.user, 'ecole'), pk=classe_id)
    
    # Calculer le classement
    eleves = classe.eleves.filter(statut='actif').order_by('prenom', 'nom')
    classement = []
    
    for eleve in eleves:
        notes = Note.objects.filter(
            eleve=eleve,
            evaluation__classe=classe,
            evaluation__trimestre=trimestre
        ).select_related('evaluation__matiere', 'evaluation')
        
        if not notes.exists():
            continue
            
        somme_moy_coef = Decimal('0')
        somme_coef = Decimal('0')
        
        # Regrouper les notes par matière avec pondération par coefficient d'évaluation
        matieres_notes = {}
        for note_obj in notes:
            matiere = note_obj.evaluation.matiere
            if matiere.id not in matieres_notes:
                matieres_notes[matiere.id] = {
                    'matiere': matiere.nom,
                    'coef_matiere': matiere.coefficient,
                    'notes_ponderees': []  # (note, coef_eval)
                }
            # Stocker la note avec son coefficient d'évaluation (absence = 0)
            coef_eval = Decimal(note_obj.evaluation.coefficient or 1)
            if note_obj.absent or note_obj.note is None:
                # Absence = 0
                matieres_notes[matiere.id]['notes_ponderees'].append(
                    (Decimal('0'), coef_eval)
                )
            else:
                matieres_notes[matiere.id]['notes_ponderees'].append(
                    (Decimal(note_obj.note), coef_eval)
                )
        
        # Calculer la moyenne pondérée par matière, puis la moyenne générale
        for matiere_data in matieres_notes.values():
            if matiere_data['notes_ponderees']:
                # Moyenne pondérée par coefficient d'évaluation
                num = Decimal('0')
                den = Decimal('0')
                for note_val, coef_eval in matiere_data['notes_ponderees']:
                    num += note_val * coef_eval
                    den += coef_eval
                if den > 0:
                    moyenne_matiere = (num / den).quantize(Decimal('0.01'))
                    # Pondération par coefficient de matière
                    somme_moy_coef += moyenne_matiere * Decimal(matiere_data['coef_matiere'])
                    somme_coef += Decimal(matiere_data['coef_matiere'])
        
        if somme_coef > 0:
            moyenne_generale = (somme_moy_coef / somme_coef).quantize(Decimal('0.01'))
            classement.append({
                'eleve': eleve,
                'moyenne': moyenne_generale,
                'mention': mention_for(moyenne_generale)
            })
    
    # Trier par moyenne décroissante
    classement.sort(key=lambda x: x['moyenne'], reverse=True)
    
    # Ajouter les rangs
    for i, item in enumerate(classement):
        item['rang'] = i + 1
    
    # Calculer la moyenne de classe
    moyenne_classe = None
    if classement:
        total_moyennes = sum(item['moyenne'] for item in classement)
        moyenne_classe = round(total_moyennes / len(classement), 2)
    
    context = {
        'classe': classe,
        'trimestre': trimestre,
        'classement': classement,
        'total_eleves': len(classement),
        'moyenne_classe': moyenne_classe
    }
    
    return render(request, 'notes/classement_classe.html', context)

@admin_required
def classement_classe_pdf(request, classe_id: int, trimestre: str = "T1"):
    """Export PDF du classement d'une classe."""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.all(), request.user, 'ecole'), pk=classe_id)
    
    # Récupérer le classement (même logique que la vue HTML)
    eleves = classe.eleves.filter(statut='actif').order_by('prenom', 'nom')
    classement = []
    
    for eleve in eleves:
        notes = Note.objects.filter(
            eleve=eleve,
            evaluation__classe=classe,
            evaluation__trimestre=trimestre
        ).select_related('evaluation__matiere', 'evaluation')
        
        if not notes.exists():
            continue
            
        somme_moy_coef = Decimal('0')
        somme_coef = Decimal('0')
        
        # Regrouper les notes par matière avec pondération par coefficient d'évaluation
        matieres_notes = {}
        for note_obj in notes:
            matiere = note_obj.evaluation.matiere
            if matiere.id not in matieres_notes:
                matieres_notes[matiere.id] = {
                    'matiere': matiere.nom,
                    'coef_matiere': matiere.coefficient,
                    'notes_ponderees': []  # (note, coef_eval)
                }
            # Stocker la note avec son coefficient d'évaluation (absence = 0)
            coef_eval = Decimal(note_obj.evaluation.coefficient or 1)
            if note_obj.absent or note_obj.note is None:
                # Absence = 0
                matieres_notes[matiere.id]['notes_ponderees'].append(
                    (Decimal('0'), coef_eval)
                )
            else:
                matieres_notes[matiere.id]['notes_ponderees'].append(
                    (Decimal(note_obj.note), coef_eval)
                )
        
        # Calculer la moyenne pondérée par matière, puis la moyenne générale
        for matiere_data in matieres_notes.values():
            if matiere_data['notes_ponderees']:
                # Moyenne pondérée par coefficient d'évaluation
                num = Decimal('0')
                den = Decimal('0')
                for note_val, coef_eval in matiere_data['notes_ponderees']:
                    num += note_val * coef_eval
                    den += coef_eval
                if den > 0:
                    moyenne_matiere = (num / den).quantize(Decimal('0.01'))
                    # Pondération par coefficient de matière
                    somme_moy_coef += moyenne_matiere * Decimal(matiere_data['coef_matiere'])
                    somme_coef += Decimal(matiere_data['coef_matiere'])
        
        if somme_coef > 0:
            moyenne_generale = (somme_moy_coef / somme_coef).quantize(Decimal('0.01'))
            classement.append({
                'eleve': eleve,
                'moyenne': moyenne_generale,
                'mention': mention_for(moyenne_generale)
            })
    
    # Trier par moyenne décroissante
    classement.sort(key=lambda x: x['moyenne'], reverse=True)
    
    # Créer le PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="classement_{classe.nom}_{trimestre}.pdf"'
    
    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    
    _apply_watermark(c, width, height)
    
    margin = 2 * cm
    y = height - margin
    
    # En-tête avec logo et coordonnées
    if getattr(classe, 'ecole', None):
        y = _draw_school_header(c, classe.ecole, y_start=y, margin=margin, page_width=width)
    
    y -= 20
    c.setFont('Helvetica-Bold', 16)
    c.drawCentredString(width/2, y, f"Classement de la classe {classe.nom} - {trimestre}")
    y -= 40
    
    c.setFont('Helvetica', 12)
    c.drawString(margin, y, f"Année scolaire: {getattr(classe, 'annee_scolaire', '')}")
    y -= 16
    c.drawString(margin, y, f"Total d'élèves classés: {len(classement)}")
    y -= 20
    
    # En-têtes du tableau
    c.setFont('Helvetica-Bold', 12)
    headers = ["Rang", "Nom et Prénom", "Matricule", "Moyenne", "Mention"]
    colw = [2*cm, 6*cm, 3*cm, 2.5*cm, 3*cm]
    x = margin
    for i, h in enumerate(headers):
        c.drawString(x, y, h)
        x += colw[i]
    
    y -= 14
    c.setFillColor(colors.lightgrey)
    c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
    c.setFillColor(colors.black)
    y -= 10
    
    # Données du classement
    c.setFont('Helvetica', 11)
    for i, item in enumerate(classement):
        if y < margin + 60:
            c.showPage()
            _apply_watermark(c, width, height)
            y = height - margin
        
        x = margin
        c.drawString(x, y, str(i + 1))  # Rang
        x += colw[0]
        c.drawString(x, y, f"{item['eleve'].prenom} {item['eleve'].nom}")  # Nom
        x += colw[1]
        c.drawString(x, y, item['eleve'].matricule or '-')  # Matricule
        x += colw[2]
        c.drawString(x, y, f"{item['moyenne']}")  # Moyenne
        x += colw[3]
        c.drawString(x, y, item['mention'] or '-')  # Mention
        
        y -= 14
    
    # Pied de page
    c.setFont('Helvetica-Oblique', 10)
    c.setFillColor(colors.darkgrey)
    from datetime import datetime
    c.drawString(margin, margin/2, f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    
    c.save()
    return response


@login_required
def imprimer_tableau_notes_pdf(request):
    """Imprimer le tableau des notes avec ajustement des colonnes sur A4 landscape"""
    from django.template.loader import render_to_string
    from weasyprint import HTML, CSS
    from django.http import HttpResponse
    
    # Récupérer les paramètres
    classe_id = request.GET.get('classe_id')
    periode = request.GET.get('periode')
    
    if not classe_id or not periode:
        return HttpResponse("Paramètres manquants", status=400)
    
    try:
        # Récupérer la classe et les données
        classe_note = get_object_or_404(ClasseNote, pk=classe_id)
        classe = classe_note.classe
        
        # Récupérer les matières
        matieres = MatiereNote.objects.filter(classe=classe_note).order_by('matiere__nom')
        
        # Calculer le classement
        from .calculs_moyennes import calculer_classement_classe
        from .calculs_intelligent import calculer_rang_intelligent
        
        # Récupérer les élèves
        from eleves.models import Eleve
        eleves = Eleve.objects.filter(classe=classe, est_actif=True).order_by('matricule')
        
        # Calculer les moyennes et rangs
        classement_resultat = calculer_classement_classe(eleves, matieres, periode, 'mensuel')
        
        # Préparer les données pour le template
        classement_data = []
        for eleve in eleves:
            # Récupérer les détails des notes par matière
            details_matieres = {}
            for matiere in matieres:
                from .calculs_moyennes import calculer_moyenne_matiere
                result = calculer_moyenne_matiere(eleve, matiere, periode, 'mensuel')
                details_matieres[matiere.id] = result
            
            # Récupérer le rang et la moyenne
            rang_num = classement_resultat['rang_map'].get(eleve.id)
            rang_str = str(rang_num) if rang_num else "-"
            moyenne = classement_resultat['moyennes_par_eleve'].get(eleve.id)
            
            # Formatter le rang avec ex-æquo si nécessaire
            if rang_num:
                from .calculs_intelligent import formater_rang_intelligent
                sexe = getattr(eleve, 'sexe', 'M') or 'M'
                rang_str = formater_rang_intelligent(rang_num, sexe)
            
            classement_data.append({
                'matricule': eleve.matricule,
                'nom_complet': eleve.nom_complet,
                'rang': rang_str,
                'moyenne': moyenne,
                'details_matieres': details_matieres,
                'sexe': getattr(eleve, 'sexe', 'M') or 'M'
            })
        
        # Trier par rang
        classement_data.sort(key=lambda x: x['rang'] if x['rang'] != '-' else '999')
        
        # Contexte pour le template
        context = {
            'classe_selectionnee': classe,
            'periode_selectionnee': periode,
            'matieres': matieres,
            'classement_data': classement_data,
        }
        
        # Générer le HTML
        html_content = render_to_string('notes/impression_tableau_notes.html', context, request=request)
        
        # Créer le PDF avec WeasyPrint
        html = HTML(string=html_content)
        css = CSS(string='''
            @page {
                size: A4 landscape;
                margin: 10mm;
            }
        ''')
        
        pdf = html.write_pdf(stylesheets=[css])
        
        # Retourner le PDF
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="tableau_notes_{classe.nom}_{periode}.pdf"'
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'impression du tableau: {str(e)}")
        return HttpResponse(f"Erreur: {str(e)}", status=500)

@admin_required
def classement_classe_excel(request, classe_id: int, trimestre: str = "T1"):
    """Export Excel du classement d'une classe."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.all(), request.user, 'ecole'), pk=classe_id)
    
    # Récupérer le classement (même logique que les autres vues)
    eleves = classe.eleves.filter(statut='actif').order_by('prenom', 'nom')
    classement = []
    
    for eleve in eleves:
        notes = Note.objects.filter(
            eleve=eleve,
            evaluation__classe=classe,
            evaluation__trimestre=trimestre
        ).select_related('evaluation__matiere', 'evaluation')
        
        if not notes.exists():
            continue
            
        somme_moy_coef = Decimal('0')
        somme_coef = Decimal('0')
        
        # Regrouper les notes par matière avec pondération par coefficient d'évaluation
        matieres_notes = {}
        for note_obj in notes:
            matiere = note_obj.evaluation.matiere
            if matiere.id not in matieres_notes:
                matieres_notes[matiere.id] = {
                    'matiere': matiere.nom,
                    'coef_matiere': matiere.coefficient,
                    'notes_ponderees': []  # (note, coef_eval)
                }
            # Stocker la note avec son coefficient d'évaluation (absence = 0)
            coef_eval = Decimal(note_obj.evaluation.coefficient or 1)
            if note_obj.absent or note_obj.note is None:
                # Absence = 0
                matieres_notes[matiere.id]['notes_ponderees'].append(
                    (Decimal('0'), coef_eval)
                )
            else:
                matieres_notes[matiere.id]['notes_ponderees'].append(
                    (Decimal(note_obj.note), coef_eval)
                )
        
        # Calculer la moyenne pondérée par matière, puis la moyenne générale
        for matiere_data in matieres_notes.values():
            if matiere_data['notes_ponderees']:
                # Moyenne pondérée par coefficient d'évaluation
                num = Decimal('0')
                den = Decimal('0')
                for note_val, coef_eval in matiere_data['notes_ponderees']:
                    num += note_val * coef_eval
                    den += coef_eval
                if den > 0:
                    moyenne_matiere = (num / den).quantize(Decimal('0.01'))
                    # Pondération par coefficient de matière
                    somme_moy_coef += moyenne_matiere * Decimal(matiere_data['coef_matiere'])
                    somme_coef += Decimal(matiere_data['coef_matiere'])
        
        if somme_coef > 0:
            moyenne_generale = (somme_moy_coef / somme_coef).quantize(Decimal('0.01'))
            classement.append({
                'eleve': eleve,
                'moyenne': moyenne_generale,
                'mention': mention_for(moyenne_generale)
            })
    
    # Trier par moyenne décroissante
    classement.sort(key=lambda x: x['moyenne'], reverse=True)
    
    # Créer le fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Classement {classe.nom} {trimestre}"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center")
    
    # En-têtes
    headers = ["Rang", "Nom", "Prénom", "Matricule", "Moyenne", "Mention"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Données
    for row, item in enumerate(classement, 2):
        ws.cell(row=row, column=1, value=row-1)  # Rang
        ws.cell(row=row, column=2, value=item['eleve'].nom)
        ws.cell(row=row, column=3, value=item['eleve'].prenom)
        ws.cell(row=row, column=4, value=item['eleve'].matricule or '-')
        ws.cell(row=row, column=5, value=float(item['moyenne']))
        ws.cell(row=row, column=6, value=item['mention'] or '-')
    
    # Ajuster la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="classement_{classe.nom}_{trimestre}.xlsx"'
    
    wb.save(response)
    return response


@login_required
def imprimer_tableau_notes_pdf(request):
    """Imprimer le tableau des notes avec ajustement des colonnes sur A4 landscape"""
    from django.template.loader import render_to_string
    from weasyprint import HTML, CSS
    from django.http import HttpResponse
    
    # Récupérer les paramètres
    classe_id = request.GET.get('classe_id')
    periode = request.GET.get('periode')
    
    if not classe_id or not periode:
        return HttpResponse("Paramètres manquants", status=400)
    
    try:
        # Récupérer la classe et les données
        classe_note = get_object_or_404(ClasseNote, pk=classe_id)
        classe = classe_note.classe
        
        # Récupérer les matières
        matieres = MatiereNote.objects.filter(classe=classe_note).order_by('matiere__nom')
        
        # Calculer le classement
        from .calculs_moyennes import calculer_classement_classe
        from .calculs_intelligent import calculer_rang_intelligent
        
        # Récupérer les élèves
        from eleves.models import Eleve
        eleves = Eleve.objects.filter(classe=classe, est_actif=True).order_by('matricule')
        
        # Calculer les moyennes et rangs
        classement_resultat = calculer_classement_classe(eleves, matieres, periode, 'mensuel')
        
        # Préparer les données pour le template
        classement_data = []
        for eleve in eleves:
            # Récupérer les détails des notes par matière
            details_matieres = {}
            for matiere in matieres:
                from .calculs_moyennes import calculer_moyenne_matiere
                result = calculer_moyenne_matiere(eleve, matiere, periode, 'mensuel')
                details_matieres[matiere.id] = result
            
            # Récupérer le rang et la moyenne
            rang_num = classement_resultat['rang_map'].get(eleve.id)
            rang_str = str(rang_num) if rang_num else "-"
            moyenne = classement_resultat['moyennes_par_eleve'].get(eleve.id)
            
            # Formatter le rang avec ex-æquo si nécessaire
            if rang_num:
                from .calculs_intelligent import formater_rang_intelligent
                sexe = getattr(eleve, 'sexe', 'M') or 'M'
                rang_str = formater_rang_intelligent(rang_num, sexe)
            
            classement_data.append({
                'matricule': eleve.matricule,
                'nom_complet': eleve.nom_complet,
                'rang': rang_str,
                'moyenne': moyenne,
                'details_matieres': details_matieres,
                'sexe': getattr(eleve, 'sexe', 'M') or 'M'
            })
        
        # Trier par rang
        classement_data.sort(key=lambda x: x['rang'] if x['rang'] != '-' else '999')
        
        # Contexte pour le template
        context = {
            'classe_selectionnee': classe,
            'periode_selectionnee': periode,
            'matieres': matieres,
            'classement_data': classement_data,
        }
        
        # Générer le HTML
        html_content = render_to_string('notes/impression_tableau_notes.html', context, request=request)
        
        # Créer le PDF avec WeasyPrint
        html = HTML(string=html_content)
        css = CSS(string='''
            @page {
                size: A4 landscape;
                margin: 10mm;
            }
        ''')
        
        pdf = html.write_pdf(stylesheets=[css])
        
        # Retourner le PDF
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="tableau_notes_{classe.nom}_{periode}.pdf"'
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'impression du tableau: {str(e)}")
        return HttpResponse(f"Erreur: {str(e)}", status=500)

@login_required
@require_school_object(model=ClasseEleve, pk_kwarg='classe_id', field_path='ecole')
def cartes_scolaires_classe(request, classe_id):
    """Interface pour générer les cartes scolaires d'une classe"""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.all(), request.user, 'ecole'), pk=classe_id)
    
    # Filtrage par école pour non-admin
    if not request.user.is_superuser:
        ecole_user = user_school(request.user)
        if ecole_user and classe.ecole != ecole_user:
            messages.error(request, "Accès non autorisé à cette classe.")
            return redirect('notes:tableau_bord')
    
    # Récupérer tous les élèves de la classe
    eleves = classe.eleves.filter(statut='ACTIF').order_by('prenom', 'nom')
    
    context = {
        'classe': classe,
        'eleves': eleves,
        'nb_eleves': eleves.count(),
    }
    
    return render(request, 'notes/cartes_scolaires.html', context)

@login_required
@require_school_object(model=ClasseEleve, pk_kwarg='classe_id', field_path='ecole')
def cartes_scolaires_pdf(request, classe_id):
    """Génère les cartes scolaires PDF pour une classe"""
    classe = get_object_or_404(filter_by_user_school(ClasseEleve.objects.all(), request.user, 'ecole'), pk=classe_id)
    
    # Filtrage par école pour non-admin
    if not request.user.is_superuser:
        ecole_user = user_school(request.user)
        if ecole_user and classe.ecole != ecole_user:
            messages.error(request, "Accès non autorisé à cette classe.")
            return redirect('notes:tableau_bord')
    
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        import io
    except ImportError:
        messages.error(request, "ReportLab requis pour générer le PDF")
        return redirect('notes:cartes_scolaires_classe', classe_id=classe_id)
    
    # Récupérer les élèves
    eleves = classe.eleves.filter(statut='ACTIF').order_by('prenom', 'nom')
    
    if not eleves.exists():
        messages.warning(request, "Aucun élève actif dans cette classe.")
        return redirect('notes:cartes_scolaires_classe', classe_id=classe_id)
    
    # Créer le PDF
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Filigrane standardisé (logo centré, rotation légère, opacité 4%)
    _apply_watermark(c, width, height, ecole=getattr(classe, 'ecole', None))
    
    # Configuration des cartes (10 cartes par page - 2x5)
    card_width = 8.5 * cm
    card_height = 5.4 * cm  # Format carte de crédit
    margin = 0.3 * cm
    spacing_x = 0.2 * cm  # Espacement horizontal réduit
    spacing_y = 0.15 * cm  # Espacement vertical réduit
    
    # Position des cartes sur la page (2 colonnes x 5 lignes)
    positions = [
        # Colonne gauche
        (margin, height - margin - card_height),  # Ligne 1
        (margin, height - margin - 2 * card_height - spacing_y),  # Ligne 2
        (margin, height - margin - 3 * card_height - 2 * spacing_y),  # Ligne 3
        (margin, height - margin - 4 * card_height - 3 * spacing_y),  # Ligne 4
        (margin, height - margin - 5 * card_height - 4 * spacing_y),  # Ligne 5
        # Colonne droite
        (margin + card_width + spacing_x, height - margin - card_height),  # Ligne 1
        (margin + card_width + spacing_x, height - margin - 2 * card_height - spacing_y),  # Ligne 2
        (margin + card_width + spacing_x, height - margin - 3 * card_height - 2 * spacing_y),  # Ligne 3
        (margin + card_width + spacing_x, height - margin - 4 * card_height - 3 * spacing_y),  # Ligne 4
        (margin + card_width + spacing_x, height - margin - 5 * card_height - 4 * spacing_y),  # Ligne 5
    ]
    
    card_count = 0
    
    for eleve in eleves:
        # Nouvelle page si nécessaire
        if card_count > 0 and card_count % 10 == 0:
            c.showPage()
            # Filigrane standardisé sur nouvelle page
            _apply_watermark(c, width, height)
        
        # Position de la carte actuelle
        pos_x, pos_y = positions[card_count % 10]
        
        # Dessiner l'arrière-plan blanc de la carte d'abord
        c.setFillColor(colors.white)
        c.rect(pos_x, pos_y, card_width, card_height, fill=1, stroke=0)
        
        # Ajouter un filigrane sur chaque carte individuelle
        c.saveState()
        try:
            from django.contrib.staticfiles import finders
            import os
            from django.conf import settings
            # Logo d'école prioritaire
            school_logo = getattr(getattr(classe, 'ecole', None), 'logo', None)
            logo_path = None
            try:
                if school_logo and hasattr(school_logo, 'path') and os.path.exists(school_logo.path):
                    logo_path = school_logo.path
            except Exception:
                logo_path = None
            if not logo_path:
                # Fallback: logo statique
                logo_path = finders.find('logos/logo.png') or os.path.join(settings.BASE_DIR, 'static', 'logos', 'logo.png')
            
            if logo_path and os.path.exists(logo_path):
                # Calculer la position centrale de la carte
                center_x = pos_x + card_width / 2
                center_y = pos_y + card_height / 2
                
                # Appliquer la transformation (rotation et opacité)
                c.translate(center_x, center_y)
                c.rotate(30)  # Rotation de 30 degrés
                
                # Taille du filigrane (plus petit pour s'adapter à la carte)
                watermark_size = min(card_width, card_height) * 0.6
                
                # Dessiner le logo en filigrane avec opacité réduite
                c.setFillAlpha(0.08)  # Opacité très faible
                c.drawImage(logo_path, -watermark_size/2, -watermark_size/2, 
                          watermark_size, watermark_size, preserveAspectRatio=True)
        except Exception:
            pass
        c.restoreState()
        
        # Dessiner le cadre de la carte
        c.setStrokeColor(colors.black)
        c.setLineWidth(2)
        c.rect(pos_x, pos_y, card_width, card_height, fill=0, stroke=1)
        
        # Logo de l'école (en haut à droite)
        logo_size = 1.2 * cm
        logo_x = pos_x + card_width - logo_size - 0.2 * cm
        logo_y = pos_y + card_height - logo_size - 0.1 * cm
        
        try:
            import os
            from django.conf import settings
            from django.contrib.staticfiles import finders
            # Priorité au logo de l'école
            school_logo = getattr(getattr(classe, 'ecole', None), 'logo', None)
            logo_path = None
            try:
                if school_logo and hasattr(school_logo, 'path') and os.path.exists(school_logo.path):
                    logo_path = school_logo.path
            except Exception:
                logo_path = None
            if not logo_path:
                logo_path = finders.find('logos/logo.png') or os.path.join(settings.BASE_DIR, 'static', 'logos', 'logo.png')
            if logo_path and os.path.exists(logo_path):
                c.drawImage(logo_path, logo_x, logo_y, logo_size, logo_size, preserveAspectRatio=True)
            else:
                # Aucun logo trouvé
                c.setFont('Helvetica', 6)
                c.setFillColor(colors.grey)
                c.drawCentredString(logo_x + logo_size/2, logo_y + logo_size/2, "LOGO")
        except Exception as e:
            # En cas d'erreur, afficher simplement le texte d'erreur
            c.setFont('Helvetica', 6)
            c.setFillColor(colors.red)
            c.drawCentredString(logo_x + logo_size/2, logo_y + logo_size/2, "ERREUR")
        
        # En-tête école (centré sur toute la largeur de la carte)
        c.setFillColor(colors.darkblue)
        ecole_nom = classe.ecole.nom.upper() if classe.ecole else "ÉCOLE"
        # Ajuster la taille de police selon la longueur du nom
        if len(ecole_nom) > 40:
            c.setFont('Helvetica-Bold', 6)
        elif len(ecole_nom) > 30:
            c.setFont('Helvetica-Bold', 7)
        else:
            c.setFont('Helvetica-Bold', 8)
        # Centrer le nom au milieu de la carte (meilleur équilibre visuel)
        c.drawCentredString(pos_x + card_width/2, pos_y + card_height - 0.8*cm, ecole_nom)
        
        # Titre "CARTE SCOLAIRE"
        c.setFont('Helvetica-Bold', 8)
        c.setFillColor(colors.red)
        c.drawCentredString(pos_x + card_width/2, pos_y + card_height - 1.3*cm, "CARTE SCOLAIRE")
        
        # Année scolaire
        annee_actuelle = datetime.now().year
        annee_scolaire = f"{annee_actuelle}-{annee_actuelle + 1}"
        c.setFont('Helvetica', 7)
        c.setFillColor(colors.black)
        c.drawCentredString(pos_x + card_width/2, pos_y + card_height - 1.7*cm, f"Année: {annee_scolaire}")
        
        # Photo de l'élève (côté gauche) — encore remontée pour un meilleur équilibre visuel
        photo_size = 2.2 * cm
        photo_x = pos_x + 0.3 * cm
        photo_y = pos_y + 1.2 * cm
        
        # Dessiner le cadre de la photo
        c.setStrokeColor(colors.grey)
        c.setLineWidth(1)
        c.rect(photo_x, photo_y, photo_size, photo_size)
        
        # Afficher la photo de l'élève si elle existe
        if eleve.photo and hasattr(eleve.photo, 'path'):
            try:
                import os
                from reportlab.lib.utils import ImageReader
                from PIL import Image
                
                # Vérifier que le fichier photo existe
                if os.path.exists(eleve.photo.path):
                    # Ouvrir et redimensionner l'image
                    with Image.open(eleve.photo.path) as img:
                        # Convertir en RGB si nécessaire
                        if img.mode != 'RGB':
                            img = img.convert('RGB')
                        
                        # Calculer les dimensions pour maintenir le ratio
                        img_width, img_height = img.size
                        ratio = min(photo_size / (img_width * 72/96), photo_size / (img_height * 72/96))
                        
                        new_width = img_width * ratio * 72/96
                        new_height = img_height * ratio * 72/96
                        
                        # Centrer l'image dans le cadre
                        img_x = photo_x + (photo_size - new_width) / 2
                        img_y = photo_y + (photo_size - new_height) / 2
                        
                        # Dessiner l'image
                        c.drawImage(ImageReader(img), img_x, img_y, new_width, new_height)
                else:
                    # Fichier photo introuvable
                    c.setFont('Helvetica', 7)
                    c.setFillColor(colors.red)
                    c.drawCentredString(photo_x + photo_size/2, photo_y + photo_size/2 - 0.1*cm, "PHOTO")
                    c.drawCentredString(photo_x + photo_size/2, photo_y + photo_size/2 - 0.3*cm, "MANQUANTE")
            except Exception as e:
                # Erreur lors du traitement de l'image
                c.setFont('Helvetica', 7)
                c.setFillColor(colors.red)
                c.drawCentredString(photo_x + photo_size/2, photo_y + photo_size/2 - 0.1*cm, "ERREUR")
                c.drawCentredString(photo_x + photo_size/2, photo_y + photo_size/2 - 0.3*cm, "PHOTO")
        else:
            # Pas de photo définie
            c.setFont('Helvetica', 8)
            c.setFillColor(colors.grey)
            c.drawCentredString(photo_x + photo_size/2, photo_y + photo_size/2 - 0.2*cm, "PHOTO")
        
        # Informations élève (côté droit)
        info_x = pos_x + 3.2 * cm
        info_y = pos_y + card_height - 2.2 * cm
        
        c.setFillColor(colors.black)
        
        # Nom et prénom
        c.setFont('Helvetica-Bold', 9)
        nom_complet = eleve.nom_complet.upper()
        if len(nom_complet) > 25:
            nom_complet = nom_complet[:25] + "..."
        c.drawString(info_x, info_y, nom_complet)
        info_y -= 0.4 * cm
        
        # Matricule
        c.setFont('Helvetica', 7)
        c.drawString(info_x, info_y, f"Matricule: {eleve.matricule}")
        info_y -= 0.3 * cm
        
        # Classe
        c.drawString(info_x, info_y, f"Classe: {classe.nom}")
        info_y -= 0.3 * cm
        
        # Date de naissance
        if eleve.date_naissance:
            date_naiss = eleve.date_naissance.strftime('%d/%m/%Y')
            c.drawString(info_x, info_y, f"Né(e) le: {date_naiss}")
            info_y -= 0.3 * cm
        
        # Contact responsable
        if hasattr(eleve, 'responsable_principal') and eleve.responsable_principal:
            resp = eleve.responsable_principal
            if resp.telephone:
                c.drawString(info_x, info_y, f"Contact: {resp.telephone}")
        
        # Pied de carte
        c.setFont('Helvetica', 6)
        c.setFillColor(colors.grey)
        c.drawString(pos_x + 0.2*cm, pos_y + 0.2*cm, "Cette carte est strictement personnelle")
        
        card_count += 1
    
    c.save()
    
    # Préparer la réponse
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    filename = f"cartes_scolaires_{classe.nom.replace(' ', '_')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(pdf)
    
    return response


@login_required
def imprimer_tableau_notes_pdf(request):
    """Imprimer le tableau des notes avec ajustement des colonnes sur A4 landscape"""
    from django.template.loader import render_to_string
    from weasyprint import HTML, CSS
    from django.http import HttpResponse
    
    # Récupérer les paramètres
    classe_id = request.GET.get('classe_id')
    periode = request.GET.get('periode')
    
    if not classe_id or not periode:
        return HttpResponse("Paramètres manquants", status=400)
    
    try:
        # Récupérer la classe et les données
        classe_note = get_object_or_404(ClasseNote, pk=classe_id)
        classe = classe_note.classe
        
        # Récupérer les matières
        matieres = MatiereNote.objects.filter(classe=classe_note).order_by('matiere__nom')
        
        # Calculer le classement
        from .calculs_moyennes import calculer_classement_classe
        from .calculs_intelligent import calculer_rang_intelligent
        
        # Récupérer les élèves
        from eleves.models import Eleve
        eleves = Eleve.objects.filter(classe=classe, est_actif=True).order_by('matricule')
        
        # Calculer les moyennes et rangs
        classement_resultat = calculer_classement_classe(eleves, matieres, periode, 'mensuel')
        
        # Préparer les données pour le template
        classement_data = []
        for eleve in eleves:
            # Récupérer les détails des notes par matière
            details_matieres = {}
            for matiere in matieres:
                from .calculs_moyennes import calculer_moyenne_matiere
                result = calculer_moyenne_matiere(eleve, matiere, periode, 'mensuel')
                details_matieres[matiere.id] = result
            
            # Récupérer le rang et la moyenne
            rang_num = classement_resultat['rang_map'].get(eleve.id)
            rang_str = str(rang_num) if rang_num else "-"
            moyenne = classement_resultat['moyennes_par_eleve'].get(eleve.id)
            
            # Formatter le rang avec ex-æquo si nécessaire
            if rang_num:
                from .calculs_intelligent import formater_rang_intelligent
                sexe = getattr(eleve, 'sexe', 'M') or 'M'
                rang_str = formater_rang_intelligent(rang_num, sexe)
            
            classement_data.append({
                'matricule': eleve.matricule,
                'nom_complet': eleve.nom_complet,
                'rang': rang_str,
                'moyenne': moyenne,
                'details_matieres': details_matieres,
                'sexe': getattr(eleve, 'sexe', 'M') or 'M'
            })
        
        # Trier par rang
        classement_data.sort(key=lambda x: x['rang'] if x['rang'] != '-' else '999')
        
        # Contexte pour le template
        context = {
            'classe_selectionnee': classe,
            'periode_selectionnee': periode,
            'matieres': matieres,
            'classement_data': classement_data,
        }
        
        # Générer le HTML
        html_content = render_to_string('notes/impression_tableau_notes.html', context, request=request)
        
        # Créer le PDF avec WeasyPrint
        html = HTML(string=html_content)
        css = CSS(string='''
            @page {
                size: A4 landscape;
                margin: 10mm;
            }
        ''')
        
        pdf = html.write_pdf(stylesheets=[css])
        
        # Retourner le PDF
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="tableau_notes_{classe.nom}_{periode}.pdf"'
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'impression du tableau: {str(e)}")
        return HttpResponse(f"Erreur: {str(e)}", status=500)

@login_required
def carte_eleve_pdf(request, matricule):
    """Génère la carte scolaire d'un élève spécifique par son matricule"""
    from django.contrib import messages
    from eleves.models import Eleve
    
    # Récupérer l'élève par matricule
    try:
        eleve = Eleve.objects.select_related('classe', 'classe__ecole', 'responsable_principal').get(matricule=matricule)
        
        # Vérifier les permissions (sauf pour superuser)
        if not request.user.is_superuser:
            if hasattr(request.user, 'profil') and request.user.profil.ecole:
                if eleve.classe.ecole != request.user.profil.ecole:
                    messages.error(request, "Vous n'avez pas accès à cet élève.")
                    return redirect('eleves:liste_eleves')
            else:
                messages.error(request, "Accès non autorisé.")
                return redirect('eleves:liste_eleves')
                
    except Eleve.DoesNotExist:
        messages.error(request, f"Aucun élève trouvé avec le matricule: {matricule}")
        return redirect('eleves:liste_eleves')
    
    # Créer le PDF avec une seule carte
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Filigrane standardisé (logo centré, rotation légère, opacité 4%)
    _apply_watermark(c, width, height)
    
    # Configuration de la carte (centrée sur la page) - Format carte bancaire standard
    card_width = 8.6 * cm  # 86mm
    card_height = 5.4 * cm  # 54mm
    
    # Position centrée sur la page
    pos_x = (width - card_width) / 2
    pos_y = (height - card_height) / 2
    
    # Dessiner l'arrière-plan blanc de la carte d'abord
    c.setFillColor(colors.white)
    c.rect(pos_x, pos_y, card_width, card_height, fill=1, stroke=0)
    
    # Ajouter un filigrane sur la carte individuelle
    c.saveState()
    try:
        from django.contrib.staticfiles import finders
        import os
        from django.conf import settings
        # Priorité au logo de l'école de l'élève
        school_logo = getattr(getattr(eleve.classe, 'ecole', None), 'logo', None)
        logo_path = None
        try:
            if school_logo and hasattr(school_logo, 'path') and os.path.exists(school_logo.path):
                logo_path = school_logo.path
        except Exception:
            logo_path = None
        if not logo_path:
            logo_path = finders.find('logos/logo.png') or os.path.join(settings.BASE_DIR, 'static', 'logos', 'logo.png')
        
        if logo_path and os.path.exists(logo_path):
            # Calculer la position centrale de la carte
            center_x = pos_x + card_width / 2
            center_y = pos_y + card_height / 2
            
            # Appliquer la transformation (rotation et opacité)
            c.translate(center_x, center_y)
            c.rotate(25)  # Rotation de 25 degrés
            
            # Taille du filigrane (agrandi pour être plus visible)
            watermark_size = min(card_width, card_height) * 0.7
            
            # Dessiner le logo en filigrane avec opacité visible
            c.setFillAlpha(0.12)  # Opacité visible (comme les tickets)
            c.drawImage(logo_path, -watermark_size/2, -watermark_size/2, 
                      watermark_size, watermark_size, preserveAspectRatio=True)
    except Exception:
        pass
    c.restoreState()
    
    # Extraire les couleurs du logo de l'école
    primary_color = '#10b981'
    light_color = '#d1fae5'
    
    try:
        if logo_path and os.path.exists(logo_path):
            from eleves.views import _extraire_couleurs_logo
            primary_color, light_color = _extraire_couleurs_logo(logo_path)
    except:
        pass
    
    # Design moderne avec couleurs personnalisées
    # Formes géométriques décoratives en arrière-plan
    c.setFillColor(colors.HexColor(light_color))
    c.setFillAlpha(0.15)
    c.circle(pos_x + card_width - 1 * cm, pos_y + card_height - 1 * cm, 2.5 * cm, stroke=0, fill=1)
    c.circle(pos_x + 0.8 * cm, pos_y + 0.8 * cm, 1.8 * cm, stroke=0, fill=1)
    c.setFillAlpha(1)
    
    # Bordure moderne avec coins arrondis
    c.setStrokeColor(colors.HexColor(primary_color))
    c.setLineWidth(2.5)
    c.roundRect(pos_x + 0.15 * cm, pos_y + 0.15 * cm, card_width - 0.3 * cm, card_height - 0.3 * cm, 8, stroke=1, fill=0)
    
    # Bande décorative en haut avec forme ondulée
    c.setFillColor(colors.HexColor(primary_color))
    path = c.beginPath()
    path.moveTo(pos_x + 0.2 * cm, pos_y + card_height - 0.2 * cm)
    path.lineTo(pos_x + card_width - 0.2 * cm, pos_y + card_height - 0.2 * cm)
    path.lineTo(pos_x + card_width - 0.2 * cm, pos_y + card_height - 1.2 * cm)
    # Courbe ondulée
    path.curveTo(pos_x + card_width * 0.75, pos_y + card_height - 1.15 * cm, 
                 pos_x + card_width * 0.5, pos_y + card_height - 1.25 * cm, 
                 pos_x + card_width * 0.25, pos_y + card_height - 1.15 * cm)
    path.curveTo(pos_x + card_width * 0.15, pos_y + card_height - 1.12 * cm, 
                 pos_x + 0.2 * cm, pos_y + card_height - 1.15 * cm, 
                 pos_x + 0.2 * cm, pos_y + card_height - 1.2 * cm)
    path.close()
    c.drawPath(path, fill=1, stroke=0)
    
    # Nom de l'école dans l'en-tête
    c.setFont("Helvetica-Bold", 9)
    c.setFillColor(colors.white)
    school_name = (eleve.classe.ecole.nom if eleve.classe.ecole else "École").upper()
    if len(school_name) > 35:
        school_name = school_name[:32] + "..."
    c.drawCentredString(pos_x + card_width/2, pos_y + card_height - 0.6 * cm, school_name)
    
    # Sous-titre "CARTE SCOLAIRE"
    c.setFont("Helvetica-Bold", 7)
    c.setFillAlpha(0.85)
    c.drawCentredString(pos_x + card_width/2, pos_y + card_height - 0.95 * cm, "CARTE SCOLAIRE")
    c.setFillAlpha(1)
    
    # Année scolaire (petit badge)
    c.setFillColor(colors.HexColor(primary_color))
    c.setFillAlpha(0.15)
    c.roundRect(pos_x + 0.25 * cm, pos_y + card_height - 1.25 * cm, 2 * cm, 0.35 * cm, 3, stroke=0, fill=1)
    c.setFillAlpha(1)
    c.setFont("Helvetica", 7)
    c.setFillColor(colors.HexColor(primary_color))
    current_year = timezone.now().year
    next_year = current_year + 1
    annee_scolaire = f"{current_year}-{next_year}"
    c.drawCentredString(pos_x + 1.25 * cm, pos_y + card_height - 1.15 * cm, annee_scolaire)
    
    # Photo de l'élève avec bordures arrondies (à gauche) - agrandi
    photo_width = 2.2 * cm
    photo_height = 2.2 * cm
    photo_x = pos_x + 0.3 * cm
    photo_y = pos_y + card_height/2 - photo_height/2
    
    # Afficher la photo ou placeholder
    try:
        print(f"DEBUG Carte - Élève: {eleve.prenom} {eleve.nom}")
        print(f"DEBUG Carte - Photo field: {eleve.photo}")
        print(f"DEBUG Carte - Has photo: {bool(eleve.photo)}")
        
        if eleve.photo:
            print(f"DEBUG Carte - Photo path: {eleve.photo.path if hasattr(eleve.photo, 'path') else 'NO PATH'}")
            if hasattr(eleve.photo, 'path'):
                print(f"DEBUG Carte - File exists: {os.path.exists(eleve.photo.path)}")
        
        if eleve.photo and hasattr(eleve.photo, 'path') and os.path.exists(eleve.photo.path):
            from PIL import Image, ImageDraw
            
            print(f"DEBUG Carte - Chargement photo: {eleve.photo.path}")
            # Ouvrir l'image
            img = Image.open(eleve.photo.path)
            print(f"DEBUG Carte - Image mode: {img.mode}, size: {img.size}")
            
            # Convertir en RGB si nécessaire
            if img.mode in ('RGBA', 'LA', 'P'):
                background = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                background.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
                img = background
            elif img.mode != 'RGB':
                img = img.convert('RGB')
            
            # Redimensionner l'image
            size = (int(photo_width * 28.35), int(photo_height * 28.35))  # Conversion cm vers pixels
            img = img.resize(size, Image.Resampling.LANCZOS)
            
            # Créer un masque avec coins arrondis
            mask = Image.new('L', size, 0)
            draw = ImageDraw.Draw(mask)
            # Rectangle avec coins arrondis (radius = 10% de la largeur)
            radius = int(size[0] * 0.15)
            draw.rounded_rectangle([(0, 0), size], radius=radius, fill=255)
            
            # Appliquer le masque
            output = Image.new('RGBA', size, (255, 255, 255, 0))
            output.paste(img, (0, 0))
            output.putalpha(mask)
            
            # Sauvegarder temporairement
            temp_buffer = io.BytesIO()
            output.save(temp_buffer, format='PNG')
            temp_buffer.seek(0)
            
            # Ombre portée
            c.setFillColor(colors.HexColor('#000000'))
            c.setFillAlpha(0.15)
            c.roundRect(photo_x + 0.05 * cm, photo_y - 0.05 * cm, photo_width, photo_height, 0.3 * cm, stroke=0, fill=1)
            c.setFillAlpha(1)
            
            # Dessiner la photo sur le PDF
            c.drawImage(temp_buffer, photo_x, photo_y, width=photo_width, height=photo_height, mask='auto')
            
            # Bordure colorée autour de la photo
            c.setStrokeColor(colors.HexColor(primary_color))
            c.setLineWidth(3)
            c.roundRect(photo_x, photo_y, photo_width, photo_height, 0.3 * cm, stroke=1, fill=0)
                
        else:
            # Placeholder si pas de photo
            c.setFillColor(colors.HexColor(light_color))
            c.roundRect(photo_x, photo_y, photo_width, photo_height, 0.3 * cm, stroke=0, fill=1)
            c.setStrokeColor(colors.HexColor(primary_color))
            c.setLineWidth(3)
            c.roundRect(photo_x, photo_y, photo_width, photo_height, 0.3 * cm, stroke=1, fill=0)
            c.setFont("Helvetica", 8)
            c.setFillColor(colors.HexColor(primary_color))
            c.drawCentredString(photo_x + photo_width/2, photo_y + photo_height/2 - 0.1 * cm, "PHOTO")
            
    except Exception as e:
        # Placeholder en cas d'erreur
        print(f"Erreur photo carte scolaire: {e}")
        c.setFillColor(colors.HexColor(light_color))
        c.roundRect(photo_x, photo_y, photo_width, photo_height, 0.3 * cm, stroke=0, fill=1)
        c.setStrokeColor(colors.HexColor(primary_color))
        c.setLineWidth(3)
        c.roundRect(photo_x, photo_y, photo_width, photo_height, 0.3 * cm, stroke=1, fill=0)
        c.setFont("Helvetica", 8)
        c.setFillColor(colors.HexColor(primary_color))
        c.drawCentredString(photo_x + photo_width/2, photo_y + photo_height/2 - 0.1 * cm, "PHOTO")
    
    # Zone d'information avec fond subtil (poussée vers la droite)
    info_box_x = photo_x + photo_width + 0.4 * cm
    info_box_y = pos_y + 0.4 * cm
    info_box_width = card_width - (photo_x + photo_width + 0.6 * cm - pos_x)
    info_box_height = card_height - 1.7 * cm
    
    c.setFillColor(colors.HexColor(light_color))
    c.setFillAlpha(0.08)
    c.roundRect(info_box_x, info_box_y, info_box_width, info_box_height, 6, stroke=0, fill=1)
    c.setFillAlpha(1)
    
    # Informations de l'élève (poussées vers la droite)
    info_x = info_box_x + 0.2 * cm
    info_y_start = pos_y + card_height/2 + 0.5 * cm
    
    # Nom complet (en majuscules et gras)
    c.setFont("Helvetica-Bold", 11)
    c.setFillColor(colors.HexColor('#1f2937'))
    nom_complet = eleve.nom_complet.upper()
    if len(nom_complet) > 16:
        nom_complet = nom_complet[:13] + "..."
    c.drawString(info_x, info_y_start, nom_complet)
    
    # Ligne décorative sous le nom
    c.setStrokeColor(colors.HexColor(primary_color))
    c.setLineWidth(1.5)
    c.line(info_x, info_y_start - 0.1 * cm, info_x + 3 * cm, info_y_start - 0.1 * cm)
    
    # Informations détaillées avec espacement
    y_info = info_y_start - 0.55 * cm
    c.setFillColor(colors.HexColor('#374151'))
    
    # Matricule avec espacement
    c.setFont("Helvetica-Bold", 9)
    c.drawString(info_x, y_info, "N° : ")
    c.setFont("Helvetica", 9)
    c.drawString(info_x + 0.9 * cm, y_info, eleve.matricule)
    
    y_info -= 0.45 * cm
    # Classe avec espacement
    c.setFont("Helvetica-Bold", 9)
    c.drawString(info_x, y_info, "Classe : ")
    c.setFont("Helvetica", 9)
    c.drawString(info_x + 1.3 * cm, y_info, eleve.classe.nom)
    
    y_info -= 0.45 * cm
    # Date de naissance avec espacement
    if eleve.date_naissance:
        date_naiss = eleve.date_naissance.strftime("%d/%m/%Y")
        c.setFont("Helvetica-Bold", 9)
        c.drawString(info_x, y_info, "Né(e) le : ")
        c.setFont("Helvetica", 8)
        c.drawString(info_x + 1.5 * cm, y_info, date_naiss)
    
    y_info -= 0.45 * cm
    # Téléphone responsable avec espacement
    if eleve.responsable_principal and eleve.responsable_principal.telephone:
        tel = eleve.responsable_principal.telephone[:14]
        c.setFont("Helvetica-Bold", 9)
        c.drawString(info_x, y_info, "Tél : ")
        c.setFont("Helvetica", 8)
        c.drawString(info_x + 0.9 * cm, y_info, tel)
    
    # Pied de page moderne
    c.setFillColor(colors.HexColor(primary_color))
    c.setFillAlpha(0.05)
    c.roundRect(pos_x + 0.3 * cm, pos_y + 0.15 * cm, card_width - 0.6 * cm, 0.35 * cm, 4, stroke=0, fill=1)
    c.setFillAlpha(1)
    
    c.setFont("Helvetica", 7)
    c.setFillColor(colors.HexColor('#6b7280'))
    c.drawCentredString(pos_x + card_width/2, pos_y + 0.23 * cm, 
                     "Cette carte est strictement personnelle")
    
    # Finaliser le PDF
    c.showPage()
    c.save()
    
    # Préparer la réponse
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    filename = f"carte_scolaire_{eleve.matricule}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def imprimer_tableau_notes_pdf(request):
    """Imprimer le tableau des notes avec ajustement des colonnes sur A4 landscape"""
    from django.template.loader import render_to_string
    from weasyprint import HTML, CSS
    from django.http import HttpResponse
    
    # Récupérer les paramètres
    classe_id = request.GET.get('classe_id')
    periode = request.GET.get('periode')
    
    if not classe_id or not periode:
        return HttpResponse("Paramètres manquants", status=400)
    
    try:
        # Récupérer la classe et les données
        classe_note = get_object_or_404(ClasseNote, pk=classe_id)
        classe = classe_note.classe
        
        # Récupérer les matières
        matieres = MatiereNote.objects.filter(classe=classe_note).order_by('matiere__nom')
        
        # Calculer le classement
        from .calculs_moyennes import calculer_classement_classe
        from .calculs_intelligent import calculer_rang_intelligent
        
        # Récupérer les élèves
        from eleves.models import Eleve
        eleves = Eleve.objects.filter(classe=classe, est_actif=True).order_by('matricule')
        
        # Calculer les moyennes et rangs
        classement_resultat = calculer_classement_classe(eleves, matieres, periode, 'mensuel')
        
        # Préparer les données pour le template
        classement_data = []
        for eleve in eleves:
            # Récupérer les détails des notes par matière
            details_matieres = {}
            for matiere in matieres:
                from .calculs_moyennes import calculer_moyenne_matiere
                result = calculer_moyenne_matiere(eleve, matiere, periode, 'mensuel')
                details_matieres[matiere.id] = result
            
            # Récupérer le rang et la moyenne
            rang_num = classement_resultat['rang_map'].get(eleve.id)
            rang_str = str(rang_num) if rang_num else "-"
            moyenne = classement_resultat['moyennes_par_eleve'].get(eleve.id)
            
            # Formatter le rang avec ex-æquo si nécessaire
            if rang_num:
                from .calculs_intelligent import formater_rang_intelligent
                sexe = getattr(eleve, 'sexe', 'M') or 'M'
                rang_str = formater_rang_intelligent(rang_num, sexe)
            
            classement_data.append({
                'matricule': eleve.matricule,
                'nom_complet': eleve.nom_complet,
                'rang': rang_str,
                'moyenne': moyenne,
                'details_matieres': details_matieres,
                'sexe': getattr(eleve, 'sexe', 'M') or 'M'
            })
        
        # Trier par rang
        classement_data.sort(key=lambda x: x['rang'] if x['rang'] != '-' else '999')
        
        # Contexte pour le template
        context = {
            'classe_selectionnee': classe,
            'periode_selectionnee': periode,
            'matieres': matieres,
            'classement_data': classement_data,
        }
        
        # Générer le HTML
        html_content = render_to_string('notes/impression_tableau_notes.html', context, request=request)
        
        # Créer le PDF avec WeasyPrint
        html = HTML(string=html_content)
        css = CSS(string='''
            @page {
                size: A4 landscape;
                margin: 10mm;
            }
        ''')
        
        pdf = html.write_pdf(stylesheets=[css])
        
        # Retourner le PDF
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="tableau_notes_{classe.nom}_{periode}.pdf"'
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'impression du tableau: {str(e)}")
        return HttpResponse(f"Erreur: {str(e)}", status=500)

_PERIODE_LIBELLES_VIEWS = {
    'OCTOBRE': 'Octobre', 'NOVEMBRE': 'Novembre', 'DECEMBRE': 'Décembre',
    'JANVIER': 'Janvier', 'FEVRIER': 'Février', 'MARS': 'Mars',
    'AVRIL': 'Avril', 'MAI': 'Mai', 'JUIN': 'Juin',
    'TRIMESTRE_1': '1er Trimestre', 'TRIMESTRE_2': '2ème Trimestre', 'TRIMESTRE_3': '3ème Trimestre',
    'SEMESTRE_1': '1er Semestre', 'SEMESTRE_2': '2ème Semestre',
    'ANNUEL_TRIM': 'Annuel', 'ANNUEL_SEM': 'Annuel',
}

def _libelle_periode_views(periode):
    """Convertit un code de période en libellé lisible (ex: SEMESTRE_1 → 1er Semestre)."""
    if not periode:
        return ''
    return _PERIODE_LIBELLES_VIEWS.get(str(periode).upper(), str(periode).replace('_', ' ').title())


def _generer_lettre_parent_inline(eleve, moyenne, classe_nom, periode, note_max, est_primaire):
    """Genere une lettre d'information aux parents pour un eleve en difficulte"""
    est_fille = getattr(eleve, 'sexe', 'M') == 'F'
    mot_fils = 'votre fille' if est_fille else 'votre fils'
    pronom_le = 'la' if est_fille else 'le'

    seuil_diff = 4 if est_primaire else 8
    seuil_critique = 3 if est_primaire else 6

    if moyenne < seuil_critique:
        urgence = 'URGENT'
    elif moyenne < seuil_diff:
        urgence = 'IMPORTANT'
    else:
        urgence = 'A NOTER'

    gravite = 'grande difficulte' if moyenne < seuil_diff else 'difficulte'

    return {
        'urgence': urgence,
        'objet': f"[{urgence}] Situation scolaire de {eleve.prenom} {eleve.nom}",
        'intro': f"Nous vous informons que {mot_fils} {eleve.prenom} {eleve.nom}, "
                f"eleve en classe de {classe_nom}, rencontre actuellement des difficultes scolaires importantes.",
        'constat': f"A l'issue de la periode {_libelle_periode_views(periode)}, {eleve.prenom} a obtenu une moyenne "
                  f"generale de {moyenne:.2f}/{note_max}, ce qui {pronom_le} place en situation de {gravite}.",
        'demandes': [
            "Verifier quotidiennement les devoirs et lecons",
            "Assurer un environnement calme et propice au travail a la maison",
            "Limiter les distractions (telephone, television, jeux video)",
            f"Encourager et valoriser les efforts de {mot_fils}, meme minimes",
            "Vous presenter a l'ecole pour un entretien avec l'enseignant(e)"
        ],
        'conclusion': f"Nous restons a votre disposition pour tout entretien. "
                     f"Votre implication est essentielle pour aider {mot_fils} a surmonter ces difficultes."
    }


def _generer_message_eleve_inline(eleve, moyenne, classe_nom, periode, note_max, est_primaire):
    """Genere un message d'encouragement pour un eleve en difficulte"""
    est_fille = getattr(eleve, 'sexe', 'M') == 'F'
    cher = 'Chere' if est_fille else 'Cher'

    return {
        'titre': f"Message personnel pour {eleve.prenom}",
        'intro': f"{cher} {eleve.prenom},",
        'constat': f"Tes resultats de ce {_libelle_periode_views(periode)} montrent que tu rencontres des difficultes. "
                  f"Ta moyenne de {moyenne:.2f}/{note_max} n'est pas a la hauteur de ce que tu peux accomplir.",
        'encouragements': [
            "Chaque eleve peut progresser avec de la volonte et du travail",
            "Tes difficultes actuelles ne definissent pas ton avenir",
            "Tes enseignants croient en toi et sont la pour t'aider"
        ],
        'conseils': [
            "Organise ton temps de travail avec un planning regulier",
            "N'hesite pas a poser des questions en classe quand tu ne comprends pas",
            "Revois tes lecons chaque soir, meme 15-20 minutes",
            "Travaille en groupe avec des camarades qui peuvent t'aider",
            "Fixe-toi des petits objectifs atteignables chaque semaine"
        ],
        'conclusion': "Nous sommes convaincus que tu peux t'ameliorer. "
                     "L'important est de ne pas baisser les bras et de demander de l'aide quand tu en as besoin."
    }


@login_required
def statistiques(request):
    """Statistiques globales de l'école"""
    from eleves.models import Ecole
    
    user_profil = getattr(request.user, 'profil', None)
    ecole = user_profil.ecole if user_profil else Ecole.objects.first()
    
    # Récupérer les classes disponibles
    if ecole:
        classes = ClasseNote.objects.filter(ecole=ecole, actif=True).order_by('nom')
    else:
        classes = ClasseNote.objects.filter(actif=True).order_by('nom')
    
    # Classe sélectionnée
    classe_id = request.GET.get('classe_id')
    classe_selectionnee = None
    if classe_id:
        try:
            classe_selectionnee = classes.get(id=classe_id)
        except ClasseNote.DoesNotExist:
            pass
    
    # Statistiques globales de l'école
    total_eleves = Eleve.objects.filter(statut='ACTIF').count()
    total_classes = ClasseEleve.objects.all().count()
    
    # Période sélectionnée (vide → défaut TRIMESTRE_1)
    periode = request.GET.get('periode') or 'TRIMESTRE_1'
    
    # Initialiser les statistiques
    nb_evalues = 0
    nb_non_evalues = 0
    nb_non_admis = 0
    nb_a_suivre = 0
    nb_excellents = 0
    nb_precaution = 0
    eleves_non_admis = []
    eleves_a_suivre = []
    eleves_excellents = []
    eleves_precaution = []
    recommandations = []
    
    # Si une classe est sélectionnée, calculer les statistiques
    if classe_selectionnee and periode:
        from decimal import Decimal
        
        # Récupérer les élèves de la classe
        classe_eleve = ClasseEleve.objects.filter(
            nom=classe_selectionnee.nom,
            annee_scolaire=classe_selectionnee.annee_scolaire,
            ecole=classe_selectionnee.ecole
        ).first()
        
        if classe_eleve:
            eleves = Eleve.objects.filter(
                classe=classe_eleve, statut='ACTIF'
            ).select_related('classe').order_by('prenom', 'nom')
        else:
            eleves = Eleve.objects.none()

        # Récupérer les matières de la classe
        matieres = MatiereNote.objects.filter(classe=classe_selectionnee, actif=True)
        
        # Détecter le niveau scolaire pour adapter le barème
        from .calculs_moyennes import detecter_niveau_scolaire
        niveau = detecter_niveau_scolaire(classe_selectionnee.nom)
        est_primaire = (niveau == 'PRIMAIRE')
        note_max = 10 if est_primaire else 20
        seuil_reussite = 5 if est_primaire else 10
        
        if eleves.exists() and matieres.exists():
            # ── Source unique de calcul : même fonction que le bulletin ──────────
            from .calculs_moyennes import calculer_moyennes_classe_optimise
            from .utils_rangs import calculer_rangs_classe_periode

            if 'TRIMESTRE' in (periode or ''):
                system_type = 'trimestre'
            elif 'SEMESTRE' in (periode or ''):
                system_type = 'semestre'
            else:
                system_type = 'mensuel'

            rangs_officiels = calculer_rangs_classe_periode(classe_selectionnee, periode, use_cache=False)
            resultats = calculer_moyennes_classe_optimise(eleves, matieres, periode, system_type)

            seuil_suivre   = seuil_reussite + (1 if est_primaire else 2)
            seuil_precaution = seuil_reussite + (2 if est_primaire else 4)

            for eleve in eleves:
                eleve_result = resultats.get(eleve.id)
                if not eleve_result:
                    nb_non_evalues += 1
                    continue

                rang_info = rangs_officiels.get(eleve.id)
                if not rang_info:
                    nb_non_evalues += 1
                    continue

                moyenne_generale = float(rang_info.get('moyenne') or 0)
                nb_evalues += 1
                eleve_data = {
                    'eleve': eleve,
                    'moyenne': moyenne_generale
                }

                if moyenne_generale < seuil_reussite:
                    nb_non_admis += 1
                    eleve_data['ecart'] = round(seuil_reussite - moyenne_generale, 2)
                    eleve_data['lettre_parent'] = _generer_lettre_parent_inline(eleve, moyenne_generale, classe_selectionnee.nom, periode, note_max, est_primaire)
                    eleve_data['message_eleve'] = _generer_message_eleve_inline(eleve, moyenne_generale, classe_selectionnee.nom, periode, note_max, est_primaire)
                    eleves_non_admis.append(eleve_data)
                elif moyenne_generale < seuil_suivre:
                    nb_a_suivre += 1
                    eleves_a_suivre.append(eleve_data)
                elif moyenne_generale < seuil_precaution:
                    nb_precaution += 1
                    eleves_precaution.append(eleve_data)
                else:
                    nb_excellents += 1
                    eleves_excellents.append(eleve_data)

            # Calculer les taux
            total_eleves_classe = eleves.count()
            # Construire stats par matière depuis les résultats canoniques
            stats_matieres_notes = {}
            for _eid, _eres in resultats.items():
                for _det in _eres.get('details_matieres', []):
                    _mat = _det.get('matiere')
                    if _mat is None:
                        continue
                    _moy = _det.get('moyenne')
                    if _moy is not None and _moy > 0:
                        stats_matieres_notes.setdefault(_mat.id, []).append(_moy)
            taux_reussite = round((nb_evalues - nb_non_admis) / nb_evalues * 100, 1) if nb_evalues > 0 else 0
            taux_echec = round(nb_non_admis / nb_evalues * 100, 1) if nb_evalues > 0 else 0
            
            # Générer des recommandations
            if nb_non_admis > 0:
                recommandations.append({
                    'type': 'DANGER',
                    'message': f'{nb_non_admis} élève(s) en difficulté (moyenne < {seuil_reussite}/{note_max}). Mise en place de soutien scolaire recommandée.',
                    'couleur': 'danger'
                })
            
            if nb_a_suivre > 0:
                recommandations.append({
                    'type': 'WARNING',
                    'message': f'{nb_a_suivre} élève(s) à suivre (moyenne entre {seuil_reussite} et {seuil_reussite + (1 if est_primaire else 2)}/{note_max}). Accompagnement personnalisé conseillé.',
                    'couleur': 'warning'
                })
            
            if nb_excellents > 0:
                recommandations.append({
                    'type': 'SUCCESS',
                    'message': f'{nb_excellents} élève(s) excellent(s) (moyenne ≥ {seuil_reussite + (2 if est_primaire else 4)}/{note_max}). Félicitations !',
                    'couleur': 'success'
                })
            
            if nb_non_evalues > 0:
                recommandations.append({
                    'type': 'INFO',
                    'message': f'{nb_non_evalues} élève(s) non évalué(s) pour cette période.',
                    'couleur': 'info'
                })
            
            if not recommandations:
                recommandations.append({
                    'type': 'INFO',
                    'message': 'Statistiques calculées avec succès.',
                    'couleur': 'info'
                })
        else:
            recommandations.append({
                'type': 'WARNING',
                'message': 'Aucune donnée disponible pour cette classe et cette période.',
                'couleur': 'warning'
            })
    else:
        recommandations.append({
            'type': 'INFO',
            'message': 'Sélectionnez une classe et une période pour voir les statistiques.',
            'couleur': 'info'
        })
    
    # Calculer les statistiques générales si des élèves sont évalués
    stats_generales = None
    if nb_evalues > 0:
        # Calculer la moyenne de classe
        total_moyennes = 0
        moyenne_max = 0
        moyenne_min = note_max if classe_selectionnee else 20
        
        for data in eleves_non_admis + eleves_a_suivre + eleves_excellents + eleves_precaution:
            moyenne = data['moyenne']
            total_moyennes += moyenne
            moyenne_max = max(moyenne_max, moyenne)
            moyenne_min = min(moyenne_min, moyenne)
        
        moyenne_classe = round(total_moyennes / nb_evalues, 2) if nb_evalues > 0 else 0
        taux_reussite = round((nb_evalues - nb_non_admis) / nb_evalues * 100, 1) if nb_evalues > 0 else 0
        
        stats_generales = {
            'total_eleves': nb_evalues,
            'moyenne_classe': moyenne_classe,
            'moyenne_max': round(moyenne_max, 2),
            'moyenne_min': round(moyenne_min, 2),
            'taux_reussite': taux_reussite
        }
        
        # Calculer les données pour les graphiques
        import json
        
        # Répartition des notes
        all_eleves = eleves_non_admis + eleves_a_suivre + eleves_excellents + eleves_precaution
        
        if est_primaire:
            excellent = len([e for e in all_eleves if e['moyenne'] >= 8])
            tres_bien = len([e for e in all_eleves if 7 <= e['moyenne'] < 8])
            bien = len([e for e in all_eleves if 6 <= e['moyenne'] < 7])
            assez_bien = len([e for e in all_eleves if 5 <= e['moyenne'] < 6])
            insuffisant = len([e for e in all_eleves if e['moyenne'] < 5])
        else:
            excellent = len([e for e in all_eleves if e['moyenne'] >= 16])
            tres_bien = len([e for e in all_eleves if 14 <= e['moyenne'] < 16])
            bien = len([e for e in all_eleves if 12 <= e['moyenne'] < 14])
            assez_bien = len([e for e in all_eleves if 10 <= e['moyenne'] < 12])
            insuffisant = len([e for e in all_eleves if e['moyenne'] < 10])
        
        repartition_json = json.dumps({
            'excellent': excellent,
            'tres_bien': tres_bien,
            'bien': bien,
            'assez_bien': assez_bien,
            'insuffisant': insuffisant
        })
        
        # Statistiques par matière (réutilise les données déjà calculées)
        stats_matieres = []
        for matiere in matieres:
            notes = stats_matieres_notes.get(matiere.id, [])
            if notes:
                stats_matieres.append({
                    'matiere': matiere.nom,
                    'moyenne': round(sum(notes) / len(notes), 2),
                    'max': round(max(notes), 2),
                    'min': round(min(notes), 2),
                    'nb_eleves': len(notes)
                })
        
        stats_matieres_json = json.dumps(stats_matieres)

        # Top 10 élèves (trié par moyenne décroissante)
        all_eleves = eleves_non_admis + eleves_a_suivre + eleves_excellents + eleves_precaution
        stats_par_eleve = sorted(all_eleves, key=lambda x: x['moyenne'], reverse=True)[:10]

        # Évolution (pour l'instant, juste la période actuelle)
        evolution_json = json.dumps([{
            'periode': periode,
            'moyenne': moyenne_classe
        }])
    else:
        repartition_json = '[]'
        stats_matieres_json = '[]'
        evolution_json = '[]'
    
    context = {
        'titre_page': 'Statistiques de l\'École',
        'ecole': ecole,
        'classes': classes,
        'classe_selectionnee': classe_selectionnee,
        'periode': periode,
        'total_eleves': total_eleves,
        'total_classes': total_classes,
        'stats_generales': stats_generales,
        'repartition_json': repartition_json,
        'stats_matieres_json': stats_matieres_json,
        'evolution_json': evolution_json,
        'stats_par_matiere': stats_matieres if stats_matieres_json != '[]' else None,
        'nb_evalues': nb_evalues,
        'nb_non_evalues': nb_non_evalues,
        'nb_non_admis': nb_non_admis,
        'nb_a_suivre': nb_a_suivre,
        'nb_excellents': nb_excellents,
        'nb_precaution': nb_precaution,
        'total_echecs': nb_non_admis,
        'taux_reussite': round((nb_evalues - nb_non_admis) / nb_evalues * 100, 1) if nb_evalues > 0 else 0,
        'taux_echec': round(nb_non_admis / nb_evalues * 100, 1) if nb_evalues > 0 else 0,
        'eleves_non_admis': eleves_non_admis,
        'eleves_a_suivre': eleves_a_suivre,
        'eleves_excellents': eleves_excellents,
        'eleves_precaution': eleves_precaution,
        'note_max': note_max if classe_selectionnee else 20,
        'est_primaire': est_primaire if classe_selectionnee else False,
        'seuil_reussite': seuil_reussite if classe_selectionnee else 10,
        'strategies': [],
        'recommandations': recommandations,
        'stats_par_eleve': stats_par_eleve if nb_evalues > 0 else [],
        'periodes': [
            # Mois
            ('OCTOBRE', 'Octobre'),
            ('NOVEMBRE', 'Novembre'),
            ('DECEMBRE', 'Décembre'),
            ('JANVIER', 'Janvier'),
            ('FEVRIER', 'Février'),
            ('MARS', 'Mars'),
            ('AVRIL', 'Avril'),
            ('MAI', 'Mai'),
            ('JUIN', 'Juin'),
            # Trimestres
            ('TRIMESTRE_1', '1er Trimestre'),
            ('TRIMESTRE_2', '2ème Trimestre'),
            ('TRIMESTRE_3', '3ème Trimestre'),
            # Semestres
            ('SEMESTRE_1', '1er Semestre'),
            ('SEMESTRE_2', '2ème Semestre'),
        ]
    }
    
    return render(request, 'notes/statistiques.html', context)

@login_required
def gerer_classes(request):
    """Gérer les classes - Liste et ajout"""
    user_profil = getattr(request.user, 'profil', None)
    ecole = user_profil.ecole if user_profil else None
    
    # Récupérer les classes
    if ecole:
        classes = ClasseNote.objects.filter(ecole=ecole).order_by('-date_creation')
    else:
        classes = ClasseNote.objects.all().order_by('-date_creation')
    
    # Statistiques
    total_classes = classes.count()
    classes_actives = classes.filter(actif=True).count()
    
    # Traitement du formulaire
    form = ClasseNoteForm(ecole=ecole)
    if request.method == 'POST':
        form = ClasseNoteForm(request.POST, ecole=ecole)
        if form.is_valid():
            classe = form.save(commit=False)
            if ecole:
                classe.ecole = ecole
            classe.cree_par = request.user
            
            # Vérifier si la classe existe déjà (utiliser les données du formulaire directement)
            classe_existante = ClasseNote.objects.filter(
                ecole=ecole,  # Utiliser l'école du profil utilisateur
                nom=classe.nom,
                annee_scolaire=classe.annee_scolaire
            ).first()
            
            if classe_existante:
                messages.error(request, f'❌ La classe "{classe.nom}" existe déjà pour l\'année scolaire {classe.annee_scolaire}.')
            else:
                try:
                    classe.save()
                    messages.success(request, f'✅ Classe "{classe.nom}" créée avec succès!')
                    return redirect('notes:gerer_classes')
                except IntegrityError:
                    messages.error(request, f'❌ Erreur: La classe "{classe.nom}" existe déjà pour cette année scolaire.')
        else:
            messages.error(request, '❌ Veuillez corriger les erreurs dans le formulaire.')
    
    context = {
        'titre_page': 'Gestion des Classes',
        'classes': classes,
        'total_classes': total_classes,
        'classes_actives': classes_actives,
        'form': form,
    }
    
    return render(request, 'notes/gerer_classes.html', context)

@login_required
def modifier_classe(request, classe_id):
    """Modifier une classe"""
    classe = get_object_or_404(ClasseNote, pk=classe_id)
    
    # Vérifier que l'utilisateur a accès à cette classe
    user_profil = getattr(request.user, 'profil', None)
    ecole = user_profil.ecole if user_profil else None
    
    # Sécurité : Vérifier que la classe appartient à l'école de l'utilisateur
    if ecole and classe.ecole != ecole:
        messages.error(request, "❌ Vous n'avez pas accès à cette classe")
        return redirect('notes:gerer_classes')
    
    if request.method == 'POST':
        form = ClasseNoteForm(request.POST, instance=classe, ecole=ecole)
        if form.is_valid():
            # Sauvegarder sans commit pour garder l'école
            instance = form.save(commit=False)
            # S'assurer que l'école reste la même
            instance.ecole = classe.ecole
            instance.save()
            messages.success(request, f'✅ Classe "{classe.nom}" modifiée avec succès!')
            return redirect('notes:gerer_classes')
        else:
            # Afficher les erreurs détaillées
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'❌ {field}: {error}')
    else:
        form = ClasseNoteForm(instance=classe, ecole=ecole)
    
    context = {
        'titre_page': 'Modifier une Classe',
        'form': form,
        'classe': classe,
    }
    
    return render(request, 'notes/modifier_classe.html', context)

@login_required
def supprimer_classe(request, classe_id):
    """Supprimer une classe avec vérification des données liées"""
    from django.http import JsonResponse
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        classe = get_object_or_404(ClasseNote, pk=classe_id)
        
        # Sécurité : Vérifier que la classe appartient à l'école de l'utilisateur
        user_profil = getattr(request.user, 'profil', None)
        ecole = user_profil.ecole if user_profil else None
        if ecole and classe.ecole != ecole:
            return JsonResponse({'success': False, 'error': 'Accès non autorisé'}, status=403)
        
        # Vérifier s'il y a des matières ou notes liées
        has_matieres = MatiereNote.objects.filter(classe=classe).exists()
        has_evaluations = Evaluation.objects.filter(matiere__classe=classe).exists()
        
        if has_matieres or has_evaluations:
            # Désactiver au lieu de supprimer
            classe.actif = False
            classe.save()
            return JsonResponse({
                'success': True,
                'message': f'Classe "{classe.nom}" désactivée (contient des données)'
            })
        else:
            # Supprimer si pas de données
            nom_classe = classe.nom
            classe.delete()
            return JsonResponse({
                'success': True,
                'message': f'Classe "{nom_classe}" supprimée avec succès'
            })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def gerer_matieres(request):
    """Gérer les matières par classe"""
    user_profil = getattr(request.user, 'profil', None)
    ecole = user_profil.ecole if user_profil else None
    
    # Récupérer les classes
    if ecole:
        classes = ClasseNote.objects.filter(ecole=ecole, actif=True).order_by('niveau', 'nom')
    else:
        classes = ClasseNote.objects.filter(actif=True).order_by('niveau', 'nom')
    
    # Filtres
    classe_id = request.GET.get('classe_id')
    classe_selectionnee = None
    matieres = []
    
    if classe_id:
        classe_selectionnee = get_object_or_404(ClasseNote, pk=classe_id)
        matieres = MatiereNote.objects.filter(classe=classe_selectionnee).order_by('nom')
    
    # Traitement du formulaire d'ajout
    form = MatiereNoteForm()
    if request.method == 'POST' and classe_selectionnee:
        form = MatiereNoteForm(request.POST)
        if form.is_valid():
            code = form.cleaned_data.get('code')
            nom = form.cleaned_data.get('nom')
            
            # Vérifier si le code existe déjà pour cette classe
            if MatiereNote.objects.filter(classe=classe_selectionnee, code=code).exists():
                messages.error(request, f'❌ Une matière avec le code "{code}" existe déjà pour cette classe.')
                return redirect(f'/notes/matieres/?classe_id={classe_id}')
            
            # Vérifier si le nom existe déjà pour cette classe
            if MatiereNote.objects.filter(classe=classe_selectionnee, nom=nom).exists():
                messages.error(request, f'❌ Une matière avec le nom "{nom}" existe déjà pour cette classe.')
                return redirect(f'/notes/matieres/?classe_id={classe_id}')
            
            matiere = form.save(commit=False)
            matiere.classe = classe_selectionnee
            matiere.cree_par = request.user
            
            # Gestion des coefficients selon le niveau
            from .calculs_moyennes import detecter_niveau_scolaire
            niveau = detecter_niveau_scolaire(classe_selectionnee.nom)
            
            if niveau == 'MATERNELLE':
                # MATERNELLE: Pas de coefficient (pas de notes numériques)
                matiere.coefficient = None
            elif niveau == 'PRIMAIRE':
                # PRIMAIRE: Coefficient = 1.0 (pas de pondération)
                matiere.coefficient = 1.0
            # COLLEGE/LYCEE: Garder le coefficient saisi par l'utilisateur
            
            matiere.save()
            messages.success(request, f'✅ Matière "{matiere.nom}" ajoutée avec succès!')
            return redirect(f'/notes/matieres/?classe_id={classe_id}')
        else:
            messages.error(request, '❌ Veuillez corriger les erreurs dans le formulaire.')
    
    # Détecter le niveau de la classe sélectionnée pour le template
    niveau_classe = None
    if classe_selectionnee:
        from .calculs_moyennes import detecter_niveau_scolaire
        niveau_classe = detecter_niveau_scolaire(classe_selectionnee.nom)
    
    context = {
        'titre_page': 'Gestion des Matières',
        'classes': classes,
        'classe_selectionnee': classe_selectionnee,
        'niveau_classe': niveau_classe,
        'matieres': matieres,
        'form': form,
    }
    
    return render(request, 'notes/gerer_matieres.html', context)

@login_required
def modifier_matiere(request, matiere_id):
    """Modifier une matière"""
    matiere = get_object_or_404(MatiereNote, pk=matiere_id)
    
    # Sécurité : Vérifier que la matière appartient à l'école de l'utilisateur
    user_profil = getattr(request.user, 'profil', None)
    ecole = user_profil.ecole if user_profil else None
    if ecole and matiere.classe.ecole != ecole:
        messages.error(request, "❌ Vous n'avez pas accès à cette matière")
        return redirect('notes:gerer_matieres')
    classe_id = matiere.classe.id
    
    if request.method == 'POST':
        form = MatiereNoteForm(request.POST, instance=matiere)
        if form.is_valid():
            code = form.cleaned_data.get('code')
            nom = form.cleaned_data.get('nom')
            
            # Vérifier si le code existe déjà pour cette classe (exclure l'instance actuelle)
            if MatiereNote.objects.filter(classe=matiere.classe, code=code).exclude(pk=matiere.pk).exists():
                messages.error(request, f'❌ Une matière avec le code "{code}" existe déjà pour cette classe.')
                return redirect(f'/notes/matieres/?classe_id={classe_id}')
            
            # Vérifier si le nom existe déjà pour cette classe (exclure l'instance actuelle)
            if MatiereNote.objects.filter(classe=matiere.classe, nom=nom).exclude(pk=matiere.pk).exists():
                messages.error(request, f'❌ Une matière avec le nom "{nom}" existe déjà pour cette classe.')
                return redirect(f'/notes/matieres/?classe_id={classe_id}')
            
            instance = form.save(commit=False)
            
            # Gestion des coefficients selon le niveau
            from .calculs_moyennes import detecter_niveau_scolaire
            niveau = detecter_niveau_scolaire(matiere.classe.nom)
            
            if niveau == 'MATERNELLE':
                # MATERNELLE: Pas de coefficient (pas de notes numériques)
                instance.coefficient = None
            elif niveau == 'PRIMAIRE':
                # PRIMAIRE: Coefficient = 1.0 (pas de pondération)
                instance.coefficient = 1.0
            # COLLEGE/LYCEE: Garder le coefficient saisi par l'utilisateur
            
            instance.save()
            messages.success(request, f'✅ Matière "{matiere.nom}" modifiée avec succès!')
            return redirect(f'/notes/matieres/?classe_id={classe_id}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'❌ {field}: {error}')
    else:
        form = MatiereNoteForm(instance=matiere)
    
    context = {
        'titre_page': 'Modifier une Matière',
        'form': form,
        'matiere': matiere,
        'classe_id': classe_id,
    }
    
    return render(request, 'notes/modifier_matiere.html', context)

@login_required
def supprimer_matiere(request, matiere_id):
    """Supprimer une matière avec vérification des données liées"""
    from django.http import JsonResponse
    from .models import NoteMensuelle, CompositionNote, AppreciationMaternelle
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        matiere = get_object_or_404(MatiereNote, pk=matiere_id)
        
        # Sécurité : Vérifier que la matière appartient à l'école de l'utilisateur
        user_profil = getattr(request.user, 'profil', None)
        ecole = user_profil.ecole if user_profil else None
        if ecole and matiere.classe.ecole != ecole:
            return JsonResponse({'success': False, 'error': 'Accès non autorisé'}, status=403)
        
        classe_id = matiere.classe.id
        
        # Vérifier TOUTES les données liées
        has_evaluations = Evaluation.objects.filter(matiere=matiere).exists()
        has_notes = NoteEleve.objects.filter(evaluation__matiere=matiere).exists()
        has_notes_mensuelles = NoteMensuelle.objects.filter(matiere=matiere).exists()
        has_compositions = CompositionNote.objects.filter(matiere=matiere).exists()
        has_appreciations = AppreciationMaternelle.objects.filter(matiere=matiere).exists()
        
        has_data = has_evaluations or has_notes or has_notes_mensuelles or has_compositions or has_appreciations
        
        if has_data:
            # Désactiver au lieu de supprimer
            matiere.actif = False
            matiere.save()
            return JsonResponse({
                'success': True,
                'message': f'Matière "{matiere.nom}" désactivée (contient des données)',
                'classe_id': classe_id
            })
        else:
            # Supprimer si pas de données
            nom_matiere = matiere.nom
            matiere.delete()
            return JsonResponse({
                'success': True,
                'message': f'Matière "{nom_matiere}" supprimée avec succès',
                'classe_id': classe_id
            })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def charger_matieres_defaut(request, classe_id):
    """Charger les matières par défaut pour une classe"""
    from notes.matieres_defaut import charger_matieres_pour_classe
    
    classe = get_object_or_404(ClasseNote, pk=classe_id)
    
    # Sécurité : Vérifier les permissions
    user_profil = getattr(request.user, 'profil', None)
    if user_profil and user_profil.ecole and classe.ecole != user_profil.ecole:
        messages.error(request, "❌ Vous n'avez pas accès à cette classe")
        return redirect('notes:gerer_matieres')
    
    # Charger les matières par défaut
    nombre_creees, nombre_existantes, erreurs = charger_matieres_pour_classe(classe, request.user)
    
    # Messages de retour
    if nombre_creees > 0:
        messages.success(request, f'✅ {nombre_creees} matière(s) créée(s) avec succès pour {classe.nom}')
    
    if nombre_existantes > 0:
        messages.info(request, f'ℹ️ {nombre_existantes} matière(s) existaient déjà')
    
    if erreurs:
        for erreur in erreurs:
            messages.warning(request, f'⚠️ {erreur}')
    
    if nombre_creees == 0 and nombre_existantes == 0:
        messages.warning(request, 'Aucune matière par défaut disponible pour ce niveau')
    
    return redirect('notes:gerer_matieres')

@login_required
def gerer_evaluations(request):
    """Gérer les évaluations"""
    return render(request, 'notes/gerer_evaluations.html', {'titre_page': 'Gestion des Évaluations'})

@login_required
def creer_evaluation(request):
    """Créer une évaluation"""
    messages.info(request, 'Fonction en cours de développement')
    return redirect('notes:gerer_evaluations')

@login_required
def gerer_eleves(request):
    """Gérer les élèves - Consultation par classe"""
    
    user_profil = getattr(request.user, 'profil', None)
    ecole = user_profil.ecole if user_profil else None
    
    # Récupérer les classes de notes
    if ecole:
        classes_notes = ClasseNote.objects.filter(ecole=ecole, actif=True).order_by('niveau', 'nom')
    else:
        classes_notes = ClasseNote.objects.filter(actif=True).order_by('niveau', 'nom')
    
    # Filtres
    classe_id = request.GET.get('classe_id')
    classe_selectionnee = None
    eleves = []
    
    if classe_id:
        classe_selectionnee = get_object_or_404(ClasseNote, pk=classe_id)
        
        # Trouver la classe d'élèves correspondante
        classe_eleve = None
        
        # Utiliser l'école de la classe sélectionnée (pas celle de l'utilisateur)
        ecole_classe = classe_selectionnee.ecole
        
        # Méthode 1 : Correspondance exacte par nom, année scolaire ET école
        try:
            # Utiliser filter().first() au lieu de get() pour éviter MultipleObjectsReturned
            classe_eleve = ClasseEleve.objects.filter(
                nom=classe_selectionnee.nom,
                annee_scolaire=classe_selectionnee.annee_scolaire,
                ecole=ecole_classe  # Utiliser l'école de la classe, pas celle de l'utilisateur
            ).first()
        except Exception:
            pass
        
        # Méthode 2 : Recherche approximative améliorée
        if not classe_eleve:
            # Extraire les chiffres et mots-clés du nom
            nom_recherche = classe_selectionnee.nom.lower()
            # Nettoyer le nom
            nom_nettoye = nom_recherche.replace('série', '').replace('année', '').replace('ème', '').replace('eme', '').strip()
            
            # Chercher dans les classes de la même année scolaire ET de la même école
            classes_similaires = ClasseEleve.objects.filter(
                annee_scolaire=classe_selectionnee.annee_scolaire,
                ecole=ecole_classe  # Filtrer par l'école de la classe
            )
            
            # Essayer de trouver une correspondance
            for classe_candidate in classes_similaires:
                nom_candidate = classe_candidate.nom.lower()
                # Vérifier si le nom nettoyé est contenu dans le nom de la classe candidate
                if nom_nettoye in nom_candidate or any(mot in nom_candidate for mot in nom_nettoye.split() if len(mot) > 2):
                    classe_eleve = classe_candidate
                    break
        
        # Méthode 3 : Recherche par niveau si disponible
        if not classe_eleve and hasattr(classe_selectionnee, 'niveau'):
            niveau_map = {
                'MATERNELLE': ['maternelle', 'petite', 'moyenne', 'grande'],
                'PRIMAIRE': ['cp', 'ce', 'cm', 'primaire', '1ère', '2ème', '3ème', '4ème', '5ème', '6ème'],
                'COLLEGE': ['7ème', '8ème', '9ème', '10ème', 'collège', 'college'],
                'LYCEE': ['11ème', '12ème', 'lycée', 'lycee', 'terminale', 'première', 'seconde']
            }
            
            niveau = classe_selectionnee.niveau
            mots_cles = niveau_map.get(niveau, [])
            nom_classe_lower = classe_selectionnee.nom.lower()
            
            # Filtrer par année scolaire ET école
            for classe_candidate in ClasseEleve.objects.filter(
                annee_scolaire=classe_selectionnee.annee_scolaire,
                ecole=ecole_classe  # Filtrer par l'école de la classe
            ):
                nom_candidate_lower = classe_candidate.nom.lower()
                # Vérifier si des mots-clés du niveau sont présents dans les deux noms
                if any(mot in nom_classe_lower and mot in nom_candidate_lower for mot in mots_cles):
                    classe_eleve = classe_candidate
                    break
        
        # Récupérer les élèves si une classe a été trouvée
        if classe_eleve:
            eleves = Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('prenom', 'nom')
    
    # Calculer les statistiques
    total_eleves = len(eleves)
    eleves_avec_notes = 0  # TODO: Calculer le nombre d'élèves avec des notes
    
    # Informations de débogage pour l'utilisateur
    classe_eleve_trouvee = None
    classes_disponibles = []
    if classe_selectionnee and not eleves:
        # Lister toutes les classes d'élèves disponibles pour aider au diagnostic
        # Filtrer par l'école de la classe sélectionnée
        classes_disponibles = list(ClasseEleve.objects.filter(
            annee_scolaire=classe_selectionnee.annee_scolaire,
            ecole=classe_selectionnee.ecole  # Filtrer par l'école de la classe
        ).values_list('nom', flat=True))
    
    context = {
        'titre_page': 'Gestion des Élèves',
        'classes': classes_notes,
        'classe_selectionnee': classe_selectionnee,
        'eleves': eleves,
        'total_eleves': total_eleves,
        'eleves_avec_notes': eleves_avec_notes,
        'classes_disponibles': classes_disponibles,
    }
    
    return render(request, 'notes/gerer_eleves.html', context)

@login_required
def saisir_notes(request):
    """Saisir les notes"""
    
    user_profil = getattr(request.user, 'profil', None)
    ecole = user_profil.ecole if user_profil else None
    
    # Récupérer les classes
    if ecole:
        classes = ClasseNote.objects.filter(ecole=ecole, actif=True).order_by('niveau', 'nom')
    else:
        classes = ClasseNote.objects.filter(actif=True).order_by('niveau', 'nom')
    
    # Paramètres de sélection
    classe_id = request.GET.get('classe_id')
    periode_classement = request.GET.get('periode', '')
    matiere_id = request.GET.get('matiere_id')
    type_note = request.GET.get('type_note', '')
    periode = request.GET.get('periode', '')
    system_type = request.GET.get('system_type', 'semestre')
    
    classe_selectionnee = None
    matiere_selectionnee = None
    matieres = []
    eleves = []
    evaluations = []
    niveau_enseignement = 'SECONDAIRE'
    est_maternelle = False
    
    # Types de notes disponibles par défaut
    types_notes_disponibles = [
        ('mensuelle', 'Note Mensuelle'),
        ('trimestrielle', 'Note Trimestrielle'),
        ('semestrielle', 'Note Semestrielle'),
        ('composition', 'Composition'),
        ('appreciation', 'Appréciation'),
    ]
    
    # Périodes disponibles par défaut
    periodes_disponibles = []

    if classe_id:
        classe_selectionnee = get_object_or_404(ClasseNote, pk=classe_id)
        niveau_enseignement = classe_selectionnee.niveau_enseignement
        matieres = MatiereNote.objects.filter(
            classe=classe_selectionnee, actif=True
        ).select_related('classe').order_by('nom')

        # Détecter le niveau scolaire pour adapter l'interface
        from .calculs_moyennes import detecter_niveau_scolaire
        niveau_detecte = detecter_niveau_scolaire(classe_selectionnee.nom)
        est_maternelle = (niveau_detecte == 'MATERNELLE')
        
        # Pour la maternelle : uniquement les appréciations (pas de notes numériques)
        if est_maternelle:
            types_notes_disponibles = [
                ('appreciation', 'Appréciation'),
            ]
            # Forcer le type de note à appreciation pour la maternelle
            if type_note != 'appreciation':
                type_note = 'appreciation'
        
        # Déterminer les périodes disponibles selon le type de note
        if not est_maternelle and not type_note:
            if periode.startswith('SEMESTRE'):
                type_note = 'semestrielle'
            elif periode.startswith('TRIMESTRE'):
                type_note = 'trimestrielle'
            else:
                type_note = 'trimestrielle'

        if type_note == 'appreciation':
            # Pour les appréciations : trimestres
            periodes_disponibles = [
                ('TRIMESTRE_1', '1er Trimestre'),
                ('TRIMESTRE_2', '2ème Trimestre'),
                ('TRIMESTRE_3', '3ème Trimestre'),
            ]
        elif type_note == 'mensuelle':
            # Pour les notes mensuelles : 9 mois de l'année scolaire
            periodes_disponibles = [
                ('OCTOBRE', 'Octobre'),
                ('NOVEMBRE', 'Novembre'),
                ('DECEMBRE', 'Décembre'),
                ('JANVIER', 'Janvier'),
                ('FEVRIER', 'Février'),
                ('MARS', 'Mars'),
                ('AVRIL', 'Avril'),
                ('MAI', 'Mai'),
                ('JUIN', 'Juin'),
            ]
        elif type_note == 'trimestrielle':
            # Pour les notes trimestrielles
            periodes_disponibles = [
                ('TRIMESTRE_1', '1er Trimestre'),
                ('TRIMESTRE_2', '2ème Trimestre'),
                ('TRIMESTRE_3', '3ème Trimestre'),
            ]
        elif type_note == 'semestrielle':
            # Pour les notes semestrielles
            periodes_disponibles = [
                ('SEMESTRE_1', '1er Semestre'),
                ('SEMESTRE_2', '2ème Semestre'),
            ]
        elif type_note == 'composition':
            # Pour les compositions : selon le système choisi
            if system_type == 'semestre':
                periodes_disponibles = [
                    ('SEMESTRE_1', '1er Semestre'),
                    ('SEMESTRE_2', '2ème Semestre'),
                ]
            else:  # trimestre
                periodes_disponibles = [
                    ('TRIMESTRE_1', '1er Trimestre'),
                    ('TRIMESTRE_2', '2ème Trimestre'),
                    ('TRIMESTRE_3', '3ème Trimestre'),
                ]
        else:
            # Par défaut : trimestres
            periodes_disponibles = [
                ('TRIMESTRE_1', '1er Trimestre'),
                ('TRIMESTRE_2', '2ème Trimestre'),
                ('TRIMESTRE_3', '3ème Trimestre'),
            ]
        
        if matiere_id:
            matiere_selectionnee = get_object_or_404(
                MatiereNote,
                pk=matiere_id,
                classe=classe_selectionnee,
                actif=True,
            )
            if periode:
                evaluations = Evaluation.objects.filter(
                    matiere=matiere_selectionnee, periode=periode
                ).select_related('matiere').order_by('date_evaluation')
            else:
                evaluations = Evaluation.objects.none()

            # Vérifier si des notes existent déjà (1 seule requête count)
            notes_existantes_count = NoteEleve.objects.filter(
                evaluation__matiere=matiere_selectionnee,
                evaluation__periode=periode
            ).count() if periode else 0
            
            # Récupérer les élèves
            try:
                # Mapping spécial pour les classes avec noms différents (même que consulter_notes)
                mapping_classes = {
                    61: 56,  # ClasseNote '12ème Année' -> ClasseEleve '12ÈME ANNÉE'
                    59: 8,   # ClasseNote '11ème Série littéraire' -> ClasseEleve '11ème série littéraire'
                }
                
                if classe_selectionnee.id in mapping_classes:
                    classe_eleve = ClasseEleve.objects.filter(
                        id=mapping_classes[classe_selectionnee.id]
                    ).first()
                else:
                    # Utiliser filter().first() au lieu de get() pour éviter MultipleObjectsReturned
                    classe_eleve = ClasseEleve.objects.filter(
                        nom=classe_selectionnee.nom,
                        annee_scolaire=classe_selectionnee.annee_scolaire,
                        ecole=classe_selectionnee.ecole  # Filtrer par l'école de la classe
                    ).first()
                
                if classe_eleve:
                    eleves = Eleve.objects.filter(
                        classe=classe_eleve, statut='ACTIF'
                    ).select_related('classe').order_by('prenom', 'nom')
                else:
                    # Recherche approximative
                    nom_recherche = classe_selectionnee.nom.lower().replace('série', '').replace('année', '').strip()
                    classes_similaires = ClasseEleve.objects.filter(
                        nom__icontains=nom_recherche,
                        ecole=classe_selectionnee.ecole
                    )
                    if classes_similaires.count() >= 1:
                        classe_eleve = classes_similaires.first()
                        eleves = Eleve.objects.filter(
                            classe=classe_eleve, statut='ACTIF'
                        ).select_related('classe').order_by('prenom', 'nom')
            except Exception:
                pass
    
    # Préparer les informations sur les notes existantes
    notes_deja_saisies = False
    nombre_notes_existantes = 0
    notes_existantes_json = '{}'
    
    if matiere_selectionnee and periode and eleves:
        import json as _json
        import logging
        logger = logging.getLogger(__name__)
        notes_map = {}
        
        # Définir les types de périodes
        periodes_mensuelles = ['OCTOBRE', 'NOVEMBRE', 'DECEMBRE', 'JANVIER', 'FEVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN']
        periodes_trimestrielles = ['TRIMESTRE_1', 'TRIMESTRE_2', 'TRIMESTRE_3']
        periodes_semestrielles = ['SEMESTRE_1', 'SEMESTRE_2']
        
        annee_scolaire = classe_selectionnee.annee_scolaire if classe_selectionnee else ''
        
        # Récupérer les IDs des élèves de la classe pour filtrer
        eleves_ids = list(eleves.values_list('id', flat=True))
        
        # DEBUG: Log des paramètres de recherche
        logger.info(f"[SAISIR_NOTES] Recherche notes: matiere={matiere_selectionnee.id}, periode={periode}, type_note={type_note}")
        logger.info(f"[SAISIR_NOTES] Année scolaire: {annee_scolaire}, Nb élèves: {len(eleves_ids)}")
        
        # 1. Chercher dans NoteMensuelle (notes mensuelles importées ou saisies)
        if periode in periodes_mensuelles or type_note == 'mensuelle':
            try:
                # Chercher les notes mensuelles pour cette matière et période
                notes_mensuelles_qs = NoteMensuelle.objects.filter(
                    matiere=matiere_selectionnee,
                    mois=periode,
                    eleve_id__in=eleves_ids  # Filtrer par les élèves de la classe
                )
                
                # DEBUG: Compter avant filtre année
                count_avant = notes_mensuelles_qs.count()
                logger.info(f"[SAISIR_NOTES] NoteMensuelle avant filtre année: {count_avant}")
                
                # Filtrer par année scolaire si disponible
                if annee_scolaire:
                    notes_mensuelles_qs = notes_mensuelles_qs.filter(annee_scolaire=annee_scolaire)
                
                # DEBUG: Compter après filtre année
                count_apres = notes_mensuelles_qs.count()
                logger.info(f"[SAISIR_NOTES] NoteMensuelle après filtre année: {count_apres}")
                
                notes_mensuelles = notes_mensuelles_qs.select_related('eleve')
                
                for n in notes_mensuelles:
                    try:
                        notes_map[n.eleve_id] = {
                            'note': float(n.note) if n.note is not None else None,
                            'absent': bool(n.absent),
                            'appreciation': None,
                            'commentaire': None,
                        }
                    except Exception:
                        continue
                
                logger.info(f"[SAISIR_NOTES] Notes map créé avec {len(notes_map)} entrées")
            except Exception as e:
                logger.error(f"[SAISIR_NOTES] Erreur récupération NoteMensuelle: {e}")
        
        # 2. Chercher dans AppreciationMaternelle (appréciations maternelle)
        if type_note == 'appreciation' and est_maternelle:
            try:
                from .models import AppreciationMaternelle
                appreciations_maternelle = AppreciationMaternelle.objects.filter(
                    matiere=matiere_selectionnee,
                    trimestre=periode,
                    eleve_id__in=eleves_ids
                )
                
                # Filtrer par année scolaire si disponible
                if annee_scolaire:
                    appreciations_maternelle = appreciations_maternelle.filter(annee_scolaire=annee_scolaire)
                
                appreciations_maternelle = appreciations_maternelle.select_related('eleve')
                
                logger.info(f"[SAISIR_NOTES] AppreciationMaternelle trouvées: {appreciations_maternelle.count()}")
                
                for app in appreciations_maternelle:
                    try:
                        notes_map[app.eleve_id] = {
                            'note': None,
                            'absent': bool(app.absent),
                            'appreciation': app.appreciation,  # A, B, C, D, E
                            'commentaire': app.observation if hasattr(app, 'observation') else None,
                        }
                    except Exception:
                        continue
                
                logger.info(f"[SAISIR_NOTES] Notes map après AppreciationMaternelle: {len(notes_map)} entrées")
            except Exception as e:
                logger.error(f"[SAISIR_NOTES] Erreur récupération AppreciationMaternelle: {e}")
        
        # 3. Chercher dans CompositionNote (notes trimestrielles/semestrielles importées)
        if periode in periodes_trimestrielles or periode in periodes_semestrielles:
            try:
                notes_composition = CompositionNote.objects.filter(
                    matiere=matiere_selectionnee,
                    periode=periode,
                    annee_scolaire=annee_scolaire,
                    eleve_id__in=eleves_ids  # Filtrer par les élèves de la classe
                ).select_related('eleve')
                
                for n in notes_composition:
                    try:
                        notes_map[n.eleve_id] = {
                            'note': float(n.note) if n.note is not None else None,
                            'absent': bool(n.absent),
                            'appreciation': None,
                            'commentaire': None,
                        }
                    except Exception:
                        continue
            except Exception:
                pass
        
        # 3. Chercher aussi dans NoteEleve (via évaluations) - peut écraser les valeurs précédentes
        if evaluations.exists():
            try:
                qs_notes = NoteEleve.objects.filter(
                    evaluation__in=evaluations,
                    eleve_id__in=eleves_ids  # Filtrer par les élèves de la classe
                ).select_related('eleve')
                
                for n in qs_notes:
                    try:
                        # Dernière valeur écrase la précédente si plusieurs evals
                        notes_map[str(n.eleve_id)] = {
                            'note': float(n.note) if getattr(n, 'note', None) is not None else None,
                            'absent': bool(getattr(n, 'absent', False)),
                            'appreciation': getattr(n, 'appreciation_finale', None),
                            'commentaire': getattr(n, 'commentaire', None),
                        }
                    except Exception:
                        continue
            except Exception:
                pass
        
        nombre_notes_existantes = len(notes_map)
        notes_deja_saisies = nombre_notes_existantes > 0
        
        try:
            notes_existantes_json = _json.dumps(notes_map)
        except Exception:
            notes_existantes_json = '{}'
    
    # Déterminer la note maximale selon le niveau
    note_sur = 20  # Par défaut
    if classe_selectionnee:
        if niveau_enseignement == 'PRIMAIRE' or 'PRIMAIRE' in classe_selectionnee.niveau:
            note_sur = 10
        else:
            note_sur = 20
    
    context = {
        'titre_page': 'Saisie des Notes',
        'classes': classes,
        'classe_selectionnee': classe_selectionnee,
        'matieres': matieres,
        'matiere_selectionnee': matiere_selectionnee,
        'eleves': eleves,
        'evaluations': evaluations,
        'type_note': type_note,
        'periode': periode,
        'periodes_disponibles': periodes_disponibles,
        'types_notes_disponibles': types_notes_disponibles,
        'system_type': system_type,
        'niveau_enseignement': niveau_enseignement,
        'est_maternelle': est_maternelle,
        'notes_existantes_json': notes_existantes_json,
        'notes_deja_saisies': notes_deja_saisies,
        'nombre_notes_existantes': nombre_notes_existantes,
        'note_sur': note_sur,
    }
    
    return render(request, 'notes/saisir_notes.html', context)

@login_required
def liste_saisie_pdf(request):
    """Générer un PDF de la liste de saisie des notes"""
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    import io
    import re
    
    # Récupérer les paramètres
    classe_id = request.GET.get('classe_id')
    matiere_id = request.GET.get('matiere_id')
    periode = request.GET.get('periode')
    type_note = request.GET.get('type_note', '')
    
    if not all([classe_id, matiere_id, periode]):
        return HttpResponse("Paramètres manquants: classe_id, matiere_id et periode sont requis", status=400)
    
    try:
        classe = get_object_or_404(ClasseNote, pk=classe_id)
        matiere = get_object_or_404(MatiereNote, pk=matiere_id)
    except Exception as e:
        return HttpResponse(f"Erreur lors de la récupération des données: {str(e)}", status=400)
    
    # Déterminer le type de notation selon le niveau
    niveau_enseignement = classe.niveau_enseignement or 'SECONDAIRE'
    niveau = classe.niveau or ''
    is_maternelle = niveau_enseignement == 'MATERNELLE'
    is_primaire = niveau_enseignement == 'PRIMAIRE' or 'PRIMAIRE' in niveau
    is_appreciation = type_note == 'appreciation'
    
    # Déterminer la note maximale
    if is_primaire:
        note_sur = 10
    else:
        note_sur = 20
    
    # Récupérer les élèves avec mapping spécial (même logique que saisir_notes et consulter_notes)
    mapping_classes = {
        61: 56,  # ClasseNote '12ème Année' -> ClasseEleve '12ÈME ANNÉE'
        59: 8,   # ClasseNote '11ème Série littéraire' -> ClasseEleve '11ème série littéraire'
    }
    
    if classe.id in mapping_classes:
        classe_eleve = ClasseEleve.objects.filter(
            id=mapping_classes[classe.id]
        ).first()
    else:
        classe_eleve = ClasseEleve.objects.filter(
            nom=classe.nom,
            annee_scolaire=classe.annee_scolaire,
            ecole=classe.ecole
        ).first()
    
    if classe_eleve:
        eleves = Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('prenom', 'nom')
    else:
        eleves = []
    
    # Créer le PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#007bff'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    # Titre
    elements.append(Paragraph(f"Liste de Saisie - {classe.nom}", title_style))
    elements.append(Paragraph(f"Matière: {matiere.nom} | Période: {periode}", styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))
    
    # En-tête du tableau selon le type
    if is_appreciation:
        # Pour les appréciations (Maternelle)
        data = [['N°', 'Matricule', 'Prénom', 'Nom', 'Appréciation', 'Commentaire', 'Absent']]
        col_widths = [1*cm, 3*cm, 4*cm, 4*cm, 4*cm, 5*cm, 2*cm]
    else:
        # Pour les notes (Primaire /10 ou Secondaire /20)
        data = [['N°', 'Matricule', 'Prénom', 'Nom', f'Note /{note_sur}', 'Absent', 'Observations']]
        col_widths = [1*cm, 3*cm, 4*cm, 4*cm, 2*cm, 2*cm, 6*cm]
    
    for idx, eleve in enumerate(eleves, 1):
        if is_appreciation:
            data.append([
                str(idx),
                eleve.matricule or '',
                eleve.prenom or '',
                eleve.nom or '',
                '',  # Appréciation à remplir
                '',  # Commentaire
                ''   # Absent à cocher
            ])
        else:
            data.append([
                str(idx),
                eleve.matricule or '',
                eleve.prenom or '',
                eleve.nom or '',
                '',  # Note à remplir
                '',  # Absent à cocher
                ''   # Observations
            ])
    
    # Style du tableau
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#007bff')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # Construire le PDF
    doc.build(elements)
    
    # Retourner la réponse
    buffer.seek(0)
    
    # Nettoyer le nom de fichier (supprimer les caractères spéciaux)
    nom_classe_clean = re.sub(r'[^\w\s-]', '', classe.nom).replace(' ', '_')
    code_matiere_clean = re.sub(r'[^\w\s-]', '', matiere.code).replace(' ', '_')
    filename = f"liste_saisie_{nom_classe_clean}_{code_matiere_clean}.pdf"
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def fiche_saisie_notes_pdf(request):
    """Generer l'ancienne fiche PDF de saisie des notes pour les professeurs."""
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from weasyprint import HTML
    import base64

    classe_id = request.GET.get('classe_id')
    matiere_id = request.GET.get('matiere_id')
    system_type = request.GET.get('system_type', 'trimestre')

    if not classe_id:
        return HttpResponse("Parametre classe_id requis", status=400)

    try:
        classe = get_object_or_404(ClasseNote, pk=classe_id)
    except Exception as e:
        return HttpResponse(f"Erreur: {str(e)}", status=400)

    matiere = None
    if matiere_id:
        try:
            matiere = get_object_or_404(MatiereNote, pk=matiere_id)
        except Exception:
            pass

    mapping_classes = {
        61: 56,
        59: 8,
    }

    if classe.id in mapping_classes:
        classe_eleve = ClasseEleve.objects.filter(id=mapping_classes[classe.id]).first()
    else:
        classe_eleve = ClasseEleve.objects.filter(
            nom=classe.nom,
            annee_scolaire=classe.annee_scolaire,
            ecole=classe.ecole
        ).first()

        if not classe_eleve:
            classe_eleve = ClasseEleve.objects.filter(
                nom__iexact=classe.nom,
                annee_scolaire=classe.annee_scolaire
            ).first()

    eleves = []
    if classe_eleve:
        eleves = list(Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('prenom', 'nom'))

    if system_type == 'semestre':
        colonnes_periode1 = ['Octobre', 'Novembre', 'Decembre', 'Janvier', 'Moy. Cours', 'Composition']
        colonnes_periode2 = ['Fevrier', 'Mars', 'Avril', 'Mai', 'Moy. Cours', 'Composition']
        periode1_nom = '1er Semestre'
        periode2_nom = '2eme Semestre'
    else:
        colonnes_periode1 = ['Octobre', 'Novembre', 'Decembre', 'Moy. Cours', 'Composition']
        colonnes_periode2 = ['Janvier', 'Fevrier', 'Mars', 'Moy. Cours', 'Composition']
        colonnes_periode3 = ['Avril', 'Mai', 'Juin', 'Moy. Cours', 'Composition']
        periode1_nom = '1er Trimestre'
        periode2_nom = '2eme Trimestre'
        periode3_nom = '3eme Trimestre'

    user_profil = getattr(request.user, 'profil', None)
    ecole = user_profil.ecole if user_profil else classe.ecole

    logo_base64 = None
    if ecole and ecole.logo:
        try:
            with ecole.logo.open('rb') as logo_file:
                logo_base64 = base64.b64encode(logo_file.read()).decode('utf-8')
        except Exception:
            pass

    context = {
        'classe': classe,
        'matiere': matiere,
        'eleves': eleves,
        'system_type': system_type,
        'ecole': ecole,
        'annee_scolaire': classe.annee_scolaire,
        'date_impression': timezone.now(),
        'logo_base64': logo_base64,
    }

    if system_type == 'semestre':
        context.update({
            'colonnes_periode1': colonnes_periode1,
            'colonnes_periode2': colonnes_periode2,
            'periode1_nom': periode1_nom,
            'periode2_nom': periode2_nom,
        })
    else:
        context.update({
            'colonnes_periode1': colonnes_periode1,
            'colonnes_periode2': colonnes_periode2,
            'colonnes_periode3': colonnes_periode3,
            'periode1_nom': periode1_nom,
            'periode2_nom': periode2_nom,
            'periode3_nom': periode3_nom,
        })

    html_content = render_to_string('notes/fiche_saisie_notes_pdf.html', context)
    pdf_file = HTML(string=html_content).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    nom_classe_clean = classe.nom.replace(' ', '_').replace('/', '-')
    filename = f"fiche_saisie_{nom_classe_clean}_{system_type}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


@login_required
def fiche_report_notes_pdf(request):
    """Generer une fiche PDF mensuelle avec toutes les matieres en colonnes."""
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from weasyprint import HTML
    import base64

    classe_id = request.GET.get('classe_id')
    periode = request.GET.get('periode', 'OCTOBRE')

    if not classe_id:
        return HttpResponse("Parametre classe_id requis", status=400)

    try:
        classe = get_object_or_404(ClasseNote, pk=classe_id)
    except Exception as e:
        return HttpResponse(f"Erreur: {str(e)}", status=400)

    matieres = list(MatiereNote.objects.filter(classe=classe, actif=True).order_by('nom'))

    mapping_classes = {
        61: 56,
        59: 8,
    }

    if classe.id in mapping_classes:
        classe_eleve = ClasseEleve.objects.filter(id=mapping_classes[classe.id]).first()
    else:
        classe_eleve = ClasseEleve.objects.filter(
            nom=classe.nom,
            annee_scolaire=classe.annee_scolaire,
            ecole=classe.ecole
        ).first()

        if not classe_eleve:
            classe_eleve = ClasseEleve.objects.filter(
                nom__iexact=classe.nom,
                annee_scolaire=classe.annee_scolaire
            ).first()

    eleves = []
    if classe_eleve:
        eleves = list(Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('prenom', 'nom'))

    user_profil = getattr(request.user, 'profil', None)
    ecole = user_profil.ecole if user_profil else classe.ecole

    logo_base64 = None
    if ecole and ecole.logo:
        try:
            with ecole.logo.open('rb') as logo_file:
                logo_base64 = base64.b64encode(logo_file.read()).decode('utf-8')
        except Exception:
            pass

    periode_display = {
        'OCTOBRE': 'Octobre',
        'NOVEMBRE': 'Novembre',
        'DECEMBRE': 'Decembre',
        'JANVIER': 'Janvier',
        'FEVRIER': 'Fevrier',
        'MARS': 'Mars',
        'AVRIL': 'Avril',
        'MAI': 'Mai',
        'JUIN': 'Juin',
        'TRIMESTRE_1': '1er Trimestre',
        'TRIMESTRE_2': '2eme Trimestre',
        'TRIMESTRE_3': '3eme Trimestre',
        'SEMESTRE_1': '1er Semestre',
        'SEMESTRE_2': '2eme Semestre',
    }.get(periode, periode)

    context = {
        'classe': classe,
        'matieres': matieres,
        'eleves': eleves,
        'ecole': ecole,
        'annee_scolaire': classe.annee_scolaire,
        'date_impression': timezone.now(),
        'logo_base64': logo_base64,
        'periode': periode,
        'periode_display': periode_display,
    }

    html_content = render_to_string('notes/fiche_report_notes_pdf.html', context)
    pdf_file = HTML(string=html_content).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    nom_classe_clean = classe.nom.replace(' ', '_').replace('/', '-')
    filename = f"fiche_mensuelle_{nom_classe_clean}_{periode}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


@login_required
def imprimer_tableau_notes_pdf(request):
    """Imprimer le tableau des notes avec ajustement des colonnes sur A4 landscape"""
    from django.template.loader import render_to_string
    from weasyprint import HTML, CSS
    from django.http import HttpResponse
    
    # Récupérer les paramètres
    classe_id = request.GET.get('classe_id')
    periode = request.GET.get('periode')
    
    if not classe_id or not periode:
        return HttpResponse("Paramètres manquants", status=400)
    
    try:
        # Récupérer la classe et les données
        classe_note = get_object_or_404(ClasseNote, pk=classe_id)
        classe = classe_note.classe
        
        # Récupérer les matières
        matieres = MatiereNote.objects.filter(classe=classe_note).order_by('matiere__nom')
        
        # Calculer le classement
        from .calculs_moyennes import calculer_classement_classe
        from .calculs_intelligent import calculer_rang_intelligent
        
        # Récupérer les élèves
        from eleves.models import Eleve
        eleves = Eleve.objects.filter(classe=classe, est_actif=True).order_by('matricule')
        
        # Calculer les moyennes et rangs
        classement_resultat = calculer_classement_classe(eleves, matieres, periode, 'mensuel')
        
        # Préparer les données pour le template
        classement_data = []
        for eleve in eleves:
            # Récupérer les détails des notes par matière
            details_matieres = {}
            for matiere in matieres:
                from .calculs_moyennes import calculer_moyenne_matiere
                result = calculer_moyenne_matiere(eleve, matiere, periode, 'mensuel')
                details_matieres[matiere.id] = result
            
            # Récupérer le rang et la moyenne
            rang_num = classement_resultat['rang_map'].get(eleve.id)
            rang_str = str(rang_num) if rang_num else "-"
            moyenne = classement_resultat['moyennes_par_eleve'].get(eleve.id)
            
            # Formatter le rang avec ex-æquo si nécessaire
            if rang_num:
                from .calculs_intelligent import formater_rang_intelligent
                sexe = getattr(eleve, 'sexe', 'M') or 'M'
                rang_str = formater_rang_intelligent(rang_num, sexe)
            
            classement_data.append({
                'matricule': eleve.matricule,
                'nom_complet': eleve.nom_complet,
                'rang': rang_str,
                'moyenne': moyenne,
                'details_matieres': details_matieres,
                'sexe': getattr(eleve, 'sexe', 'M') or 'M'
            })
        
        # Trier par rang
        classement_data.sort(key=lambda x: x['rang'] if x['rang'] != '-' else '999')
        
        # Contexte pour le template
        context = {
            'classe_selectionnee': classe,
            'periode_selectionnee': periode,
            'matieres': matieres,
            'classement_data': classement_data,
        }
        
        # Générer le HTML
        html_content = render_to_string('notes/impression_tableau_notes.html', context, request=request)
        
        # Créer le PDF avec WeasyPrint
        html = HTML(string=html_content)
        css = CSS(string='''
            @page {
                size: A4 landscape;
                margin: 10mm;
            }
        ''')
        
        pdf = html.write_pdf(stylesheets=[css])
        
        # Retourner le PDF
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="tableau_notes_{classe.nom}_{periode}.pdf"'
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'impression du tableau: {str(e)}")
        return HttpResponse(f"Erreur: {str(e)}", status=500)

@login_required
def sauvegarder_notes(request):
    """Sauvegarder les notes saisies avec support des transactions"""
    from django.http import JsonResponse
    from eleves.models import Eleve
    from django.db import transaction
    import json
    from decimal import Decimal, InvalidOperation
    import logging
    
    logger = logging.getLogger(__name__)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)
    
    try:
        data = json.loads(request.body)
        notes_data = data.get('notes', [])
        evaluation_id = data.get('evaluation_id')
        matiere_id = data.get('matiere_id')
        periode = data.get('periode')
        
        # Validation des paramètres
        if not all([matiere_id, periode]):
            return JsonResponse({'success': False, 'error': 'Paramètres manquants (matière ou période)'}, status=400)
        
        # Récupérer la matière
        matiere = get_object_or_404(MatiereNote, pk=matiere_id)
        
        # Sécurité : Vérifier que la matière appartient à l'école de l'utilisateur
        user_profil = getattr(request.user, 'profil', None)
        ecole = user_profil.ecole if user_profil else None
        if ecole and matiere.classe.ecole != ecole:
            return JsonResponse({'success': False, 'error': 'Accès non autorisé'}, status=403)
        
        # Détecter si c'est une classe de maternelle (appréciations uniquement)
        from .calculs_moyennes import detecter_niveau_scolaire
        niveau_detecte = detecter_niveau_scolaire(matiere.classe.nom)
        est_maternelle = (niveau_detecte == 'MATERNELLE')
        
        # Vérifier si c'est une sauvegarde d'appréciations
        est_appreciation = any('appreciation' in note_data for note_data in notes_data)
        
        # Pour les appréciations (maternelle), on n'a pas besoin d'évaluation
        evaluation = None
        if not est_appreciation and not est_maternelle:
            # Créer l'évaluation si elle n'existe pas (pour les notes numériques uniquement)
            if not evaluation_id:
                # Déterminer le type d'évaluation selon la période
                if periode in ['OCTOBRE', 'NOVEMBRE', 'DECEMBRE', 'JANVIER', 'FEVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN']:
                    type_eval = 'DEVOIR'
                    titre_eval = f"Note {periode.capitalize()} - {matiere.nom}"
                elif periode.startswith('TRIMESTRE'):
                    type_eval = 'COMPOSITION'
                    titre_eval = f"Composition {periode.replace('_', ' ')} - {matiere.nom}"
                else:
                    type_eval = 'COMPOSITION'
                    titre_eval = f"Composition {periode.replace('_', ' ')} - {matiere.nom}"
                
                # Utiliser le coefficient de la matière, ou 1 par défaut si None
                coef = matiere.coefficient if matiere.coefficient is not None else 1
                
                # Utiliser filter().first() pour éviter MultipleObjectsReturned
                evaluation = Evaluation.objects.filter(
                    matiere=matiere,
                    periode=periode
                ).first()
                if not evaluation:
                    evaluation = Evaluation.objects.create(
                        matiere=matiere,
                        periode=periode,
                        titre=titre_eval,
                        type_evaluation=type_eval,
                        date_evaluation=timezone.now().date(),
                        note_sur=20 if matiere.classe.niveau_enseignement == 'SECONDAIRE' else 10,
                        coefficient=coef,
                        cree_par=request.user,
                    )
                    logger.info(f"Évaluation créée: {evaluation.id}")
                else:
                    logger.info(f"Évaluation récupérée: {evaluation.id}")
            else:
                evaluation = get_object_or_404(Evaluation, pk=evaluation_id)
        
        notes_sauvegardees = 0
        notes_modifiees = 0
        erreurs = []
        notes_details = []
        
        # Utiliser une transaction pour garantir l'intégrité des données
        with transaction.atomic():
            for note_data in notes_data:
                try:
                    eleve_id = note_data.get('eleve_id')
                    absent = note_data.get('absent', False)
                    
                    if not eleve_id:
                        continue
                    
                    # Nettoyer l'ID des caractères non numériques (espaces insécables, etc.)
                    import re
                    eleve_id_clean = re.sub(r'[^\d]', '', str(eleve_id))
                    if not eleve_id_clean:
                        erreurs.append(f"ID élève invalide: {eleve_id}")
                        continue
                    
                    eleve = Eleve.objects.get(pk=int(eleve_id_clean))
                    
                    # Traiter selon le type de note
                    if 'appreciation' in note_data:
                        # Appréciation (pour maternelle) - utiliser AppreciationMaternelle
                        from notes.models import AppreciationMaternelle
                        
                        appreciation = note_data.get('appreciation', '').strip()
                        commentaire = note_data.get('commentaire', '').strip()
                        
                        if not appreciation and not absent:
                            continue
                        
                        # Déterminer le trimestre depuis la période
                        trimestre = periode if periode.startswith('TRIMESTRE') else 'TRIMESTRE_1'
                        
                        note_obj, created = AppreciationMaternelle.objects.update_or_create(
                            eleve=eleve,
                            matiere=matiere,
                            trimestre=trimestre,
                            annee_scolaire=matiere.classe.annee_scolaire,
                            defaults={
                                'appreciation': appreciation if appreciation else None,
                                'commentaire': commentaire if commentaire else None,
                                'absent': absent,
                                'cree_par': request.user,
                            }
                        )
                    else:
                        # Note numérique
                        note_value = str(note_data.get('note', '')).strip()
                        
                        # Définir les périodes par type
                        periodes_mensuelles = ['OCTOBRE', 'NOVEMBRE', 'DECEMBRE', 'JANVIER', 'FEVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN']
                        periodes_trimestrielles = ['TRIMESTRE_1', 'TRIMESTRE_2', 'TRIMESTRE_3']
                        periodes_semestrielles = ['SEMESTRE_1', 'SEMESTRE_2']
                        
                        # Déterminer la note maximale
                        note_sur = 20
                        if matiere.classe.niveau_enseignement == 'PRIMAIRE':
                            note_sur = 10
                        
                        # Valider la note
                        note_decimal = None
                        if note_value and not absent:
                            try:
                                note_decimal = Decimal(note_value.replace(',', '.'))
                                if note_decimal < 0 or note_decimal > note_sur:
                                    erreurs.append(f"{eleve.nom} {eleve.prenom}: Note invalide (doit être entre 0 et {note_sur})")
                                    continue
                            except (InvalidOperation, ValueError, TypeError):
                                erreurs.append(f"{eleve.nom} {eleve.prenom}: Format de note invalide")
                                continue
                        elif not absent:
                            # Pas de note et pas absent, on ignore
                            continue
                        
                        # Sauvegarder dans le bon modèle selon la période
                        annee_scolaire = matiere.classe.annee_scolaire
                        
                        if periode.upper() in periodes_mensuelles:
                            # Sauvegarder dans NoteMensuelle
                            note_obj, created = NoteMensuelle.objects.update_or_create(
                                eleve=eleve,
                                matiere=matiere,
                                mois=periode.upper(),
                                annee_scolaire=annee_scolaire,
                                defaults={
                                    'note': note_decimal if not absent else Decimal('0'),
                                    'absent': absent,
                                    'cree_par': request.user,
                                }
                            )
                        elif periode.upper() in periodes_trimestrielles or periode.upper() in periodes_semestrielles:
                            # Sauvegarder dans CompositionNote
                            note_obj, created = CompositionNote.objects.update_or_create(
                                eleve=eleve,
                                matiere=matiere,
                                periode=periode.upper(),
                                annee_scolaire=annee_scolaire,
                                defaults={
                                    'note': note_decimal if not absent else Decimal('0'),
                                    'absent': absent,
                                    'cree_par': request.user,
                                }
                            )
                        else:
                            # Fallback: Sauvegarder dans NoteEleve via Evaluation
                            if evaluation is None:
                                erreurs.append(f"{eleve.nom} {eleve.prenom}: Période non reconnue et pas d'évaluation")
                                continue
                            note_obj, created = NoteEleve.objects.update_or_create(
                                eleve=eleve,
                                evaluation=evaluation,
                                defaults={
                                    'note': note_decimal if not absent else Decimal('0'),
                                    'absent': absent,
                                    'cree_par': request.user,
                                }
                            )
                    
                    if created:
                        notes_sauvegardees += 1
                    else:
                        notes_modifiees += 1
                    
                    # Ajouter les détails de la note sauvegardée
                    note_detail = {
                        'eleve_id': eleve_id,
                        'eleve_nom': f"{eleve.nom} {eleve.prenom}",
                        'absent': note_obj.absent,
                        'created': created
                    }
                    
                    # Ajouter les champs spécifiques selon le type d'objet
                    if hasattr(note_obj, 'appreciation'):
                        # AppreciationMaternelle
                        note_detail['appreciation'] = note_obj.appreciation
                        note_detail['note'] = None
                    else:
                        # NoteEleve
                        note_detail['note'] = float(note_obj.note) if note_obj.note else None
                        note_detail['appreciation'] = None
                    
                    notes_details.append(note_detail)
                    
                except Eleve.DoesNotExist:
                    erreurs.append(f"Élève ID {eleve_id} introuvable")
                except Exception as e:
                    logger.error(f"Erreur lors de la sauvegarde de la note pour l'élève {eleve_id}: {str(e)}")
                    erreurs.append(f"Erreur: {str(e)}")
        
        # Préparer la réponse
        total_notes = notes_sauvegardees + notes_modifiees
        message_parts = []
        
        if notes_sauvegardees > 0:
            message_parts.append(f"{notes_sauvegardees} nouvelle(s) note(s) ajoutée(s)")
        if notes_modifiees > 0:
            message_parts.append(f"{notes_modifiees} note(s) modifiée(s)")
        
        message = " et ".join(message_parts) if message_parts else "Aucune note à sauvegarder"
        
        response_data = {
            'success': True,
            'notes_sauvegardees': notes_sauvegardees,
            'notes_modifiees': notes_modifiees,
            'total': total_notes,
            'message': f'✅ {message}',
            'notes_details': notes_details,
            'evaluation_id': evaluation.id if evaluation else None
        }
        
        if erreurs:
            response_data['erreurs'] = erreurs
            response_data['message'] += f' ⚠️ {len(erreurs)} erreur(s) détectée(s)'
        
        logger.info(f"Sauvegarde terminée: {total_notes} notes traitées, {len(erreurs)} erreurs")
        
        # Invalider le cache des rangs/classements et des moyennes
        if total_notes > 0:
            try:
                from .utils_rangs import invalider_cache_rangs
                invalider_cache_rangs(matiere.classe, periode)
                # Invalider aussi le cache calculer_moyennes_classe_optimise
                # (pattern: moy_classe_<id>_<periode>_<system_type>_*)
                # Django locmem cache ne supporte pas delete_pattern,
                # on vide tout le préfixe via une clé sentinelle incrémentale
                _version_key = f"moy_version_classe_{matiere.classe.id}"
                cache.set(_version_key, cache.get(_version_key, 0) + 1, 86400)
                logger.info(f"Cache invalidé pour classe {matiere.classe.id}, période {periode}")
            except Exception as e:
                logger.warning(f"Erreur invalidation cache: {e}")
        
        return JsonResponse(response_data)
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Données JSON invalides'}, status=400)
    except Exception as e:
        logger.error(f"Erreur lors de la sauvegarde des notes: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Erreur serveur: {str(e)}'}, status=500)

@login_required
def supprimer_notes(request):
    """Supprimer les notes d'une évaluation ou d'une période spécifique.
    
    Gère tous les types de notes:
    - NoteMensuelle (notes mensuelles)
    - CompositionNote (compositions trimestrielles/semestrielles)
    - AppreciationMaternelle (appréciations maternelle)
    - NoteEleve (notes d'évaluation classiques)
    """
    from django.http import JsonResponse
    from eleves.models import Eleve
    from django.db import transaction
    from .models import NoteMensuelle, CompositionNote, AppreciationMaternelle
    import json
    import logging
    
    logger = logging.getLogger(__name__)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)
    
    try:
        data = json.loads(request.body)
        matiere_id = data.get('matiere_id')
        periode = data.get('periode')
        eleve_ids = data.get('eleve_ids', [])  # Liste optionnelle d'élèves spécifiques
        type_note = data.get('type_note', '')  # Type de note optionnel
        
        # Validation des paramètres
        if not all([matiere_id, periode]):
            return JsonResponse({'success': False, 'error': 'Paramètres manquants (matière ou période)'}, status=400)
        
        # Récupérer la matière
        matiere = get_object_or_404(MatiereNote, pk=matiere_id)

        # ── Sécurité: vérifier que la matière appartient à l'école de l'utilisateur ──
        user_profil = getattr(request.user, 'profil', None)
        ecole_user = user_profil.ecole if user_profil else None
        if ecole_user and matiere.classe.ecole != ecole_user:
            return JsonResponse({'success': False, 'error': 'Accès non autorisé'}, status=403)

        # Définir les périodes par type
        periodes_mensuelles = ['OCTOBRE', 'NOVEMBRE', 'DECEMBRE', 'JANVIER', 'FEVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN']
        periodes_trimestrielles = ['TRIMESTRE_1', 'TRIMESTRE_2', 'TRIMESTRE_3']
        periodes_semestrielles = ['SEMESTRE_1', 'SEMESTRE_2']
        
        notes_supprimees = 0
        details = []
        
        # Utiliser une transaction pour garantir l'intégrité des données
        with transaction.atomic():
            # 1. Supprimer les notes mensuelles si période mensuelle
            if periode.upper() in periodes_mensuelles:
                query = NoteMensuelle.objects.filter(matiere=matiere, mois=periode.upper())
                if eleve_ids:
                    query = query.filter(eleve_id__in=eleve_ids)
                count = query.count()
                if count > 0:
                    query.delete()
                    notes_supprimees += count
                    details.append(f"{count} note(s) mensuelle(s)")
                    logger.info(f"Suppression de {count} NoteMensuelle pour {matiere.nom}, mois {periode}")
            
            # 2. Supprimer les compositions si période trimestrielle/semestrielle
            if periode.upper() in periodes_trimestrielles or periode.upper() in periodes_semestrielles:
                query = CompositionNote.objects.filter(matiere=matiere, periode=periode.upper())
                if eleve_ids:
                    query = query.filter(eleve_id__in=eleve_ids)
                count = query.count()
                if count > 0:
                    query.delete()
                    notes_supprimees += count
                    details.append(f"{count} composition(s)")
                    logger.info(f"Suppression de {count} CompositionNote pour {matiere.nom}, période {periode}")
            
            # 3. Supprimer les appréciations maternelle si période trimestrielle
            if periode.upper() in periodes_trimestrielles:
                query = AppreciationMaternelle.objects.filter(matiere=matiere, trimestre=periode.upper())
                if eleve_ids:
                    query = query.filter(eleve_id__in=eleve_ids)
                count = query.count()
                if count > 0:
                    query.delete()
                    notes_supprimees += count
                    details.append(f"{count} appréciation(s) maternelle")
                    logger.info(f"Suppression de {count} AppreciationMaternelle pour {matiere.nom}, période {periode}")
            
            # 4. Supprimer les notes d'évaluation classiques
            evaluations = Evaluation.objects.filter(matiere=matiere, periode=periode)
            if evaluations.exists():
                notes_query = NoteEleve.objects.filter(evaluation__in=evaluations)
                if eleve_ids:
                    notes_query = notes_query.filter(eleve_id__in=eleve_ids)
                count = notes_query.count()
                if count > 0:
                    notes_query.delete()
                    notes_supprimees += count
                    details.append(f"{count} note(s) d'évaluation")
                    logger.info(f"Suppression de {count} NoteEleve pour {matiere.nom}, période {periode}")
        
        if notes_supprimees == 0:
            return JsonResponse({
                'success': False, 
                'error': f'Aucune note trouvée pour la matière "{matiere.nom}" et la période "{periode}"'
            }, status=404)
        
        return JsonResponse({
            'success': True,
            'message': f'✅ {notes_supprimees} note(s) supprimée(s) avec succès',
            'notes_supprimees': notes_supprimees,
            'details': details
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Données JSON invalides'}, status=400)
    except Exception as e:
        logger.error(f"Erreur lors de la suppression des notes: {str(e)}")
        return JsonResponse({'success': False, 'error': f'Erreur serveur: {str(e)}'}, status=500)

@login_required
def consulter_notes(request):
    """Consulter les notes - Vue complète par classe"""
    from decimal import Decimal
    
    user_profil = getattr(request.user, 'profil', None)
    ecole = user_profil.ecole if user_profil else None
    
    # Récupérer les classes
    if ecole:
        classes = ClasseNote.objects.filter(ecole=ecole, actif=True).order_by('niveau', 'nom')
    else:
        classes = ClasseNote.objects.filter(actif=True).order_by('niveau', 'nom')
    
    # Paramètres de sélection
    classe_id = request.GET.get('classe_id')
    periode_classement = request.GET.get('periode', '')
    
    classe_selectionnee = None
    matieres = []
    eleves_toutes_notes = []
    niveau_enseignement = 'SECONDAIRE'
    periodes_disponibles = []
    est_maternelle = False
    est_primaire = False
    
    if classe_id:
        classe_selectionnee = get_object_or_404(ClasseNote, pk=classe_id)
        niveau_enseignement = classe_selectionnee.niveau_enseignement
        matieres = MatiereNote.objects.filter(classe=classe_selectionnee, actif=True).order_by('nom')
        
        # Détecter le niveau scolaire pour adapter l'interface
        from .calculs_moyennes import detecter_niveau_scolaire
        niveau_detecte = detecter_niveau_scolaire(classe_selectionnee.nom)
        est_maternelle = (niveau_detecte == 'MATERNELLE')
        est_primaire = (niveau_detecte == 'PRIMAIRE')
        
        # Pour la maternelle : rediriger automatiquement vers TRIMESTRE_1 si pas de période
        if est_maternelle and not periode_classement:
            from django.shortcuts import redirect
            return redirect(f'/notes/consulter/?classe_id={classe_id}&periode=TRIMESTRE_1')
        
        # Déterminer les périodes disponibles selon le niveau
        if est_maternelle:
            # Pour la maternelle : uniquement les trimestres (appréciations)
            periodes_disponibles = [
                ('TRIMESTRE_1', '1er Trimestre'),
                ('TRIMESTRE_2', '2ème Trimestre'),
                ('TRIMESTRE_3', '3ème Trimestre'),
            ]
        else:
            # Pour les autres niveaux : toutes les périodes possibles
            periodes_disponibles = [
                # Mois
                ('OCTOBRE', 'Octobre'),
                ('NOVEMBRE', 'Novembre'),
                ('DECEMBRE', 'Décembre'),
                ('JANVIER', 'Janvier'),
                ('FEVRIER', 'Février'),
                ('MARS', 'Mars'),
                ('AVRIL', 'Avril'),
                ('MAI', 'Mai'),
                ('JUIN', 'Juin'),
                # Trimestres
                ('TRIMESTRE_1', '1er Trimestre'),
                ('TRIMESTRE_2', '2ème Trimestre'),
                ('TRIMESTRE_3', '3ème Trimestre'),
                # Semestres
                ('SEMESTRE_1', '1er Semestre'),
                ('SEMESTRE_2', '2ème Semestre'),
                # Résultat annuel
                ('ANNUEL_TRIM', 'Résultat Annuel (Trimestres)'),
                ('ANNUEL_SEM', 'Résultat Annuel (Semestres)'),
            ]

        # Récupérer les élèves
        try:
            # Mapping spécial pour les classes avec noms différents
            mapping_classes = {
                61: 56,  # ClasseNote '12ème Année' -> ClasseEleve '12ÈME ANNÉE'
                59: 8,   # ClasseNote '11ème Série littéraire' -> ClasseEleve '11ème série littéraire'
            }
            
            if classe_selectionnee.id in mapping_classes:
                classe_eleve = ClasseEleve.objects.filter(
                    id=mapping_classes[classe_selectionnee.id]
                ).first()
            else:
                # Utiliser filter().first() au lieu de get() pour éviter MultipleObjectsReturned
                classe_eleve = ClasseEleve.objects.filter(
                    nom=classe_selectionnee.nom,
                    annee_scolaire=classe_selectionnee.annee_scolaire,
                    ecole=classe_selectionnee.ecole  # Filtrer par l'école de la classe
                ).first()
            
            if classe_eleve:
                # OPTIMISATION: Pré-charger les relations pour éviter N+1
                eleves = Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').select_related('classe').order_by('prenom', 'nom')
            else:
                eleves = []
            
            # Utiliser directement les données centralisées pour garantir cohérence totale
            from .utils_rangs import calculer_rangs_classe_periode
            
            # Déterminer la période à utiliser
            if periode_classement:
                periode_pour_calcul = periode_classement
            else:
                # Pour la maternelle, utiliser TRIMESTRE_1 par défaut
                if est_maternelle:
                    periode_pour_calcul = 'TRIMESTRE_1'
                else:
                    periode_pour_calcul = 'OCTOBRE'
                    periodes_codes = [p[0] for p in periodes_disponibles]
                    if 'OCTOBRE' not in periodes_codes:
                        periode_pour_calcul = periodes_disponibles[0][0]
                # IMPORTANT: Mettre à jour periode_classement pour la recherche des notes
                periode_classement = periode_pour_calcul
            
            # OPTIMISATION: Calculer les rangs avec le cache activé pour performance
            rangs_dict = calculer_rangs_classe_periode(classe_selectionnee, periode_pour_calcul, use_cache=True)
            
            # Définir les listes de périodes pour la classification
            periodes_mensuelles = ['OCTOBRE', 'NOVEMBRE', 'DECEMBRE', 'JANVIER', 'FEVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN']
            periodes_trimestrielles = ['TRIMESTRE_1', 'TRIMESTRE_2', 'TRIMESTRE_3']
            periodes_semestrielles = ['SEMESTRE_1', 'SEMESTRE_2']
            periodes_annuelles = ['ANNUEL_TRIM', 'ANNUEL_SEM']

            # Import des modèles nécessaires (une seule fois)
            from .models import NoteMensuelle, CompositionNote, AppreciationMaternelle
            
            # OPTIMISATION v3.0: Pré-charger TOUTES les notes en une seule requête
            eleves_ids = list(eleves.values_list('id', flat=True))
            matieres_ids = list(matieres.values_list('id', flat=True))
            
            # Pré-charger les notes mensuelles
            notes_mensuelles_dict = {}
            if periode_classement in periodes_mensuelles:
                notes_mensuelles_qs = NoteMensuelle.objects.filter(
                    eleve_id__in=eleves_ids,
                    matiere_id__in=matieres_ids,
                    mois=periode_classement,
                    annee_scolaire=classe_selectionnee.annee_scolaire
                ).values('eleve_id', 'matiere_id', 'note', 'absent')
                for n in notes_mensuelles_qs:
                    notes_mensuelles_dict[(n['eleve_id'], n['matiere_id'])] = n
            
            # Pré-charger les compositions
            compositions_dict = {}
            if periode_classement in periodes_trimestrielles or periode_classement in periodes_semestrielles:
                compositions_qs = CompositionNote.objects.filter(
                    eleve_id__in=eleves_ids,
                    matiere_id__in=matieres_ids,
                    periode=periode_classement,
                    annee_scolaire=classe_selectionnee.annee_scolaire
                ).values('eleve_id', 'matiere_id', 'note', 'absent')
                for c in compositions_qs:
                    compositions_dict[(c['eleve_id'], c['matiere_id'])] = c
            
            # Pré-charger les appréciations maternelle (même logique que calculer_rangs_maternelle)
            appreciations_dict = {}
            if est_maternelle:
                # D'abord essayer avec l'année scolaire exacte - NE PAS filtrer par eleve_id
                appreciations_qs = AppreciationMaternelle.objects.filter(
                    matiere__in=matieres,
                    trimestre=periode_classement,
                    annee_scolaire=classe_selectionnee.annee_scolaire
                ).values('eleve_id', 'matiere_id', 'appreciation', 'commentaire', 'absent')
                
                # Si aucune appréciation trouvée, essayer sans filtre année scolaire
                if not appreciations_qs.exists():
                    appreciations_qs = AppreciationMaternelle.objects.filter(
                        matiere__in=matieres,
                        trimestre=periode_classement
                    ).values('eleve_id', 'matiere_id', 'appreciation', 'commentaire', 'absent')
                
                for a in appreciations_qs:
                    appreciations_dict[(a['eleve_id'], a['matiere_id'])] = a
            
            # Pour chaque élève, récupérer toutes ses notes/appréciations pour l'affichage
            for eleve in eleves:
                notes_par_matiere = {}
                
                # Récupérer les données centralisées
                rang_info = rangs_dict.get(eleve.id)
                moyenne_generale = float(rang_info['moyenne']) if rang_info else None
                rang = rang_info['rang'] if rang_info else '-'
                
                # Pour la maternelle : récupérer les appréciations au lieu des notes
                if est_maternelle:
                    for matiere in matieres:
                        # Créer une évaluation factice pour l'affichage
                        eval_factice = type('EvalFactice', (), {
                            'titre': periode_classement,
                            'periode': periode_classement,
                            'coefficient': 1,
                            'date_evaluation': None
                        })()
                        
                        notes_matiere = {
                            'evaluations': [eval_factice],
                            'notes': [],
                            'moyenne': None,
                            'appreciation': None
                        }
                        
                        # OPTIMISATION: Utiliser le dictionnaire pré-chargé
                        appreciation_data = appreciations_dict.get((eleve.id, matiere.id))
                        appreciation_value = appreciation_data['appreciation'] if appreciation_data else None
                        commentaire_value = appreciation_data['commentaire'] if appreciation_data else None
                        absent_value = appreciation_data['absent'] if appreciation_data else False
                        
                        # Ajouter l'appréciation trouvée
                        notes_matiere['notes'].append({
                            'evaluation': eval_factice,
                            'note': None,
                            'appreciation': appreciation_value,
                            'appreciation_display': dict(AppreciationMaternelle.APPRECIATION_CHOICES).get(appreciation_value, '-') if appreciation_value else '-',
                            'commentaire': commentaire_value,
                            'absent': absent_value,
                        })
                        
                        notes_matiere['appreciation'] = appreciation_value
                        notes_par_matiere[matiere.id] = notes_matiere
                    
                    # Pour la maternelle, utiliser les données du calcul des rangs (pourcentage d'acquisition)
                    rang_info = rangs_dict.get(eleve.id)
                    if rang_info:
                        moyenne_generale = float(rang_info['moyenne'])  # Pourcentage d'acquisition
                        rang = rang_info['rang']
                    else:
                        moyenne_generale = None
                        rang = '-'
                
                elif periode_classement in periodes_mensuelles:
                    # Système mensuel - OPTIMISATION: utiliser le dictionnaire pré-chargé
                    for matiere in matieres:
                        # Créer une évaluation factice pour l'affichage
                        eval_factice = type('EvalFactice', (), {
                            'titre': periode_classement,
                            'periode': periode_classement,
                            'coefficient': 1,
                            'date_evaluation': None
                        })()
                        
                        notes_matiere = {
                            'evaluations': [eval_factice],
                            'notes': [],
                            'moyenne': None
                        }
                        
                        # OPTIMISATION: Utiliser le dictionnaire pré-chargé
                        note_data = notes_mensuelles_dict.get((eleve.id, matiere.id))
                        note_value = note_data['note'] if note_data else None
                        absent_value = note_data['absent'] if note_data else False
                        
                        # Ajouter la note trouvée ou une note vide
                        notes_matiere['notes'].append({
                            'evaluation': eval_factice,
                            'note': note_value,
                            'absent': absent_value,
                        })
                        
                        # Calculer la moyenne pour cette matière
                        if note_value is not None and not absent_value:
                            notes_matiere['moyenne'] = float(note_value)
                        else:
                            notes_matiere['moyenne'] = 0.0
                        
                        notes_par_matiere[matiere.id] = notes_matiere
                elif periode_classement in periodes_annuelles:
                    # Résultat annuel - moyenne de chaque matière sur les 3 trimestres ou 2 semestres
                    from .calculs_moyennes import calculer_moyenne_annuelle_matiere
                    system_type_annuel = 'annuel_trimestriel' if periode_classement == 'ANNUEL_TRIM' else 'annuel_semestriel'
                    for matiere in matieres:
                        eval_factice = type('EvalFactice', (), {
                            'titre': 'Moy. Annuelle',
                            'periode': periode_classement,
                            'coefficient': 1,
                            'date_evaluation': None
                        })()

                        notes_matiere = {
                            'evaluations': [eval_factice],
                            'notes': [],
                            'moyenne': None
                        }

                        result_matiere = calculer_moyenne_annuelle_matiere(eleve, matiere, system_type_annuel)
                        moyenne_annuelle_matiere = result_matiere.get('moyenne_annuelle')

                        notes_matiere['notes'].append({
                            'evaluation': eval_factice,
                            'note': moyenne_annuelle_matiere,
                            'absent': False,
                        })

                        notes_matiere['moyenne'] = moyenne_annuelle_matiere if moyenne_annuelle_matiere is not None else 0.0

                        notes_par_matiere[matiere.id] = notes_matiere
                else:
                    # Système trimestriel/semestriel - OPTIMISATION: utiliser le dictionnaire pré-chargé
                    for matiere in matieres:
                        # Créer une évaluation factice pour l'affichage
                        eval_factice = type('EvalFactice', (), {
                            'titre': periode_classement,
                            'periode': periode_classement,
                            'coefficient': 1,
                            'date_evaluation': None
                        })()

                        notes_matiere = {
                            'evaluations': [eval_factice],
                            'notes': [],
                            'moyenne': None
                        }

                        # OPTIMISATION: Utiliser le dictionnaire pré-chargé
                        compo_data = compositions_dict.get((eleve.id, matiere.id))
                        note_value = compo_data['note'] if compo_data else None
                        absent_value = compo_data['absent'] if compo_data else False

                        # Ajouter la note trouvée ou une note vide
                        notes_matiere['notes'].append({
                            'evaluation': eval_factice,
                            'note': note_value,
                            'absent': absent_value,
                        })

                        # Calculer la moyenne pour cette matière
                        if note_value is not None and not absent_value:
                            notes_matiere['moyenne'] = float(note_value)
                        else:
                            notes_matiere['moyenne'] = 0.0

                        notes_par_matiere[matiere.id] = notes_matiere

                eleves_toutes_notes.append({
                    'eleve': eleve,
                    'notes_par_matiere': notes_par_matiere,
                    'moyenne_generale': moyenne_generale,  # Utiliser la moyenne centralisée
                    'rang': rang,  # Utiliser le rang centralisé
                })
                
        except ClasseEleve.DoesNotExist:
            pass
    
    # Trier par rang pour l'affichage (élèves avec rang d'abord)
    eleves_avec_rang = [e for e in eleves_toutes_notes if e['rang'] != '-']
    eleves_sans_rang = [e for e in eleves_toutes_notes if e['rang'] == '-']
    eleves_avec_rang.sort(key=lambda x: x['moyenne_generale'], reverse=True)
    eleves_toutes_notes = eleves_avec_rang + eleves_sans_rang
    
    # Récupérer les évaluations pour les en-têtes (simplifiées)
    evaluations_par_matiere = {}
    periodes_mensuelles_list = ['OCTOBRE', 'NOVEMBRE', 'DECEMBRE', 'JANVIER', 'FEVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN']
    periodes_trimestrielles_list = ['TRIMESTRE_1', 'TRIMESTRE_2', 'TRIMESTRE_3']
    periodes_semestrielles_list = ['SEMESTRE_1', 'SEMESTRE_2']
    periodes_annuelles_list = ['ANNUEL_TRIM', 'ANNUEL_SEM']

    for matiere in matieres:
        if periode_classement in periodes_mensuelles_list:
            # Pour les mois: une seule colonne "Note"
            eval_factice = type('EvalFactice', (), {
                'id': f'mensuel_{matiere.id}',
                'titre': 'Note',
                'periode': periode_classement,
                'coefficient': 1,
                'date_evaluation': None
            })()
            evaluations_par_matiere[matiere.id] = [eval_factice]
        elif periode_classement in periodes_trimestrielles_list or periode_classement in periodes_semestrielles_list:
            # Pour les trimestres/semestres: une seule colonne "Compo"
            eval_factice = type('EvalFactice', (), {
                'id': f'compo_{matiere.id}',
                'titre': 'Compo',
                'periode': periode_classement,
                'coefficient': 1,
                'date_evaluation': None
            })()
            evaluations_par_matiere[matiere.id] = [eval_factice]
        elif periode_classement in periodes_annuelles_list:
            # Pour le résultat annuel: une seule colonne "Moy. Annuelle"
            eval_factice = type('EvalFactice', (), {
                'id': f'annuel_{matiere.id}',
                'titre': 'Moy. Annuelle',
                'periode': periode_classement,
                'coefficient': 1,
                'date_evaluation': None
            })()
            evaluations_par_matiere[matiere.id] = [eval_factice]
        else:
            # Par défaut: une colonne générique
            eval_factice = type('EvalFactice', (), {
                'id': f'note_{matiere.id}',
                'titre': 'Note',
                'periode': periode_classement or 'OCTOBRE',
                'coefficient': 1,
                'date_evaluation': None
            })()
            evaluations_par_matiere[matiere.id] = [eval_factice]
    
    # Garder la période sélectionnée (ne pas écraser avec la première période)
    periode_selectionnee = request.GET.get('periode', '')
    if not periode_selectionnee and periodes_disponibles:
        periode_selectionnee = periodes_disponibles[0][0]  # Code de la première période
    
    context = {
        'titre_page': 'Consultation des Notes',
        'classes': classes,
        'classe_selectionnee': classe_selectionnee,
        'matieres': matieres,
        'periodes_disponibles': periodes_disponibles,
        'periode_classement': periode_selectionnee,
        'periode_selectionnee': periode_selectionnee,  # Alias pour le template
        'eleves_toutes_notes': eleves_toutes_notes,
        'evaluations_par_matiere': evaluations_par_matiere,
        'niveau_enseignement': niveau_enseignement,
        'est_maternelle': est_maternelle,
        'est_primaire': est_primaire,
    }
    
    return render(request, 'notes/consulter_notes.html', context)

@login_required
def bulletin_dynamique(request):
    """Bulletin dynamique - Génération de bulletins personnalisés"""
    from decimal import Decimal
    from django.contrib import messages
    
    user_profil = getattr(request.user, 'profil', None)
    ecole = user_profil.ecole if user_profil else None
    
    # Récupérer les classes
    if ecole:
        classes = ClasseNote.objects.filter(ecole=ecole, actif=True).order_by('niveau', 'nom')
    else:
        classes = ClasseNote.objects.filter(actif=True).order_by('niveau', 'nom')
    
    # Paramètres de sélection avec nettoyage
    def nettoyer_parametre_numerique(param):
        """Nettoie un paramètre numérique en supprimant les espaces et caractères invalides"""
        if not param:
            return None
        
        # Convertir en string si nécessaire
        if not isinstance(param, str):
            param = str(param)
        
        # Remplacer tous les types d'espaces (y compris les espaces insécables) par rien
        import re
        param_nettoye = re.sub(r'[\s\u00A0\u2000-\u200F\u2028-\u202F\u205F\u3000]', '', param)
        
        # Supprimer les caractères non numériques sauf le signe moins
        param_nettoye = ''.join(c for c in param_nettoye if c.isdigit() or c == '-')
        
        try:
            return int(param_nettoye) if param_nettoye else None
        except ValueError:
            return None
    
    def valider_et_corriger_eleve_id(eleve_id, classe_id):
        """Valide et corrige l'ID de l'élève si nécessaire"""
        if not eleve_id or not isinstance(eleve_id, int):
            return None
        
        # Essayer de trouver l'élève avec l'ID donné
        try:
            return Eleve.objects.get(pk=eleve_id)
        except Eleve.DoesNotExist:
            # Si l'élève n'existe pas, essayer des variations
            # Cas spécial : si l'ID ressemble à 1xxx mais qu'on a 1 xxx (espace)
            if eleve_id > 1000 and eleve_id < 20000:
                # Essayer avec un zéro supplémentaire
                with_zero = int(f"1{str(eleve_id)[1:]}")
                try:
                    return Eleve.objects.get(pk=with_zero)
                except Eleve.DoesNotExist:
                    pass
            
            # Essayer avec des zéros devant
            padded_id = int(f"{eleve_id:05d}")  # Au moins 5 chiffres
            try:
                return Eleve.objects.get(pk=padded_id)
            except Eleve.DoesNotExist:
                pass
            
            return None
    
    classe_id = nettoyer_parametre_numerique(request.GET.get('classe_id'))
    eleve_id = nettoyer_parametre_numerique(request.GET.get('eleve_id'))
    periode = request.GET.get('periode', '').strip()
    system_type = request.GET.get('system_type', 'trimestre').strip()  # mensuel, trimestre, semestre, annuel_trimestriel, annuel_semestriel
    
    # Valider que la période correspond au system_type sélectionné
    # Si la période ne correspond pas, la réinitialiser
    if periode:
        periodes_valides = {
            'mensuel': ['OCTOBRE', 'NOVEMBRE', 'DECEMBRE', 'JANVIER', 'FEVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN'],
            'trimestre': ['TRIMESTRE_1', 'TRIMESTRE_2', 'TRIMESTRE_3'],
            'semestre': ['SEMESTRE_1', 'SEMESTRE_2'],
            'annuel_trimestriel': ['ANNUEL_TRIM'],
            'annuel_semestriel': ['ANNUEL_SEM'],
            'annuel': ['ANNUEL_TRIM'],
        }
        if system_type in periodes_valides and periode not in periodes_valides.get(system_type, []):
            periode = ''  # Réinitialiser la période si elle ne correspond pas au system_type
    
    classe_selectionnee = None
    eleves = []
    eleve_selectionne = None
    matieres = []
    niveau_enseignement = 'SECONDAIRE'
    periodes_disponibles = []
    bulletin_data = None
    
    if classe_id:
        classe_selectionnee = get_object_or_404(ClasseNote, pk=classe_id)
        # Utiliser detecter_niveau_scolaire pour une détection cohérente avec les bulletins PDF
        from .calculs_moyennes import detecter_niveau_scolaire
        niveau_enseignement = detecter_niveau_scolaire(classe_selectionnee.nom)
        matieres = MatiereNote.objects.filter(classe=classe_selectionnee, actif=True).order_by('nom')
        
        # Déterminer les périodes disponibles selon le système
        if system_type == 'mensuel':
            periodes_disponibles = [
                ('OCTOBRE', 'Octobre'),
                ('NOVEMBRE', 'Novembre'),
                ('DECEMBRE', 'Décembre'),
                ('JANVIER', 'Janvier'),
                ('FEVRIER', 'Février'),
                ('MARS', 'Mars'),
                ('AVRIL', 'Avril'),
                ('MAI', 'Mai'),
                ('JUIN', 'Juin'),
            ]
        elif system_type == 'trimestre':
            periodes_disponibles = [
                ('TRIMESTRE_1', '1er Trimestre'),
                ('TRIMESTRE_2', '2ème Trimestre'),
                ('TRIMESTRE_3', '3ème Trimestre'),
            ]
        elif system_type == 'semestre':
            periodes_disponibles = [
                ('SEMESTRE_1', '1er Semestre'),
                ('SEMESTRE_2', '2ème Semestre'),
            ]
        elif system_type == 'annuel_trimestriel':
            # Bulletin annuel basé sur les trimestres (T1+T2+T3)/3
            periodes_disponibles = [
                ('ANNUEL_TRIM', 'Bulletin Annuel (Trimestres)'),
            ]
        elif system_type == 'annuel_semestriel':
            # Bulletin annuel basé sur les semestres (S1+S2)/2
            periodes_disponibles = [
                ('ANNUEL_SEM', 'Bulletin Annuel (Semestres)'),
            ]
        elif system_type == 'annuel':
            # Ancien système annuel (redirige vers trimestriel par défaut)
            periodes_disponibles = [
                ('ANNUEL_TRIM', 'Bulletin Annuel (Trimestres)'),
            ]
        
        # Récupérer les élèves de la classe
        # Essayer d'abord une correspondance exacte
        classe_eleve = ClasseEleve.objects.filter(
            nom=classe_selectionnee.nom,
            annee_scolaire=classe_selectionnee.annee_scolaire,
            ecole=classe_selectionnee.ecole
        ).first()
        
        if not classe_eleve:
            # Si pas de correspondance exacte, essayer avec insensibilité à la casse
            classe_eleve = ClasseEleve.objects.filter(
                nom__iexact=classe_selectionnee.nom,
                annee_scolaire=classe_selectionnee.annee_scolaire,
                ecole=classe_selectionnee.ecole
            ).first()
        
        if classe_eleve:
            eleves = Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('prenom', 'nom')
        else:
            try:
                # Recherche approximative
                pass
                eleves = Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('prenom', 'nom')
            except (ClasseEleve.DoesNotExist, ClasseEleve.MultipleObjectsReturned):
                eleves = []
        
        # Préparer un bulletin vide avec toutes les matières (dès la sélection de la classe)
        if matieres:
            # Déterminer le titre de la période
            titre_periode = ''
            for code, libelle in periodes_disponibles:
                if code == periode:
                    titre_periode = libelle
                    break
            
            # Déterminer le type de bulletin
            type_bulletin = system_type
            
            bulletin_data = {
                'eleve': None,
                'classe': classe_selectionnee,
                'periode': periode,
                'system_type': system_type,
                'type_bulletin': type_bulletin,
                'titre_periode': titre_periode,
                'titre_moyenne': 'Moyenne Continue',
                'titre_composition': 'Composition',
                'matieres_notes': [],
                'moyenne_generale': None,
                'rang': None,
                'mention': None,
                'appreciation': '',
                'appreciation_generale': '',
                'effectif': len(eleves),
                'mois_libelle': titre_periode,
            }
            
            # Initialiser eleve_selectionne
            eleve_selectionne = None
            
            # Si un élève est sélectionné
            if eleve_id and isinstance(eleve_id, int) and eleve_id > 0:
                eleve_selectionne = valider_et_corriger_eleve_id(eleve_id, classe_id)
                if eleve_selectionne:
                    bulletin_data['eleve'] = eleve_selectionne
                    # Si l'ID a été corrigé, informer l'utilisateur
                    if eleve_selectionne.pk != eleve_id:
                        messages.info(request, f"ID d'élève corrigé : {eleve_id} → {eleve_selectionne.pk}")
                else:
                    # L'élève n'existe pas, réinitialiser la sélection
                    eleve_id = None
                    messages.warning(request, f"L'élève avec l'ID {eleve_id} n'a pas été trouvé.")
            
            # Détecter le niveau scolaire
            from .calculs_moyennes import detecter_niveau_scolaire
            niveau_detecte = detecter_niveau_scolaire(classe_selectionnee.nom)
            est_maternelle = (niveau_detecte == 'MATERNELLE')
            est_primaire = (niveau_detecte == 'PRIMAIRE')
            
            # Pour chaque matière, préparer la structure
            total_points = Decimal('0')
            total_coefficients = Decimal('0')
            
            # NOUVEAU: Totaux par colonne pour les détails trimestriels/semestriels
            totaux_colonnes = {
                'octobre': Decimal('0'),
                'novembre': Decimal('0'), 
                'decembre': Decimal('0'),
                'janvier': Decimal('0'),
                'fevrier': Decimal('0'),
                'mars': Decimal('0'),
                'avril': Decimal('0'),
                'mai': Decimal('0'),
                'juin': Decimal('0'),
                'moyenne_continue': Decimal('0'),
                'composition': Decimal('0'),
                'count_moyenne_continue': 0,
                'count_composition': 0
            }
            
            # Calculer les totaux des coefficients (même sans élève sélectionné)
            # Pour la maternelle, les coefficients sont None, donc on utilise 1 par défaut
            for matiere in matieres:
                coef = matiere.coefficient if matiere.coefficient is not None else Decimal('1')
                total_coefficients += coef
            
            for matiere in matieres:
                # Initialiser les variables
                moyenne_continue = None
                note_composition = None
                moyennes_mensuelles = []
                appreciation_maternelle = None
                commentaire_maternelle = None
                absent_maternelle = False
                
                # Pour la maternelle : récupérer les appréciations au lieu des notes
                if est_maternelle and periode and eleve_selectionne:
                    from .models import AppreciationMaternelle
                    try:
                        appreciation_obj = AppreciationMaternelle.objects.get(
                            eleve=eleve_selectionne,
                            matiere=matiere,
                            trimestre=periode,
                            annee_scolaire=classe_selectionnee.annee_scolaire
                        )
                        appreciation_maternelle = appreciation_obj.appreciation
                        commentaire_maternelle = appreciation_obj.commentaire
                        absent_maternelle = appreciation_obj.absent
                    except AppreciationMaternelle.DoesNotExist:
                        pass
                
                # Si une période est sélectionnée, récupérer les données (pour les autres niveaux)
                elif periode and eleve_selectionne:
                    # ============================================================
                    # UTILISER LA FONCTION CENTRALISÉE INTELLIGENTE
                    # Elle gère tous les types: mensuel, trimestre, semestre, annuel
                    # ============================================================
                    from .calculs_moyennes import calculer_bulletin_intelligent, calculer_moyenne_periode_guineenne
                    
                    result_intelligent = calculer_bulletin_intelligent(
                        eleve_selectionne, matiere, periode, system_type
                    )
                    
                    moyenne_continue = result_intelligent['moyenne_continue']
                    note_composition = result_intelligent['note_composition']
                    moyennes_mensuelles = result_intelligent['moyennes_mensuelles']
                
                # Calculer la moyenne de la matière selon le système guinéen
                # Secondaire: moyenne = 40% cours + 60% composition. Primaire: composition.
                # Si mensuel: moyenne = moyenne_continue uniquement (pas de composition)
                # Si annuel: moyenne = moyenne des périodes (T1+T2+T3)/3 ou (S1+S2)/2
                moyenne_matiere = None
                
                if system_type in ['annuel_trimestriel', 'annuel_semestriel']:
                    # Pour les bulletins annuels, moyenne_continue contient déjà la moyenne annuelle
                    moyenne_matiere = moyenne_continue
                elif system_type == 'mensuel':
                    moyenne_matiere = moyenne_continue
                elif moyenne_continue is not None and note_composition is not None:
                    # Formule guinéenne centralisée.
                    moyenne_matiere = round(calculer_moyenne_periode_guineenne(
                        moyenne_continue,
                        note_composition,
                        'PRIMAIRE' if est_primaire else 'SECONDAIRE'
                    ), 2)
                elif note_composition is not None:
                    # Seulement la composition
                    moyenne_matiere = note_composition
                elif moyenne_continue is not None:
                    # Seulement la moyenne continue
                    moyenne_matiere = moyenne_continue
                
                # Calculer les points
                # PRIMAIRE: Pas de coefficients (tous égaux à 1)
                # SECONDAIRE (Collège/Lycée): Avec coefficients
                # RÈGLE PÉDAGOGIQUE: Toutes les matières comptent (sans notes = 0)
                if est_primaire:
                    # Primaire: coefficient = 1 pour toutes les matières
                    coefficient_effectif = Decimal('1')
                else:
                    # Secondaire: utiliser le coefficient de la matière
                    coefficient_effectif = matiere.coefficient if matiere.coefficient and matiere.coefficient > 0 else Decimal('1')
                
                # Si pas de notes, compter comme 0 (ne pas favoriser l'élève)
                moyenne_matiere_calcul = moyenne_matiere if moyenne_matiere is not None else 0.0
                points = round(moyenne_matiere_calcul * float(coefficient_effectif), 2)
                total_points += Decimal(str(moyenne_matiere_calcul)) * coefficient_effectif
                
                # Préparer les notes pour l'affichage
                notes_matiere = []
                if system_type in ['annuel_trimestriel', 'annuel_semestriel']:
                    # Bulletin annuel: afficher les moyennes par période
                    if moyennes_mensuelles:
                        for moy_periode in moyennes_mensuelles:
                            notes_matiere.append({
                                'note': moy_periode['moyenne'],
                                'absent': moy_periode['absent'],
                                'libelle': moy_periode['libelle'],
                                'type': 'periode'
                            })
                    
                    # Ajouter la moyenne annuelle
                    notes_matiere.append({
                        'note': moyenne_matiere,
                        'absent': False,
                        'libelle': 'Moy. Annuelle',
                        'type': 'annuelle'
                    })
                    
                elif system_type in ['trimestre', 'semestre']:
                    # NOUVEAU: Inclure les moyennes mensuelles détaillées
                    if moyennes_mensuelles:
                        # Ajouter les moyennes mensuelles
                        for moy_mens in moyennes_mensuelles:
                            notes_matiere.append({
                                'note': moy_mens['moyenne'],
                                'absent': moy_mens['absent'],
                                'libelle': moy_mens['libelle'],
                                'type': 'mensuelle'
                            })
                    
                    # Ajouter la moyenne continue calculée
                    notes_matiere.append({
                        'note': moyenne_continue,
                        'absent': False,
                        'libelle': 'Moy. Continue',
                        'type': 'continue'
                    })
                    
                    # Ajouter la composition
                    notes_matiere.append({
                        'note': note_composition,
                        'absent': False,
                        'libelle': 'Composition',
                        'type': 'composition'
                    })
                    
                elif system_type == 'mensuel':
                    notes_matiere = [
                        {'note': moyenne_continue, 'absent': False, 'libelle': 'Moyenne', 'type': 'mensuelle'}
                    ]
                
                # Déterminer le coefficient à afficher
                # PRIMAIRE: coefficient = 1 (pas de pondération)
                # SECONDAIRE: coefficient réel de la matière
                coefficient_affiche = 1 if est_primaire else (matiere.coefficient if matiere.coefficient else 1)
                
                bulletin_data['matieres_notes'].append({
                    'matiere': matiere,
                    'notes': notes_matiere,
                    'moyennes_mensuelles': moyennes_mensuelles,  # NOUVEAU: Détails mensuels
                    'moyenne_continue': moyenne_continue,
                    'note_composition': note_composition,
                    'moyenne': moyenne_matiere,
                    'coefficient': coefficient_affiche,
                    'points': points,
                    'total': points,  # Alias pour compatibilité
                    # Données spécifiques maternelle
                    'appreciation': appreciation_maternelle,
                    'commentaire': commentaire_maternelle,
                    'absent': absent_maternelle,
                    # Indicateur de niveau pour le template
                    'est_primaire': est_primaire,
                })
                
                # NOUVEAU: Accumuler les totaux par colonne
                if system_type in ['trimestre', 'semestre'] and moyennes_mensuelles:
                    # Ajouter les moyennes mensuelles aux totaux
                    for moy_mens in moyennes_mensuelles:
                        if not moy_mens['absent'] and moy_mens['moyenne'] is not None:
                            mois_lower = moy_mens['libelle'].lower()
                            if 'oct' in mois_lower:
                                totaux_colonnes['octobre'] += Decimal(str(moy_mens['moyenne']))
                            elif 'nov' in mois_lower:
                                totaux_colonnes['novembre'] += Decimal(str(moy_mens['moyenne']))
                            elif 'déc' in mois_lower or 'dec' in mois_lower:
                                totaux_colonnes['decembre'] += Decimal(str(moy_mens['moyenne']))
                            elif 'jan' in mois_lower:
                                totaux_colonnes['janvier'] += Decimal(str(moy_mens['moyenne']))
                            elif 'fév' in mois_lower or 'fev' in mois_lower:
                                totaux_colonnes['fevrier'] += Decimal(str(moy_mens['moyenne']))
                            elif 'mar' in mois_lower:
                                totaux_colonnes['mars'] += Decimal(str(moy_mens['moyenne']))
                            elif 'avr' in mois_lower or 'avr' in mois_lower:
                                totaux_colonnes['avril'] += Decimal(str(moy_mens['moyenne']))
                            elif 'mai' in mois_lower:
                                totaux_colonnes['mai'] += Decimal(str(moy_mens['moyenne']))
                            elif 'jui' in mois_lower and 'juin' in mois_lower:
                                totaux_colonnes['juin'] += Decimal(str(moy_mens['moyenne']))
                    
                    # Ajouter la moyenne continue
                    if moyenne_continue is not None:
                        totaux_colonnes['moyenne_continue'] += Decimal(str(moyenne_continue))
                        totaux_colonnes['count_moyenne_continue'] += 1
                    
                    # Ajouter la composition
                    if note_composition is not None:
                        totaux_colonnes['composition'] += Decimal(str(note_composition))
                        totaux_colonnes['count_composition'] += 1
                        
                elif system_type == 'mensuel' and moyenne_continue is not None:
                    # Pour les bulletins mensuels, ajouter à la moyenne continue
                    totaux_colonnes['moyenne_continue'] += Decimal(str(moyenne_continue))
                    totaux_colonnes['count_moyenne_continue'] += 1
            
            # Ajouter les totaux au bulletin
            bulletin_data['total_points'] = round(float(total_points), 2) if total_points > 0 else None
            bulletin_data['total_coefficients'] = float(total_coefficients) if total_coefficients > 0 else None
            
            # NOUVEAU: Ajouter les totaux par colonne
            bulletin_data['totaux_colonnes'] = {}
            
            # Calculer les moyennes pour les totaux par colonne
            if totaux_colonnes['count_moyenne_continue'] > 0:
                bulletin_data['totaux_colonnes']['moyenne_continue'] = round(
                    float(totaux_colonnes['moyenne_continue'] / totaux_colonnes['count_moyenne_continue']), 2
                )
            else:
                bulletin_data['totaux_colonnes']['moyenne_continue'] = None
                
            if totaux_colonnes['count_composition'] > 0:
                bulletin_data['totaux_colonnes']['composition'] = round(
                    float(totaux_colonnes['composition'] / totaux_colonnes['count_composition']), 2
                )
            else:
                bulletin_data['totaux_colonnes']['composition'] = None
            
            # Ajouter les totaux mensuels (sans moyenne, juste la somme)
            bulletin_data['totaux_colonnes']['octobre'] = round(float(totaux_colonnes['octobre']), 2) if totaux_colonnes['octobre'] > 0 else None
            bulletin_data['totaux_colonnes']['novembre'] = round(float(totaux_colonnes['novembre']), 2) if totaux_colonnes['novembre'] > 0 else None
            bulletin_data['totaux_colonnes']['decembre'] = round(float(totaux_colonnes['decembre']), 2) if totaux_colonnes['decembre'] > 0 else None
            bulletin_data['totaux_colonnes']['janvier'] = round(float(totaux_colonnes['janvier']), 2) if totaux_colonnes['janvier'] > 0 else None
            bulletin_data['totaux_colonnes']['fevrier'] = round(float(totaux_colonnes['fevrier']), 2) if totaux_colonnes['fevrier'] > 0 else None
            bulletin_data['totaux_colonnes']['mars'] = round(float(totaux_colonnes['mars']), 2) if totaux_colonnes['mars'] > 0 else None
            bulletin_data['totaux_colonnes']['avril'] = round(float(totaux_colonnes['avril']), 2) if totaux_colonnes['avril'] > 0 else None
            bulletin_data['totaux_colonnes']['mai'] = round(float(totaux_colonnes['mai']), 2) if totaux_colonnes['mai'] > 0 else None
            bulletin_data['totaux_colonnes']['juin'] = round(float(totaux_colonnes['juin']), 2) if totaux_colonnes['juin'] > 0 else None
            
            # Calculer la moyenne générale et le rang
            if eleve_selectionne:
                from decimal import Decimal as _Dec
                from .utils_rangs import calculer_rangs_classe_periode
                
                # Pour la maternelle : utiliser le système de calcul basé sur les appréciations
                if est_maternelle and periode:
                    rangs_dict = calculer_rangs_classe_periode(classe_selectionnee, periode, use_cache=False)
                    rang_info = rangs_dict.get(eleve_selectionne.id)
                    
                    if rang_info:
                        # Taux d'acquisition en pourcentage
                        taux_acquisition = float(rang_info['moyenne'])
                        bulletin_data['moyenne_generale'] = taux_acquisition
                        bulletin_data['rang'] = f"{rang_info['rang']}/{rang_info['total_eleves']}"
                        
                        # Appréciation basée sur le taux d'acquisition
                        if taux_acquisition >= 90:
                            bulletin_data['mention'] = 'Excellent'
                        elif taux_acquisition >= 75:
                            bulletin_data['mention'] = 'Très Bien'
                        elif taux_acquisition >= 60:
                            bulletin_data['mention'] = 'Bien'
                        elif taux_acquisition >= 50:
                            bulletin_data['mention'] = 'Assez Bien'
                        else:
                            bulletin_data['mention'] = 'À encourager'
                        
                        bulletin_data['appreciation'] = f"Bon trimestre {eleve_selectionne.prenom}. Continue ainsi !"
                    else:
                        bulletin_data['moyenne_generale'] = 0
                        bulletin_data['rang'] = "-"
                        bulletin_data['mention'] = 'Non évalué'
                        bulletin_data['appreciation'] = ''
                    
                    # Récupérer les analyses et recommandations du BulletinMaternelle
                    from .models import BulletinMaternelle
                    bulletin_maternelle = BulletinMaternelle.objects.filter(
                        eleve=eleve_selectionne,
                        classe=classe_selectionnee,
                        trimestre=periode,
                        annee_scolaire=classe_selectionnee.annee_scolaire
                    ).first()
                    
                    if bulletin_maternelle:
                        # Récupérer les analyses sélectionnées
                        analyses_codes = bulletin_maternelle.analyses or []
                        analyses_dict = dict(BulletinMaternelle.ANALYSES_CHOICES)
                        bulletin_data['analyses_selectionnees'] = [analyses_dict.get(code, code) for code in analyses_codes]
                        
                        # Récupérer les recommandations sélectionnées
                        recommandations_codes = bulletin_maternelle.recommandations or []
                        recommandations_dict = dict(BulletinMaternelle.RECOMMANDATIONS_CHOICES)
                        bulletin_data['recommandations_selectionnees'] = [recommandations_dict.get(code, code) for code in recommandations_codes]
                        
                        # Appréciation générale personnalisée
                        if bulletin_maternelle.appreciation_generale:
                            bulletin_data['appreciation'] = bulletin_maternelle.appreciation_generale
                    else:
                        bulletin_data['analyses_selectionnees'] = []
                        bulletin_data['recommandations_selectionnees'] = []
                
                # Pour les autres niveaux : calcul classique
                elif total_coefficients > 0:
                    moyenne_generale = round(float(total_points / total_coefficients), 2)
                    bulletin_data['moyenne_generale'] = moyenne_generale
                    # Source unique: recalcul centralise des rangs et moyennes.
                    # Le modele Classement peut etre un ancien instantane; il ne doit
                    # pas remplacer le calcul courant du bulletin.
                    if periode:
                        rangs_dict = calculer_rangs_classe_periode(classe_selectionnee, periode, use_cache=False)

                        rang_info = rangs_dict.get(eleve_selectionne.id)
                        if rang_info:
                            moyenne_eleve = float(rang_info['moyenne'])
                            moyenne_dec = _Dec(str(moyenne_eleve))
                            bulletin_data['moyenne_generale'] = moyenne_eleve
                            bulletin_data['mention'] = obtenir_mention_intelligente(moyenne_dec, niveau_detecte)
                            bulletin_data['appreciation'] = obtenir_appreciation_intelligente(moyenne_dec, eleve_selectionne.prenom, niveau_detecte)
                            bulletin_data['rang'] = f"{rang_info['rang']}/{rang_info['total_eleves']}"
                        else:
                            bulletin_data['mention'] = None
                            bulletin_data['appreciation'] = obtenir_appreciation_intelligente(_Dec('0'), eleve_selectionne.prenom, niveau_detecte)
                            bulletin_data['rang'] = "-"
                    else:
                        moyenne_dec = _Dec(str(moyenne_generale))
                        bulletin_data['mention'] = obtenir_mention_intelligente(moyenne_dec, niveau_detecte)
                        bulletin_data['appreciation'] = obtenir_appreciation_intelligente(moyenne_dec, eleve_selectionne.prenom, niveau_detecte)
                        bulletin_data['rang'] = "-"
    
    # Déterminer est_maternelle et est_primaire pour le contexte
    est_maternelle_ctx = False
    est_primaire_ctx = False
    mode_saisie_notes = 'mixte'  # Par défaut: notes mensuelles + compositions
    has_notes_mensuelles = True
    has_compositions = True
    
    if classe_selectionnee:
        from .calculs_moyennes import detecter_niveau_scolaire, detecter_notes_mensuelles_classe
        niveau_detecte_ctx = detecter_niveau_scolaire(classe_selectionnee.nom)
        est_maternelle_ctx = (niveau_detecte_ctx == 'MATERNELLE')
        est_primaire_ctx = (niveau_detecte_ctx == 'PRIMAIRE')
        
        # Détecter le mode de saisie des notes (mensuel, composition_seule, mixte)
        detection_result = detecter_notes_mensuelles_classe(classe_selectionnee, periode)
        mode_saisie_notes = detection_result['mode_saisie']
        has_notes_mensuelles = detection_result['has_notes_mensuelles']
        has_compositions = detection_result['has_compositions']
    
    # Base de notation: 10 pour primaire, 20 pour secondaire
    base_notation = 10 if est_primaire_ctx else 20
    
    context = {
        'titre_page': 'Bulletin Dynamique',
        'classes': classes,
        'classe_selectionnee': classe_selectionnee,
        'eleves': eleves,
        'eleve_selectionne': eleve_selectionne,
        'matieres': matieres,
        'periodes_disponibles': periodes_disponibles,
        'periode': periode,
        'periode_selectionnee': periode,  # Alias pour le template
        'system_type': system_type,
        'niveau_enseignement': niveau_enseignement,
        'est_maternelle': est_maternelle_ctx,
        'est_primaire': est_primaire_ctx,  # NOUVEAU: Pour masquer les coefficients en primaire
        'base_notation': base_notation,  # 10 pour primaire, 20 pour secondaire
        'bulletin_data': bulletin_data,
        'ecole': ecole,
        'annee_scolaire': classe_selectionnee.annee_scolaire if classe_selectionnee else '',
        # NOUVEAU: Mode de saisie pour masquer colonnes mensuelles si seulement compositions
        'mode_saisie_notes': mode_saisie_notes,
        'has_notes_mensuelles': has_notes_mensuelles,
        'has_compositions': has_compositions,
    }
    
    return render(request, 'notes/bulletin_dynamique.html', context)

@login_required
def sauvegarder_appreciations_maternelle(request):
    """Sauvegarder les appréciations pour la maternelle/garderie"""
    from django.http import JsonResponse
    import json
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        from .models import AppreciationMaternelle, MatiereNote
        from eleves.models import Eleve
        
        # Format 1: Liste d'appréciations (pour import en masse)
        if 'appreciations' in data and isinstance(data['appreciations'], list):
            appreciations_data = data.get('appreciations', [])
            
            if not appreciations_data:
                return JsonResponse({'success': False, 'error': 'Aucune appréciation à sauvegarder'})
            
            saved_count = 0
            errors = []
            
            for item in appreciations_data:
                try:
                    eleve_id = item.get('eleve_id')
                    matiere_id = item.get('matiere_id')
                    trimestre = item.get('trimestre')
                    appreciation = item.get('appreciation')
                    commentaire = item.get('commentaire', '')
                    absent = item.get('absent', False)
                    
                    if not all([eleve_id, matiere_id, trimestre]):
                        continue

                    eleve = Eleve.objects.get(pk=eleve_id)
                    matiere = MatiereNote.objects.get(pk=matiere_id)
                    # ── Sécurité: vérifier école ──
                    user_profil = getattr(request.user, 'profil', None)
                    ecole_user = user_profil.ecole if user_profil else None
                    if ecole_user and matiere.classe.ecole != ecole_user:
                        errors.append(f"Accès non autorisé à la matière {matiere_id}")
                        continue

                    annee_scolaire = matiere.classe.annee_scolaire
                    
                    obj, created = AppreciationMaternelle.objects.update_or_create(
                        eleve=eleve,
                        matiere=matiere,
                        trimestre=trimestre,
                        annee_scolaire=annee_scolaire,
                        defaults={
                            'appreciation': appreciation if not absent else '',
                            'commentaire': commentaire,
                            'absent': absent,
                            'cree_par': request.user,
                        }
                    )
                    saved_count += 1
                    
                except Eleve.DoesNotExist:
                    errors.append(f"Élève {eleve_id} non trouvé")
                except MatiereNote.DoesNotExist:
                    errors.append(f"Matière {matiere_id} non trouvée")
                except Exception as e:
                    errors.append(str(e))
            
            if errors:
                return JsonResponse({
                    'success': True,
                    'message': f'{saved_count} appréciations sauvegardées avec {len(errors)} erreurs',
                    'errors': errors[:5]
                })
            
            return JsonResponse({
                'success': True,
                'message': f'{saved_count} appréciations sauvegardées avec succès'
            })
        
        # Format 2: Appréciations par trimestre (depuis le template)
        else:
            eleve_id = data.get('eleve_id')
            matiere_id = data.get('matiere_id')
            annee_scolaire = data.get('annee_scolaire')
            appreciations = data.get('appreciations', {})
            
            if not all([eleve_id, matiere_id]):
                return JsonResponse({'success': False, 'error': 'Élève et matière requis'})
            
            try:
                eleve = Eleve.objects.get(pk=eleve_id)
                matiere = MatiereNote.objects.get(pk=matiere_id)

                # ── Sécurité: vérifier école ──
                user_profil = getattr(request.user, 'profil', None)
                ecole_user = user_profil.ecole if user_profil else None
                if ecole_user and matiere.classe.ecole != ecole_user:
                    return JsonResponse({'success': False, 'error': 'Accès non autorisé'}, status=403)

                if not annee_scolaire:
                    annee_scolaire = matiere.classe.annee_scolaire

                saved_count = 0

                # Trimestre 1
                if 'trimestre1' in appreciations:
                    t1 = appreciations['trimestre1']
                    obj, created = AppreciationMaternelle.objects.update_or_create(
                        eleve=eleve,
                        matiere=matiere,
                        trimestre='TRIMESTRE_1',
                        annee_scolaire=annee_scolaire,
                        defaults={
                            'appreciation': t1.get('appreciation', ''),
                            'commentaire': t1.get('commentaire', ''),
                            'absent': t1.get('absent', False),
                            'cree_par': request.user,
                        }
                    )
                    saved_count += 1
                
                # Trimestre 2
                if 'trimestre2' in appreciations:
                    t2 = appreciations['trimestre2']
                    obj, created = AppreciationMaternelle.objects.update_or_create(
                        eleve=eleve,
                        matiere=matiere,
                        trimestre='TRIMESTRE_2',
                        annee_scolaire=annee_scolaire,
                        defaults={
                            'appreciation': t2.get('appreciation', ''),
                            'commentaire': t2.get('commentaire', ''),
                            'absent': t2.get('absent', False),
                            'cree_par': request.user,
                        }
                    )
                    saved_count += 1
                
                # Trimestre 3
                if 'trimestre3' in appreciations:
                    t3 = appreciations['trimestre3']
                    obj, created = AppreciationMaternelle.objects.update_or_create(
                        eleve=eleve,
                        matiere=matiere,
                        trimestre='TRIMESTRE_3',
                        annee_scolaire=annee_scolaire,
                        defaults={
                            'appreciation': t3.get('appreciation', ''),
                            'commentaire': t3.get('commentaire', ''),
                            'absent': t3.get('absent', False),
                            'cree_par': request.user,
                        }
                    )
                    saved_count += 1
                
                return JsonResponse({
                    'success': True,
                    'message': f'{saved_count} appréciations sauvegardées avec succès'
                })
                
            except Eleve.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Élève non trouvé'})
            except MatiereNote.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Matière non trouvée'})
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Données JSON invalides'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def saisie_notes_simple(request):
    """Saisie notes simple"""
    return render(request, 'notes/saisie_notes_simple.html', {'titre_page': 'Saisie Notes Simple'})

@login_required
def imprimer_tableau_notes_html(request):
    """Imprimer le tableau des notes avec ajustement des colonnes sur A4 landscape (version navigateur)"""
    from django.template.loader import render_to_string
    from django.http import HttpResponse
    import logging
    
    logger = logging.getLogger(__name__)
    
    # Récupérer les paramètres
    classe_id = request.GET.get('classe_id')
    periode = request.GET.get('periode')
    
    if not classe_id or not periode:
        return HttpResponse("Paramètres manquants", status=400)
    
    try:
        # Récupérer la classe et les données
        classe_note = get_object_or_404(ClasseNote, pk=classe_id)
        # Vérifier si ClasseNote a un attribut classe ou classe_eleve
        if hasattr(classe_note, 'classe'):
            classe = classe_note.classe
        elif hasattr(classe_note, 'classe_eleve'):
            classe = classe_note.classe_eleve
        else:
            # Si aucun attribut, utiliser directement l'objet
            classe = classe_note
        
        # Récupérer les matières
        matieres = MatiereNote.objects.filter(classe=classe_note).order_by('nom')
        
        # Calculer le classement
        from .calculs_moyennes import calculer_classement_classe
        from .calculs_intelligent import calculer_rang_intelligent
        
        # Récupérer les élèves
        from eleves.models import Eleve
        # Si classe_note est une ClasseNote, utiliser classe_note.classe
        if hasattr(classe_note, 'classe') and classe_note.classe:
            eleves = Eleve.objects.filter(classe=classe_note.classe, statut='actif').order_by('matricule')
        else:
            # Sinon, essayer de récupérer les élèves directement
            eleves = Eleve.objects.filter(statut='actif').order_by('matricule')
        
        # Calculer les moyennes et rangs
        classement_resultat = calculer_classement_classe(eleves, matieres, periode, 'mensuel')
        
        # Préparer les données pour le template
        classement_data = []
        for eleve in eleves:
            # Récupérer les détails des notes par matière
            details_matieres = {}
            for matiere in matieres:
                from .calculs_moyennes import calculer_moyenne_matiere
                result = calculer_moyenne_matiere(eleve, matiere, periode, 'mensuel')
                details_matieres[matiere.id] = result
            
            # Récupérer le rang et la moyenne
            rang_num = classement_resultat['rang_map'].get(eleve.id)
            rang_str = str(rang_num) if rang_num else "-"
            moyenne = classement_resultat['moyennes_par_eleve'].get(eleve.id)
            
            # Formatter le rang avec ex-æquo si nécessaire
            if rang_num:
                from .calculs_intelligent import formater_rang_intelligent
                sexe = getattr(eleve, 'sexe', 'M') or 'M'
                rang_str = formater_rang_intelligent(rang_num, sexe)
            
            classement_data.append({
                'matricule': eleve.matricule,
                'nom_complet': eleve.nom_complet,
                'rang': rang_str,
                'moyenne': moyenne,
                'details_matieres': details_matieres,
                'sexe': getattr(eleve, 'sexe', 'M') or 'M'
            })
        
        # Trier par rang
        classement_data.sort(key=lambda x: x['rang'] if x['rang'] != '-' else '999')
        
        # Contexte pour le template
        context = {
            'classe_selectionnee': classe,
            'periode_selectionnee': periode,
            'matieres': matieres,
            'classement_data': classement_data,
        }
        
        # Générer le HTML
        html_content = render_to_string('notes/impression_tableau_notes.html', context, request=request)
        
        # Retourner le HTML (le navigateur gérera l'impression)
        response = HttpResponse(html_content, content_type='text/html')
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'impression du tableau: {str(e)}")
        return HttpResponse(f"Erreur: {str(e)}", status=500)

@login_required
def saisie_notes_simple(request):
    """Saisie notes simple - Système Guinéen avec chargement des notes importées"""
    from eleves.models import Eleve, Classe as ClasseEleve
    from .models import ClasseNote, MatiereNote, NoteMensuelle, CompositionNote, AppreciationMaternelle
    from .calculs_moyennes import detecter_niveau_scolaire
    
    ecole = _get_ecole(request)
    
    # Récupérer les classes disponibles
    classes = ClasseNote.objects.filter(ecole=ecole, actif=True).order_by('nom') if ecole else ClasseNote.objects.none()
    
    # Paramètres de recherche
    classe_id = request.GET.get('classe_id')
    eleve_id = request.GET.get('eleve_id')
    matiere_id = request.GET.get('matiere_id')
    
    classe_selectionnee = None
    eleve_selectionne = None
    matiere_selectionnee = None
    eleves = []
    matieres = []
    notes_mensuelles = {}
    compositions = {}
    appreciations = {}
    niveau_enseignement = 'PRIMAIRE'
    system_type = 'trimestre'
    
    # Liste des mois
    mois_list = [
        ('OCTOBRE', 'Octobre'),
        ('NOVEMBRE', 'Novembre'),
        ('DECEMBRE', 'Décembre'),
        ('JANVIER', 'Janvier'),
        ('FEVRIER', 'Février'),
        ('MARS', 'Mars'),
        ('AVRIL', 'Avril'),
        ('MAI', 'Mai'),
        ('JUIN', 'Juin'),
    ]
    
    if classe_id:
        classe_selectionnee = ClasseNote.objects.filter(id=classe_id, ecole=ecole).first()
        
        if classe_selectionnee:
            # Détecter le niveau d'enseignement
            niveau_enseignement = detecter_niveau_scolaire(classe_selectionnee.nom)
            
            # Récupérer les élèves de la classe
            classe_eleve = ClasseEleve.objects.filter(
                nom=classe_selectionnee.nom,
                annee_scolaire=classe_selectionnee.annee_scolaire,
                ecole=ecole
            ).first()
            
            if classe_eleve:
                eleves = Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('prenom', 'nom')
            
            # Récupérer les matières
            matieres = MatiereNote.objects.filter(classe=classe_selectionnee, actif=True).order_by('nom')
    
    if eleve_id and classe_selectionnee:
        eleve_selectionne = Eleve.objects.filter(id=eleve_id).first()
    
    if matiere_id and classe_selectionnee:
        matiere_selectionnee = MatiereNote.objects.filter(id=matiere_id, classe=classe_selectionnee).first()
    
    # Charger les notes existantes (importées ou saisies)
    if eleve_selectionne and matiere_selectionnee:
        # Charger les notes mensuelles depuis NoteMensuelle
        notes_mensuelles_qs = NoteMensuelle.objects.filter(
            eleve=eleve_selectionne,
            matiere=matiere_selectionnee,
            annee_scolaire=classe_selectionnee.annee_scolaire
        )
        
        for nm in notes_mensuelles_qs:
            if nm.note is not None:
                notes_mensuelles[nm.mois] = float(nm.note)
            elif nm.absent:
                notes_mensuelles[nm.mois] = 'ABS'
        
        # Charger les compositions
        compositions_qs = CompositionNote.objects.filter(
            eleve=eleve_selectionne,
            matiere=matiere_selectionnee,
            annee_scolaire=classe_selectionnee.annee_scolaire
        )
        
        for comp in compositions_qs:
            if comp.note is not None:
                if 'TRIMESTRE_1' in comp.periode or 'SEMESTRE_1' in comp.periode:
                    compositions['composition1'] = float(comp.note)
                elif 'TRIMESTRE_2' in comp.periode or 'SEMESTRE_2' in comp.periode:
                    compositions['composition2'] = float(comp.note)
                elif 'TRIMESTRE_3' in comp.periode:
                    compositions['composition3'] = float(comp.note)
        
        # Charger les appréciations maternelle
        if niveau_enseignement == 'MATERNELLE':
            appreciations_qs = AppreciationMaternelle.objects.filter(
                eleve=eleve_selectionne,
                matiere=matiere_selectionnee,
                annee_scolaire=classe_selectionnee.annee_scolaire
            )
            
            for app in appreciations_qs:
                if 'TRIMESTRE_1' in app.trimestre:
                    appreciations['appreciation1'] = app.appreciation
                elif 'TRIMESTRE_2' in app.trimestre:
                    appreciations['appreciation2'] = app.appreciation
                elif 'TRIMESTRE_3' in app.trimestre:
                    appreciations['appreciation3'] = app.appreciation
    
    # Compter les notes existantes pour afficher un indicateur
    notes_existantes_count = 0
    if classe_selectionnee and matiere_selectionnee:
        notes_existantes_count = NoteMensuelle.objects.filter(
            matiere=matiere_selectionnee,
            annee_scolaire=classe_selectionnee.annee_scolaire
        ).exclude(note__isnull=True).count()
    
    # Liste des codes de mois pour le JavaScript
    mois_codes = ['OCTOBRE', 'NOVEMBRE', 'DECEMBRE', 'JANVIER', 'FEVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN']
    
    context = {
        'titre_page': 'Saisie Notes - Système Guinéen',
        'classes': classes,
        'classe_selectionnee': classe_selectionnee,
        'eleves': eleves,
        'eleve_selectionne': eleve_selectionne,
        'matieres': matieres,
        'matiere_selectionnee': matiere_selectionnee,
        'notes_mensuelles': notes_mensuelles,
        'composition1': compositions.get('composition1'),
        'composition2': compositions.get('composition2'),
        'composition3': compositions.get('composition3'),
        'appreciation1': appreciations.get('appreciation1'),
        'appreciation2': appreciations.get('appreciation2'),
        'appreciation3': appreciations.get('appreciation3'),
        'niveau_enseignement': niveau_enseignement,
        'system_type': system_type,
        'mois_list': mois_list,
        'mois_codes': mois_codes,
        'notes_existantes_count': notes_existantes_count,
    }
    
    return render(request, 'notes/saisie_notes_simple.html', context)

@login_required
def sauvegarder_notes_guineen(request):
    """Sauvegarder notes guinéen - Notes mensuelles et compositions"""
    from django.http import JsonResponse
    from .models import ClasseNote, MatiereNote, NoteMensuelle, CompositionNote
    from eleves.models import Eleve
    import json
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)
    
    try:
        data = json.loads(request.body)
        eleve_id = data.get('eleve_id')
        matiere_id = data.get('matiere_id')
        annee_scolaire = data.get('annee_scolaire')
        notes_mois = data.get('notes_mois', {})  # Format: {mois: {note: X, absent: bool}}
        compositions = data.get('compositions', {})  # Format: {compositionX: {note: X, absent: bool}}
        
        eleve = Eleve.objects.get(id=eleve_id)
        matiere = MatiereNote.objects.get(id=matiere_id)

        # ── Sécurité: vérifier que l'élève et la matière appartiennent à l'école ──
        user_profil = getattr(request.user, 'profil', None)
        ecole_user = user_profil.ecole if user_profil else None
        if ecole_user and matiere.classe.ecole != ecole_user:
            return JsonResponse({'success': False, 'error': 'Accès non autorisé à cette matière'}, status=403)
        if ecole_user and eleve.classe and eleve.classe.ecole != ecole_user:
            return JsonResponse({'success': False, 'error': 'Accès non autorisé à cet élève'}, status=403)

        # Utiliser l'année scolaire de la matière si non fournie
        if not annee_scolaire:
            annee_scolaire = matiere.classe.annee_scolaire

        from .calculs_moyennes import detecter_niveau_scolaire
        niveau_note = detecter_niveau_scolaire(matiere.classe.nom)
        note_max = Decimal('10') if niveau_note == 'PRIMAIRE' else Decimal('20')

        saved_count = 0
        updated_count = 0

        # Sauvegarder les notes mensuelles
        for mois, note_data in notes_mois.items():
            if note_data is not None:
                try:
                    note_value = note_data.get('note') if isinstance(note_data, dict) else note_data
                    absent = note_data.get('absent', False) if isinstance(note_data, dict) else False
                    
                    if absent or (note_value is not None and note_value != ''):
                        note_decimal = Decimal('0') if absent else Decimal(str(note_value).replace(',', '.'))
                        if note_decimal < 0 or note_decimal > note_max:
                            return JsonResponse({
                                'success': False,
                                'error': f'Note invalide: elle doit être entre 0 et {note_max}'
                            }, status=400)
                        obj, created = NoteMensuelle.objects.update_or_create(
                            eleve=eleve,
                            matiere=matiere,
                            mois=mois.upper(),
                            annee_scolaire=annee_scolaire,
                            defaults={
                                'note': note_decimal,
                                'absent': absent
                            }
                        )
                        if created:
                            saved_count += 1
                        else:
                            updated_count += 1
                except (ValueError, InvalidOperation):
                    continue
        
        # Sauvegarder les compositions
        periode_mapping = {
            'composition1': 'TRIMESTRE_1',
            'composition2': 'TRIMESTRE_2',
            'composition3': 'TRIMESTRE_3',
        }
        
        for key, comp_data in compositions.items():
            if comp_data is not None and key in periode_mapping:
                try:
                    note_value = comp_data.get('note') if isinstance(comp_data, dict) else comp_data
                    absent = comp_data.get('absent', False) if isinstance(comp_data, dict) else False
                    
                    if absent or (note_value is not None and note_value != ''):
                        note_decimal = Decimal('0') if absent else Decimal(str(note_value).replace(',', '.'))
                        if note_decimal < 0 or note_decimal > note_max:
                            return JsonResponse({
                                'success': False,
                                'error': f'Note invalide: elle doit être entre 0 et {note_max}'
                            }, status=400)
                        obj, created = CompositionNote.objects.update_or_create(
                            eleve=eleve,
                            matiere=matiere,
                            periode=periode_mapping[key],
                            annee_scolaire=annee_scolaire,
                            defaults={
                                'note': note_decimal,
                                'absent': absent
                            }
                        )
                        if created:
                            saved_count += 1
                        else:
                            updated_count += 1
                except (ValueError, InvalidOperation):
                    continue
        
        # Invalider le cache des rangs
        try:
            # Essayer d'invalider le cache avec pattern
            cache_key = f"rangs_classe_{matiere.classe.id}_*"
            if hasattr(cache, 'delete_pattern'):
                cache.delete_pattern(cache_key)
            else:
                # Sinon, invalider les clés connues
                for periode in ['OCTOBRE', 'NOVEMBRE', 'DECEMBRE', 'JANVIER', 'FEVRIER', 'MARS', 'AVRIL', 'MAI', 'JUIN']:
                    cache.delete(f"rangs_classe_{matiere.classe.id}_periode_{periode}")
        except Exception:
            pass
        
        return JsonResponse({
            'success': True,
            'message': f'{saved_count} note(s) créée(s), {updated_count} mise(s) à jour',
            'saved': saved_count,
            'updated': updated_count
        })
        
    except Eleve.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Élève non trouvé'}, status=404)
    except MatiereNote.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Matière non trouvée'}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Données JSON invalides'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def imprimer_tableau_notes_pdf(request):
    """Imprimer le tableau des notes avec ajustement des colonnes sur A4 landscape - Supporte maternelle"""
    from django.template.loader import render_to_string
    from weasyprint import HTML, CSS
    from django.http import HttpResponse
    from .models import AppreciationMaternelle
    from .calculs_moyennes import detecter_niveau_scolaire
    from .utils_rangs import calculer_rangs_classe_periode
    
    classe_id = request.GET.get('classe_id')
    periode = request.GET.get('periode')
    
    if not classe_id or not periode:
        return HttpResponse("Paramètres manquants", status=400)
    
    try:
        classe_note = get_object_or_404(ClasseNote, pk=classe_id)
        matieres = MatiereNote.objects.filter(classe=classe_note, actif=True).order_by('nom')
        
        # Détecter si maternelle
        niveau_detecte = detecter_niveau_scolaire(classe_note.nom)
        est_maternelle = (niveau_detecte == 'MATERNELLE')
        
        # Récupérer les élèves
        from eleves.models import Eleve, Classe as ClasseEleve
        classe_eleve = ClasseEleve.objects.filter(
            nom=classe_note.nom,
            annee_scolaire=classe_note.annee_scolaire,
            ecole=classe_note.ecole
        ).first()
        
        eleves = Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('prenom', 'nom') if classe_eleve else []
        rangs_dict = calculer_rangs_classe_periode(classe_note, periode, use_cache=True)
        
        classement_data = []
        
        if est_maternelle:
            # Récupérer appréciations maternelle (même logique que consulter_notes)
            appreciations_qs = AppreciationMaternelle.objects.filter(
                matiere__in=matieres,
                trimestre=periode,
                annee_scolaire=classe_note.annee_scolaire
            ).values('eleve_id', 'matiere_id', 'appreciation')
            
            if not appreciations_qs.exists():
                appreciations_qs = AppreciationMaternelle.objects.filter(
                    matiere__in=matieres,
                    trimestre=periode
                ).values('eleve_id', 'matiere_id', 'appreciation')
            
            appreciations_dict = {(a['eleve_id'], a['matiere_id']): a for a in appreciations_qs}
            
            for eleve in eleves:
                details_matieres = {}
                for matiere in matieres:
                    app_data = appreciations_dict.get((eleve.id, matiere.id))
                    details_matieres[matiere.id] = {
                        'note_display': app_data['appreciation'] if app_data and app_data['appreciation'] else '-'
                    }
                
                rang_info = rangs_dict.get(eleve.id)
                classement_data.append({
                    'matricule': eleve.matricule,
                    'prenom': eleve.prenom,
                    'nom': eleve.nom,
                    'rang': rang_info['rang'] if rang_info else '-',
                    'moyenne': float(rang_info['moyenne']) if rang_info else None,
                    'details_matieres': details_matieres,
                })
        else:
            from .calculs_moyennes import calculer_classement_classe, calculer_moyenne_matiere
            classement_resultat = calculer_classement_classe(eleves, matieres, periode, 'mensuel')
            
            for eleve in eleves:
                details_matieres = {}
                for matiere in matieres:
                    result = calculer_moyenne_matiere(eleve, matiere, periode, 'mensuel')
                    details_matieres[matiere.id] = result
                
                rang_info = rangs_dict.get(eleve.id)
                classement_data.append({
                    'matricule': eleve.matricule,
                    'prenom': eleve.prenom,
                    'nom': eleve.nom,
                    'rang': rang_info['rang'] if rang_info else '-',
                    'moyenne': float(rang_info['moyenne']) if rang_info else None,
                    'details_matieres': details_matieres,
                })
        
        # Trier par rang numérique
        def sort_rang(x):
            r = x['rang']
            if r == '-': return 999
            return int(r.replace('er', '').replace('ère', '').replace('ème', ''))
        classement_data.sort(key=sort_rang)
        
        context = {
            'classe_note': classe_note,
            'classe_selectionnee': classe_note,
            'periode_selectionnee': periode,
            'matieres': matieres,
            'classement_data': classement_data,
            'est_maternelle': est_maternelle,
            'annee_scolaire': classe_note.annee_scolaire,
        }
        
        html_content = render_to_string('notes/impression_tableau_notes.html', context, request=request)
        html = HTML(string=html_content)
        css = CSS(string='@page { size: A4 landscape; margin: 10mm; }')
        pdf = html.write_pdf(stylesheets=[css])
        
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="tableau_notes_{classe_note.nom}_{periode}.pdf"'
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'impression du tableau: {str(e)}")
        return HttpResponse(f"Erreur: {str(e)}", status=500)


# ============================================================================
# BULLETIN MATERNELLE V2 - Basé sur AppreciationMaternelle
# ============================================================================

@login_required
def saisie_bulletin_maternelle(request, eleve_id, classe_id, trimestre):
    """Interface de saisie des analyses et recommandations pour le bulletin maternelle"""
    from .models import AppreciationMaternelle, BulletinMaternelle
    from eleves.models import Eleve
    
    eleve = get_object_or_404(Eleve, id=eleve_id)
    classe_note = get_object_or_404(ClasseNote, id=classe_id)
    
    # Récupérer ou créer le bulletin
    bulletin, created = BulletinMaternelle.objects.get_or_create(
        eleve=eleve,
        classe=classe_note,
        trimestre=trimestre,
        annee_scolaire=classe_note.annee_scolaire,
        defaults={'cree_par': request.user}
    )
    
    # Récupérer les appréciations existantes
    matieres = MatiereNote.objects.filter(classe=classe_note, actif=True).order_by('nom')
    appreciations = AppreciationMaternelle.objects.filter(
        eleve=eleve,
        matiere__in=matieres,
        trimestre=trimestre,
        annee_scolaire=classe_note.annee_scolaire
    ).select_related('matiere')
    
    if request.method == 'POST':
        # Sauvegarder les analyses
        analyses_selectionnees = request.POST.getlist('analyses')
        bulletin.analyses = analyses_selectionnees
        
        # Sauvegarder les recommandations
        recommandations_selectionnees = request.POST.getlist('recommandations')
        bulletin.recommandations = recommandations_selectionnees
        
        # Appréciation générale
        bulletin.appreciation_generale = request.POST.get('appreciation_generale', '')
        bulletin.save()
        
        messages.success(request, f"Bulletin de {eleve.prenom} {eleve.nom} sauvegardé!")
        
        if 'generer_pdf' in request.POST:
            return redirect('notes:bulletin_maternelle_v2_pdf', eleve_id=eleve_id, classe_id=classe_id, trimestre=trimestre)
        
        return redirect('notes:consulter_notes')
    
    context = {
        'eleve': eleve,
        'classe_note': classe_note,
        'trimestre': trimestre,
        'trimestre_display': dict(BulletinMaternelle.TRIMESTRE_CHOICES).get(trimestre, trimestre),
        'bulletin': bulletin,
        'matieres': matieres,
        'appreciations': {a.matiere_id: a for a in appreciations},
        'analyses_choices': BulletinMaternelle.ANALYSES_CHOICES,
        'recommandations_choices': BulletinMaternelle.RECOMMANDATIONS_CHOICES,
    }
    
    return render(request, 'notes/maternelle/saisie_bulletin_v2.html', context)


@login_required
def bulletin_maternelle_v2(request, eleve_id, classe_id, trimestre):
    """Affiche le bulletin maternelle V2 en HTML"""
    from .models import AppreciationMaternelle, BulletinMaternelle
    from .utils_rangs import calculer_rangs_classe_periode
    from eleves.models import Eleve
    
    eleve = get_object_or_404(Eleve, id=eleve_id)
    classe_note = get_object_or_404(ClasseNote, id=classe_id)
    
    # Récupérer le bulletin (analyses/recommandations)
    bulletin = BulletinMaternelle.objects.filter(
        eleve=eleve,
        classe=classe_note,
        trimestre=trimestre,
        annee_scolaire=classe_note.annee_scolaire
    ).first()
    
    # Récupérer les appréciations par matière
    matieres = MatiereNote.objects.filter(classe=classe_note, actif=True).order_by('nom')
    appreciations = AppreciationMaternelle.objects.filter(
        eleve=eleve,
        matiere__in=matieres,
        trimestre=trimestre,
        annee_scolaire=classe_note.annee_scolaire
    ).select_related('matiere')
    
    # Si pas d'appréciations avec année, essayer sans
    if not appreciations.exists():
        appreciations = AppreciationMaternelle.objects.filter(
            eleve=eleve,
            matiere__in=matieres,
            trimestre=trimestre
        ).select_related('matiere')
    
    # Préparer les notes avec lettres et mentions
    notes_data = []
    for app in appreciations:
        notes_data.append({
            'matiere': app.matiere,
            'lettre': app.appreciation,
            'mention': dict(AppreciationMaternelle.APPRECIATION_CHOICES).get(app.appreciation, ''),
            'absent': app.absent
        })
    
    # Calculer moyenne et rang
    rangs_dict = calculer_rangs_classe_periode(classe_note, trimestre, use_cache=True)
    rang_info = rangs_dict.get(eleve.id, {})
    
    context = {
        'eleve': eleve,
        'classe': classe_note,
        'ecole': classe_note.ecole,
        'trimestre': trimestre,
        'trimestre_display': dict(BulletinMaternelle.TRIMESTRE_CHOICES).get(trimestre, trimestre),
        'notes': notes_data,
        'moyenne': rang_info.get('moyenne'),
        'rang': rang_info.get('rang', '-'),
        'bulletin': bulletin,
        'analyses_selectionnees': bulletin.get_analyses_display() if bulletin else [],
        'recommandations_selectionnees': bulletin.get_recommandations_display() if bulletin else [],
    }
    
    return render(request, 'notes/maternelle/bulletin_v2.html', context)


@login_required
def bulletin_maternelle_v2_pdf(request, eleve_id, classe_id, trimestre):
    """Génère le bulletin maternelle V2 en PDF"""
    from django.template.loader import render_to_string
    from weasyprint import HTML
    from .models import AppreciationMaternelle, BulletinMaternelle
    from .utils_rangs import calculer_rangs_classe_periode
    from eleves.models import Eleve
    import base64
    import os
    
    eleve = get_object_or_404(Eleve, id=eleve_id)
    classe_note = get_object_or_404(ClasseNote, id=classe_id)
    
    # Récupérer le bulletin
    bulletin = BulletinMaternelle.objects.filter(
        eleve=eleve, classe=classe_note, trimestre=trimestre,
        annee_scolaire=classe_note.annee_scolaire
    ).first()
    
    # Récupérer les appréciations
    matieres = MatiereNote.objects.filter(classe=classe_note, actif=True).order_by('nom')
    appreciations = AppreciationMaternelle.objects.filter(
        eleve=eleve, matiere__in=matieres, trimestre=trimestre,
        annee_scolaire=classe_note.annee_scolaire
    ).select_related('matiere')
    
    if not appreciations.exists():
        appreciations = AppreciationMaternelle.objects.filter(
            eleve=eleve, matiere__in=matieres, trimestre=trimestre
        ).select_related('matiere')
    
    # Préparer les notes
    notes_data = []
    for app in appreciations:
        notes_data.append({
            'matiere': app.matiere,
            'lettre': app.appreciation,
            'mention': dict(AppreciationMaternelle.APPRECIATION_CHOICES).get(app.appreciation, ''),
            'note': _lettre_vers_note(app.appreciation),
            'absent': app.absent
        })
    
    # Calculer moyenne et rang
    rangs_dict = calculer_rangs_classe_periode(classe_note, trimestre, use_cache=True)
    rang_info = rangs_dict.get(eleve.id, {})
    moyenne_pourcentage = rang_info.get('moyenne')  # Pourcentage d'acquisition (ex: 87.5%)
    
    # Déterminer lettre et mention générales basées sur le pourcentage
    lettre_generale = None
    mention_generale = ''
    if moyenne_pourcentage:
        note_sur_10 = float(moyenne_pourcentage) / 10  # 87.5% -> 8.75
        lettre_generale = _note_vers_lettre(note_sur_10)
        mention_generale = dict(AppreciationMaternelle.APPRECIATION_CHOICES).get(lettre_generale, '') if lettre_generale else ''
    
    # Encoder logo et photo
    ecole = classe_note.ecole
    logo_base64 = ''
    if ecole.logo:
        try:
            if os.path.exists(ecole.logo.path):
                with open(ecole.logo.path, 'rb') as f:
                    logo_base64 = base64.b64encode(f.read()).decode('utf-8')
        except: pass
    
    photo_base64 = ''
    if hasattr(eleve, 'photo') and eleve.photo:
        try:
            if os.path.exists(eleve.photo.path):
                with open(eleve.photo.path, 'rb') as f:
                    photo_base64 = base64.b64encode(f.read()).decode('utf-8')
        except: pass
    
    context = {
        'eleve': eleve,
        'classe': classe_note,
        'ecole': ecole,
        'evaluation': {'trimestre': trimestre, 'annee_scolaire': classe_note.annee_scolaire,
                       'get_trimestre_display': dict(BulletinMaternelle.TRIMESTRE_CHOICES).get(trimestre, trimestre)},
        'notes': notes_data,
        'moyenne_pourcentage': f"{float(moyenne_pourcentage):.1f}%" if moyenne_pourcentage else None,
        'lettre_generale': lettre_generale,
        'mention_generale': mention_generale,
        'rang': rang_info.get('rang', '-'),
        'total_eleves': rang_info.get('total_eleves'),
        'analyses_selectionnees': bulletin.get_analyses_display() if bulletin else [],
        'recommandations_selectionnees': bulletin.get_recommandations_display() if bulletin else [],
        'logo_base64': logo_base64,
        'photo_base64': photo_base64,
        'date_impression': timezone.now(),
    }
    
    html_content = render_to_string('notes/maternelle/bulletin_pdf.html', context)
    pdf_file = HTML(string=html_content).write_pdf()
    
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="bulletin_{eleve.nom}_{eleve.prenom}_{trimestre}.pdf"'
    return response


from .utils_maternelle import lettre_vers_note as _lettre_vers_note, note_vers_lettre as _note_vers_lettre


@login_required
def bulletins_classe_maternelle_v2_pdf(request):
    """Génère tous les bulletins maternelle V2 d'une classe en un seul PDF"""
    from django.template.loader import render_to_string
    from weasyprint import HTML
    from .models import AppreciationMaternelle, BulletinMaternelle
    from .utils_rangs import calculer_rangs_classe_periode
    from eleves.models import Eleve, Classe
    import base64
    import os
    
    classe_id = request.GET.get('classe')
    trimestre = request.GET.get('trimestre', 'TRIMESTRE_1')
    
    if not classe_id:
        messages.error(request, "Veuillez sélectionner une classe")
        return redirect('notes:consulter_notes')
    
    classe_note = get_object_or_404(ClasseNote, id=classe_id)
    
    # Mapping des classes spéciales
    mapping_classes = {
        61: 56,
        59: 8,
    }
    
    # Récupérer les élèves de la classe avec logique améliorée
    classe_eleves = None
    eleves = []
    
    try:
        if classe_note.id in mapping_classes:
            classe_eleves = Classe.objects.filter(id=mapping_classes[classe_note.id]).first()
        else:
            # Essayer avec nom exact, année et école
            classe_eleves = Classe.objects.filter(
                nom=classe_note.nom,
                annee_scolaire=classe_note.annee_scolaire,
                ecole=classe_note.ecole
            ).first()
            
            if not classe_eleves:
                # Essayer sans le filtre école
                classe_eleves = Classe.objects.filter(
                    nom__iexact=classe_note.nom,
                    annee_scolaire=classe_note.annee_scolaire
                ).first()
            
            if not classe_eleves:
                # Essayer avec une correspondance partielle du nom
                classe_eleves = Classe.objects.filter(
                    nom__icontains=classe_note.nom.split()[0] if classe_note.nom else '',
                    annee_scolaire=classe_note.annee_scolaire
                ).first()
            
            if not classe_eleves:
                # Dernier essai: chercher par nom uniquement (toutes années)
                classe_eleves = Classe.objects.filter(
                    nom__iexact=classe_note.nom
                ).order_by('-annee_scolaire').first()
        
        if classe_eleves:
            eleves = list(Eleve.objects.filter(
                classe=classe_eleves,
                statut='ACTIF'
            ).order_by('prenom', 'nom'))
    except Exception as e:
        eleves = []
    
    if not eleves:
        debug_info = f"Classe: {classe_note.nom}, Année: {classe_note.annee_scolaire}"
        if classe_eleves:
            debug_info += f", Classe élèves trouvée: ID={classe_eleves.id}"
        else:
            debug_info += ", Aucune classe élèves correspondante trouvée"
        messages.warning(request, f"Aucun élève trouvé dans cette classe. ({debug_info})")
        return redirect('notes:consulter_notes')
    
    # Encoder le logo
    ecole = classe_note.ecole
    logo_base64 = ''
    if ecole and ecole.logo:
        try:
            if os.path.exists(ecole.logo.path):
                with open(ecole.logo.path, 'rb') as f:
                    logo_base64 = base64.b64encode(f.read()).decode('utf-8')
        except: pass
    
    # Calculer les rangs pour toute la classe
    rangs_dict = calculer_rangs_classe_periode(classe_note, trimestre, use_cache=True)
    
    # Récupérer les matières de la classe
    matieres = MatiereNote.objects.filter(classe=classe_note, actif=True).order_by('nom')
    
    # OPTIMISATION: Précharger TOUS les bulletins et appréciations en une seule requête
    eleves_ids = [e.id for e in eleves]
    
    # Précharger tous les bulletins de la classe en une requête
    all_bulletins = BulletinMaternelle.objects.filter(
        eleve_id__in=eleves_ids, classe=classe_note, trimestre=trimestre,
        annee_scolaire=classe_note.annee_scolaire
    )
    bulletins_dict = {b.eleve_id: b for b in all_bulletins}
    
    # Précharger toutes les appréciations en une requête
    all_appreciations = AppreciationMaternelle.objects.filter(
        eleve_id__in=eleves_ids, matiere__in=matieres, trimestre=trimestre,
        annee_scolaire=classe_note.annee_scolaire
    ).select_related('matiere')
    
    # Si pas d'appréciations avec année, essayer sans
    if not all_appreciations.exists():
        all_appreciations = AppreciationMaternelle.objects.filter(
            eleve_id__in=eleves_ids, matiere__in=matieres, trimestre=trimestre
        ).select_related('matiere')
    
    # Organiser les appréciations par élève
    appreciations_par_eleve = {}
    for app in all_appreciations:
        if app.eleve_id not in appreciations_par_eleve:
            appreciations_par_eleve[app.eleve_id] = {}
        appreciations_par_eleve[app.eleve_id][app.matiere.id] = app
    
    # Préparer les données pour chaque élève
    bulletins_data = []
    for eleve in eleves:
        # Récupérer le bulletin depuis le cache
        bulletin = bulletins_dict.get(eleve.id)
        
        # Récupérer les appréciations depuis le cache
        appreciations_dict = appreciations_par_eleve.get(eleve.id, {})
        
        # Préparer les notes - TOUJOURS inclure toutes les matières
        notes_data = []
        
        # Pour chaque matière, créer une entrée (avec ou sans appréciation)
        for matiere in matieres:
            appreciation = appreciations_dict.get(matiere.id)
            
            if appreciation:
                # Si appréciation existe, l'utiliser
                notes_data.append({
                    'matiere': matiere,
                    'lettre': appreciation.appreciation,
                    'mention': dict(AppreciationMaternelle.APPRECIATION_CHOICES).get(appreciation.appreciation, ''),
                    'note': _lettre_vers_note(appreciation.appreciation),
                    'absent': appreciation.absent
                })
            else:
                # Si pas d'appréciation, créer une entrée vide pour cocher manuellement
                notes_data.append({
                    'matiere': matiere,
                    'lettre': None,
                    'mention': '',
                    'note': None,
                    'absent': False
                })
        
        # Calculer moyenne et rang
        rang_info = rangs_dict.get(eleve.id, {})
        moyenne_pourcentage = rang_info.get('moyenne')  # Pourcentage d'acquisition (ex: 87.5%)
        
        # Déterminer lettre et mention générales basées sur le pourcentage
        # Conversion: pourcentage -> note sur 10 -> lettre
        lettre_generale = None
        mention_generale = ''
        if moyenne_pourcentage:
            note_sur_10 = float(moyenne_pourcentage) / 10  # 87.5% -> 8.75
            lettre_generale = _note_vers_lettre(note_sur_10)
            mention_generale = dict(AppreciationMaternelle.APPRECIATION_CHOICES).get(lettre_generale, '') if lettre_generale else ''
        
        # Photo de l'élève
        photo_base64 = ''
        if hasattr(eleve, 'photo') and eleve.photo:
            try:
                if os.path.exists(eleve.photo.path):
                    with open(eleve.photo.path, 'rb') as f:
                        photo_base64 = base64.b64encode(f.read()).decode('utf-8')
            except: pass
        
        bulletins_data.append({
            'eleve': eleve,
            'notes': notes_data,
            'moyenne_pourcentage': f"{float(moyenne_pourcentage):.1f}%" if moyenne_pourcentage else None,
            'lettre_generale': lettre_generale,
            'mention_generale': mention_generale,
            'rang': rang_info.get('rang', '-'),
            'total_eleves': rang_info.get('total_eleves', len(eleves)),
            'analyses_selectionnees': bulletin.get_analyses_display() if bulletin else [],
            'recommandations_selectionnees': bulletin.get_recommandations_display() if bulletin else [],
            'photo_base64': photo_base64,
            'bulletin': bulletin,
        })
    
    # Déterminer le nom du trimestre
    trimestre_display = {
        'TRIMESTRE_1': '1er Trimestre',
        'TRIMESTRE_2': '2ème Trimestre',
        'TRIMESTRE_3': '3ème Trimestre',
        'SEMESTRE_1': '1er Semestre',
        'SEMESTRE_2': '2ème Semestre',
    }.get(trimestre, trimestre)
    
    context = {
        'bulletins': bulletins_data,
        'classe': classe_note,
        'ecole': ecole,
        'trimestre': trimestre,
        'trimestre_display': trimestre_display,
        'annee_scolaire': classe_note.annee_scolaire,
        'logo_base64': logo_base64,
        'date_impression': timezone.now(),
    }
    
    html_content = render_to_string('notes/maternelle/bulletins_classe_pdf.html', context)
    pdf_file = HTML(string=html_content).write_pdf()
    
    response = HttpResponse(pdf_file, content_type='application/pdf')
    nom_classe_clean = classe_note.nom.replace(' ', '_').replace('/', '-')
    response['Content-Disposition'] = f'inline; filename="bulletins_{nom_classe_clean}_{trimestre}.pdf"'
    return response


@login_required
def fiches_recommandations_pdf(request):
    """Génère les fiches de recommandations vierges pour tous les élèves d'une classe maternelle"""
    from django.template.loader import render_to_string
    from weasyprint import HTML
    from eleves.models import Eleve, Classe
    import base64
    import os
    
    classe_id = request.GET.get('classe')
    trimestre = request.GET.get('trimestre', 'TRIMESTRE_1')
    
    if not classe_id:
        messages.error(request, "Veuillez sélectionner une classe")
        return redirect('notes:consulter_notes')
    
    classe_note = get_object_or_404(ClasseNote, id=classe_id)
    
    # Mapping des classes spéciales
    mapping_classes = {
        61: 56,
        59: 8,
    }
    
    # Récupérer les élèves de la classe
    classe_eleves = None
    eleves = []
    
    try:
        if classe_note.id in mapping_classes:
            classe_eleves = Classe.objects.filter(id=mapping_classes[classe_note.id]).first()
        else:
            # Essayer avec nom exact, année et école
            classe_eleves = Classe.objects.filter(
                nom=classe_note.nom,
                annee_scolaire=classe_note.annee_scolaire,
                ecole=classe_note.ecole
            ).first()
            
            if not classe_eleves:
                # Essayer sans le filtre école
                classe_eleves = Classe.objects.filter(
                    nom__iexact=classe_note.nom,
                    annee_scolaire=classe_note.annee_scolaire
                ).first()
            
            if not classe_eleves:
                # Essayer avec une correspondance partielle du nom
                classe_eleves = Classe.objects.filter(
                    nom__icontains=classe_note.nom.split()[0] if classe_note.nom else '',
                    annee_scolaire=classe_note.annee_scolaire
                ).first()
            
            if not classe_eleves:
                # Dernier essai: chercher par nom uniquement (toutes années)
                classe_eleves = Classe.objects.filter(
                    nom__iexact=classe_note.nom
                ).order_by('-annee_scolaire').first()
        
        if classe_eleves:
            eleves = list(Eleve.objects.filter(
                classe=classe_eleves,
                statut='ACTIF'
            ).order_by('prenom', 'nom'))
    except Exception as e:
        eleves = []
    
    if not eleves:
        # Message de debug plus informatif
        debug_info = f"Classe: {classe_note.nom}, Année: {classe_note.annee_scolaire}"
        if classe_eleves:
            debug_info += f", Classe élèves trouvée: ID={classe_eleves.id}"
        else:
            debug_info += ", Aucune classe élèves correspondante trouvée"
        messages.warning(request, f"Aucun élève trouvé dans cette classe. ({debug_info})")
        return redirect('notes:consulter_notes')
    
    # Encoder le logo
    ecole = classe_note.ecole
    logo_base64 = ''
    if ecole.logo:
        try:
            if os.path.exists(ecole.logo.path):
                with open(ecole.logo.path, 'rb') as f:
                    logo_base64 = base64.b64encode(f.read()).decode('utf-8')
        except: pass
    
    # Préparer les données pour chaque élève
    eleves_data = []
    for eleve in eleves:
        eleves_data.append({
            'eleve': eleve,
            'nom': eleve.nom,
            'prenom': eleve.prenom,
        })
    
    # Déterminer le nom du trimestre
    trimestre_display = {
        'TRIMESTRE_1': '1er Trimestre',
        'TRIMESTRE_2': '2ème Trimestre', 
        'TRIMESTRE_3': '3ème Trimestre',
        'SEMESTRE_1': '1er Semestre',
        'SEMESTRE_2': '2ème Semestre',
    }.get(trimestre, trimestre)
    
    context = {
        'eleves': eleves_data,
        'classe': classe_note,
        'ecole': ecole,
        'trimestre': trimestre,
        'trimestre_display': trimestre_display,
        'annee_scolaire': classe_note.annee_scolaire,
        'logo_base64': logo_base64,
        'date_impression': timezone.now(),
    }
    
    # Générer le HTML
    html_content = render_to_string('notes/maternelle/fiches_recommandations_pdf.html', context)
    
    # Générer le PDF
    pdf_file = HTML(string=html_content).write_pdf()
    
    # Créer la réponse
    response = HttpResponse(pdf_file, content_type='application/pdf')
    filename = f"fiches_recommandations_{classe_note.nom}_{trimestre}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    
    return response


@login_required
def bulletin_maternelle_modele2_pdf(request, eleve_id, classe_id, trimestre):
    """Génère le bulletin maternelle Modèle 2 (format tableau avec activités) en PDF"""
    from django.template.loader import render_to_string
    from weasyprint import HTML
    from .models import AppreciationMaternelle, BulletinMaternelle
    from .utils_rangs import calculer_rangs_classe_periode
    from eleves.models import Eleve, Classe as ClasseEleve
    import base64
    import os
    
    eleve = get_object_or_404(Eleve, id=eleve_id)
    classe_note = get_object_or_404(ClasseNote, id=classe_id)
    
    # Récupérer le bulletin
    bulletin = BulletinMaternelle.objects.filter(
        eleve=eleve, classe=classe_note, trimestre=trimestre,
        annee_scolaire=classe_note.annee_scolaire
    ).first()
    
    # Récupérer les appréciations
    matieres = MatiereNote.objects.filter(classe=classe_note, actif=True).order_by('nom')
    appreciations = AppreciationMaternelle.objects.filter(
        eleve=eleve, matiere__in=matieres, trimestre=trimestre,
        annee_scolaire=classe_note.annee_scolaire
    ).select_related('matiere')
    
    if not appreciations.exists():
        appreciations = AppreciationMaternelle.objects.filter(
            eleve=eleve, matiere__in=matieres, trimestre=trimestre
        ).select_related('matiere')
    
    # Créer un dictionnaire des notes par nom de matière normalisé
    notes_dict = {}
    for app in appreciations:
        nom_matiere = app.matiere.nom.lower().strip()
        notes_dict[nom_matiere] = {
            'appreciation': app.appreciation,
            'observation': app.commentaire if hasattr(app, 'commentaire') else '',
            'absent': app.absent
        }
    
    # Fonction helper pour trouver une matière par mots-clés (recherche intelligente)
    def trouver_matiere(*keywords):
        # Normaliser les mots-clés
        keywords_lower = [kw.lower() for kw in keywords]
        for key, value in notes_dict.items():
            key_normalized = key.lower().replace('-', ' ').replace('_', ' ')
            for kw in keywords_lower:
                if kw in key_normalized:
                    return value
        return {}
    
    # Mapper les matières vers les activités du modèle 2 avec recherche flexible et intelligente
    notes_mapped = {
        'francais_lecture': trouver_matiere('lecture', 'lect'),
        'francais_graphisme': trouver_matiere('graphisme', 'écriture', 'ecriture', 'graph'),
        'francais_recitation': trouver_matiere('récitation', 'recitation', 'récit', 'poésie', 'poesie'),
        'francais_oral': trouver_matiere('oral', 'langage', 'expression orale', 'lang'),
        'maths_numeration': trouver_matiere('numération', 'numeration', 'numer', 'nombre'),
        'maths_geometrie': trouver_matiere('géométrie', 'geometrie', 'géom', 'geom', 'forme'),
        'maths_prenumerations': trouver_matiere('prénumération', 'prenumeration', 'pré-numération', 'pre-num', 'prénum'),
        'explorer_civique': trouver_matiere('civique', 'moral', 'explorer', 'instruction', 'icm', 'éducation civique'),
        'expression_coloriage': trouver_matiere('coloriage', 'dessin', 'expression', 'création', 'art', 'plastique', 'color'),
        'activites_espace': trouver_matiere('espace', 'spatial', 'repérage'),
        'activites_temps': trouver_matiere('temps', 'temporel', 'chronolog'),
    }
    
    # Calculer rang et effectif
    rangs_dict = calculer_rangs_classe_periode(classe_note, trimestre, use_cache=True)
    rang_info = rangs_dict.get(eleve.id, {})
    
    # Compter les élèves
    classe_eleve = ClasseEleve.objects.filter(
        nom=classe_note.nom,
        annee_scolaire=classe_note.annee_scolaire,
        ecole=classe_note.ecole
    ).first()
    effectif = Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').count() if classe_eleve else 0
    
    # Encoder logo
    ecole = classe_note.ecole
    logo_base64 = ''
    if ecole and ecole.logo:
        try:
            if os.path.exists(ecole.logo.path):
                with open(ecole.logo.path, 'rb') as f:
                    logo_base64 = base64.b64encode(f.read()).decode('utf-8')
        except: pass
    
    context = {
        'eleve': eleve,
        'classe': classe_note,
        'ecole': ecole,
        'trimestre': trimestre,
        'trimestre_display': dict(BulletinMaternelle.TRIMESTRE_CHOICES).get(trimestre, trimestre),
        'annee_scolaire': classe_note.annee_scolaire,
        'notes_dict': notes_mapped,
        'rang': rang_info.get('rang', ''),
        'effectif': effectif,
        'appreciation_generale': bulletin.appreciation_generale if bulletin else '',
        'enseignant': '',
        'logo_base64': logo_base64,
        'date_impression': timezone.now(),
    }
    
    html_content = render_to_string('notes/maternelle/bulletin_modele2_pdf.html', context)
    pdf_file = HTML(string=html_content).write_pdf()
    
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="bulletin_m2_{eleve.nom}_{eleve.prenom}_{trimestre}.pdf"'
    return response


@login_required
def bulletins_classe_maternelle_modele2_pdf(request):
    """Génère tous les bulletins maternelle Modèle 2 d'une classe en un seul PDF"""
    from django.template.loader import render_to_string
    from weasyprint import HTML
    from .models import AppreciationMaternelle, BulletinMaternelle
    from .utils_rangs import calculer_rangs_classe_periode
    from eleves.models import Eleve, Classe as ClasseEleve
    import base64
    import os
    
    classe_id = request.GET.get('classe')
    trimestre = request.GET.get('trimestre', 'TRIMESTRE_1')
    
    if not classe_id:
        messages.error(request, "Veuillez sélectionner une classe")
        return redirect('notes:consulter_notes')
    
    classe_note = get_object_or_404(ClasseNote, id=classe_id)
    
    # Récupérer les élèves
    classe_eleve = ClasseEleve.objects.filter(
        nom=classe_note.nom,
        annee_scolaire=classe_note.annee_scolaire,
        ecole=classe_note.ecole
    ).first()
    
    if not classe_eleve:
        classe_eleve = ClasseEleve.objects.filter(
            nom__iexact=classe_note.nom,
            annee_scolaire=classe_note.annee_scolaire
        ).first()
    
    eleves = list(Eleve.objects.filter(classe=classe_eleve, statut='ACTIF').order_by('prenom', 'nom')) if classe_eleve else []
    
    if not eleves:
        messages.warning(request, "Aucun élève trouvé dans cette classe.")
        return redirect('notes:consulter_notes')
    
    # Récupérer les rangs
    rangs_dict = calculer_rangs_classe_periode(classe_note, trimestre, use_cache=True)
    effectif = len(eleves)
    
    # Récupérer les matières
    matieres = list(MatiereNote.objects.filter(classe=classe_note, actif=True).order_by('nom'))
    
    # Encoder logo UNE SEULE FOIS
    ecole = classe_note.ecole
    logo_base64 = ''
    if ecole and ecole.logo:
        try:
            if os.path.exists(ecole.logo.path):
                with open(ecole.logo.path, 'rb') as f:
                    logo_base64 = base64.b64encode(f.read()).decode('utf-8')
        except: pass
    
    # OPTIMISATION: Précharger TOUS les bulletins et appréciations en masse
    eleves_ids = [e.id for e in eleves]
    
    # Précharger tous les bulletins
    all_bulletins = BulletinMaternelle.objects.filter(
        eleve_id__in=eleves_ids, classe=classe_note, trimestre=trimestre,
        annee_scolaire=classe_note.annee_scolaire
    )
    bulletins_dict = {b.eleve_id: b for b in all_bulletins}
    
    # Précharger toutes les appréciations
    all_appreciations = list(AppreciationMaternelle.objects.filter(
        eleve_id__in=eleves_ids, matiere__in=matieres, trimestre=trimestre,
        annee_scolaire=classe_note.annee_scolaire
    ).select_related('matiere'))
    
    if not all_appreciations:
        all_appreciations = list(AppreciationMaternelle.objects.filter(
            eleve_id__in=eleves_ids, matiere__in=matieres, trimestre=trimestre
        ).select_related('matiere'))
    
    # Dictionnaire de conversion des appréciations en texte complet
    APPRECIATION_TEXTE = {
        'A+': 'Excellent',
        'A': 'Très bien',
        'B+': 'Bien',
        'B': 'Assez bien',
        'B-': 'Moyen',
        'C': 'Passable',
        'D': 'Éprouve des difficultés',
    }
    
    # Organiser les appréciations par élève
    appreciations_par_eleve = {}
    for app in all_appreciations:
        if app.eleve_id not in appreciations_par_eleve:
            appreciations_par_eleve[app.eleve_id] = {}
        nom_matiere = app.matiere.nom.lower().strip()
        appreciations_par_eleve[app.eleve_id][nom_matiere] = {
            'appreciation': app.appreciation,
            'appreciation_texte': APPRECIATION_TEXTE.get(app.appreciation, app.appreciation or ''),
            'observation': app.commentaire if hasattr(app, 'commentaire') else '',
            'absent': app.absent
        }
    
    # Générer les bulletins
    bulletins_html = []
    for eleve in eleves:
        # Récupérer depuis le cache
        bulletin = bulletins_dict.get(eleve.id)
        notes_dict = appreciations_par_eleve.get(eleve.id, {})
        
        # Fonction helper pour trouver une matière par mots-clés (recherche intelligente)
        def trouver_matiere(*keywords):
            keywords_lower = [kw.lower() for kw in keywords]
            for key, value in notes_dict.items():
                key_normalized = key.lower().replace('-', ' ').replace('_', ' ')
                for kw in keywords_lower:
                    if kw in key_normalized:
                        return value
            return {}
        
        # Mapper les matières avec recherche flexible et intelligente
        notes_mapped = {
            'francais_lecture': trouver_matiere('lecture', 'lect'),
            'francais_graphisme': trouver_matiere('graphisme', 'écriture', 'ecriture', 'graph'),
            'francais_recitation': trouver_matiere('récitation', 'recitation', 'récit', 'poésie', 'poesie'),
            'francais_oral': trouver_matiere('oral', 'langage', 'expression orale', 'lang'),
            'maths_numeration': trouver_matiere('numération', 'numeration', 'numer', 'nombre'),
            'maths_geometrie': trouver_matiere('géométrie', 'geometrie', 'géom', 'geom', 'forme'),
            'maths_prenumerations': trouver_matiere('prénumération', 'prenumeration', 'pré-numération', 'pre-num', 'prénum'),
            'explorer_civique': trouver_matiere('civique', 'moral', 'explorer', 'instruction', 'icm', 'éducation civique'),
            'expression_coloriage': trouver_matiere('coloriage', 'dessin', 'expression', 'création', 'art', 'plastique', 'color'),
            'activites_espace': trouver_matiere('espace', 'spatial', 'repérage'),
            'activites_temps': trouver_matiere('temps', 'temporel', 'chronolog'),
        }
        
        rang_info = rangs_dict.get(eleve.id, {})
        
        context = {
            'eleve': eleve,
            'classe': classe_note,
            'ecole': ecole,
            'trimestre': trimestre,
            'trimestre_display': dict(BulletinMaternelle.TRIMESTRE_CHOICES).get(trimestre, trimestre),
            'annee_scolaire': classe_note.annee_scolaire,
            'notes_dict': notes_mapped,
            'rang': rang_info.get('rang', ''),
            'effectif': effectif,
            'appreciation_generale': bulletin.appreciation_generale if bulletin else '',
            'enseignant': '',
            'logo_base64': logo_base64,
        }
        
        bulletin_html = render_to_string('notes/maternelle/bulletin_modele2_pdf.html', context)
        bulletins_html.append(bulletin_html)
    
    # Combiner tous les bulletins avec saut de page
    combined_html = '<div style="page-break-after: always;"></div>'.join(bulletins_html)
    
    pdf_file = HTML(string=combined_html).write_pdf()
    
    response = HttpResponse(pdf_file, content_type='application/pdf')
    nom_classe_clean = classe_note.nom.replace(' ', '_').replace('/', '-')
    response['Content-Disposition'] = f'inline; filename="bulletins_m2_{nom_classe_clean}_{trimestre}.pdf"'
    return response
