"""Vues du module Recouvrement.

Le module regroupe, derrière un tableau de bord général, les sous-modules
existants (dépenses, logistique, fournitures, bibliothèque) et quatre
nouveaux registres : cuisine, documents, versements et abonnements
informatique. Les trois premiers partagent la même structure (date,
désignation/lieu, montant, observation) : leurs vues sont donc génériques et
paramétrées par la table `MODULES`.
"""

import io
from datetime import timedelta
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Sum
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from eleves.models import Eleve
from salaires.models import AffectationClasse, Enseignant, EtatSalaire
from utilisateurs.permissions import (
    can_add_expenses, can_delete_expenses, can_modify_expenses,
)
from utilisateurs.utils import filter_by_user_school, user_is_superadmin, user_school

from .models import Depense
from .models_recouvrement import (
    AbonnementInformatique, DepenseCuisine, DepenseDocument, Versement,
)
from .forms_recouvrement import (
    AbonnementInformatiqueForm, DepenseCuisineForm, DepenseDocumentForm,
    VersementForm,
)

# ---------------------------------------------------------------------------
# Configuration des sous-modules « simples »
# ---------------------------------------------------------------------------

MODULES = {
    'cuisine': {
        'model': DepenseCuisine,
        'form': DepenseCuisineForm,
        'titre': 'Dépenses de la cuisine',
        'singulier': 'Dépense de cuisine',
        'description': "Achats et frais engagés pour la cuisine de l'établissement",
        'icone': 'fa-utensils',
        'couleur': 'success',
        'champ_libelle': 'designation',
        'entete_libelle': 'Désignation',
        'fichier': 'depenses_cuisine',
    },
    'document': {
        'model': DepenseDocument,
        'form': DepenseDocumentForm,
        'titre': 'Dépenses de documents',
        'singulier': 'Dépense de document',
        'description': "Impressions, reprographie et fournitures administratives",
        'icone': 'fa-file-invoice',
        'couleur': 'info',
        'champ_libelle': 'designation',
        'entete_libelle': 'Désignation',
        'fichier': 'depenses_documents',
    },
    'versement': {
        'model': Versement,
        'form': VersementForm,
        'titre': 'Versements',
        'singulier': 'Versement',
        'description': "Versements d'espèces vers la banque, le coffre ou la direction",
        'icone': 'fa-hand-holding-dollar',
        'couleur': 'primary',
        'champ_libelle': 'lieu_versement',
        'entete_libelle': 'Lieu de versement',
        'fichier': 'versements',
    },
}


def _config_module(module: str) -> dict:
    """Configuration du sous-module, 404 si la clé d'URL est inconnue."""
    try:
        return MODULES[module]
    except KeyError:
        raise Http404("Sous-module de recouvrement inconnu.")


def _queryset_module(request, config):
    """Queryset du sous-module, cloisonné à l'école de l'utilisateur."""
    qs = config['model'].objects.all()
    return filter_by_user_school(qs, request.user, 'ecole')


def _objet_module(request, config, pk):
    """Objet du sous-module accessible à l'utilisateur, sinon 404."""
    return get_object_or_404(_queryset_module(request, config), pk=pk)


def _parse_date(valeur):
    """Convertit une date de formulaire (AAAA-MM-JJ) ; None si invalide."""
    from datetime import datetime
    if not valeur:
        return None
    try:
        return datetime.strptime(valeur.strip(), '%Y-%m-%d').date()
    except (ValueError, AttributeError):
        return None


def _filtrer_operations(request, config, qs):
    """Applique recherche texte et bornes de dates communes aux sous-modules."""
    recherche = (request.GET.get('q') or '').strip()
    if recherche:
        qs = qs.filter(
            Q(**{f"{config['champ_libelle']}__icontains": recherche})
            | Q(observation__icontains=recherche)
        )

    date_debut = _parse_date(request.GET.get('du'))
    date_fin = _parse_date(request.GET.get('au'))
    if date_debut:
        qs = qs.filter(date__gte=date_debut)
    if date_fin:
        qs = qs.filter(date__lte=date_fin)

    return qs, {'q': recherche, 'du': request.GET.get('du') or '', 'au': request.GET.get('au') or ''}


def _total(qs) -> Decimal:
    return qs.aggregate(total=Sum('montant'))['total'] or Decimal('0')


def _serie_mensuelle(qs, mois=12):
    """Totaux des `mois` derniers mois, du plus ancien au plus récent."""
    from django.db.models.functions import TruncMonth

    aujourdhui = timezone.localdate()
    premier_mois = (aujourdhui.replace(day=1) - timedelta(days=31 * (mois - 1))).replace(day=1)
    agregats = (
        qs.filter(date__gte=premier_mois)
        .annotate(mois=TruncMonth('date'))
        .values('mois')
        .annotate(total=Sum('montant'))
    )
    totaux = {a['mois']: a['total'] or Decimal('0') for a in agregats if a['mois']}

    libelles_mois = [
        'janv.', 'févr.', 'mars', 'avr.', 'mai', 'juin',
        'juil.', 'août', 'sept.', 'oct.', 'nov.', 'déc.',
    ]
    serie = []
    curseur = premier_mois
    while curseur <= aujourdhui.replace(day=1):
        serie.append({
            'libelle': f"{libelles_mois[curseur.month - 1]} {curseur.year}",
            'total': int(totaux.get(curseur, Decimal('0')) or 0),
        })
        curseur = (curseur.replace(day=28) + timedelta(days=4)).replace(day=1)
    return serie


def _statistiques_operations(qs) -> dict:
    """Indicateurs affichés en haut du tableau de bord d'un sous-module."""
    aujourdhui = timezone.localdate()
    debut_mois = aujourdhui.replace(day=1)
    total = _total(qs)
    nombre = qs.count()
    return {
        'nombre': nombre,
        'total': int(total),
        'total_mois': int(_total(qs.filter(date__gte=debut_mois))),
        'total_annee': int(_total(qs.filter(date__year=aujourdhui.year))),
        'moyenne': int(total / nombre) if nombre else 0,
    }


# ---------------------------------------------------------------------------
# Tableau de bord général
# ---------------------------------------------------------------------------

@login_required
def hub_recouvrement(request):
    """Accueil du module : indicateurs consolidés puis cartes des sous-modules."""
    aujourdhui = timezone.localdate()
    debut_mois = aujourdhui.replace(day=1)

    # Dépenses classiques : l'école y est portée par le profil du créateur.
    depenses_qs = Depense.objects.all()
    if not user_is_superadmin(request.user):
        ecole = user_school(request.user)
        depenses_qs = (
            depenses_qs.filter(cree_par__profil__ecole=ecole) if ecole
            else Depense.objects.none()
        )

    cartes_chiffrees = []
    total_general = Decimal('0')
    total_mois = Decimal('0')

    depenses_total = depenses_qs.aggregate(t=Sum('montant_ttc'))['t'] or Decimal('0')
    depenses_mois = depenses_qs.filter(date_facture__gte=debut_mois).aggregate(
        t=Sum('montant_ttc')
    )['t'] or Decimal('0')
    total_general += depenses_total
    total_mois += depenses_mois

    for cle, config in MODULES.items():
        qs = _queryset_module(request, config)
        total = _total(qs)
        montant_mois = _total(qs.filter(date__gte=debut_mois))
        # Les versements sont des entrées de caisse : ils ne gonflent pas le
        # total dépensé affiché en tête de page.
        if cle != 'versement':
            total_general += total
            total_mois += montant_mois
        cartes_chiffrees.append({
            'cle': cle,
            'config': config,
            'nombre': qs.count(),
            'total': int(total),
            'total_mois': int(montant_mois),
        })

    abonnements_qs = filter_by_user_school(
        AbonnementInformatique.objects.all(), request.user, 'eleve__classe__ecole'
    )
    abonnements_actifs = abonnements_qs.filter(date_fin__gte=aujourdhui).count()
    abonnements_expires = abonnements_qs.filter(date_fin__lt=aujourdhui).count()
    abonnements_alerte = abonnements_qs.filter(
        date_fin__gte=aujourdhui,
        date_fin__lte=aujourdhui + timedelta(days=AbonnementInformatique.ALERTE_JOURS_DEFAUT),
    ).count()

    # Salaires enseignants payés
    salaires_payes_qs = filter_by_user_school(
        EtatSalaire.objects.filter(paye=True), request.user, 'enseignant__ecole'
    )
    salaires_total = salaires_payes_qs.aggregate(t=Sum('salaire_net'))['t'] or Decimal('0')
    salaires_mois = salaires_payes_qs.filter(
        periode__annee=aujourdhui.year, periode__mois=aujourdhui.month
    ).aggregate(t=Sum('salaire_net'))['t'] or Decimal('0')
    salaires_enseignants_count = salaires_payes_qs.values('enseignant_id').distinct().count()

    context = {
        'titre_page': 'Recouvrement',
        'total_general': int(total_general),
        'total_mois': int(total_mois),
        'depenses_nombre': depenses_qs.count(),
        'depenses_total': int(depenses_total),
        'depenses_mois': int(depenses_mois),
        'cartes_chiffrees': cartes_chiffrees,
        'versements_total': next(
            (c['total'] for c in cartes_chiffrees if c['cle'] == 'versement'), 0
        ),
        'abonnements_total': abonnements_qs.count(),
        'abonnements_actifs': abonnements_actifs,
        'abonnements_expires': abonnements_expires,
        'abonnements_alerte': abonnements_alerte,
        'abonnements_montant': int(_total(abonnements_qs)),
        'salaires_total': int(salaires_total),
        'salaires_mois': int(salaires_mois),
        'salaires_enseignants_count': salaires_enseignants_count,
    }
    return render(request, 'depenses/recouvrement/hub.html', context)


# ---------------------------------------------------------------------------
# Sous-modules simples : cuisine, documents, versements
# ---------------------------------------------------------------------------

@login_required
def dashboard_module(request, module):
    """Tableau de bord d'un sous-module simple : indicateurs + registre filtré."""
    config = _config_module(module)
    qs, filtres = _filtrer_operations(request, config, _queryset_module(request, config))
    qs = qs.select_related('cree_par')

    stats = _statistiques_operations(qs)
    serie = _serie_mensuelle(qs)
    maximum_serie = max([p['total'] for p in serie] or [0]) or 1

    paginator = Paginator(qs, 25)
    page = paginator.get_page(request.GET.get('page'))

    context = {
        'module': module,
        'config': config,
        'titre_page': config['titre'],
        'operations': page,
        'page_obj': page,
        'stats': stats,
        'serie': [dict(p, hauteur=int(p['total'] * 100 / maximum_serie)) for p in serie],
        'filtres': filtres,
    }
    return render(request, 'depenses/recouvrement/module_dashboard.html', context)


@login_required
@can_add_expenses
def ajouter_operation(request, module):
    """Création d'une opération : l'école et l'auteur sont posés côté serveur."""
    config = _config_module(module)
    if request.method == 'POST':
        form = config['form'](request.POST)
        if form.is_valid():
            operation = form.save(commit=False)
            operation.cree_par = request.user
            operation.ecole = user_school(request.user)
            operation.save()
            messages.success(request, f"{config['singulier']} enregistrée.")
            return redirect('depenses:recouvrement_dashboard_module', module=module)
        messages.error(request, "Le formulaire contient des erreurs.")
    else:
        form = config['form']()

    context = {
        'module': module,
        'config': config,
        'form': form,
        'titre_page': f"Nouvelle {config['singulier'].lower()}",
        'mode': 'creation',
    }
    return render(request, 'depenses/recouvrement/module_form.html', context)


@login_required
@can_modify_expenses
def modifier_operation(request, module, pk):
    config = _config_module(module)
    operation = _objet_module(request, config, pk)

    if request.method == 'POST':
        form = config['form'](request.POST, instance=operation)
        if form.is_valid():
            form.save()
            messages.success(request, f"{config['singulier']} modifiée.")
            return redirect('depenses:recouvrement_dashboard_module', module=module)
        messages.error(request, "Le formulaire contient des erreurs.")
    else:
        form = config['form'](instance=operation)

    context = {
        'module': module,
        'config': config,
        'form': form,
        'operation': operation,
        'titre_page': f"Modifier — {config['singulier'].lower()}",
        'mode': 'modification',
    }
    return render(request, 'depenses/recouvrement/module_form.html', context)


@login_required
@can_delete_expenses
def supprimer_operation(request, module, pk):
    config = _config_module(module)
    operation = _objet_module(request, config, pk)

    if request.method == 'POST':
        operation.delete()
        messages.success(request, f"{config['singulier']} supprimée.")
        return redirect('depenses:recouvrement_dashboard_module', module=module)

    context = {
        'module': module,
        'config': config,
        'operation': operation,
        'titre_page': f"Supprimer — {config['singulier'].lower()}",
    }
    return render(request, 'depenses/recouvrement/confirm_delete.html', context)


# ---------------------------------------------------------------------------
# Exports (Excel / PDF)
# ---------------------------------------------------------------------------

def _reponse_excel(wb, nom_fichier: str) -> HttpResponse:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    horodatage = timezone.now().strftime('%Y%m%d')
    response['Content-Disposition'] = f'attachment; filename="{nom_fichier}_{horodatage}.xlsx"'
    return response


def _classeur(titre: str, entetes: list, lignes: list):
    """Classeur openpyxl avec en-tête en gras et colonnes dimensionnées."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = titre[:31]

    ws.append(entetes)
    entete_fond = PatternFill('solid', fgColor='D9E1F2')
    for cellule in ws[1]:
        cellule.font = Font(bold=True)
        cellule.fill = entete_fond
        cellule.alignment = Alignment(horizontal='center')

    for ligne in lignes:
        ws.append(ligne)

    for index in range(1, len(entetes) + 1):
        longueurs = [len(str(entetes[index - 1]))]
        longueurs += [len(str(l[index - 1])) for l in lignes if len(l) >= index]
        ws.column_dimensions[get_column_letter(index)].width = min(max(longueurs) + 4, 45)

    ws.freeze_panes = 'A2'
    return wb


def _reponse_pdf_tableau(titre, sous_titre, entetes, lignes, nom_fichier, largeurs=None):
    """Tableau PDF paysage avec filigrane, en-tête et pied de page paginé."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=14 * mm,
        title=titre,
    )

    styles = getSampleStyleSheet()
    style_titre = ParagraphStyle(
        'TitreRecouvrement', parent=styles['Title'], fontSize=15, spaceAfter=2,
    )
    style_sous_titre = ParagraphStyle(
        'SousTitreRecouvrement', parent=styles['Normal'],
        fontSize=9, textColor=colors.HexColor('#555555'), alignment=1,
    )
    style_cellule = ParagraphStyle(
        'CelluleRecouvrement', parent=styles['Normal'], fontSize=8, leading=10,
    )

    donnees = [entetes]
    for ligne in lignes:
        donnees.append([Paragraph(str(v), style_cellule) for v in ligne])

    table = Table(donnees, colWidths=largeurs, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8.5),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#b0b0b0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f4f6f8')]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))

    def _decor(canvas_pdf, document):
        try:
            from ecole_moderne.pdf_utils import draw_logo_watermark
            draw_logo_watermark(canvas_pdf, document.pagesize[0], document.pagesize[1], opacity=0.04)
        except Exception:
            pass
        canvas_pdf.setFont('Helvetica', 7.5)
        canvas_pdf.setFillColor(colors.HexColor('#666666'))
        canvas_pdf.drawString(12 * mm, 8 * mm, f"Édité le {timezone.localtime():%d/%m/%Y à %H:%M}")
        canvas_pdf.drawRightString(
            document.pagesize[0] - 12 * mm, 8 * mm, f"Page {document.page}"
        )

    elements = [Paragraph(titre, style_titre)]
    if sous_titre:
        elements.append(Paragraph(sous_titre, style_sous_titre))
    elements.append(Spacer(1, 8))
    elements.append(table)

    doc.build(elements, onFirstPage=_decor, onLaterPages=_decor)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    horodatage = timezone.now().strftime('%Y%m%d')
    response['Content-Disposition'] = f'inline; filename="{nom_fichier}_{horodatage}.pdf"'
    return response


def _lignes_module(config, qs):
    """Lignes d'export d'un sous-module, dans l'ordre des en-têtes."""
    return [
        [
            operation.date.strftime('%d/%m/%Y'),
            getattr(operation, config['champ_libelle']) or '',
            int(operation.montant or 0),
            operation.observation or '',
        ]
        for operation in qs
    ]


def _libelle_periode(filtres) -> str:
    if filtres.get('du') and filtres.get('au'):
        return f"Période du {filtres['du']} au {filtres['au']}"
    if filtres.get('du'):
        return f"À partir du {filtres['du']}"
    if filtres.get('au'):
        return f"Jusqu'au {filtres['au']}"
    return "Toutes les opérations"


@login_required
def export_module_excel(request, module):
    """Export Excel du registre, filtres de la page repris tels quels."""
    config = _config_module(module)
    qs, _ = _filtrer_operations(request, config, _queryset_module(request, config))

    entetes = ['Date', config['entete_libelle'], 'Montant (GNF)', 'Observation']
    lignes = _lignes_module(config, qs)
    lignes.append(['', 'TOTAL', int(_total(qs)), ''])

    return _reponse_excel(_classeur(config['titre'], entetes, lignes), config['fichier'])


@login_required
def export_module_pdf(request, module):
    """Export PDF du registre, mêmes filtres que l'écran."""
    from reportlab.lib.units import mm

    config = _config_module(module)
    qs, filtres = _filtrer_operations(request, config, _queryset_module(request, config))

    entetes = ['Date', config['entete_libelle'], 'Montant (GNF)', 'Observation']
    lignes = [
        [l[0], l[1], f"{l[2]:,.0f}".replace(',', ' '), l[3]]
        for l in _lignes_module(config, qs)
    ]
    total = _total(qs)
    lignes.append(['', 'TOTAL', f"{total:,.0f}".replace(',', ' '), f"{qs.count()} opération(s)"])

    sous_titre = _libelle_periode(filtres)
    return _reponse_pdf_tableau(
        config['titre'], sous_titre, entetes, lignes, config['fichier'],
        largeurs=[25 * mm, 85 * mm, 35 * mm, 120 * mm],
    )


# ---------------------------------------------------------------------------
# Abonnements informatique
# ---------------------------------------------------------------------------

def _queryset_abonnements(request):
    qs = AbonnementInformatique.objects.select_related(
        'eleve', 'eleve__classe', 'eleve__classe__ecole'
    )
    return filter_by_user_school(qs, request.user, 'eleve__classe__ecole')


def _filtrer_abonnements(request, qs):
    """Recherche par matricule/nom et filtre de statut (actif, alerte, expiré)."""
    recherche = (request.GET.get('q') or '').strip()
    if recherche:
        qs = qs.filter(
            Q(eleve__matricule__icontains=recherche)
            | Q(eleve__nom__icontains=recherche)
            | Q(eleve__prenom__icontains=recherche)
        )

    aujourdhui = timezone.localdate()
    statut = (request.GET.get('statut') or '').strip().lower()
    if statut == 'actif':
        qs = qs.filter(date_fin__gte=aujourdhui)
    elif statut == 'expire':
        qs = qs.filter(date_fin__lt=aujourdhui)
    elif statut == 'alerte':
        qs = qs.filter(
            date_fin__gte=aujourdhui,
            date_fin__lte=aujourdhui + timedelta(days=AbonnementInformatique.ALERTE_JOURS_DEFAUT),
        )

    classe_id = (request.GET.get('classe') or '').strip()
    if classe_id.isdigit():
        qs = qs.filter(eleve__classe_id=int(classe_id))

    return qs, {'q': recherche, 'statut': statut, 'classe': classe_id}


@login_required
def dashboard_informatique(request):
    """Tableau de bord des abonnements informatique, avec alertes de fin."""
    qs = _queryset_abonnements(request)
    aujourdhui = timezone.localdate()
    seuil = aujourdhui + timedelta(days=AbonnementInformatique.ALERTE_JOURS_DEFAUT)

    actifs = qs.filter(date_fin__gte=aujourdhui)
    expires = qs.filter(date_fin__lt=aujourdhui)
    alertes = actifs.filter(date_fin__lte=seuil).order_by('date_fin')

    debut_mois = aujourdhui.replace(day=1)
    serie = _serie_mensuelle(qs)
    maximum_serie = max([p['total'] for p in serie] or [0]) or 1

    context = {
        'titre_page': 'Abonnements informatique',
        'total': qs.count(),
        'actifs': actifs.count(),
        'expires': expires.count(),
        'alertes_nombre': alertes.count(),
        'montant_total': int(_total(qs)),
        'montant_mois': int(_total(qs.filter(date__gte=debut_mois))),
        'alertes': alertes.select_related('eleve', 'eleve__classe')[:15],
        'expires_recents': expires.order_by('-date_fin').select_related(
            'eleve', 'eleve__classe'
        )[:10],
        'derniers': qs.order_by('-date_creation')[:10],
        'serie': [dict(p, hauteur=int(p['total'] * 100 / maximum_serie)) for p in serie],
    }
    return render(request, 'depenses/recouvrement/informatique_dashboard.html', context)


@login_required
def liste_informatique(request):
    """Liste des abonnements avec recherche par matricule et filtres."""
    qs, filtres = _filtrer_abonnements(request, _queryset_abonnements(request))
    qs = qs.order_by('date_fin', 'eleve__nom')

    paginator = Paginator(qs, 25)
    page = paginator.get_page(request.GET.get('page'))

    classes = []
    try:
        from eleves.models import Classe
        classes = filter_by_user_school(Classe.objects.all(), request.user, 'ecole').order_by('nom')
    except Exception:
        classes = []

    context = {
        'titre_page': 'Abonnements informatique',
        'abonnements': page,
        'page_obj': page,
        'filtres': filtres,
        'classes': classes,
        'total_filtre': int(_total(qs)),
        'nombre_filtre': qs.count(),
    }
    return render(request, 'depenses/recouvrement/informatique_liste.html', context)


def _eleves_autorises(request):
    return filter_by_user_school(
        Eleve.objects.select_related('classe'), request.user, 'classe__ecole'
    ).order_by('nom', 'prenom')


@login_required
@can_add_expenses
def ajouter_abonnement_informatique(request):
    """Nouvel abonnement : l'élève se choisit par matricule ou par nom."""
    eleves = _eleves_autorises(request)

    if request.method == 'POST':
        form = AbonnementInformatiqueForm(request.POST)
        form.fields['eleve'].queryset = eleves
        if form.is_valid():
            abonnement = form.save(commit=False)
            abonnement.cree_par = request.user
            abonnement.save()
            messages.success(
                request,
                f"Abonnement informatique enregistré pour {abonnement.eleve}."
            )
            return redirect('depenses:recouvrement_informatique_liste')
        messages.error(request, "Le formulaire contient des erreurs.")
    else:
        form = AbonnementInformatiqueForm()
        form.fields['eleve'].queryset = eleves
        eleve_id = request.GET.get('eleve')
        if eleve_id and str(eleve_id).isdigit():
            form.initial['eleve'] = int(eleve_id)

    context = {
        'titre_page': "Nouvel abonnement informatique",
        'form': form,
        'mode': 'creation',
    }
    return render(request, 'depenses/recouvrement/informatique_form.html', context)


@login_required
@can_modify_expenses
def modifier_abonnement_informatique(request, pk):
    abonnement = get_object_or_404(_queryset_abonnements(request), pk=pk)
    eleves = _eleves_autorises(request)

    if request.method == 'POST':
        form = AbonnementInformatiqueForm(request.POST, instance=abonnement)
        form.fields['eleve'].queryset = eleves
        if form.is_valid():
            form.save()
            messages.success(request, "Abonnement informatique modifié.")
            return redirect('depenses:recouvrement_informatique_liste')
        messages.error(request, "Le formulaire contient des erreurs.")
    else:
        form = AbonnementInformatiqueForm(instance=abonnement)
        form.fields['eleve'].queryset = eleves

    context = {
        'titre_page': "Modifier l'abonnement informatique",
        'form': form,
        'abonnement': abonnement,
        'mode': 'modification',
    }
    return render(request, 'depenses/recouvrement/informatique_form.html', context)


@login_required
@can_delete_expenses
def supprimer_abonnement_informatique(request, pk):
    abonnement = get_object_or_404(_queryset_abonnements(request), pk=pk)

    if request.method == 'POST':
        abonnement.delete()
        messages.success(request, "Abonnement informatique supprimé.")
        return redirect('depenses:recouvrement_informatique_liste')

    context = {
        'titre_page': "Supprimer l'abonnement informatique",
        'abonnement': abonnement,
        'config': {'titre': 'Abonnements informatique', 'icone': 'fa-desktop'},
        'operation': abonnement,
        'module': None,
    }
    return render(request, 'depenses/recouvrement/informatique_confirm_delete.html', context)


@login_required
def rechercher_eleve_informatique(request):
    """Recherche d'élèves par matricule ou nom, pour le formulaire (JSON)."""
    recherche = (request.GET.get('q') or '').strip()
    resultats = []
    if len(recherche) >= 2:
        eleves = _eleves_autorises(request).filter(
            Q(matricule__icontains=recherche)
            | Q(nom__icontains=recherche)
            | Q(prenom__icontains=recherche)
        )[:20]
        resultats = [
            {
                'id': eleve.id,
                'matricule': eleve.matricule,
                'nom_complet': f"{eleve.prenom} {eleve.nom}".strip(),
                'classe': getattr(eleve.classe, 'nom', ''),
            }
            for eleve in eleves
        ]
    return JsonResponse({'resultats': resultats})


def _lignes_abonnements(qs):
    lignes = []
    for abonnement in qs:
        eleve = abonnement.eleve
        lignes.append([
            abonnement.numero_carte,
            eleve.matricule or '',
            f"{eleve.prenom} {eleve.nom}".strip(),
            getattr(eleve.classe, 'nom', ''),
            abonnement.date.strftime('%d/%m/%Y'),
            int(abonnement.montant or 0),
            abonnement.date_debut.strftime('%d/%m/%Y'),
            abonnement.date_fin.strftime('%d/%m/%Y'),
            abonnement.jours_restants,
            abonnement.statut_libelle,
            abonnement.observation or '',
        ])
    return lignes


@login_required
def export_informatique_excel(request):
    qs, _ = _filtrer_abonnements(request, _queryset_abonnements(request))
    qs = qs.order_by('date_fin', 'eleve__nom')

    entetes = [
        'N° carte', 'Matricule', 'Élève', 'Classe', 'Date', 'Montant (GNF)',
        'Début', 'Fin', 'Jours restants', 'Statut', 'Observation',
    ]
    lignes = _lignes_abonnements(qs)
    lignes.append(['', '', 'TOTAL', '', '', int(_total(qs)), '', '', '', '', ''])

    return _reponse_excel(
        _classeur('Abonnements informatique', entetes, lignes), 'abonnements_informatique'
    )


@login_required
def export_informatique_pdf(request):
    from reportlab.lib.units import mm

    qs, filtres = _filtrer_abonnements(request, _queryset_abonnements(request))
    qs = qs.order_by('date_fin', 'eleve__nom')

    entetes = [
        'N° carte', 'Matricule', 'Élève', 'Classe', 'Montant (GNF)',
        'Début', 'Fin', 'Jours rest.', 'Statut',
    ]
    lignes = []
    for ligne in _lignes_abonnements(qs):
        lignes.append([
            ligne[0], ligne[1], ligne[2], ligne[3],
            f"{ligne[5]:,.0f}".replace(',', ' '),
            ligne[6], ligne[7], ligne[8], ligne[9],
        ])
    lignes.append([
        '', '', 'TOTAL', '', f"{_total(qs):,.0f}".replace(',', ' '),
        '', '', '', f"{qs.count()} abonné(s)",
    ])

    libelles_statut = {
        'actif': 'Abonnements actifs', 'expire': 'Abonnements expirés',
        'alerte': "Abonnements proches de l'échéance",
    }
    sous_titre = libelles_statut.get(filtres.get('statut'), 'Tous les abonnements')
    if filtres.get('q'):
        sous_titre += f" — recherche « {filtres['q']} »"

    return _reponse_pdf_tableau(
        'Abonnements informatique', sous_titre, entetes, lignes,
        'abonnements_informatique',
        largeurs=[24 * mm, 26 * mm, 55 * mm, 28 * mm, 28 * mm, 24 * mm, 24 * mm, 22 * mm, 26 * mm],
    )


@login_required
def carte_abonnement_informatique(request, pk):
    """Carte d'abonnement au format carte de visite (85,6 × 54 mm)."""
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    from reportlab.pdfgen import canvas as pdf_canvas

    abonnement = get_object_or_404(_queryset_abonnements(request), pk=pk)
    eleve = abonnement.eleve
    ecole = getattr(getattr(eleve, 'classe', None), 'ecole', None)

    largeur, hauteur = 85.6 * mm, 54 * mm
    buffer = io.BytesIO()
    c = pdf_canvas.Canvas(buffer, pagesize=(largeur, hauteur))
    c.setTitle(f"Carte abonnement informatique {abonnement.numero_carte}")

    couleur_statut = {
        'ACTIF': colors.HexColor('#1e7e34'),
        'BIENTOT': colors.HexColor('#b8860b'),
        'EXPIRE': colors.HexColor('#b02a37'),
    }[abonnement.statut]

    # Fond et bandeau
    c.setFillColor(colors.HexColor('#f7f9fc'))
    c.rect(0, 0, largeur, hauteur, stroke=0, fill=1)
    c.setFillColor(colors.HexColor('#1f3864'))
    c.rect(0, hauteur - 13 * mm, largeur, 13 * mm, stroke=0, fill=1)

    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 9)
    c.drawString(5 * mm, hauteur - 6 * mm, "CARTE D'ABONNEMENT INFORMATIQUE")
    c.setFont('Helvetica', 6.5)
    nom_ecole = getattr(ecole, 'nom', '') or ''
    if nom_ecole:
        c.drawString(5 * mm, hauteur - 10 * mm, nom_ecole[:52])

    # Photo de l'élève (ou initiales)
    photo_largeur, photo_hauteur = 18 * mm, 22 * mm
    x_photo, y_photo = largeur - photo_largeur - 5 * mm, hauteur - 13 * mm - photo_hauteur - 3 * mm
    photo_dessinee = False
    photo = getattr(eleve, 'photo', None)
    if photo:
        # `.path` lève une exception quand aucun fichier n'est associé : on ne
        # l'interroge donc qu'après avoir vérifié que le champ est renseigné.
        try:
            import os
            chemin_photo = photo.path
            if os.path.exists(chemin_photo):
                c.drawImage(
                    ImageReader(chemin_photo), x_photo, y_photo,
                    width=photo_largeur, height=photo_hauteur,
                    preserveAspectRatio=True, mask='auto',
                )
                photo_dessinee = True
        except Exception:
            photo_dessinee = False
    if not photo_dessinee:
        c.setStrokeColor(colors.HexColor('#b0b0b0'))
        c.setFillColor(colors.HexColor('#e9ecef'))
        c.rect(x_photo, y_photo, photo_largeur, photo_hauteur, stroke=1, fill=1)
        initiales = ''.join(
            partie[0].upper()
            for partie in f"{eleve.prenom} {eleve.nom}".split()[:2] if partie
        ) or 'E'
        c.setFillColor(colors.HexColor('#6c757d'))
        c.setFont('Helvetica-Bold', 13)
        c.drawCentredString(x_photo + photo_largeur / 2, y_photo + photo_hauteur / 2 - 4, initiales)

    # Identité et abonnement
    y = hauteur - 18 * mm
    c.setFillColor(colors.HexColor('#1f3864'))
    c.setFont('Helvetica-Bold', 10)
    c.drawString(5 * mm, y, f"{eleve.prenom} {eleve.nom}".strip()[:28])

    c.setFillColor(colors.HexColor('#333333'))
    c.setFont('Helvetica', 7)
    y -= 5 * mm

    def ligne(libelle, valeur):
        nonlocal y
        c.setFont('Helvetica-Bold', 7)
        c.drawString(5 * mm, y, f"{libelle} :")
        c.setFont('Helvetica', 7)
        c.drawString(25 * mm, y, str(valeur)[:26])
        y -= 4.2 * mm

    ligne('Matricule', eleve.matricule or '—')
    ligne('Classe', getattr(eleve.classe, 'nom', '—') or '—')
    ligne('Montant', f"{int(abonnement.montant or 0):,}".replace(',', ' ') + ' GNF')
    ligne('Validité', f"{abonnement.date_debut:%d/%m/%Y} → {abonnement.date_fin:%d/%m/%Y}")

    # Pied : numéro de carte et statut
    c.setFillColor(colors.HexColor('#666666'))
    c.setFont('Helvetica', 6.5)
    c.drawString(5 * mm, 4 * mm, f"N° {abonnement.numero_carte}")

    c.setFillColor(couleur_statut)
    c.roundRect(largeur - 30 * mm, 3 * mm, 25 * mm, 5.5 * mm, 2, stroke=0, fill=1)
    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 6.5)
    libelle_statut = abonnement.statut_libelle.upper()
    if abonnement.statut == 'ACTIF':
        libelle_statut = f"ACTIF · {abonnement.jours_restants} J"
    c.drawCentredString(largeur - 17.5 * mm, 5 * mm, libelle_statut)

    c.showPage()
    c.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = (
        f'inline; filename="carte_informatique_{abonnement.numero_carte}.pdf"'
    )
    return response


# ---------------------------------------------------------------------------
# Recouvrement des salaires enseignants
# ---------------------------------------------------------------------------

MOIS_NOMS = [
    '', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
    'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre',
]

NIVEAUX_ENSEIGNANT = ('Maternelle', 'Primaire', 'Collège', 'Lycée', 'Autre')


def _niveau_enseignants(enseignant_ids):
    """Associe à chaque enseignant secondaire son niveau réel (Collège/Lycée)
    d'après sa classe affectée la plus récente.
    """
    niveau_classe = {}
    affectations = (
        AffectationClasse.objects
        .filter(enseignant_id__in=enseignant_ids)
        .select_related('classe')
        .order_by('-date_debut')
    )
    for affectation in affectations:
        if affectation.enseignant_id in niveau_classe:
            continue
        if affectation.classe:
            niveau_classe[affectation.enseignant_id] = affectation.classe.niveau or ''
    return niveau_classe


def _niveau_label(enseignant, niveau_classe_par_enseignant):
    if enseignant.type_enseignant == 'MATERNELLE':
        return 'Maternelle'
    if enseignant.type_enseignant == 'PRIMAIRE':
        return 'Primaire'
    if enseignant.type_enseignant == 'SECONDAIRE':
        niveau_classe = niveau_classe_par_enseignant.get(enseignant.id, '')
        if niveau_classe.startswith('LYCEE'):
            return 'Lycée'
        return 'Collège'
    return 'Autre'


def _collecter_salaires_payes(request):
    """Construit le tableau pivot (enseignant × mois) des salaires payés,
    cloisonné à l'école de l'utilisateur, ainsi que les indicateurs du
    tableau de bord.
    """
    etats_qs = filter_by_user_school(
        EtatSalaire.objects.filter(paye=True).select_related('enseignant', 'periode'),
        request.user, 'enseignant__ecole',
    )

    colonnes = [
        {'annee': annee, 'mois': mois, 'label': f"{MOIS_NOMS[mois]} {annee}"}
        for annee, mois in sorted(set(etats_qs.values_list('periode__annee', 'periode__mois')))
    ]
    index_colonne = {(c['annee'], c['mois']): i for i, c in enumerate(colonnes)}

    enseignant_ids = list(etats_qs.values_list('enseignant_id', flat=True).distinct())
    niveau_classe_par_enseignant = _niveau_enseignants(enseignant_ids)

    lignes_par_enseignant = {}
    for etat in etats_qs:
        ens = etat.enseignant
        ligne = lignes_par_enseignant.get(ens.id)
        if ligne is None:
            ligne = {
                'enseignant': ens,
                'niveau': _niveau_label(ens, niveau_classe_par_enseignant),
                'montants': [Decimal('0')] * len(colonnes),
                'total': Decimal('0'),
            }
            lignes_par_enseignant[ens.id] = ligne
        idx = index_colonne[(etat.periode.annee, etat.periode.mois)]
        ligne['montants'][idx] += etat.salaire_net
        ligne['total'] += etat.salaire_net

    lignes = sorted(
        lignes_par_enseignant.values(),
        key=lambda r: (r['enseignant'].nom, r['enseignant'].prenoms),
    )

    totaux_par_niveau = {niveau: Decimal('0') for niveau in NIVEAUX_ENSEIGNANT}
    for ligne in lignes:
        totaux_par_niveau[ligne['niveau']] += ligne['total']

    evolution_mensuelle = [
        {
            'label': c['label'],
            'total': sum((l['montants'][i] for l in lignes), Decimal('0')),
        }
        for i, c in enumerate(colonnes)
    ]
    maximum_evolution = max([e['total'] for e in evolution_mensuelle] or [Decimal('0')]) or Decimal('1')
    for e in evolution_mensuelle:
        e['hauteur'] = int((e['total'] * 100) / maximum_evolution)

    aujourdhui = timezone.localdate()
    idx_mois_courant = index_colonne.get((aujourdhui.year, aujourdhui.month))
    total_mois_courant = (
        sum((l['montants'][idx_mois_courant] for l in lignes), Decimal('0'))
        if idx_mois_courant is not None else Decimal('0')
    )
    total_general = sum((l['total'] for l in lignes), Decimal('0'))

    return {
        'colonnes': colonnes,
        'lignes': lignes,
        'totaux_par_niveau': totaux_par_niveau,
        'evolution_mensuelle': evolution_mensuelle,
        'total_mois_courant': total_mois_courant,
        'total_general': total_general,
        'nb_enseignants': len(lignes),
    }


@login_required
def salaires_dashboard(request):
    """Tableau de bord Recouvrement > Salaires enseignants."""
    donnees = _collecter_salaires_payes(request)

    context = {
        'titre_page': 'Salaires enseignants',
        **donnees,
    }
    return render(request, 'depenses/recouvrement/salaires_dashboard.html', context)


@login_required
def export_salaires_excel(request):
    """Export Excel du registre des salaires payés (pivot enseignant × mois)."""
    donnees = _collecter_salaires_payes(request)

    entetes = ['Enseignant', 'Niveau'] + [c['label'] for c in donnees['colonnes']] + ['Total']
    lignes = []
    for ligne in donnees['lignes']:
        lignes.append(
            [ligne['enseignant'].nom_complet, ligne['niveau']]
            + [float(m) for m in ligne['montants']]
            + [float(ligne['total'])]
        )

    wb = _classeur('Salaires enseignants', entetes, lignes)
    return _reponse_excel(wb, 'salaires_enseignants')
