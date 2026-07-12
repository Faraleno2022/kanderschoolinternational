from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count, F, DecimalField, ExpressionWrapper
from django.db import models
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import datetime, date
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from .models_logistique import (
    CategorieArticle, Article, BienEtablissement, 
    MouvementStock, Inventaire, LigneInventaire
)
from .forms import (
    CategorieArticleForm, ArticleForm, BienEtablissementForm, MouvementStockForm,
    InventaireForm, LigneInventaireForm
)


@login_required
def dashboard_logistique(request):
    """Dashboard principal de la logistique"""
    from utilisateurs.utils import user_school

    ecole = user_school(request.user)

    # Filtres de base par école
    articles_qs = Article.objects.filter(actif=True)
    biens_qs = BienEtablissement.objects.filter(actif=True)
    mouvements_qs = MouvementStock.objects.all()
    if ecole:
        articles_qs = articles_qs.filter(cree_par__profil__ecole=ecole)
        biens_qs = biens_qs.filter(cree_par__profil__ecole=ecole)
        mouvements_qs = mouvements_qs.filter(cree_par__profil__ecole=ecole)

    # Statistiques générales
    total_articles = articles_qs.count()
    total_biens = biens_qs.count()

    # Valeur totale du stock
    valeur_article = ExpressionWrapper(
        F('stock_actuel') * F('prix_unitaire'),
        output_field=DecimalField(max_digits=20, decimal_places=0),
    )
    valeur_stock = articles_qs.aggregate(total=Sum(valeur_article))

    # Articles en alerte (stock minimum)
    articles_alerte = articles_qs.filter(
        stock_actuel__lte=models.F('stock_minimum')
    ).count()

    # Derniers mouvements
    derniers_mouvements = mouvements_qs.select_related(
        'article', 'cree_par'
    ).order_by('-date_mouvement')[:10]

    # Répartition par catégorie
    repartition_categories = CategorieArticle.objects.annotate(
        nb_articles=Count('articles'),
        valeur_totale=Sum(
            ExpressionWrapper(
                F('articles__stock_actuel') * F('articles__prix_unitaire'),
                output_field=DecimalField(max_digits=20, decimal_places=0),
            )
        ),
    ).filter(actif=True)

    # Biens nécessitant une maintenance
    biens_maintenance = biens_qs.filter(
        date_prochaine_maintenance__lte=date.today()
    ).count()
    
    context = {
        'titre_page': 'Dashboard Logistique',
        'total_articles': total_articles,
        'total_biens': total_biens,
        'valeur_stock': valeur_stock.get('total', 0) or 0,
        'articles_alerte': articles_alerte,
        'derniers_mouvements': derniers_mouvements,
        'repartition_categories': repartition_categories,
        'biens_maintenance': biens_maintenance,
    }
    
    return render(request, 'depenses/logistique/dashboard.html', context)


@login_required
def liste_articles(request):
    """Liste des articles en stock"""
    from utilisateurs.utils import user_school

    # Filtres
    q = request.GET.get('q', '')
    categorie_id = request.GET.get('categorie', '')
    etat = request.GET.get('etat', '')
    alerte = request.GET.get('alerte', '')

    articles = Article.objects.select_related('categorie').filter(actif=True)
    # Sécurité : filtrer par école
    ecole = user_school(request.user)
    if ecole:
        articles = articles.filter(cree_par__profil__ecole=ecole)
    
    if q:
        articles = articles.filter(
            Q(code_article__icontains=q) |
            Q(nom__icontains=q) |
            Q(marque__icontains=q) |
            Q(reference__icontains=q)
        )
    
    if categorie_id:
        articles = articles.filter(categorie_id=categorie_id)
    
    if etat:
        articles = articles.filter(etat=etat)
    
    if alerte == 'oui':
        articles = articles.filter(stock_actuel__lte=models.F('stock_minimum'))
    
    categories = CategorieArticle.objects.filter(actif=True)
    
    context = {
        'titre_page': 'Stock & Articles',
        'articles': articles,
        'categories': categories,
        'q': q,
        'categorie_id': categorie_id,
        'etat': etat,
        'alerte': alerte,
    }
    
    return render(request, 'depenses/logistique/liste_articles.html', context)


@login_required
def liste_biens(request):
    """Liste des biens de l'établissement"""
    from utilisateurs.utils import user_school

    # Filtres
    q = request.GET.get('q', '')
    type_bien = request.GET.get('type_bien', '')
    etat = request.GET.get('etat', '')

    biens = BienEtablissement.objects.filter(actif=True)
    # Sécurité : filtrer par école
    ecole = user_school(request.user)
    if ecole:
        biens = biens.filter(cree_par__profil__ecole=ecole)
    
    if q:
        biens = biens.filter(
            Q(code_bien__icontains=q) |
            Q(nom__icontains=q) |
            Q(localisation__icontains=q)
        )
    
    if type_bien:
        biens = biens.filter(type_bien=type_bien)
    
    if etat:
        biens = biens.filter(etat=etat)
    
    context = {
        'titre_page': 'Biens de l\'Établissement',
        'biens': biens,
        'q': q,
        'type_bien': type_bien,
        'etat': etat,
    }
    
    return render(request, 'depenses/logistique/liste_biens.html', context)


@login_required
def creer_bien(request):
    """Créer un bien de l'établissement"""
    
    if request.method == 'POST':
        form = BienEtablissementForm(request.POST, request.FILES)
        if form.is_valid():
            bien = form.save(commit=False)
            bien.cree_par = request.user
            
            # Générer le code du bien si non fourni
            if not bien.code_bien:
                today = date.today()
                prefix = f"BIEN-{today.strftime('%Y%m%d')}"
                last_bien = BienEtablissement.objects.filter(
                    code_bien__startswith=prefix
                ).order_by('-code_bien').first()
                
                if last_bien:
                    last_num = int(last_bien.code_bien.split('-')[-1])
                    bien.code_bien = f"{prefix}-{last_num + 1:04d}"
                else:
                    bien.code_bien = f"{prefix}-0001"
            
            bien.save()
            messages.success(request, f'Bien "{bien.nom}" créé avec succès.')
            return redirect('depenses:liste_biens')
    else:
        form = BienEtablissementForm()
    
    context = {
        'titre_page': 'Nouveau Bien',
        'form': form,
    }
    
    return render(request, 'depenses/logistique/form_bien.html', context)


@login_required
def modifier_bien(request, bien_id):
    """Modifier un bien de l'établissement"""
    
    bien = get_object_or_404(BienEtablissement, pk=bien_id)
    
    if request.method == 'POST':
        form = BienEtablissementForm(request.POST, request.FILES, instance=bien)
        if form.is_valid():
            form.save()
            messages.success(request, f'Bien "{bien.nom}" modifié avec succès.')
            return redirect('depenses:liste_biens')
    else:
        form = BienEtablissementForm(instance=bien)
    
    context = {
        'titre_page': 'Modifier Bien',
        'form': form,
        'bien': bien,
    }
    
    return render(request, 'depenses/logistique/form_bien.html', context)


@login_required
def liste_mouvements(request):
    """Liste des mouvements de stock"""
    from utilisateurs.utils import user_school

    # Filtres
    article_id = request.GET.get('article', '')
    type_mouvement = request.GET.get('type', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')

    mouvements = MouvementStock.objects.select_related(
        'article', 'cree_par'
    ).all()
    # Sécurité : filtrer par école
    ecole = user_school(request.user)
    if ecole:
        mouvements = mouvements.filter(cree_par__profil__ecole=ecole)
    
    if article_id:
        mouvements = mouvements.filter(article_id=article_id)
    
    if type_mouvement:
        mouvements = mouvements.filter(type_mouvement=type_mouvement)
    
    if date_debut:
        mouvements = mouvements.filter(date_mouvement__gte=date_debut)
    
    if date_fin:
        mouvements = mouvements.filter(date_mouvement__lte=date_fin)
    
    articles = Article.objects.filter(actif=True)
    
    context = {
        'titre_page': 'Mouvements de Stock',
        'mouvements': mouvements,
        'articles': articles,
        'article_id': article_id,
        'type_mouvement': type_mouvement,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    
    return render(request, 'depenses/logistique/liste_mouvements.html', context)


@login_required
def liste_inventaires(request):
    """Liste des inventaires"""
    
    inventaires = Inventaire.objects.select_related(
        'cree_par', 'valide_par'
    ).order_by('-date_inventaire')
    
    context = {
        'titre_page': 'Inventaires',
        'inventaires': inventaires,
    }
    
    return render(request, 'depenses/logistique/liste_inventaires.html', context)


@login_required
def creer_mouvement(request):
    """Créer un mouvement de stock"""
    
    if request.method == 'POST':
        form = MouvementStockForm(request.POST)
        if form.is_valid():
            mouvement = form.save(commit=False)
            mouvement.cree_par = request.user
            
            # Générer le numéro de mouvement
            today = date.today()
            prefix = f"MVT-{today.strftime('%Y%m%d')}"
            last_mvt = MouvementStock.objects.filter(
                numero_mouvement__startswith=prefix
            ).order_by('-numero_mouvement').first()
            
            if last_mvt:
                last_num = int(last_mvt.numero_mouvement.split('-')[-1])
                mouvement.numero_mouvement = f"{prefix}-{last_num + 1:04d}"
            else:
                mouvement.numero_mouvement = f"{prefix}-0001"
            
            mouvement.save()
            messages.success(request, 'Mouvement de stock créé avec succès.')
            return redirect('depenses:liste_mouvements')
    else:
        form = MouvementStockForm()
    
    context = {
        'titre_page': 'Nouveau Mouvement',
        'form': form,
    }
    
    return render(request, 'depenses/logistique/form_mouvement.html', context)


def _generer_code(modele, champ, prefixe):
    """Génère un code séquentiel du type PREFIXE-YYYYMMDD-0001."""
    today = date.today()
    base = f"{prefixe}-{today.strftime('%Y%m%d')}"
    dernier = modele.objects.filter(
        **{f"{champ}__startswith": base}
    ).order_by(f"-{champ}").first()
    if dernier:
        try:
            num = int(getattr(dernier, champ).split('-')[-1]) + 1
        except (ValueError, IndexError):
            num = 1
    else:
        num = 1
    return f"{base}-{num:04d}"


# ===== CRUD ARTICLES =====

@login_required
def detail_article(request, article_id):
    """Détail d'un article avec historique des mouvements"""
    article = get_object_or_404(Article, pk=article_id)
    mouvements = article.mouvements.select_related('cree_par').order_by('-date_mouvement')[:50]

    context = {
        'titre_page': f"Article : {article.nom}",
        'article': article,
        'mouvements': mouvements,
    }
    return render(request, 'depenses/logistique/detail_article.html', context)


@login_required
def creer_article(request):
    """Créer un article en stock"""
    if request.method == 'POST':
        form = ArticleForm(request.POST, request.FILES)
        if form.is_valid():
            article = form.save(commit=False)
            article.cree_par = request.user
            if not article.code_article:
                article.code_article = _generer_code(Article, 'code_article', 'ART')
            article.save()
            messages.success(request, f'Article "{article.nom}" créé avec succès.')
            return redirect('depenses:detail_article', article_id=article.pk)
    else:
        form = ArticleForm()

    context = {
        'titre_page': 'Nouvel Article',
        'form': form,
    }
    return render(request, 'depenses/logistique/form_article.html', context)


@login_required
def modifier_article(request, article_id):
    """Modifier un article en stock"""
    article = get_object_or_404(Article, pk=article_id)

    if request.method == 'POST':
        form = ArticleForm(request.POST, request.FILES, instance=article)
        if form.is_valid():
            form.save()
            messages.success(request, f'Article "{article.nom}" modifié avec succès.')
            return redirect('depenses:detail_article', article_id=article.pk)
    else:
        form = ArticleForm(instance=article)

    context = {
        'titre_page': f"Modifier : {article.nom}",
        'form': form,
        'article': article,
    }
    return render(request, 'depenses/logistique/form_article.html', context)


@login_required
def supprimer_article(request, article_id):
    """Désactiver (suppression logique) un article"""
    article = get_object_or_404(Article, pk=article_id)
    if request.method == 'POST':
        article.actif = False
        article.save(update_fields=['actif'])
        messages.success(request, f'Article "{article.nom}" archivé.')
        return redirect('depenses:liste_articles')

    context = {
        'titre_page': 'Archiver un article',
        'article': article,
    }
    return render(request, 'depenses/logistique/confirmer_suppression_article.html', context)


# ===== CATÉGORIES D'ARTICLES =====

@login_required
def gestion_categories_articles(request):
    """Liste et création des catégories d'articles"""
    if request.method == 'POST':
        form = CategorieArticleForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Catégorie créée avec succès.')
            return redirect('depenses:gestion_categories_articles')
    else:
        form = CategorieArticleForm()

    categories = CategorieArticle.objects.annotate(
        nb_articles=Count('articles')
    ).order_by('type_categorie', 'nom')

    context = {
        'titre_page': "Catégories d'articles",
        'form': form,
        'categories': categories,
    }
    return render(request, 'depenses/logistique/categories_articles.html', context)


@login_required
def modifier_categorie_article(request, categorie_id):
    """Modifier une catégorie d'article"""
    categorie = get_object_or_404(CategorieArticle, pk=categorie_id)
    if request.method == 'POST':
        form = CategorieArticleForm(request.POST, instance=categorie)
        if form.is_valid():
            form.save()
            messages.success(request, 'Catégorie modifiée avec succès.')
            return redirect('depenses:gestion_categories_articles')
    else:
        form = CategorieArticleForm(instance=categorie)

    categories = CategorieArticle.objects.annotate(
        nb_articles=Count('articles')
    ).order_by('type_categorie', 'nom')

    context = {
        'titre_page': "Modifier la catégorie",
        'form': form,
        'categories': categories,
        'categorie_edition': categorie,
    }
    return render(request, 'depenses/logistique/categories_articles.html', context)


# ===== INVENTAIRES =====

@login_required
def creer_inventaire(request):
    """Démarrer un inventaire : crée l'inventaire et une ligne par article actif"""
    if request.method == 'POST':
        form = InventaireForm(request.POST)
        if form.is_valid():
            inventaire = form.save(commit=False)
            inventaire.cree_par = request.user
            inventaire.numero_inventaire = _generer_code(Inventaire, 'numero_inventaire', 'INV')
            inventaire.save()

            # Générer les lignes pour tous les articles actifs
            articles = Article.objects.filter(actif=True)
            for article in articles:
                LigneInventaire.objects.create(
                    inventaire=inventaire,
                    article=article,
                    stock_theorique=article.stock_actuel,
                    stock_physique=article.stock_actuel,
                    prix_unitaire=article.prix_unitaire,
                    valeur_theorique=article.stock_actuel * article.prix_unitaire,
                    valeur_physique=article.stock_actuel * article.prix_unitaire,
                )
            inventaire.nombre_articles = articles.count()
            inventaire.save(update_fields=['nombre_articles'])

            messages.success(request, f'Inventaire {inventaire.numero_inventaire} démarré.')
            return redirect('depenses:detail_inventaire', inventaire_id=inventaire.pk)
    else:
        form = InventaireForm(initial={'date_inventaire': date.today()})

    context = {
        'titre_page': 'Nouvel Inventaire',
        'form': form,
    }
    return render(request, 'depenses/logistique/form_inventaire.html', context)


@login_required
def detail_inventaire(request, inventaire_id):
    """Saisie du stock physique et synthèse d'un inventaire"""
    inventaire = get_object_or_404(Inventaire, pk=inventaire_id)
    lignes = inventaire.lignes.select_related('article').order_by('article__nom')

    if request.method == 'POST' and inventaire.statut == 'EN_COURS':
        for ligne in lignes:
            valeur = request.POST.get(f'stock_physique_{ligne.pk}')
            if valeur is not None and valeur != '':
                try:
                    ligne.stock_physique = int(valeur)
                    ligne.save()  # recalcule écart et valeurs (cf. LigneInventaire.save)
                except (ValueError, TypeError):
                    continue
        # Recalcul de la synthèse
        agg = lignes.aggregate(valeur=Sum('valeur_physique'))
        inventaire.valeur_totale = agg['valeur'] or 0
        inventaire.ecarts_detectes = lignes.exclude(ecart=0).count()
        inventaire.statut = 'TERMINE'
        inventaire.save(update_fields=['valeur_totale', 'ecarts_detectes', 'statut'])
        messages.success(request, 'Saisie enregistrée. Inventaire prêt à être validé.')
        return redirect('depenses:detail_inventaire', inventaire_id=inventaire.pk)

    context = {
        'titre_page': f"Inventaire {inventaire.numero_inventaire}",
        'inventaire': inventaire,
        'lignes': lignes,
    }
    return render(request, 'depenses/logistique/detail_inventaire.html', context)


@login_required
def valider_inventaire(request, inventaire_id):
    """Valider l'inventaire et ajuster le stock réel via des mouvements"""
    inventaire = get_object_or_404(Inventaire, pk=inventaire_id)

    if request.method == 'POST':
        if inventaire.statut in ('VALIDE', 'ANNULE'):
            messages.warning(request, 'Cet inventaire est déjà clôturé.')
            return redirect('depenses:detail_inventaire', inventaire_id=inventaire.pk)

        lignes_ecart = inventaire.lignes.exclude(ecart=0).select_related('article')
        for ligne in lignes_ecart:
            # Ajuster le stock de l'article à la valeur physique constatée
            mvt = MouvementStock(
                article=ligne.article,
                type_mouvement='AJUSTEMENT',
                motif='INVENTAIRE',
                quantite=ligne.stock_physique,
                prix_unitaire=ligne.prix_unitaire,
                destinataire='',
                document_reference=inventaire.numero_inventaire,
                observations=f"Ajustement inventaire (écart {ligne.ecart})",
                cree_par=request.user,
            )
            mvt.numero_mouvement = _generer_code(MouvementStock, 'numero_mouvement', 'MVT')
            mvt.save()  # met à jour le stock de l'article (type AJUSTEMENT)

        inventaire.statut = 'VALIDE'
        inventaire.valide_par = request.user
        inventaire.date_validation = timezone.now()
        inventaire.save(update_fields=['statut', 'valide_par', 'date_validation'])
        messages.success(request, f'Inventaire {inventaire.numero_inventaire} validé. Stock ajusté.')

    return redirect('depenses:detail_inventaire', inventaire_id=inventaire.pk)


@login_required
def export_stock_excel(request):
    """Exporter le stock en Excel"""
    
    # Créer le workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Articles"
    
    # En-têtes
    headers = [
        'Code Article', 'Nom', 'Catégorie', 'Stock Actuel', 
        'Stock Minimum', 'Prix Unitaire', 'Valeur Stock', 'État', 'Emplacement'
    ]
    
    # Style des en-têtes
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Données
    articles = Article.objects.select_related('categorie').filter(actif=True)
    
    for row, article in enumerate(articles, 2):
        ws.cell(row=row, column=1, value=article.code_article)
        ws.cell(row=row, column=2, value=article.nom)
        ws.cell(row=row, column=3, value=article.categorie.nom)
        ws.cell(row=row, column=4, value=article.stock_actuel)
        ws.cell(row=row, column=5, value=article.stock_minimum)
        ws.cell(row=row, column=6, value=float(article.prix_unitaire))
        ws.cell(row=row, column=7, value=float(article.valeur_stock))
        ws.cell(row=row, column=8, value=article.get_etat_display())
        ws.cell(row=row, column=9, value=article.emplacement)
    
    # Ajuster la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2
    
    # Réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=stock_articles_{date.today()}.xlsx'
    
    wb.save(response)
    return response
