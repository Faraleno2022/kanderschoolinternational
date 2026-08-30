"""Listes et exports des élèves inscrits ou réinscrits."""

from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from eleves.utils_annee import get_annee_active
from paiements.models import EcheancierPaiement
from utilisateurs.utils import filter_by_user_school, user_school
from ecole_moderne.pdf_utils import draw_logo_watermark


VALID_NATURES = {
    EcheancierPaiement.NATURE_INSCRIPTION: ('inscrits', 'Inscription'),
    EcheancierPaiement.NATURE_REINSCRIPTION: ('réinscrits', 'Réinscription'),
}


def _admission_queryset(request, nature):
    if nature not in VALID_NATURES:
        nature = EcheancierPaiement.NATURE_INSCRIPTION
    queryset = (
        EcheancierPaiement.objects
        .select_related('eleve', 'ecole_reference', 'classe_reference')
        .filter(nature_frais=nature, eleve__est_dans_corbeille=False)
    )
    queryset = filter_by_user_school(queryset, request.user, 'ecole_reference')

    selected_year = (request.GET.get('annee_scolaire') or '').strip()
    school = user_school(request.user)
    if not selected_year and school:
        selected_year = get_annee_active(request, school) or ''
    if selected_year:
        queryset = queryset.filter(annee_scolaire=selected_year)

    selected_class = (request.GET.get('classe') or '').strip()
    if selected_class.isdigit():
        queryset = queryset.filter(classe_reference_id=int(selected_class))
    else:
        selected_class = ''

    query = (request.GET.get('q') or '').strip()
    if query:
        queryset = queryset.filter(
            Q(eleve__matricule__icontains=query)
            | Q(eleve__nom__icontains=query)
            | Q(eleve__prenom__icontains=query)
            | Q(classe_reference__nom__icontains=query)
        )

    return queryset.order_by('-date_creation', '-pk'), {
        'q': query,
        'selected_year': selected_year,
        'selected_class': selected_class,
        'school': school,
    }


def _filter_options(request):
    base = filter_by_user_school(
        EcheancierPaiement.objects.all(), request.user, 'ecole_reference',
    )
    years = base.values_list('annee_scolaire', flat=True).distinct().order_by('-annee_scolaire')
    classes = (
        base.exclude(classe_reference=None)
        .values('classe_reference_id', 'classe_reference__nom')
        .distinct().order_by('classe_reference__nom')
    )
    return years, classes


@login_required
def liste_eleves_par_nature(request, nature):
    queryset, filters = _admission_queryset(request, nature)
    paginator = Paginator(queryset, 30)
    page_obj = paginator.get_page(request.GET.get('page'))
    slug, label = VALID_NATURES[nature]
    years, classes = _filter_options(request)
    return render(request, 'paiements/liste_eleves_admission.html', {
        'page_obj': page_obj,
        'nature': nature,
        'nature_slug': slug,
        'nature_label': label,
        'titre_page': f'Élèves {slug}',
        'annees': years,
        'classes': classes,
        **filters,
    })


@login_required
def liste_eleves_inscrits(request):
    return liste_eleves_par_nature(request, EcheancierPaiement.NATURE_INSCRIPTION)


@login_required
def liste_eleves_reinscrits(request):
    return liste_eleves_par_nature(request, EcheancierPaiement.NATURE_REINSCRIPTION)


def _filename(nature, extension):
    slug = 'inscrits' if nature == EcheancierPaiement.NATURE_INSCRIPTION else 'reinscrits'
    return f'eleves_{slug}_{timezone.localdate():%Y-%m-%d}.{extension}'


def _export_excel(request, nature):
    queryset, filters = _admission_queryset(request, nature)
    _slug, label = VALID_NATURES[nature]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = label
    sheet.merge_cells('A1:I1')
    sheet['A1'] = f'LISTE DES ÉLÈVES - {label.upper()}'
    sheet['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    sheet['A1'].fill = PatternFill('solid', fgColor='1F4E78')
    sheet['A1'].alignment = Alignment(horizontal='center')
    sheet.append(['Matricule', 'Nom', 'Prénom', 'Sexe', 'Classe', 'École', 'Année', 'Frais dû', 'Frais payé'])
    for cell in sheet[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='5B9BD5')
        cell.alignment = Alignment(horizontal='center')
    for schedule in queryset:
        sheet.append([
            schedule.eleve.matricule,
            schedule.eleve.nom,
            schedule.eleve.prenom,
            schedule.eleve.get_sexe_display(),
            schedule.classe_reference.nom if schedule.classe_reference else '-',
            schedule.ecole_reference.nom if schedule.ecole_reference else '-',
            schedule.annee_scolaire,
            int(schedule.frais_inscription_du or 0),
            int(schedule.frais_inscription_paye or 0),
        ])
    for column in ('H', 'I'):
        for cell in sheet[column][2:]:
            cell.number_format = '#,##0 "GNF"'
    widths = [18, 22, 22, 12, 25, 35, 14, 18, 18]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = 'A3'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{_filename(nature, "xlsx")}"'
    workbook.save(response)
    return response


def _export_pdf(request, nature):
    queryset, filters = _admission_queryset(request, nature)
    _slug, label = VALID_NATURES[nature]
    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer, pagesize=landscape(A4), rightMargin=1.1 * cm,
        leftMargin=1.1 * cm, topMargin=1.1 * cm, bottomMargin=1.2 * cm,
        title=f'Élèves - {label}',
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f'<b>LISTE DES ÉLÈVES — {label.upper()}</b>', styles['Title']),
        Paragraph(
            f"Année scolaire : {filters['selected_year'] or 'Toutes'} &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Édité le {timezone.localtime():%d/%m/%Y à %H:%M}",
            styles['Normal'],
        ),
        Spacer(1, 0.35 * cm),
    ]
    data = [['Matricule', 'Élève', 'Sexe', 'Classe', 'École', 'Année', 'Frais dû', 'Frais payé']]
    for schedule in queryset:
        data.append([
            schedule.eleve.matricule,
            schedule.eleve.nom_complet,
            schedule.eleve.get_sexe_display(),
            schedule.classe_reference.nom if schedule.classe_reference else '-',
            schedule.ecole_reference.nom if schedule.ecole_reference else '-',
            schedule.annee_scolaire,
            f"{int(schedule.frais_inscription_du or 0):,}".replace(',', ' '),
            f"{int(schedule.frais_inscription_paye or 0):,}".replace(',', ' '),
        ])
    table = Table(data, repeatRows=1, colWidths=[2.5*cm, 4.2*cm, 1.8*cm, 3.5*cm, 5*cm, 2.3*cm, 2.6*cm, 2.6*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), .35, colors.HexColor('#B8C4CE')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F6F9')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (-2, 1), (-1, -1), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(table)
    school = filters['school']

    def decorate_page(canvas, doc):
        try:
            draw_logo_watermark(canvas, doc.pagesize[0], doc.pagesize[1], ecole=school)
        except Exception:
            pass
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.drawRightString(doc.pagesize[0] - 1.1*cm, .55*cm, f'Page {doc.page}')
        canvas.restoreState()

    document.build(story, onFirstPage=decorate_page, onLaterPages=decorate_page)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{_filename(nature, "pdf")}"'
    return response


@login_required
def export_eleves_inscrits_excel(request):
    return _export_excel(request, EcheancierPaiement.NATURE_INSCRIPTION)


@login_required
def export_eleves_inscrits_pdf(request):
    return _export_pdf(request, EcheancierPaiement.NATURE_INSCRIPTION)


@login_required
def export_eleves_reinscrits_excel(request):
    return _export_excel(request, EcheancierPaiement.NATURE_REINSCRIPTION)


@login_required
def export_eleves_reinscrits_pdf(request):
    return _export_pdf(request, EcheancierPaiement.NATURE_REINSCRIPTION)
