from django.conf import settings
from django.test import TestCase, override_settings
from django.urls import reverse
from django.contrib.auth import get_user_model
from datetime import date

from eleves.models import Ecole, Classe, Eleve, Responsable
from paiements.models import Paiement, TypePaiement, ModePaiement, EcheancierPaiement, Relance
from utilisateurs.models import Profil


TEST_MIDDLEWARE = [
    middleware for middleware in settings.MIDDLEWARE
    if middleware != 'ecole_moderne.licence_middleware.LicenceMiddleware'
]


@override_settings(MIDDLEWARE=TEST_MIDDLEWARE)
class SchoolFilteringTests(TestCase):
    def setUp(self):
        # Schools (provide required fields)
        self.ecole1 = Ecole.objects.create(
            nom="Ecole A",
            adresse="Adresse A",
            telephone="+224620000001",
            directeur="Dir A",
        )
        self.ecole2 = Ecole.objects.create(
            nom="Ecole B",
            adresse="Adresse B",
            telephone="+224620000002",
            directeur="Dir B",
        )
        # Classes (provide required niveau + annee_scolaire)
        self.classe1 = Classe.objects.create(nom="C1", ecole=self.ecole1, niveau="PRIMAIRE_1", annee_scolaire="2024-2025")
        self.classe2 = Classe.objects.create(nom="C2", ecole=self.ecole2, niveau="PRIMAIRE_1", annee_scolaire="2024-2025")

        # Responsables
        self.resp1 = Responsable.objects.create(prenom="P1", nom="R1", relation="PERE", telephone="+224620000011", adresse="Adr1")
        self.resp2 = Responsable.objects.create(prenom="P2", nom="R2", relation="PERE", telephone="+224620000012", adresse="Adr2")

        # Students (provide all required fields)
        self.eleve1 = Eleve.objects.create(
            nom="Alpha",
            prenom="A",
            matricule="A-001",
            classe=self.classe1,
            sexe='M',
            date_naissance=date(2015, 1, 1),
            lieu_naissance="Conakry",
            date_inscription=date(2024, 9, 1),
            responsable_principal=self.resp1,
        )
        self.eleve2 = Eleve.objects.create(
            nom="Bravo",
            prenom="B",
            matricule="B-001",
            classe=self.classe2,
            sexe='F',
            date_naissance=date(2015, 2, 2),
            lieu_naissance="Conakry",
            date_inscription=date(2024, 9, 1),
            responsable_principal=self.resp2,
        )
        # Payment metadata
        self.type_insc = TypePaiement.objects.create(nom="Frais d'inscription")
        self.mode_espece = ModePaiement.objects.create(nom="Espèces")
        # Payments
        self.paiement1 = Paiement.objects.create(
            eleve=self.eleve1,
            type_paiement=self.type_insc,
            mode_paiement=self.mode_espece,
            montant=30000,
            statut='VALIDE',
            date_paiement=date(2024, 9, 10),
        )
        self.paiement2 = Paiement.objects.create(
            eleve=self.eleve2,
            type_paiement=self.type_insc,
            mode_paiement=self.mode_espece,
            montant=30000,
            statut='VALIDE',
            date_paiement=date(2024, 9, 11),
        )
        self.echeancier1 = EcheancierPaiement.objects.create(
            eleve=self.eleve1, annee_scolaire="2024-2025",
            frais_inscription_du=50000, frais_inscription_paye=30000,
            date_echeance_inscription=date(2024, 9, 1),
            date_echeance_tranche_1=date(2025, 1, 15),
            date_echeance_tranche_2=date(2025, 3, 15),
            date_echeance_tranche_3=date(2025, 5, 15),
        )
        self.echeancier2 = EcheancierPaiement.objects.create(
            eleve=self.eleve2, annee_scolaire="2024-2025",
            frais_inscription_du=60000, frais_inscription_paye=30000,
            date_echeance_inscription=date(2024, 9, 1),
            date_echeance_tranche_1=date(2025, 1, 15),
            date_echeance_tranche_2=date(2025, 3, 15),
            date_echeance_tranche_3=date(2025, 5, 15),
        )
        Relance.objects.create(
            eleve=self.eleve1, canal="SMS", statut="ENVOYEE",
            message="Rappel école A", solde_estime=20000,
        )
        Relance.objects.create(
            eleve=self.eleve2, canal="SMS", statut="ENVOYEE",
            message="Rappel école B", solde_estime=30000,
        )
        # Users
        User = get_user_model()
        self.user1 = User.objects.create_user(username="u1", password="pass12345")
        self.user2 = User.objects.create_user(username="u2", password="pass12345")
        Profil.objects.update_or_create(
            user=self.user1,
            defaults={
                'role': 'COMPTABLE', 'ecole': self.ecole1,
                'telephone': "+224620000021", 'peut_consulter_rapports': True,
                'is_validated': True,
            },
        )
        Profil.objects.update_or_create(
            user=self.user2,
            defaults={
                'role': 'COMPTABLE', 'ecole': self.ecole2,
                'telephone': "+224620000022", 'peut_consulter_rapports': True,
                'is_validated': True,
            },
        )
        self.user1.refresh_from_db()
        self.user2.refresh_from_db()

    def login1(self):
        self.client.logout()
        self.client.force_login(self.user1)

    def login2(self):
        self.client.logout()
        self.client.force_login(self.user2)

    def test_api_paiements_list_filtered_by_school(self):
        self.login1()
        url = reverse("paiements:api_paiements_list")
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        ids = [r["id"] for r in data.get("results", [])]
        self.assertIn(self.paiement1.id, ids)
        self.assertNotIn(self.paiement2.id, ids)

    def test_rapport_comptable_filtre_par_ecole(self):
        self.login1()
        response = self.client.get(reverse("paiements:rapport_comptable"), {
            "date_debut": "2024-01-01", "date_fin": "2026-12-31",
        })
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.eleve1.nom_complet)
        self.assertNotContains(response, self.eleve2.nom_complet)

    def test_exports_rapport_comptable(self):
        self.login1()
        params = {"date_debut": "2024-01-01", "date_fin": "2026-12-31"}
        excel = self.client.get(reverse("paiements:export_rapport_comptable_excel"), params)
        self.assertEqual(excel.status_code, 200)
        self.assertIn("spreadsheetml", excel["Content-Type"])
        self.assertTrue(excel.content.startswith(b"PK"))

        pdf = self.client.get(reverse("paiements:export_rapport_comptable_pdf"), params)
        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(pdf["Content-Type"], "application/pdf")
        self.assertTrue(pdf.content.startswith(b"%PDF"))

    def test_export_recap_par_classe_excel_filtre_par_ecole(self):
        from io import BytesIO
        from openpyxl import load_workbook

        self.login1()
        response = self.client.get(reverse("paiements:export_recap_par_classe_excel"))
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.content.startswith(b"PK"))

        workbook = load_workbook(BytesIO(response.content), read_only=True)
        lignes = list(workbook.active.iter_rows(values_only=True))
        self.assertTrue(any(self.ecole1.nom in ligne for ligne in lignes[1:]))
        self.assertFalse(any(self.ecole2.nom in ligne for ligne in lignes[1:]))

    def test_api_paiement_detail_for_other_school_is_404(self):
        self.login1()
        url = reverse("paiements:api_paiement_detail", kwargs={"pk": self.paiement2.id})
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 404)

    def test_ajax_eleve_info_other_school_is_404(self):
        self.login1()
        url = reverse("paiements:ajax_eleve_info")
        resp = self.client.get(url, {"matricule": self.eleve2.matricule})
        self.assertEqual(resp.status_code, 404)

    def test_ajax_eleve_info_own_school_ok(self):
        self.login1()
        url = reverse("paiements:ajax_eleve_info")
        resp = self.client.get(url, {"matricule": self.eleve1.matricule})
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertTrue(data.get("success"))
        self.assertEqual(data.get("eleve", {}).get("id"), self.eleve1.id)

    def test_annuler_remise_paiement_other_school_is_404(self):
        """Un comptable ne peut pas annuler les remises d'un paiement d'une autre école."""
        self.login1()  # user1 -> ecole1
        url = reverse("paiements:annuler_remise_paiement", kwargs={"paiement_id": self.paiement2.id})
        resp = self.client.post(url)
        self.assertEqual(resp.status_code, 404)

    def test_annuler_remise_paiement_unique_other_school_is_404(self):
        """Même avec un remise_id arbitraire, l'accès à un paiement d'une autre école doit renvoyer 404."""
        self.login1()  # user1 -> ecole1
        # Pas besoin de créer une remise réelle: la vue vérifie d'abord l'accès au paiement
        url = reverse("paiements:annuler_remise_paiement_unique", kwargs={"paiement_id": self.paiement2.id, "remise_id": 999})
        resp = self.client.post(url)
        self.assertEqual(resp.status_code, 404)

    # --- Nouvelles vérifications des vues HTML protégées par require_school_object ---
    def test_detail_paiement_other_school_is_404(self):
        self.login1()
        url = reverse("paiements:detail_paiement", kwargs={"paiement_id": self.paiement2.id})
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 404)

    def test_generer_recu_pdf_other_school_is_404(self):
        self.login1()
        url = reverse("paiements:generer_recu_pdf", kwargs={"paiement_id": self.paiement2.id})
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 404)

    def test_echeancier_eleve_other_school_is_404(self):
        self.login1()
        url = reverse("paiements:echeancier_eleve", kwargs={"eleve_id": self.eleve2.id})
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 404)

    def test_relancer_eleve_other_school_is_404(self):
        self.login1()
        url = reverse("paiements:relancer_eleve", kwargs={"eleve_id": self.eleve2.id})
        # GET simple, on ne vérifie que la protection d'accès (pas les side-effects)
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 404)
