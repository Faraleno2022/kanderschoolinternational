from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import RequestFactory, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from eleves.models import Classe, Ecole, Eleve, Responsable
from paiements.models import (
    EcheancierPaiement,
    ModePaiement,
    Paiement,
    PaiementRemise,
    Relance,
    RemiseReduction,
    TypePaiement,
)
from paiements.rapports_professionnels import (
    collect_accounting_data,
    collect_recovery_data,
)
from paiements.tests.support import TEST_MIDDLEWARE


@override_settings(MIDDLEWARE=TEST_MIDDLEWARE)
class ProfessionalReportsTests(TestCase):
    cutoff = date(2026, 2, 15)

    def setUp(self):
        self.user = get_user_model().objects.create_superuser(
            username='admin-rapports-professionnels',
            email='rapports@example.com',
            password='mot-de-passe-test',
            first_name='Aïssatou',
            last_name='Camara',
        )
        self.client.force_login(self.user)
        self.factory = RequestFactory()
        self.ecole = Ecole.objects.create(
            nom='Groupe scolaire Démonstration & Associés',
            adresse='Conakry',
            telephone='+224620000401',
            directeur='Direction générale',
        )
        self.classe = Classe.objects.create(
            ecole=self.ecole,
            nom='8ème A',
            niveau='COLLEGE_8',
            annee_scolaire='2025-2026',
        )
        self.responsable = Responsable.objects.create(
            prenom='Mamadou',
            nom='Diallo',
            relation='PERE',
            telephone='+224620000402',
            adresse='Conakry',
        )
        self.students = [
            self._student('RAP-001', 'Aminata'),
            self._student('RAP-002', 'Boubacar'),
            self._student('RAP-003', 'Fatoumata'),
        ]
        self._schedule(self.students[0], paid=Decimal('80000'))
        self._schedule(self.students[1], paid=Decimal('0'))
        self._schedule(self.students[2], paid=Decimal('320000'))

        self.payment_type = TypePaiement.objects.create(nom='Scolarité annuelle rapport')
        self.payment_mode = ModePaiement.objects.create(nom='Mobile Money rapport')
        self.valid_partial = self._payment(
            self.students[0], 'RAP-REC-001', '80000', 'VALIDE', date(2026, 1, 15),
        )
        self._payment(
            self.students[2], 'RAP-REC-002', '320000', 'VALIDE', date(2026, 1, 16),
        )
        self._payment(
            self.students[1], 'RAP-REC-003', '20000', 'EN_ATTENTE', date(2026, 1, 20),
        )
        self._payment(
            self.students[1], 'RAP-REC-004', '30000', 'REMBOURSE', date(2026, 1, 22),
        )

        discount = RemiseReduction.objects.create(
            nom='Remise sociale rapport',
            type_remise='MONTANT_FIXE',
            valeur=Decimal('10000'),
            motif='SOCIALE',
            date_debut=date(2025, 9, 1),
            date_fin=date(2026, 6, 30),
        )
        PaiementRemise.objects.create(
            paiement=self.valid_partial,
            remise=discount,
            montant_remise=Decimal('10000'),
        )

        old_reminder = Relance.objects.create(
            eleve=self.students[1],
            canal='WHATSAPP',
            message='Merci de régulariser la situation.',
            statut='ENVOYEE',
            solde_estime=Decimal('320000'),
            cree_par=self.user,
        )
        Relance.objects.filter(pk=old_reminder.pk).update(
            date_creation=timezone.make_aware(datetime(2026, 1, 20, 10, 0)),
        )
        Relance.objects.create(
            eleve=self.students[1],
            canal='SMS',
            message='Relance postérieure à la date d’arrêt.',
            statut='ECHEC',
            solde_estime=Decimal('320000'),
            cree_par=self.user,
        )

    def _student(self, matricule, prenom):
        return Eleve.objects.create(
            matricule=matricule,
            prenom=prenom,
            nom='Rapport',
            sexe='F',
            date_naissance=date(2013, 1, 1),
            classe=self.classe,
            date_inscription=date(2025, 9, 1),
            responsable_principal=self.responsable,
        )

    def _schedule(self, student, paid):
        return EcheancierPaiement.objects.create(
            eleve=student,
            annee_scolaire='2025-2026',
            frais_inscription_du=Decimal('20000'),
            frais_inscription_paye=min(paid, Decimal('20000')),
            tranche_1_due=Decimal('100000'),
            tranche_1_payee=min(max(paid - Decimal('20000'), 0), Decimal('100000')),
            tranche_2_due=Decimal('100000'),
            tranche_2_payee=min(max(paid - Decimal('120000'), 0), Decimal('100000')),
            tranche_3_due=Decimal('100000'),
            tranche_3_payee=min(max(paid - Decimal('220000'), 0), Decimal('100000')),
            date_echeance_inscription=date(2025, 9, 30),
            date_echeance_tranche_1=date(2026, 1, 10),
            date_echeance_tranche_2=date(2026, 3, 5),
            date_echeance_tranche_3=date(2026, 5, 5),
        )

    def _payment(self, student, receipt, amount, status, payment_date):
        return Paiement.objects.create(
            eleve=student,
            type_paiement=self.payment_type,
            mode_paiement=self.payment_mode,
            numero_recu=receipt,
            montant=Decimal(amount),
            date_paiement=payment_date,
            statut=status,
            reference_externe=f'TXN-{receipt}',
            cree_par=self.user,
            valide_par=self.user if status == 'VALIDE' else None,
        )

    def _request(self, **params):
        defaults = {
            'classe_id': str(self.classe.pk),
            'au': self.cutoff.isoformat(),
        }
        defaults.update(params)
        request = self.factory.get('/paiements/liste/', defaults)
        request.user = self.user
        return request

    def test_synthese_comptable_regroupe_statuts_remises_et_rapprochements(self):
        data = collect_accounting_data(self._request())

        self.assertEqual(data['payment_count'], 4)
        self.assertEqual(data['validated_count'], 2)
        self.assertEqual(data['total_validated'], Decimal('400000'))
        self.assertEqual(data['total_discounts'], Decimal('10000'))
        self.assertEqual(data['total_coverage'], Decimal('410000'))
        self.assertEqual(data['by_status']['EN_ATTENTE']['count'], 1)
        self.assertEqual(data['by_status']['REMBOURSE']['amount'], Decimal('30000'))
        self.assertEqual(data['by_mode']['Mobile Money rapport']['amount'], Decimal('400000'))
        self.assertEqual(
            data['by_mode']['Mobile Money rapport']['reference_missing'], 0,
        )
        self.assertEqual(data['discount_by_reason']['Réduction sociale'], Decimal('10000'))
        self.assertEqual(data['by_component']['inscription']['amount'], Decimal('40000'))
        self.assertEqual(data['by_component']['tranche_1']['amount'], Decimal('160000'))
        self.assertEqual(data['by_component']['tranche_2']['amount'], Decimal('100000'))
        self.assertEqual(data['by_component']['tranche_3']['amount'], Decimal('100000'))
        self.assertEqual(data['unallocated_total'], Decimal('0'))

    def test_periode_future_est_automatiquement_arretee_aujourdhui(self):
        future = timezone.localdate().replace(year=timezone.localdate().year + 1)

        data = collect_accounting_data(self._request(au=future.isoformat()))

        self.assertTrue(data['period_adjusted'])
        self.assertEqual(data['end'], timezone.localdate())
        self.assertEqual(data['cutoff'], timezone.localdate())
        self.assertIn(timezone.localdate().strftime('%d/%m/%Y'), data['period_label'])

    def test_ventilation_separe_reinscription_et_tranches(self):
        reinscription_type = TypePaiement.objects.create(
            nom='Réinscription + Tranche 1 rapport',
        )
        self._payment(
            self.students[1], 'RAP-REC-005', '80000', 'VALIDE', date(2026, 1, 25),
        )
        payment = Paiement.objects.get(numero_recu='RAP-REC-005')
        payment.type_paiement = reinscription_type
        payment.save(update_fields=['type_paiement'])

        data = collect_accounting_data(self._request())

        row = next(item for item in data['payment_rows'] if item['receipt'] == 'RAP-REC-005')
        self.assertEqual(row['allocation']['reinscription'], Decimal('20000'))
        self.assertEqual(row['allocation']['inscription'], Decimal('0'))
        self.assertEqual(row['allocation']['tranche_1'], Decimal('60000'))

    def test_recouvrement_est_reconstruit_a_la_date_arret(self):
        data = collect_recovery_data(self._request())

        self.assertEqual(data['schedule_count'], 3)
        self.assertEqual(data['total_due'], Decimal('960000'))
        self.assertEqual(data['total_cash'], Decimal('400000'))
        self.assertEqual(data['total_discount'], Decimal('10000'))
        self.assertEqual(data['total_balance'], Decimal('550000'))
        self.assertEqual(data['total_overdue'], Decimal('150000'))
        self.assertEqual(data['total_upcoming'], Decimal('200000'))
        self.assertEqual(data['settled_count'], 1)
        self.assertEqual(data['partial_count'], 1)
        self.assertEqual(data['unpaid_count'], 1)
        self.assertEqual(data['overdue_count'], 2)
        self.assertEqual(data['aging']['31-60 jours']['count'], 2)
        self.assertEqual(data['aging']['31-60 jours']['amount'], Decimal('130000'))
        self.assertEqual(data['aging']['Plus de 90 jours']['count'], 1)
        self.assertEqual(data['aging']['Plus de 90 jours']['amount'], Decimal('20000'))
        self.assertEqual(len(data['period_relances']), 1)
        self.assertEqual(data['priority_rows'][0]['reminder_count'], 1)
        self.assertEqual(data['reminder_by_channel']['WhatsApp']['sent'], 1)

    def test_remise_ciblee_solde_eleve_et_est_precisee_dans_les_exports(self):
        ligne_remise = PaiementRemise.objects.get(paiement=self.valid_partial)
        ligne_remise.portee_tranches = '1'
        ligne_remise.save(update_fields=['portee_tranches'])
        self._payment(
            self.students[0], 'RAP-REC-006', '230000', 'VALIDE',
            date(2026, 1, 25),
        )

        data = collect_recovery_data(self._request())
        eleve = next(
            item for item in data['student_rows']
            if item['matricule'] == 'RAP-001'
        )
        self.assertEqual(eleve['cash'], Decimal('310000'))
        self.assertEqual(eleve['discount'], Decimal('10000'))
        self.assertEqual(eleve['coverage'], Decimal('320000'))
        self.assertEqual(eleve['balance'], Decimal('0'))
        self.assertEqual(eleve['status'], 'Soldé avec remise')
        self.assertAlmostEqual(float(eleve['discount_rate']), 10 / 3, places=2)
        self.assertIn("L'élève est soldé grâce au paiement et à la remise", eleve['settlement_note'])

        params = {'classe_id': self.classe.pk, 'au': self.cutoff.isoformat()}
        excel_response = self.client.get(
            reverse('paiements:export_recouvrement_excel'), params,
        )
        workbook = load_workbook(BytesIO(excel_response.content), data_only=True)
        sheet = workbook['Portefeuille élèves']
        row = next(
            values for values in sheet.iter_rows(min_row=6, values_only=True)
            if values[0] == 'RAP-001'
        )
        self.assertEqual(row[7], 10000)
        self.assertAlmostEqual(row[8], 1 / 30)
        self.assertEqual(row[9], 320000)
        self.assertEqual(row[10], 0)
        self.assertEqual(row[13], 'Soldé avec remise')
        self.assertIn('Remise appliquée', row[14])

    def test_exports_pdf_et_excel_sont_disponibles(self):
        params = {'classe_id': self.classe.pk, 'au': self.cutoff.isoformat()}
        for route in ('export_comptabilite_pdf', 'export_recouvrement_pdf'):
            with self.subTest(route=route):
                response = self.client.get(reverse(f'paiements:{route}'), params)
                self.assertEqual(response.status_code, 200)
                self.assertEqual(response['Content-Type'], 'application/pdf')
                self.assertTrue(response.content.startswith(b'%PDF'))
                self.assertGreater(len(response.content), 5000)

        accounting_excel = self.client.get(
            reverse('paiements:export_comptabilite_excel'), params,
        )
        self.assertEqual(accounting_excel.status_code, 200)
        accounting_workbook = load_workbook(
            BytesIO(accounting_excel.content), data_only=True,
        )
        self.assertEqual(
            accounting_workbook.sheetnames,
            [
                'Synthèse', 'Journal validé', 'Affectations', 'Statuts',
                'Ventilations', 'Remises',
            ],
        )
        self.assertEqual(
            accounting_workbook['Journal validé'].cell(5, 12).value,
            'Inscription',
        )
        self.assertEqual(
            accounting_workbook['Affectations'].cell(12, 3).value,
            400000,
        )

        recovery_excel = self.client.get(
            reverse('paiements:export_recouvrement_excel'), params,
        )
        self.assertEqual(recovery_excel.status_code, 200)
        workbook = load_workbook(BytesIO(recovery_excel.content), data_only=True)
        self.assertEqual(
            workbook.sheetnames,
            [
                'Synthèse', 'Portefeuille élèves', 'Classes', 'Balance âgée',
                'Priorités', 'Relances', 'Journal relances',
            ],
        )
        self.assertEqual(workbook['Portefeuille élèves'].cell(5, 1).value, 'Matricule')
        self.assertEqual(workbook['Portefeuille élèves'].max_row, 8)
        self.assertEqual(workbook['Journal relances'].max_row, 6)

    def test_exports_modes_encaissement_regroupent_les_paiements_valides(self):
        mode_especes = ModePaiement.objects.create(nom='Espèces rapport')
        paiement_especes = self._payment(
            self.students[1], 'RAP-REC-ESPECES', '50000', 'VALIDE',
            date(2026, 1, 24),
        )
        paiement_especes.mode_paiement = mode_especes
        paiement_especes.reference_externe = ''
        paiement_especes.save(update_fields=['mode_paiement', 'reference_externe'])
        paiement_attente = self._payment(
            self.students[1], 'RAP-REC-ATTENTE', '90000', 'EN_ATTENTE',
            date(2026, 1, 25),
        )
        paiement_attente.mode_paiement = mode_especes
        paiement_attente.save(update_fields=['mode_paiement'])

        params = {'classe_id': self.classe.pk, 'au': self.cutoff.isoformat()}
        response = self.client.get(
            reverse('paiements:export_modes_encaissement_excel'), params,
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.content.startswith(b'PK'))
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        self.assertEqual(
            workbook.sheetnames,
            ['Synthèse par mode', 'Détail des encaissements'],
        )
        summary_rows = list(
            workbook['Synthèse par mode'].iter_rows(min_row=6, values_only=True)
        )
        summary = {row[0]: row for row in summary_rows}
        self.assertEqual(summary['Espèces rapport'][1], 1)
        self.assertEqual(summary['Espèces rapport'][2], 50000)
        self.assertEqual(summary['Mobile Money rapport'][1], 2)
        self.assertEqual(summary['Mobile Money rapport'][2], 400000)
        self.assertAlmostEqual(summary['Mobile Money rapport'][3], 400000 / 450000)
        self.assertEqual(summary['TOTAL'][1], 3)
        self.assertEqual(summary['TOTAL'][2], 450000)

        receipts = {
            row[2]
            for row in workbook['Détail des encaissements'].iter_rows(
                min_row=6, values_only=True,
            )
        }
        self.assertIn('RAP-REC-ESPECES', receipts)
        self.assertNotIn('RAP-REC-ATTENTE', receipts)

        pdf_response = self.client.get(
            reverse('paiements:export_modes_encaissement_pdf'), params,
        )
        self.assertEqual(pdf_response.status_code, 200)
        self.assertEqual(pdf_response['Content-Type'], 'application/pdf')
        self.assertTrue(pdf_response.content.startswith(b'%PDF'))
        self.assertGreater(len(pdf_response.content), 5000)

    def test_filtres_invalides_retournent_une_erreur_400(self):
        response = self.client.get(
            reverse('paiements:export_recouvrement_pdf'),
            {'classe_id': self.classe.pk, 'du': '2026-03-01', 'au': '2026-02-01'},
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn('date de début', response.content.decode('utf-8'))

    def test_les_rapports_sont_proteges_par_la_permission(self):
        simple_user = get_user_model().objects.create_user(
            username='sans-permission-rapport',
            password='mot-de-passe-test',
        )
        simple_user.profil.peut_consulter_rapports = False
        simple_user.profil.save(update_fields=['peut_consulter_rapports'])
        self.client.force_login(simple_user)

        for route in (
            'export_recouvrement_pdf',
            'export_modes_encaissement_pdf',
            'export_modes_encaissement_excel',
        ):
            with self.subTest(route=route):
                response = self.client.get(reverse(f'paiements:{route}'))
                self.assertEqual(response.status_code, 403)
