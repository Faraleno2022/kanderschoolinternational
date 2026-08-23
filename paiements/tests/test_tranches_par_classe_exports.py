from datetime import date
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook

from eleves.models import Classe, Ecole, Eleve, GrilleTarifaire
from paiements.models import (
    EcheancierPaiement, ModePaiement, Paiement, PaiementRemise,
    RemiseReduction, TypePaiement,
)

from .support import TEST_MIDDLEWARE


@override_settings(MIDDLEWARE=TEST_MIDDLEWARE)
class TranchesParClasseExportsTests(TestCase):
    def setUp(self):
        self.ecole = Ecole.objects.create(
            nom='Ecole export tranches', adresse='Conakry',
            telephone='+224620000701', directeur='Direction',
        )
        self.classe = Classe.objects.create(
            ecole=self.ecole, nom='Classe export', niveau='PRIMAIRE_1',
            annee_scolaire='2025-2026',
        )
        GrilleTarifaire.objects.create(
            ecole=self.ecole, niveau='PRIMAIRE_1',
            annee_scolaire='2025-2026',
            frais_inscription=Decimal('30000'),
            frais_reinscription=Decimal('20000'),
            tranche_1=Decimal('100000'),
            tranche_2=Decimal('110000'),
            tranche_3=Decimal('120000'),
        )
        self.mode = ModePaiement.objects.create(nom='Especes export')
        self.type_inscription = TypePaiement.objects.create(nom='Inscription')
        self.type_reinscription = TypePaiement.objects.create(
            nom='Réinscription + Tranche 1'
        )
        self._creer_eleve(
            'EXP-INS', Decimal('30000'), Decimal('30000'),
            Decimal('50000'), self.type_inscription,
        )
        self._creer_eleve(
            'EXP-REI', Decimal('20000'), Decimal('20000'),
            Decimal('60000'), self.type_reinscription,
        )
        self.user = get_user_model().objects.create_superuser(
            username='admin_export_tranches',
            email='admin.export.tranches@example.com',
            password='pass12345',
        )
        self.client.force_login(self.user)

    def _creer_eleve(self, matricule, admission_due, admission_paid, t1_paid, type_paiement):
        eleve = Eleve.objects.create(
            matricule=matricule, prenom='Eleve', nom=matricule,
            sexe='F', date_naissance=date(2018, 1, 1),
            lieu_naissance='Conakry', classe=self.classe,
            date_inscription=date(2025, 9, 1),
        )
        EcheancierPaiement.objects.create(
            eleve=eleve, annee_scolaire='2025-2026',
            frais_inscription_du=admission_due,
            tranche_1_due=Decimal('100000'),
            tranche_2_due=Decimal('110000'),
            tranche_3_due=Decimal('120000'),
            frais_inscription_paye=admission_paid,
            tranche_1_payee=t1_paid,
            date_echeance_inscription=date(2025, 9, 1),
            date_echeance_tranche_1=date(2026, 1, 15),
            date_echeance_tranche_2=date(2026, 3, 15),
            date_echeance_tranche_3=date(2026, 5, 15),
        )
        Paiement.objects.create(
            eleve=eleve, type_paiement=type_paiement,
            mode_paiement=self.mode,
            montant=admission_paid + t1_paid,
            date_paiement=date(2025, 9, 10), statut='VALIDE',
        )

    def _params(self):
        return {
            'classe': self.classe.id,
            'annee_scolaire': '2025-2026',
        }

    def test_excel_separe_inscription_et_reinscription(self):
        response = self.client.get(
            reverse('paiements:export_tranches_par_classe_excel'),
            self._params(),
        )
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        sheet = workbook[self.classe.nom]
        self.assertEqual(
            list(next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))),
            [
                'Élève', 'Inscription payée', 'Réinscription payée',
                'Tranche 1 payée', 'Tranche 2 payée', 'Tranche 3 payée',
                'Total dû', 'Total encaissé', 'Remise', 'Remise (%)',
                'Reste', 'Situation', 'Précision remise',
            ],
        )
        lignes = list(sheet.iter_rows(min_row=3, values_only=True))
        inscription = next(row for row in lignes if 'EXP-INS' in row[0])
        self.assertEqual(inscription[1], 30000)
        self.assertEqual(inscription[2], 0)
        self.assertEqual(inscription[3], 50000)

        reinscription = next(row for row in lignes if 'EXP-REI' in row[0])
        self.assertEqual(reinscription[1], 0)
        self.assertEqual(reinscription[2], 20000)
        self.assertEqual(reinscription[3], 60000)
        self.assertEqual(reinscription[7], 80000)

    def test_pdf_contient_la_colonne_reinscription(self):
        from reportlab.platypus import Table as RealTable

        tables = []

        def capturer_table(data, *args, **kwargs):
            tables.append(data)
            return RealTable(data, *args, **kwargs)

        with patch('reportlab.platypus.Table', side_effect=capturer_table):
            response = self.client.get(
                reverse('paiements:export_tranches_par_classe_pdf'),
                self._params(),
            )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.content.startswith(b'%PDF'))
        self.assertTrue(tables)
        self.assertEqual(
            tables[0][0][2].getPlainText(), 'Réinscription payée',
        )
        # La réinscription n'est jamais recopiée dans la colonne inscription.
        ligne_reinscription = next(
            row for row in tables[0][1:]
            if 'EXP-REI' in row[0].getPlainText()
        )
        self.assertEqual(ligne_reinscription[1], '0')
        self.assertEqual(ligne_reinscription[2], '20 000')

    def test_remise_solde_eleve_et_apparait_dans_pdf_et_excel(self):
        eleve = self.classe.eleves.get(matricule='EXP-REI')
        echeancier = eleve.echeancier
        echeancier.tranche_1_payee = Decimal('100000')
        echeancier.tranche_2_payee = Decimal('110000')
        echeancier.save(update_fields=['tranche_1_payee', 'tranche_2_payee'])

        paiement = Paiement.objects.create(
            eleve=eleve, type_paiement=self.type_reinscription,
            mode_paiement=self.mode, montant=Decimal('150000'),
            date_paiement=date(2026, 2, 10), statut='VALIDE',
        )
        remise = RemiseReduction.objects.create(
            nom='Remise T3 export 100%', type_remise='POURCENTAGE',
            valeur=Decimal('100'), motif='SOCIALE',
            date_debut=date(2025, 9, 1), date_fin=date(2026, 8, 31),
        )
        PaiementRemise.objects.create(
            paiement=paiement, remise=remise,
            montant_remise=Decimal('120000'), portee_tranches='3',
            deduite_du_paiement=True,
        )

        excel_response = self.client.get(
            reverse('paiements:export_tranches_par_classe_excel'),
            self._params(),
        )
        workbook = load_workbook(BytesIO(excel_response.content), data_only=True)
        sheet = workbook[self.classe.nom]
        ligne = next(
            row for row in sheet.iter_rows(min_row=3, values_only=True)
            if row[0] and 'EXP-REI' in row[0]
        )
        self.assertEqual(ligne[7], 230000)
        self.assertEqual(ligne[8], 120000)
        # Le taux exporté est exactement celui sélectionné (100 %), pas le
        # rapport artificiel 120 000 / 330 000 = 36,4 %.
        self.assertEqual(ligne[9], 1)
        self.assertEqual(ligne[10], 0)
        self.assertEqual(ligne[11], 'Soldé avec remise')
        self.assertIn('Élève soldé', ligne[12])
        self.assertNotIn('120 000', ligne[12])

        from reportlab.platypus import Table as RealTable

        tables = []

        def capturer_table(data, *args, **kwargs):
            tables.append(data)
            return RealTable(data, *args, **kwargs)

        with patch('reportlab.platypus.Table', side_effect=capturer_table):
            pdf_response = self.client.get(
                reverse('paiements:export_tranches_par_classe_pdf'),
                self._params(),
            )
        self.assertEqual(pdf_response.status_code, 200)
        self.assertEqual(pdf_response['Content-Type'], 'application/pdf')
        self.assertTrue(pdf_response.content.startswith(b'%PDF'))
        self.assertGreater(len(pdf_response.content), 1000)

        ligne_pdf = next(
            row for row in tables[0][1:]
            if 'EXP-REI' in row[0].getPlainText()
        )
        self.assertEqual(ligne_pdf[8], '120 000')
        self.assertEqual(ligne_pdf[9], '100.0 %')
        self.assertEqual(ligne_pdf[10], '0')
        self.assertIn('Remise appliquée', ligne_pdf[11].getPlainText())
        self.assertNotIn('120 000', ligne_pdf[11].getPlainText())
        # Le montant de la remise n'est recopié dans aucune autre colonne.
        self.assertEqual(
            sum('120 000' in getattr(cell, 'getPlainText', lambda: str(cell))()
                for cell in ligne_pdf),
            1,
        )

    def _creer_cas_remise_non_deduite(self):
        eleve = Eleve.objects.create(
            matricule='EXP-REM-ND', prenom='Fara', nom='Leno',
            sexe='M', date_naissance=date(2018, 1, 1),
            lieu_naissance='Conakry', classe=self.classe,
            date_inscription=date(2025, 9, 1),
        )
        EcheancierPaiement.objects.create(
            eleve=eleve, annee_scolaire='2025-2026',
            frais_inscription_du=Decimal('30000'),
            tranche_1_due=Decimal('100000'),
            tranche_2_due=Decimal('110000'),
            tranche_3_due=Decimal('120000'),
            frais_inscription_paye=Decimal('30000'),
            tranche_1_payee=Decimal('100000'),
            date_echeance_inscription=date(2025, 9, 1),
            date_echeance_tranche_1=date(2026, 1, 15),
            date_echeance_tranche_2=date(2026, 3, 15),
            date_echeance_tranche_3=date(2026, 5, 15),
        )
        paiement = Paiement.objects.create(
            eleve=eleve, type_paiement=self.type_inscription,
            mode_paiement=self.mode, montant=Decimal('130000'),
            date_paiement=date(2025, 9, 10), statut='VALIDE',
        )
        remise = RemiseReduction.objects.create(
            nom='Remise scolarité 10%', type_remise='POURCENTAGE',
            valeur=Decimal('10'), motif='SOCIALE',
            date_debut=date(2025, 9, 1), date_fin=date(2026, 8, 31),
        )
        PaiementRemise.objects.create(
            paiement=paiement, remise=remise,
            montant_remise=Decimal('10000'), portee_tranches='1',
            deduite_du_paiement=False,
        )
        return paiement

    def test_remise_non_deduite_ne_devient_pas_un_paiement_t2_pdf_excel(self):
        self._creer_cas_remise_non_deduite()

        excel_response = self.client.get(
            reverse('paiements:export_tranches_par_classe_excel'),
            self._params(),
        )
        workbook = load_workbook(BytesIO(excel_response.content), data_only=True)
        sheet = workbook[self.classe.nom]
        ligne = next(
            row for row in sheet.iter_rows(min_row=3, values_only=True)
            if row[7] == 130000 and row[8] == 10000
        )
        self.assertEqual(ligne[1:6], (30000, 0, 100000, 0, 0))
        self.assertEqual(ligne[7], 130000)
        self.assertEqual(ligne[8], 10000)
        self.assertEqual(ligne[9], 0.1)
        self.assertEqual(ligne[10], 220000)

        from reportlab.platypus import Table as RealTable
        tables = []

        def capturer_table(data, *args, **kwargs):
            tables.append(data)
            return RealTable(data, *args, **kwargs)

        with patch('reportlab.platypus.Table', side_effect=capturer_table):
            pdf_response = self.client.get(
                reverse('paiements:export_tranches_par_classe_pdf'),
                self._params(),
            )

        self.assertEqual(pdf_response.status_code, 200)
        ligne_pdf = next(
            row for row in tables[0][1:]
            if row[7] == '130 000' and row[8] == '10 000'
        )
        self.assertEqual(ligne_pdf[1:6], ['30 000', '0', '100 000', '0', '0'])
        self.assertEqual(ligne_pdf[8], '10 000')
        self.assertEqual(ligne_pdf[9], '10.0 %')
        self.assertEqual(ligne_pdf[10], '220 000')

    def test_recu_affiche_la_remise_une_fois_sans_fausse_t2(self):
        paiement = self._creer_cas_remise_non_deduite()
        from reportlab.pdfgen.canvas import Canvas as RealCanvas

        textes = []

        def canvas_capture(*args, **kwargs):
            pdf_canvas = RealCanvas(*args, **kwargs)
            draw_string = pdf_canvas.drawString

            def capturer_draw_string(x, y, texte, *draw_args, **draw_kwargs):
                textes.append(str(texte))
                return draw_string(x, y, texte, *draw_args, **draw_kwargs)

            pdf_canvas.drawString = capturer_draw_string
            return pdf_canvas

        with patch('paiements.views.canvas.Canvas', side_effect=canvas_capture):
            response = self.client.get(reverse(
                'paiements:generer_recu_pdf',
                kwargs={'paiement_id': paiement.id},
            ))

        self.assertEqual(response.status_code, 200)
        debut = textes.index('Affectation du paiement')
        fin = textes.index("Informations de l'élève")
        affectation = textes[debut:fin]
        self.assertIn('Inscription: 30 000 GNF', affectation)
        self.assertIn('1ère tranche: 100 000 GNF', affectation)
        self.assertFalse(any('2ème tranche' in texte for texte in affectation))
        self.assertFalse(any('3ème tranche' in texte for texte in affectation))
        self.assertFalse(any(texte.startswith('Total remises') for texte in textes))
        self.assertFalse(any(texte.startswith('Montant net payé') for texte in textes))
        self.assertIn('Remises appliquées', textes)
        self.assertIn('- Remise scolarité 10% (T1) : -10 000 GNF', textes)
