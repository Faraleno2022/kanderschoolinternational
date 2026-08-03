from datetime import date
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook

from eleves.models import Classe, Ecole, Eleve, GrilleTarifaire, Responsable
from paiements.models import EcheancierPaiement

from .support import TEST_MIDDLEWARE


@override_settings(MIDDLEWARE=TEST_MIDDLEWARE)
class InscriptionReinscriptionReportingTests(TestCase):
    def setUp(self):
        self.ecole = Ecole.objects.create(
            nom="Ecole ventilation",
            adresse="Conakry",
            telephone="+224620000301",
            directeur="Direction",
        )
        self.classe_mixte = Classe.objects.create(
            ecole=self.ecole,
            nom="Classe mixte",
            niveau="PRIMAIRE_1",
            annee_scolaire="2025-2026",
        )
        self.classe_reinscription = Classe.objects.create(
            ecole=self.ecole,
            nom="Classe reinscription",
            niveau="PRIMAIRE_1",
            annee_scolaire="2025-2026",
        )
        GrilleTarifaire.objects.create(
            ecole=self.ecole,
            niveau="PRIMAIRE_1",
            annee_scolaire="2025-2026",
            frais_inscription=Decimal("30000"),
            frais_reinscription=Decimal("20000"),
        )
        responsable = Responsable.objects.create(
            prenom="Parent",
            nom="Test",
            relation="PERE",
            telephone="+224620000302",
            adresse="Conakry",
        )
        self._create_student("VENT-INS", self.classe_mixte, responsable, Decimal("30000"))
        self._create_student("VENT-RE1", self.classe_mixte, responsable, Decimal("20000"))
        self._create_student("VENT-RE2", self.classe_reinscription, responsable, Decimal("20000"))

        User = get_user_model()
        self.user = User.objects.create_superuser(
            username="admin_ventilation",
            email="admin.ventilation@example.com",
            password="pass12345",
        )
        self.client.force_login(self.user)

    def _create_student(self, matricule, classe, responsable, admission_due):
        eleve = Eleve.objects.create(
            matricule=matricule,
            prenom="Eleve",
            nom=matricule,
            sexe="F",
            date_naissance=date(2018, 1, 1),
            lieu_naissance="Conakry",
            classe=classe,
            date_inscription=date(2025, 9, 1),
            responsable_principal=responsable,
        )
        EcheancierPaiement.objects.create(
            eleve=eleve,
            annee_scolaire="2025-2026",
            frais_inscription_du=admission_due,
            date_echeance_inscription=date(2025, 9, 1),
            date_echeance_tranche_1=date(2026, 1, 15),
            date_echeance_tranche_2=date(2026, 3, 15),
            date_echeance_tranche_3=date(2026, 5, 15),
        )

    def test_totaux_et_detail_ne_comptent_jamais_deux_fois_la_reinscription(self):
        response = self.client.get(reverse("paiements:liste_paiements"))

        self.assertEqual(response.status_code, 200)
        totals = response.context["totaux_du"]
        self.assertEqual(totals["frais_inscription_total"], 30000)
        self.assertEqual(totals["frais_reinscription_total"], 40000)
        self.assertEqual(totals["du_global_net"], 70000)
        self.assertAlmostEqual(totals["frais_reinscription_pct"], 40000 / 70000 * 100)

        rows = {
            row["classe_id"]: row
            for row in response.context["totaux_du_detail_classes"]
        }
        mixed = rows[self.classe_mixte.id]
        self.assertEqual(mixed["frais_inscription_total"], 30000)
        self.assertEqual(mixed["frais_reinscription_total"], 20000)
        self.assertEqual(mixed["du_global_net"], 50000)
        self.assertEqual(mixed["frais_reinscription_pct"], 40.0)

        reinscription_only = rows[self.classe_reinscription.id]
        self.assertEqual(reinscription_only["frais_inscription_total"], 0)
        self.assertEqual(reinscription_only["frais_reinscription_total"], 20000)
        self.assertEqual(reinscription_only["du_global_net"], 20000)
        self.assertEqual(reinscription_only["frais_reinscription_pct"], 100.0)

    def test_export_excel_utilise_la_meme_ventilation(self):
        response = self.client.get(reverse("paiements:export_recap_par_classe_excel"))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        rows = {
            row[1]: row
            for row in workbook.active.iter_rows(min_row=2, values_only=True)
        }
        mixed = rows[self.classe_mixte.nom]
        self.assertEqual(mixed[4], 30000)
        self.assertEqual(mixed[5], 20000)
        self.assertEqual(mixed[6], 40.0)
        self.assertEqual(mixed[7], 50000)

        reinscription_only = rows[self.classe_reinscription.nom]
        self.assertEqual(reinscription_only[4], 0)
        self.assertEqual(reinscription_only[5], 20000)
        self.assertEqual(reinscription_only[6], 100.0)
        self.assertEqual(reinscription_only[7], 20000)
