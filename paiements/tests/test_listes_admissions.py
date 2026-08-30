from datetime import date
from decimal import Decimal
from io import BytesIO

from django.contrib.auth.models import User
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook

from eleves.models import Classe, Ecole, Eleve
from paiements.models import EcheancierPaiement

from .support import TEST_MIDDLEWARE


@override_settings(MIDDLEWARE=TEST_MIDDLEWARE)
class ListesAdmissionsTests(TestCase):
    def setUp(self):
        self.ecole = self._school('École admissions', '+224620004001')
        self.other_school = self._school('École extérieure', '+224620004002')
        self.classe = self._class(self.ecole, '1ère A')
        self.other_class = self._class(self.other_school, '1ère B')
        self.user = User.objects.create_user('admissions', password='secret')
        self.user.profil.role = 'COMPTABLE'
        self.user.profil.telephone = '+224620004003'
        self.user.profil.ecole = self.ecole
        self.user.profil.is_validated = True
        self.user.profil.save()
        self.inscrit = self._student(self.classe, 'ADM-001', 'Inscrit')
        self.reinscrit = self._student(self.classe, 'ADM-002', 'Réinscrit')
        self.external = self._student(self.other_class, 'ADM-EXT', 'Externe')
        self._schedule(self.inscrit, EcheancierPaiement.NATURE_INSCRIPTION)
        self._schedule(self.reinscrit, EcheancierPaiement.NATURE_REINSCRIPTION)
        self._schedule(self.external, EcheancierPaiement.NATURE_INSCRIPTION)
        self.client.force_login(self.user)

    @staticmethod
    def _school(name, phone):
        return Ecole.objects.create(
            nom=name, adresse='Conakry', telephone=phone,
            directeur='Direction', etat='VALIDE',
        )

    @staticmethod
    def _class(school, name):
        return Classe.objects.create(
            ecole=school, nom=name, niveau='PRIMAIRE_1',
            annee_scolaire='2026-2027',
        )

    @staticmethod
    def _student(classe, matricule, first_name):
        return Eleve.objects.create(
            matricule=matricule, prenom=first_name, nom='Élève',
            sexe='F', classe=classe,
        )

    @staticmethod
    def _schedule(student, nature):
        return EcheancierPaiement.objects.create(
            eleve=student,
            annee_scolaire='2026-2027',
            nature_frais=nature,
            frais_inscription_du=Decimal('50000'),
            frais_inscription_paye=Decimal('50000'),
            date_echeance_inscription=date(2026, 7, 1),
            date_echeance_tranche_1=date(2026, 9, 1),
            date_echeance_tranche_2=date(2027, 1, 1),
            date_echeance_tranche_3=date(2027, 3, 1),
        )

    def test_listes_separent_inscription_reinscription_et_ecole(self):
        inscrits = self.client.get(reverse('paiements:liste_eleves_inscrits'))
        self.assertContains(inscrits, 'ADM-001')
        self.assertNotContains(inscrits, 'ADM-002')
        self.assertNotContains(inscrits, 'ADM-EXT')

        reinscrits = self.client.get(reverse('paiements:liste_eleves_reinscrits'))
        self.assertContains(reinscrits, 'ADM-002')
        self.assertNotContains(reinscrits, 'ADM-001')

    def test_exports_excel_et_pdf(self):
        excel = self.client.get(reverse('paiements:export_eleves_inscrits_excel'))
        self.assertEqual(excel.status_code, 200)
        workbook = load_workbook(BytesIO(excel.content))
        values = [row[0].value for row in workbook.active.iter_rows(min_row=3)]
        self.assertIn('ADM-001', values)
        self.assertNotIn('ADM-002', values)
        self.assertNotIn('ADM-EXT', values)

        pdf = self.client.get(reverse('paiements:export_eleves_reinscrits_pdf'))
        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(pdf['Content-Type'], 'application/pdf')
        self.assertTrue(pdf.content.startswith(b'%PDF'))
