from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden
from django.db.models import Prefetch
from django.utils import timezone
from datetime import date, datetime
from decimal import Decimal
from functools import partial
from types import SimpleNamespace
import unicodedata
from xml.sax.saxutils import escape

from eleves.models import Classe, GrilleTarifaire
from eleves.utils_annee import get_annee_active
from paiements.allocation import allocate_cash_and_discounts
from paiements.models import Paiement, PaiementRemise
from utilisateurs.utils import user_is_admin, user_school
from rapports.utils import _draw_header_and_watermark, _get_logo_path

# ReportLab
# ReportLab: fera l'objet d'un import différé dans la vue PDF


def _annee_vers_dates(annee_scolaire: str):
    try:
        deb, fin = annee_scolaire.split('-')
        an_deb = int(deb)
        an_fin = int(fin)
        return an_deb, an_fin
    except Exception:
        # Fallback année en cours selon rentrée (Septembre)
        today = timezone.now().date()
        y = today.year
        if today.month >= 9:
            return y, y + 1
        return y - 1, y


def _normaliser_type_paiement(value):
    """Normalise un libellé pour reconnaître inscription/réinscription."""
    value = unicodedata.normalize('NFKD', str(value or ''))
    return ''.join(
        caractere for caractere in value
        if not unicodedata.combining(caractere)
    ).lower().replace('-', ' ').strip()


def _paiements_valides_de_periode(eleve, annee_scolaire):
    paiements = list(getattr(eleve, '_paiements_valides_export', ()))
    if not annee_scolaire:
        return paiements

    an_deb, an_fin = _annee_vers_dates(annee_scolaire)
    debut = date(an_deb, 9, 1)
    fin = date(an_fin, 8, 31)
    return [
        paiement for paiement in paiements
        if debut <= paiement.date_paiement <= fin
    ]


def _est_reinscription(echeancier, paiements, grille):
    """Détermine la nature du poste d'admission sans le compter deux fois."""
    types = [
        _normaliser_type_paiement(p.type_paiement.nom)
        for p in paiements
        if getattr(p, 'type_paiement_id', None)
    ]
    a_reinscription = any('reinscription' in nom for nom in types)
    a_inscription = any(
        'inscription' in nom and 'reinscription' not in nom
        for nom in types
    )
    if a_reinscription != a_inscription:
        return a_reinscription

    if grille is not None:
        montant_reinscription = Decimal(grille.frais_reinscription or 0)
        montant_admission = Decimal(echeancier.frais_inscription_du or 0)
        if montant_reinscription > 0 and montant_admission == montant_reinscription:
            return True
    return False


def _remises_des_paiements(paiements):
    """Retourne les remises validées préchargées avec les paiements exportés."""
    remises = []
    for paiement in paiements:
        prechargees = getattr(paiement, '_remises_export', None)
        if prechargees is None:
            prechargees = paiement.remises.select_related('remise').all()
        remises.extend(prechargees)
    return remises


def _pourcentage_remises_selectionne(remises):
    """Retourne le taux réellement choisi, jamais un taux recalculé.

    Les remises fixes n'ont pas de pourcentage et produisent donc ``None``.
    Lorsque plusieurs remises en pourcentage sont liées aux paiements d'un
    élève, leurs taux sont additionnés comme les montants accordés.
    """
    taux = [
        Decimal(remise.remise.valeur or 0)
        for remise in remises
        if getattr(remise, 'remise_id', None)
        and getattr(remise.remise, 'type_remise', None) == 'POURCENTAGE'
    ]
    return sum(taux, Decimal('0')) if taux else None


def _allocation_encaissement_affichee(echeancier, montant_encaisse, remises):
    """Ventile uniquement l'argent réellement encaissé dans les colonnes payées.

    Une remise non déduite du reçu augmente la couverture globale, mais ne doit
    jamais repousser artificiellement une partie du paiement vers la tranche
    suivante. Seules les remises explicitement déduites du montant du reçu sont
    pré-affectées avant le paiement net, puisque cet argent n'a pas été encaissé.
    """
    remises_deduites = [
        remise for remise in remises
        if getattr(remise, 'deduite_du_paiement', False)
    ]
    return allocate_cash_and_discounts(
        echeancier,
        montant_encaisse,
        remises_deduites,
    )['cash_allocation']


def _precision_remise(
    total_du, cash, discount, discount_applied, balance, discount_rate=None,
):
    if total_du <= 0:
        status = 'À contrôler'
    elif balance <= 0:
        status = 'Soldé avec remise' if discount > 0 else 'Soldé'
    elif cash + discount_applied > 0:
        status = 'Paiement partiel'
    else:
        status = 'À payer'

    if discount > 0:
        # Le montant et le taux ont chacun leur colonne dédiée. Ne pas les
        # répéter ici évite qu'une remise soit confondue avec un encaissement.
        precision = "Remise appliquée."
        if total_du > 0 and balance <= 0:
            precision += " Élève soldé grâce au paiement et à la remise."
        elif total_du > 0:
            precision += f" Solde restant : {balance:,.0f} GNF.".replace(',', ' ')
        discount_unapplied = max(discount - discount_applied, Decimal('0'))
        if discount_unapplied > 0:
            precision += " Une partie de la remise reste non imputée et doit être contrôlée."
    else:
        precision = status

    return {
        'remise': discount,
        'remise_imputee': discount_applied,
        'remise_pct': discount_rate,
        'reste': max(balance, Decimal('0')),
        'situation': status,
        'precision': precision,
    }


def _ventiler_sans_echeancier(paiements, grille):
    """Reconstitue la répartition en reproduisant l'ordre du moteur d'allocation."""
    total_paye = sum((Decimal(p.montant or 0) for p in paiements), Decimal('0'))
    remises = _remises_des_paiements(paiements)
    total_remise = sum(
        (Decimal(remise.montant_remise or 0) for remise in remises),
        Decimal('0'),
    )
    types = [
        _normaliser_type_paiement(p.type_paiement.nom)
        for p in paiements
        if getattr(p, 'type_paiement_id', None)
    ]
    est_reinscription = any('reinscription' in nom for nom in types)

    if grille is None:
        # Sans grille ni échéancier, seul le total encaissé est certain.
        return {
            'inscription': Decimal('0'), 'reinscription': Decimal('0'),
            't1': Decimal('0'), 't2': Decimal('0'), 't3': Decimal('0'),
            'total_du': Decimal('0'), 'tuition_due': Decimal('0'),
            'total_paye': total_paye,
            **_precision_remise(
                Decimal('0'), total_paye, total_remise,
                Decimal('0'), Decimal('0'),
                _pourcentage_remises_selectionne(remises),
            ),
        }

    admission_due = Decimal((
        grille.frais_reinscription if est_reinscription
        else grille.frais_inscription
    ) or 0)
    dus = [
        admission_due,
        Decimal(grille.tranche_1 or 0),
        Decimal(grille.tranche_2 or 0),
        Decimal(grille.tranche_3 or 0),
    ]
    proxy = SimpleNamespace(
        frais_inscription_du=dus[0], frais_inscription_paye=Decimal('0'),
        tranche_1_due=dus[1], tranche_1_payee=Decimal('0'),
        tranche_2_due=dus[2], tranche_2_payee=Decimal('0'),
        tranche_3_due=dus[3], tranche_3_payee=Decimal('0'),
    )
    coverage = allocate_cash_and_discounts(proxy, total_paye, remises)
    cash_allocation = _allocation_encaissement_affichee(
        proxy, total_paye, remises,
    )
    total_du = sum(dus, Decimal('0'))
    tuition_due = sum(dus[1:], Decimal('0'))

    return {
        'inscription': (
            Decimal('0') if est_reinscription
            else cash_allocation['inscription']
        ),
        'reinscription': (
            cash_allocation['inscription'] if est_reinscription
            else Decimal('0')
        ),
        't1': cash_allocation['tranche_1'],
        't2': cash_allocation['tranche_2'],
        't3': cash_allocation['tranche_3'],
        'total_du': total_du,
        'tuition_due': tuition_due,
        'total_paye': total_paye,
        **_precision_remise(
            total_du,
            coverage['cash_applied'],
            coverage['discount_recorded'],
            coverage['discount_applied'],
            coverage['balance'],
            _pourcentage_remises_selectionne(remises),
        ),
    }


def _lignes_classe(classe, annee_scolaire):
    """Retourne les mêmes lignes comptables pour les exports PDF et Excel."""
    grille = GrilleTarifaire.objects.filter(
        ecole=classe.ecole,
        niveau=classe.niveau,
        annee_scolaire=annee_scolaire or classe.annee_scolaire,
    ).first()
    paiements_valides = (
        Paiement.objects
        .filter(statut='VALIDE')
        .select_related('type_paiement')
        .prefetch_related(Prefetch(
            'remises',
            queryset=PaiementRemise.objects.select_related('remise'),
            to_attr='_remises_export',
        ))
        .order_by('date_paiement', 'id')
    )
    eleves = classe.eleves.prefetch_related('echeanciers').prefetch_related(
        Prefetch(
            'paiements', queryset=paiements_valides,
            to_attr='_paiements_valides_export',
        )
    ).order_by('nom', 'prenom')
    if hasattr(eleves.model, 'est_dans_corbeille'):
        eleves = eleves.filter(est_dans_corbeille=False)

    for eleve in eleves:
        paiements = _paiements_valides_de_periode(
            eleve, annee_scolaire or classe.annee_scolaire,
        )
        echeancier = getattr(eleve, 'echeancier', None)
        if echeancier is not None and (
            not annee_scolaire or echeancier.annee_scolaire == annee_scolaire
        ):
            remises = _remises_des_paiements(paiements)
            paiement_total = sum(
                (Decimal(p.montant or 0) for p in paiements),
                Decimal('0'),
            )
            # Les paiements validés sont l'autorité pour le total encaissé.
            # Les champs *_payee de certains anciens échéanciers peuvent aussi
            # contenir une remise et ne doivent donc servir qu'en fallback.
            cash_source = (
                paiement_total
                if paiements
                else Decimal(echeancier.total_paye or 0)
            )
            coverage = allocate_cash_and_discounts(
                echeancier, cash_source, remises,
            )
            cash_allocation = _allocation_encaissement_affichee(
                echeancier, cash_source, remises,
            )
            admission_paye = cash_allocation['inscription']
            reinscription = _est_reinscription(
                echeancier, paiements, grille,
            )
            inscription_payee = (
                Decimal('0') if reinscription else admission_paye
            )
            reinscription_payee = (
                admission_paye if reinscription else Decimal('0')
            )
            t1 = cash_allocation['tranche_1']
            t2 = cash_allocation['tranche_2']
            t3 = cash_allocation['tranche_3']
            total_du = Decimal(echeancier.total_du or 0)
            tuition_due = sum((
                Decimal(echeancier.tranche_1_due or 0),
                Decimal(echeancier.tranche_2_due or 0),
                Decimal(echeancier.tranche_3_due or 0),
            ), Decimal('0'))
            valeurs = {
                'inscription': inscription_payee,
                'reinscription': reinscription_payee,
                't1': t1, 't2': t2, 't3': t3,
                'total_du': total_du,
                'tuition_due': tuition_due,
                'total_paye': cash_source,
                **_precision_remise(
                    total_du,
                    coverage['cash_applied'],
                    coverage['discount_recorded'],
                    coverage['discount_applied'],
                    coverage['balance'],
                    _pourcentage_remises_selectionne(remises),
                ),
            }
        else:
            valeurs = _ventiler_sans_echeancier(paiements, grille)

        valeurs['eleve'] = (
            getattr(eleve, 'nom_complet', None)
            or f"{getattr(eleve, 'prenom', '')} {getattr(eleve, 'nom', '')}".strip()
        )
        yield valeurs


@login_required
def export_tranches_par_classe_pdf(request):
    """Export PDF des tranches par classe avec logo entête et filigrane.

    Filtres GET:
    - ecole: id de l'école
    - classe: id de la classe
    - annee_scolaire: ex '2024-2025'

    Respecte la séparation par école pour les non-admins.
    """
    # Contrôle d'accès: Admin ou Comptable uniquement
    is_admin = user_is_admin(request.user)
    is_comptable = False
    try:
        if hasattr(request.user, 'profil'):
            is_comptable = (getattr(request.user.profil, 'role', None) == 'COMPTABLE')
    except Exception:
        is_comptable = False
    if not (is_admin or is_comptable):
        return HttpResponseForbidden("Accès refusé: vous n'avez pas l'autorisation d'exporter ce rapport.")

    # Lecture et validation des paramètres
    raw_ecole = (request.GET.get('ecole') or '').strip()
    raw_classe = (request.GET.get('classe') or request.GET.get('classe_id') or '').strip()
    annee_scolaire = (request.GET.get('annee_scolaire') or '').strip()

    def parse_int(value):
        try:
            return int(value)
        except Exception:
            return None

    ecole_id = parse_int(raw_ecole) if raw_ecole else None
    classe_id = parse_int(raw_classe) if raw_classe else None

    # Scope classes (filtrées par année active)
    classes = Classe.objects.select_related('ecole').all()
    ecole_user = user_school(request.user)
    annee_active = get_annee_active(request, ecole_user) if ecole_user else None
    restreindre = not user_is_admin(request.user) and ecole_user is not None
    if restreindre:
        classes = classes.filter(ecole=ecole_user)
    elif ecole_id:
        classes = classes.filter(ecole_id=ecole_id)
    if classe_id:
        classes = classes.filter(id=classe_id)
    if annee_active and not annee_scolaire:
        classes = classes.filter(annee_scolaire=annee_active)
    elif annee_scolaire:
        classes = classes.filter(annee_scolaire=annee_scolaire)

    # Anti-abus: limiter le nombre de classes exportées en une requête
    classes = list(classes.order_by('ecole__nom', 'niveau', 'nom')[:200])
    ecoles_exportees = {
        classe.ecole_id: classe.ecole
        for classe in classes
        if getattr(classe, 'ecole_id', None)
    }
    ecole_entete = (
        next(iter(ecoles_exportees.values()))
        if len(ecoles_exportees) == 1 else None
    )

    # Préparer réponse PDF
    response = HttpResponse(content_type='application/pdf')
    suffix = datetime.now().strftime('%Y%m%d')
    response['Content-Disposition'] = f'attachment; filename="tranches_par_classe_{suffix}.pdf"'

    # Import différé de ReportLab pour éviter les erreurs si non installé
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import Image, SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
    except Exception:
        return HttpResponse("ReportLab n'est pas installé. Veuillez exécuter: pip install reportlab", status=500)

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=20, leftMargin=20, topMargin=60, bottomMargin=30
    )
    elements = []
    styles = getSampleStyleSheet()
    cell = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=8, leading=9)
    header_cell = ParagraphStyle(
        'HeaderCell', parent=cell, fontName='Helvetica-Bold',
        textColor=colors.white, fontSize=6.4, leading=7.2,
    )

    titre = 'Tranches par classe'
    if annee_scolaire:
        titre += f" – Année {annee_scolaire}"
    elements.append(Paragraph(titre, styles['Title']))
    elements.append(Spacer(1, 0.5*cm))

    header = [
        'Élève', 'Inscription payée', 'Réinscription payée',
        'Tranche 1 payée', 'Tranche 2 payée', 'Tranche 3 payée',
        'Total dû', 'Total encaissé', 'Remise', 'Remise %',
        'Reste', 'Situation / précision',
    ]

    def P(x):
        safe_value = escape(str(x or '')).replace('\n', '<br/>')
        return Paragraph(safe_value, cell)

    def H(x):
        return Paragraph(escape(str(x or '')), header_cell)

    # Parcours des classes
    for classe in classes:
        # Pour un export multi-écoles, chaque section porte explicitement le
        # logo de son établissement. Pour une seule école, le logo est répété
        # dans l'en-tête de chaque page par le callback ci-dessous.
        if len(ecoles_exportees) > 1:
            logo_path = _get_logo_path(classe.ecole)
            titre_ecole = Paragraph(
                escape(getattr(classe.ecole, 'nom', '') or 'Établissement'),
                styles['Heading2'],
            )
            if logo_path:
                logo = Image(logo_path, width=1.8*cm, height=1.1*cm)
                marque = Table(
                    [[logo, titre_ecole]],
                    colWidths=[2.2*cm, 23.5*cm],
                    style=TableStyle([
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('LEFTPADDING', (0, 0), (-1, -1), 0),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                        ('TOPPADDING', (0, 0), (-1, -1), 0),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                    ]),
                )
                elements.append(marque)
            else:
                elements.append(titre_ecole)

        # Titre de la classe
        titre_classe = f"Classe: {classe.nom} – {getattr(classe.ecole, 'nom', '')}"
        elements.append(Paragraph(titre_classe, styles['Heading2']))
        elements.append(Spacer(1, 0.2*cm))

        data = [[H(label) for label in header]]
        lignes = list(_lignes_classe(classe, annee_scolaire))

        for ligne in lignes:
            data.append([
                P(ligne['eleve']),
                f"{ligne['inscription']:,}".replace(',', ' '),
                f"{ligne['reinscription']:,}".replace(',', ' '),
                f"{ligne['t1']:,}".replace(',', ' '),
                f"{ligne['t2']:,}".replace(',', ' '),
                f"{ligne['t3']:,}".replace(',', ' '),
                f"{ligne['total_du']:,}".replace(',', ' '),
                f"{ligne['total_paye']:,}".replace(',', ' '),
                f"{ligne['remise']:,}".replace(',', ' '),
                (
                    f"{ligne['remise_pct']:.1f} %"
                    if ligne['remise_pct'] is not None else '—'
                ),
                f"{ligne['reste']:,}".replace(',', ' '),
                P(
                    ligne['situation']
                    if ligne['precision'] == ligne['situation']
                    else f"{ligne['situation']}\n{ligne['precision']}"
                ),
            ])

        if lignes:
            total_remise = sum((ligne['remise'] for ligne in lignes), Decimal('0'))
            data.append([
                P('TOTAL CLASSE'), '', '', '', '', '',
                f"{sum((ligne['total_du'] for ligne in lignes), Decimal('0')):,}".replace(',', ' '),
                f"{sum((ligne['total_paye'] for ligne in lignes), Decimal('0')):,}".replace(',', ' '),
                f"{total_remise:,}".replace(',', ' '), '—',
                f"{sum((ligne['reste'] for ligne in lignes), Decimal('0')):,}".replace(',', ' '),
                P(f"{sum(1 for ligne in lignes if ligne['situation'].startswith('Soldé'))} élève(s) soldé(s)"),
            ])

        # Construire la table pour la classe
        col_widths = (
            [3.6*cm] + [1.75*cm] * 5 + [2.05*cm] * 3
            + [1.55*cm, 2.15*cm, 4.1*cm]
        )
        table = Table(data, repeatRows=1, colWidths=col_widths)
        table_commands = [
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#174A6E')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 8),
            ('FONTSIZE', (0,1), (-1,-1), 6.5),
            ('ALIGN', (1,1), (-1,-1), 'RIGHT'),
            ('ALIGN', (0,0), (0,-1), 'LEFT'),
            ('ALIGN', (-1,1), (-1,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 2),
            ('RIGHTPADDING', (0,0), (-1,-1), 2),
            ('TOPPADDING', (0,0), (-1,-1), 1),
            ('BOTTOMPADDING', (0,0), (-1,-1), 1),
        ]
        if lignes:
            table_commands.extend([
                ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#DCEAF3')),
                ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
            ])
        table.setStyle(TableStyle(table_commands))
        elements.append(table)
        elements.append(Spacer(1, 0.6*cm))

    # Construire le document avec l'en-tête et le logo de l'école sélectionnée.
    # `partial` évite de perdre l'école quand ReportLab appelle le callback.
    dessiner_entete = partial(
        _draw_header_and_watermark,
        ecole=ecole_entete,
        titre_override='Tranches par classe',
    )
    doc.build(
        elements,
        onFirstPage=dessiner_entete,
        onLaterPages=dessiner_entete,
    )
    return response

@login_required
def export_tranches_par_classe_excel(request):
    """Export Excel identique au PDF des tranches par classe.

    Filtres GET facultatifs: ecole, classe/classe_id, annee_scolaire.
    Respecte la séparation par école pour non-admin.
    """
    # Contrôle d'accès
    is_admin = user_is_admin(request.user)
    is_comptable = False
    try:
        if hasattr(request.user, 'profil'):
            is_comptable = (getattr(request.user.profil, 'role', None) == 'COMPTABLE')
    except Exception:
        is_comptable = False
    if not (is_admin or is_comptable):
        return HttpResponseForbidden("Accès refusé: vous n'avez pas l'autorisation d'exporter ce rapport.")

    # Import openpyxl
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return HttpResponse("OpenPyXL n'est pas installé. Veuillez exécuter: pip install openpyxl", status=500)

    raw_ecole = (request.GET.get('ecole') or '').strip()
    raw_classe = (request.GET.get('classe') or request.GET.get('classe_id') or '').strip()
    annee_scolaire = (request.GET.get('annee_scolaire') or '').strip()

    def parse_int(value):
        try:
            return int(value)
        except Exception:
            return None

    ecole_id = parse_int(raw_ecole) if raw_ecole else None
    classe_id = parse_int(raw_classe) if raw_classe else None

    classes = Classe.objects.select_related('ecole').all()
    ecole_user = user_school(request.user)
    annee_active_xl = get_annee_active(request, ecole_user) if ecole_user else None
    restreindre = not user_is_admin(request.user) and ecole_user is not None
    if restreindre:
        classes = classes.filter(ecole=ecole_user)
    elif ecole_id:
        classes = classes.filter(ecole_id=ecole_id)
    if classe_id:
        classes = classes.filter(id=classe_id)
    if annee_active_xl and not annee_scolaire:
        classes = classes.filter(annee_scolaire=annee_active_xl)
    elif annee_scolaire:
        classes = classes.filter(annee_scolaire=annee_scolaire)
    classes = classes.order_by('ecole__nom', 'niveau', 'nom')[:200]

    wb = Workbook()
    ws_index = wb.active
    ws_index.title = 'Index'
    index_title = 'Tranches par classe'
    if annee_scolaire:
        index_title += f" - Année {annee_scolaire}"
    ws_index.append([index_title, None, None])
    ws_index.merge_cells('A1:C1')
    ws_index.cell(1, 1).font = Font(bold=True, size=14, color='174A6E')
    ws_index.cell(1, 1).alignment = Alignment(horizontal='center')
    ws_index.append(['École', 'Classe', 'Feuille'])
    for cell in ws_index[2]:
        cell.fill = PatternFill('solid', fgColor='174A6E')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    ws_index.column_dimensions['A'].width = 44
    ws_index.column_dimensions['B'].width = 26
    ws_index.column_dimensions['C'].width = 28
    ws_index.freeze_panes = 'A3'
    ws_index.sheet_view.showGridLines = False

    headers = [
        'Élève', 'Inscription payée', 'Réinscription payée',
        'Tranche 1 payée', 'Tranche 2 payée', 'Tranche 3 payée',
        'Total dû', 'Total encaissé', 'Remise', 'Remise (%)',
        'Reste', 'Situation', 'Précision remise',
    ]

    for idx, classe in enumerate(classes, start=1):
        sheet_name = f"{classe.nom[:25]}"  # Limite Excel <=31
        ws = wb.create_sheet(title=sheet_name)
        ws.append([f"Classe: {classe.nom} – {getattr(classe.ecole, 'nom', '')}"])
        ws.append(headers)

        lignes = list(_lignes_classe(classe, annee_scolaire))
        for ligne in lignes:
            ws.append([
                ligne['eleve'],
                int(ligne['inscription']), int(ligne['reinscription']),
                int(ligne['t1']), int(ligne['t2']), int(ligne['t3']),
                int(ligne['total_du']), int(ligne['total_paye']),
                int(ligne['remise']),
                (
                    float(ligne['remise_pct'] / Decimal('100'))
                    if ligne['remise_pct'] is not None else None
                ),
                int(ligne['reste']),
                ligne['situation'], ligne['precision'],
            ])

        if lignes:
            total_remise = sum((ligne['remise'] for ligne in lignes), Decimal('0'))
            ws.append([
                'TOTAL CLASSE', None, None, None, None, None,
                int(sum((ligne['total_du'] for ligne in lignes), Decimal('0'))),
                int(sum((ligne['total_paye'] for ligne in lignes), Decimal('0'))),
                int(total_remise), None,
                int(sum((ligne['reste'] for ligne in lignes), Decimal('0'))),
                f"{sum(1 for ligne in lignes if ligne['situation'].startswith('Soldé'))} élève(s) soldé(s)",
                'Les remises sont présentées séparément et déduites uniquement du reste.',
            ])

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
        ws.cell(1, 1).font = Font(bold=True, size=14, color='174A6E')
        ws.cell(1, 1).alignment = Alignment(horizontal='center')
        for cell in ws[2]:
            cell.fill = PatternFill('solid', fgColor='174A6E')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True,
            )
        ws.row_dimensions[2].height = 34
        if lignes:
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill('solid', fgColor='DCEAF3')
                cell.font = Font(bold=True)

        # Ajuster largeur colonnes simple
        for col in range(1, 14):
            if col == 1:
                width = 25
            elif col in (12, 13):
                width = 26 if col == 12 else 52
            else:
                width = 16
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = f"A2:M{max(ws.max_row - 1, 2)}"
        for row in range(3, ws.max_row + 1):
            ws.cell(row, 10).number_format = '0.0%'
            ws.cell(row, 13).alignment = Alignment(wrap_text=True, vertical='top')
            for col in (2, 3, 4, 5, 6, 7, 8, 9, 11):
                ws.cell(row, col).number_format = '#,##0'
        ws.sheet_view.showGridLines = False

        # Index line
        ws_index.append([getattr(classe.ecole, 'nom', ''), classe.nom, sheet_name])

    # Supprimer la feuille par défaut si vide
    if ws_index.max_row == 2:
        ws_index.append(['Aucune classe'])
    else:
        ws_index.auto_filter.ref = f"A2:C{ws_index.max_row}"

    from io import BytesIO
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    resp = HttpResponse(stream.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    suffix = datetime.now().strftime('%Y%m%d')
    filename = f'tranches_par_classe_{suffix}.xlsx'
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp
