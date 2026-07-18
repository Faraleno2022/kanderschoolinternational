from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.db.models import Count, Sum
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from eleves.models import Classe
from utilisateurs.permissions import can_view_reports
from utilisateurs.utils import filter_by_user_school, user_school

from .models import EcheancierPaiement, Paiement, Relance


def _parse_date(value, default):
    try:
        return datetime.strptime(value, "%Y-%m-%d").date() if value else default
    except (TypeError, ValueError):
        return default


def _retard_echeancier(echeancier, date_reference):
    """Montant arrivé à échéance et non payé à la date de référence."""
    retard = Decimal("0")
    postes = (
        (echeancier.date_echeance_inscription, echeancier.frais_inscription_du, echeancier.frais_inscription_paye),
        (echeancier.date_echeance_tranche_1, echeancier.tranche_1_due, echeancier.tranche_1_payee),
        (echeancier.date_echeance_tranche_2, echeancier.tranche_2_due, echeancier.tranche_2_payee),
        (echeancier.date_echeance_tranche_3, echeancier.tranche_3_due, echeancier.tranche_3_payee),
    )
    for echeance, montant_du, montant_paye in postes:
        if echeance and echeance <= date_reference:
            retard += max(Decimal("0"), (montant_du or 0) - (montant_paye or 0))
    return retard


def _rapport_data(request):
    aujourd_hui = timezone.localdate()
    debut_defaut = aujourd_hui.replace(day=1)
    date_debut = _parse_date(request.GET.get("date_debut"), debut_defaut)
    date_fin = _parse_date(request.GET.get("date_fin"), aujourd_hui)
    if date_debut > date_fin:
        date_debut, date_fin = date_fin, date_debut

    classe_id = (request.GET.get("classe") or "").strip()
    statut = (request.GET.get("statut") or "VALIDE").strip().upper()
    statuts_valides = {code for code, _ in Paiement.STATUT_CHOICES}
    if statut not in statuts_valides and statut != "TOUS":
        statut = "VALIDE"

    classes = filter_by_user_school(
        Classe.objects.select_related("ecole").order_by("annee_scolaire", "nom"),
        request.user,
        "ecole",
    )
    classe_selectionnee = None
    if classe_id:
        try:
            classe_selectionnee = classes.filter(pk=int(classe_id)).first()
        except (TypeError, ValueError):
            classe_selectionnee = None

    paiements = filter_by_user_school(
        Paiement.objects.select_related(
            "eleve", "eleve__classe", "eleve__classe__ecole", "type_paiement", "mode_paiement"
        ),
        request.user,
        "eleve__classe__ecole",
    ).filter(date_paiement__range=(date_debut, date_fin))
    if statut != "TOUS":
        paiements = paiements.filter(statut=statut)
    if classe_selectionnee:
        paiements = paiements.filter(eleve__classe=classe_selectionnee)
    paiements = paiements.order_by("eleve__classe__nom", "-date_paiement", "eleve__nom")

    echeanciers = filter_by_user_school(
        EcheancierPaiement.objects.select_related("eleve", "eleve__classe", "eleve__classe__ecole"),
        request.user,
        "eleve__classe__ecole",
    )
    if classe_selectionnee:
        echeanciers = echeanciers.filter(eleve__classe=classe_selectionnee)

    retards = []
    for echeancier in echeanciers.order_by("eleve__classe__nom", "eleve__nom", "eleve__prenom"):
        montant_retard = _retard_echeancier(echeancier, date_fin)
        if montant_retard > 0:
            retards.append({"echeancier": echeancier, "montant_retard": montant_retard})

    relances = filter_by_user_school(
        Relance.objects.select_related("eleve", "eleve__classe", "eleve__classe__ecole"),
        request.user,
        "eleve__classe__ecole",
    ).filter(date_creation__date__range=(date_debut, date_fin))
    if classe_selectionnee:
        relances = relances.filter(eleve__classe=classe_selectionnee)
    relances = relances.order_by("eleve__classe__nom", "-date_creation")

    paiements_list = list(paiements)
    relances_list = list(relances)
    total_paiements = sum((p.montant or Decimal("0") for p in paiements_list), Decimal("0"))
    total_retards = sum((r["montant_retard"] for r in retards), Decimal("0"))

    modes = []
    modes_agreges = paiements.values("mode_paiement__nom").annotate(
        nombre=Count("id"), montant=Sum("montant")
    ).order_by("mode_paiement__nom")
    for ligne in modes_agreges:
        montant = ligne["montant"] or Decimal("0")
        modes.append({
            "nom": ligne["mode_paiement__nom"] or "Non renseigné",
            "nombre": ligne["nombre"],
            "montant": montant,
            "pourcentage": (montant * Decimal("100") / total_paiements) if total_paiements else Decimal("0"),
        })

    classes_stats = {}
    for paiement in paiements_list:
        cle = paiement.eleve.classe_id
        ligne = classes_stats.setdefault(cle, {
            "classe": paiement.eleve.classe, "paiements": 0, "montant": Decimal("0"),
            "retards": Decimal("0"), "relances": 0,
        })
        ligne["paiements"] += 1
        ligne["montant"] += paiement.montant or Decimal("0")
    for retard in retards:
        classe = retard["echeancier"].eleve.classe
        ligne = classes_stats.setdefault(classe.id, {
            "classe": classe, "paiements": 0, "montant": Decimal("0"),
            "retards": Decimal("0"), "relances": 0,
        })
        ligne["retards"] += retard["montant_retard"]
    for relance in relances_list:
        classe = relance.eleve.classe
        ligne = classes_stats.setdefault(classe.id, {
            "classe": classe, "paiements": 0, "montant": Decimal("0"),
            "retards": Decimal("0"), "relances": 0,
        })
        ligne["relances"] += 1

    ecole = classe_selectionnee.ecole if classe_selectionnee else user_school(request.user)
    return {
        "titre_page": "Rapport comptable consolidé",
        "ecole": ecole,
        "classes": classes,
        "classe_selectionnee": classe_selectionnee,
        "classe_id": str(classe_selectionnee.id) if classe_selectionnee else "",
        "date_debut": date_debut,
        "date_fin": date_fin,
        "statut": statut,
        "paiements": paiements_list,
        "retards": retards,
        "relances": relances_list,
        "modes": modes,
        "classes_stats": sorted(classes_stats.values(), key=lambda x: (x["classe"].annee_scolaire, x["classe"].nom)),
        "total_paiements": total_paiements,
        "total_retards": total_retards,
        "nombre_paiements": len(paiements_list),
        "nombre_retards": len(retards),
        "nombre_relances": len(relances_list),
    }


@can_view_reports
def rapport_comptable(request):
    return render(request, "paiements/rapport_comptable.html", _rapport_data(request))


def _filename(data, extension):
    classe = data["classe_selectionnee"]
    portee = classe.nom if classe else "etablissement"
    propre = "".join(c if c.isalnum() or c in "-_" else "_" for c in portee)
    return f"rapport_comptable_{propre}_{data['date_debut']}_{data['date_fin']}.{extension}"


@can_view_reports
def export_rapport_comptable_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    data = _rapport_data(request)
    wb = Workbook()
    ws = wb.active
    ws.title = "Synthèse"
    entete = PatternFill("solid", fgColor="1F4E78")
    blanc = Font(color="FFFFFF", bold=True)

    ws.append([data["ecole"].nom if data["ecole"] else "Tous les établissements"])
    ws.append(["Rapport comptable consolidé"])
    ws.append(["Période", data["date_debut"].strftime("%d/%m/%Y"), data["date_fin"].strftime("%d/%m/%Y")])
    ws.append(["Portée", data["classe_selectionnee"].nom if data["classe_selectionnee"] else "Tout l'établissement"])
    ws.append([])
    ws.append(["Indicateur", "Valeur"])
    ws.append(["Nombre de paiements", data["nombre_paiements"]])
    ws.append(["Montant encaissé (GNF)", data["total_paiements"]])
    ws.append(["Élèves en retard", data["nombre_retards"]])
    ws.append(["Retard total (GNF)", data["total_retards"]])
    ws.append(["Nombre de relances", data["nombre_relances"]])
    ws.append([])
    ws.append(["Mode de paiement", "Opérations", "Montant (GNF)", "% des encaissements"])
    for mode in data["modes"]:
        ws.append([mode["nom"], mode["nombre"], mode["montant"], float(mode["pourcentage"] / 100)])
    for cell in ws[13]:
        cell.fill, cell.font = entete, blanc
    for row in range(14, ws.max_row + 1):
        ws.cell(row, 4).number_format = "0.00%"

    feuilles = (
        ("Paiements", ["Date", "Reçu", "Matricule", "Élève", "Classe", "Type", "Mode", "Montant", "Statut"], [
            [p.date_paiement, p.numero_recu, p.eleve.matricule, p.eleve.nom_complet, p.eleve.classe.nom,
             p.type_paiement.nom, p.mode_paiement.nom, p.montant, p.get_statut_display()]
            for p in data["paiements"]
        ]),
        ("Retards", ["Matricule", "Élève", "Classe", "Année scolaire", "Retard (GNF)"], [
            [r["echeancier"].eleve.matricule, r["echeancier"].eleve.nom_complet,
             r["echeancier"].eleve.classe.nom, r["echeancier"].annee_scolaire, r["montant_retard"]]
            for r in data["retards"]
        ]),
        ("Relances", ["Date", "Matricule", "Élève", "Classe", "Canal", "Statut", "Solde estimé"], [
            [timezone.localtime(r.date_creation).replace(tzinfo=None), r.eleve.matricule, r.eleve.nom_complet,
             r.eleve.classe.nom, r.get_canal_display(), r.get_statut_display(), r.solde_estime]
            for r in data["relances"]
        ]),
        ("Par classe", ["Classe", "Année", "Paiements", "Encaissé", "Retards", "Relances"], [
            [r["classe"].nom, r["classe"].annee_scolaire, r["paiements"], r["montant"], r["retards"], r["relances"]]
            for r in data["classes_stats"]
        ]),
    )
    for titre, colonnes, lignes in feuilles:
        feuille = wb.create_sheet(titre)
        feuille.append(colonnes)
        for cell in feuille[1]:
            cell.fill, cell.font, cell.alignment = entete, blanc, Alignment(horizontal="center")
        for ligne in lignes:
            feuille.append(ligne)
        feuille.freeze_panes = "A2"
        feuille.auto_filter.ref = feuille.dimensions
        for colonne in range(1, feuille.max_column + 1):
            largeur = max((len(str(feuille.cell(row, colonne).value or "")) for row in range(1, feuille.max_row + 1)), default=10)
            feuille.column_dimensions[get_column_letter(colonne)].width = min(max(largeur + 2, 12), 40)

    sortie = BytesIO()
    wb.save(sortie)
    response = HttpResponse(
        sortie.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{_filename(data, "xlsx")}"'
    return response


@can_view_reports
def export_rapport_comptable_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    data = _rapport_data(request)
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{_filename(data, "pdf")}"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=10 * mm, bottomMargin=10 * mm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(data["ecole"].nom if data["ecole"] else "Tous les établissements", styles["Title"]),
        Paragraph("Rapport comptable consolidé", styles["Heading1"]),
        Paragraph(
            f"Période du {data['date_debut']:%d/%m/%Y} au {data['date_fin']:%d/%m/%Y} — "
            f"{data['classe_selectionnee'].nom if data['classe_selectionnee'] else 'Tout l’établissement'}",
            styles["Normal"],
        ), Spacer(1, 5 * mm),
    ]

    def tableau(titre, entetes, lignes, largeurs=None):
        elements.append(Paragraph(titre, styles["Heading2"]))
        contenu = [entetes] + (lignes or [["Aucune donnée"] + [""] * (len(entetes) - 1)])
        t = Table(contenu, repeatRows=1, colWidths=largeurs)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F6F9")]),
        ]))
        elements.extend([t, Spacer(1, 4 * mm)])

    tableau("Indicateurs", ["Paiements", "Encaissé (GNF)", "Élèves en retard", "Retard (GNF)", "Relances"], [[
        data["nombre_paiements"], f"{data['total_paiements']:,.0f}", data["nombre_retards"],
        f"{data['total_retards']:,.0f}", data["nombre_relances"],
    ]])
    tableau("Rapprochement par mode de paiement", ["Mode", "Opérations", "Montant (GNF)", "%"], [
        [m["nom"], m["nombre"], f"{m['montant']:,.0f}", f"{m['pourcentage']:.2f}%"] for m in data["modes"]
    ])
    tableau("Synthèse par classe", ["Classe", "Année", "Paiements", "Encaissé", "Retards", "Relances"], [
        [r["classe"].nom, r["classe"].annee_scolaire, r["paiements"], f"{r['montant']:,.0f}",
         f"{r['retards']:,.0f}", r["relances"]] for r in data["classes_stats"]
    ])
    elements.append(PageBreak())
    tableau("Détail des paiements", ["Date", "Reçu", "Matricule", "Élève", "Classe", "Type", "Mode", "Montant"], [
        [f"{p.date_paiement:%d/%m/%Y}", p.numero_recu, p.eleve.matricule, p.eleve.nom_complet,
         p.eleve.classe.nom, p.type_paiement.nom, p.mode_paiement.nom, f"{p.montant:,.0f}"]
        for p in data["paiements"]
    ])
    tableau("Retards de paiement", ["Matricule", "Élève", "Classe", "Année", "Retard (GNF)"], [
        [r["echeancier"].eleve.matricule, r["echeancier"].eleve.nom_complet,
         r["echeancier"].eleve.classe.nom, r["echeancier"].annee_scolaire, f"{r['montant_retard']:,.0f}"]
        for r in data["retards"]
    ])
    tableau("Relances", ["Date", "Matricule", "Élève", "Classe", "Canal", "Statut", "Solde estimé"], [
        [timezone.localtime(r.date_creation).strftime("%d/%m/%Y %H:%M"), r.eleve.matricule, r.eleve.nom_complet,
         r.eleve.classe.nom, r.get_canal_display(), r.get_statut_display(), f"{r.solde_estime:,.0f}"]
        for r in data["relances"]
    ])
    doc.build(elements)
    return response
