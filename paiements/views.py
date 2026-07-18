from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.cache import cache_page
from django.views.decorators.vary import vary_on_cookie
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods, require_POST
from django.contrib import messages
from django.db import transaction, IntegrityError
from django.db.models import Q, F, Sum, Count, Value, DecimalField, ExpressionWrapper, Case, When, OuterRef, Subquery
from django.db.models.functions import Coalesce, Greatest, Least
from django.http import JsonResponse, HttpResponse, Http404
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.core.cache import cache
from django.core.paginator import Paginator
from decimal import Decimal
from datetime import date, datetime, timedelta
from io import BytesIO
import os
import logging
import urllib.parse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.utils import ImageReader
except Exception:
    canvas = None
    A4 = (595.27, 841.89)
    ImageReader = None
from ecole_moderne.pdf_utils import draw_logo_watermark
from ecole_moderne.security_decorators import require_school_object

from .models import Paiement, EcheancierPaiement, TypePaiement, ModePaiement, RemiseReduction, PaiementRemise, Relance, TwilioInboundMessage
from eleves.models import Eleve, GrilleTarifaire, Classe
from eleves.utils_annee import get_annee_active
from .forms import PaiementForm, EcheancierForm, RechercheForm
from .remise_forms import PaiementRemiseForm, CalculateurRemiseForm
from utilisateurs.utils import user_is_admin, user_is_superadmin, filter_by_user_school, user_school
from utilisateurs.permissions import has_permission, get_user_permissions, can_add_payments, can_modify_payments, can_delete_payments, can_validate_payments, can_view_reports, can_apply_discounts
from .notifications import (
    send_payment_receipt,
    send_enrollment_confirmation,
    send_relance_notification,
    send_retard_notification,
)

def ensure_echeancier_for_eleve(eleve: "Eleve", *, created_by=None, prefer_reinscription: bool = False) -> "EcheancierPaiement":
    """Crée (silencieusement) un `EcheancierPaiement` pour l'élève s'il n'existe pas.

    - Utilise `eleves.GrilleTarifaire` pour pré-remplir les montants dus et l'année scolaire
    - Définit des dates d'échéance par défaut: inscription=today, T1=15/01, T2=15/03, T3=15/05
    - Retourne l'échéancier existant ou nouvellement créé
    """
    try:
        ech = getattr(eleve, 'echeancier', None)
    except Exception:
        ech = None

    # Si un échéancier existe mais semble vide (tous les dues = 0), on tentera de le renseigner via la grille
    if ech is not None:
        try:
            total_du_exist = int((ech.frais_inscription_du or 0) + (ech.tranche_1_due or 0) + (ech.tranche_2_due or 0) + (ech.tranche_3_due or 0))
        except Exception:
            total_du_exist = 0
        if total_du_exist > 0:
            return ech
        # sinon on essaiera plus bas de charger une grille et de mettre à jour cet échéancier

    # Déterminer/charger la grille tarifaire applicable
    try:
        niveau = getattr(eleve.classe, 'niveau', None)
        ecole = getattr(eleve.classe, 'ecole', None)
        annee_classe = getattr(eleve.classe, 'annee_scolaire', None)
    except Exception:
        niveau = None
        ecole = None
        annee_classe = None

    try:
        from datetime import date as _d
        today_d = _d.today()
    except Exception:
        from datetime import date as _d
        today_d = _d.today()

    annee_scolaire_def = f"{today_d.year}-{today_d.year+1}" if today_d.month >= 9 else f"{today_d.year-1}-{today_d.year}"
    grille = None
    try:
        if ecole and niveau:
            # 1) Grille exacte sur l'année de la classe
            if annee_classe:
                grille = GrilleTarifaire.objects.filter(ecole=ecole, niveau=niveau, annee_scolaire=annee_classe).first()
            # 2) Sinon année scolaire par défaut
            if grille is None:
                grille = GrilleTarifaire.objects.filter(ecole=ecole, niveau=niveau, annee_scolaire=annee_scolaire_def).first()
            # 3) Sinon la plus récente
            if grille is None:
                grille = GrilleTarifaire.objects.filter(ecole=ecole, niveau=niveau).order_by('-annee_scolaire').first()
    except Exception:
        grille = None

    # Préparer les champs
    if grille:
        annee_scol = grille.annee_scolaire
        fi = (grille.frais_reinscription or 0) if prefer_reinscription else (grille.frais_inscription or 0)
        t1 = grille.tranche_1 or 0
        t2 = grille.tranche_2 or 0
        t3 = grille.tranche_3 or 0
    else:
        annee_scol = annee_classe or annee_scolaire_def
        fi = 0
        t1 = 0
        t2 = 0
        t3 = 0

    # Dates d'échéance par défaut (priorité aux valeurs de la grille si présentes)
    try:
        try:
            annee_debut = int(str(annee_scol).split('-')[0])
        except Exception:
            annee_debut = today_d.year if today_d.month >= 9 else today_d.year - 1
        annee_fin = annee_debut + 1
        from datetime import date as _d
        # Par défaut génériques
        default_insc = today_d
        default_t1 = _d(annee_fin, 1, 15)
        default_t2 = _d(annee_fin, 3, 15)
        default_t3 = _d(annee_fin, 5, 15)
        # Surcharges via grille si disponibles
        if grille is not None:
            d_insc = getattr(grille, 'date_echeance_inscription_defaut', None) or default_insc
            d_t1 = getattr(grille, 'date_echeance_tranche_1_defaut', None) or default_t1
            d_t2 = getattr(grille, 'date_echeance_tranche_2_defaut', None) or default_t2
            d_t3 = getattr(grille, 'date_echeance_tranche_3_defaut', None) or default_t3
        else:
            d_insc = default_insc
            d_t1 = default_t1
            d_t2 = default_t2
            d_t3 = default_t3
    except Exception:
        d_insc = today_d
        d_t1 = today_d
        d_t2 = today_d
        d_t3 = today_d

    # Mettre à jour un échéancier existant vide, sinon créer
    if ech is not None:
        try:
            ech.annee_scolaire = annee_scol
            ech.frais_inscription_du = fi
            ech.tranche_1_due = t1
            ech.tranche_2_due = t2
            ech.tranche_3_due = t3
            # Initialiser les dates si absentes
            if not getattr(ech, 'date_echeance_inscription', None):
                ech.date_echeance_inscription = d_insc
            if not getattr(ech, 'date_echeance_tranche_1', None):
                ech.date_echeance_tranche_1 = d_t1
            if not getattr(ech, 'date_echeance_tranche_2', None):
                ech.date_echeance_tranche_2 = d_t2
            if not getattr(ech, 'date_echeance_tranche_3', None):
                ech.date_echeance_tranche_3 = d_t3
            if created_by and getattr(created_by, 'is_authenticated', False) and not getattr(ech, 'cree_par', None):
                ech.cree_par = created_by
            ech.save()
        except Exception:
            logging.getLogger(__name__).exception("Mise à jour de l'échéancier existant échouée")
        return ech
    else:
        try:
            with transaction.atomic():
                ech = EcheancierPaiement.objects.create(
                    eleve=eleve,
                    annee_scolaire=annee_scol,
                    frais_inscription_du=fi,
                    tranche_1_due=t1,
                    tranche_2_due=t2,
                    tranche_3_due=t3,
                    date_echeance_inscription=d_insc,
                    date_echeance_tranche_1=d_t1,
                    date_echeance_tranche_2=d_t2,
                    date_echeance_tranche_3=d_t3,
                    cree_par=created_by if created_by and getattr(created_by, 'is_authenticated', False) else None,
                )
            return ech
        except IntegrityError:
            # Race condition: un autre thread/requête a créé l'échéancier entre-temps
            # → récupérer l'échéancier existant
            logging.getLogger(__name__).info(
                "Échéancier déjà créé par un autre processus pour l'élève %s, récupération.", eleve.id
            )
            return EcheancierPaiement.objects.filter(eleve=eleve).first()

@login_required
def ajax_montant_suggere(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Méthode invalide'}, status=405)
    try:
        eleve_id = request.POST.get('eleve_id')
        type_id = request.POST.get('type_id')
        if not eleve_id or not type_id:
            return JsonResponse({'ok': False, 'error': 'Paramètres manquants'}, status=400)

        # Charger élève (filtré par l'école de l'utilisateur si non admin)
        eleve_qs = Eleve.objects.select_related('classe', 'classe__ecole')
        eleve_qs = filter_by_user_school(eleve_qs, request.user, 'classe__ecole')
        eleve = get_object_or_404(eleve_qs, pk=int(eleve_id))

        type_pmt = get_object_or_404(TypePaiement, pk=int(type_id))
        type_nom = (type_pmt.nom or '').strip().lower()

        # Assurer un échéancier
        ech = getattr(eleve, 'echeancier', None)
        if not ech:
            prefer_reinsc = ('réinscription' in type_nom) or ('reinscription' in type_nom)
            ech = ensure_echeancier_for_eleve(eleve, created_by=request.user, prefer_reinscription=prefer_reinsc)
        if not ech:
            return JsonResponse({'ok': False, 'error': "Aucun échéancier disponible pour l'élève."}, status=400)

        # Récup montants dus/payés
        try:
            fi_due = int(ech.frais_inscription_du or 0)
            fi_pay = int(ech.frais_inscription_paye or 0)
            t1_due = int(ech.tranche_1_due or 0)
            t1_pay = int(ech.tranche_1_payee or 0)
            t2_due = int(ech.tranche_2_due or 0)
            t2_pay = int(ech.tranche_2_payee or 0)
            t3_due = int(ech.tranche_3_due or 0)
            t3_pay = int(ech.tranche_3_payee or 0)
        except Exception:
            fi_due = fi_pay = t1_due = t1_pay = t2_due = t2_pay = t3_due = t3_pay = 0

        rfi = max(0, fi_due - fi_pay)
        rt1 = max(0, t1_due - t1_pay)
        rt2 = max(0, t2_due - t2_pay)
        rt3 = max(0, t3_due - t3_pay)

        suggested = 0
        description = ''
        # Types combinés prioritairement
        if ((('inscription' in type_nom) and ('annuel' in type_nom)) or ((('réinscription' in type_nom) or ('reinscription' in type_nom)) and ('annuel' in type_nom))):
            suggested = rfi + rt1 + rt2 + rt3
            description = "frais d'inscription/réinscription + Annuel (reste)"
        elif ((('inscription' in type_nom) or ('réinscription' in type_nom) or ('reinscription' in type_nom)) and ('tranche 1 + tranche 2' in type_nom or 'tranche1 + tranche2' in type_nom)):
            suggested = rfi + rt1 + rt2
            description = "frais d'inscription/réinscription + Tranche 1 + Tranche 2 (reste)"
        elif ((('inscription' in type_nom) or ('réinscription' in type_nom) or ('reinscription' in type_nom)) and ('tranche 1' in type_nom or '1ère tranche' in type_nom or '1ere tranche' in type_nom)):
            suggested = rfi + rt1
            description = "frais d'inscription/réinscription + Tranche 1 (reste)"
        elif ('inscription' in type_nom) or ('réinscription' in type_nom) or ('reinscription' in type_nom):
            suggested = rfi
            description = "frais d'inscription/réinscription (reste)"
        elif ('tranche 1 + tranche 2 + tranche 3' in type_nom or 'tranche1 + tranche2 + tranche3' in type_nom):
            suggested = rt1 + rt2 + rt3
            description = "Tranche 1 + Tranche 2 + Tranche 3 (reste)"
        elif 'tranche 1 + tranche 2' in type_nom or 'tranche1 + tranche2' in type_nom:
            suggested = rt1 + rt2
            description = "Tranche 1 + Tranche 2 (reste)"
        elif 'tranche 2 + tranche 3' in type_nom or 'tranche2 + tranche3' in type_nom:
            suggested = rt2 + rt3
            description = "Tranche 2 + Tranche 3 (reste)"
        elif 'tranche 1' in type_nom or '1ère tranche' in type_nom or '1ere tranche' in type_nom:
            suggested = rt1
            description = "1ère tranche (reste)"
        elif 'tranche 2' in type_nom or '2ème tranche' in type_nom or '2eme tranche' in type_nom:
            suggested = rt2
            description = "2ème tranche (reste)"
        elif 'tranche 3' in type_nom or '3ème tranche' in type_nom or '3eme tranche' in type_nom:
            suggested = rt3
            description = "3ème tranche (reste)"
        elif 'scolarité' in type_nom:
            suggested = rt1 + rt2 + rt3
            description = "Scolarité (reste)"

        breakdown = {
            'fi_restant': rfi,
            't1_restant': rt1,
            't2_restant': rt2,
            't3_restant': rt3,
            'description': description,
        }
        return JsonResponse({'ok': True, 'suggested': int(suggested or 0), 'breakdown': breakdown})
    except Exception:
        logging.getLogger(__name__).exception("ajax_montant_suggere failed")
        return JsonResponse({'ok': False, 'error': 'Erreur interne'}, status=500)

def _allocate_payment_to_echeancier(paiement: "Paiement") -> None:
    """Alloue intelligemment le montant d'un paiement dans l'échéancier de l'élève.

    Règles:
    - Allouer d'abord les frais d'inscription si encore dûs (fi_due - fi_payee)
    - Puis répartition séquentielle: T1 -> T2 -> T3
    - Si le type contient 'annuel', répartir proportionnellement entre tranches restantes après inscription
    - Ne jamais dépasser les montants dus par tranche
    - Utilise Decimal partout pour éviter les pertes de précision
    """
    _ZERO = Decimal('0')

    try:
        eleve = paiement.eleve

        # Tout le bloc d'allocation est atomique avec verrouillage (select_for_update)
        with transaction.atomic():
            # Verrouiller l'échéancier pour éviter les écritures concurrentes
            ech = EcheancierPaiement.objects.select_for_update().filter(eleve=eleve).first()
            if not ech:
                ech = ensure_echeancier_for_eleve(eleve, created_by=getattr(paiement, 'cree_par', None))
                if ech:
                    # Re-verrouiller après création
                    ech = EcheancierPaiement.objects.select_for_update().filter(pk=ech.pk).first()

            if not ech:
                logging.getLogger(__name__).error(
                    "Impossible de créer/verrouiller l'échéancier pour l'élève %s", eleve.id
                )
                return

            montant = Decimal(str(paiement.montant or 0))
            if montant <= _ZERO:
                return

            # Récupérer dû et payé actuels (garder en Decimal)
            fi_due = Decimal(str(ech.frais_inscription_du or 0))
            fi_payee = Decimal(str(ech.frais_inscription_paye or 0))
            t1_due = Decimal(str(ech.tranche_1_due or 0))
            t1_payee = Decimal(str(ech.tranche_1_payee or 0))
            t2_due = Decimal(str(ech.tranche_2_due or 0))
            t2_payee = Decimal(str(ech.tranche_2_payee or 0))
            t3_due = Decimal(str(ech.tranche_3_due or 0))
            t3_payee = Decimal(str(ech.tranche_3_payee or 0))

            type_nom = (getattr(paiement.type_paiement, 'nom', '') or '').strip().lower()

            remaining = montant
            changed = False

            # 1) Inscription d'abord si due, quel que soit le type (règle métier: priorité à l'inscription)
            manque_insc = max(_ZERO, fi_due - fi_payee)
            take = min(remaining, manque_insc)
            if take > _ZERO:
                ech.frais_inscription_paye = fi_payee + take
                remaining -= take
                fi_payee += take
                changed = True

            # Helper pour allocation séquentielle
            def alloc_seq(current_due, current_paid):
                nonlocal remaining, changed
                manque = max(_ZERO, Decimal(str(current_due or 0)) - Decimal(str(current_paid or 0)))
                if manque <= _ZERO or remaining <= _ZERO:
                    return _ZERO
                take_local = min(remaining, manque)
                remaining -= take_local
                changed = True
                return take_local

            # 2) Mode proportionnel si 'annuel' est indiqué (après inscription)
            if 'annuel' in type_nom and remaining > _ZERO:
                r1 = max(_ZERO, t1_due - t1_payee)
                r2 = max(_ZERO, t2_due - t2_payee)
                r3 = max(_ZERO, t3_due - t3_payee)
                total_rest = r1 + r2 + r3
                if total_rest > _ZERO:
                    a_repartir = min(remaining, total_rest)

                    from decimal import ROUND_DOWN
                    p1 = (a_repartir * r1 / total_rest).quantize(Decimal('1'), rounding=ROUND_DOWN) if r1 > _ZERO else _ZERO
                    p2 = (a_repartir * r2 / total_rest).quantize(Decimal('1'), rounding=ROUND_DOWN) if r2 > _ZERO else _ZERO
                    p3 = a_repartir - (p1 + p2)

                    take1 = min(p1, r1) if p1 > _ZERO else _ZERO
                    take2 = min(p2, r2) if p2 > _ZERO else _ZERO
                    take3 = min(p3, r3) if p3 > _ZERO else _ZERO

                    # Redistribuer le reste après plafonnement
                    reste_cap = a_repartir - (take1 + take2 + take3)
                    if reste_cap > _ZERO:
                        for idx, (tk, rk) in enumerate([(take1, r1), (take2, r2), (take3, r3)]):
                            dispo = rk - tk
                            if dispo > _ZERO and reste_cap > _ZERO:
                                extra = min(reste_cap, dispo)
                                if idx == 0:
                                    take1 += extra
                                elif idx == 1:
                                    take2 += extra
                                else:
                                    take3 += extra
                                reste_cap -= extra

                    if take1 > _ZERO:
                        ech.tranche_1_payee = t1_payee + take1
                        t1_payee += take1
                        changed = True
                    if take2 > _ZERO:
                        ech.tranche_2_payee = t2_payee + take2
                        t2_payee += take2
                        changed = True
                    if take3 > _ZERO:
                        ech.tranche_3_payee = t3_payee + take3
                        t3_payee += take3
                        changed = True

                    remaining -= (take1 + take2 + take3)

            # 3) Allocation séquentielle pour le reste (ou si non-annuel)
            if remaining > _ZERO:
                take = alloc_seq(t1_due, t1_payee)
                if take > _ZERO:
                    ech.tranche_1_payee = t1_payee + take
                    t1_payee += take

            if remaining > _ZERO:
                take = alloc_seq(t2_due, t2_payee)
                if take > _ZERO:
                    ech.tranche_2_payee = t2_payee + take
                    t2_payee += take

            if remaining > _ZERO:
                take = alloc_seq(t3_due, t3_payee)
                if take > _ZERO:
                    ech.tranche_3_payee = t3_payee + take
                    t3_payee += take

            # Sauvegarder seulement si des changements ont été effectués
            if changed:
                ech.save()

                # Mettre à jour le statut/global de l'échéancier après allocation
                try:
                    _auto_validate_echeancier_for_eleve(eleve)
                except Exception:
                    logging.getLogger(__name__).exception("Auto-validation après allocation")

            # Log pour debug si montant non alloué
            if remaining > _ZERO:
                logging.getLogger(__name__).warning(
                    "Allocation incomplète: %s GNF non alloués pour paiement %s "
                    "(élève %s, type '%s')",
                    remaining, paiement.id, eleve.id, type_nom
                )

    except Exception:
        logging.getLogger(__name__).exception("Erreur allocation paiement -> échéancier")

def _allocate_combined_payment(paiement: "Paiement", echeancier: "EcheancierPaiement" = None) -> None:
    """CompatibilitÃ© avec les anciens tests et appels internes."""
    _allocate_payment_to_echeancier(paiement)


def _sum_validated_payments_and_remises(eleve):
    """Retourne (paiements_valides, remises_valides) sans double comptage SQL."""
    paiement_total = (
        Paiement.objects
        .filter(eleve=eleve, statut='VALIDE')
        .aggregate(total=Sum('montant'))
        .get('total') or 0
    )
    remise_total = (
        PaiementRemise.objects
        .filter(paiement__eleve=eleve, paiement__statut='VALIDE')
        .aggregate(total=Sum('montant_remise'))
        .get('total') or 0
    )
    return int(paiement_total or 0), int(remise_total or 0)


def _auto_validate_echeancier_for_eleve(eleve: "Eleve") -> None:
    """Synchronise l'échéancier de l'élève avec les paiements VALIDÉS avant impression du reçu.

    Règles conservatrices:
    - Si la somme des paiements validés + remises couvre le total dû -> statut = PAYE_COMPLET
      et on aligne les champs *_payee sur les *_due pour cohérence d'affichage.
    - Si couverture = 0 -> statut = A_PAYER (pas d'allocation détaillée effectuée ici)
    - Sinon -> statut = PAYE_PARTIEL (sans répartir finement par tranche)

    Cette fonction évite les incohérences si l'allocation manuelle par tranche a été oubliée.
    """
    try:
        # Récupérer l'échéancier (sans exception si absent)
        echeancier = getattr(eleve, 'echeancier', None)
        if echeancier is None:
            echeancier = EcheancierPaiement.objects.filter(eleve=eleve).first()
        if not echeancier:
            return

        # Totaux dus
        total_du = int((echeancier.frais_inscription_du or 0)
                       + (echeancier.tranche_1_due or 0)
                       + (echeancier.tranche_2_due or 0)
                       + (echeancier.tranche_3_due or 0))

        # Paiements validés et remises appliquées sur des paiements
        sum_montant, sum_remises = _sum_validated_payments_and_remises(eleve)

        couverture = max(0, sum_montant + sum_remises)

        # Déterminer le nouveau statut avec gestion du retard
        # Calcul de l'exigible (sommes dont la date d'échéance est passée ou aujourd'hui)
        from django.utils import timezone as _tz
        today = _tz.localdate() if hasattr(_tz, 'localdate') else date.today()
        exigible = 0
        if echeancier.date_echeance_inscription and echeancier.date_echeance_inscription <= today:
            exigible += int(echeancier.frais_inscription_du or 0)
        if echeancier.date_echeance_tranche_1 and echeancier.date_echeance_tranche_1 <= today:
            exigible += int(echeancier.tranche_1_due or 0)
        if echeancier.date_echeance_tranche_2 and echeancier.date_echeance_tranche_2 <= today:
            exigible += int(echeancier.tranche_2_due or 0)
        if echeancier.date_echeance_tranche_3 and echeancier.date_echeance_tranche_3 <= today:
            exigible += int(echeancier.tranche_3_due or 0)

        # Allocation conservatrice sur les tranches SANS jamais réduire l'existant
        # 1) total déjà indiqué comme payé dans l'échéancier
        old_insc = int(echeancier.frais_inscription_paye or 0)
        old_t1 = int(echeancier.tranche_1_payee or 0)
        old_t2 = int(echeancier.tranche_2_payee or 0)
        old_t3 = int(echeancier.tranche_3_payee or 0)
        current_total_paid = max(0, old_insc + old_t1 + old_t2 + old_t3)

        # 2) Incrément à répartir basé sur la couverture réelle
        increment = max(0, couverture - current_total_paid)
        remaining = increment

        def _alloc(due: int, paid: int, remaining_local: int):
            due_i = int(due or 0)
            paid_i = int(paid or 0)
            room = max(0, due_i - paid_i)
            take = min(room, max(0, int(remaining_local)))
            return paid_i + take, remaining_local - take

        changed = False
        if remaining > 0:
            # Ordre: inscription -> T1 -> T2 -> T3
            new_insc, remaining = _alloc(echeancier.frais_inscription_du, old_insc, remaining)
            new_t1, remaining = _alloc(echeancier.tranche_1_due, old_t1, remaining)
            new_t2, remaining = _alloc(echeancier.tranche_2_due, old_t2, remaining)
            new_t3, remaining = _alloc(echeancier.tranche_3_due, old_t3, remaining)

            if new_insc != old_insc:
                echeancier.frais_inscription_paye = new_insc
                changed = True
            if new_t1 != old_t1:
                echeancier.tranche_1_payee = new_t1
                changed = True
            if new_t2 != old_t2:
                echeancier.tranche_2_payee = new_t2
                changed = True
            if new_t3 != old_t3:
                echeancier.tranche_3_payee = new_t3
                changed = True

        # Somme payée effective bornée au total dû
        paye_effectif = min(couverture, total_du)

        if total_du <= 0:
            new_statut = 'PAYE_COMPLET'
        elif paye_effectif >= total_du:
            new_statut = 'PAYE_COMPLET'
        elif exigible > 0 and paye_effectif < exigible:
            new_statut = 'EN_RETARD'
        elif paye_effectif <= 0:
            new_statut = 'A_PAYER'
        else:
            new_statut = 'PAYE_PARTIEL'

        # Appliquer le statut et éventuellement aligner les montants payés si totalement soldé
        # 'changed' peut déjà être True si allocation ci-dessus a modifié des champs
        if echeancier.statut != new_statut:
            echeancier.statut = new_statut
            changed = True
        if new_statut == 'PAYE_COMPLET':
            # Aligner les montants payés pour refléter le soldé complet
            if echeancier.frais_inscription_paye != echeancier.frais_inscription_du:
                echeancier.frais_inscription_paye = echeancier.frais_inscription_du
                changed = True
            if echeancier.tranche_1_payee != echeancier.tranche_1_due:
                echeancier.tranche_1_payee = echeancier.tranche_1_due
                changed = True
            if echeancier.tranche_2_payee != echeancier.tranche_2_due:
                echeancier.tranche_2_payee = echeancier.tranche_2_due
                changed = True
            if echeancier.tranche_3_payee != echeancier.tranche_3_due:
                echeancier.tranche_3_payee = echeancier.tranche_3_due
                changed = True

        if changed:
            echeancier.save()
    except Exception:
        # Ne jamais bloquer l'impression du reçu à cause de cette étape
        logging.getLogger(__name__).exception("Erreur lors de la validation automatique de l'échéancier")


def _is_valid_twilio_request(request):
    """Valide la signature Twilio sur les webhooks entrants.

    Utilise le RequestValidator officiel de Twilio pour vérifier que la
    requête provient bien de Twilio (en-tête X-Twilio-Signature).
    Si TWILIO_AUTH_TOKEN n'est pas configuré, rejette toutes les requêtes.
    """
    from django.conf import settings as django_settings
    auth_token = getattr(django_settings, 'TWILIO_AUTH_TOKEN', '')
    if not auth_token:
        logging.getLogger(__name__).warning(
            "TWILIO_AUTH_TOKEN non configuré — requête Twilio rejetée"
        )
        return False
    try:
        from twilio.request_validator import RequestValidator
        validator = RequestValidator(auth_token)
        # Reconstituer l'URL complète telle que vue par Twilio
        url = request.build_absolute_uri()
        signature = request.META.get('HTTP_X_TWILIO_SIGNATURE', '')
        return validator.validate(url, request.POST.dict(), signature)
    except ImportError:
        logging.getLogger(__name__).error(
            "Le package twilio n'est pas installé — validation impossible"
        )
        return False
    except Exception:
        logging.getLogger(__name__).exception("Erreur lors de la validation Twilio")
        return False


@csrf_exempt
@require_http_methods(["POST"])
def twilio_inbound(request):
    """Réception des messages entrants Twilio (SMS/WhatsApp).
{{ ... }}
    Journalise les données utiles et répond 200.
    """
    if not _is_valid_twilio_request(request):
        return HttpResponse("Invalid signature", status=403)
    try:
        data = request.POST.dict()
    except Exception:
        data = {}
    # Champs utiles possibles: From, To, Body, SmsSid, MessageSid, WaId, NumMedia, etc.
    logging.getLogger(__name__).info("Twilio inbound message: %s", data)
    # Persist inbound message
    try:
        from_number = (data.get('From') or '').strip()
        to_number = (data.get('To') or '').strip()
        body = data.get('Body')
        message_sid = data.get('MessageSid') or data.get('SmsSid')
        wa_id = data.get('WaId')
        try:
            num_media = int(data.get('NumMedia') or 0)
        except Exception:
            num_media = 0
        channel = 'WHATSAPP' if from_number.lower().startswith('whatsapp:') else 'SMS'
        TwilioInboundMessage.objects.update_or_create(
            message_sid=message_sid,
            defaults={
                'channel': channel,
                'from_number': from_number,
                'to_number': to_number,
                'body': body,
                'wa_id': wa_id,
                'num_media': num_media,
                'raw_data': data,
            }
        )
    except Exception:
        logging.getLogger(__name__).exception("Erreur lors de l'enregistrement du message entrant Twilio")
    return JsonResponse({"status": "ok"})

@csrf_exempt
@require_http_methods(["POST"]) 
def twilio_status_callback(request):
    """Réception des callbacks de statut Twilio (optionnel).
    Journalise l'événement et répond 200.
    """
    if not _is_valid_twilio_request(request):
        return HttpResponse("Invalid signature", status=403)
    try:
        data = request.POST.dict()
    except Exception:
        data = {}
    logging.getLogger(__name__).info("Twilio status callback: %s", data)
    # Persist status update if MessageSid is present
    try:
        message_sid = data.get('MessageSid') or data.get('SmsSid')
        if message_sid:
            status = data.get('MessageStatus') or data.get('SmsStatus')
            error_code = data.get('ErrorCode')
            error_message = data.get('ErrorMessage')
            from django.utils import timezone as _tz
            obj, created = TwilioInboundMessage.objects.get_or_create(message_sid=message_sid, defaults={'raw_data': data})
            obj.delivery_status = status
            obj.error_code = str(error_code) if error_code is not None else obj.error_code
            obj.error_message = error_message or obj.error_message
            obj.status_updated_at = _tz.now()
            # Conserver dernières données brutes utiles
            try:
                merged = obj.raw_data or {}
                merged.update(data)
                obj.raw_data = merged
            except Exception:
                obj.raw_data = data
            obj.save()
    except Exception:
        logging.getLogger(__name__).exception("Erreur lors de l'enregistrement du status callback Twilio")
    return JsonResponse({"status": "ok"})

# ---------------------------------------------------------------
# Tableau de bord Paiements – statistiques réelles + listes
# ---------------------------------------------------------------

def _compute_stats(user):
    """Calcule les statistiques affichées sur le tableau de bord en respectant l'école de l'utilisateur (sauf admin).
    Retourne un dict: total_paiements_mois, nombre_paiements_mois, eleves_en_retard, paiements_en_attente.
    """
    try:
        from django.utils import timezone as _tz
        today = _tz.localdate() if hasattr(_tz, 'localdate') else date.today()
    except Exception:
        today = date.today()

    # Début du mois courant
    try:
        month_start = today.replace(day=1)
    except Exception:
        # fallback simple
        month_start = date(today.year, today.month, 1)

    # Somme des paiements validés sur le mois (DateField -> filtre inclusif par bornes)
    _qs_total_mois = Paiement.objects.filter(
        statut='VALIDE',
        date_paiement__gte=month_start,
        date_paiement__lte=today,
    )
    _qs_total_mois = filter_by_user_school(_qs_total_mois, user, 'eleve__classe__ecole')
    total_mois = (_qs_total_mois.aggregate(total=Sum('montant'))['total'] or 0)

    # Nombre de paiements (tous statuts) ce mois
    _qs_nb = Paiement.objects.filter(
        date_paiement__gte=month_start,
        date_paiement__lte=today,
    )
    _qs_nb = filter_by_user_school(_qs_nb, user, 'eleve__classe__ecole')
    nb_paiements_mois = _qs_nb.count()

    # Élèves en retard: calcul simplifié pour éviter les erreurs de colonnes manquantes
    # Utilise une approche plus robuste qui ne dépend pas des colonnes date_echeance_*
    try:
        # Méthode simplifiée: comparer total dû vs total payé
        _qs_retard = EcheancierPaiement.objects.annotate(
            total_du=F('frais_inscription_du') + F('tranche_1_due') + F('tranche_2_due') + F('tranche_3_due'),
            total_paye=F('frais_inscription_paye') + F('tranche_1_payee') + F('tranche_2_payee') + F('tranche_3_payee')
        ).filter(total_du__gt=F('total_paye'))
        _qs_retard = filter_by_user_school(_qs_retard, user, 'eleve__classe__ecole')
        eleves_retard_count = _qs_retard.count()
    except Exception:
        eleves_retard_count = 0

    # Paiements en attente
    _qs_attente = Paiement.objects.filter(statut='EN_ATTENTE')
    _qs_attente = filter_by_user_school(_qs_attente, user, 'eleve__classe__ecole')
    en_attente_count = _qs_attente.count()

    return {
        'total_paiements_mois': int(total_mois or 0),
        'nombre_paiements_mois': int(nb_paiements_mois or 0),
        'eleves_en_retard': int(eleves_retard_count or 0),
        'paiements_en_attente': int(en_attente_count or 0),
    }


@login_required
def tableau_bord_paiements(request):
    """Affiche le tableau de bord des paiements avec stats et listes utiles."""
    if not _template_exists('paiements/tableau_bord.html'):
        return HttpResponse('Tableau de bord paiements (template manquant)')

    stats = _compute_stats(request.user)

    # Paiements récents: derniers validés d'abord, sinon tout, sur 30 jours sinon fallback 20 derniers
    try:
        from django.utils import timezone as _tz
        today = _tz.localdate() if hasattr(_tz, 'localdate') else date.today()
    except Exception:
        today = date.today()

    try:
        from datetime import timedelta
        last_30 = today - timedelta(days=30)
    except Exception:
        last_30 = today

    paiements_recents_qs = (
        Paiement.objects
        .select_related('eleve', 'type_paiement', 'mode_paiement')
        .filter(date_paiement__gte=last_30)
        .order_by('-date_paiement', '-date_creation')
    )
    paiements_recents_qs = filter_by_user_school(paiements_recents_qs, request.user, 'eleve__classe__ecole')
    if paiements_recents_qs.count() == 0:
        paiements_recents_qs = (
            Paiement.objects
            .select_related('eleve', 'type_paiement', 'mode_paiement')
            .order_by('-date_paiement', '-date_creation')
        )
        paiements_recents_qs = filter_by_user_school(paiements_recents_qs, request.user, 'eleve__classe__ecole')
    paiements_recents = list(paiements_recents_qs[:20])

    # Top élèves en retard (montant de retard décroissant)
    exigible_expr = (
        Case(
            When(date_echeance_inscription__lte=today, then=F('frais_inscription_du')),
            default=Value(0),
            output_field=DecimalField(max_digits=10, decimal_places=0),
        )
        + Case(
            When(date_echeance_tranche_1__lte=today, then=F('tranche_1_due')),
            default=Value(0),
            output_field=DecimalField(max_digits=10, decimal_places=0),
        )
        + Case(
            When(date_echeance_tranche_2__lte=today, then=F('tranche_2_due')),
            default=Value(0),
            output_field=DecimalField(max_digits=10, decimal_places=0),
        )
        + Case(
            When(date_echeance_tranche_3__lte=today, then=F('tranche_3_due')),
            default=Value(0),
            output_field=DecimalField(max_digits=10, decimal_places=0),
        )
    )
    remises_expr = Coalesce(
        Sum('eleve__paiements__remises__montant_remise', filter=Q(eleve__paiements__statut='VALIDE')),
        Value(0),
        output_field=DecimalField(max_digits=10, decimal_places=0),
    )
    # Les remises ne compensent que les montants exigibles au jour J
    remises_applicables = Least(remises_expr, exigible_expr)
    paye_effectif_expr = (
        F('frais_inscription_paye') + F('tranche_1_payee') + F('tranche_2_payee') + F('tranche_3_payee')
        + remises_applicables
    )
    retard_expr = ExpressionWrapper(exigible_expr - paye_effectif_expr, output_field=DecimalField(max_digits=10, decimal_places=0))
    eleves_en_retard = (
        EcheancierPaiement.objects
        .select_related('eleve', 'eleve__classe', 'eleve__classe__ecole')
        .annotate(retard_db=retard_expr)
        .filter(retard_db__gt=0)
    )
    eleves_en_retard = filter_by_user_school(eleves_en_retard, request.user, 'eleve__classe__ecole').order_by('-retard_db')[:10]

    ecole_for_annee = user_school(request.user) if not user_is_admin(request.user) else None
    annee_active = get_annee_active(request, ecole_for_annee) if ecole_for_annee else None
    echeanciers_direction_qs = (
        EcheancierPaiement.objects
        .select_related('eleve', 'eleve__classe', 'eleve__classe__ecole')
        .annotate(
            remises_valides=Coalesce(
                Sum('eleve__paiements__remises__montant_remise', filter=Q(eleve__paiements__statut='VALIDE')),
                Value(0),
                output_field=DecimalField(max_digits=12, decimal_places=0),
            )
        )
    )
    echeanciers_direction_qs = filter_by_user_school(echeanciers_direction_qs, request.user, 'eleve__classe__ecole')
    if annee_active:
        echeanciers_direction_qs = echeanciers_direction_qs.filter(eleve__classe__annee_scolaire=annee_active)

    finance_direction = {
        'annee_active': annee_active or '',
        'eleves_suivis': 0,
        'eleves_soldes': 0,
        'eleves_non_soldes': 0,
        'total_du': 0,
        'total_encaisse': 0,
        'reste_a_encaisser': 0,
        'retard_total': 0,
        'prevision_30j': 0,
        'taux_recouvrement': 0,
        'taux_recouvrement_bar': 0,
    }
    classes_map = {}
    date_limite_prevision = today + timedelta(days=30)

    def _component_values(echeancier):
        return [
            (
                int(echeancier.frais_inscription_du or 0),
                int(echeancier.frais_inscription_paye or 0),
                echeancier.date_echeance_inscription,
            ),
            (
                int(echeancier.tranche_1_due or 0),
                int(echeancier.tranche_1_payee or 0),
                echeancier.date_echeance_tranche_1,
            ),
            (
                int(echeancier.tranche_2_due or 0),
                int(echeancier.tranche_2_payee or 0),
                echeancier.date_echeance_tranche_2,
            ),
            (
                int(echeancier.tranche_3_due or 0),
                int(echeancier.tranche_3_payee or 0),
                echeancier.date_echeance_tranche_3,
            ),
        ]

    for echeancier in echeanciers_direction_qs:
        remises = int(getattr(echeancier, 'remises_valides', 0) or 0)
        components = _component_values(echeancier)
        total_du = sum(due for due, _paye, _echeance in components)
        total_paye_brut = sum(paye for _due, paye, _echeance in components) + remises
        total_paye = min(total_du, total_paye_brut)
        reste = max(total_du - total_paye, 0)

        exigible = sum(due for due, _paye, echeance in components if echeance and echeance <= today)
        retard = max(exigible - total_paye, 0)
        prevision = sum(
            max(due - paye, 0)
            for due, paye, echeance in components
            if echeance and today < echeance <= date_limite_prevision
        )

        finance_direction['eleves_suivis'] += 1
        finance_direction['total_du'] += total_du
        finance_direction['total_encaisse'] += total_paye
        finance_direction['reste_a_encaisser'] += reste
        finance_direction['retard_total'] += retard
        finance_direction['prevision_30j'] += prevision
        if total_du > 0 and reste <= 0:
            finance_direction['eleves_soldes'] += 1
        elif total_du > 0:
            finance_direction['eleves_non_soldes'] += 1

        classe = echeancier.eleve.classe
        classe_id = classe.id if classe else 0
        classe_nom = classe.nom if classe else 'Sans classe'
        ecole_nom = classe.ecole.nom if classe and classe.ecole else ''
        row = classes_map.setdefault(classe_id, {
            'classe_id': classe_id,
            'classe_nom': classe_nom,
            'ecole_nom': ecole_nom,
            'eleves_count': 0,
            'total_du': 0,
            'total_encaisse': 0,
            'reste': 0,
            'retard': 0,
            'taux': 0,
        })
        row['eleves_count'] += 1
        row['total_du'] += total_du
        row['total_encaisse'] += total_paye
        row['reste'] += reste
        row['retard'] += retard

    if finance_direction['total_du'] > 0:
        finance_direction['taux_recouvrement'] = round(
            finance_direction['total_encaisse'] / finance_direction['total_du'] * 100,
            1,
        )
        finance_direction['taux_recouvrement_bar'] = min(
            100,
            max(0, int(round(finance_direction['taux_recouvrement']))),
        )

    classes_a_risque = []
    for row in classes_map.values():
        if row['total_du'] > 0:
            row['taux'] = round(row['total_encaisse'] / row['total_du'] * 100, 1)
        classes_a_risque.append(row)
    classes_a_risque = sorted(classes_a_risque, key=lambda item: (item['retard'], item['reste']), reverse=True)[:8]

    modes_encaissement_qs = Paiement.objects.filter(
        statut='VALIDE',
        date_paiement__gte=today.replace(day=1),
        date_paiement__lte=today,
    )
    modes_encaissement_qs = filter_by_user_school(modes_encaissement_qs, request.user, 'eleve__classe__ecole')
    modes_encaissement = (
        modes_encaissement_qs
        .values('mode_paiement__nom')
        .annotate(
            total=Coalesce(
                Sum('montant'),
                Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
                output_field=DecimalField(max_digits=12, decimal_places=0),
            ),
            nombre=Count('id'),
        )
        .order_by('-total')[:6]
    )

    context = {
        'titre_page': 'Tableau de bord des paiements',
        'stats': stats,
        'paiements_recents': paiements_recents,
        'eleves_en_retard': eleves_en_retard,
        'finance_direction': finance_direction,
        'classes_a_risque': classes_a_risque,
        'modes_encaissement': modes_encaissement,
    }
    return render(request, 'paiements/tableau_bord.html', context)

@login_required
def liste_paiements(request):
    """Liste des paiements optimisée avec cache intelligent et requêtes optimisées"""
    from ecole_moderne.performance_config import get_cached_or_set, CACHE_TTL, OptimizedQueryMixin
    
    titre_page = "Liste des paiements"
    q = (request.GET.get('q') or '').strip()
    statut = (request.GET.get('statut') or '').strip()
    annee_filtre = (request.GET.get('annee') or '').strip()
    page = request.GET.get('page') or 1

    # Cache de l'école utilisateur
    user_school_cache_key = f'user_school_{request.user.id}'
    user_school_obj = cache.get(user_school_cache_key)
    if user_school_obj is None and not user_is_admin(request.user):
        from utilisateurs.utils import user_school
        user_school_obj = user_school(request.user)
        if user_school_obj:
            cache.set(user_school_cache_key, user_school_obj, CACHE_TTL['user_school'])

    # Année scolaire active (utilisée par défaut si pas de filtre explicite)
    ecole_for_annee = user_school_obj or user_school(request.user) if not user_is_admin(request.user) else None
    annee_active = get_annee_active(request, ecole_for_annee) if ecole_for_annee else None

    # Queryset optimisé avec prefetch
    qs = OptimizedQueryMixin.get_optimized_paiements_queryset(user_school_obj)

    # Restreindre par école de l'utilisateur (sauf admin)
    if not user_is_admin(request.user) and user_school_obj:
        qs = qs.filter(eleve__classe__ecole=user_school_obj)

    # Filtre par année scolaire (via la classe de l'élève)
    if annee_filtre:
        qs = qs.filter(eleve__classe__annee_scolaire=annee_filtre)
    elif annee_active:
        # Par défaut, montrer uniquement l'année active
        qs = qs.filter(eleve__classe__annee_scolaire=annee_active)

    # Filtre recherche optimisé
    if q:
        qs = qs.filter(
            Q(numero_recu__icontains=q) |
            Q(reference_externe__icontains=q) |
            Q(observations__icontains=q) |
            Q(eleve__nom__icontains=q) |
            Q(eleve__prenom__icontains=q) |
            Q(eleve__matricule__icontains=q)
        )

    # Appliquer filtre par statut
    if statut:
        qs = qs.filter(statut=statut)

    # Calcul des totaux dynamiques (adaptés aux filtres en place)
    try:
        from django.utils import timezone as _tz
        today = _tz.localdate() if hasattr(_tz, 'localdate') else date.today()
    except Exception:
        today = date.today()
    try:
        month_start = today.replace(day=1)
    except Exception:
        month_start = date(today.year, today.month, 1)

    qs_effectif = qs
    qs_non_annule = qs_effectif.exclude(statut='ANNULE')

    total_paiements = qs_non_annule.count()
    montant_total = int(qs_non_annule.aggregate(total=Sum('montant'))['total'] or 0)
    montant_total_valide = int(qs_non_annule.filter(statut='VALIDE').aggregate(total=Sum('montant'))['total'] or 0)

    en_attente_qs = qs_effectif.filter(statut='EN_ATTENTE')
    total_en_attente = en_attente_qs.count()
    montant_en_attente = int(en_attente_qs.aggregate(total=Sum('montant'))['total'] or 0)

    ce_mois_qs = qs_non_annule.filter(date_paiement__gte=month_start, date_paiement__lte=today)
    total_ce_mois = ce_mois_qs.count()
    montant_ce_mois = int(ce_mois_qs.aggregate(total=Sum('montant'))['total'] or 0)
    montant_ce_mois_valide = int(ce_mois_qs.filter(statut='VALIDE').aggregate(total=Sum('montant'))['total'] or 0)

    # Montants annulés
    annule_qs = qs_effectif.filter(statut='ANNULE')
    total_annule = annule_qs.count()
    montant_annule = int(annule_qs.aggregate(total=Sum('montant'))['total'] or 0)

    # Reste à payer basé sur les élèves actifs (depuis les échéanciers)
    eleves_actifs_qs = Eleve.objects.filter(statut='ACTIF').select_related('classe', 'classe__ecole')
    eleves_actifs_qs = filter_by_user_school(eleves_actifs_qs, request.user, 'classe__ecole')
    if annee_filtre:
        eleves_actifs_qs = eleves_actifs_qs.filter(classe__annee_scolaire=annee_filtre)
    elif annee_active:
        eleves_actifs_qs = eleves_actifs_qs.filter(classe__annee_scolaire=annee_active)
    eleves_actifs_count = eleves_actifs_qs.count()
    eche_actifs_qs = EcheancierPaiement.objects.filter(eleve__in=eleves_actifs_qs)
    reste_a_payer_agg = eche_actifs_qs.aggregate(
        total_du=Coalesce(
            Sum(
                Coalesce(F('frais_inscription_du'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)), output_field=DecimalField(max_digits=12, decimal_places=0))
                + Coalesce(F('tranche_1_due'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)), output_field=DecimalField(max_digits=12, decimal_places=0))
                + Coalesce(F('tranche_2_due'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)), output_field=DecimalField(max_digits=12, decimal_places=0))
                + Coalesce(F('tranche_3_due'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)), output_field=DecimalField(max_digits=12, decimal_places=0)),
                output_field=DecimalField(max_digits=12, decimal_places=0),
            ),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        ),
        total_paye=Coalesce(
            Sum(
                Coalesce(F('frais_inscription_paye'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)), output_field=DecimalField(max_digits=12, decimal_places=0))
                + Coalesce(F('tranche_1_payee'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)), output_field=DecimalField(max_digits=12, decimal_places=0))
                + Coalesce(F('tranche_2_payee'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)), output_field=DecimalField(max_digits=12, decimal_places=0))
                + Coalesce(F('tranche_3_payee'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)), output_field=DecimalField(max_digits=12, decimal_places=0)),
                output_field=DecimalField(max_digits=12, decimal_places=0),
            ),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        ),
    )
    total_du_actifs = int(reste_a_payer_agg.get('total_du') or 0)
    total_paye_actifs = int(reste_a_payer_agg.get('total_paye') or 0)
    reste_a_payer = max(total_du_actifs - total_paye_actifs, 0)

    # Calculs supplémentaires: Dû scolarité net après remises + frais d'inscription (réels depuis l'échéancier)
    eleves_qs = Eleve.objects.select_related('classe', 'classe__ecole').all()
    eleves_qs = filter_by_user_school(eleves_qs, request.user, 'classe__ecole')
    if q:
        eleves_qs = eleves_qs.filter(
            Q(nom__icontains=q) | Q(prenom__icontains=q) | Q(matricule__icontains=q)
            | Q(classe__nom__icontains=q) | Q(classe__ecole__nom__icontains=q)
            | Q(paiements__numero_recu__icontains=q) | Q(paiements__reference_externe__icontains=q)
            | Q(paiements__observations__icontains=q)
        ).distinct()

    # Toujours compter les élèves restreints à l'école de l'utilisateur
    eleves_count = eleves_qs.count()

    eche_qs = EcheancierPaiement.objects.filter(eleve__in=eleves_qs)
    dues_sco_expr = (
        Coalesce(
            F('tranche_1_due'),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        )
        + Coalesce(
            F('tranche_2_due'),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        )
        + Coalesce(
            F('tranche_3_due'),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        )
    )
    remises_expr = Coalesce(
        Sum('eleve__paiements__remises__montant_remise', filter=Q(eleve__paiements__statut='VALIDE')),
        Value(0),
        output_field=DecimalField(max_digits=12, decimal_places=0),
    )
    # Annoter montant de réinscription dû par échéancier (en comparant à la grille tarifaire)
    try:
        reinsc_subq = GrilleTarifaire.objects.filter(
            ecole=OuterRef('eleve__classe__ecole'),
            niveau=OuterRef('eleve__classe__niveau'),
            annee_scolaire=OuterRef('annee_scolaire'),
        ).values('frais_reinscription')[:1]
        eche_qs = eche_qs.annotate(
            reinsc_due=Case(
                When(frais_inscription_du=Subquery(reinsc_subq), then=F('frais_inscription_du')),
                default=Value(0),
                output_field=DecimalField(max_digits=12, decimal_places=0),
            )
        )
    except Exception:
        # En cas d'erreur d'annotation, fallback sans champ dédié
        pass

    aggr_du = eche_qs.aggregate(
        dues_sco=Coalesce(
            Sum(dues_sco_expr, output_field=DecimalField(max_digits=12, decimal_places=0)),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        ),
        remises=remises_expr,
    )
    dues_sco_total = int(aggr_du.get('dues_sco') or 0)
    remises_total = int(aggr_du.get('remises') or 0)
    du_sco_net = max(dues_sco_total - remises_total, 0)
    frais_inscription_total = int(
        eche_qs.aggregate(
            total=Coalesce(
                Sum(F('frais_inscription_du'), output_field=DecimalField(max_digits=12, decimal_places=0)),
                Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
                output_field=DecimalField(max_digits=12, decimal_places=0),
            )
        ).get('total')
        or 0
    )
    du_global_net = du_sco_net + frais_inscription_total

    # Total global de réinscription (échéanciers dont le poste inscription correspond à la réinscription de la grille)
    try:
        reinsc_total = int(
            eche_qs.aggregate(
                total=Coalesce(
                    Sum(F('reinsc_due'), output_field=DecimalField(max_digits=12, decimal_places=0)),
                    Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
                    output_field=DecimalField(max_digits=12, decimal_places=0),
                )
            ).get('total') or 0
        )
    except Exception:
        reinsc_total = 0
    # Ratio global (éviter division par 0)
    try:
        reinsc_ratio = float(reinsc_total) / float(frais_inscription_total) * 100.0 if int(frais_inscription_total) > 0 else 0.0
    except Exception:
        reinsc_ratio = 0.0

    # Détail par école/classe (filtre libre appliqué aux élèves)
    detail_qs = (
        eche_qs
        .values(
            'eleve__classe__ecole__id', 'eleve__classe__ecole__nom',
            'eleve__classe__id', 'eleve__classe__nom'
        )
        .annotate(
            eleves_count=Count('eleve', distinct=True),
            dues_sco_sum=Coalesce(
                Sum(dues_sco_expr, output_field=DecimalField(max_digits=12, decimal_places=0)),
                Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
                output_field=DecimalField(max_digits=12, decimal_places=0),
            ),
            remises_sum=remises_expr,
            frais_insc_sum=Coalesce(
                Sum(F('frais_inscription_du'), output_field=DecimalField(max_digits=12, decimal_places=0)),
                Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
                output_field=DecimalField(max_digits=12, decimal_places=0),
            ),
            reinsc_sum=Coalesce(
                Sum(F('reinsc_due'), output_field=DecimalField(max_digits=12, decimal_places=0)),
                Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
                output_field=DecimalField(max_digits=12, decimal_places=0),
            ),
        )
        .order_by('eleve__classe__ecole__nom', 'eleve__classe__nom')
    )
    totaux_du_detail_classes = []
    for row in detail_qs:
        dues = int(row.get('dues_sco_sum') or 0)
        rem = int(row.get('remises_sum') or 0)
        net_sco = max(dues - rem, 0)
        cnt = int(row.get('eleves_count') or 0)
        insc = int(row.get('frais_insc_sum') or 0)
        reinsc = int(row.get('reinsc_sum') or 0)
        tot = net_sco + insc
        # Ratio par classe (réinscription / inscription)
        try:
            reinsc_pct = float(reinsc) / float(insc) * 100.0 if insc > 0 else 0.0
        except Exception:
            reinsc_pct = 0.0
        totaux_du_detail_classes.append({
            'ecole_id': row.get('eleve__classe__ecole__id'),
            'ecole_nom': row.get('eleve__classe__ecole__nom'),
            'classe_id': row.get('eleve__classe__id'),
            'classe_nom': row.get('eleve__classe__nom'),
            'eleves_count': cnt,
            'du_sco_net': net_sco,
            'frais_inscription_total': insc,
            'frais_reinscription_total': reinsc,
            'frais_reinscription_pct': reinsc_pct,
            'du_global_net': tot,
        })

    # Pagination
    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(page)

    context = {
        'titre_page': titre_page,
        'q': q,
        'statut': statut,
        'annee_filtre': annee_filtre or (annee_active or ''),
        'annee_active': annee_active or '',
        'paiements': page_obj.object_list,
        'page_obj': page_obj,
        # Totaux pour l'UI (utilisés par _paiements_resultats.html)
        'totaux': {
            'total_paiements': int(total_paiements or 0),
            'montant_total': int(montant_total or 0),
            'montant_total_valide': int(montant_total_valide or 0),
            'total_en_attente': int(total_en_attente or 0),
            'montant_en_attente': int(montant_en_attente or 0),
            'total_ce_mois': int(total_ce_mois or 0),
            'montant_ce_mois': int(montant_ce_mois or 0),
            'montant_ce_mois_valide': int(montant_ce_mois_valide or 0),
            'total_annule': int(total_annule or 0),
            'montant_annule': int(montant_annule or 0),
            'eleves_actifs_count': int(eleves_actifs_count or 0),
            'reste_a_payer': int(reste_a_payer or 0),
            'total_paye_actifs': int(total_paye_actifs or 0),
        },
        'totaux_du': {
            'eleves_count': int(eleves_count or 0),
            'du_sco_net': int(du_sco_net or 0),
            'frais_inscription_total': int(frais_inscription_total or 0),
            'frais_reinscription_total': int(reinsc_total or 0),
            'frais_reinscription_pct': reinsc_ratio,
            'du_global_net': int(du_global_net or 0),
        },
        'totaux_du_detail_classes': totaux_du_detail_classes,
        # Alerte relance: compte global des élèves en retard (filtré par école)
        'eleves_en_retard': _compute_stats(request.user).get('eleves_en_retard', 0),
    }

    # Réponse partielle pour les requêtes AJAX (utilisé par la recherche/pagination dynamique)
    try:
        is_ajax = request.headers.get('x-requested-with') == 'XMLHttpRequest'
    except Exception:
        is_ajax = False
    if is_ajax and _template_exists('paiements/_paiements_resultats.html'):
        return render(request, 'paiements/_paiements_resultats.html', context)

    template = 'paiements/liste_paiements.html' if _template_exists('paiements/liste_paiements.html') else None
    if template:
        return render(request, template, context)
    return HttpResponse('Liste des paiements')

@login_required
def export_recap_par_classe_excel(request):
    """Exporte en Excel le récapitulatif par classe (dû scolarité net, inscription, réinscription, total, ratio)."""
    # Reprendre la logique de liste_paiements pour les agrégations rapides
    q = request.GET.get('q', '').strip()
    eche_qs = EcheancierPaiement.objects.select_related(
        'eleve', 'eleve__classe', 'eleve__classe__ecole'
    )
    eche_qs = filter_by_user_school(eche_qs, request.user, 'eleve__classe__ecole')
    if q:
        eche_qs = eche_qs.filter(
            Q(eleve__nom__icontains=q) | Q(eleve__prenom__icontains=q) |
            Q(eleve__classe__nom__icontains=q) | Q(eleve__classe__ecole__nom__icontains=q)
        )

    # Calculer les montants dus une seule fois par échéancier. L'ancienne
    # agrégation joignait les échéanciers aux paiements puis aux remises :
    # MySQL pouvait refuser la requête et les jointures multipliaient les dus.
    echeanciers = list(eche_qs)
    eleve_ids = [echeancier.eleve_id for echeancier in echeanciers]

    grille_map = {
        (grille.ecole_id, grille.niveau, grille.annee_scolaire): grille.frais_reinscription or Decimal('0')
        for grille in GrilleTarifaire.objects.filter(
            ecole_id__in={e.eleve.classe.ecole_id for e in echeanciers}
        )
    } if echeanciers else {}

    remises_qs = PaiementRemise.objects.filter(
        paiement__statut='VALIDE', paiement__eleve_id__in=eleve_ids
    )
    remises_qs = filter_by_user_school(
        remises_qs, request.user, 'paiement__eleve__classe__ecole'
    )
    remises_par_classe = {
        row['paiement__eleve__classe_id']: row['total'] or Decimal('0')
        for row in remises_qs.values('paiement__eleve__classe_id').annotate(total=Sum('montant_remise'))
    }

    details = {}
    for echeancier in echeanciers:
        classe = echeancier.eleve.classe
        row = details.setdefault(classe.id, {
            'ecole': classe.ecole.nom,
            'classe': classe.nom,
            'eleves_count': 0,
            'dues_sco_sum': Decimal('0'),
            'frais_insc_sum': Decimal('0'),
            'reinsc_sum': Decimal('0'),
        })
        row['eleves_count'] += 1
        row['dues_sco_sum'] += (
            (echeancier.tranche_1_due or 0)
            + (echeancier.tranche_2_due or 0)
            + (echeancier.tranche_3_due or 0)
        )
        frais_inscription = echeancier.frais_inscription_du or Decimal('0')
        row['frais_insc_sum'] += frais_inscription
        frais_reinscription = grille_map.get(
            (classe.ecole_id, classe.niveau, echeancier.annee_scolaire), Decimal('0')
        )
        if frais_reinscription and frais_inscription == frais_reinscription:
            row['reinsc_sum'] += frais_inscription

    detail_rows = sorted(details.items(), key=lambda item: (item[1]['ecole'], item[1]['classe']))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Récap par classe'

    headers = [
        'École', 'Classe', '# Élèves',
        'Dû scolarité net', 'Inscription', 'Réinscription', 'Réinscription %', 'Total dû net'
    ]
    ws.append(headers)

    for classe_id, row in detail_rows:
        dues = int(row.get('dues_sco_sum') or 0)
        rem = int(remises_par_classe.get(classe_id, 0) or 0)
        net_sco = max(dues - rem, 0)
        insc = int(row.get('frais_insc_sum') or 0)
        reinsc = int(row.get('reinsc_sum') or 0)
        tot = net_sco + insc
        pct = (reinsc / insc * 100.0) if insc > 0 else 0.0
        ws.append([
            row.get('ecole'),
            row.get('classe'),
            int(row.get('eleves_count') or 0),
            net_sco, insc, reinsc, round(pct, 2), tot
        ])

    # Styles simples
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = 22

    from django.http import HttpResponse as DjHttp
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    resp = DjHttp(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename="recap_par_classe.xlsx"'
    return resp

@login_required
@require_school_object(Paiement, pk_kwarg='paiement_id', field_path='eleve__classe__ecole')
def detail_paiement(request, paiement_id:int):
    """Affiche le détail d'un paiement.

    Contexte pour `templates/paiements/detail_paiement.html`:
      - titre_page: str
      - paiement: instance `Paiement`
      - is_admin: bool
      - user_permissions: dict avec `can_validate_payments`
    """
    paiement_qs = Paiement.objects.select_related(
        'eleve', 'type_paiement', 'mode_paiement',
        'eleve__classe', 'eleve__classe__ecole',
    )
    paiement_qs = filter_by_user_school(paiement_qs, request.user, 'eleve__classe__ecole')
    paiement = get_object_or_404(paiement_qs, pk=paiement_id)

    # Préparer les informations de permissions utilisées dans le template
    try:
        perms_ctx = get_user_permissions(request.user)
    except Exception:
        perms_ctx = {}

    # Total des remises appliquées sur ce paiement
    try:
        remises_total = (
            paiement.remises.aggregate(total=Sum('montant_remise')).get('total') or 0
        )
    except Exception:
        remises_total = 0

    # Déterminer si l'utilisateur est comptable pour l'affichage UI (les actions restent protégées côté serveur)
    try:
        role_user = getattr(getattr(request.user, 'profil', None), 'role', None)
        is_comptable_flag = (role_user == 'COMPTABLE')
    except Exception:
        is_comptable_flag = False

    context = {
        'titre_page': f"Détail du paiement #{paiement.id}",
        'paiement': paiement,
        'is_admin': user_is_admin(request.user) if request.user.is_authenticated else False,
        'user_permissions': perms_ctx,
        'is_comptable': is_comptable_flag,
        'remises_total': int(remises_total or 0),
    }
    return render(request, 'paiements/detail_paiement.html', context)

@login_required
def ajouter_paiement(request, eleve_id:int=None):
    """Créer un paiement.
    - GET: affiche le formulaire `templates/paiements/form_paiement.html`
    - POST: enregistre le paiement en statut EN_ATTENTE
    """
    titre_page = "Ajouter un paiement"
    action = "Enregistrer"

    eleve = None
    initial = {}
    if eleve_id:
        eleve_qs = Eleve.objects.select_related('classe', 'classe__ecole')
        eleve_qs = filter_by_user_school(eleve_qs, request.user, 'classe__ecole')
        eleve = get_object_or_404(eleve_qs, pk=eleve_id)
        initial['eleve'] = eleve

    if request.method == 'POST':
        form = PaiementForm(request.POST)
        if form.is_valid():
            # Pré-valider la cohérence métier avant d'enregistrer
            paiement: Paiement = form.save(commit=False)

            # Vérifier que l'élève du paiement est bien dans l'école de l'utilisateur (sauf admin)
            if not user_is_admin(request.user):
                try:
                    ecole_user = user_school(request.user)
                    ecole_pmt = getattr(getattr(getattr(paiement.eleve, 'classe', None), 'ecole', None), 'id', None)
                    if ecole_user is None or (getattr(ecole_user, 'id', None) != ecole_pmt):
                        messages.error(request, "Accès refusé: cet élève n'appartient pas à votre école.")
                        return redirect('paiements:liste_paiements')
                except Exception:
                    messages.error(request, "Accès refusé: impossible de valider l'appartenance de l'élève à votre école.")
                    return redirect('paiements:liste_paiements')

            # Récupérer/assurer l'échéancier de l'élève pour les contrôles
            try:
                ech = getattr(paiement.eleve, 'echeancier', None)
            except Exception:
                ech = None
            if not ech:
                try:
                    ech = ensure_echeancier_for_eleve(
                        paiement.eleve,
                        created_by=request.user if request.user.is_authenticated else None,
                    )
                except Exception:
                    ech = None

            # Si on ne parvient pas à obtenir un échéancier, on empêche un enregistrement potentiellement incohérent
            if not ech:
                messages.error(request, "Impossible de récupérer l'échéancier de l'élève. Réessayez ou créez-le d'abord.")
                return render(request, 'paiements/form_paiement.html', {
                    'titre_page': titre_page,
                    'action': action,
                    'form': form,
                    'eleve': eleve,
                })

            # 1) Validation du montant saisi vs montant du type de paiement
            type_nom = (getattr(paiement.type_paiement, 'nom', '') or '').strip().lower()
            try:
                fi_due = int(ech.frais_inscription_du or 0)
                fi_payee = int(ech.frais_inscription_paye or 0)
                t1_due = int(ech.tranche_1_due or 0)
                t1_payee = int(ech.tranche_1_payee or 0)
                t2_due = int(ech.tranche_2_due or 0)
                t2_payee = int(ech.tranche_2_payee or 0)
                t3_due = int(ech.tranche_3_due or 0)
                t3_payee = int(ech.tranche_3_payee or 0)
            except Exception:
                fi_due = fi_payee = t1_due = t1_payee = t2_due = t2_payee = t3_due = t3_payee = 0

            # Validation du montant selon le type de paiement
            montant_saisi = int(paiement.montant or 0)
            montant_attendu = 0
            type_description = ""
            
            # IMPORTANT: évaluer d'abord les types combinés pour éviter que 'inscription' seul ne matche
            if ('inscription' in type_nom and 'annuel' in type_nom):
                # Frais d'inscription + Annuel (T1+T2+T3)
                montant_attendu = fi_due + t1_due + t2_due + t3_due
                type_description = "frais d'inscription + Annuel"
            elif ('inscription' in type_nom and ('tranche 1 + tranche 2' in type_nom or 'tranche1 + tranche2' in type_nom)):
                # Frais d'inscription + T1 + T2
                montant_attendu = fi_due + t1_due + t2_due
                type_description = "frais d'inscription + Tranche 1 + Tranche 2"
            elif ('inscription' in type_nom and ('tranche 1' in type_nom or '1ère tranche' in type_nom or '1ere tranche' in type_nom)):
                # Frais d'inscription + T1
                montant_attendu = fi_due + t1_due
                type_description = "frais d'inscription + Tranche 1"
            elif 'inscription' in type_nom:
                montant_attendu = fi_due
                type_description = "frais d'inscription"
            elif ('tranche 1 + tranche 2 + tranche 3' in type_nom or 'tranche1 + tranche2 + tranche3' in type_nom):
                montant_attendu = t1_due + t2_due + t3_due
                type_description = "Tranche 1 + Tranche 2 + Tranche 3"
            elif 'tranche 1 + tranche 2' in type_nom or 'tranche1 + tranche2' in type_nom:
                montant_attendu = t1_due + t2_due
                type_description = "Tranche 1 + Tranche 2"
            elif 'tranche 2 + tranche 3' in type_nom or 'tranche2 + tranche3' in type_nom:
                montant_attendu = t2_due + t3_due
                type_description = "Tranche 2 + Tranche 3"
            elif 'tranche 1' in type_nom or '1ère tranche' in type_nom or '1ere tranche' in type_nom:
                montant_attendu = t1_due
                type_description = "1ère tranche"
            elif 'tranche 2' in type_nom or '2ème tranche' in type_nom or '2eme tranche' in type_nom:
                montant_attendu = t2_due
                type_description = "2ème tranche"
            elif 'tranche 3' in type_nom or '3ème tranche' in type_nom or '3eme tranche' in type_nom:
                montant_attendu = t3_due
                type_description = "3ème tranche"
            
            # Vérifier si le montant correspond au type sélectionné
            if montant_attendu > 0 and montant_saisi != montant_attendu:
                # Paiements partiels: autoriser sans confirmation pour une tranche simple
                is_single_tranche = type_description in ["1ère tranche", "2ème tranche", "3ème tranche"]
                if montant_saisi < montant_attendu and is_single_tranche:
                    # Autoriser directement le paiement partiel de tranche
                    pass
                elif montant_saisi < montant_attendu:
                    # Demander confirmation pour paiement partiel (inscription ou types combinés)
                    confirmation_partiel = request.POST.get('confirmation_paiement_partiel')
                    if not confirmation_partiel:
                        # Message d'avertissement et demande de confirmation
                        from django.utils.safestring import mark_safe
                        message_html = mark_safe(
                            f'<span style="color: #dc3545; font-weight: bold; font-size: 1.1em;">'
                            f'⚠️ ATTENTION: Le montant saisi ({montant_saisi:,} GNF) est inférieur au montant standard '
                            f'pour {type_description} ({montant_attendu:,} GNF).</span><br>'
                            f'<strong>S\'agit-il d\'un paiement partiel ?</strong> Si oui, confirmez ci-dessous.'
                        )
                        messages.error(request, message_html)
                        return render(request, 'paiements/form_paiement.html', {
                            'titre_page': titre_page,
                            'action': action,
                            'form': form,
                            'eleve': eleve,
                            'montant_attendu': montant_attendu,
                            'montant_saisi': montant_saisi,
                            'type_description': type_description,
                            'show_partial_confirmation': True,
                        })
                else:
                    # Montant supérieur au montant standard: autoriser.
                    # Raison: pour les types combinés et même pour certaines tranches,
                    # on souhaite permettre que l'excédent soit alloué à la tranche suivante
                    # (allocation intelligente lors de la validation). Les contrôles
                    # anti-surpaiement par groupe et le plafond global empêcheront tout excès réel.
                    pass

            # 2) Bloquer si la tranche ciblée est déjà soldée ou risque de sur-paiement
            
            # Vérification pour l'inscription (seulement si type = inscription seule, pas combiné)
            if ('inscription' in type_nom) and not (
                'tranche' in type_nom or 'annuel' in type_nom
            ):
                if (fi_due > 0) and (fi_payee >= fi_due):
                    from django.utils.safestring import mark_safe
                    message_html = mark_safe(
                        f'<span style="color: #dc3545; font-weight: bold; font-size: 1.1em;">'
                        f'❌ ERREUR: L\'inscription est déjà totalement payée pour cet élève.</span><br>'
                        f'<strong>Montant dû:</strong> {fi_due:,} GNF | <strong>Déjà payé:</strong> {fi_payee:,} GNF<br>'
                        f'<strong>Aucune somme supplémentaire n\'est autorisée pour l\'inscription.</strong>'
                    )
                    messages.error(request, message_html)
                    return render(request, 'paiements/form_paiement.html', {
                        'titre_page': titre_page,
                        'action': action,
                        'form': form,
                        'eleve': eleve,
                    })
                elif (fi_due > 0) and ((fi_payee + montant_saisi) > fi_due):
                    from django.utils.safestring import mark_safe
                    sur_paiement = (fi_payee + montant_saisi) - fi_due
                    montant_max = fi_due - fi_payee
                    message_html = mark_safe(
                        f'<span style="color: #dc3545; font-weight: bold; font-size: 1.1em;">'
                        f'❌ ERREUR: Sur-paiement détecté pour l\'inscription!</span><br>'
                        f'<strong>Montant dû:</strong> {fi_due:,} GNF | <strong>Déjà payé:</strong> {fi_payee:,} GNF<br>'
                        f'<strong>Montant saisi:</strong> {montant_saisi:,} GNF | <strong>Sur-paiement:</strong> {sur_paiement:,} GNF<br>'
                        f'<strong>Montant maximum autorisé:</strong> {montant_max:,} GNF'
                    )
                    messages.error(request, message_html)
                    return render(request, 'paiements/form_paiement.html', {
                        'titre_page': titre_page,
                        'action': action,
                        'form': form,
                        'eleve': eleve,
                    })
            
            # Vérification pour Tranche 1 + Tranche 2
            elif ('tranche 1 + tranche 2' in type_nom or 'tranche1 + tranche2' in type_nom):
                # Vérifier si les deux tranches sont complètement soldées
                if ((t1_due > 0) and (t1_payee >= t1_due)) and ((t2_due > 0) and (t2_payee >= t2_due)):
                    from django.utils.safestring import mark_safe
                    message_html = mark_safe(
                        f'<span style="color: #dc3545; font-weight: bold; font-size: 1.1em;">'
                        f'❌ ERREUR: Les tranches 1 et 2 sont déjà totalement payées pour cet élève.</span><br>'
                        f'<strong>Tranche 1 - Dû:</strong> {t1_due:,} GNF | <strong>Payé:</strong> {t1_payee:,} GNF<br>'
                        f'<strong>Tranche 2 - Dû:</strong> {t2_due:,} GNF | <strong>Payé:</strong> {t2_payee:,} GNF'
                    )
                    messages.error(request, message_html)
                    return render(request, 'paiements/form_paiement.html', {
                        'titre_page': titre_page,
                        'action': action,
                        'form': form,
                        'eleve': eleve,
                    })
                # Vérifier uniquement les sur-paiements (pas les paiements partiels)
                elif ((t1_due + t2_due) > 0) and (((t1_payee + t2_payee) + montant_saisi) > (t1_due + t2_due)):
                    from django.utils.safestring import mark_safe
                    total_paye = t1_payee + t2_payee
                    total_du = t1_due + t2_due
                    sur_paiement = (total_paye + montant_saisi) - total_du
                    montant_max = total_du - total_paye
                    message_html = mark_safe(
                        f'<span style="color: #dc3545; font-weight: bold; font-size: 1.1em;">'
                        f'❌ ERREUR: Sur-paiement détecté pour Tranche 1 + Tranche 2!</span><br>'
                        f'<strong>Total dû (T1+T2):</strong> {total_du:,} GNF | <strong>Déjà payé:</strong> {total_paye:,} GNF<br>'
                        f'<strong>Montant saisi:</strong> {montant_saisi:,} GNF | <strong>Sur-paiement:</strong> {sur_paiement:,} GNF<br>'
                        f'<strong>Montant maximum autorisé:</strong> {montant_max:,} GNF'
                    )
                    messages.error(request, message_html)
                    return render(request, 'paiements/form_paiement.html', {
                        'titre_page': titre_page,
                        'action': action,
                        'form': form,
                        'eleve': eleve,
                    })
            
            # Vérification pour Tranche 2 + Tranche 3
            elif ('tranche 2 + tranche 3' in type_nom or 'tranche2 + tranche3' in type_nom):
                # Vérifier si les deux tranches sont complètement soldées
                if ((t2_due > 0) and (t2_payee >= t2_due)) and ((t3_due > 0) and (t3_payee >= t3_due)):
                    from django.utils.safestring import mark_safe
                    message_html = mark_safe(
                        f'<span style="color: #dc3545; font-weight: bold; font-size: 1.1em;">'
                        f'❌ ERREUR: Les tranches 2 et 3 sont déjà totalement payées pour cet élève.</span><br>'
                        f'<strong>Tranche 2 - Dû:</strong> {t2_due:,} GNF | <strong>Payé:</strong> {t2_payee:,} GNF<br>'
                        f'<strong>Tranche 3 - Dû:</strong> {t3_due:,} GNF | <strong>Payé:</strong> {t3_payee:,} GNF'
                    )
                    messages.error(request, message_html)
                    return render(request, 'paiements/form_paiement.html', {
                        'titre_page': titre_page,
                        'action': action,
                        'form': form,
                        'eleve': eleve,
                    })
                # Vérifier uniquement les sur-paiements (pas les paiements partiels)
                elif ((t2_due + t3_due) > 0) and (((t2_payee + t3_payee) + montant_saisi) > (t2_due + t3_due)):
                    from django.utils.safestring import mark_safe
                    total_paye = t2_payee + t3_payee
                    total_du = t2_due + t3_due
                    sur_paiement = (total_paye + montant_saisi) - total_du
                    montant_max = total_du - total_paye
                    message_html = mark_safe(
                        f'<span style="color: #dc3545; font-weight: bold; font-size: 1.1em;">'
                        f'❌ ERREUR: Sur-paiement détecté pour Tranche 2 + Tranche 3!</span><br>'
                        f'<strong>Total dû (T2+T3):</strong> {total_du:,} GNF | <strong>Déjà payé:</strong> {total_paye:,} GNF<br>'
                        f'<strong>Montant saisi:</strong> {montant_saisi:,} GNF | <strong>Sur-paiement:</strong> {sur_paiement:,} GNF<br>'
                        f'<strong>Montant maximum autorisé:</strong> {montant_max:,} GNF'
                    )
                    messages.error(request, message_html)
                    return render(request, 'paiements/form_paiement.html', {
                        'titre_page': titre_page,
                        'action': action,
                        'form': form,
                        'eleve': eleve,
                    })
            
            # Vérification pour Tranche 1 + Tranche 2 + Tranche 3
            elif ('tranche 1 + tranche 2 + tranche 3' in type_nom or 'tranche1 + tranche2 + tranche3' in type_nom):
                # Vérifier si les trois tranches sont complètement soldées
                if ((t1_due > 0) and (t1_payee >= t1_due)) and ((t2_due > 0) and (t2_payee >= t2_due)) and ((t3_due > 0) and (t3_payee >= t3_due)):
                    from django.utils.safestring import mark_safe
                    message_html = mark_safe(
                        f'<span style="color: #dc3545; font-weight: bold; font-size: 1.1em;">'
                        f'❌ ERREUR: Les tranches 1, 2 et 3 sont déjà totalement payées pour cet élève.</span><br>'
                        f'<strong>Tranche 1 - Dû:</strong> {t1_due:,} GNF | <strong>Payé:</strong> {t1_payee:,} GNF<br>'
                        f'<strong>Tranche 2 - Dû:</strong> {t2_due:,} GNF | <strong>Payé:</strong> {t2_payee:,} GNF<br>'
                        f'<strong>Tranche 3 - Dû:</strong> {t3_due:,} GNF | <strong>Payé:</strong> {t3_payee:,} GNF'
                    )
                    messages.error(request, message_html)
                    return render(request, 'paiements/form_paiement.html', {
                        'titre_page': titre_page,
                        'action': action,
                        'form': form,
                        'eleve': eleve,
                    })
                # Vérifier uniquement les sur-paiements (pas les paiements partiels)
                elif ((t1_due + t2_due + t3_due) > 0) and (((t1_payee + t2_payee + t3_payee) + montant_saisi) > (t1_due + t2_due + t3_due)):
                    from django.utils.safestring import mark_safe
                    total_paye = t1_payee + t2_payee + t3_payee
                    total_du = t1_due + t2_due + t3_due
                    sur_paiement = (total_paye + montant_saisi) - total_du
                    montant_max = total_du - total_paye
                    message_html = mark_safe(
                        f'<span style="color: #dc3545; font-weight: bold; font-size: 1.1em;">'
                        f'❌ ERREUR: Sur-paiement détecté pour Tranche 1 + Tranche 2 + Tranche 3!</span><br>'
                        f'<strong>Total dû (T1+T2+T3):</strong> {total_du:,} GNF | <strong>Déjà payé:</strong> {total_paye:,} GNF<br>'
                        f'<strong>Montant saisi:</strong> {montant_saisi:,} GNF | <strong>Sur-paiement:</strong> {sur_paiement:,} GNF<br>'
                        f'<strong>Montant maximum autorisé:</strong> {montant_max:,} GNF'
                    )
                    messages.error(request, message_html)
                    return render(request, 'paiements/form_paiement.html', {
                        'titre_page': titre_page,
                        'action': action,
                        'form': form,
                        'eleve': eleve,
                    })

            # Vérification pour la 1ère tranche
            elif ('tranche 1' in type_nom or '1ère tranche' in type_nom or '1ere tranche' in type_nom):
                # Bloquer uniquement si complètement soldée
                if (t1_due > 0) and (t1_payee >= t1_due):
                    from django.utils.safestring import mark_safe
                    message_html = mark_safe(
                        f'<span style="color: #dc3545; font-weight: bold; font-size: 1.1em;">'
                        f'❌ ERREUR: La 1ère tranche est déjà totalement payée pour cet élève.</span><br>'
                        f'<strong>Montant dû:</strong> {t1_due:,} GNF | <strong>Déjà payé:</strong> {t1_payee:,} GNF'
                    )
                    messages.error(request, message_html)
                    return render(request, 'paiements/form_paiement.html', {
                        'titre_page': titre_page,
                        'action': action,
                        'form': form,
                        'eleve': eleve,
                    })
                # Gestion intelligente des sur-paiements avec proposition de paiement partiel
                elif (t1_due > 0) and ((t1_payee + montant_saisi) > t1_due):
                    # Autoriser un dépassement si celui-ci correspond exactement au reste d'inscription à payer
                    fi_remaining = max(0, fi_due - fi_payee)
                    sur_paiement_calcule = (t1_payee + montant_saisi) - t1_due
                    
                    if fi_remaining > 0 and sur_paiement_calcule <= fi_remaining:
                        # On laisse passer: l'allocation automatique versera d'abord l'inscription puis T1
                        pass
                    else:
                        # Vérifier si l'utilisateur a confirmé le paiement partiel pour la tranche suivante
                        confirmation_partiel_suivant = request.POST.get('confirmation_paiement_partiel_suivant')
                        
                        if not confirmation_partiel_suivant:
                            # Calculer combien pourrait aller à la tranche suivante
                            t2_remaining = max(0, t2_due - t2_payee)
                            t3_remaining = max(0, t3_due - t3_payee)
                            
                            from django.utils.safestring import mark_safe
                            sur_paiement = (t1_payee + montant_saisi) - t1_due
                            montant_max_t1 = t1_due - t1_payee
                            
                            # Proposer l'allocation vers les tranches suivantes
                            suggestion_html = ""
                            if t2_remaining > 0:
                                montant_vers_t2 = min(sur_paiement, t2_remaining)
                                suggestion_html = f'<br><strong>💡 Suggestion:</strong> {montant_max_t1:,} GNF pour T1 + {montant_vers_t2:,} GNF comme acompte T2'
                            elif t3_remaining > 0:
                                montant_vers_t3 = min(sur_paiement, t3_remaining)
                                suggestion_html = f'<br><strong>💡 Suggestion:</strong> {montant_max_t1:,} GNF pour T1 + {montant_vers_t3:,} GNF comme acompte T3'
                            
                            message_html = mark_safe(
                                f'<span style="color: #f39c12; font-weight: bold; font-size: 1.1em;">'
                                f'⚠️ ATTENTION: Montant supérieur à la 1ère tranche!</span><br>'
                                f'<strong>Montant dû T1:</strong> {t1_due:,} GNF | <strong>Déjà payé T1:</strong> {t1_payee:,} GNF<br>'
                                f'<strong>Montant saisi:</strong> {montant_saisi:,} GNF | <strong>Excédent:</strong> {sur_paiement:,} GNF<br>'
                                f'<strong>Montant maximum T1:</strong> {montant_max_t1:,} GNF'
                                f'{suggestion_html}<br><br>'
                                f'<strong>Voulez-vous utiliser l\'excédent comme acompte sur une tranche suivante ?</strong>'
                            )
                            messages.warning(request, message_html)
                            return render(request, 'paiements/form_paiement.html', {
                                'titre_page': titre_page,
                                'action': action,
                                'form': form,
                                'eleve': eleve,
                                'show_partial_next_confirmation': True,
                                'montant_t1_max': montant_max_t1,
                                'excedent': sur_paiement,
                                't2_remaining': t2_remaining,
                                't3_remaining': t3_remaining,
                            })
                        else:
                            # L'utilisateur a confirmé, on laisse passer pour allocation intelligente
                            pass
            
            # Vérification pour la 2ème tranche
            elif ('tranche 2' in type_nom or '2ème tranche' in type_nom or '2eme tranche' in type_nom):
                # Bloquer uniquement si complètement soldée
                if (t2_due > 0) and (t2_payee >= t2_due):
                    from django.utils.safestring import mark_safe
                    message_html = mark_safe(
                        f'<span style="color: #dc3545; font-weight: bold; font-size: 1.1em;">'
                        f'❌ ERREUR: La 2ème tranche est déjà totalement payée pour cet élève.</span><br>'
                        f'<strong>Montant dû:</strong> {t2_due:,} GNF | <strong>Déjà payé:</strong> {t2_payee:,} GNF'
                    )
                    messages.error(request, message_html)
                    return render(request, 'paiements/form_paiement.html', {
                        'titre_page': titre_page,
                        'action': action,
                        'form': form,
                        'eleve': eleve,
                    })
                # Gestion intelligente des sur-paiements T2 avec proposition vers T3
                elif (t2_due > 0) and ((t2_payee + montant_saisi) > t2_due):
                    # Vérifier si l'utilisateur a confirmé le paiement partiel pour la tranche suivante
                    confirmation_partiel_suivant = request.POST.get('confirmation_paiement_partiel_suivant')
                    
                    if not confirmation_partiel_suivant:
                        # Calculer combien pourrait aller à la tranche suivante
                        t3_remaining = max(0, t3_due - t3_payee)
                        
                        from django.utils.safestring import mark_safe
                        sur_paiement = (t2_payee + montant_saisi) - t2_due
                        montant_max_t2 = t2_due - t2_payee
                        
                        # Proposer l'allocation vers T3 si disponible
                        if t3_remaining > 0:
                            montant_vers_t3 = min(sur_paiement, t3_remaining)
                            suggestion_html = f'<br><strong>💡 Suggestion:</strong> {montant_max_t2:,} GNF pour T2 + {montant_vers_t3:,} GNF comme acompte T3'
                            
                            message_html = mark_safe(
                                f'<span style="color: #f39c12; font-weight: bold; font-size: 1.1em;">'
                                f'⚠️ ATTENTION: Montant supérieur à la 2ème tranche!</span><br>'
                                f'<strong>Montant dû T2:</strong> {t2_due:,} GNF | <strong>Déjà payé T2:</strong> {t2_payee:,} GNF<br>'
                                f'<strong>Montant saisi:</strong> {montant_saisi:,} GNF | <strong>Excédent:</strong> {sur_paiement:,} GNF<br>'
                                f'<strong>Montant maximum T2:</strong> {montant_max_t2:,} GNF'
                                f'{suggestion_html}<br><br>'
                                f'<strong>Voulez-vous utiliser l\'excédent comme acompte sur la 3ème tranche ?</strong>'
                            )
                            messages.warning(request, message_html)
                            return render(request, 'paiements/form_paiement.html', {
                                'titre_page': titre_page,
                                'action': action,
                                'form': form,
                                'eleve': eleve,
                                'show_partial_next_confirmation': True,
                                'montant_t2_max': montant_max_t2,
                                'excedent': sur_paiement,
                                't3_remaining': t3_remaining,
                                'tranche_source': 'T2',
                            })
                        else:
                            # Aucune tranche suivante disponible, bloquer le sur-paiement
                            message_html = mark_safe(
                                f'<span style="color: #dc3545; font-weight: bold; font-size: 1.1em;">'
                                f'❌ ERREUR: Sur-paiement détecté pour la 2ème tranche!</span><br>'
                                f'<strong>Montant dû:</strong> {t2_due:,} GNF | <strong>Déjà payé:</strong> {t2_payee:,} GNF<br>'
                                f'<strong>Montant saisi:</strong> {montant_saisi:,} GNF | <strong>Sur-paiement:</strong> {sur_paiement:,} GNF<br>'
                                f'<strong>Montant maximum autorisé:</strong> {montant_max_t2:,} GNF<br>'
                                f'<em>Aucune tranche suivante disponible pour l\'excédent.</em>'
                            )
                            messages.error(request, message_html)
                            return render(request, 'paiements/form_paiement.html', {
                                'titre_page': titre_page,
                                'action': action,
                                'form': form,
                                'eleve': eleve,
                            })
                    else:
                        # L'utilisateur a confirmé, on laisse passer pour allocation intelligente
                        pass
            
            # Vérification pour la 3ème tranche
            elif ('tranche 3' in type_nom or '3ème tranche' in type_nom or '3eme tranche' in type_nom):
                # Bloquer uniquement si complètement soldée
                if (t3_due > 0) and (t3_payee >= t3_due):
                    from django.utils.safestring import mark_safe
                    message_html = mark_safe(
                        f'<span style="color: #dc3545; font-weight: bold; font-size: 1.1em;">'
                        f'❌ ERREUR: La 3ème tranche est déjà totalement payée pour cet élève.</span><br>'
                        f'<strong>Montant dû:</strong> {t3_due:,} GNF | <strong>Déjà payé:</strong> {t3_payee:,} GNF'
                    )
                    messages.error(request, message_html)
                    return render(request, 'paiements/form_paiement.html', {
                        'titre_page': titre_page,
                        'action': action,
                        'form': form,
                        'eleve': eleve,
                    })
                # Bloquer strictement les sur-paiements pour T3 (dernière tranche)
                elif (t3_due > 0) and ((t3_payee + montant_saisi) > t3_due):
                    from django.utils.safestring import mark_safe
                    sur_paiement = (t3_payee + montant_saisi) - t3_due
                    montant_max = t3_due - t3_payee
                    message_html = mark_safe(
                        f'<span style="color: #dc3545; font-weight: bold; font-size: 1.1em;">'
                        f'❌ ERREUR: Sur-paiement détecté pour la 3ème tranche!</span><br>'
                        f'<strong>Montant dû:</strong> {t3_due:,} GNF | <strong>Déjà payé:</strong> {t3_payee:,} GNF<br>'
                        f'<strong>Montant saisi:</strong> {montant_saisi:,} GNF | <strong>Sur-paiement:</strong> {sur_paiement:,} GNF<br>'
                        f'<strong>Montant maximum autorisé:</strong> {montant_max:,} GNF<br>'
                        f'<em>Aucune tranche suivante disponible pour l\'excédent.</em>'
                    )
                    messages.error(request, message_html)
                    return render(request, 'paiements/form_paiement.html', {
                        'titre_page': titre_page,
                        'action': action,
                        'form': form,
                        'eleve': eleve,
                    })

            # 2) Bloquer les sur-paiements par rapport au total annuel dû (incluant inscription + tranches)
            try:
                total_du = int((ech.frais_inscription_du or 0) + (ech.tranche_1_due or 0) + (ech.tranche_2_due or 0) + (ech.tranche_3_due or 0))
            except Exception:
                total_du = 0

            try:
                # Ne compter que les paiements VALIDE pour le plafond global afin
                # d'éviter de bloquer des saisies lorsque des paiements sont encore EN_ATTENTE.
                # Les contrôles de sur-paiement par tranche/groupes ci-dessus empêchent déjà
                # les excès au niveau détaillé.
                aggs = (
                    Paiement.objects
                    .filter(eleve=paiement.eleve, statut='VALIDE')
                    .aggregate(sum_montant=Sum('montant'))
                )
                deja_saisi = int(aggs.get('sum_montant') or 0)
            except Exception:
                deja_saisi = 0

            try:
                remises_valides = (
                    Paiement.objects
                    .filter(eleve=paiement.eleve, statut='VALIDE')
                    .aggregate(total=Sum('remises__montant_remise'))
                    .get('total') or 0
                )
                remises_valides = int(remises_valides)
            except Exception:
                remises_valides = 0

            montant_soumis = int(paiement.montant or 0)
            restant_global = max(0, total_du - max(0, deja_saisi + remises_valides))
            if montant_soumis > restant_global:
                # Message précis avec le plafond autorisé restant
                try:
                    montant_autorise = max(0, restant_global)
                except Exception:
                    montant_autorise = 0
                messages.error(
                    request,
                    f"Montant trop élevé: le reste total à payer pour cet élève est de {montant_autorise:,} GNF. Veuillez saisir un montant inférieur ou égal.",
                )
                return render(request, 'paiements/form_paiement.html', {
                    'titre_page': titre_page,
                    'action': action,
                    'form': form,
                    'eleve': eleve,
                })

            # Si tout est cohérent, on peut enregistrer
            with transaction.atomic():
                # Attacher l'utilisateur créateur si connecté
                if request.user.is_authenticated:
                    paiement.cree_par = request.user
                # Statut par défaut reste EN_ATTENTE (défini dans le modèle)
                paiement.save()
                # Auto-création de l'échéancier s'il n'existe pas, puis synchro/validation
                try:
                    _auto_validate_echeancier_for_eleve(paiement.eleve)
                except Exception:
                    logging.getLogger(__name__).exception("Auto-validation échéancier après enregistrement paiement")
            # Notifications: reçu paiement (WhatsApp + SMS) et, si inscription, confirmation d'inscription
            try:
                send_payment_receipt(paiement.eleve, paiement)
                type_nom = (getattr(paiement.type_paiement, 'nom', '') or '').strip().lower()
                if 'inscription' in type_nom:
                    send_enrollment_confirmation(paiement.eleve, paiement)
            except Exception:
                logging.getLogger(__name__).exception("Erreur lors de l'envoi des notifications Twilio")
            messages.success(request, "Paiement enregistré avec succès.")
            # Rediriger vers la page échéancier de l'élève
            return redirect('paiements:echeancier_eleve', eleve_id=paiement.eleve_id)
        else:
            messages.error(request, "Veuillez corriger les erreurs du formulaire.")
    else:
        form = PaiementForm(initial=initial)
        # Si l'élève est imposé, fixer la valeur initiale proprement
        if eleve:
            form.fields['eleve'].initial = eleve

    context = {
        'titre_page': titre_page,
        'action': action,
        'form': form,
        'eleve': eleve,
    }
    return render(request, 'paiements/form_paiement.html', context)

@login_required
@require_POST
@require_school_object(Paiement, pk_kwarg='paiement_id', field_path='eleve__classe__ecole')
def valider_paiement(request, paiement_id:int):
    """Valide un paiement en le passant au statut VALIDE.

    - Vérifie les permissions: admin ou can_validate_payments
    - Met à jour: statut, date_validation, valide_par, date_modification
    - Optionnel: tente d'allouer le paiement à l'échéancier si une fonction utilitaire existe
    - Notifie le responsable (WhatsApp/SMS) avec le reçu
    """
    paiement_qs = filter_by_user_school(
        Paiement.objects.select_related('type_paiement', 'mode_paiement', 'eleve', 'eleve__classe', 'eleve__classe__ecole'),
        request.user, 'eleve__classe__ecole'
    )
    paiement = get_object_or_404(paiement_qs, pk=paiement_id)

    # Contrôle serveur strict: seuls admin ou détenteurs de la permission explicite peuvent valider
    if not request.user.is_authenticated or not (user_is_admin(request.user) or has_permission(request.user, 'peut_valider_paiements')):
        messages.error(request, "Vous n'avez pas l'autorisation de valider ce paiement.")
        return redirect('paiements:detail_paiement', paiement_id=paiement.id)

    if paiement.statut == 'VALIDE':
        messages.info(request, "Ce paiement est déjà validé.")
        return redirect('paiements:detail_paiement', paiement_id=paiement.id)

    with transaction.atomic():
        paiement.statut = 'VALIDE'
        try:
            paiement.date_validation = timezone.now()
        except Exception:
            from django.utils import timezone as _tz
            paiement.date_validation = _tz.now()
        paiement.valide_par = request.user
        try:
            paiement.date_modification = timezone.now()
        except Exception:
            pass
        paiement.save()

        # Allocation intelligente à l'échéancier
        try:
            _allocate_payment_to_echeancier(paiement)
        except Exception:
            logging.getLogger(__name__).exception("Erreur lors de l'allocation du paiement à l'échéancier")

        # S'assurer que l'échéancier existe et synchroniser le statut (incl. EN_RETARD)
        try:
            ensure_echeancier_for_eleve(paiement.eleve, created_by=request.user if request.user.is_authenticated else None)
            _auto_validate_echeancier_for_eleve(paiement.eleve)
        except Exception:
            logging.getLogger(__name__).exception("Erreur ensure/auto-validate échéancier après validation du paiement")

    # Envoyer le reçu de paiement après validation
    try:
        send_payment_receipt(paiement.eleve, paiement)
    except Exception:
        logging.getLogger(__name__).exception("Erreur lors de l'envoi du reçu après validation")

    messages.success(request, "Paiement validé avec succès.")
    return redirect('paiements:detail_paiement', paiement_id=paiement.id)

@login_required
@require_POST
@require_school_object(Eleve, pk_kwarg='eleve_id', field_path='classe__ecole')
def relancer_eleve(request, eleve_id:int):
    """Crée une relance et envoie la notification (WhatsApp/SMS) au responsable.
    GET params optionnels:
      - canal: SMS | WHATSAPP (par défaut WHATSAPP)
      - message: texte personnalisé
    """
    eleve_qs = Eleve.objects.select_related('classe')
    eleve_qs = filter_by_user_school(eleve_qs, request.user, 'classe__ecole')
    eleve = get_object_or_404(eleve_qs, pk=eleve_id)
    canal = (request.POST.get('canal') or request.GET.get('canal') or 'WHATSAPP').upper()
    message_txt = (request.POST.get('message') or request.GET.get('message') or '').strip()
    next_url = (request.POST.get('next') or request.GET.get('next') or '').strip()

    # Solde estimé depuis l'échéancier
    try:
        echeancier = getattr(eleve, 'echeancier', None)
        solde_estime = echeancier.solde_restant if echeancier else 0
    except Exception:
        solde_estime = 0

    if not message_txt:
        classe_nom = eleve.classe.nom if eleve.classe else ''
        try:
            echeancier = getattr(eleve, 'echeancier', None)
            solde_txt = f"{int(echeancier.solde_restant or 0):,}".replace(",", " ") if echeancier else "0"
        except Exception:
            solde_txt = "0"
        message_txt = (
            f"Bonjour Cher Parent,\n\n"
            f"Nous vous rappelons que la situation financière de {eleve.nom_complet} "
            f"({eleve.matricule}) en classe {classe_nom} présente un reste à payer de {solde_txt} GNF.\n"
            "Merci de bien vouloir régulariser ou contacter l'administration.\n\n"
            "La Direction"
        )

    with transaction.atomic():
        relance = Relance.objects.create(
            eleve=eleve,
            canal=canal if canal in {c for c, _ in Relance.CANAL_CHOICES} else 'AUTRE',
            message=message_txt,
            statut='ENREGISTREE',
            solde_estime=solde_estime or 0,
            cree_par=request.user if request.user.is_authenticated else None,
        )
    try:
        send_relance_notification(relance)
        messages.success(request, "Relance créée et notification envoyée.")
    except Exception:
        logging.getLogger(__name__).exception("Erreur lors de l'envoi de la relance Twilio")
        messages.warning(request, "Relance créée mais l'envoi de la notification a échoué.")

    if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        return redirect(next_url)
    return redirect('paiements:echeancier_eleve', eleve_id=eleve.id)

@login_required
@require_POST
def envoyer_notifs_retards(request):
    """Envoie des notifications de retard aux responsables des élèves avec solde > 0.
    Action manuelle: GET uniquement, simple résumé via messages.
    """
    if not request.user.is_authenticated:
        return HttpResponse(status=403)
    # Optionnel: restreindre aux admins/permissions
    if not (user_is_admin(request.user) or can_view_reports(request.user)):
        return HttpResponse(status=403)

    # Annoter montant de retard au niveau DB: (exigible - (payé + remises))
    try:
        from django.utils import timezone as _tz
        today = _tz.localdate() if hasattr(_tz, 'localdate') else date.today()
    except Exception:
        today = date.today()

    exigible_expr = (
        Case(
            When(date_echeance_inscription__lte=today, then=F('frais_inscription_du')),
            default=Value(0),
            output_field=DecimalField(max_digits=10, decimal_places=0),
        )
        + Case(
            When(date_echeance_tranche_1__lte=today, then=F('tranche_1_due')),
            default=Value(0),
            output_field=DecimalField(max_digits=10, decimal_places=0),
        )
        + Case(
            When(date_echeance_tranche_2__lte=today, then=F('tranche_2_due')),
            default=Value(0),
            output_field=DecimalField(max_digits=10, decimal_places=0),
        )
        + Case(
            When(date_echeance_tranche_3__lte=today, then=F('tranche_3_due')),
            default=Value(0),
            output_field=DecimalField(max_digits=10, decimal_places=0),
        )
    )
    remises_expr = Coalesce(
        Sum('eleve__paiements__remises__montant_remise', filter=Q(eleve__paiements__statut='VALIDE')),
        Value(0),
        output_field=DecimalField(max_digits=10, decimal_places=0),
    )
    # Limiter l'effet des remises au montant actuellement exigible
    remises_applicables = Least(remises_expr, exigible_expr)
    paye_effectif_expr = (
        F('frais_inscription_paye') + F('tranche_1_payee') + F('tranche_2_payee') + F('tranche_3_payee')
        + remises_applicables
    )
    retard_expr = ExpressionWrapper(exigible_expr - paye_effectif_expr, output_field=DecimalField(max_digits=10, decimal_places=0))
    qs = (
        EcheancierPaiement.objects.select_related('eleve', 'eleve__classe')
        .annotate(retard=retard_expr)
        .filter(retard__gt=0)
    )
    qs = filter_by_user_school(qs, request.user, 'eleve__classe__ecole')
    envoyes = 0
    for ech in qs[:500]:  # sécurité: batch max 500
        try:
            send_retard_notification(ech.eleve, ech.retard)
            envoyes += 1
        except Exception:
            logging.getLogger(__name__).exception("Échec envoi retard pour %s", getattr(ech.eleve, 'nom_complet', 'eleve'))
            continue
    messages.info(request, f"Notifications de retard envoyées: {envoyes} (sur {qs.count()} éligibles)")
    # Rediriger vers relances ou tableau de bord
    return redirect('paiements:liste_relances')

@login_required
def liste_relances(request):
    """Liste des relances avec filtres et pagination."""
    titre_page = "Liste des relances"
    q = (request.GET.get('q') or '').strip()
    canal = (request.GET.get('canal') or '').strip().upper()
    statut = (request.GET.get('statut') or '').strip().upper()
    # Filtrer par élève si fourni (depuis le bouton Alertes de la liste des paiements)
    eleve_id = (request.GET.get('eleve_id') or '').strip()

    qs = (
        Relance.objects.select_related('eleve', 'eleve__classe')
        .order_by('-date_creation')
    )
    qs = filter_by_user_school(qs, request.user, 'eleve__classe__ecole')
    if q:
        qs = qs.filter(
            Q(eleve__nom__icontains=q)
            | Q(eleve__prenom__icontains=q)
            | Q(eleve__matricule__icontains=q)
            | Q(message__icontains=q)
        )
    if canal:
        qs = qs.filter(canal=canal)
    if statut:
        qs = qs.filter(statut=statut)
    if eleve_id:
        try:
            qs = qs.filter(eleve_id=int(eleve_id))
        except Exception:
            # Si la conversion échoue, on ignore le filtre pour ne pas casser la vue
            pass

    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get('page') or 1)

    context = {
        'titre_page': titre_page,
        'q': q,
        'canal': canal,
        'statut': statut,
        'eleve_id': eleve_id,
        'page_obj': page_obj,
    }
    template = 'paiements/relances.html' if _template_exists('paiements/relances.html') else None
    if template:
        return render(request, template, context)
    return HttpResponse('Liste des relances')

@login_required
@require_school_object(Eleve, pk_kwarg='eleve_id', field_path='classe__ecole')
def echeancier_eleve(request, eleve_id:int):
    """Affiche l'échéancier et l'historique des paiements d'un élève.

    Contexte fourni au template `templates/paiements/echeancier_eleve.html`:
      - titre_page: Titre de la page
      - eleve: instance d'`Eleve`
      - echeancier: instance d'`EcheancierPaiement` (ou None si non créé)
      - paiements: queryset des `Paiement` liés à l'élève (ordonnés récents d'abord)
      - today: date du jour (timezone-aware localdate si dispo)
    """
    # Récupération de l'élève avec sa classe/école pour l'en-tête (restreint par école)
    eleve_qs = Eleve.objects.select_related('classe', 'classe__ecole')
    eleve_qs = filter_by_user_school(eleve_qs, request.user, 'classe__ecole')
    eleve = get_object_or_404(eleve_qs, pk=eleve_id)

    # Échéancier (peut ne pas exister encore)
    try:
        echeancier = getattr(eleve, 'echeancier', None)
    except Exception:
        echeancier = None

    # Historique des paiements (les plus récents d'abord)
    paiements = (
        Paiement.objects
        .select_related('type_paiement', 'mode_paiement')
        .filter(eleve=eleve)
        .order_by('-date_paiement', '-date_creation')
    )

    # Date du jour pour l'affichage des retards
    try:
        from django.utils import timezone as _tz
        today = _tz.localdate() if hasattr(_tz, 'localdate') else date.today()
    except Exception:
        today = date.today()

    finance_eleve = None
    if echeancier:
        remises_total = int(
            PaiementRemise.objects
            .filter(paiement__eleve=eleve, paiement__statut='VALIDE')
            .aggregate(total=Coalesce(
                Sum('montant_remise'),
                Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
                output_field=DecimalField(max_digits=12, decimal_places=0),
            ))
            .get('total') or 0
        )

        postes = [
            {
                'code': 'INSCRIPTION',
                'libelle': "Frais d'inscription",
                'du': int(echeancier.frais_inscription_du or 0),
                'paye': int(echeancier.frais_inscription_paye or 0),
                'echeance': echeancier.date_echeance_inscription,
            },
            {
                'code': 'TRANCHE_1',
                'libelle': '1ère tranche',
                'du': int(echeancier.tranche_1_due or 0),
                'paye': int(echeancier.tranche_1_payee or 0),
                'echeance': echeancier.date_echeance_tranche_1,
            },
            {
                'code': 'TRANCHE_2',
                'libelle': '2ème tranche',
                'du': int(echeancier.tranche_2_due or 0),
                'paye': int(echeancier.tranche_2_payee or 0),
                'echeance': echeancier.date_echeance_tranche_2,
            },
            {
                'code': 'TRANCHE_3',
                'libelle': '3ème tranche',
                'du': int(echeancier.tranche_3_due or 0),
                'paye': int(echeancier.tranche_3_payee or 0),
                'echeance': echeancier.date_echeance_tranche_3,
            },
        ]
        total_du = sum(poste['du'] for poste in postes)
        total_paye_brut = sum(poste['paye'] for poste in postes)
        total_couvert = min(total_du, total_paye_brut + remises_total)
        reste_a_payer = max(total_du - total_couvert, 0)
        exigible = sum(poste['du'] for poste in postes if poste['echeance'] and poste['echeance'] <= today)
        retard_reel = max(exigible - total_couvert, 0)
        taux_paye = round((total_couvert / total_du * 100), 1) if total_du > 0 else 0

        postes_non_soldes = [
            {
                **poste,
                'reste': max(poste['du'] - poste['paye'], 0),
                'en_retard': bool(poste['echeance'] and poste['echeance'] < today and poste['paye'] < poste['du']),
            }
            for poste in postes
            if poste['du'] > poste['paye']
        ]
        postes_non_soldes.sort(key=lambda item: (not item['en_retard'], item['echeance'] or today))
        prochain_paiement = postes_non_soldes[0] if postes_non_soldes else None

        responsable = getattr(eleve, 'responsable_principal', None) or getattr(eleve, 'responsable_secondaire', None)
        telephone_parent = getattr(responsable, 'telephone', '') if responsable else ''

        def _format_whatsapp_number(numero):
            clean = (numero or '').replace(' ', '').replace('-', '').replace('.', '')
            if not clean:
                return ''
            if clean.startswith('00'):
                clean = '+' + clean[2:]
            elif clean.startswith('224'):
                clean = '+' + clean
            elif not clean.startswith('+'):
                clean = '+224' + clean
            return clean

        ecole = eleve.classe.ecole if eleve.classe else None
        nom_ecole = ecole.nom if ecole else 'École'
        tel_ecole = ecole.tous_telephones if ecole else ''
        message_relance = (
            f"Bonjour Cher Parent,\n\n"
            f"Voici la situation financière de {eleve.prenom} {eleve.nom} ({eleve.matricule}) à {nom_ecole}.\n"
            f"Total dû : {total_du:,.0f} GNF\n"
            f"Total payé/remises : {total_couvert:,.0f} GNF\n"
            f"Reste à payer : {reste_a_payer:,.0f} GNF\n"
        ).replace(',', ' ')
        if retard_reel > 0:
            message_relance += f"Montant en retard : {retard_reel:,.0f} GNF\n".replace(',', ' ')
        if prochain_paiement:
            date_txt = prochain_paiement['echeance'].strftime('%d/%m/%Y') if prochain_paiement['echeance'] else 'non définie'
            message_relance += (
                f"Prochain paiement attendu : {prochain_paiement['libelle']} - "
                f"{prochain_paiement['reste']:,.0f} GNF, échéance {date_txt}.\n"
            ).replace(',', ' ')
        message_relance += "\nMerci de régulariser la situation. La Direction"
        if tel_ecole:
            message_relance += f"\nContact école : {tel_ecole}"

        whatsapp_number = _format_whatsapp_number(telephone_parent)
        whatsapp_relance_link = ''
        if whatsapp_number:
            whatsapp_relance_link = f"https://wa.me/{whatsapp_number.replace('+', '')}?text={urllib.parse.quote(message_relance)}"

        finance_eleve = {
            'total_du': total_du,
            'total_paye_brut': total_paye_brut,
            'remises_total': remises_total,
            'total_couvert': total_couvert,
            'reste_a_payer': reste_a_payer,
            'retard_reel': retard_reel,
            'taux_paye': taux_paye,
            'taux_paye_bar': min(100, max(0, int(round(taux_paye)))),
            'prochain_paiement': prochain_paiement,
            'message_relance': message_relance,
            'telephone_parent': telephone_parent,
            'whatsapp_relance_link': whatsapp_relance_link,
        }

    context = {
        'titre_page': "Échéancier des paiements",
        'eleve': eleve,
        'echeancier': echeancier,
        'paiements': paiements,
        'today': today,
        'finance_eleve': finance_eleve,
    }
    return render(request, 'paiements/echeancier_eleve.html', context)

@login_required
@require_school_object(Eleve, pk_kwarg='eleve_id', field_path='classe__ecole')
def creer_echeancier(request, eleve_id:int):
    """Créer ou éditer l'échéancier d'un élève.

    - Si un échéancier existe déjà: redirige vers la page d'échéancier avec message.
    - GET: affiche `templates/paiements/form_echeancier.html` pré-rempli si possible par la grille tarifaire.
    - POST: valide et enregistre l'échéancier puis redirige vers la page d'échéancier de l'élève.
    """
    eleve_qs = Eleve.objects.select_related('classe', 'classe__ecole')
    eleve_qs = filter_by_user_school(eleve_qs, request.user, 'classe__ecole')
    eleve = get_object_or_404(eleve_qs, pk=eleve_id)

    # Si un échéancier existe déjà, on informe et on redirige
    if getattr(eleve, 'echeancier', None):
        messages.info(request, "Un échéancier existe déjà pour cet élève.")
        return redirect('paiements:echeancier_eleve', eleve_id=eleve.id)

    # Pré-remplissage depuis la grille tarifaire si disponible
    initial = {}
    try:
        niveau = getattr(eleve.classe, 'niveau', None)
        ecole = getattr(eleve.classe, 'ecole', None)
        # Année scolaire préférée: celle de la classe de l'élève, sinon calcul par date
        today = date.today()
        annee_scolaire_def = f"{today.year}-{today.year+1}" if today.month >= 9 else f"{today.year-1}-{today.year}"
        annee_classe = getattr(eleve.classe, 'annee_scolaire', None)
        from eleves.models import GrilleTarifaire as _Grille
        grille = None
        # 1) Essayer l'année de la classe si présente
        if annee_classe:
            grille = _Grille.objects.filter(ecole=ecole, niveau=niveau, annee_scolaire=annee_classe).first()
            if grille is None:
                messages.info(request, f"Aucune grille trouvée pour l'année {annee_classe}. Recherche d'une autre année...")
        # 2) Sinon essayer l'année par défaut calculée (ou si 1) a échoué et diffère)
        if grille is None:
            if not annee_classe or (annee_classe and annee_classe != annee_scolaire_def):
                grille = _Grille.objects.filter(ecole=ecole, niveau=niveau, annee_scolaire=annee_scolaire_def).first()
                if grille and annee_classe and annee_classe != annee_scolaire_def:
                    messages.info(request, f"Utilisation de la grille {grille.annee_scolaire} (aucune pour {annee_classe}).")
        # 3) Fallback: prendre la plus récente disponible pour l'école/niveau
        if grille is None:
            grille = _Grille.objects.filter(ecole=ecole, niveau=niveau).order_by('-annee_scolaire').first()
            if grille:
                messages.warning(request, f"Grille exacte introuvable. Utilisation de la plus récente: {grille.annee_scolaire}.")

        if grille:
            initial.update({
                'annee_scolaire': grille.annee_scolaire,
                'frais_inscription_du': grille.frais_inscription,
                'tranche_1_due': grille.tranche_1,
                'tranche_2_due': grille.tranche_2,
                'tranche_3_due': grille.tranche_3,
            })
        # Proposer des dates d'échéance par défaut
        try:
            # Inscription: aujourd'hui, puis jalons (janvier/mars) pour les tranches
            from datetime import date as _d
            today_d = _d.today()
            initial.setdefault('date_echeance_inscription', today_d)
            # 15 janvier, 15 mars, 15 mai de l'année de fin de l'année scolaire (annee_debut + 1)
            annee_scol = (initial.get('annee_scolaire') or annee_scolaire_def)
            try:
                annee_debut = int(str(annee_scol).split('-')[0])
            except Exception:
                annee_debut = today_d.year
            annee_fin = annee_debut + 1
            initial.setdefault('date_echeance_tranche_1', _d(annee_fin, 1, 15))
            initial.setdefault('date_echeance_tranche_2', _d(annee_fin, 3, 15))
            # Dernière tranche: 15 mai
            initial.setdefault('date_echeance_tranche_3', _d(annee_fin, 5, 15))
        except Exception:
            pass
    except Exception:
        grille = None

    if request.method == 'POST':
        form = EcheancierForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                ech: EcheancierPaiement = form.save(commit=False)
                ech.eleve = eleve
                if request.user.is_authenticated:
                    ech.cree_par = request.user
                ech.save()
            messages.success(request, "Échéancier créé avec succès.")
            return redirect('paiements:echeancier_eleve', eleve_id=eleve.id)
        else:
            messages.error(request, "Veuillez corriger les erreurs du formulaire.")
    else:
        form = EcheancierForm(initial=initial)

    context = {
        'titre_page': "Créer un échéancier",
        'eleve': eleve,
        'form': form,
        'grille': grille if 'grille' in locals() else None,
        'action': 'Créer',
    }
    return render(request, 'paiements/form_echeancier.html', context)

@login_required
@require_school_object(Eleve, pk_kwarg='eleve_id', field_path='classe__ecole')
def assurer_echeancier(request, eleve_id: int):
    """Assure la création automatique de l'échéancier si manquant, puis redirige vers la page échéancier.

    Utilise `ensure_echeancier_for_eleve()` pour créer silencieusement à partir de la grille tarifaire.
    """
    eleve = get_object_or_404(Eleve.objects.select_related('classe', 'classe__ecole'), pk=eleve_id)
    try:
        ensure_echeancier_for_eleve(
            eleve,
            created_by=request.user if getattr(request.user, 'is_authenticated', False) else None,
        )
        # Synchroniser le statut juste après
        _auto_validate_echeancier_for_eleve(eleve)
        messages.success(request, "Échéancier créé/mis à jour automatiquement.")
    except Exception:
        logging.getLogger(__name__).exception("Erreur lors de l'assurance de l'échéancier")
        messages.error(request, "Impossible de créer automatiquement l'échéancier. Veuillez réessayer ou le créer manuellement.")
    return redirect('paiements:echeancier_eleve', eleve_id=eleve.id)

@login_required
@require_school_object(Eleve, pk_kwarg='eleve_id', field_path='classe__ecole')
def valider_echeancier(request, eleve_id: int):
    """Valide/synchronise l'échéancier d'un élève sur la base des paiements validés.

    - POST requis pour exécuter l'action
    - Vérifie l'autorisation via `can_validate_payments`
    - Utilise `_auto_validate_echeancier_for_eleve` pour ajuster le statut et montants payés si nécessaire
    - Redirige ensuite vers la page `echeancier_eleve`
    """
    # Autorisation
    if not has_permission(request.user, 'peut_valider_paiements'):
        messages.error(request, "Vous n'avez pas l'autorisation de valider les échéanciers.")
        return redirect('paiements:echeancier_eleve', eleve_id=eleve_id)

    # Méthode HTTP
    if request.method != 'POST':
        messages.warning(request, "Action invalide: la validation doit être envoyée en POST.")
        return redirect('paiements:echeancier_eleve', eleve_id=eleve_id)

    eleve = get_object_or_404(Eleve.objects.select_related('classe', 'classe__ecole'), pk=eleve_id)
    try:
        with transaction.atomic():
            # S'assurer qu'un échéancier existe d'abord
            ensure_echeancier_for_eleve(
                eleve,
                created_by=request.user if getattr(request.user, 'is_authenticated', False) else None,
            )
            # Puis synchroniser/valider
            _auto_validate_echeancier_for_eleve(eleve)
        messages.success(request, "Échéancier validé et synchronisé avec les paiements.")
    except Exception:
        logging.getLogger(__name__).exception("Erreur lors de la validation/synchronisation de l'échéancier")
        messages.error(request, "Une erreur est survenue lors de la validation de l'échéancier.")
    # Nouveau flux: si l'élève a un paiement récent en attente sans remise, rediriger vers son détail
    try:
        paiement_en_attente = (
            Paiement.objects
            .filter(eleve=eleve, statut='EN_ATTENTE')
            .order_by('-date_paiement', '-date_creation', '-id')
            .first()
        )
    except Exception:
        paiement_en_attente = None

    if paiement_en_attente:
        try:
            nb_remises = paiement_en_attente.remises.count()
        except Exception:
            nb_remises = 0
        if (nb_remises or 0) == 0:
            messages.info(request, "Aucune remise appliquée: veuillez valider le paiement en attente.")
            return redirect('paiements:detail_paiement', paiement_id=paiement_en_attente.id)

    return redirect('paiements:echeancier_eleve', eleve_id=eleve.id)

@login_required
@require_school_object(Paiement, pk_kwarg='paiement_id', field_path='eleve__classe__ecole')
def generer_recu_pdf(request, paiement_id:int):
    """Génère un reçu PDF téléchargeable pour un paiement validé.

    - Ajoute un filigrane via `ecole_moderne/pdf_utils.draw_logo_watermark`
    - Inclut les informations clés du paiement et de l'élève
    - Liste les remises appliquées et affiche le total des remises
    """
    paiement_qs = Paiement.objects.select_related('eleve', 'type_paiement', 'mode_paiement', 'eleve__classe', 'eleve__classe__ecole')
    paiement_qs = filter_by_user_school(paiement_qs, request.user, 'eleve__classe__ecole')
    paiement = get_object_or_404(paiement_qs, pk=paiement_id)

    # Optionnel: n'autoriser le reçu que pour les paiements validés
    if getattr(paiement, 'statut', 'EN_ATTENTE') != 'VALIDE':
        messages.warning(request, "Le reçu n'est disponible que pour les paiements validés.")
        return redirect('paiements:detail_paiement', paiement_id=paiement.id)

    if canvas is None:
        return HttpResponse("La génération de PDF n'est pas disponible sur ce serveur (ReportLab manquant).", status=500)

    # Valider/synchroniser l'échéancier de l'élève avant génération du reçu
    try:
        with transaction.atomic():
            _auto_validate_echeancier_for_eleve(paiement.eleve)
    except Exception:
        logging.getLogger(__name__).exception("Validation automatique de l'échéancier avant reçu échouée")

    # Calcul total remises
    remises_total = paiement.remises.aggregate(total=Sum('montant_remise')).get('total') or 0

    # Préparer le buffer et le canvas
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Filigrane: toujours actif pour les reçus PDF, spécifique à l'école du paiement
    try:
        ecole_obj = getattr(getattr(paiement.eleve, 'classe', None), 'ecole', None)
        draw_logo_watermark(c, width, height, ecole=ecole_obj)
    except Exception:
        pass

    # Déterminer libellé Inscription/Réinscription pour l'affichage (structure inchangée)
    try:
        _type_nom = (getattr(paiement.type_paiement, 'nom', '') or '').strip().lower()
    except Exception:
        _type_nom = ''
    label_insc = "Réinscription" if ('réinscription' in _type_nom or 'reinscription' in _type_nom) else "Inscription"

    # Mise en page simple
    left = 40
    top = height - 40
    line_h = 18

    def draw_line(text, x=left, y=None, bold=False):
        nonlocal top
        if y is None:
            y = top
        font_name = 'Helvetica-Bold' if bold else 'Helvetica'
        c.setFont(font_name, 11)
        c.drawString(x, y, text)
        top = y - line_h

    # Logo en en-tête (côté gauche) — logo de l'école si disponible
    try:
        logo_path = None
        try:
            ecole_obj = getattr(getattr(paiement.eleve, 'classe', None), 'ecole', None)
            import os
            if ecole_obj is not None and hasattr(ecole_obj, 'logo'):
                school_logo_path = getattr(getattr(ecole_obj, 'logo', None), 'path', None)
                if school_logo_path and os.path.exists(school_logo_path):
                    logo_path = school_logo_path
        except Exception:
            logo_path = None

        # Fallback vers le logo statique global si aucun logo d'école
        if not logo_path:
            from django.contrib.staticfiles import finders
            logo_path = finders.find('logos/logo.png')

        if logo_path and ImageReader is not None:
            try:
                logo_img = ImageReader(logo_path)
                logo_w, logo_h = 80, 80
                c.drawImage(logo_img, left, top - logo_h, width=logo_w, height=logo_h, preserveAspectRatio=True, mask='auto')
                
                # Titre à côté du logo
                c.setFont('Helvetica-Bold', 18)
                c.drawString(left + logo_w + 20, top - 25, "REÇU DE PAIEMENT")
                
                # Nom de l'école sous le titre
                c.setFont('Helvetica-Bold', 11)
                ecole_nom = getattr(ecole_obj, 'nom', "")
                c.drawString(left + logo_w + 20, top - 45, ecole_nom)
                # Coordonnées de l'école (sur une ou deux lignes, si disponibles)
                adr = getattr(ecole_obj, 'adresse', '') or ''
                tel = getattr(ecole_obj, 'telephone', '') or ''
                email = getattr(ecole_obj, 'email', '') or ''
                info_x = left + logo_w + 20
                y_info = top - 62
                c.setFont('Helvetica', 9)
                try:
                    c.setFillGray(0.3)
                except Exception:
                    pass
                if adr:
                    c.drawString(info_x, y_info, f"Adresse: {adr}")
                    y_info -= 12
                # Afficher téléphone et email sur des lignes séparées
                if tel:
                    c.drawString(info_x, y_info, f"Tél: {tel}")
                    y_info -= 12
                if email:
                    c.drawString(info_x, y_info, f"Email: {email}")
                    y_info -= 12
                try:
                    c.setFillGray(0.0)
                except Exception:
                    pass
                
                # Ajuster le top après le bloc en-tête
                top -= (logo_h + 26)
            except Exception:
                # Fallback sans logo
                c.setFont('Helvetica-Bold', 18)
                c.drawString(left, top, "REÇU DE PAIEMENT")
                top -= 15
                c.setFont('Helvetica-Bold', 11)
                c.drawString(left, top, (getattr(ecole_obj, 'nom', "")))
                # Coordonnées
                adr = getattr(ecole_obj, 'adresse', '') or ''
                tel = getattr(ecole_obj, 'telephone', '') or ''
                email = getattr(ecole_obj, 'email', '') or ''
                c.setFont('Helvetica', 9)
                try:
                    c.setFillGray(0.3)
                except Exception:
                    pass
                top -= 14
                if adr:
                    c.drawString(left, top, f"Adresse: {adr}")
                    top -= 12
                # Afficher téléphone et email sur des lignes séparées
                if tel:
                    c.drawString(left, top, f"Tél: {tel}")
                    top -= 12
                if email:
                    c.drawString(left, top, f"Email: {email}")
                    top -= 12
                try:
                    c.setFillGray(0.0)
                except Exception:
                    pass
                top -= 10
        else:
            # Fallback sans logo
            c.setFont('Helvetica-Bold', 18)
            c.drawString(left, top, "REÇU DE PAIEMENT")
            top -= 15
            c.setFont('Helvetica-Bold', 11)
            c.drawString(left, top, (getattr(ecole_obj, 'nom', "")))
            # Coordonnées
            adr = getattr(ecole_obj, 'adresse', '') or ''
            tel = getattr(ecole_obj, 'telephone', '') or ''
            email = getattr(ecole_obj, 'email', '') or ''
            c.setFont('Helvetica', 9)
            try:
                c.setFillGray(0.3)
            except Exception:
                pass
            top -= 14
            if adr:
                c.drawString(left, top, f"Adresse: {adr}")
                top -= 12
            # Afficher téléphone et email sur des lignes séparées
            if tel:
                c.drawString(left, top, f"Tél: {tel}")
                top -= 12
            if email:
                c.drawString(left, top, f"Email: {email}")
                top -= 12
            try:
                c.setFillGray(0.0)
            except Exception:
                pass
            top -= 10
    except Exception:
        # Fallback en cas d'erreur
        c.setFont('Helvetica-Bold', 18)
        c.drawString(left, top, "REÇU DE PAIEMENT")
        top -= 15
        c.setFont('Helvetica-Bold', 11)
        c.drawString(left, top, (getattr(ecole_obj, 'nom', "")))
        # Coordonnées
        adr = getattr(ecole_obj, 'adresse', '') or ''
        tel = getattr(ecole_obj, 'telephone', '') or ''
        email = getattr(ecole_obj, 'email', '') or ''
        dirc = getattr(ecole_obj, 'directeur', '') or ''
        c.setFont('Helvetica', 9)
        try:
            c.setFillGray(0.3)
        except Exception:
            pass
        top -= 14
        if adr:
            c.drawString(left, top, f"Adresse: {adr}")
            top -= 12
        # Afficher téléphone et email sur des lignes séparées
        if tel:
            c.drawString(left, top, f"Tél: {tel}")
            top -= 12
        if email:
            c.drawString(left, top, f"Email: {email}")
            top -= 12
        if dirc:
            c.drawString(left, top, f"Directeur: {dirc}")
            top -= 12
        try:
            c.setFillGray(0.0)
        except Exception:
            pass
        top -= 10

    # Photo élève (en haut à droite si disponible) ou placeholder avec initiales si absente
    try:
        img_drawn = False
        img_w, img_h = 100, 100
        x_img = width - 40 - img_w
        y_img = height - 40 - img_h
        if ImageReader is not None:
            photo_path = getattr(getattr(paiement.eleve, 'photo', None), 'path', None)
            if photo_path and os.path.exists(photo_path):
                try:
                    img = ImageReader(photo_path)
                    c.drawImage(img, x_img, y_img, width=img_w, height=img_h, preserveAspectRatio=True, mask='auto')
                    img_drawn = True
                except Exception:
                    img_drawn = False
        if not img_drawn:
            # Dessiner un placeholder avec initiales
            nom_complet = str(getattr(paiement.eleve, 'nom_complet', '') or '').strip()
            initiales = ''.join([p[0].upper() for p in nom_complet.split()[:2]]) or 'E'
            c.setLineWidth(1)
            try:
                c.roundRect(x_img, y_img, img_w, img_h, 8)
            except Exception:
                c.rect(x_img, y_img, img_w, img_h)
            c.setFont('Helvetica-Bold', 24)
            c.drawCentredString(x_img + img_w/2, y_img + img_h/2 - 8, initiales)
            c.setFont('Helvetica', 8)
            c.drawCentredString(x_img + img_w/2, y_img + 6, "Pas de photo")
        # Afficher le nom de l'élève sous l'image/placeholder
        try:
            nom_aff = str(getattr(paiement.eleve, 'nom_complet', '') or '').strip()
            if nom_aff:
                c.setFont('Helvetica', 9)
                c.drawCentredString(x_img + img_w/2, y_img - 12, nom_aff)
        except Exception:
            pass
    except Exception:
        # En cas de problème avec le rendu de la photo/placeholder, ne pas bloquer la génération du reçu
        pass

    # Informations paiement
    draw_line(f"Numéro de reçu : {paiement.numero_recu}", bold=True)
    draw_line(f"Date de paiement : {paiement.date_paiement.strftime('%d/%m/%Y')}")
    draw_line(f"Type de paiement : {paiement.type_paiement.nom}")
    draw_line(f"Mode de paiement : {paiement.mode_paiement.nom}")
    if getattr(paiement, 'reference_externe', None):
        draw_line(f"Référence externe : {paiement.reference_externe}")
    if getattr(paiement, 'observations', None):
        # Limiter l'observation à une ligne raisonnable pour le reçu
        obs = str(paiement.observations).strip()
        if obs:
            draw_line(f"Observations : {obs}")
    # Calculer le montant global annuel à payer
    try:
        echeancier = getattr(paiement.eleve, 'echeancier', None)
        if echeancier:
            montant_global_annuel = int(
                (echeancier.frais_inscription_du or 0) +
                (echeancier.tranche_1_due or 0) +
                (echeancier.tranche_2_due or 0) +
                (echeancier.tranche_3_due or 0)
            )
        else:
            montant_global_annuel = 0
    except Exception:
        montant_global_annuel = 0
    
    # Afficher le montant global annuel
    if montant_global_annuel > 0:
        draw_line(f"Montant global annuel : {str(f'{montant_global_annuel:,}').replace(',', ' ')} GNF", bold=True)
        top -= 5  # Petit espace
    
    # Afficher le montant payé
    draw_line(f"Montant payé : {str(f'{paiement.montant:,.0f}').replace(',', ' ')} GNF", bold=True)

    if remises_total and int(remises_total) > 0:
        draw_line(f"Total remises : -{str(f'{int(remises_total):,}').replace(',', ' ')} GNF")
    # Montant net (jamais négatif)
    montant_net = max(0, int(paiement.montant - (remises_total or 0)))
    draw_line(f"Montant net payé : {str(f'{montant_net:,}').replace(',', ' ')} GNF", bold=True)

    # Affectation du paiement courant sur les tranches (simulation déterministe)
    # Objectif: montrer, pour CE reçu, quelle partie couvre Inscription/T1/T2/T3
    try:
        echeancier_for_alloc = getattr(paiement.eleve, 'echeancier', None)
    except Exception:
        echeancier_for_alloc = None
    if echeancier_for_alloc:
        try:
            # Restants initiaux égaux aux dus de l'échéancier
            rest_insc = int(echeancier_for_alloc.frais_inscription_du or 0)
            rest_t1 = int(echeancier_for_alloc.tranche_1_due or 0)
            rest_t2 = int(echeancier_for_alloc.tranche_2_due or 0)
            rest_t3 = int(echeancier_for_alloc.tranche_3_due or 0)

            # Parcourir tous les paiements validés (y compris celui-ci) dans l'ordre
            paiements_valides = (
                Paiement.objects
                .filter(eleve=paiement.eleve, statut='VALIDE')
                .order_by('date_paiement', 'date_creation', 'id')
            )

            allocations = {}
            for p in paiements_valides.iterator():
                # Couverture de ce paiement = montant + remises sur CE paiement
                try:
                    rem_p = p.remises.aggregate(total=Sum('montant_remise')).get('total') or 0
                except Exception:
                    rem_p = 0
                reste_a_repartir = max(0, int(p.montant) + int(rem_p))

                a_insc = a_t1 = a_t2 = a_t3 = 0
                if reste_a_repartir and rest_insc > 0:
                    a = min(rest_insc, reste_a_repartir)
                    a_insc = a
                    rest_insc -= a
                    reste_a_repartir -= a
                if reste_a_repartir and rest_t1 > 0:
                    a = min(rest_t1, reste_a_repartir)
                    a_t1 = a
                    rest_t1 -= a
                    reste_a_repartir -= a
                if reste_a_repartir and rest_t2 > 0:
                    a = min(rest_t2, reste_a_repartir)
                    a_t2 = a
                    rest_t2 -= a
                    reste_a_repartir -= a
                if reste_a_repartir and rest_t3 > 0:
                    a = min(rest_t3, reste_a_repartir)
                    a_t3 = a
                    rest_t3 -= a
                    reste_a_repartir -= a

                allocations[p.id] = (a_insc, a_t1, a_t2, a_t3)

            if allocations.get(paiement.id):
                top -= 6
                draw_line("Affectation du paiement", bold=True)
                a_insc, a_t1, a_t2, a_t3 = allocations[paiement.id]
                draw_line(f"{label_insc}: {str(f'{int(a_insc):,}').replace(',', ' ')} GNF")
                draw_line(f"1ère tranche: {str(f'{int(a_t1):,}').replace(',', ' ')} GNF")
                draw_line(f"2ème tranche: {str(f'{int(a_t2):,}').replace(',', ' ')} GNF")
                draw_line(f"3ème tranche: {str(f'{int(a_t3):,}').replace(',', ' ')} GNF")
        except Exception:
            pass

    # Élève
    top -= 6
    draw_line("Informations de l'élève", bold=True)
    draw_line(f"Nom : {paiement.eleve.nom_complet}")
    if getattr(paiement.eleve, 'matricule', None):
        draw_line(f"Matricule : {paiement.eleve.matricule}")
    if getattr(paiement.eleve, 'classe', None):
        draw_line(f"Classe : {paiement.eleve.classe}")

    # Échéances (si disponibles sur l'échéancier de l'élève)
    try:
        echeancier = getattr(paiement.eleve, 'echeancier', None)
    except Exception:
        echeancier = None
    if echeancier:
        top -= 6
        draw_line("Échéances", bold=True)
        try:
            def _fmt_amount(v):
                try:
                    return str(f"{int(v or 0):,}").replace(',', ' ')
                except Exception:
                    return str(v or 0)
            def _fmt_date(d):
                try:
                    return d.strftime('%d/%m/%Y') if d else ''
                except Exception:
                    return str(d) if d else ''
            # Inscription / Réinscription (libellé dynamique, structure inchangée)
            draw_line(f"{label_insc}: {_fmt_amount(echeancier.frais_inscription_du)} GNF - Échéance: {_fmt_date(echeancier.date_echeance_inscription)}")
            # Tranches
            draw_line(f"1ère tranche: {_fmt_amount(echeancier.tranche_1_due)} GNF - Échéance: {_fmt_date(echeancier.date_echeance_tranche_1)}")
            draw_line(f"2ème tranche: {_fmt_amount(echeancier.tranche_2_due)} GNF - Échéance: {_fmt_date(echeancier.date_echeance_tranche_2)}")
            draw_line(f"3ème tranche: {_fmt_amount(echeancier.tranche_3_due)} GNF - Échéance: {_fmt_date(echeancier.date_echeance_tranche_3)}")
        except Exception:
            pass

        # Restes à payer par tranche
        try:
            def _reste(due, paye):
                try:
                    return max(0, int((due or 0) - (paye or 0)))
                except Exception:
                    return 0
            # Calcul global basé sur les paiements validés: somme(montants) - somme(remises)
            try:
                total_du = int((echeancier.frais_inscription_du or 0) + (echeancier.tranche_1_due or 0) + (echeancier.tranche_2_due or 0) + (echeancier.tranche_3_due or 0))
            except Exception:
                total_du = 0

            try:
                sum_montant, sum_remises = _sum_validated_payments_and_remises(paiement.eleve)
            except Exception:
                sum_montant = 0
                sum_remises = 0

            # Calcul de la couverture: montants payés + remises validées (les remises couvrent une partie du dû)
            couverture_validee = max(0, int(sum_montant) + int(sum_remises))
            # Inclure le paiement courant s'il n'est pas encore validé (montant + remises sur ce reçu)
            try:
                couverture_courante = max(0, int(paiement.montant) + int(remises_total or 0))
            except Exception:
                couverture_courante = 0
            couverture_effective = couverture_validee + (couverture_courante if paiement.statut != 'VALIDE' else 0)
            tout_solde = (total_du <= couverture_effective)
            solde_global = max(0, int(total_du - couverture_effective))

            top -= 6
            # Solde global restant
            draw_line(f"Solde global restant : {str(f'{solde_global:,}').replace(',', ' ')} GNF", bold=True)
            draw_line("Restes à payer par tranche", bold=True)
            if tout_solde:
                r_insc = r_t1 = r_t2 = r_t3 = 0
            else:
                r_insc = _reste(echeancier.frais_inscription_du, echeancier.frais_inscription_paye)
                r_t1 = _reste(echeancier.tranche_1_due, echeancier.tranche_1_payee)
                r_t2 = _reste(echeancier.tranche_2_due, echeancier.tranche_2_payee)
                r_t3 = _reste(echeancier.tranche_3_due, echeancier.tranche_3_payee)
            draw_line(f"{label_insc}: {str(f'{r_insc:,}').replace(',', ' ')} GNF")
            draw_line(f"1ère tranche: {str(f'{r_t1:,}').replace(',', ' ')} GNF")
            draw_line(f"2ème tranche: {str(f'{r_t2:,}').replace(',', ' ')} GNF")
            draw_line(f"3ème tranche: {str(f'{r_t3:,}').replace(',', ' ')} GNF")
        except Exception:
            pass

    # Remises détaillées
    if remises_total and int(remises_total) > 0:
        top -= 6
        draw_line("Remises appliquées", bold=True)
        for pr in paiement.remises.select_related('remise').all():
            nom = getattr(pr.remise, 'nom', 'Remise')
            montant = str(f"{int(pr.montant_remise):,}").replace(',', ' ')
            draw_line(f"- {nom} : -{montant} GNF")

    # Bloc signatures
    top -= 20
    c.setFont('Helvetica-Bold', 11)
    c.drawString(left, top, "Signatures")
    top -= 16
    # Lignes de signature (caissier et responsable)
    sig_line_y = top
    c.setLineWidth(0.8)
    try:
        from reportlab.lib import colors
        c.setStrokeColor(colors.grey)
    except Exception:
        pass
    # Caissier à gauche
    c.line(left, sig_line_y, left + 200, sig_line_y)
    c.setFont('Helvetica', 10)
    c.drawString(left, sig_line_y - 14, "Caissier(e)")
    # Responsable à droite
    right_x = left + 260
    c.setLineWidth(0.8)
    c.line(right_x, sig_line_y, right_x + 200, sig_line_y)
    c.setFont('Helvetica', 10)
    c.drawString(right_x, sig_line_y - 14, "Responsable")
    # Restaurer couleur par défaut
    try:
        c.setStrokeColorRGB(0, 0, 0)
    except Exception:
        pass

    # Pied de page
    c.setFont('Helvetica', 9)
    c.drawRightString(width - 40, 30, f"Généré le {timezone.now().strftime('%d/%m/%Y %H:%M')}")

    c.showPage()
    c.save()

    pdf = buffer.getvalue()
    buffer.close()

    filename = f"Recu_{paiement.numero_recu}.pdf"
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@login_required
def export_liste_paiements_excel(request):
    """Exporte en Excel la liste des paiements selon les filtres (q, statut).
    Colonnes: Élève, Classe, École, Type, Montant, Mode, Date, Statut, N° Reçu, Observations
    """
    q = (request.GET.get('q') or '').strip()
    statut = (request.GET.get('statut') or '').strip()

    # Construire le queryset cohérent avec la liste
    qs = (
        Paiement.objects
        .select_related('eleve', 'eleve__classe', 'eleve__classe__ecole', 'type_paiement', 'mode_paiement')
        .exclude(statut='ANNULE')
        .order_by('-date_paiement', '-date_creation')
    )
    # Sécurité: restreindre aux paiements de l'école de l'utilisateur (sauf admin)
    qs = filter_by_user_school(qs, request.user, 'eleve__classe__ecole')
    if q:
        qs = qs.filter(
            Q(numero_recu__icontains=q)
            | Q(reference_externe__icontains=q)
            | Q(observations__icontains=q)
            | Q(eleve__nom__icontains=q)
            | Q(eleve__prenom__icontains=q)
            | Q(eleve__matricule__icontains=q)
        )
    if statut:
        qs = qs.filter(statut=statut)

    # Créer le classeur
    wb = Workbook()
    ws = wb.active
    ws.title = 'Paiements'

    headers = [
        'Élève', 'Classe', 'École', 'Type', 'Montant (GNF)', 'Mode', 'Date', 'Statut', 'N° Reçu', 'Observations'
    ]
    ws.append(headers)

    # Styles
    header_fill = PatternFill(start_color='007bff', end_color='007bff', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='DDDDDD')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border_all

    # Lignes
    row_idx = 2
    for p in qs.iterator():
        eleve_nom = f"{getattr(p.eleve, 'nom', '')} {getattr(p.eleve, 'prenom', '')}".strip()
        classe_nom = getattr(getattr(p.eleve, 'classe', None), 'nom', '')
        ecole_nom = getattr(getattr(getattr(p.eleve, 'classe', None), 'ecole', None), 'nom', '')
        type_nom = getattr(p.type_paiement, 'nom', '')
        mode_nom = getattr(p.mode_paiement, 'nom', '')
        date_val = getattr(p, 'date_paiement', None)
        statut_txt = getattr(p, 'statut', '')
        recu = getattr(p, 'numero_recu', '')
        obs = getattr(p, 'observations', '') or ''

        ws.cell(row=row_idx, column=1, value=eleve_nom)
        ws.cell(row=row_idx, column=2, value=classe_nom)
        ws.cell(row=row_idx, column=3, value=ecole_nom)

        ws.cell(row=row_idx, column=4, value=type_nom)
        montant_cell = ws.cell(row=row_idx, column=5, value=float(p.montant or 0))
        montant_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        ws.cell(row=row_idx, column=6, value=mode_nom)

        date_cell = ws.cell(row=row_idx, column=7, value=date_val)
        date_cell.number_format = 'DD/MM/YYYY'
        ws.cell(row=row_idx, column=8, value=statut_txt)
        ws.cell(row=row_idx, column=9, value=recu)
        ws.cell(row=row_idx, column=10, value=obs)

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = border_all
            if col in (1, 2, 3, 4, 6, 8, 9, 10):
                ws.cell(row=row_idx, column=col).alignment = Alignment(vertical='top')

        row_idx += 1

    # Ajustement des largeurs de colonnes
    widths = [22, 14, 18, 18, 16, 14, 12, 12, 12, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Ligne de total montant
    if row_idx > 2:
        total_label_cell = ws.cell(row=row_idx, column=4, value='Total:')
        total_label_cell.font = Font(bold=True)
        total_cell = ws.cell(row=row_idx, column=5, value=f"=SUM(E2:E{row_idx-1})")
        total_cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        total_cell.font = Font(bold=True)

    # Réponse HTTP
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"paiements_{ts}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
@user_passes_test(lambda u: (
    getattr(u, 'is_staff', False)
    or (
        hasattr(u, 'profil') and getattr(getattr(u, 'profil', None), 'role', None) in ['ADMIN', 'COMPTABLE', 'DIRECTEUR']
    )
))
def rapport_remises(request):
    """Rapport des remises avec agrégations par élève et filtres période/recherche.

    Contexte pour `templates/paiements/rapport_remises.html`:
      - rows: liste de dicts avec paiements/élève et champs: nb_remises, total_remise
      - total_global: somme de toutes les remises listées
      - q, date_debut, date_fin: filtres saisis
    """
    titre_page = "Rapport des remises"
    q = (request.GET.get('q') or '').strip()
    date_debut = (request.GET.get('date_debut') or '').strip()
    date_fin = (request.GET.get('date_fin') or '').strip()

    # Base queryset sur les remises liées à des paiements validés
    rem_qs = PaiementRemise.objects.select_related('paiement', 'paiement__eleve')
    rem_qs = rem_qs.filter(paiement__statut='VALIDE')
    # Sécurité: restreindre aux remises liées aux paiements de l'école de l'utilisateur
    rem_qs = filter_by_user_school(rem_qs, request.user, 'paiement__eleve__classe__ecole')

    # Filtre période sur la date du paiement si fournie
    try:
        if date_debut:
            rem_qs = rem_qs.filter(paiement__date_paiement__gte=date_debut)
        if date_fin:
            rem_qs = rem_qs.filter(paiement__date_paiement__lte=date_fin)
    except Exception:
        # En cas de format invalide, ignorer silencieusement
        pass

    # Filtre recherche simple sur élève
    if q:
        rem_qs = rem_qs.filter(
            Q(paiement__eleve__nom__icontains=q)
            | Q(paiement__eleve__prenom__icontains=q)
            | Q(paiement__eleve__matricule__icontains=q)
            | Q(paiement__eleve__classe__nom__icontains=q)
        )

    # Agrégations par élève
    rows = (
        rem_qs
        .values(
            'paiement__eleve__id',
            'paiement__eleve__prenom',
            'paiement__eleve__nom',
            'paiement__eleve__matricule',
            'paiement__eleve__classe__nom',
        )
        .annotate(
            nb_remises=Count('id'),
            total_remise=Coalesce(Sum('montant_remise'), Value(0, output_field=DecimalField(max_digits=10, decimal_places=0)))
        )
        .order_by('-total_remise')
    )

    total_global = 0
    try:
        total_global = int(rem_qs.aggregate(s=Coalesce(Sum('montant_remise'), Value(0)))['s'] or 0)
    except Exception:
        total_global = 0

    context = {
        'titre_page': titre_page,
        'q': q,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'rows': rows,
        'total_global': total_global,
    }
    template = 'paiements/rapport_remises.html' if _template_exists('paiements/rapport_remises.html') else None
    if template:
        return render(request, template, context)
    return HttpResponse('Rapport remises')

@login_required
@user_passes_test(lambda u: u.is_staff or (hasattr(u, 'profil') and u.profil.role in ['ADMIN', 'COMPTABLE', 'DIRECTEUR']))
def liste_eleves_soldes(request):
    """Liste des élèves soldés en tenant compte des remises (hors frais d'inscription).

    Règles:
    - Frais d'inscription (30 000 GNF) non impactés par les remises.
    - Remises s'appliquent uniquement à la scolarité (tranches 1..3).
    - Élève considéré soldé si: net_du = inscription_du + max(tranches_du - remises_totales, 0) est payé.
    """
    from django.utils import timezone as _tz
    today = _tz.localdate() if hasattr(_tz, 'localdate') else date.today()

    # Déterminer dynamiquement l'année scolaire par défaut en fonction de la date courante
    # Septembre → Août: AAAA-AAAA+1, sinon AAAA-1-AAAA
    try:
        annee_dyn = f"{today.year}-{today.year+1}" if today.month >= 9 else f"{today.year-1}-{today.year}"
    except Exception:
        annee_dyn = "2025-2026"
    annee = (request.GET.get('annee') or annee_dyn).strip()
    ecole_id = (request.GET.get('ecole_id') or '').strip()
    classe_id = (request.GET.get('classe_id') or '').strip()
    q = (request.GET.get('q') or '').strip()

    # Base queryset
    qs = (
        EcheancierPaiement.objects
        .select_related('eleve', 'eleve__classe', 'eleve__classe__ecole')
    )

    # Restreindre à l'année scolaire sélectionnée
    try:
        qs = qs.filter(annee_scolaire=annee)
    except Exception:
        pass

    # Sécurité: restreindre aux élèves de l'école de l'utilisateur (sauf admin)
    qs = filter_by_user_school(qs, request.user, 'eleve__classe__ecole')

    # Filtres école/classe
    if ecole_id:
        qs = qs.filter(eleve__classe__ecole_id=ecole_id)
    if classe_id:
        qs = qs.filter(eleve__classe_id=classe_id)
    if q:
        qs = qs.filter(
            Q(eleve__nom__icontains=q) | Q(eleve__prenom__icontains=q) | Q(eleve__matricule__icontains=q)
        )

    # Expressions de calcul
    dues_sco = (
        Coalesce(
            F('tranche_1_due'),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        )
        + Coalesce(
            F('tranche_2_due'),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        )
        + Coalesce(
            F('tranche_3_due'),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        )
    )
    paye_total = (
        Coalesce(
            F('frais_inscription_paye'),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        )
        + Coalesce(
            F('tranche_1_payee'),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        )
        + Coalesce(
            F('tranche_2_payee'),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        )
        + Coalesce(
            F('tranche_3_payee'),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        )
    )
    # Déterminer la période de l'année scolaire pour restreindre les remises aux paiements de l'année
    try:
        annee_debut = int(annee.split('-')[0])
        periode_debut = date(annee_debut, 9, 1)
        periode_fin = date(annee_debut + 1, 8, 31)
    except Exception:
        # Fallback simple si parsing échoue: limiter à l'année civile courante
        annee_debut = today.year if today.month >= 9 else today.year - 1
        periode_debut = date(annee_debut, 9, 1)
        periode_fin = date(annee_debut + 1, 8, 31)
    
    # Spécifique 2025-2026: début au 14/08/2025 pour inclure les enregistrements d'août
    try:
        if annee == "2025-2026":
            periode_debut = date(2025, 8, 14)
    except Exception:
        pass
    # Note: on ne bascule plus automatiquement à l'année précédente avant le 1er septembre.
    # La période par défaut 2025-2026 est maintenue même si today < 1er septembre 2025.
    # Éviter une plage inversée: si today < periode_debut, on fixe periode_fin = periode_debut.
    # Sinon, on cape la fin de période à aujourd'hui pour éviter une fin future.
    try:
        if today < periode_debut:
            periode_fin = periode_debut
        elif periode_fin > today:
            periode_fin = today
    except Exception:
        pass

    remises_total = Coalesce(
        Sum(
            'eleve__paiements__remises__montant_remise',
            filter=(
                Q(eleve__paiements__statut='VALIDE') &
                Q(eleve__paiements__date_paiement__gte=periode_debut) &
                Q(eleve__paiements__date_paiement__lte=periode_fin)
            ),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        ),
        Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
        output_field=DecimalField(max_digits=12, decimal_places=0),
    )
    # Montant payé effectif pour décider du statut SOLDÉ
    # Utilise les champs cumulés de l'échéancier (payés) + remises sur la période
    # Cela évite qu'un filtrage de période fasse disparaître des élèves soldés.
    paye_effectif = ExpressionWrapper(
        Coalesce(F('frais_inscription_paye'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)))
        + Coalesce(F('tranche_1_payee'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)))
        + Coalesce(F('tranche_2_payee'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)))
        + Coalesce(F('tranche_3_payee'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)))
        + remises_total,
        output_field=DecimalField(max_digits=12, decimal_places=0),
    )
    net_sco_du = Greatest(
        Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
        ExpressionWrapper(dues_sco - remises_total, output_field=DecimalField(max_digits=12, decimal_places=0))
    )
    net_du = ExpressionWrapper(
        Coalesce(
            F('frais_inscription_du'),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        )
        + net_sco_du,
        output_field=DecimalField(max_digits=12, decimal_places=0),
    )
    # Solde calculé: net dû (inscription + scolarité nette des remises) - payé effectif cumulé
    solde_calc = ExpressionWrapper(net_du - paye_effectif, output_field=DecimalField(max_digits=12, decimal_places=0))

    qs = qs.annotate(
        total_du_calc=net_du,
        total_paye_calc=paye_effectif,
        solde_calcule=solde_calc,
        total_remises_calc=remises_total,
    ).order_by('eleve__classe__nom', 'eleve__nom', 'eleve__prenom')

    # Élèves soldés: solde <= 0
    all_vals = list(qs.values("pk", "solde_calcule"))
    soldes_pks_list = [r["pk"] for r in all_vals if (r["solde_calcule"] or 0) <= 0]
    qs_soldes = qs.filter(pk__in=soldes_pks_list)

    # Totaux
    aggr = qs_soldes.aggregate(
        du=Coalesce(
            Sum('total_du_calc', output_field=DecimalField(max_digits=12, decimal_places=0)),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        ),
        paye=Coalesce(
            Sum('total_paye_calc', output_field=DecimalField(max_digits=12, decimal_places=0)),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        ),
        solde=Coalesce(
            Sum('solde_calcule', output_field=DecimalField(max_digits=12, decimal_places=0)),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        ),
        remises=Coalesce(
            Sum('total_remises_calc', output_field=DecimalField(max_digits=12, decimal_places=0)),
            Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        ),
    )

    # Pagination
    paginator = Paginator(qs_soldes, 25)
    page_obj = paginator.get_page(request.GET.get('page') or 1)

    # Options d'écoles/classes
    ecoles_qs = []
    try:
        from eleves.models import Ecole
        ecoles_qs = Ecole.objects.all().order_by('nom')
        # Restreindre la liste des écoles pour les non-admins
        if not (getattr(request.user, 'is_superuser', False) or getattr(request.user, 'is_staff', False)):
            ecole_id_user = getattr(getattr(request.user, 'profil', None), 'ecole_id', None)
            if ecole_id_user:
                ecoles_qs = ecoles_qs.filter(pk=ecole_id_user)
            else:
                ecoles_qs = ecoles_qs.none()
    except Exception:
        ecoles_qs = []
    ecole_paiement = user_school(request.user)
    annee_active_p = get_annee_active(request, ecole_paiement) if ecole_paiement else None
    classes = Classe.objects.select_related('ecole').all().order_by('ecole__nom', 'nom')
    classes = filter_by_user_school(classes, request.user, 'ecole')
    if annee_active_p:
        classes = classes.filter(annee_scolaire=annee_active_p)

    # Proposer quelques années autour de l'année active pour la sélection
    try:
        annees_options = [
            f"{annee_debut - 1}-{annee_debut}",
            f"{annee_debut}-{annee_debut + 1}",
            f"{annee_debut + 1}-{annee_debut + 2}",
        ]
    except Exception:
        annees_options = [annee]

    context = {
        'annee': annee,
        'annees_options': annees_options,
        'ecoles': ecoles_qs,
        'classes': classes,
        'ecole_id': ecole_id,
        'classe_id': classe_id,
        'q': q,
        'page_obj': page_obj,
        'totaux': {
            'du': int(aggr['du'] or 0),
            'paye': int(aggr['paye'] or 0),
            'solde': int(aggr['solde'] or 0),
            'remises': int(aggr['remises'] or 0),
        },
        'periode_debut': periode_debut,
        'periode_fin': periode_fin,
    }
    template = 'paiements/eleves_soldes.html' if _template_exists('paiements/eleves_soldes.html') else None
    if template:
        return render(request, template, context)
    return HttpResponse('Élèves soldés')

@login_required
@user_passes_test(lambda u: u.is_staff or (hasattr(u, 'profil') and u.profil.role in ['ADMIN', 'COMPTABLE', 'DIRECTEUR']))
def eleves_soldes_simple(request):
    """Version simplifiée et robuste: affiche les élèves soldés pour l'année sélectionnée.

    Conserve les mêmes clés de contexte attendues par `templates/paiements/eleves_soldes.html`.
    Règle soldé: (inscription_du + max(tranches_du - remises, 0)) - (inscription_payée + tranches_payées + remises) <= 0
    """
    from django.utils import timezone as _tz
    from django.contrib import messages as django_messages
    today = _tz.localdate() if hasattr(_tz, 'localdate') else date.today()

    # Année scolaire par défaut: Septembre→Août
    annee_dyn = f"{today.year}-{today.year+1}" if today.month >= 9 else f"{today.year-1}-{today.year}"
    annee = (request.GET.get('annee') or annee_dyn).strip()
    ecole_id = (request.GET.get('ecole_id') or '').strip()
    classe_id = (request.GET.get('classe_id') or '').strip()
    q = (request.GET.get('q') or '').strip()

    # Période de l'année
    try:
        annee_debut = int(annee.split('-')[0])
        periode_debut = date(annee_debut, 9, 1)
        periode_fin = date(annee_debut + 1, 8, 31)
    except Exception:
        annee_debut = today.year if today.month >= 9 else today.year - 1
        periode_debut = date(annee_debut, 9, 1)
        periode_fin = date(annee_debut + 1, 8, 31)
    if annee == "2025-2026":
        periode_debut = date(2025, 8, 14)
    if today < periode_debut:
        periode_fin = periode_debut
    elif periode_fin > today:
        periode_fin = today

    # Options filtres (toujours disponibles même en cas d'erreur)
    try:
        from eleves.models import Ecole
        ecoles_qs = Ecole.objects.all().order_by('nom')
        if not (getattr(request.user, 'is_superuser', False) or getattr(request.user, 'is_staff', False)):
            ecole_id_user = getattr(getattr(request.user, 'profil', None), 'ecole_id', None)
            ecoles_qs = ecoles_qs.filter(pk=ecole_id_user) if ecole_id_user else ecoles_qs.none()
    except Exception:
        ecoles_qs = []
    
    try:
        ecole_resume = user_school(request.user)
        annee_active_r = get_annee_active(request, ecole_resume) if ecole_resume else None
        classes = Classe.objects.select_related('ecole').all().order_by('ecole__nom', 'nom')
        classes = filter_by_user_school(classes, request.user, 'ecole')
        if annee_active_r:
            classes = classes.filter(annee_scolaire=annee_active_r)
    except Exception:
        classes = []

    try:
        annees_options = [
            f"{annee_debut - 1}-{annee_debut}",
            f"{annee_debut}-{annee_debut + 1}",
            f"{annee_debut + 1}-{annee_debut + 2}",
        ]
    except Exception:
        annees_options = [annee]

    # Vérifier si la table EcheancierPaiement existe et a les bonnes colonnes
    try:
        # Test simple pour vérifier que la table existe avec les colonnes requises
        test_qs = EcheancierPaiement.objects.values('frais_inscription_du', 'tranche_1_due')[:1]
        list(test_qs)  # Force l'exécution de la requête
    except Exception as e:
        # Table ou colonnes manquantes - afficher un message d'erreur
        django_messages.warning(request, "La table des échéanciers n'est pas configurée. Veuillez exécuter les migrations: python manage.py migrate paiements")
        context = {
            'annee': annee,
            'annees_options': annees_options,
            'ecoles': ecoles_qs,
            'classes': classes,
            'ecole_id': ecole_id,
            'classe_id': classe_id,
            'q': q,
            'page_obj': None,
            'totaux': {'du': 0, 'paye': 0, 'solde': 0, 'remises': 0},
            'periode_debut': periode_debut,
            'periode_fin': periode_fin,
            'erreur_migration': True,
        }
        return render(request, 'paiements/eleves_soldes.html', context)

    # Base queryset restreinte par année et école utilisateur
    qs = EcheancierPaiement.objects.select_related('eleve', 'eleve__classe', 'eleve__classe__ecole')
    qs = qs.filter(annee_scolaire=annee)
    qs = filter_by_user_school(qs, request.user, 'eleve__classe__ecole')

    if ecole_id:
        qs = qs.filter(eleve__classe__ecole_id=ecole_id)
    if classe_id:
        qs = qs.filter(eleve__classe_id=classe_id)
    if q:
        qs = qs.filter(
            Q(eleve__nom__icontains=q) | Q(eleve__prenom__icontains=q) | Q(eleve__matricule__icontains=q)
        )

    # Expressions
    dues_sco = (
        Coalesce(F('tranche_1_due'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)))
        + Coalesce(F('tranche_2_due'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)))
        + Coalesce(F('tranche_3_due'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)))
    )
    remises_total = Coalesce(
        Sum(
            'eleve__paiements__remises__montant_remise',
            filter=(
                Q(eleve__paiements__statut='VALIDE') &
                Q(eleve__paiements__date_paiement__gte=periode_debut) &
                Q(eleve__paiements__date_paiement__lte=periode_fin)
            ),
            output_field=DecimalField(max_digits=12, decimal_places=0),
        ),
        Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
    )
    # Calcul du montant de réinscription dû (basé sur la grille tarifaire)
    try:
        from eleves.models import GrilleTarifaire
        reinsc_subq = GrilleTarifaire.objects.filter(
            ecole=OuterRef('eleve__classe__ecole'),
            niveau=OuterRef('eleve__classe__niveau'),
            annee_scolaire=OuterRef('annee_scolaire'),
        ).values('frais_reinscription')[:1]
        
        # Annotation pour identifier les frais de réinscription
        qs = qs.annotate(
            reinsc_due=Case(
                When(frais_inscription_du=Subquery(reinsc_subq), then=F('frais_inscription_du')),
                default=Value(0),
                output_field=DecimalField(max_digits=12, decimal_places=0),
            )
        )
    except Exception:
        # Fallback si pas de grille tarifaire
        qs = qs.annotate(
            reinsc_due=Value(0, output_field=DecimalField(max_digits=12, decimal_places=0))
        )

    paye_effectif = ExpressionWrapper(
        Coalesce(F('frais_inscription_paye'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)))
        + Coalesce(F('tranche_1_payee'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)))
        + Coalesce(F('tranche_2_payee'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)))
        + Coalesce(F('tranche_3_payee'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)))
        + remises_total,
        output_field=DecimalField(max_digits=12, decimal_places=0),
    )
    net_sco_du = Greatest(
        Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
        ExpressionWrapper(dues_sco - remises_total, output_field=DecimalField(max_digits=12, decimal_places=0))
    )
    
    # Total dû incluant inscription/réinscription + scolarité nette
    net_du = ExpressionWrapper(
        Coalesce(F('frais_inscription_du'), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)))
        + net_sco_du,
        output_field=DecimalField(max_digits=12, decimal_places=0),
    )
    solde_calc = ExpressionWrapper(net_du - paye_effectif, output_field=DecimalField(max_digits=12, decimal_places=0))

    qs = qs.annotate(
        total_du_calc=net_du,
        total_paye_calc=paye_effectif,
        solde_calcule=solde_calc,
        total_remises_calc=remises_total,
    ).order_by('eleve__classe__nom', 'eleve__nom', 'eleve__prenom')
    all_vals = list(qs.values("pk", "solde_calcule"))
    soldes_pks_list = [r["pk"] for r in all_vals if (r["solde_calcule"] or 0) <= 0]
    qs_soldes = qs.filter(pk__in=soldes_pks_list)

    aggr = qs_soldes.aggregate(
        du=Coalesce(Sum('total_du_calc', output_field=DecimalField(max_digits=12, decimal_places=0)), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0))),
        paye=Coalesce(Sum('total_paye_calc', output_field=DecimalField(max_digits=12, decimal_places=0)), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0))),
        solde=Coalesce(Sum('solde_calcule', output_field=DecimalField(max_digits=12, decimal_places=0)), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0))),
        remises=Coalesce(Sum('total_remises_calc', output_field=DecimalField(max_digits=12, decimal_places=0)), Value(0, output_field=DecimalField(max_digits=12, decimal_places=0))),
    )

    paginator = Paginator(qs_soldes, 25)
    page_obj = paginator.get_page(request.GET.get('page') or 1)

    context = {
        'annee': annee,
        'annees_options': annees_options,
        'ecoles': ecoles_qs,
        'classes': classes,
        'ecole_id': ecole_id,
        'classe_id': classe_id,
        'q': q,
        'page_obj': page_obj,
        'totaux': {
            'du': int(aggr.get('du') or 0),
            'paye': int(aggr.get('paye') or 0),
            'solde': int(aggr.get('solde') or 0),
            'remises': int(aggr.get('remises') or 0),
        },
        'periode_debut': periode_debut,
        'periode_fin': periode_fin,
    }
    return render(request, 'paiements/eleves_soldes.html', context)

@login_required
def ajax_eleve_info(request):
    """Retourne des informations élève + échéancier pour le formulaire paiement.
    Attend un paramètre `matricule` (GET). Utilisé par `templates/paiements/form_paiement.html`.
    """
    matricule = request.GET.get('matricule') or request.POST.get('matricule')
    if not matricule:
        return JsonResponse({'success': False, 'error': 'Matricule requis.'}, status=400)

    try:
        eleve_qs = Eleve.objects.select_related('classe', 'classe__ecole')
        # Sécurité: restreindre aux élèves de l'école de l'utilisateur
        eleve_qs = filter_by_user_school(eleve_qs, request.user, 'classe__ecole')
        eleve = eleve_qs.get(matricule__iexact=matricule)
    except Eleve.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Élève introuvable.'}, status=404)

    # Construire la réponse
    # Sécuriser l'accès à l'URL de la photo (FieldFile.url peut lever une exception si vide)
    photo_url = ''
    try:
        photo_field = getattr(eleve, 'photo', None)
        if photo_field and getattr(photo_field, 'name', ''):
            photo_url = photo_field.url
    except Exception:
        photo_url = ''

    data = {
        'success': True,
        'eleve': {
            'id': eleve.id,
            'matricule': getattr(eleve, 'matricule', ''),
            'nom': getattr(eleve, 'nom', ''),
            'prenom': getattr(eleve, 'prenom', ''),
            'classe': getattr(eleve.classe, 'nom', '') if getattr(eleve, 'classe', None) else '',
            'ecole': getattr(eleve.classe.ecole, 'nom', '') if getattr(eleve, 'classe', None) and getattr(eleve.classe, 'ecole', None) else '',
            'photo_url': photo_url,
        },
        'echeancier': None,
        'has_echeancier': False,
    }

    # Échéancier (si présent)
    try:
        echeancier = getattr(eleve, 'echeancier', None)
    except Exception:
        echeancier = None

    if echeancier:
        data['echeancier'] = {
            'inscription_du': int(echeancier.frais_inscription_du or 0),
            'inscription_paye': int(echeancier.frais_inscription_paye or 0),
            'tranche_1_du': int(echeancier.tranche_1_due or 0),
            'tranche_1_paye': int(echeancier.tranche_1_payee or 0),
            'tranche_2_du': int(echeancier.tranche_2_due or 0),
            'tranche_2_paye': int(echeancier.tranche_2_payee or 0),
            'tranche_3_du': int(echeancier.tranche_3_due or 0),
            'tranche_3_paye': int(echeancier.tranche_3_payee or 0),
            'total_du': int(echeancier.total_du or 0),
            'total_paye': int(echeancier.total_paye or 0),
            'reste_a_payer': int((echeancier.total_du or 0) - (echeancier.total_paye or 0)),
        }
        data['has_echeancier'] = True

    return JsonResponse(data)

@login_required
def ajax_classes_par_ecole(request):
    return JsonResponse({'ok': True, 'classes': []})

@login_required
def ajax_statistiques_paiements(request):
    """Endpoint AJAX minimal pour statistiques paiements.
    Fourni pour satisfaire le routage; peut être enrichi ultérieurement.
    """
    try:
        base = filter_by_user_school(Paiement.objects.all(), request.user, 'eleve__classe__ecole')
        total = base.count()
        montant_total = int(base.aggregate(total=Sum('montant'))['total'] or 0)
    except Exception:
        total = 0
        montant_total = 0
    return JsonResponse({'success': True, 'total': total, 'montant_total': montant_total})

@login_required
@require_http_methods(["GET", "POST"])
def ajax_calculer_remise(request):
    """Calcule un aperçu de remise. Implémentation basique pour compatibilité UI.
    Accepte un paramètre 'montant' et retourne le même montant en sortie par défaut.
    """
    montant_raw = request.GET.get('montant') or request.POST.get('montant') or '0'
    try:
        montant = int(float(str(montant_raw).replace(' ', '').replace(',', '.')))
    except Exception:
        montant = 0
    return JsonResponse({
        'success': True,
        'montant_initial': montant,
        'montant_apres_remise': montant,
        'details': [],
    })

@login_required
@can_apply_discounts
@require_school_object(Paiement, pk_kwarg='paiement_id', field_path='eleve__classe__ecole')
def appliquer_remise_paiement(request, paiement_id:int):
    """Affiche et traite le formulaire d'application de remises pour un paiement."""
    paiement = get_object_or_404(
        Paiement.objects.select_related('eleve', 'type_paiement', 'mode_paiement'),
        pk=paiement_id,
    )

    # Seuls les paiements en attente peuvent être modifiés
    if getattr(paiement, 'statut', 'EN_ATTENTE') != 'EN_ATTENTE':
        messages.warning(request, "Seuls les paiements en attente peuvent recevoir des remises.")
        return redirect('paiements:detail_paiement', paiement_id=paiement.id)

    # Base scolarité = T1+T2+T3 (hors inscription)
    base_scolarite = 0
    try:
        ech = getattr(paiement.eleve, 'echeancier', None)
        if ech:
            base_scolarite = int((ech.tranche_1_due or 0) + (ech.tranche_2_due or 0) + (ech.tranche_3_due or 0))
        if not base_scolarite:
            # Fallback via grille tarifaire de la classe
            try:
                from eleves.models import GrilleTarifaire as _Grille
                classe = getattr(paiement.eleve, 'classe', None)
                ecole = getattr(classe, 'ecole', None)
                niveau = getattr(classe, 'niveau', None)
                annee = getattr(classe, 'annee_scolaire', None)
                grille = None
                if ecole and niveau and annee:
                    grille = _Grille.objects.filter(ecole=ecole, niveau=niveau, annee_scolaire=annee).first()
                if not grille and ecole and niveau:
                    grille = _Grille.objects.filter(ecole=ecole, niveau=niveau).order_by('-annee_scolaire').first()
                if grille:
                    base_scolarite = int((grille.tranche_1 or 0) + (grille.tranche_2 or 0) + (grille.tranche_3 or 0))
            except Exception:
                pass
    except Exception:
        base_scolarite = 0

    if request.method == 'POST':
        form = PaiementRemiseForm(request.POST, paiement=paiement)
        if form.is_valid():
            remises = form.cleaned_data.get('remises') or []
            pct_str = form.cleaned_data.get('pourcentage_scolarite') or ''
            try:
                pct_value = int(pct_str) if str(pct_str).isdigit() else 0
            except Exception:
                pct_value = 0
            # Si aucune remise n'est sélectionnée, ne rien modifier et afficher une erreur
            if not remises and pct_value <= 0:
                messages.error(request, "Aucune remise sélectionnée. Aucune modification n'a été effectuée.")
                try:
                    remises_existantes = list(paiement.remises.select_related('remise').all())
                except Exception:
                    remises_existantes = []
                context = {
                    'paiement': paiement,
                    'form': form,
                    'remises_existantes': remises_existantes,
                    'base_scolarite': int(base_scolarite or 0),
                }
                return render(request, 'paiements/appliquer_remise.html', context)

            # Avertissement serveur quand 100% est sélectionné (affiché même sans JS)
            if pct_value == 100:
                messages.warning(
                    request,
                    "Attention: vous appliquez 100% de remise scolarité. Cela annulera entièrement la scolarité (T1+T2+T3) pour l'année en cours. Cette règle est applicable à toutes les classes. Vérifiez l'autorisation avant de confirmer."
                )

            # pourcentage_scolarite est un aperçu UI, on ne le persiste pas ici faute de modèle dédié
            with transaction.atomic():
                # Remplacer les remises existantes par la sélection
                PaiementRemise.objects.filter(paiement=paiement).delete()
                created = 0
                for remise in remises:
                    try:
                        montant_remise = remise.calculer_remise(paiement.montant)
                    except Exception:
                        montant_remise = 0
                    PaiementRemise.objects.create(
                        paiement=paiement,
                        remise=remise,
                        montant_remise=montant_remise,
                    )
                    created += 1

                # Appliquer également la remise scolarité (%) si choisie
                if pct_value > 0:
                    from datetime import date
                    annee = paiement.date_paiement.year
                    # Chercher une remise existante "Remise scolarité X%" active et couvrant la date
                    nom_remise = f"Remise scolarité {pct_value}%"
                    remise_pct = RemiseReduction.objects.filter(
                        nom=nom_remise,
                        type_remise='POURCENTAGE',
                        valeur=pct_value,
                        actif=True,
                        date_debut__lte=paiement.date_paiement,
                        date_fin__gte=paiement.date_paiement,
                    ).first()
                    if not remise_pct:
                        # Créer une remise "technique" pour l'année en cours
                        remise_pct = RemiseReduction.objects.create(
                            nom=nom_remise,
                            type_remise='POURCENTAGE',
                            valeur=pct_value,
                            motif='AUTRE',
                            description="Remise scolarité variable (technique)",
                            date_debut=date(annee, 1, 1),
                            date_fin=date(annee, 12, 31),
                            actif=True,
                        )
                    # 3% s'applique sur base scolarité (T1+T2+T3), pas sur le montant du paiement
                    try:
                        montant_remise_pct = (base_scolarite * pct_value) / 100
                    except Exception:
                        montant_remise_pct = (paiement.montant * pct_value) / 100
                    # Ne jamais dépasser le montant du paiement
                    try:
                        from decimal import Decimal as _D
                        montant_remise_pct = min(_D(montant_remise_pct), _D(paiement.montant))
                    except Exception:
                        try:
                            montant_remise_pct = min(float(montant_remise_pct), float(paiement.montant))
                        except Exception:
                            pass
                    PaiementRemise.objects.create(
                        paiement=paiement,
                        remise=remise_pct,
                        montant_remise=montant_remise_pct,
                    )
                    created += 1
            messages.success(request, f"Remises appliquées: {created}.")
            return redirect('paiements:detail_paiement', paiement_id=paiement.id)
        else:
            messages.error(request, "Veuillez corriger les erreurs du formulaire de remises.")
    else:
        form = PaiementRemiseForm(paiement=paiement)

    # Remises déjà liées au paiement (pour affichage et cases cochées)
    try:
        remises_existantes = list(paiement.remises.select_related('remise').all())
    except Exception:
        remises_existantes = []

    context = {
        'paiement': paiement,
        'form': form,
        'remises_existantes': remises_existantes,
        'base_scolarite': int(base_scolarite or 0),
    }
    return render(request, 'paiements/appliquer_remise.html', context)

@login_required
def calculateur_remise(request):
    return HttpResponse('Calculateur de remise (placeholder)')

@login_required
@require_POST
@can_apply_discounts
@require_school_object(Paiement, pk_kwarg='paiement_id', field_path='eleve__classe__ecole')
def annuler_remise_paiement(request, paiement_id:int, remise_id:int=None):
    """Annule les remises appliquées à un paiement.

    - Si remise_id est fourni: supprime uniquement cette remise
    - Sinon: supprime toutes les remises du paiement
    """
    # Sécurité: restreindre l'accès au paiement à l'école de l'utilisateur
    paiement = get_object_or_404(
        filter_by_user_school(
            Paiement.objects.select_related('eleve', 'eleve__classe'),
            request.user,
            'eleve__classe__ecole'
        ),
        pk=paiement_id,
    )
    try:
        if remise_id:
            PaiementRemise.objects.filter(paiement=paiement, id=remise_id).delete()
            messages.success(request, "Remise supprimée.")
        else:
            PaiementRemise.objects.filter(paiement=paiement).delete()
            messages.success(request, "Toutes les remises de ce paiement ont été supprimées.")
    except Exception:
        messages.error(request, "Impossible d'annuler la remise.")
    return redirect('paiements:detail_paiement', paiement_id=paiement.id)

@login_required
def export_paiements_periode_excel(request):
    """Exporte les paiements entre deux dates (du, au) en Excel.
    Paramètres: ?du=YYYY-MM-DD&au=YYYY-MM-DD&statut=VALIDE|EN_ATTENTE|... (optionnel)
    """
    du = request.GET.get('du')
    au = request.GET.get('au')
    statut = (request.GET.get('statut') or '').strip()

    qs = Paiement.objects.select_related('eleve', 'type_paiement', 'mode_paiement')
    # Sécurité: restreindre aux paiements de l'école de l'utilisateur
    qs = filter_by_user_school(qs, request.user, 'eleve__classe__ecole')
    # Filtres période
    try:
        if du:
            qs = qs.filter(date_paiement__gte=du)
        if au:
            qs = qs.filter(date_paiement__lte=au)
    except Exception:
        pass
    if statut:
        qs = qs.filter(statut=statut)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Paiements'
    headers = ['Élève', 'Matricule', 'Classe', 'École', 'Type', 'Montant', 'Mode', 'Date', 'Statut', 'N° Reçu']
    ws.append(headers)
    for p in qs.order_by('date_paiement', 'id'):
        ws.append([
            f"{getattr(p.eleve, 'nom', '')} {getattr(p.eleve, 'prenom', '')}",
            getattr(p.eleve, 'matricule', ''),
            getattr(getattr(p.eleve, 'classe', None), 'nom', ''),
            getattr(getattr(getattr(p.eleve, 'classe', None), 'ecole', None), 'nom', ''),
            getattr(p.type_paiement, 'nom', ''),
            int(p.montant or 0),
            getattr(p.mode_paiement, 'nom', ''),
            getattr(p, 'date_paiement', None).strftime('%Y-%m-%d') if getattr(p, 'date_paiement', None) else '',
            p.statut,
            p.numero_recu or '',
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = 'paiements_periode.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
def rapport_retards(request):
    """Rapport des élèves en retard de paiement (montant exigible > payé+remises).
    Filtres: ?classe_id=&ecole_id=&du=&au=
    """
    from django.utils import timezone as _tz
    today = _tz.localdate() if hasattr(_tz, 'localdate') else date.today()

    exigible_expr = (
        Case(
            When(date_echeance_inscription__lte=today, then=F('frais_inscription_du')),
            default=Value(0), output_field=DecimalField(max_digits=12, decimal_places=0),
        )
        + Case(
            When(date_echeance_tranche_1__lte=today, then=F('tranche_1_due')),
            default=Value(0), output_field=DecimalField(max_digits=12, decimal_places=0),
        )
        + Case(
            When(date_echeance_tranche_2__lte=today, then=F('tranche_2_due')),
            default=Value(0), output_field=DecimalField(max_digits=12, decimal_places=0),
        )
        + Case(
            When(date_echeance_tranche_3__lte=today, then=F('tranche_3_due')),
            default=Value(0), output_field=DecimalField(max_digits=12, decimal_places=0),
        )
    )
    remises_expr = Coalesce(
        Sum('eleve__paiements__remises__montant_remise', filter=Q(eleve__paiements__statut='VALIDE')),
        Value(0), output_field=DecimalField(max_digits=12, decimal_places=0),
    )
    remises_applicables = Least(remises_expr, exigible_expr)
    paye_effectif_expr = (
        F('frais_inscription_paye') + F('tranche_1_payee') + F('tranche_2_payee') + F('tranche_3_payee') + remises_applicables
    )
    retard_expr = ExpressionWrapper(exigible_expr - paye_effectif_expr, output_field=DecimalField(max_digits=12, decimal_places=0))

    qs = (EcheancierPaiement.objects
          .select_related('eleve', 'eleve__classe', 'eleve__classe__ecole')
          .annotate(retard=retard_expr))
    # Sécurité: restreindre aux échéanciers de l'école de l'utilisateur
    qs = filter_by_user_school(qs, request.user, 'eleve__classe__ecole')
    qs = qs.filter(retard__gt=0).order_by('-retard')

    context = {'titre_page': 'Rapport des retards', 'items': qs}
    if _template_exists('rapports/liste_rapports.html'):
        return render(request, 'rapports/liste_rapports.html', context)
    return HttpResponse(f"Retards: {qs.count()} élèves en retard")

@login_required
def rapport_encaissements(request):
    """Rapport des encaissements entre ?du=&au=, somme et décompte par statut."""
    du = request.GET.get('du')
    au = request.GET.get('au')
    qs = Paiement.objects.all()
    # Sécurité: restreindre aux paiements de l'école de l'utilisateur
    qs = filter_by_user_school(qs, request.user, 'eleve__classe__ecole')
    try:
        if du:
            qs = qs.filter(date_paiement__gte=du)
        if au:
            qs = qs.filter(date_paiement__lte=au)
    except Exception:
        pass
    total = int(qs.aggregate(total=Sum('montant'))['total'] or 0)
    par_statut = list(qs.values('statut').annotate(count=Count('id'), somme=Coalesce(Sum('montant'), Value(0))).order_by('statut'))
    context = {'titre_page': 'Rapport des encaissements', 'total': total, 'par_statut': par_statut}
    if _template_exists('rapports/tableau_bord.html'):
        return render(request, 'rapports/tableau_bord.html', context)
    return JsonResponse({'total': total, 'par_statut': par_statut})

@login_required
def api_paiements_list(request):
    """API JSON: liste des paiements avec filtres simples (?q=&statut=&limit=)."""
    q = (request.GET.get('q') or '').strip()
    statut = (request.GET.get('statut') or '').strip()
    try:
        limit = int(request.GET.get('limit') or 50)
    except Exception:
        limit = 50
    qs = Paiement.objects.select_related('eleve', 'type_paiement', 'mode_paiement')
    # Sécurité: restreindre aux paiements de l'école de l'utilisateur
    qs = filter_by_user_school(qs, request.user, 'eleve__classe__ecole')
    if q:
        qs = qs.filter(
            Q(numero_recu__icontains=q) | Q(reference_externe__icontains=q) | Q(observations__icontains=q)
            | Q(eleve__nom__icontains=q) | Q(eleve__prenom__icontains=q) | Q(eleve__matricule__icontains=q)
        )
    if statut:
        qs = qs.filter(statut=statut)
    data = []
    for p in qs.order_by('-date_paiement', '-id')[:limit]:
        data.append({
            'id': p.id,
            'eleve': {
                'id': getattr(p.eleve, 'id', None),
                'matricule': getattr(p.eleve, 'matricule', ''),
                'nom': getattr(p.eleve, 'nom', ''),
                'prenom': getattr(p.eleve, 'prenom', ''),
            },
            'type': getattr(p.type_paiement, 'nom', ''),
            'mode': getattr(p.mode_paiement, 'nom', ''),
            'montant': int(p.montant or 0),
            'date': getattr(p, 'date_paiement', None).strftime('%Y-%m-%d') if getattr(p, 'date_paiement', None) else None,
            'statut': p.statut,
            'numero_recu': p.numero_recu,
        })
    return JsonResponse({'results': data})

@login_required
def api_paiement_detail(request, pk:int):
    """API JSON: détail d'un paiement"""
    p = get_object_or_404(
        filter_by_user_school(
            Paiement.objects.select_related('eleve', 'type_paiement', 'mode_paiement'),
            request.user,
            'eleve__classe__ecole'
        ),
        pk=pk,
    )
    data = {
        'id': p.id,
        'eleve': {
            'id': getattr(p.eleve, 'id', None),
            'matricule': getattr(p.eleve, 'matricule', ''),
            'nom': getattr(p.eleve, 'nom', ''),
            'prenom': getattr(p.eleve, 'prenom', ''),
        },
        'type': getattr(p.type_paiement, 'nom', ''),
        'mode': getattr(p.mode_paiement, 'nom', ''),
        'montant': int(p.montant or 0),
        'date': getattr(p, 'date_paiement', None).strftime('%Y-%m-%d') if getattr(p, 'date_paiement', None) else None,
        'statut': p.statut,
        'numero_recu': p.numero_recu,
        'remises_total': int(p.remises.aggregate(total=Sum('montant_remise')).get('total') or 0) if hasattr(p, 'remises') else 0,
    }
    return JsonResponse(data)

def _template_exists(path:str)->bool:
    """Utilitaire léger: détecte si un template existe dans le chargeur Django."""
    try:
        from django.template.loader import get_template
        get_template(path)
        return True
    except Exception:
        return False


# ========== VUES POUR LES NOTES DE RAPPEL ==========

@login_required
def generer_note_rappel_pdf(request, eleve_id):
    """Génère une note de rappel de paiement pour un élève"""
    from .note_rappel_generator import generer_note_rappel_eleve
    
    # Récupérer l'élève
    eleve = get_object_or_404(Eleve, id=eleve_id)
    
    # Vérifier les permissions
    if not user_is_admin(request.user):
        ecole_user = user_school(request.user)
        if ecole_user != eleve.classe.ecole:
            messages.error(request, "Vous n'avez pas accès à cet élève.")
            return redirect('eleves:detail_eleve', eleve_id=eleve_id)
    
    # Créer la réponse PDF
    response = HttpResponse(content_type='application/pdf')
    filename = f"note_rappel_{eleve.matricule}_{datetime.now().strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Générer le PDF
    generer_note_rappel_eleve(eleve, response)
    
    # Log de l'action
    messages.success(request, f"Note de rappel générée pour {eleve.nom_complet}")
    
    return response


@login_required
def generer_notes_rappel_classe_pdf(request, classe_id):
    """Génère les notes de rappel pour tous les élèves ayant des impayés dans une classe"""
    from .note_rappel_generator import generer_note_rappel_eleve
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, PageBreak
    from io import BytesIO
    
    # Récupérer la classe
    classe = get_object_or_404(Classe, id=classe_id)
    
    # Vérifier les permissions
    if not user_is_admin(request.user):
        ecole_user = user_school(request.user)
        if ecole_user != classe.ecole:
            messages.error(request, "Vous n'avez pas accès à cette classe.")
            return redirect('eleves:classe_detail', classe_id=classe_id)
    
    # Récupérer les élèves avec des impayés (optimisé: 1 requête au lieu de N+1)
    from .models import ConfigurationPaiement
    eleves_avec_impayes = []

    try:
        config = ConfigurationPaiement.objects.get(classe=classe)
        montant_total = config.montant_inscription + config.montant_scolarite

        # Annoter les paiements validés pour éviter N+1 requêtes
        eleves_qs = (
            Eleve.objects.filter(classe=classe, statut='ACTIF')
            .select_related('classe', 'classe__ecole')
            .annotate(
                _total_paye=Coalesce(
                    Sum('paiements__montant', filter=Q(paiements__statut='VALIDE')),
                    Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
                    output_field=DecimalField(max_digits=12, decimal_places=0),
                ),
            )
        )
        for eleve in eleves_qs:
            reste = montant_total - eleve._total_paye
            if reste > 0:
                eleves_avec_impayes.append(eleve)

    except ConfigurationPaiement.DoesNotExist:
        messages.warning(request, "Configuration de paiement non définie pour cette classe.")
        eleves_avec_impayes = list(
            Eleve.objects.filter(classe=classe, statut='ACTIF')
            .select_related('classe', 'classe__ecole')
        )
    
    if not eleves_avec_impayes:
        messages.info(request, "Aucun élève avec des impayés dans cette classe.")
        return redirect('eleves:classe_detail', classe_id=classe_id)
    
    # Fusionner les PDFs individuels
    try:
        from PyPDF2 import PdfMerger
    except ImportError:
        try:
            from pypdf import PdfMerger
        except ImportError:
            # Fallback: un seul PDF
            response = HttpResponse(content_type='application/pdf')
            filename = f"notes_rappel_{classe.nom}_{datetime.now().strftime('%Y%m%d')}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            generer_note_rappel_eleve(eleves_avec_impayes[0], response)
            messages.warning(request, "Module PDF fusion indisponible. Seule la première note a été générée.")
            return response

    merger = PdfMerger()
    for eleve in eleves_avec_impayes:
        buf = BytesIO()
        generer_note_rappel_eleve(eleve, buf)
        buf.seek(0)
        merger.append(buf)

    response = HttpResponse(content_type='application/pdf')
    filename = f"notes_rappel_{classe.nom}_{datetime.now().strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    merger.write(response)
    merger.close()

    messages.success(request, f"{len(eleves_avec_impayes)} notes de rappel générées pour la classe {classe.nom}")

    return response


@login_required
def generer_toutes_notes_rappel_pdf(request):
    """Génère les notes de rappel PDF pour TOUS les élèves avec impayés (toutes classes)."""
    from .note_rappel_generator import generer_note_rappel_eleve
    from .models import ConfigurationPaiement

    # Permissions
    if not user_is_admin(request.user) and not can_view_reports(request.user):
        return HttpResponse(status=403)

    # Élèves selon école de l'utilisateur
    if user_is_superadmin(request.user):
        eleves_qs = Eleve.objects.filter(statut='ACTIF')
    else:
        ecole_user = user_school(request.user)
        if ecole_user is None:
            messages.warning(request, "Aucune école associée à votre compte.")
            return redirect('paiements:liste_eleves_impayes')
        eleves_qs = Eleve.objects.filter(classe__ecole=ecole_user, statut='ACTIF')

    # Construire un dictionnaire classe_id -> montant total (évite Subquery incompatible MySQL $)
    config_map = {}
    for conf in ConfigurationPaiement.objects.all():
        config_map[conf.classe_id] = (conf.montant_inscription or 0) + (conf.montant_scolarite or 0)

    eleves_qs = (
        eleves_qs
        .select_related('classe', 'classe__ecole', 'responsable_principal')
        .annotate(
            _total_paye=Coalesce(
                Sum('paiements__montant', filter=Q(paiements__statut='VALIDE')),
                Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
                output_field=DecimalField(max_digits=12, decimal_places=0),
            ),
        )
        .order_by('classe__nom', 'nom', 'prenom')
    )

    eleves_avec_impayes = []
    for eleve in eleves_qs[:500]:
        config_total = config_map.get(eleve.classe_id, 0)
        if config_total <= 0:
            continue
        total_paye = int(eleve._total_paye)
        reste = config_total - total_paye
        if reste <= 0:
            continue
        eleve._config_total = config_total
        eleve._total_paye_val = total_paye
        eleve._reste_a_payer = reste
        eleves_avec_impayes.append(eleve)

    if not eleves_avec_impayes:
        messages.info(request, "Aucun élève avec des impayés.")
        return redirect('paiements:liste_eleves_impayes')

    # Fusionner les PDFs individuels avec PyPDF2/pypdf
    try:
        from PyPDF2 import PdfMerger
    except ImportError:
        try:
            from pypdf import PdfMerger
        except ImportError:
            # Fallback: générer un seul PDF (le premier élève)
            response = HttpResponse(content_type='application/pdf')
            filename = f"notes_rappel_tous_{datetime.now().strftime('%Y%m%d')}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            generer_note_rappel_eleve(eleves_avec_impayes[0], response)
            messages.warning(request, "Module PDF fusion indisponible. Seule la première note a été générée.")
            return response

    merger = PdfMerger()
    for eleve in eleves_avec_impayes:
        buf = BytesIO()
        generer_note_rappel_eleve(eleve, buf)
        buf.seek(0)
        merger.append(buf)

    response = HttpResponse(content_type='application/pdf')
    filename = f"notes_rappel_tous_{datetime.now().strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    merger.write(response)
    merger.close()

    messages.success(request, f"{len(eleves_avec_impayes)} notes de rappel générées.")
    return response


def liste_eleves_impayes(request):
    """Affiche la liste des élèves avec des impayés.

    Optimisé: utilise des annotations DB au lieu de boucler avec N+1 requêtes.
    """
    from .models import ConfigurationPaiement

    # IMPORTANT: Seul le superuser peut voir toutes les écoles
    if user_is_superadmin(request.user):
        eleves = Eleve.objects.filter(statut='ACTIF')
    else:
        ecole_user = user_school(request.user)
        if ecole_user is None:
            eleves = Eleve.objects.none()
        else:
            eleves = Eleve.objects.filter(classe__ecole=ecole_user, statut='ACTIF')

    # Construire un dictionnaire classe_id -> montant total (évite Subquery incompatible MySQL $)
    config_map = {}
    for conf in ConfigurationPaiement.objects.all():
        config_map[conf.classe_id] = (conf.montant_inscription or 0) + (conf.montant_scolarite or 0)

    eleves = (
        eleves
        .select_related('classe', 'classe__ecole')
        .annotate(
            _total_paye=Coalesce(
                Sum('paiements__montant', filter=Q(paiements__statut='VALIDE')),
                Value(0, output_field=DecimalField(max_digits=12, decimal_places=0)),
                output_field=DecimalField(max_digits=12, decimal_places=0),
            ),
        )
        .order_by('classe__nom', 'nom', 'prenom')
    )

    eleves_avec_soldes = []
    for eleve in eleves:
        montant_total = config_map.get(eleve.classe_id, 0)
        if montant_total <= 0:
            continue
        montant_paye = int(eleve._total_paye)
        reste = montant_total - montant_paye
        if reste <= 0:
            continue
        eleves_avec_soldes.append({
            'eleve': eleve,
            'montant_total': montant_total,
            'montant_paye': montant_paye,
            'reste_a_payer': reste,
            'pourcentage_paye': int((montant_paye / montant_total * 100)) if montant_total > 0 else 0,
        })

    # Trier par reste à payer décroissant
    eleves_avec_soldes.sort(key=lambda x: x['reste_a_payer'], reverse=True)

    context = {
        'eleves_avec_soldes': eleves_avec_soldes,
        'total_impayes': sum(e['reste_a_payer'] for e in eleves_avec_soldes),
    }

    return render(request, 'paiements/liste_eleves_impayes.html', context)
