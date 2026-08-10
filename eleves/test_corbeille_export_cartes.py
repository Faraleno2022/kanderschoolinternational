"""Tests des fonctions demandées : cartes 8/A4, export réimportable, corbeille."""
import re
from datetime import date, timedelta
from io import BytesIO

from django.conf import settings
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook

from bus.models import AbonnementBus, AbonnementCantine

from .models import Classe, Ecole, Eleve, Responsable


TEST_MIDDLEWARE = tuple(
    item for item in settings.MIDDLEWARE
    if item != 'ecole_moderne.licence_middleware.LicenceMiddleware'
)


def compter_pages_pdf(contenu):
    """Compte les pages d'un PDF sans dépendance externe."""
    return len(re.findall(rb'/Type\s*/Page[^s]', contenu))


@override_settings(MIDDLEWARE=TEST_MIDDLEWARE)
class CartesHuitParPageTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_superuser('admin-cartes', 'admin@test.local', 'secret')
        self.client.force_login(self.user)
        self.ecole = Ecole.objects.create(
            nom='École Fonctions', adresse='Conakry', telephone='+224622000101',
            directeur='Direction', etat='VALIDE',
        )
        self.classe = Classe.objects.create(
            ecole=self.ecole, nom='1ère primaire A', niveau='PRIMAIRE_1',
            code_matricule='T8', annee_scolaire='2026-2027',
        )
        self.pere = Responsable.objects.create(
            prenom='Mamadou', nom='Diallo', relation='PERE',
            telephone='+224622000111', adresse='Kaloum', email='pere@test.local',
        )
        self.mere = Responsable.objects.create(
            prenom='Fatoumata', nom='Bah', relation='MERE',
            telephone='+224622000112', adresse='Kaloum',
        )
        expiration = date.today() + timedelta(days=30)
        for index in range(9):
            eleve = Eleve.objects.create(
                matricule=f'T8-{index + 1:03d}', prenom=f'Eleve{index}',
                nom='Test', sexe='M', classe=self.classe, statut='ACTIF',
                date_naissance=date(2016, 3, 12), lieu_naissance='Conakry',
                responsable_principal=self.pere, responsable_secondaire=self.mere,
            )
            AbonnementBus.objects.create(
                eleve=eleve, montant=10000, date_expiration=expiration,
                statut='ACTIF', zone='Zone A', point_arret='Arrêt A',
            )
            AbonnementCantine.objects.create(
                eleve=eleve, montant=10000, date_expiration=expiration,
                statut='ACTIF', type_repas='DEJEUNER', periodicite='MENSUEL',
            )

    def test_cartes_retrait_bus_cantine_et_scolaires_sont_huit_par_page(self):
        urls = (
            reverse('eleves:tickets_retrait_classe_pdf', args=[self.classe.id]),
            reverse('eleves:tickets_bus_classe_pdf', args=[self.classe.id]),
            reverse('eleves:tickets_cantine_classe_pdf', args=[self.classe.id]),
            reverse('eleves:cartes_scolaires_classe_pdf', args=[self.classe.id]),
        )
        for url in urls:
            with self.subTest(url=url):
                response = self.client.get(url)
                self.assertEqual(response.status_code, 200)
                # 9 élèves => 8 cartes sur la première page, 1 sur la seconde.
                self.assertEqual(compter_pages_pdf(response.content), 2)

    def test_export_complet_est_reimportable_sans_classe_forcee(self):
        response = self.client.get(reverse('eleves:exporter_tous_eleves_modele'))
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook['Élèves']
        headers = [cell.value for cell in sheet[1]]
        self.assertEqual(
            headers,
            [
                'École', 'Classe', 'Année scolaire', 'Matricule', 'Prénom', 'Nom',
                'Sexe', 'Date de Naissance', 'Lieu de Naissance',
                'Nom du Père/Tuteur', 'Prénom du Père/Tuteur', 'Téléphone Principal',
                'Adresse', 'Nom de la Mère', 'Prénom de la Mère',
                'Téléphone Secondaire', 'Email',
            ],
        )
        self.assertEqual(sheet.cell(2, headers.index('Classe') + 1).value, self.classe.nom)
        self.assertEqual(sheet.cell(2, headers.index('École') + 1).value, self.ecole.nom)

        # Le fichier exporté doit être réimportable sans choisir de classe.
        upload = SimpleUploadedFile(
            'eleves_reimportables.xlsx', response.content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        import_response = self.client.post(reverse('eleves:importer_eleves'), {
            'classe_id': '', 'generer_matricules': 'on', 'fichier': upload,
        })
        self.assertRedirects(import_response, reverse('eleves:gestion_classes'))
        self.assertEqual(Eleve.objects.filter(matricule='T8-001').count(), 1)
        self.assertEqual(Eleve.objects.count(), 9)

    def test_carte_cantine_individuelle(self):
        eleve = Eleve.objects.get(matricule='T8-001')
        response = self.client.get(reverse('eleves:ticket_cantine_pdf', args=[eleve.id]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertEqual(compter_pages_pdf(response.content), 1)

    def test_suppression_place_dans_corbeille_et_restaure_sans_perte(self):
        from paiements.models import ModePaiement, Paiement, TypePaiement

        eleve = Eleve.objects.get(matricule='T8-001')
        type_paiement = TypePaiement.objects.create(nom='Test corbeille')
        mode = ModePaiement.objects.create(nom='Espèces test')
        paiement = Paiement.objects.create(
            eleve=eleve, type_paiement=type_paiement, mode_paiement=mode,
            montant=100000, date_paiement=date.today(), statut='VALIDE',
        )
        abonnements_avant = eleve.abonnements_cantine.count()

        # Confirmation utilisateur -> corbeille (aucune donnée détruite).
        response = self.client.post(reverse('eleves:supprimer_eleve', args=[eleve.id]))
        self.assertRedirects(response, reverse('eleves:liste_eleves'))
        eleve.refresh_from_db()
        self.assertTrue(eleve.est_dans_corbeille)
        self.assertTrue(Paiement.objects.filter(pk=paiement.pk).exists())
        self.assertEqual(eleve.abonnements_cantine.count(), abonnements_avant)
        self.assertNotContains(self.client.get(reverse('eleves:liste_eleves')), 'T8-001')
        self.assertContains(self.client.get(reverse('eleves:corbeille_eleves')), 'T8-001')

        # Restauration : statut d'origine récupéré.
        response = self.client.post(reverse('eleves:restaurer_eleve', args=[eleve.id]))
        self.assertRedirects(response, reverse('eleves:corbeille_eleves'))
        eleve.refresh_from_db()
        self.assertFalse(eleve.est_dans_corbeille)
        self.assertEqual(eleve.statut, 'ACTIF')
        self.assertTrue(Paiement.objects.filter(pk=paiement.pk).exists())

    def test_suppression_dans_admin_django_envoie_en_corbeille(self):
        eleve = Eleve.objects.get(matricule='T8-002')
        url = reverse('admin:eleves_eleve_delete', args=[eleve.pk])
        response = self.client.post(url, {'post': 'yes'}, follow=True)
        self.assertEqual(response.status_code, 200)
        eleve.refresh_from_db()
        self.assertTrue(eleve.est_dans_corbeille)
        # L'élève reste visible dans la corbeille de l'admin.
        corbeille = self.client.get(reverse('admin:eleves_elevecorbeille_changelist'))
        self.assertContains(corbeille, 'T8-002')
