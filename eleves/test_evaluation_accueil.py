from datetime import date, timedelta
from io import BytesIO

from django.contrib.auth.models import User
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from paiements.tests.support import TEST_MIDDLEWARE

from .models import Classe, Ecole, Eleve


@override_settings(MIDDLEWARE=TEST_MIDDLEWARE)
class EvaluationAccueilTests(TestCase):
    def setUp(self):
        self.ecole = Ecole.objects.create(
            nom='École évaluation', adresse='Conakry',
            telephone='+224620003001', directeur='Direction', etat='VALIDE',
        )
        self.classe = Classe.objects.create(
            ecole=self.ecole, nom='1ère A', niveau='PRIMAIRE_1',
            annee_scolaire='2026-2027',
        )
        self.user = User.objects.create_user('evaluation', password='secret')
        self.user.profil.role = 'SECRETAIRE'
        self.user.profil.telephone = '+224620003002'
        self.user.profil.ecole = self.ecole
        self.user.profil.is_validated = True
        self.user.profil.save()
        self.old = Eleve.objects.create(
            matricule='EVAL-001', prenom='Ancien', nom='Élève',
            sexe='M', classe=self.classe,
        )
        self.new = Eleve.objects.create(
            matricule='EVAL-002', prenom='Nouveau', nom='Élève',
            sexe='F', classe=self.classe,
        )
        Eleve.objects.filter(pk=self.old.pk).update(
            date_creation=timezone.now() - timedelta(days=2),
        )
        self.client.force_login(self.user)

    def test_liste_place_le_dernier_ajoute_en_premier(self):
        response = self.client.get(reverse('eleves:liste_eleves'))
        ids = [student.pk for student in response.context['page_obj'].object_list]
        self.assertEqual(ids[:2], [self.new.pk, self.old.pk])

    def test_pointage_et_filtre_du_test_accueil(self):
        response = self.client.post(
            reverse('eleves:pointer_test_accueil', args=[self.new.pk]),
            {'evalue': '1'},
        )
        self.assertRedirects(response, reverse('eleves:liste_eleves'))
        self.new.refresh_from_db()
        self.assertTrue(self.new.test_accueil_evalue)
        self.assertIsNotNone(self.new.date_evaluation_accueil)

        evaluated = self.client.get(reverse('eleves:liste_eleves'), {'evaluation': 'evalue'})
        ids = [student.pk for student in evaluated.context['page_obj'].object_list]
        self.assertEqual(ids, [self.new.pk])

        not_evaluated = self.client.get(
            reverse('eleves:liste_eleves'), {'evaluation': 'non_evalue'},
        )
        ids = [student.pk for student in not_evaluated.context['page_obj'].object_list]
        self.assertEqual(ids, [self.old.pk])

    def test_exports_excel_et_pdf_respectent_le_statut(self):
        self.new.test_accueil_evalue = True
        self.new.date_evaluation_accueil = timezone.now()
        self.new.save(update_fields=['test_accueil_evalue', 'date_evaluation_accueil'])

        excel = self.client.get(reverse('eleves:export_tests_accueil_excel', args=['evalues']))
        self.assertEqual(excel.status_code, 200)
        workbook = load_workbook(BytesIO(excel.content))
        values = [row[0].value for row in workbook.active.iter_rows(min_row=3)]
        self.assertIn('EVAL-002', values)
        self.assertNotIn('EVAL-001', values)

        pdf = self.client.get(reverse('eleves:export_tests_accueil_pdf', args=['non-evalues']))
        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(pdf['Content-Type'], 'application/pdf')
        self.assertTrue(pdf.content.startswith(b'%PDF'))
