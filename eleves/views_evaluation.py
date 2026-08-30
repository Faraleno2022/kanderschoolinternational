"""Pointage et exports du test d'accueil des élèves."""

from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from ecole_moderne.pdf_utils import draw_logo_watermark
from eleves.models import Eleve
from eleves.utils_annee import get_annee_active
from utilisateurs.utils import filter_by_user_school, user_school


def _evaluation_value(value):
    return str(value or '').lower() in {'1', 'true', 'evalue', 'oui'}


def _evaluation_queryset(request, evaluated):
    queryset = (
        Eleve.objects.select_related('classe', 'classe__ecole')
        .filter(est_dans_corbeille=False, test_accueil_evalue=evaluated)
    )
    queryset = filter_by_user_school(queryset, request.user, 'classe__ecole')
    school = user_school(request.user)
    if school:
        active_year = get_annee_active(request, school)
        if active_year:
            queryset = queryset.filter(classe__annee_scolaire=active_year)
    class_id = (request.GET.get('classe_id') or '').strip()
    if class_id.isdigit():
        queryset = queryset.filter(classe_id=int(class_id))
    query = (request.GET.get('recherche') or request.GET.get('q') or '').strip()
    if query:
        queryset = queryset.filter(
            Q(matricule__icontains=query)
            | Q(nom__icontains=query)
            | Q(prenom__icontains=query)
            | Q(classe__nom__icontains=query)
        )
    return queryset.order_by('-date_creation', '-id'), school


@login_required
@require_POST
def pointer_test_accueil(request, eleve_id):
    queryset = filter_by_user_school(
        Eleve.objects.filter(est_dans_corbeille=False),
        request.user,
        'classe__ecole',
    )
    student = get_object_or_404(queryset, pk=eleve_id)
    evaluated = _evaluation_value(request.POST.get('evalue'))
    student.test_accueil_evalue = evaluated
    student.date_evaluation_accueil = timezone.now() if evaluated else None
    student.save(update_fields=[
        'test_accueil_evalue', 'date_evaluation_accueil', 'date_modification',
    ])
    state = 'évalué' if evaluated else 'non évalué'
    messages.success(request, f"{student.nom_complet} est maintenant marqué(e) {state}.")
    next_url = request.POST.get('next') or reverse('eleves:liste_eleves')
    if not url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        next_url = reverse('eleves:liste_eleves')
    return redirect(next_url)


def _filename(evaluated, extension):
    state = 'evalues' if evaluated else 'non_evalues'
    return f'tests_accueil_{state}_{timezone.localdate():%Y-%m-%d}.{extension}'


@login_required
def export_tests_accueil_excel(request, statut):
    evaluated = statut == 'evalues'
    queryset, _school = _evaluation_queryset(request, evaluated)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Tests accueil'
    title = "ÉLÈVES ÉVALUÉS" if evaluated else "ÉLÈVES NON ÉVALUÉS"
    sheet.merge_cells('A1:H1')
    sheet['A1'] = f"TEST D'ACCUEIL — {title}"
    sheet['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    sheet['A1'].fill = PatternFill('solid', fgColor='1F4E78')
    sheet['A1'].alignment = Alignment(horizontal='center')
    sheet.append(['Matricule', 'Nom', 'Prénom', 'Sexe', 'Classe', 'École', 'Année', 'Date évaluation'])
    for cell in sheet[2]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='5B9BD5')
    for student in queryset:
        sheet.append([
            student.matricule, student.nom, student.prenom,
            student.get_sexe_display(), student.classe.nom,
            student.classe.ecole.nom, student.classe.annee_scolaire,
            timezone.localtime(student.date_evaluation_accueil).strftime('%d/%m/%Y %H:%M')
            if student.date_evaluation_accueil else '-',
        ])
    for index, width in enumerate([18, 22, 22, 12, 26, 35, 14, 22], 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = 'A3'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{_filename(evaluated, "xlsx")}"'
    workbook.save(response)
    return response


@login_required
def export_tests_accueil_pdf(request, statut):
    evaluated = statut == 'evalues'
    queryset, school = _evaluation_queryset(request, evaluated)
    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer, pagesize=landscape(A4), rightMargin=1.2*cm,
        leftMargin=1.2*cm, topMargin=1.1*cm, bottomMargin=1.2*cm,
    )
    styles = getSampleStyleSheet()
    state_label = 'ÉLÈVES ÉVALUÉS' if evaluated else 'ÉLÈVES NON ÉVALUÉS'
    story = [
        Paragraph(f"<b>TEST D'ACCUEIL — {state_label}</b>", styles['Title']),
        Paragraph(f"Édité le {timezone.localtime():%d/%m/%Y à %H:%M}", styles['Normal']),
        Spacer(1, .35*cm),
    ]
    data = [['Matricule', 'Élève', 'Sexe', 'Classe', 'École', 'Année', 'Date évaluation']]
    for student in queryset:
        data.append([
            student.matricule, student.nom_complet, student.get_sexe_display(),
            student.classe.nom, student.classe.ecole.nom,
            student.classe.annee_scolaire,
            timezone.localtime(student.date_evaluation_accueil).strftime('%d/%m/%Y %H:%M')
            if student.date_evaluation_accueil else '-',
        ])
    table = Table(data, repeatRows=1, colWidths=[2.7*cm, 4.5*cm, 2*cm, 4.2*cm, 6*cm, 2.5*cm, 3.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), .35, colors.HexColor('#B8C4CE')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F6F9')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(table)

    def decorate_page(canvas, doc):
        try:
            draw_logo_watermark(canvas, doc.pagesize[0], doc.pagesize[1], ecole=school)
        except Exception:
            pass
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.drawRightString(doc.pagesize[0] - 1.2*cm, .55*cm, f'Page {doc.page}')
        canvas.restoreState()

    document.build(story, onFirstPage=decorate_page, onLaterPages=decorate_page)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{_filename(evaluated, "pdf")}"'
    return response
