from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth import login as auth_login
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.conf import settings
from django.core.paginator import Paginator
from django.db.models import Q, Count, Case, When, IntegerField
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.urls import reverse
from django.utils import timezone
from django.utils.text import slugify
from django.db import transaction
from django.core.cache import cache
from django.views.decorators.cache import never_cache
from datetime import date
import os
import io
import logging

logger = logging.getLogger(__name__)
from .models import Eleve, Responsable, Classe, Ecole, HistoriqueEleve, GrilleTarifaire
from .forms import EleveForm, ResponsableForm, RechercheEleveForm, ClasseForm, EcoleForm
from utilisateurs.forms import SignupInlineForm
from utilisateurs.models import JournalActivite
from utilisateurs.utils import user_is_admin, user_is_superadmin, filter_by_user_school, user_school
from .utils_annee import get_annee_active
from django.views.decorators.cache import cache_page
from django.views.decorators.vary import vary_on_cookie

# ReportLab pour génération PDF
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Utilitaire PDF partagé (filigrane)
from ecole_moderne.pdf_utils import draw_logo_watermark
from ecole_moderne.security_decorators import delete_permission_required

# Excel
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None
    get_column_letter = None

@login_required
def liste_eleves(request):
    """Vue optimisée pour afficher la liste des élèves avec cache intelligent"""
    from ecole_moderne.query_optimizer import QueryOptimizer, PaginationOptimizer
    from ecole_moderne.decorators import cache_user_data
    
    form_recherche = RechercheEleveForm(request.GET or None)
    
    # Cache de l'école utilisateur
    user_school_cache_key = f'user_school_{request.user.id}'
    user_school_obj = cache.get(user_school_cache_key)
    if user_school_obj is None and not user_is_admin(request.user):
        user_school_obj = user_school(request.user)
        if user_school_obj:
            cache.set(user_school_cache_key, user_school_obj, 300)
    
    # Queryset optimisé avec relations pré-chargées
    eleves = QueryOptimizer.get_optimized_eleves(
        school=user_school_obj if not user_is_admin(request.user) else None,
        with_payments=True,
        with_classes=True
    )

    # Filtrer par année scolaire active
    if not user_is_admin(request.user) and user_school_obj:
        annee_active = get_annee_active(request, user_school_obj)
        if annee_active:
            eleves = eleves.filter(classe__annee_scolaire=annee_active)

    # Application des filtres optimisés
    if form_recherche.is_valid():
        recherche = form_recherche.cleaned_data.get('recherche')
        if recherche:
            eleves = eleves.filter(
                Q(matricule__icontains=recherche) |
                Q(nom__icontains=recherche) |
                Q(prenom__icontains=recherche) |
                Q(classe__nom__icontains=recherche) |
                Q(classe__ecole__nom__icontains=recherche) |
                Q(responsable_principal__nom__icontains=recherche) |
                Q(responsable_principal__prenom__icontains=recherche)
            )
    
    # Filtre par classe optimisé
    classe_id = request.GET.get('classe_id')
    if classe_id:
        try:
            int(classe_id)
            eleves = eleves.filter(classe_id=classe_id)
        except (TypeError, ValueError):
            classe_id = None
    
    # Statistiques optimisées avec cache
    stats_cache_key = f'eleves_stats_{request.user.id}_{hash(str(eleves.query))}'
    stats = cache.get(stats_cache_key)
    
    if stats is None:
        stats = eleves.aggregate(
            total_eleves=Count('id'),
            eleves_actifs=Count(Case(When(statut='ACTIF', then=1), output_field=IntegerField())),
            eleves_suspendus=Count(Case(When(statut='SUSPENDU', then=1), output_field=IntegerField())),
            eleves_exclus=Count(Case(When(statut='EXCLU', then=1), output_field=IntegerField()))
        )
        cache.set(stats_cache_key, stats, 120)  # Cache 2 minutes
    
    # Pagination optimisée
    page_number = request.GET.get('page', 1)
    page_obj, paginator = PaginationOptimizer.optimize_pagination(
        eleves.order_by('nom', 'prenom'), 
        page_number, 
        per_page=15
    )
    
    # Log de l'activité
    JournalActivite.objects.create(
        user=request.user,
        action='CONSULTATION',
        type_objet='ELEVE',
        description=f"Consultation de la liste des élèves (page {page_number or 1})",
        adresse_ip=request.META.get('REMOTE_ADDR', ''),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )
    
    # Liste des classes pour export (restreinte si besoin)
    # IMPORTANT: Seul le superuser peut voir toutes les écoles
    if user_is_superadmin(request.user):
        classes = (
            Classe.objects.select_related('ecole')
            .filter(ecole__etat='VALIDE')
            .order_by('ecole__nom', 'niveau', 'nom')
        )
    else:
        # Tous les autres utilisateurs (y compris ADMIN d'école) ne voient que leur école
        user_ecole = user_school(request.user)
        if user_ecole is None:
            classes = Classe.objects.none()
        else:
            annee_active = get_annee_active(request, user_ecole)
            qs_filter = {'ecole': user_ecole, 'ecole__etat': 'VALIDE'}
            if annee_active:
                qs_filter['annee_scolaire'] = annee_active
            classes = (
                Classe.objects.select_related('ecole')
                .filter(**qs_filter)
                .order_by('niveau', 'nom')
            )
            # Fallback: si l'école de l'utilisateur n'est pas encore validée, proposer quand même ses classes
            if not classes.exists():
                qs_fallback = {'ecole': user_ecole}
                if annee_active:
                    qs_fallback['annee_scolaire'] = annee_active
                classes = (
                    Classe.objects.select_related('ecole')
                    .filter(**qs_fallback)
                    .order_by('niveau', 'nom')
                )

    context = {
        'page_obj': page_obj,
        'form_recherche': form_recherche,
        'stats': stats,
        'titre_page': 'Gestion des Élèves',
        'classes': classes,
        # Conserver la sélection actuelle de classe dans l'UI
        'selected_classe_id': str(classe_id) if classe_id else '',
    }

    # Rendu partiel pour la recherche dynamique
    if request.GET.get('partial') == '1' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        response = render(request, 'eleves/partials/_liste_eleves_zone.html', context)
        # Eviter le cache sur le fragment
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response

    # Rendu complet de la page
    response = render(request, 'eleves/liste_eleves.html', context)
    # Headers anti-cache pour s'assurer que la liste est toujours fraîche
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response

def creer_ecole(request):
    """Vue (hors admin Django) pour créer une école avec logo.
    Accessible à tout utilisateur authentifié. L'école est créée en statut EN_ATTENTE
    pour validation par un administrateur qui configurera ensuite les échéances.
    """
    signup_form = None
    if request.method == 'POST':
        form = EcoleForm(request.POST, request.FILES)
        # Si l'utilisateur n'est pas connecté, traiter l'inscription intégrée en amont
        if not request.user.is_authenticated:
            signup_form = SignupInlineForm(request.POST)
            if signup_form.is_valid():
                # Créer l'utilisateur et le connecter
                try:
                    user = signup_form.save(commit=False)
                    # Enregistrer l'email saisi si fourni
                    try:
                        user.email = (signup_form.cleaned_data.get('email') or '').strip()
                    except Exception:
                        pass
                    # Marquer le compte comme inactif en attente de validation par un administrateur
                    user.is_active = False
                    user.save()
                    
                    # Créer le profil utilisateur avec validation en attente (éviter les doublons)
                    from utilisateurs.models import Profil
                    profil, created = Profil.objects.get_or_create(
                        user=user,
                        defaults={
                            'role': 'DIRECTEUR',  # Par défaut, créateur d'école = directeur
                            'telephone': '',  # À compléter lors de la validation
                            'is_validated': False,  # En attente de validation administrative
                            'actif': False
                        }
                    )
                    
                    # Ne PAS connecter l'utilisateur. Informer de la validation requise.
                    messages.success(request, "Compte créé. Un administrateur doit valider votre demande avant connexion.")
                except Exception as e:
                    messages.error(request, f"Erreur lors de la création du compte: {e}")
                    return render(request, 'eleves/creer_ecole.html', {
                        'form': form,
                        'signup_form': signup_form,
                        'titre_page': "Créer une École",
                    })
            else:
                # Renvoyer le formulaire avec erreurs sans tenter la création d'école
                messages.error(request, "Veuillez corriger les erreurs du formulaire de compte.")
                return render(request, 'eleves/creer_ecole.html', {
                    'form': form,
                    'signup_form': signup_form,
                    'titre_page': "Créer une École",
                })

        if form.is_valid():
            try:
                ecole = form.save(commit=False)
                # Envoi en validation
                try:
                    # Si un compte vient d'être créé (anonyme au départ), lier l'école à ce compte
                    if not request.user.is_authenticated and signup_form and signup_form.is_valid():
                        # On relit l'utilisateur créé ci-dessus via le username du formulaire
                        from django.contrib.auth import get_user_model
                        UserModel = get_user_model()
                        created_username = signup_form.cleaned_data.get('username')
                        try:
                            created_user = UserModel.objects.get(username=created_username)
                        except UserModel.DoesNotExist:
                            created_user = None
                        ecole.created_by = created_user
                    else:
                        ecole.created_by = request.user if request.user.is_authenticated else None
                except Exception:
                    pass
                # etat par défaut EN_ATTENTE via modèle
                ecole.save()
                
                # Traiter la création de classes fournies dans ce formulaire (si présentes)
                try:
                    from decimal import Decimal
                    classe_noms = request.POST.getlist('classe_nom[]')
                    classe_niveaux = request.POST.getlist('classe_niveau[]')
                    classe_annees = request.POST.getlist('classe_annee[]')
                    classe_caps = request.POST.getlist('classe_capacite[]')
                    for idx, nom in enumerate(classe_noms):
                        nom = (nom or '').strip()
                        niveau = (classe_niveaux[idx] if idx < len(classe_niveaux) else '').strip()
                        annee = (classe_annees[idx] if idx < len(classe_annees) else '').strip()
                        cap_raw = (classe_caps[idx] if idx < len(classe_caps) else '30')
                        capacite = 30
                        try:
                            capacite = int(cap_raw or '30')
                        except Exception:
                            capacite = 30
                        if nom and niveau and annee:
                            try:
                                Classe.objects.create(ecole=ecole, nom=nom, niveau=niveau, annee_scolaire=annee, capacite_max=capacite)
                            except Exception:
                                pass
                except Exception:
                    pass

                # Traiter la création des grilles tarifaires fournies (si présentes)
                try:
                    niveaux = request.POST.getlist('grille_niveau[]')
                    annees = request.POST.getlist('grille_annee[]')
                    frais_insp_list = request.POST.getlist('grille_inscription[]')
                    frais_reinsp_list = request.POST.getlist('grille_reinscription[]')
                    t1_list = request.POST.getlist('grille_t1[]')
                    t2_list = request.POST.getlist('grille_t2[]')
                    t3_list = request.POST.getlist('grille_t3[]')
                    p1_list = request.POST.getlist('grille_p1[]')
                    p2_list = request.POST.getlist('grille_p2[]')
                    p3_list = request.POST.getlist('grille_p3[]')
                    # Nouvelles dates d'échéance par défaut (optionnelles)
                    ech_insc_list = request.POST.getlist('grille_ech_insc[]')
                    ech_t1_list = request.POST.getlist('grille_ech_t1[]')
                    ech_t2_list = request.POST.getlist('grille_ech_t2[]')
                    ech_t3_list = request.POST.getlist('grille_ech_t3[]')
                    for i in range(len(niveaux)):
                        niveau = (niveaux[i] or '').strip()
                        annee = (annees[i] if i < len(annees) else '').strip()
                        if not niveau or not annee:
                            continue
                        try:
                            from datetime import date as _date
                            def _pdate(s):
                                try:
                                    s = (s or '').strip()
                                    return _date.fromisoformat(s) if s else None
                                except Exception:
                                    return None
                            d_insc = _pdate(ech_insc_list[i] if i < len(ech_insc_list) else None)
                            d_t1 = _pdate(ech_t1_list[i] if i < len(ech_t1_list) else None)
                            d_t2 = _pdate(ech_t2_list[i] if i < len(ech_t2_list) else None)
                            d_t3 = _pdate(ech_t3_list[i] if i < len(ech_t3_list) else None)
                            # Validation ordre croissant si plusieurs dates sont fournies
                            seq = [d for d in [d_insc, d_t1, d_t2, d_t3] if d is not None]
                            if len(seq) > 1 and seq != sorted(seq):
                                messages.error(request, f"Échéances invalides pour la grille {annee} {niveau}: l'ordre doit être Inscription ≤ T1 ≤ T2 ≤ T3.")
                            else:
                                GrilleTarifaire.objects.update_or_create(
                                    ecole=ecole,
                                    niveau=niveau,
                                    annee_scolaire=annee,
                                    defaults={
                                        'frais_inscription': Decimal((frais_insp_list[i] if i < len(frais_insp_list) else '0') or '0'),
                                        'frais_reinscription': Decimal((frais_reinsp_list[i] if i < len(frais_reinsp_list) else '0') or '0'),
                                        'tranche_1': Decimal((t1_list[i] if i < len(t1_list) else '0') or '0'),
                                        'tranche_2': Decimal((t2_list[i] if i < len(t2_list) else '0') or '0'),
                                        'tranche_3': Decimal((t3_list[i] if i < len(t3_list) else '0') or '0'),
                                        'periode_1': (p1_list[i] if i < len(p1_list) else "À l'inscription") or "À l'inscription",
                                        'periode_2': (p2_list[i] if i < len(p2_list) else 'Début janvier') or 'Début janvier',
                                        'periode_3': (p3_list[i] if i < len(p3_list) else 'Début mars') or 'Début mars',
                                        'date_echeance_inscription_defaut': d_insc,
                                        'date_echeance_tranche_1_defaut': d_t1,
                                        'date_echeance_tranche_2_defaut': d_t2,
                                        'date_echeance_tranche_3_defaut': d_t3,
                                    }
                                )
                        except Exception:
                            pass
                except Exception:
                    pass

                # Si aucune classe/grille fournie, créer automatiquement les classes et grilles par défaut
                try:
                    from .utils import creer_classes_et_grilles_par_defaut
                    # Ne créer les défauts que si aucune grille n'a été soumise
                    a_deja_grilles = GrilleTarifaire.objects.filter(ecole=ecole).exists()
                    if not a_deja_grilles:
                        result = creer_classes_et_grilles_par_defaut(ecole)
                        if result['total_classes'] > 0:
                            messages.info(request, f"{result['total_classes']} classes et {result['total_grilles']} grilles tarifaires créées automatiquement.")
                except Exception as e:
                    messages.warning(request, f"École créée mais erreur lors de la création des classes par défaut: {e}")
                if request.user.is_authenticated:
                    messages.success(request, f"École '{ecole.nom}' créée en brouillon. Vous pouvez maintenant ajouter les classes et les échéances, puis soumettre pour validation.")
                    return redirect('eleves:configurer_ecole', ecole_id=ecole.id)
            except Exception as e:
                messages.error(request, f"Erreur lors de la création de l'école: {e}")
        else:
            messages.error(request, "Veuillez corriger les erreurs du formulaire.")
    else:
        form = EcoleForm()
        if not request.user.is_authenticated:
            signup_form = SignupInlineForm()

    return render(request, 'eleves/creer_ecole.html', {
        'form': form,
        'signup_form': signup_form,
        'titre_page': "Créer une École",
        'action': 'Créer',
    })


def creer_ecole_confirmation(request):
    """Redirige vers l'accueil après la confirmation de demande de création d'école."""
    try:
        messages.info(request, "Demande soumise. Un administrateur procédera à la validation.")
    except Exception:
        pass
    return redirect('home')


def _user_can_edit_school(request, ecole: Ecole) -> bool:
    """Autoriser édition si admin ou créateur, ou si école en BROUILLON/EN_ATTENTE."""
    if user_is_admin(request.user):
        return True
    try:
        if ecole.created_by_id and request.user.is_authenticated and ecole.created_by_id == request.user.id:
            return True
        # Cas d'une école créée anonymement: autoriser la configuration tant que BROUILLON
        if not ecole.created_by_id and ecole.etat == 'BROUILLON':
            return True
    except Exception:
        pass
    return False


def configurer_ecole(request, ecole_id: int):
    """Assistant de configuration d'une école (brouillon):
    - Ajouter des classes
    - Définir les grilles tarifaires
    - Soumettre pour validation (etat -> EN_ATTENTE)
    Accès: admin ou créateur tant que non validée.
    """
    ecole = get_object_or_404(Ecole, pk=ecole_id)

    # Droit d'accès: admin ou créateur. Après validation, vue en lecture seule.
    can_edit = _user_can_edit_school(request, ecole) and ecole.etat in ("BROUILLON", "EN_ATTENTE")

    message_info = None
    if request.method == 'POST' and can_edit:
        action = request.POST.get('action')
        try:
            if action == 'add_classe':
                nom = (request.POST.get('classe_nom') or '').strip()
                niveau = request.POST.get('classe_niveau')
                annee = (request.POST.get('classe_annee') or '').strip()
                capacite = int(request.POST.get('classe_capacite') or '30')
                if not nom or not niveau or not annee:
                    messages.error(request, "Veuillez renseigner nom, niveau et année scolaire pour la classe.")
                else:
                    Classe.objects.create(ecole=ecole, nom=nom, niveau=niveau, annee_scolaire=annee, capacite_max=capacite)
                    messages.success(request, f"Classe '{nom}' ajoutée.")
                return redirect('eleves:configurer_ecole', ecole_id=ecole.id)

            if action == 'add_grille':
                niveau = request.POST.get('grille_niveau')
                annee = (request.POST.get('grille_annee') or '').strip()
                frais_insp = request.POST.get('grille_inscription') or '0'
                frais_reinsp = request.POST.get('grille_reinscription') or '0'
                t1 = request.POST.get('grille_t1') or '0'
                t2 = request.POST.get('grille_t2') or '0'
                t3 = request.POST.get('grille_t3') or '0'
                p1 = request.POST.get('grille_p1') or "À l'inscription"
                p2 = request.POST.get('grille_p2') or 'Début janvier'
                p3 = request.POST.get('grille_p3') or 'Début mars'
                # Dates d'échéance par défaut (optionnelles)
                ech_insc = request.POST.get('grille_ech_insc')
                ech_t1 = request.POST.get('grille_ech_t1')
                ech_t2 = request.POST.get('grille_ech_t2')
                ech_t3 = request.POST.get('grille_ech_t3')
                if not niveau or not annee:
                    messages.error(request, "Veuillez renseigner le niveau et l'année scolaire pour la grille.")
                else:
                    from decimal import Decimal
                    from datetime import date as _date
                    def _pdate_one(s):
                        try:
                            s = (s or '').strip()
                            return _date.fromisoformat(s) if s else None
                        except Exception:
                            return None
                    d_insc = _pdate_one(ech_insc)
                    d_t1 = _pdate_one(ech_t1)
                    d_t2 = _pdate_one(ech_t2)
                    d_t3 = _pdate_one(ech_t3)
                    seq = [d for d in [d_insc, d_t1, d_t2, d_t3] if d is not None]
                    if len(seq) > 1 and seq != sorted(seq):
                        messages.error(request, "Échéances invalides: l'ordre doit être Inscription ≤ T1 ≤ T2 ≤ T3.")
                        return redirect('eleves:configurer_ecole', ecole_id=ecole.id)
                    GrilleTarifaire.objects.update_or_create(
                        ecole=ecole, niveau=niveau, annee_scolaire=annee,
                        defaults={
                            'frais_inscription': Decimal(frais_insp or '0'),
                            'frais_reinscription': Decimal(frais_reinsp or '0'),
                            'tranche_1': Decimal(t1 or '0'),
                            'tranche_2': Decimal(t2 or '0'),
                            'tranche_3': Decimal(t3 or '0'),
                            'periode_1': p1,
                            'periode_2': p2,
                            'periode_3': p3,
                            'date_echeance_inscription_defaut': d_insc,
                            'date_echeance_tranche_1_defaut': d_t1,
                            'date_echeance_tranche_2_defaut': d_t2,
                            'date_echeance_tranche_3_defaut': d_t3,
                        }
                    )
                    messages.success(request, f"Grille tarifaire {annee} pour le niveau sélectionné enregistrée.")
                return redirect('eleves:configurer_ecole', ecole_id=ecole.id)

            if action == 'submit_validation':
                ecole.etat = 'EN_ATTENTE'
                ecole.save(update_fields=['etat'])
                # Afficher une page de confirmation dédiée qui redirige ensuite vers l'accueil
                return render(request, 'eleves/confirmation_redirection.html', {
                    'message': "Demande soumise. Un administrateur procédera à la validation.",
                    'redirect_url': reverse('home'),
                    'titre_page': 'Demande envoyée'
                })

        except Exception as e:
            messages.error(request, f"Erreur: {e}")
            return redirect('eleves:configurer_ecole', ecole_id=ecole.id)

    # Données d'affichage (filtrées par année active)
    annee_active = get_annee_active(request, ecole)
    if annee_active:
        classes = Classe.objects.filter(ecole=ecole, annee_scolaire=annee_active).order_by('niveau', 'nom')
        grilles = GrilleTarifaire.objects.filter(ecole=ecole, annee_scolaire=annee_active).order_by('niveau')
    else:
        classes = Classe.objects.filter(ecole=ecole).order_by('niveau', 'nom')
        grilles = GrilleTarifaire.objects.filter(ecole=ecole).order_by('annee_scolaire', 'niveau')

    # Niveaux affichables (depuis modèle Classe)
    niveaux = getattr(Classe, 'NIVEAUX_CHOICES', [])

    return render(request, 'eleves/configurer_ecole.html', {
        'ecole': ecole,
        'can_edit': can_edit,
        'classes': classes,
        'grilles': grilles,
        'niveaux': niveaux,
        'titre_page': f"Configurer l'École - {ecole.nom}",
    })

@login_required
def detail_eleve(request, eleve_id):
    """Vue pour afficher les détails d'un élève"""
    qs = Eleve.objects.select_related(
        'classe', 'classe__ecole', 'responsable_principal', 'responsable_secondaire'
    ).prefetch_related('paiements', 'historique')
    if not user_is_admin(request.user):
        qs = filter_by_user_school(qs, request.user, 'classe__ecole')
    eleve = get_object_or_404(qs, id=eleve_id)
    
    # Statistiques des paiements
    paiements_stats = {
        'total_paiements': eleve.paiements.count(),
        'paiements_valides': eleve.paiements.filter(statut='VALIDE').count(),
        'montant_total': sum(p.montant for p in eleve.paiements.filter(statut='VALIDE')),
    }
    
    # Historique récent
    historique_recent = eleve.historique.all()[:10]
    
    # Log de l'activité
    JournalActivite.objects.create(
        user=request.user,
        action='CONSULTATION',
        type_objet='ELEVE',
        objet_id=eleve.id,
        description=f"Consultation du profil de {eleve.nom_complet}",
        adresse_ip=request.META.get('REMOTE_ADDR', ''),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )
    
    context = {
        'eleve': eleve,
        'paiements_stats': paiements_stats,
        'historique_recent': historique_recent,
        'titre_page': f'Profil de {eleve.nom_complet}'
    }
    
    return render(request, 'eleves/detail_eleve.html', context)

@login_required
@never_cache
def ajouter_eleve(request):
    """Vue optimisée pour ajouter un nouvel élève avec enregistrement ultra-rapide"""
    # Cache de l'école utilisateur pour éviter les requêtes répétées
    user_school_cache_key = f'user_school_{request.user.id}'
    user_school_obj = cache.get(user_school_cache_key)
    
    if user_school_obj is None and not user_is_admin(request.user):
        user_school_obj = user_school(request.user)
        if user_school_obj:
            cache.set(user_school_cache_key, user_school_obj, 300)  # Cache 5 minutes
    
    # Vérification d'accès rapide
    if not user_is_admin(request.user) and user_school_obj is None:
        return render(request, 'utilisateurs/acces_refuse_ecole.html', status=403)
    
    if request.method == 'POST':
        # Optimisation: Créer les formulaires seulement si nécessaire
        form = EleveForm(request.POST, request.FILES, user=request.user)
        
        # Cache des classes pour éviter les requêtes répétées (filtrées par année active)
        if not user_is_admin(request.user):
            annee_active = get_annee_active(request, user_school_obj)
            classes_cache_key = f'classes_ecole_{user_school_obj.id}_{annee_active}'
            classes_qs = cache.get(classes_cache_key)
            if classes_qs is None:
                qs = Classe.objects.filter(ecole=user_school_obj).select_related('ecole')
                if annee_active:
                    qs = qs.filter(annee_scolaire=annee_active)
                classes_qs = qs
                cache.set(classes_cache_key, classes_qs, 600)  # Cache 10 minutes
            form.fields['classe'].queryset = classes_qs
        
        # Gestion optimisée des responsables
        responsable_principal_form = None
        responsable_secondaire_form = None
        
        creer_resp_principal = request.POST.get('responsable_principal_nouveau') == 'on'
        creer_resp_secondaire = request.POST.get('responsable_secondaire_nouveau') == 'on'
        
        if creer_resp_principal:
            responsable_principal_form = ResponsableForm(request.POST, prefix='resp_principal', user=request.user)
        
        if creer_resp_secondaire:
            responsable_secondaire_form = ResponsableForm(request.POST, prefix='resp_secondaire', user=request.user)
        
        # Validation rapide
        form_valide = form.is_valid()
        resp_principal_valide = responsable_principal_form.is_valid() if responsable_principal_form else True
        resp_secondaire_valide = responsable_secondaire_form.is_valid() if responsable_secondaire_form else True
        
        # Exigences minimales côté serveur pour un nouveau responsable principal
        # (évite les erreurs de contrainte NOT NULL au moment du save())
        if creer_resp_principal and responsable_principal_form:
            try:
                cd = responsable_principal_form.cleaned_data if responsable_principal_form.is_valid() else {}
            except Exception:
                cd = {}
            required_min_fields = ['prenom', 'nom', 'relation', 'telephone', 'adresse']
            for field in required_min_fields:
                val = (cd.get(field) or '').strip() if isinstance(cd.get(field), str) else cd.get(field)
                if not val:
                    responsable_principal_form.add_error(field, "Ce champ est obligatoire.")
                    resp_principal_valide = False
        
        # Responsable principal maintenant optionnel - pas de validation obligatoire
        # Les écoles peuvent ajouter des élèves sans responsable si nécessaire
        
        if form_valide and resp_principal_valide and resp_secondaire_valide:
            # Transaction atomique pour performance maximale
            try:
                with transaction.atomic():
                    # Créer les responsables en une seule transaction
                    if responsable_principal_form:
                        responsable_principal = responsable_principal_form.save()
                        form.instance.responsable_principal = responsable_principal
                    
                    if responsable_secondaire_form:
                        responsable_secondaire = responsable_secondaire_form.save()
                        form.instance.responsable_secondaire = responsable_secondaire
                    
                    # Sauvegarder l'élève
                    eleve = form.save(commit=False)
                    eleve.cree_par = request.user
                    
                    # Gérer la saisie manuelle du matricule
                    saisie_manuelle = form.cleaned_data.get('saisie_manuelle_matricule', False)
                    if saisie_manuelle and form.cleaned_data.get('matricule'):
                        # Si saisie manuelle, le matricule a déjà été validé dans le formulaire
                        eleve.matricule = form.cleaned_data['matricule']
                        # Marquer pour éviter la génération automatique
                        eleve._skip_matricule_generation = True
                    
                    eleve.save()
                    
                    # Créer historique et journal en batch (plus rapide)
                    HistoriqueEleve.objects.create(
                        eleve=eleve,
                        action='CREATION',
                        description=f"Création du profil de {eleve.prenom} {eleve.nom}",
                        utilisateur=request.user
                    )
                    
                    # Log optimisé
                    JournalActivite.objects.create(
                        user=request.user,
                        action='CREATION',
                        type_objet='ELEVE',
                        objet_id=eleve.id,
                        description=f"Création de l'élève {eleve.prenom} {eleve.nom} (matricule: {eleve.matricule})",
                        adresse_ip=request.META.get('REMOTE_ADDR', ''),
                        user_agent=request.META.get('HTTP_USER_AGENT', '')[:200]  # Limiter la taille
                    )
                    
                    # Invalider les caches pertinents
                    cache.delete_many([
                        f'stats_eleves_{request.user.id}',
                        f'classes_ecole_{user_school_obj.id if user_school_obj else "admin"}'
                    ])
                    
                messages.success(request, f"L'eleve {eleve.prenom} {eleve.nom} a ete ajoute avec succes (Matricule: {eleve.matricule}). Vous pouvez ajouter un autre eleve.")
                # Rediriger vers le formulaire d'ajout en preservant la classe selectionnee
                url = reverse('eleves:ajouter_eleve')
                classe_id = form.cleaned_data.get('classe')
                if classe_id:
                    url += f'?classe_id={classe_id.id if hasattr(classe_id, "id") else classe_id}'
                return redirect(url)
                
            except Exception as e:
                logger.exception("Erreur lors de l'enregistrement d'un élève")
                messages.error(request, "Une erreur est survenue lors de l'enregistrement.")
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        # GET: Initialisation optimisée des formulaires
        form = EleveForm(user=request.user)

        # Cache des classes pour le formulaire GET (filtrées par année active)
        if not user_is_admin(request.user) and user_school_obj:
            annee_active = get_annee_active(request, user_school_obj)
            classes_cache_key = f'classes_ecole_{user_school_obj.id}_{annee_active}'
            classes_qs = cache.get(classes_cache_key)
            if classes_qs is None:
                qs = Classe.objects.filter(ecole=user_school_obj).select_related('ecole')
                if annee_active:
                    qs = qs.filter(annee_scolaire=annee_active)
                classes_qs = qs
                cache.set(classes_cache_key, classes_qs, 600)
            form.fields['classe'].queryset = classes_qs

        # Pre-selectionner la classe si passee en parametre (apres ajout d'un eleve)
        classe_id_param = request.GET.get('classe_id')
        if classe_id_param:
            try:
                form.fields['classe'].initial = int(classe_id_param)
            except (ValueError, TypeError):
                pass

        responsable_principal_form = ResponsableForm(prefix='resp_principal', user=request.user)
        responsable_secondaire_form = ResponsableForm(prefix='resp_secondaire', user=request.user)
    
    # Statistiques optimisées avec cache
    stats_cache_key = f'stats_eleves_{request.user.id}'
    stats = cache.get(stats_cache_key)
    
    if stats is None:
        # Requête optimisée avec agrégation
        from django.db.models import Count, Case, When, IntegerField
        
        eleves_qs = Eleve.objects.all()
        if not user_is_admin(request.user) and user_school_obj:
            eleves_qs = eleves_qs.filter(classe__ecole=user_school_obj)
        
        # Agrégation en une seule requête
        stats_result = eleves_qs.aggregate(
            total_eleves=Count('id'),
            eleves_actifs=Count(Case(When(statut='ACTIF', then=1), output_field=IntegerField())),
            eleves_exclus=Count(Case(When(statut='EXCLU', then=1), output_field=IntegerField()))
        )
        
        stats = {
            'total_eleves': stats_result['total_eleves'] or 0,
            'eleves_actifs': stats_result['eleves_actifs'] or 0,
            'eleves_exclus': stats_result['eleves_exclus'] or 0,
        }
        
        # Cache les stats pour 2 minutes
        cache.set(stats_cache_key, stats, 120)
    
    context = {
        'form': form,
        'responsable_principal_form': responsable_principal_form,
        'responsable_secondaire_form': responsable_secondaire_form,
        'stats': stats,
        'titre_page': 'Ajouter un Élève',
        'action': 'Ajouter'
    }
    
    return render(request, 'eleves/ajouter_eleve.html', context)

@login_required
def modifier_eleve(request, eleve_id):
    """Vue pour modifier un élève existant"""
    qs = Eleve.objects.all()
    if not user_is_admin(request.user):
        qs = filter_by_user_school(qs, request.user, 'classe__ecole')
    
    try:
        eleve = qs.get(id=eleve_id)
    except Eleve.DoesNotExist:
        messages.error(request, f"L'élève avec l'ID {eleve_id} n'existe pas.")
        return redirect('eleves:liste_eleves')
    
    if request.method == 'POST':
        form = EleveForm(request.POST, request.FILES, instance=eleve)
        if not user_is_admin(request.user):
            try:
                ecole_u = user_school(request.user)
                qs = Classe.objects.filter(ecole=ecole_u)
                annee_active = get_annee_active(request, ecole_u)
                if annee_active:
                    qs = qs.filter(annee_scolaire=annee_active)
                form.fields['classe'].queryset = qs
            except Exception:
                pass

        # --- Formulaires responsables (edition inline) ---
        resp1_form = ResponsableForm(
            request.POST if request.POST.get('edit_resp1') else None,
            instance=eleve.responsable_principal,
            prefix='resp1'
        ) if eleve.responsable_principal else None

        resp2_form = ResponsableForm(
            request.POST if request.POST.get('edit_resp2') else None,
            instance=eleve.responsable_secondaire,
            prefix='resp2'
        ) if eleve.responsable_secondaire else None

        # Formulaire pour creer un nouveau responsable principal
        new_resp1_form = ResponsableForm(
            request.POST if request.POST.get('new_resp1') else None,
            prefix='new_resp1'
        )
        # Formulaire pour creer un nouveau responsable secondaire
        new_resp2_form = ResponsableForm(
            request.POST if request.POST.get('new_resp2') else None,
            prefix='new_resp2'
        )

        if form.is_valid():
            # Détecter les changements
            changements = []
            for field in form.changed_data:
                if field in form.fields:
                    ancien_val = getattr(eleve, field, '')
                    nouveau_val = form.cleaned_data[field]
                    changements.append(f"{form.fields[field].label}: {ancien_val} -> {nouveau_val}")

            # Sauvegarder les responsables modifies
            if request.POST.get('edit_resp1') and resp1_form and resp1_form.is_valid():
                resp1_form.save()
                changements.append("Responsable principal: informations mises a jour")

            if request.POST.get('edit_resp2') and resp2_form and resp2_form.is_valid():
                resp2_form.save()
                changements.append("Responsable secondaire: informations mises a jour")

            # Creer un nouveau responsable principal
            if request.POST.get('new_resp1') and new_resp1_form.is_valid():
                new_nom = new_resp1_form.cleaned_data.get('nom', '').strip()
                new_prenom = new_resp1_form.cleaned_data.get('prenom', '').strip()
                if new_nom or new_prenom:
                    new_resp = new_resp1_form.save()
                    form.instance.responsable_principal = new_resp
                    changements.append(f"Nouveau responsable principal cree: {new_prenom} {new_nom}")

            # Creer un nouveau responsable secondaire
            if request.POST.get('new_resp2') and new_resp2_form.is_valid():
                new_nom = new_resp2_form.cleaned_data.get('nom', '').strip()
                new_prenom = new_resp2_form.cleaned_data.get('prenom', '').strip()
                if new_nom or new_prenom:
                    new_resp = new_resp2_form.save()
                    form.instance.responsable_secondaire = new_resp
                    changements.append(f"Nouveau responsable secondaire cree: {new_prenom} {new_nom}")

            # Passer l'utilisateur actuel pour la génération automatique du matricule
            eleve = form.save(commit=False)
            eleve._current_user = request.user

            # Gérer la saisie manuelle du matricule (pour la modification)
            saisie_manuelle = form.cleaned_data.get('saisie_manuelle_matricule', False)
            if saisie_manuelle and form.cleaned_data.get('matricule'):
                eleve.matricule = form.cleaned_data['matricule']
                eleve._skip_matricule_generation = True

            eleve.save()

            # Afficher le résultat du transfert de notes (si changement de classe)
            transfert_info = getattr(eleve, '_transfert_info', None)
            if transfert_info is not None:
                nb_trans = transfert_info.get('transferees', 0)
                nb_ign = transfert_info.get('ignorees', 0)
                if transfert_info.get('classe_note_manquante'):
                    messages.warning(request, f"Attention : les notes de {eleve.prenom} {eleve.nom} n'ont pas pu etre transferees (aucune configuration de notes trouvee pour la nouvelle classe).")
                elif transfert_info.get('matieres_manquantes'):
                    messages.warning(request, f"Attention : les notes de {eleve.prenom} {eleve.nom} n'ont pas pu etre transferees (la nouvelle classe n'a aucune matiere configuree).")
                elif nb_trans > 0 and nb_ign == 0:
                    messages.success(request, f"{nb_trans} note(s) transferee(s) automatiquement vers la nouvelle classe.")
                elif nb_trans > 0 and nb_ign > 0:
                    messages.warning(request, f"{nb_trans} note(s) transferee(s) vers la nouvelle classe, mais {nb_ign} note(s) n'ont pas pu etre transferees (matieres sans equivalent dans la nouvelle classe).")
                elif nb_trans == 0 and nb_ign > 0:
                    messages.warning(request, f"Attention : {nb_ign} note(s) n'ont pas pu etre transferees (matieres sans equivalent dans la nouvelle classe).")

            # Créer l'historique si des changements ont été effectués
            if changements:
                try:
                    HistoriqueEleve.objects.create(
                        eleve=eleve,
                        action='MODIFICATION',
                        description=f"Modification: {', '.join(changements)}",
                        utilisateur=request.user
                    )
                except Exception as e:
                    logger.error(f"Error creating history: {e}")

                # Log de l'activité
                try:
                    JournalActivite.objects.create(
                        user=request.user,
                        action='MODIFICATION',
                        type_objet='ELEVE',
                        objet_id=eleve.id,
                        description=f"Modification de l'eleve {eleve.nom_complet}: {', '.join(changements)}",
                        adresse_ip=request.META.get('REMOTE_ADDR', ''),
                        user_agent=request.META.get('HTTP_USER_AGENT', '')
                    )
                except Exception as e:
                    logger.error(f"Error creating activity log: {e}")

            # Message de succès détaillé
            if changements:
                nb_changements = len(changements)
                message_changements = f" ({nb_changements} modification{'s' if nb_changements > 1 else ''} effectuee{'s' if nb_changements > 1 else ''})"
                messages.success(
                    request,
                    f"Les informations de {eleve.prenom} {eleve.nom} ont ete mises a jour avec succes{message_changements}."
                )
            else:
                messages.success(request, f"Les informations de {eleve.prenom} {eleve.nom} ont ete sauvegardees.")

            # Rediriger vers la page de modification pour voir les messages
            return redirect('eleves:modifier_eleve', eleve_id=eleve.id)
        else:
            # Formulaire invalide: informer l'utilisateur des erreurs
            erreurs = []
            try:
                for champ, msgs in list(form.errors.items())[:5]:
                    libelle = form.fields.get(champ).label if champ in form.fields else champ
                    erreurs.append(f"{libelle}: {', '.join([str(m) for m in msgs])}")
            except Exception:
                pass
            if erreurs:
                messages.error(request, "Le formulaire contient des erreurs: " + " | ".join(erreurs))
            else:
                messages.error(request, "Le formulaire est invalide. Veuillez corriger les erreurs et reessayer.")
    else:
        form = EleveForm(instance=eleve)
        if not user_is_admin(request.user):
            try:
                ecole_u = user_school(request.user)
                qs = Classe.objects.filter(ecole=ecole_u)
                annee_active = get_annee_active(request, ecole_u)
                if annee_active:
                    qs = qs.filter(annee_scolaire=annee_active)
                form.fields['classe'].queryset = qs
            except Exception:
                pass

        # Formulaires responsables pour edition inline (GET)
        resp1_form = ResponsableForm(
            instance=eleve.responsable_principal, prefix='resp1'
        ) if eleve.responsable_principal else None

        resp2_form = ResponsableForm(
            instance=eleve.responsable_secondaire, prefix='resp2'
        ) if eleve.responsable_secondaire else None

        new_resp1_form = ResponsableForm(prefix='new_resp1')
        new_resp2_form = ResponsableForm(prefix='new_resp2')

    context = {
        'form': form,
        'eleve': eleve,
        'resp1_form': locals().get('resp1_form'),
        'resp2_form': locals().get('resp2_form'),
        'new_resp1_form': locals().get('new_resp1_form'),
        'new_resp2_form': locals().get('new_resp2_form'),
        'titre_page': f'Modifier {eleve.nom_complet}',
        'action': 'Modifier'
    }

    return render(request, 'eleves/modifier_eleve_simple.html', context)

@login_required
def _get_classe_or_403(request, classe_id):
    qs = Classe.objects.select_related('ecole')
    if not user_is_admin(request.user):
        qs = qs.filter(ecole=user_school(request.user))
    return get_object_or_404(qs, id=classe_id)

@login_required
@vary_on_cookie
def export_eleves_classe_pdf(request, classe_id):
    """Exporte la liste des élèves d'une classe en PDF."""
    classe = _get_classe_or_403(request, classe_id)
    eleves = Eleve.objects.select_related('classe', 'responsable_principal').filter(classe=classe).order_by('nom', 'prenom')

    # Log activité
    JournalActivite.objects.create(
        user=request.user,
        action='EXPORT',
        type_objet='ELEVE',
        description=f"Export PDF élèves - Classe {classe.nom} ({classe.ecole.nom})",
        adresse_ip=request.META.get('REMOTE_ADDR', ''),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )

    response = HttpResponse(content_type='application/pdf')
    filename = f"eleves_{slugify(classe.ecole.nom)}_{slugify(classe.nom)}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    c = canvas.Canvas(response, pagesize=landscape(A4))
    width, height = landscape(A4)
    c.setPageCompression(1)
    
    # Ajouter le filigrane (spécifique à l'école de la classe)
    try:
        from ecole_moderne.pdf_utils import draw_logo_watermark
        draw_logo_watermark(c, width, height, ecole=getattr(classe, 'ecole', None))
    except Exception:
        pass

    # Polices (Calibri/Arial si dispo, sinon Helvetica par défaut)
    font_name = 'Helvetica'
    font_bold = 'Helvetica-Bold'
    
    try:
        calibri_path = 'C:/Windows/Fonts/calibri.ttf'
        calibri_bold_path = 'C:/Windows/Fonts/calibrib.ttf'
        if os.path.exists(calibri_path) and os.path.exists(calibri_bold_path):
            pdfmetrics.registerFont(TTFont('MainFont', calibri_path))
            pdfmetrics.registerFont(TTFont('MainFont-Bold', calibri_bold_path))
            font_name = 'MainFont'
            font_bold = 'MainFont-Bold'
        else:
            arial_path = 'C:/Windows/Fonts/arial.ttf'
            arial_bold_path = 'C:/Windows/Fonts/arialbd.ttf'
            if os.path.exists(arial_path) and os.path.exists(arial_bold_path):
                pdfmetrics.registerFont(TTFont('MainFont', arial_path))
                pdfmetrics.registerFont(TTFont('MainFont-Bold', arial_bold_path))
                font_name = 'MainFont'
                font_bold = 'MainFont-Bold'
    except Exception:
        # Utiliser les polices par défaut de ReportLab
        pass

    # Filigrane standardisé
    draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)

    margin = 2*cm
    y = height - margin

    # En-tête
    c.setFont(font_bold, 16)
    c.drawString(margin, y, f"Liste des élèves - {classe.ecole.nom}")
    y -= 18
    c.setFont(font_name, 12)
    c.drawString(margin, y, f"Classe: {classe.nom}")
    y -= 10
    c.setFillColor(colors.grey)
    c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
    c.setFillColor(colors.black)
    y -= 18

    # En-têtes du tableau
    headers = ["Matricule", "Nom", "Sexe", "Date Naissance", "Responsable", "Téléphone"]
    col_widths = [3.0*cm, 7.0*cm, 2.0*cm, 3.0*cm, 5.0*cm, 3.5*cm]
    c.setFont(font_bold, 11)
    x = margin
    for i, htxt in enumerate(headers):
        c.drawString(x, y, htxt)
        x += col_widths[i]
    y -= 14
    c.setFillColor(colors.lightgrey)
    c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
    c.setFillColor(colors.black)
    y -= 8

    c.setFont(font_name, 10)
    for e in eleves:
        # Saut de page si nécessaire
        if y < margin + 40:
            c.showPage()
            # Filigrane sur chaque nouvelle page (école de la classe)
            draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=getattr(classe, 'ecole', None))
            y = height - margin
            c.setFont(font_bold, 11)
            x = margin
            for i, htxt in enumerate(headers):
                c.drawString(x, y, htxt)
                x += col_widths[i]
            y -= 18
            c.setFont(font_name, 10)

        x = margin
        values = [
            e.matricule or '',
            f"{e.nom} {e.prenom}",
            e.get_sexe_display() if hasattr(e, 'get_sexe_display') else getattr(e, 'sexe', ''),
            e.date_naissance.strftime('%d/%m/%Y') if getattr(e, 'date_naissance', None) else '',
            e.responsable_principal.nom_complet if e.responsable_principal else '',
            e.responsable_principal.telephone if e.responsable_principal else '',
        ]
        for i, val in enumerate(values):
            c.drawString(x, y, str(val))
            x += col_widths[i]
        y -= 14

    c.showPage()
    c.save()
    return response

@login_required
def export_eleves_classe_excel(request, classe_id):
    """Exporte la liste des élèves d'une classe en Excel (.xlsx)."""
    if Workbook is None or get_column_letter is None:
        return HttpResponse("Erreur: openpyxl n'est pas installé sur le serveur.", status=500)

    classe = _get_classe_or_403(request, classe_id)
    eleves = Eleve.objects.select_related('classe', 'responsable_principal').filter(classe=classe).order_by('nom', 'prenom')

    # Log activité
    JournalActivite.objects.create(
        user=request.user,
        action='EXPORT',
        type_objet='ELEVE',
        description=f"Export Excel élèves - Classe {classe.nom} ({classe.ecole.nom})",
        adresse_ip=request.META.get('REMOTE_ADDR', ''),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Élèves"

        # En-têtes avec style
        headers = ["Matricule", "Nom complet", "Sexe", "Date de naissance", "Responsable principal", "Téléphone"]
        ws.append(headers)
        
        # Style pour les en-têtes
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Données des élèves
        for e in eleves:
            ws.append([
                e.matricule or '',
                f"{e.nom} {e.prenom}",
                e.get_sexe_display() if hasattr(e, 'get_sexe_display') else getattr(e, 'sexe', ''),
                e.date_naissance.strftime('%d/%m/%Y') if getattr(e, 'date_naissance', None) else '',
                e.responsable_principal.nom_complet if e.responsable_principal else '',
                e.responsable_principal.telephone if e.responsable_principal else '',
            ])

        # Largeur des colonnes optimisée
        widths = [15, 30, 10, 18, 30, 18]
        for idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w

        # Réponse HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"eleves_{slugify(classe.ecole.nom)}_{slugify(classe.nom)}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
        
    except Exception as e:
        logger.exception("Erreur lors de la génération du fichier Excel")
        return HttpResponse("Une erreur est survenue lors de la génération du fichier Excel.", status=500)

@login_required
@vary_on_cookie
def export_tous_eleves_pdf(request):
    """Exporte la liste de tous les élèves en PDF."""
    # IMPORTANT: Seul le superuser peut voir toutes les écoles
    if user_is_superadmin(request.user):
        eleves = Eleve.objects.select_related('classe', 'classe__ecole', 'responsable_principal').all()
    else:
        # Tous les autres utilisateurs (y compris ADMIN d'école) ne voient que leur école
        ecole = user_school(request.user)
        if ecole is None:
            eleves = Eleve.objects.none()
        else:
            eleves = Eleve.objects.select_related('classe', 'classe__ecole', 'responsable_principal').filter(
                classe__ecole=ecole
            )
    
    eleves = eleves.order_by('classe__ecole__nom', 'classe__nom', 'nom', 'prenom')

    # Log activité
    JournalActivite.objects.create(
        user=request.user,
        action='EXPORT',
        type_objet='ELEVE',
        description="Export PDF - Tous les élèves",
        adresse_ip=request.META.get('REMOTE_ADDR', ''),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )

    response = HttpResponse(content_type='application/pdf')
    filename = "tous_les_eleves.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    try:
        c = canvas.Canvas(response, pagesize=landscape(A4))
        width, height = landscape(A4)
        c.setPageCompression(1)
        
        # Ajouter le filigrane (au périmètre de l'utilisateur)
        try:
            draw_logo_watermark(c, width, height, ecole=user_school(request.user))
        except Exception:
            pass

        # Configuration des polices
        font_name = 'Helvetica'
        font_bold = 'Helvetica-Bold'
        
        try:
            calibri_path = 'C:/Windows/Fonts/calibri.ttf'
            calibri_bold_path = 'C:/Windows/Fonts/calibrib.ttf'
            if os.path.exists(calibri_path) and os.path.exists(calibri_bold_path):
                pdfmetrics.registerFont(TTFont('MainFont', calibri_path))
                pdfmetrics.registerFont(TTFont('MainFont-Bold', calibri_bold_path))
                font_name = 'MainFont'
                font_bold = 'MainFont-Bold'
        except Exception:
            pass

        # Filigrane standardisé (logo centré, rotation, opacité faible)
        draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5)

        margin = 2*cm
        y = height - margin

        # En-tête principal
        c.setFont(font_bold, 18)
        c.drawString(margin, y, "Liste complète des élèves")
        y -= 25
        
        c.setFont(font_name, 12)
        from datetime import datetime
        c.drawString(margin, y, f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}")
        y -= 15
        
        c.setFillColor(colors.grey)
        c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
        c.setFillColor(colors.black)
        y -= 25

        # En-têtes du tableau
        headers = ["École", "Classe", "Matricule", "Nom", "Responsable", "Téléphone"]
        col_widths = [4.5*cm, 4.5*cm, 3*cm, 5.5*cm, 5*cm, 3*cm]
        
        current_ecole = None
        
        for eleve in eleves:
            # Nouvelle école
            if current_ecole != eleve.classe.ecole.nom:
                if y < margin + 80:
                    c.showPage()
                    # Filigrane sur chaque nouvelle page (au périmètre de l'utilisateur)
                    draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=user_school(request.user))
                    y = height - margin
                
                current_ecole = eleve.classe.ecole.nom
                
                # Titre de l'école
                c.setFont(font_bold, 14)
                c.drawString(margin, y, f"École: {current_ecole}")
                y -= 20
                
                # En-têtes du tableau
                c.setFont(font_bold, 10)
                x = margin
                for i, header in enumerate(headers[1:]):  # Skip "École" pour cette section
                    c.drawString(x, y, header)
                    x += col_widths[i+1]
                y -= 15
                
                c.setFillColor(colors.lightgrey)
                c.rect(margin, y-2, width-2*margin, 1, fill=1, stroke=0)
                c.setFillColor(colors.black)
                y -= 8
            
            # Vérifier l'espace pour une nouvelle ligne
            if y < margin + 40:
                c.showPage()
                # Filigrane sur chaque nouvelle page (au périmètre de l'utilisateur)
                draw_logo_watermark(c, width, height, opacity=0.04, rotate=30, scale=1.5, ecole=user_school(request.user))
                y = height - margin
                
                # Répéter le titre de l'école et les en-têtes
                c.setFont(font_bold, 14)
                c.drawString(margin, y, f"École: {current_ecole} (suite)")
                y -= 20
                
                c.setFont(font_bold, 10)
                x = margin
                for i, header in enumerate(headers[1:]):
                    c.drawString(x, y, header)
                    x += col_widths[i+1]
                y -= 18
            
            # Ligne de données
            c.setFont(font_name, 9)
            x = margin
            
            # Récupérer le téléphone du responsable
            telephone = ''
            if eleve.responsable_principal:
                telephone = eleve.responsable_principal.telephone or ''
            
            values = [
                eleve.classe.nom,
                eleve.matricule or '',
                f"{eleve.nom} {eleve.prenom}",
                eleve.responsable_principal.nom_complet if eleve.responsable_principal else '',
                telephone,
            ]
            
            # Limites de caractères par colonne pour éviter le chevauchement
            max_chars = [27, 13, 28, 25, 13]  # Classe, Matricule, Nom, Responsable, Téléphone
            
            for i, val in enumerate(values):
                # Tronquer si trop long selon la colonne
                text = str(val)
                if len(text) > max_chars[i]:
                    text = text[:max_chars[i]-3] + '...'
                c.drawString(x, y, text)
                x += col_widths[i+1]
            y -= 12

        c.showPage()
        c.save()
        return response
        
    except Exception as e:
        return HttpResponse(f"Erreur lors de la génération du PDF: {str(e)}", status=500)

@login_required
def export_tous_eleves_excel(request):
    """Exporte la liste de tous les élèves en Excel (.xlsx)."""
    if Workbook is None or get_column_letter is None:
        return HttpResponse("Erreur: openpyxl n'est pas installé sur le serveur.", status=500)

    # IMPORTANT: Seul le superuser peut voir toutes les écoles
    if user_is_superadmin(request.user):
        eleves = Eleve.objects.select_related('classe', 'classe__ecole', 'responsable_principal').all()
    else:
        # Tous les autres utilisateurs (y compris ADMIN d'école) ne voient que leur école
        ecole = user_school(request.user)
        if ecole is None:
            eleves = Eleve.objects.none()
        else:
            eleves = Eleve.objects.select_related('classe', 'classe__ecole', 'responsable_principal').filter(
                classe__ecole=ecole
            )
    
    eleves = eleves.order_by('classe__ecole__nom', 'classe__nom', 'nom', 'prenom')

    # Log activité
    JournalActivite.objects.create(
        user=request.user,
        action='EXPORT',
        type_objet='ELEVE',
        description="Export Excel - Tous les élèves",
        adresse_ip=request.META.get('REMOTE_ADDR', ''),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Tous les élèves"

        # En-têtes avec style
        headers = ["École", "Classe", "Matricule", "Nom complet", "Sexe", "Date de naissance", "Responsable principal", "Téléphone"]
        ws.append(headers)
        
        # Style pour les en-têtes
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Données des élèves
        for eleve in eleves:
            ws.append([
                eleve.classe.ecole.nom,
                eleve.classe.nom,
                eleve.matricule or '',
                f"{eleve.nom} {eleve.prenom}",
                eleve.get_sexe_display() if hasattr(eleve, 'get_sexe_display') else getattr(eleve, 'sexe', ''),
                eleve.date_naissance.strftime('%d/%m/%Y') if getattr(eleve, 'date_naissance', None) else '',
                eleve.responsable_principal.nom_complet if eleve.responsable_principal else '',
                eleve.responsable_principal.telephone if eleve.responsable_principal else '',
            ])

        # Largeur des colonnes optimisée
        widths = [20, 15, 15, 25, 10, 15, 25, 15]
        for idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w

        # Réponse HTTP
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = "tous_les_eleves.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
        
    except Exception as e:
        logger.exception("Erreur lors de la génération du fichier Excel")
        return HttpResponse("Une erreur est survenue lors de la génération du fichier Excel.", status=500)

@login_required
def supprimer_eleve(request, eleve_id):
    """Vue pour supprimer un élève avec ses paiements et abonnements (avec code de vérification)"""
    # Permettre l'accès aux utilisateurs connectés (la sécurité est assurée par le code de vérification)
    # Les permissions spécifiques (soft delete vs suppression définitive) sont gérées dans le traitement
    
    qs = Eleve.objects.all()
    if not user_is_admin(request.user):
        qs = filter_by_user_school(qs, request.user, 'classe__ecole')
    
    try:
        eleve = qs.get(id=eleve_id)
    except Eleve.DoesNotExist:
        messages.error(request, f"L'élève avec l'ID {eleve_id} n'existe pas ou a déjà été supprimé.")
        return redirect('eleves:liste_eleves')

    nom_complet = f"{eleve.prenom} {eleve.nom}"
    matricule = eleve.matricule
    
    # Vérifier la permission de suppression définitive
    peut_supprimer_definitivement = user_is_admin(request.user) or (
        hasattr(request.user, 'profil') and 
        request.user.profil.peut_supprimer_eleves_definitivement
    )
    
    # Compter les éléments associés
    paiements_count = eleve.paiements.count()
    abonnements_bus_count = eleve.abonnements_bus.count()
    abonnements_cantine_count = eleve.abonnements_cantine.count()
    
    if request.method == 'POST':
        # Vérifier le code de sécurité
        code_verification = request.POST.get('code_verification', '').strip()
        suppression_definitive = request.POST.get('suppression_definitive') == 'on'
        
        # Pour les admins, toujours activer la suppression définitive par défaut
        if user_is_admin(request.user):
            # Si l'admin n'a pas explicitement décoché, on force la suppression définitive
            suppression_definitive = request.POST.get('suppression_definitive') != 'off'
            # Log pour debug
            import logging
            logger = logging.getLogger(__name__)
            logger.info(f"Admin {request.user.username} - Suppression définitive: {suppression_definitive}")
        
        from django.conf import settings as django_settings
        expected_code = django_settings.SECURITY_VERIFICATION_CODE
        if not expected_code or code_verification != expected_code:
            messages.error(request, "Code de vérification incorrect. Suppression annulée.")
            return render(request, 'eleves/confirmer_suppression.html', {
                'eleve': eleve,
                'paiements_count': paiements_count,
                'titre_page': f'Supprimer {nom_complet}'
            })
        
        # Vérifier la permission pour suppression définitive
        if suppression_definitive and not peut_supprimer_definitivement:
            messages.error(request, "Vous n'avez pas la permission de supprimer définitivement un élève.")
            return redirect('eleves:detail_eleve', eleve_id=eleve.id)
        
        # Procéder à la suppression avec le code correct
        from django.db import transaction
        try:
            with transaction.atomic():
                if suppression_definitive and peut_supprimer_definitivement:
                    # Suppression définitive pour les utilisateurs autorisés
                    # Collecter les informations avant suppression
                    paiements_supprimes = []
                    for paiement in eleve.paiements.all():
                        paiements_supprimes.append(f"{paiement.numero_recu} - {paiement.montant} GNF")
                    
                    abonnements_bus_supprimes = []
                    for abo in eleve.abonnements_bus.all():
                        abonnements_bus_supprimes.append(f"{abo.get_periodicite_display()} - {abo.montant} GNF")
                    
                    abonnements_cantine_supprimes = []
                    for abo in eleve.abonnements_cantine.all():
                        abonnements_cantine_supprimes.append(f"{abo.get_periodicite_display()} - {abo.montant} GNF")
                    
                    # Créer l'entrée dans la corbeille avant suppression
                    from administration.models import SystemLog
                    SystemLog.objects.create(
                        action='SUPPRESSION_DEFINITIVE',
                        description=f"Suppression définitive de l'élève {nom_complet} (matricule: {matricule}) avec {paiements_count} paiement(s), {abonnements_bus_count} abonnement(s) bus et {abonnements_cantine_count} abonnement(s) cantine",
                        user=request.user,
                        ip_address=request.META.get('REMOTE_ADDR', ''),
                        details={
                            'eleve_id': eleve.id,
                            'matricule': matricule,
                            'nom_complet': nom_complet,
                            'classe': str(eleve.classe),
                            'paiements_supprimes': paiements_supprimes,
                            'abonnements_bus_supprimes': abonnements_bus_supprimes,
                            'abonnements_cantine_supprimes': abonnements_cantine_supprimes,
                            'verification_code_used': True,
                            'user_agent': request.META.get('HTTP_USER_AGENT', '')
                        }
                    )
                    
                    # Supprimer les paiements
                    eleve.paiements.all().delete()
                    
                    # Supprimer les abonnements bus
                    eleve.abonnements_bus.all().delete()
                    
                    # Supprimer les abonnements cantine
                    eleve.abonnements_cantine.all().delete()
                    
                    # Supprimer l'élève définitivement
                    eleve.delete()
                    
                    total_elements = paiements_count + abonnements_bus_count + abonnements_cantine_count
                    messages.success(request, f"L'élève {nom_complet} et tous ses éléments associés ({total_elements} au total) ont été supprimés définitivement et sauvegardés dans la corbeille.")
                else:
                    # Soft delete - changer le statut au lieu de supprimer
                    eleve.statut = 'EXCLU'
                    eleve.save()

                    # Créer l'historique
                    HistoriqueEleve.objects.create(
                        eleve=eleve,
                        action='EXCLUSION',
                        description=f"Exclusion de l'élève {nom_complet} avec {paiements_count} paiement(s)",
                        utilisateur=request.user
                    )

                    # Log de l'activité
                    JournalActivite.objects.create(
                        user=request.user,
                        action='SUPPRESSION',
                        type_objet='ELEVE',
                        objet_id=eleve.id,
                        description=f"Exclusion de l'élève {nom_complet} (matricule: {matricule}) avec {paiements_count} paiement(s)",
                        adresse_ip=request.META.get('REMOTE_ADDR', ''),
                        user_agent=request.META.get('HTTP_USER_AGENT', '')
                    )
                    
                    messages.success(request, f"L'élève {nom_complet} a été exclu (soft delete).")
                    
        except Exception as e:
            messages.error(request, f"Erreur lors de la suppression: {e}")
            return redirect('eleves:detail_eleve', eleve_id=eleve.id)

        return redirect('eleves:liste_eleves')
    
    # Afficher le formulaire de confirmation
    return render(request, 'eleves/confirmer_suppression.html', {
        'eleve': eleve,
        'paiements_count': paiements_count,
        'abonnements_bus_count': abonnements_bus_count,
        'abonnements_cantine_count': abonnements_cantine_count,
        'peut_supprimer_definitivement': peut_supprimer_definitivement,
        'titre_page': f'Supprimer {nom_complet}'
    })

@login_required
@require_http_methods(["POST"])
def supprimer_eleves_masse(request):
    """Vue pour supprimer définitivement plusieurs élèves en masse"""
    # Vérifier la permission de suppression définitive
    peut_supprimer_definitivement = user_is_admin(request.user) or (
        hasattr(request.user, 'profil') and 
        request.user.profil.peut_supprimer_eleves_definitivement
    )
    
    if not peut_supprimer_definitivement:
        messages.error(request, "Vous n'avez pas la permission de supprimer définitivement des élèves.")
        return redirect('eleves:liste_eleves')
    
    # Vérifier le code de sécurité
    code_verification = request.POST.get('code_verification', '').strip()
    from django.conf import settings as django_settings
    expected_code = django_settings.SECURITY_VERIFICATION_CODE
    if not expected_code or code_verification != expected_code:
        messages.error(request, "Code de vérification incorrect. Suppression annulée.")
        return redirect('eleves:liste_eleves')
    
    # Récupérer les IDs des élèves
    eleve_ids_str = request.POST.get('eleve_ids', '')
    if not eleve_ids_str:
        messages.error(request, "Aucun élève sélectionné.")
        return redirect('eleves:liste_eleves')
    
    try:
        eleve_ids = [int(x) for x in eleve_ids_str.split(',') if x.strip()]
    except ValueError:
        messages.error(request, "IDs d'élèves invalides.")
        return redirect('eleves:liste_eleves')
    
    if not eleve_ids:
        messages.error(request, "Aucun élève sélectionné.")
        return redirect('eleves:liste_eleves')
    
    # Récupérer les élèves
    qs = Eleve.objects.filter(id__in=eleve_ids)
    if not user_is_admin(request.user):
        qs = filter_by_user_school(qs, request.user, 'classe__ecole')
    
    eleves = list(qs)
    if not eleves:
        messages.error(request, "Aucun élève trouvé avec les IDs fournis.")
        return redirect('eleves:liste_eleves')
    
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        with transaction.atomic():
            eleves_supprimes = []
            total_paiements = 0
            total_abonnements_bus = 0
            total_abonnements_cantine = 0
            
            for eleve in eleves:
                nom_complet = f"{eleve.prenom} {eleve.nom}"
                matricule = eleve.matricule
                
                paiements_count = eleve.paiements.count()
                abonnements_bus_count = eleve.abonnements_bus.count()
                abonnements_cantine_count = eleve.abonnements_cantine.count()
                
                # Collecter les informations avant suppression
                paiements_supprimes = []
                for paiement in eleve.paiements.all():
                    paiements_supprimes.append(f"{paiement.numero_recu} - {paiement.montant} GNF")
                
                abonnements_bus_supprimes = []
                for abo in eleve.abonnements_bus.all():
                    abonnements_bus_supprimes.append(f"{abo.get_periodicite_display()} - {abo.montant} GNF")
                
                abonnements_cantine_supprimes = []
                for abo in eleve.abonnements_cantine.all():
                    abonnements_cantine_supprimes.append(f"{abo.get_periodicite_display()} - {abo.montant} GNF")
                
                eleves_supprimes.append({
                    'eleve_id': eleve.id,
                    'matricule': matricule,
                    'nom_complet': nom_complet,
                    'classe': str(eleve.classe),
                    'paiements_supprimes': paiements_supprimes,
                    'abonnements_bus_supprimes': abonnements_bus_supprimes,
                    'abonnements_cantine_supprimes': abonnements_cantine_supprimes,
                })
                
                total_paiements += paiements_count
                total_abonnements_bus += abonnements_bus_count
                total_abonnements_cantine += abonnements_cantine_count
                
                # Supprimer les éléments associés
                eleve.paiements.all().delete()
                eleve.abonnements_bus.all().delete()
                eleve.abonnements_cantine.all().delete()
                
                # Supprimer l'élève
                eleve.delete()
                
                logger.info(f"Suppression définitive en masse: {nom_complet} ({matricule}) par {request.user.username}")
            
            # Créer l'entrée dans la corbeille
            from administration.models import SystemLog
            SystemLog.objects.create(
                action='SUPPRESSION_DEFINITIVE',
                description=f"Suppression définitive en masse de {len(eleves_supprimes)} élève(s) avec {total_paiements} paiement(s), {total_abonnements_bus} abonnement(s) bus et {total_abonnements_cantine} abonnement(s) cantine",
                user=request.user,
                ip_address=request.META.get('REMOTE_ADDR', ''),
                details={
                    'type': 'suppression_masse',
                    'nombre_eleves': len(eleves_supprimes),
                    'eleves_supprimes': eleves_supprimes,
                    'verification_code_used': True,
                    'user_agent': request.META.get('HTTP_USER_AGENT', '')
                }
            )
            
            # Log dans le journal d'activité
            noms = ', '.join([e['nom_complet'] for e in eleves_supprimes])
            JournalActivite.objects.create(
                user=request.user,
                action='SUPPRESSION',
                type_objet='ELEVE',
                description=f"Suppression définitive en masse de {len(eleves_supprimes)} élève(s): {noms}",
                adresse_ip=request.META.get('REMOTE_ADDR', ''),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
            messages.success(request, f"{len(eleves_supprimes)} élève(s) ont été supprimés définitivement avec toutes leurs données associées.")
            
    except Exception as e:
        logger.error(f"Erreur lors de la suppression en masse: {e}")
        messages.error(request, f"Erreur lors de la suppression en masse: {e}")
    
    return redirect('eleves:liste_eleves')


@login_required
def gestion_classes(request):
    """Vue pour gérer les classes — filtrée par année scolaire active"""
    ecole_user = user_school(request.user)
    annee_active = get_annee_active(request, ecole_user) if ecole_user else None

    classes = Classe.objects.select_related('ecole').annotate(
        eleves_count=Count('eleves', filter=Q(eleves__statut='ACTIF'))
    ).order_by('ecole__nom', 'niveau', 'nom')

    if not user_is_admin(request.user):
        classes = classes.filter(ecole=ecole_user)

    # Filtrer par année active
    if annee_active:
        classes = classes.filter(annee_scolaire=annee_active)

    # Statistiques
    stats = {
        'total_classes': classes.count(),
        'total_eleves': sum(c.eleves_count for c in classes),
        'classes_par_ecole': {}
    }

    ecoles_iter = Ecole.objects.all()
    if not user_is_admin(request.user) and ecole_user:
        ecoles_iter = ecoles_iter.filter(id=ecole_user.id)
    for ecole in ecoles_iter:
        classes_ecole = classes.filter(ecole=ecole)
        stats['classes_par_ecole'][ecole.nom] = {
            'nombre_classes': classes_ecole.count(),
            'nombre_eleves': sum(c.eleves_count for c in classes_ecole)
        }

    context = {
        'classes': classes,
        'stats': stats,
        'annee_active': annee_active,
        'titre_page': f'Gestion des Classes — {annee_active}' if annee_active else 'Gestion des Classes',
    }

    return render(request, 'eleves/gestion_classes.html', context)

@login_required
def ajax_classes_par_ecole(request, ecole_id):
    """Vue AJAX pour récupérer les classes d'une école"""
    try:
        # Non-admin: ne peut demander que sa propre école
        if not user_is_admin(request.user):
            if str(user_school(request.user).id) != str(ecole_id):
                return JsonResponse({'success': False, 'error': "Accès non autorisé à cette école."}, status=403)
        ecole = get_object_or_404(Ecole, id=ecole_id)
        annee_active = get_annee_active(request, ecole)
        qs = Classe.objects.filter(ecole=ecole)
        if annee_active:
            qs = qs.filter(annee_scolaire=annee_active)
        classes = qs.values('id', 'nom')
        return JsonResponse({
            'success': True,
            'classes': list(classes)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def ajax_statistiques_eleves(request):
    """Vue AJAX pour récupérer les statistiques des élèves"""
    try:
        # IMPORTANT: Seul le superuser peut voir toutes les écoles
        if user_is_superadmin(request.user):
            eleves = Eleve.objects.all()
        else:
            ecole = user_school(request.user)
            if ecole is None:
                eleves = Eleve.objects.none()
            else:
                eleves = Eleve.objects.filter(classe__ecole=ecole)
        stats = {
            'total_eleves': eleves.count(),
            'eleves_actifs': eleves.filter(statut='ACTIF').count(),
            'eleves_exclus': eleves.filter(statut='EXCLU').count(),
        }
        
        return JsonResponse({
            'success': True,
            'stats': stats
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def statistiques_eleves(request):
    """Vue pour afficher les statistiques complètes des élèves"""
    from django.db.models import Sum, Avg, Max, Min
    from datetime import datetime, date
    from dateutil.relativedelta import relativedelta
    
    # 1. STATISTIQUES GÉNÉRALES
    # IMPORTANT: Seul le superuser peut voir toutes les écoles
    if user_is_superadmin(request.user):
        eleves_base = Eleve.objects.all()
        classes_base = Classe.objects.all()
        responsables_base = Responsable.objects.all()
        ecoles_base = Ecole.objects.all()
    else:
        # Tous les autres utilisateurs (y compris ADMIN d'école) ne voient que leur école
        ecole_u = user_school(request.user)
        if ecole_u is None:
            # Sécurité: si pas d'école assignée, retourner des querysets vides
            eleves_base = Eleve.objects.none()
            classes_base = Classe.objects.none()
            responsables_base = Responsable.objects.none()
            ecoles_base = Ecole.objects.none()
        else:
            annee_active = get_annee_active(request, ecole_u)
            qs_classe_filter = {'ecole': ecole_u}
            if annee_active:
                qs_classe_filter['annee_scolaire'] = annee_active
            classes_base = Classe.objects.filter(**qs_classe_filter)
            eleves_base = Eleve.objects.filter(classe__in=classes_base)
            # Filtrer les responsables par école
            responsables_base = Responsable.objects.filter(
                Q(eleves_principal__classe__ecole=ecole_u) |
                Q(eleves_secondaire__classe__ecole=ecole_u)
            ).distinct()
            ecoles_base = Ecole.objects.filter(id=ecole_u.id)
    total_eleves = eleves_base.count()
    stats_generales = {
        'total_eleves': total_eleves,
        'eleves_actifs': eleves_base.filter(statut='ACTIF').count(),
        'eleves_suspendus': eleves_base.filter(statut='SUSPENDU').count(),
        'eleves_exclus': eleves_base.filter(statut='EXCLU').count(),
        'eleves_transferes': eleves_base.filter(statut='TRANSFERE').count(),
        'eleves_diplomes': eleves_base.filter(statut='DIPLOME').count(),
        'total_ecoles': ecoles_base.count(),
        'total_classes': classes_base.count(),
        'total_responsables': responsables_base.count(),
    }
    
    # 2. STATISTIQUES DÉMOGRAPHIQUES
    stats_demographiques = {
        'garcons': eleves_base.filter(sexe='M').count(),
        'filles': eleves_base.filter(sexe='F').count(),
        'pourcentage_garcons': 0,
        'pourcentage_filles': 0,
    }
    
    if total_eleves > 0:
        stats_demographiques['pourcentage_garcons'] = round((stats_demographiques['garcons'] / total_eleves) * 100, 1)
        stats_demographiques['pourcentage_filles'] = round((stats_demographiques['filles'] / total_eleves) * 100, 1)
    
    # 3. STATISTIQUES D'ÂGE
    eleves_avec_age = eleves_base.exclude(date_naissance__isnull=True)
    ages = []
    for eleve in eleves_avec_age:
        age = relativedelta(date.today(), eleve.date_naissance).years
        ages.append(age)
    
    stats_age = {
        'age_moyen': round(sum(ages) / len(ages), 1) if ages else 0,
        'age_min': min(ages) if ages else 0,
        'age_max': max(ages) if ages else 0,
        'eleves_moins_10': len([a for a in ages if a < 10]),
        'eleves_10_15': len([a for a in ages if 10 <= a <= 15]),
        'eleves_plus_15': len([a for a in ages if a > 15]),
    }
    
    # 4. RÉPARTITION PAR ÉCOLE (détaillée)
    stats_par_ecole = []
    for ecole in ecoles_base:
        eleves_ecole = eleves_base.filter(classe__ecole=ecole)
        classes_ecole = classes_base.filter(ecole=ecole)
        
        ecole_stats = {
            'ecole': ecole,
            'total_eleves': eleves_ecole.count(),
            'eleves_actifs': eleves_ecole.filter(statut='ACTIF').count(),
            'garcons': eleves_ecole.filter(sexe='M').count(),
            'filles': eleves_ecole.filter(sexe='F').count(),
            'total_classes': classes_ecole.count(),
            'classes_actives': classes_ecole.filter(eleves__isnull=False).distinct().count(),
            'moyenne_eleves_par_classe': 0,
        }
        
        if ecole_stats['total_classes'] > 0:
            ecole_stats['moyenne_eleves_par_classe'] = round(ecole_stats['total_eleves'] / ecole_stats['total_classes'], 1)
        
        # Pourcentages pour cette école
        if ecole_stats['total_eleves'] > 0:
            ecole_stats['pourcentage_garcons'] = round((ecole_stats['garcons'] / ecole_stats['total_eleves']) * 100, 1)
            ecole_stats['pourcentage_filles'] = round((ecole_stats['filles'] / ecole_stats['total_eleves']) * 100, 1)
        else:
            ecole_stats['pourcentage_garcons'] = 0
            ecole_stats['pourcentage_filles'] = 0
        
        stats_par_ecole.append(ecole_stats)
    
    # 5. RÉPARTITION PAR NIVEAU (détaillée)
    stats_par_niveau = []
    total_pour_pourcentage = total_eleves if total_eleves > 0 else 1
    
    for niveau_code, niveau_nom in Classe.NIVEAUX_CHOICES:
        eleves_niveau = eleves_base.filter(classe__niveau=niveau_code)
        count = eleves_niveau.count()
        
        if count > 0:
            niveau_stats = {
                'niveau_code': niveau_code,
                'niveau_nom': niveau_nom,
                'total_eleves': count,
                'garcons': eleves_niveau.filter(sexe='M').count(),
                'filles': eleves_niveau.filter(sexe='F').count(),
                'actifs': eleves_niveau.filter(statut='ACTIF').count(),
                'pourcentage': round((count / total_pour_pourcentage) * 100, 1),
                # Utiliser classes_base pour respecter le filtrage par école
                'classes': classes_base.filter(niveau=niveau_code, eleves__isnull=False).distinct().count(),
            }
            stats_par_niveau.append(niveau_stats)
    
    # 6. STATISTIQUES PAR CLASSE (top 10)
    stats_par_classe = []
    classes_avec_eleves = classes_base.annotate(
        nb_eleves=Count('eleves')
    ).filter(nb_eleves__gt=0).order_by('-nb_eleves')[:10]
    
    for classe in classes_avec_eleves:
        eleves_classe = eleves_base.filter(classe=classe)
        classe_stats = {
            'classe': classe,
            'total_eleves': eleves_classe.count(),
            'garcons': eleves_classe.filter(sexe='M').count(),
            'filles': eleves_classe.filter(sexe='F').count(),
            'actifs': eleves_classe.filter(statut='ACTIF').count(),
        }
        stats_par_classe.append(classe_stats)
    
    # 7. STATISTIQUES TEMPORELLES
    current_year = datetime.now().year
    current_month = datetime.now().month
    
    # Utiliser eleves_base pour respecter le filtrage par école
    stats_temporelles = {
        'inscriptions_cette_annee': eleves_base.filter(date_inscription__year=current_year).count(),
        'inscriptions_ce_mois': eleves_base.filter(
            date_inscription__year=current_year,
            date_inscription__month=current_month
        ).count(),
        'inscriptions_cette_semaine': eleves_base.filter(
            date_inscription__gte=date.today() - relativedelta(days=7)
        ).count(),
    }
    
    # Évolution mensuelle (6 derniers mois)
    evolution_mensuelle = []
    for i in range(6):
        mois_date = date.today() - relativedelta(months=i)
        # Utiliser eleves_base pour respecter le filtrage par école
        nb_inscriptions = eleves_base.filter(
            date_inscription__year=mois_date.year,
            date_inscription__month=mois_date.month
        ).count()
        
        evolution_mensuelle.append({
            'mois': mois_date.strftime('%B %Y'),
            'mois_court': mois_date.strftime('%b'),
            'inscriptions': nb_inscriptions
        })
    
    evolution_mensuelle.reverse()  # Du plus ancien au plus récent
    
    # 8. STATISTIQUES DE RESPONSABLES
    stats_responsables = {
        'total_responsables': responsables_base.count(),
        'responsables_principaux': eleves_base.values('responsable_principal').distinct().count(),
        'responsables_secondaires': eleves_base.filter(responsable_secondaire__isnull=False).values('responsable_secondaire').distinct().count(),
        'eleves_avec_deux_responsables': eleves_base.filter(responsable_secondaire__isnull=False).count(),
        'eleves_avec_un_responsable': eleves_base.filter(responsable_secondaire__isnull=True).count(),
    }
    
    # Répartition par relation
    relations_stats = []
    for relation_code, relation_nom in Responsable.RELATION_CHOICES:
        # Utiliser responsables_base pour respecter le filtrage par école
        count = responsables_base.filter(relation=relation_code).count()
        if count > 0:
            relations_stats.append({
                'relation': relation_nom,
                'count': count,
                'pourcentage': round((count / stats_responsables['total_responsables']) * 100, 1) if stats_responsables['total_responsables'] > 0 else 0
            })
    
    # 9. STATISTIQUES FINANCIÈRES (basiques)
    from paiements.models import Paiement
    
    paiements_qs = Paiement.objects.all()
    if not user_is_admin(request.user):
        paiements_qs = paiements_qs.filter(eleve__classe__ecole=user_school(request.user))
    stats_financieres = {
        'eleves_avec_paiements': eleves_base.filter(paiements__isnull=False).distinct().count(),
        'eleves_sans_paiements': eleves_base.filter(paiements__isnull=True).count(),
        'total_paiements': paiements_qs.count(),
        'paiements_valides': paiements_qs.filter(statut='VALIDE').count(),
        'paiements_en_attente': paiements_qs.filter(statut='EN_ATTENTE').count(),
    }
    
    if stats_financieres['total_paiements'] > 0:
        stats_financieres['taux_validation'] = round(
            (stats_financieres['paiements_valides'] / stats_financieres['total_paiements']) * 100, 1
        )
    else:
        stats_financieres['taux_validation'] = 0
    
    # 10. INDICATEURS DE PERFORMANCE
    indicateurs = {
        'taux_activite': round((stats_generales['eleves_actifs'] / total_eleves) * 100, 1) if total_eleves > 0 else 0,
        'taux_retention': round(((total_eleves - stats_generales['eleves_exclus'] - stats_generales['eleves_transferes']) / total_eleves) * 100, 1) if total_eleves > 0 else 0,
        'ratio_eleves_classes': round(total_eleves / stats_generales['total_classes'], 1) if stats_generales['total_classes'] > 0 else 0,
        'ratio_eleves_responsables': round(total_eleves / stats_responsables['total_responsables'], 1) if stats_responsables['total_responsables'] > 0 else 0,
    }
    
    context = {
        'stats_generales': stats_generales,
        'stats_demographiques': stats_demographiques,
        'stats_age': stats_age,
        'stats_par_ecole': stats_par_ecole,
        'stats_par_niveau': stats_par_niveau,
        'stats_par_classe': stats_par_classe,
        'stats_temporelles': stats_temporelles,
        'evolution_mensuelle': evolution_mensuelle,
        'stats_responsables': stats_responsables,
        'relations_stats': relations_stats,
        'stats_financieres': stats_financieres,
        'indicateurs': indicateurs,
        'titre_page': 'Statistiques Complètes des Élèves'
    }
    
    return render(request, 'eleves/statistiques.html', context)

@login_required
def fiche_inscription_pdf(request, eleve_id):
    """Génère la fiche d'inscription d'un élève en PDF"""
    qs = Eleve.objects.select_related(
        'classe', 'classe__ecole', 'responsable_principal', 'responsable_secondaire'
    )
    if not user_is_admin(request.user):
        qs = filter_by_user_school(qs, request.user, 'classe__ecole')
    eleve = get_object_or_404(qs, id=eleve_id)
    
    # Log de l'activité
    JournalActivite.objects.create(
        user=request.user,
        action='IMPRESSION',
        type_objet='ELEVE',
        objet_id=eleve.id,
        description=f"Impression fiche d'inscription PDF de {eleve.nom_complet}",
        adresse_ip=request.META.get('REMOTE_ADDR', ''),
        user_agent=request.META.get('HTTP_USER_AGENT', '')
    )
    
    # Créer la réponse HTTP pour le PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="fiche_inscription_{eleve.matricule}.pdf"'
    
    # Créer le PDF
    c = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    
    # Ajouter le filigrane
    try:
        from ecole_moderne.pdf_utils import draw_logo_watermark
        draw_logo_watermark(c, width, height)
    except Exception:
        pass
    
    # Configuration des polices avec détection cross-platform
    main_font_registered = False
    try:
        # Chemins possibles pour les polices selon l'OS
        font_paths = []
        
        # Windows
        if os.name == 'nt':
            font_paths.extend([
                ('C:/Windows/Fonts/calibri.ttf', 'C:/Windows/Fonts/calibrib.ttf'),
                ('C:/Windows/Fonts/arial.ttf', 'C:/Windows/Fonts/arialbd.ttf'),
            ])
        
        # Linux/Unix (PythonAnywhere, Ubuntu, etc.)
        else:
            font_paths.extend([
                ('/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf', 
                 '/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf'),
                ('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf', 
                 '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'),
                ('/System/Library/Fonts/Arial.ttf', '/System/Library/Fonts/Arial Bold.ttf'),  # macOS
            ])
        
        # Essayer chaque paire de polices
        for regular_path, bold_path in font_paths:
            if os.path.exists(regular_path) and os.path.exists(bold_path):
                try:
                    pdfmetrics.registerFont(TTFont('MainFont', regular_path))
                    pdfmetrics.registerFont(TTFont('MainFont-Bold', bold_path))
                    main_font_registered = True
                    break
                except Exception:
                    continue
        
        # Si aucune police système trouvée, utiliser les polices par défaut de ReportLab
        if not main_font_registered:
            # Les polices Helvetica sont déjà disponibles par défaut dans ReportLab
            # On n'a pas besoin de les enregistrer, juste de s'assurer qu'elles existent
            pass
            
    except Exception:
        # En cas d'erreur, s'assurer que les alias existent
        main_font_registered = False
    
    # Compression PDF pour meilleure qualité
    c.setPageCompression(1)
    
    # Filigrane avec logo de l'école (même taille que les autres exports PDF)
    c.saveState()
    try:
        # Chemin vers le logo de l'école s'il existe
        school_logo = getattr(eleve.classe.ecole, 'logo', None)
        logo_path = ''
        try:
            if school_logo and hasattr(school_logo, 'path') and os.path.exists(school_logo.path):
                logo_path = school_logo.path
            else:
                logo_path = os.path.join('static', 'logos', 'logo.png')
        except Exception:
            logo_path = os.path.join('static', 'logos', 'logo.png')
        if os.path.exists(logo_path):
            # Taille ~150% de la largeur de page (comme dans rapports/utils.py)
            wm_width = width * 1.5
            wm_height = wm_width  # carré approximatif, preserveAspectRatio activera le ratio réel
            wm_x = (width - wm_width) / 2
            wm_y = (height - wm_height) / 2
            
            # Opacité visible mais discrète (comme dans les reçus de paiement)
            try:
                c.setFillAlpha(0.15)
            except Exception:
                pass
            
            # Légère rotation pour l'effet filigrane
            c.translate(width / 2.0, height / 2.0)
            c.rotate(30)
            c.translate(-width / 2.0, -height / 2.0)
            
            c.drawImage(logo_path, wm_x, wm_y, width=wm_width, height=wm_height, preserveAspectRatio=True, mask='auto')
        else:
            # Fallback vers texte si logo non trouvé
            c.setFillAlpha(0.04)
            try:
                c.setFont("MainFont-Bold", 60)
            except:
                c.setFont("Helvetica-Bold", 60)
            c.rotate(45)
            c.drawString(200, -100, eleve.classe.ecole.nom.upper())
            c.rotate(-45)
    finally:
        c.restoreState()
    
    # En-tête avec logo de l'école
    y = height - 2*cm
    
    # Logo en en-tête (côté gauche)
    try:
        from django.contrib.staticfiles import finders
        school_logo = getattr(eleve.classe.ecole, 'logo', None)
        logo_path = None
        if school_logo and hasattr(school_logo, 'path') and os.path.exists(school_logo.path):
            logo_path = school_logo.path
        else:
            logo_path = finders.find('logos/logo.png')
        
        if logo_path:
            try:
                logo_w, logo_h = 80, 80
                c.drawImage(logo_path, 2*cm, y - logo_h, width=logo_w, height=logo_h, preserveAspectRatio=True, mask='auto')
                
                # Titre à côté du logo
                try:
                    c.setFont("MainFont-Bold", 18)
                except:
                    c.setFont("Helvetica-Bold", 18)
                
                text = eleve.classe.ecole.nom.upper()
                c.drawString(2*cm + logo_w + 20, y - 25, text)
                
                # Sous-titre
                try:
                    c.setFont("MainFont-Bold", 16)
                except:
                    c.setFont("Helvetica-Bold", 16)
                
                text = "FICHE D'INSCRIPTION"
                c.drawString(2*cm + logo_w + 20, y - 50, text)
                
                y -= (logo_h + 20)
            except Exception:
                # Fallback sans logo
                try:
                    c.setFont("MainFont-Bold", 16)
                except:
                    c.setFont("Helvetica-Bold", 16)
                
                text = eleve.classe.ecole.nom.upper()
                text_width = c.stringWidth(text, "MainFont-Bold", 16) if "MainFont-Bold" in c.getAvailableFonts() else c.stringWidth(text, "Helvetica-Bold", 16)
                c.drawString((width - text_width) / 2, y, text)
                y -= 0.8*cm
                
                text = "FICHE D'INSCRIPTION"
                text_width = c.stringWidth(text, "MainFont-Bold", 16) if "MainFont-Bold" in c.getAvailableFonts() else c.stringWidth(text, "Helvetica-Bold", 16)
                c.drawString((width - text_width) / 2, y, text)
        else:
            # Fallback sans logo
            try:
                c.setFont("MainFont-Bold", 16)
            except:
                c.setFont("Helvetica-Bold", 16)
            
            text = eleve.classe.ecole.nom.upper()
            text_width = c.stringWidth(text, "MainFont-Bold", 16) if "MainFont-Bold" in c.getAvailableFonts() else c.stringWidth(text, "Helvetica-Bold", 16)
            c.drawString((width - text_width) / 2, y, text)
            y -= 0.8*cm
            
            text = "FICHE D'INSCRIPTION"
            text_width = c.stringWidth(text, "MainFont-Bold", 16) if "MainFont-Bold" in c.getAvailableFonts() else c.stringWidth(text, "Helvetica-Bold", 16)
            c.drawString((width - text_width) / 2, y, text)
    except Exception:
        # Fallback complet en cas d'erreur
        try:
            c.setFont("MainFont-Bold", 16)
        except:
            c.setFont("Helvetica-Bold", 16)
        
        text = eleve.classe.ecole.nom.upper()
        text_width = c.stringWidth(text, "MainFont-Bold", 16) if "MainFont-Bold" in c.getAvailableFonts() else c.stringWidth(text, "Helvetica-Bold", 16)
        c.drawString((width - text_width) / 2, y, text)
        y -= 0.8*cm
        
        text = "FICHE D'INSCRIPTION"
        text_width = c.stringWidth(text, "MainFont-Bold", 16) if "MainFont-Bold" in c.getAvailableFonts() else c.stringWidth(text, "Helvetica-Bold", 16)
        c.drawString((width - text_width) / 2, y, text)
    
    y -= 1.5*cm
    
    # Photo de l'élève (si disponible)
    photo_x = width - 4*cm
    photo_y = y - 3*cm
    photo_width = 3*cm
    photo_height = 4*cm
    
    if eleve.photo:
        try:
            c.drawImage(eleve.photo.path, photo_x, photo_y, width=photo_width, height=photo_height)
        except Exception:
            # Placeholder si l'image ne peut pas être chargée
            c.rect(photo_x, photo_y, photo_width, photo_height)
            try:
                c.setFont("MainFont", 12)
            except:
                c.setFont("Helvetica", 12)
            text = "Photo"
            text_width = c.stringWidth(text, "MainFont", 12) if "MainFont" in c.getAvailableFonts() else c.stringWidth(text, "Helvetica", 12)
            c.drawString(photo_x + (photo_width - text_width)/2, photo_y + photo_height/2, text)
    else:
        # Placeholder pour photo
        c.rect(photo_x, photo_y, photo_width, photo_height)
        try:
            c.setFont("MainFont", 12)
        except:
            c.setFont("Helvetica", 12)
        text = "Photo"
        text_width = c.stringWidth(text, "MainFont", 12) if "MainFont" in c.getAvailableFonts() else c.stringWidth(text, "Helvetica", 12)
        c.drawString(photo_x + (photo_width - text_width)/2, photo_y + photo_height/2, text)
    
    # Informations de l'élève
    try:
        c.setFont("MainFont-Bold", 14)
    except:
        c.setFont("Helvetica-Bold", 14)
    
    # Section Informations personnelles
    c.drawString(2*cm, y, "INFORMATIONS PERSONNELLES")
    y -= 0.8*cm
    
    try:
        c.setFont("MainFont", 12)
    except:
        c.setFont("Helvetica", 12)
    
    # Colonne gauche
    left_col = 2*cm
    right_col = 10*cm
    line_height = 0.6*cm
    
    date_inscription_str = eleve.date_inscription.strftime('%d/%m/%Y') if eleve.date_inscription else "Non renseignée"
    date_naissance_str = eleve.date_naissance.strftime('%d/%m/%Y') if eleve.date_naissance else "Non renseignée"
    age_str = f"{eleve.age} ans" if eleve.age is not None else "Non renseigné"

    c.drawString(left_col, y, f"Matricule: {eleve.matricule}")
    c.drawString(right_col, y, f"Date d'inscription: {date_inscription_str}")
    y -= line_height

    c.drawString(left_col, y, f"Nom: {eleve.nom}")
    c.drawString(right_col, y, f"Prénom: {eleve.prenom}")
    y -= line_height

    c.drawString(left_col, y, f"Sexe: {eleve.get_sexe_display()}")
    c.drawString(right_col, y, f"Date de naissance: {date_naissance_str}")
    y -= line_height

    c.drawString(left_col, y, f"Lieu de naissance: {eleve.lieu_naissance}")
    c.drawString(right_col, y, f"Âge: {age_str}")
    y -= line_height
    
    c.drawString(left_col, y, f"Statut: {eleve.get_statut_display()}")
    y -= line_height * 1.5
    
    # Section Informations scolaires
    try:
        c.setFont("MainFont-Bold", 14)
    except:
        c.setFont("Helvetica-Bold", 14)
    
    c.drawString(left_col, y, "INFORMATIONS SCOLAIRES")
    y -= 0.8*cm
    
    try:
        c.setFont("MainFont", 12)
    except:
        c.setFont("Helvetica", 12)
    
    c.drawString(left_col, y, f"École: {eleve.classe.ecole.nom}")
    y -= line_height
    
    c.drawString(left_col, y, f"Classe: {eleve.classe.nom}")
    c.drawString(right_col, y, f"Niveau: {eleve.classe.get_niveau_display()}")
    y -= line_height
    
    c.drawString(left_col, y, f"Année scolaire: {eleve.classe.annee_scolaire}")
    y -= line_height * 1.5
    
    # Section Responsable principal
    try:
        c.setFont("MainFont-Bold", 14)
    except:
        c.setFont("Helvetica-Bold", 14)
    
    c.drawString(left_col, y, "RESPONSABLE PRINCIPAL")
    y -= 0.8*cm
    
    try:
        c.setFont("MainFont", 12)
    except:
        c.setFont("Helvetica", 12)
    
    resp_principal = eleve.responsable_principal
    c.drawString(left_col, y, f"Nom complet: {resp_principal.nom_complet}")
    c.drawString(right_col, y, f"Relation: {resp_principal.get_relation_display()}")
    y -= line_height
    
    c.drawString(left_col, y, f"Téléphone: {resp_principal.telephone}")
    if resp_principal.email:
        c.drawString(right_col, y, f"Email: {resp_principal.email}")
    y -= line_height
    
    if resp_principal.profession:
        c.drawString(left_col, y, f"Profession: {resp_principal.profession}")
        y -= line_height
    
    if resp_principal.adresse:
        c.drawString(left_col, y, f"Adresse: {resp_principal.adresse}")
        y -= line_height
    
    # Section Responsable secondaire (si existe)
    if eleve.responsable_secondaire:
        y -= line_height * 0.5
        
        try:
            c.setFont("MainFont-Bold", 14)
        except:
            c.setFont("Helvetica-Bold", 14)
        
        c.drawString(left_col, y, "RESPONSABLE SECONDAIRE")
        y -= 0.8*cm
        
        try:
            c.setFont("MainFont", 12)
        except:
            c.setFont("Helvetica", 12)
        
        resp_secondaire = eleve.responsable_secondaire
        c.drawString(left_col, y, f"Nom complet: {resp_secondaire.nom_complet}")
        c.drawString(right_col, y, f"Relation: {resp_secondaire.get_relation_display()}")
        y -= line_height
        
        c.drawString(left_col, y, f"Téléphone: {resp_secondaire.telephone}")
        if resp_secondaire.email:
            c.drawString(right_col, y, f"Email: {resp_secondaire.email}")
        y -= line_height
        
        if resp_secondaire.profession:
            c.drawString(left_col, y, f"Profession: {resp_secondaire.profession}")
            y -= line_height
        
        if resp_secondaire.adresse:
            c.drawString(left_col, y, f"Adresse: {resp_secondaire.adresse}")
            y -= line_height
    
    # Pied de page
    y = 3*cm
    try:
        c.setFont("MainFont", 10)
    except:
        c.setFont("Helvetica", 10)
    
    text = f"Fiche générée le {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
    text_width = c.stringWidth(text, "MainFont", 10) if "MainFont" in c.getAvailableFonts() else c.stringWidth(text, "Helvetica", 10)
    c.drawString((width - text_width) / 2, y, text)
    y -= 0.5*cm
    
    text = f"Système de Gestion Scolaire - {eleve.classe.ecole.nom}"
    text_width = c.stringWidth(text, "MainFont", 10) if "MainFont" in c.getAvailableFonts() else c.stringWidth(text, "Helvetica", 10)
    c.drawString((width - text_width) / 2, y, text)
    
    # Finaliser le PDF
    c.save()
    
    return response

@login_required
def ajax_rechercher_responsable_telephone(request):
    """Vue AJAX optimisée pour rechercher un responsable par numéro de téléphone (cache géré manuellement)"""
    if request.method != 'GET':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    telephone = request.GET.get('telephone', '').strip()
    if not telephone or len(telephone) < 3:  # Minimum 3 caractères
        return JsonResponse({'success': False, 'error': 'Numéro de téléphone requis (min. 3 caractères)'})
    
    try:
        # Cache de l'école utilisateur
        user_school_cache_key = f'user_school_{request.user.id}'
        user_school_obj = cache.get(user_school_cache_key)
        
        if user_school_obj is None and not user_is_admin(request.user):
            user_school_obj = user_school(request.user)
            if user_school_obj:
                cache.set(user_school_cache_key, user_school_obj, 300)
        
        # Cache de la recherche
        search_cache_key = f'search_resp_{request.user.id}_{telephone}'
        responsables_list = cache.get(search_cache_key)
        
        if responsables_list is None:
            # Base queryset optimisée
            base_qs = Responsable.objects.select_related()
            
            # Filtrer par école pour les non-admins
            if not user_is_admin(request.user) and user_school_obj:
                base_qs = base_qs.filter(
                    Q(eleves_principal__classe__ecole=user_school_obj) | 
                    Q(eleves_secondaire__classe__ecole=user_school_obj)
                )
            elif not user_is_admin(request.user):
                base_qs = base_qs.none()
            
            # Recherche optimisée par téléphone
            search_conditions = Q(telephone__icontains=telephone)
            if len(telephone) >= 8:
                search_conditions |= Q(telephone__endswith=telephone[-8:])
            
            responsables = base_qs.filter(search_conditions).distinct().values(
                'id', 'prenom', 'nom', 'relation', 'telephone', 'email', 'profession', 'adresse'
            )[:5]
            
            responsables_list = list(responsables)
            
            # Cache le résultat pour 30 secondes
            cache.set(search_cache_key, responsables_list, 30)
        
        # Log optimisé (asynchrone si possible)
        if len(responsables_list) > 0:  # Log seulement si résultats trouvés
            try:
                JournalActivite.objects.create(
                    user=request.user,
                    action='RECHERCHE',
                    type_objet='RESPONSABLE',
                    description=f"Recherche responsable par téléphone: {telephone} ({len(responsables_list)} résultats)",
                    adresse_ip=request.META.get('REMOTE_ADDR', ''),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')[:200]  # Limiter la taille
                )
            except:
                pass  # Ne pas faire échouer la recherche pour un problème de log
        
        return JsonResponse({
            'success': True,
            'responsables': responsables_list,
            'count': len(responsables_list)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors de la recherche: {str(e)}'
        })


@login_required
def generer_ticket_retrait_pdf(request, eleve_id):
    """Génère un ticket de retrait pour un élève du primaire"""
    eleve = get_object_or_404(
        Eleve.objects.select_related('classe', 'classe__ecole', 'responsable_principal'),
        id=eleve_id
    )
    
    # Vérifier les permissions
    if not user_is_admin(request.user):
        user_school_obj = user_school(request.user)
        if not user_school_obj or eleve.classe.ecole != user_school_obj:
            messages.error(request, "Vous n'avez pas accès à cet élève.")
            return redirect('eleves:liste_eleves')
    
    # Vérifier que c'est un élève du primaire
    niveau = eleve.classe.niveau.upper() if eleve.classe.niveau else ''
    if not any(x in niveau for x in ['PRIMAIRE', 'PN', 'MATERNELLE', 'GARDERIE']):
        messages.warning(request, "Les tickets de retrait sont réservés aux élèves du primaire et de la maternelle.")
        return redirect('eleves:detail_eleve', eleve_id=eleve_id)
    
    # Créer le PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="ticket_retrait_{eleve.matricule}.pdf"'
    
    # Format carte bancaire standard (86mm x 54mm)
    from reportlab.lib.units import mm
    width, height = 86*mm, 54*mm
    
    c = canvas.Canvas(response, pagesize=(width, height))
    
    # Enregistrer les polices
    try:
        pdfmetrics.registerFont(TTFont('Arial', 'C:/Windows/Fonts/arial.ttf'))
        pdfmetrics.registerFont(TTFont('Arial-Bold', 'C:/Windows/Fonts/arialbd.ttf'))
        main_font = 'Arial'
        main_font_bold = 'Arial-Bold'
    except:
        main_font = 'Helvetica'
        main_font_bold = 'Helvetica-Bold'
    
    _dessiner_ticket_retrait(c, eleve, 0, 0, width, height, main_font, main_font_bold)
    c.showPage()
    c.save()

    # Log de l'action
    try:
        JournalActivite.objects.create(
            user=request.user,
            action='GENERATION_PDF',
            type_objet='TICKET_RETRAIT',
            description=f"Ticket de retrait genere pour {eleve.prenom} {eleve.nom} ({eleve.matricule})",
            adresse_ip=request.META.get('REMOTE_ADDR', ''),
            user_agent=request.META.get('HTTP_USER_AGENT', '')[:200]
        )
    except:
        pass

    return response

    # Extraire les couleurs du logo
    primary_color = '#3b82f6'
    light_color = '#dbeafe'
    
    try:
        if eleve.classe.ecole.logo and hasattr(eleve.classe.ecole.logo, 'path'):
            logo_path = eleve.classe.ecole.logo.path
            if os.path.exists(logo_path):
                primary_color, light_color = _extraire_couleurs_logo(logo_path)
    except:
        pass
    
    # Fond blanc
    c.setFillColor(colors.white)
    c.rect(0, 0, width, height, stroke=0, fill=1)
    
    # Ajouter le logo en filigrane au centre
    c.saveState()
    try:
        logo_path = None
        
        # Récupérer le logo de l'école
        if eleve.classe and eleve.classe.ecole and eleve.classe.ecole.logo:
            try:
                if hasattr(eleve.classe.ecole.logo, 'path'):
                    if os.path.exists(eleve.classe.ecole.logo.path):
                        logo_path = eleve.classe.ecole.logo.path
            except Exception:
                pass
        
        if logo_path and os.path.exists(logo_path):
            # Centrer le logo
            c.translate(width/2, height/2)
            c.rotate(25)  # Rotation légère
            
            # Taille du filigrane (plus grand)
            watermark_size = min(width, height) * 0.7
            
            # Dessiner le logo en filigrane avec opacité visible
            c.setFillAlpha(0.12)
            c.drawImage(logo_path, -watermark_size/2, -watermark_size/2, 
                      watermark_size, watermark_size, preserveAspectRatio=True)
    except Exception:
        pass
    c.restoreState()
    
    # Formes géométriques décoratives en arrière-plan (cercles)
    c.setFillColor(colors.HexColor(light_color))
    c.setFillAlpha(0.15)
    c.circle(width - 20, height - 20, 40, stroke=0, fill=1)
    c.circle(15, 15, 30, stroke=0, fill=1)
    c.setFillAlpha(1)
    
    # Bordure avec coins arrondis
    c.setStrokeColor(colors.HexColor(primary_color))
    c.setLineWidth(2.5)
    c.roundRect(2, 2, width-4, height-4, 10, stroke=1, fill=0)
    
    # Bande diagonale décorative en haut à gauche
    c.saveState()
    c.translate(0, height)
    c.rotate(-25)
    c.setFillColor(colors.HexColor(primary_color))
    c.setFillAlpha(0.1)
    c.rect(-10, -15, 60, 25, stroke=0, fill=1)
    c.setFillAlpha(1)
    c.restoreState()
    
    # En-tête rempli avec couleur (rectangle plein)
    c.setFillColor(colors.HexColor(primary_color))
    c.roundRect(5, height-45, width-10, 40, 8, stroke=0, fill=1)
    
    # Logo de l'école en haut à droite (petit format visible)
    try:
        logo_path_visible = None
        if eleve.classe and eleve.classe.ecole and eleve.classe.ecole.logo:
            try:
                if hasattr(eleve.classe.ecole.logo, 'path'):
                    if os.path.exists(eleve.classe.ecole.logo.path):
                        logo_path_visible = eleve.classe.ecole.logo.path
            except Exception:
                pass
        
        if logo_path_visible and os.path.exists(logo_path_visible):
            from PIL import Image, ImageDraw
            
            # Position et taille du logo
            logo_size = 25
            logo_x = width - logo_size - 8
            logo_y = height - logo_size - 10
            
            # Ouvrir et traiter l'image
            img = Image.open(logo_path_visible)
            if img.mode != 'RGB':
                img = img.convert('RGB')
            
            # Créer une image circulaire
            size = (logo_size, logo_size)
            img = img.resize(size, Image.Resampling.LANCZOS)
            
            # Créer un masque circulaire
            mask = Image.new('L', size, 0)
            draw = ImageDraw.Draw(mask)
            draw.ellipse((0, 0) + size, fill=255)
            
            # Appliquer le masque
            output = Image.new('RGB', size, (255, 255, 255))
            output.paste(img, (0, 0))
            output.putalpha(mask)
            
            # Sauvegarder temporairement
            temp_buffer = io.BytesIO()
            output.save(temp_buffer, format='PNG')
            temp_buffer.seek(0)
            
            # Cercle blanc derrière le logo
            c.setFillColor(colors.white)
            c.circle(logo_x + logo_size/2, logo_y + logo_size/2, logo_size/2 + 2, stroke=0, fill=1)
            
            # Dessiner le logo
            c.drawImage(temp_buffer, logo_x, logo_y, width=logo_size, height=logo_size, mask='auto')
    except Exception:
        pass
    
    # Titre
    c.setFillColor(colors.white)
    c.setFont(main_font_bold, 13)
    c.drawCentredString(width/2, height-20, "TICKET DE RETRAIT")
    
    # Sous-titre école
    c.setFont(main_font, 7)
    c.setFillAlpha(0.9)
    c.drawCentredString(width/2, height-32, eleve.classe.ecole.nom[:50])
    c.setFillAlpha(1)
    
    # Photo de l'élève (côté droit) - affichage direct uniquement si disponible
    photo_x = width - 45
    photo_y = height/2 - 8
    photo_radius = 30
    
    # Afficher la photo UNIQUEMENT si elle existe
    if eleve.photo:
        try:
            from PIL import Image, ImageDraw
            photo_path = eleve.photo.path
            if os.path.exists(photo_path):
                # Ombre de la photo
                c.setFillColor(colors.HexColor('#000000'))
                c.setFillAlpha(0.1)
                c.circle(photo_x + 1, photo_y - 1, photo_radius + 2, stroke=0, fill=1)
                c.setFillAlpha(1)
                
                # Cercle de fond blanc avec double bordure
                c.setFillColor(colors.white)
                c.circle(photo_x, photo_y, photo_radius + 2, stroke=0, fill=1)
                c.setStrokeColor(colors.HexColor(primary_color))
                c.setLineWidth(3)
                c.circle(photo_x, photo_y, photo_radius, stroke=1, fill=1)
                
                # Ouvrir l'image avec PIL
                img = Image.open(photo_path)
                
                # Convertir en RGB si nécessaire
                if img.mode in ('RGBA', 'LA', 'P'):
                    background = Image.new('RGB', img.size, (255, 255, 255))
                    if img.mode == 'P':
                        img = img.convert('RGBA')
                    background.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
                    img = background
                elif img.mode != 'RGB':
                    img = img.convert('RGB')
                
                # Calculer la taille en pixels
                from reportlab.lib.units import cm
                pixel_size = int(photo_radius * 2 * 28.35)
                size = (pixel_size, pixel_size)
                
                # Redimensionner
                img = img.resize(size, Image.Resampling.LANCZOS)
                
                # Créer un masque circulaire
                mask = Image.new('L', size, 0)
                draw = ImageDraw.Draw(mask)
                draw.ellipse((0, 0, size[0], size[1]), fill=255)
                
                # Appliquer le masque
                output = Image.new('RGBA', size, (255, 255, 255, 0))
                output.paste(img, (0, 0))
                output.putalpha(mask)
                
                # Sauvegarder temporairement
                temp_buffer = io.BytesIO()
                output.save(temp_buffer, format='PNG')
                temp_buffer.seek(0)
                
                # Dessiner sur le PDF avec ImageReader
                from reportlab.lib.utils import ImageReader
                img_reader = ImageReader(temp_buffer)
                c.drawImage(img_reader, photo_x - photo_radius, photo_y - photo_radius, 
                           width=photo_radius * 2, height=photo_radius * 2, mask='auto')
        except Exception as e:
            # Ne rien afficher en cas d'erreur
            print(f"Erreur photo retrait: {e}")
            pass
    # Si pas de photo, on n'affiche rien (pas de cercle ni de placeholder)
    
    # Carte d'information avec fond coloré (style moderne)
    info_box_x = 8
    info_box_y = 8
    info_box_width = width - 50
    info_box_height = height - 52
    
    c.setFillColor(colors.HexColor(light_color))
    c.setFillAlpha(0.08)
    c.roundRect(info_box_x, info_box_y, info_box_width, info_box_height, 8, stroke=0, fill=1)
    c.setFillAlpha(1)
    
    # Informations de l'élève (poussées vers la droite)
    x_start = 12
    y_start = height/2 + 5
    
    # Nom complet (taille encore réduite)
    c.setFillColor(colors.HexColor('#1f2937'))
    c.setFont(main_font_bold, 9)
    nom_complet = f"{eleve.prenom} {eleve.nom}".upper()
    if len(nom_complet) > 20:
        nom_complet = nom_complet[:17] + "..."
    c.drawString(x_start, y_start, nom_complet)
    
    # Ligne décorative sous le nom (ajustée à la longueur du nom)
    nom_width = c.stringWidth(nom_complet, main_font_bold, 9)
    c.setStrokeColor(colors.HexColor(primary_color))
    c.setLineWidth(2)
    c.line(x_start, y_start - 2, x_start + nom_width, y_start - 2)
    
    # Informations détaillées avec espacement
    y = y_start - 16
    c.setFillColor(colors.HexColor('#374151'))
    
    # Matricule avec espacement après deux-points
    c.setFont(main_font_bold, 9)
    c.drawString(x_start, y, "N° :  ")
    c.setFont(main_font, 9)
    c.drawString(x_start + 25, y, eleve.matricule or "N/A")
    
    y -= 13
    # Classe avec espacement après deux-points
    c.setFont(main_font_bold, 9)
    c.drawString(x_start, y, "Classe :  ")
    c.setFont(main_font, 9)
    c.drawString(x_start + 35, y, eleve.classe.nom)
    
    y -= 13
    # Responsable avec espacement après deux-points
    c.setFont(main_font_bold, 9)
    c.drawString(x_start, y, "Parent :  ")
    c.setFont(main_font, 8)
    if eleve.responsable_principal:
        resp_nom = f"{eleve.responsable_principal.prenom} {eleve.responsable_principal.nom}"
        if len(resp_nom) > 18:
            resp_nom = resp_nom[:15] + "..."
        c.drawString(x_start + 35, y, resp_nom)
    else:
        c.drawString(x_start + 35, y, "Non renseigné")
    
    y -= 13
    # Téléphone du parent avec espacement après deux-points
    c.setFont(main_font_bold, 9)
    c.drawString(x_start, y, "Tél :  ")
    c.setFont(main_font, 8)
    if eleve.responsable_principal and eleve.responsable_principal.telephone:
        tel = eleve.responsable_principal.telephone[:14]
        c.drawString(x_start + 25, y, tel)
    else:
        c.drawString(x_start + 25, y, "Non renseigné")
    
    # Pied de page moderne
    c.setFillColor(colors.HexColor(primary_color))
    c.setFillAlpha(0.05)
    c.roundRect(6, 4, width-12, 9, 4, stroke=0, fill=1)
    c.setFillAlpha(1)
    
    c.setFont(main_font, 6)
    c.setFillColor(colors.HexColor('#6b7280'))
    c.drawCentredString(width/2, 6, f"Généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')}")
    
    c.showPage()
    c.save()
    
    # Log de l'action
    try:
        JournalActivite.objects.create(
            user=request.user,
            action='GENERATION_PDF',
            type_objet='TICKET_RETRAIT',
            description=f"Ticket de retrait généré pour {eleve.prenom} {eleve.nom} ({eleve.matricule})",
            adresse_ip=request.META.get('REMOTE_ADDR', ''),
            user_agent=request.META.get('HTTP_USER_AGENT', '')[:200]
        )
    except:
        pass
    
    return response


@login_required
def generer_ticket_bus_pdf(request, eleve_id):
    """Génère un ticket d'abonnement bus pour un élève"""
    from bus.models import AbonnementBus
    
    eleve = get_object_or_404(
        Eleve.objects.select_related('classe', 'classe__ecole', 'responsable_principal'),
        id=eleve_id
    )
    
    # Vérifier les permissions
    if not user_is_admin(request.user):
        user_school_obj = user_school(request.user)
        if not user_school_obj or eleve.classe.ecole != user_school_obj:
            messages.error(request, "Vous n'avez pas accès à cet élève.")
            return redirect('eleves:liste_eleves')
    
    # Récupérer l'abonnement bus actif
    abonnement = AbonnementBus.objects.filter(
        eleve=eleve,
        statut='ACTIF'
    ).order_by('-date_debut').first()
    
    if not abonnement:
        messages.warning(request, "Cet élève n'a pas d'abonnement bus actif.")
        return redirect('eleves:detail_eleve', eleve_id=eleve_id)
    
    # Créer le PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="ticket_bus_{eleve.matricule}.pdf"'
    
    # Format carte bancaire standard (86mm x 54mm)
    from reportlab.lib.units import mm
    width, height = 86*mm, 54*mm
    
    c = canvas.Canvas(response, pagesize=(width, height))
    
    # Enregistrer les polices
    try:
        pdfmetrics.registerFont(TTFont('Arial', 'C:/Windows/Fonts/arial.ttf'))
        pdfmetrics.registerFont(TTFont('Arial-Bold', 'C:/Windows/Fonts/arialbd.ttf'))
        main_font = 'Arial'
        main_font_bold = 'Arial-Bold'
    except:
        main_font = 'Helvetica'
        main_font_bold = 'Helvetica-Bold'
    
    _dessiner_ticket_bus(c, eleve, abonnement, 0, 0, width, height, main_font, main_font_bold)
    c.showPage()
    c.save()

    # Log de l'action
    try:
        JournalActivite.objects.create(
            user=request.user,
            action='GENERATION_PDF',
            type_objet='TICKET_BUS',
            description=f"Ticket bus genere pour {eleve.prenom} {eleve.nom} ({eleve.matricule})",
            adresse_ip=request.META.get('REMOTE_ADDR', ''),
            user_agent=request.META.get('HTTP_USER_AGENT', '')[:200]
        )
    except:
        pass

    return response

    # Extraire les couleurs du logo (thème orange pour bus)
    primary_color = '#f59e0b'
    light_color = '#fef3c7'
    
    try:
        if eleve.classe.ecole.logo and hasattr(eleve.classe.ecole.logo, 'path'):
            logo_path = eleve.classe.ecole.logo.path
            if os.path.exists(logo_path):
                primary_color, light_color = _extraire_couleurs_logo(logo_path)
    except:
        pass
    
    # Fond blanc
    c.setFillColor(colors.white)
    c.rect(0, 0, width, height, stroke=0, fill=1)
    
    # Ajouter le logo en filigrane au centre
    c.saveState()
    try:
        logo_path = None
        
        # Récupérer le logo de l'école
        if eleve.classe and eleve.classe.ecole and eleve.classe.ecole.logo:
            try:
                if hasattr(eleve.classe.ecole.logo, 'path'):
                    if os.path.exists(eleve.classe.ecole.logo.path):
                        logo_path = eleve.classe.ecole.logo.path
            except Exception:
                pass
        
        if logo_path and os.path.exists(logo_path):
            # Centrer le logo
            c.translate(width/2, height/2)
            c.rotate(25)  # Rotation légère
            
            # Taille du filigrane
            watermark_size = min(width, height) * 0.7
            
            # Dessiner le logo en filigrane avec opacité visible
            c.setFillAlpha(0.12)
            c.drawImage(logo_path, -watermark_size/2, -watermark_size/2, 
                      watermark_size, watermark_size, preserveAspectRatio=True)
    except Exception:
        pass
    c.restoreState()
    
    # Formes géométriques décoratives en arrière-plan (cercles)
    c.setFillColor(colors.HexColor(light_color))
    c.setFillAlpha(0.15)
    c.circle(width - 20, height - 20, 40, stroke=0, fill=1)
    c.circle(15, 15, 30, stroke=0, fill=1)
    c.setFillAlpha(1)
    
    # Bordure avec coins arrondis
    c.setStrokeColor(colors.HexColor(primary_color))
    c.setLineWidth(2.5)
    c.roundRect(2, 2, width-4, height-4, 10, stroke=1, fill=0)
    
    # Bande diagonale décorative en haut à gauche
    c.saveState()
    c.translate(0, height)
    c.rotate(-25)
    c.setFillColor(colors.HexColor(primary_color))
    c.setFillAlpha(0.1)
    c.rect(-10, -15, 60, 25, stroke=0, fill=1)
    c.setFillAlpha(1)
    c.restoreState()
    
    # En-tête rempli avec couleur (rectangle plein)
    c.setFillColor(colors.HexColor(primary_color))
    c.roundRect(5, height-45, width-10, 40, 8, stroke=0, fill=1)
    
    # Logo de l'école en haut à droite (petit format visible)
    try:
        logo_path_visible = None
        if eleve.classe and eleve.classe.ecole and eleve.classe.ecole.logo:
            try:
                if hasattr(eleve.classe.ecole.logo, 'path'):
                    if os.path.exists(eleve.classe.ecole.logo.path):
                        logo_path_visible = eleve.classe.ecole.logo.path
            except Exception:
                pass
        
        if logo_path_visible and os.path.exists(logo_path_visible):
            from PIL import Image, ImageDraw
            
            # Position et taille du logo
            logo_size = 25
            logo_x = width - logo_size - 8
            logo_y = height - logo_size - 10
            
            # Ouvrir et traiter l'image
            img = Image.open(logo_path_visible)
            if img.mode != 'RGB':
                img = img.convert('RGB')
            
            # Créer une image circulaire
            size = (logo_size, logo_size)
            img = img.resize(size, Image.Resampling.LANCZOS)
            
            # Créer un masque circulaire
            mask = Image.new('L', size, 0)
            draw = ImageDraw.Draw(mask)
            draw.ellipse((0, 0) + size, fill=255)
            
            # Appliquer le masque
            output = Image.new('RGB', size, (255, 255, 255))
            output.paste(img, (0, 0))
            output.putalpha(mask)
            
            # Sauvegarder temporairement
            temp_buffer = io.BytesIO()
            output.save(temp_buffer, format='PNG')
            temp_buffer.seek(0)
            
            # Cercle blanc derrière le logo
            c.setFillColor(colors.white)
            c.circle(logo_x + logo_size/2, logo_y + logo_size/2, logo_size/2 + 2, stroke=0, fill=1)
            
            # Dessiner le logo
            c.drawImage(temp_buffer, logo_x, logo_y, width=logo_size, height=logo_size, mask='auto')
    except Exception:
        pass
    
    # Titre
    c.setFillColor(colors.white)
    c.setFont(main_font_bold, 13)
    c.drawCentredString(width/2, height-20, "ABONNEMENT BUS")
    
    # Sous-titre école
    c.setFont(main_font, 7)
    c.setFillAlpha(0.9)
    c.drawCentredString(width/2, height-32, eleve.classe.ecole.nom[:50])
    c.setFillAlpha(1)
    
    # Photo de l'élève (côté droit) - affichage direct uniquement si disponible
    photo_x = width - 45
    photo_y = height/2 - 8
    photo_radius = 30
    
    # Afficher la photo UNIQUEMENT si elle existe
    if eleve.photo:
        try:
            from PIL import Image, ImageDraw
            photo_path = eleve.photo.path
            if os.path.exists(photo_path):
                # Ombre de la photo
                c.setFillColor(colors.HexColor('#000000'))
                c.setFillAlpha(0.1)
                c.circle(photo_x + 1, photo_y - 1, photo_radius + 2, stroke=0, fill=1)
                c.setFillAlpha(1)
                
                # Cercle de fond blanc avec double bordure
                c.setFillColor(colors.white)
                c.circle(photo_x, photo_y, photo_radius + 2, stroke=0, fill=1)
                c.setStrokeColor(colors.HexColor(primary_color))
                c.setLineWidth(3)
                c.circle(photo_x, photo_y, photo_radius, stroke=1, fill=1)
                
                # Ouvrir l'image avec PIL
                img = Image.open(photo_path)
                
                # Convertir en RGB si nécessaire
                if img.mode in ('RGBA', 'LA', 'P'):
                    background = Image.new('RGB', img.size, (255, 255, 255))
                    if img.mode == 'P':
                        img = img.convert('RGBA')
                    background.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
                    img = background
                elif img.mode != 'RGB':
                    img = img.convert('RGB')
                
                # Calculer la taille en pixels
                from reportlab.lib.units import cm
                pixel_size = int(photo_radius * 2 * 28.35)
                size = (pixel_size, pixel_size)
                
                # Redimensionner
                img = img.resize(size, Image.Resampling.LANCZOS)
                
                # Créer un masque circulaire
                mask = Image.new('L', size, 0)
                draw = ImageDraw.Draw(mask)
                draw.ellipse((0, 0, size[0], size[1]), fill=255)
                
                # Appliquer le masque
                output = Image.new('RGBA', size, (255, 255, 255, 0))
                output.paste(img, (0, 0))
                output.putalpha(mask)
                
                # Sauvegarder temporairement
                temp_buffer = io.BytesIO()
                output.save(temp_buffer, format='PNG')
                temp_buffer.seek(0)
                
                # Dessiner sur le PDF avec ImageReader
                from reportlab.lib.utils import ImageReader
                img_reader = ImageReader(temp_buffer)
                c.drawImage(img_reader, photo_x - photo_radius, photo_y - photo_radius, 
                           width=photo_radius * 2, height=photo_radius * 2, mask='auto')
        except Exception as e:
            # Ne rien afficher en cas d'erreur
            print(f"Erreur photo bus: {e}")
            pass
    # Si pas de photo, on n'affiche rien (pas de cercle ni de placeholder)
    
    # Carte d'information avec fond coloré (style moderne)
    info_box_x = 8
    info_box_y = 8
    info_box_width = width - 50
    info_box_height = height - 52
    
    c.setFillColor(colors.HexColor(light_color))
    c.setFillAlpha(0.08)
    c.roundRect(info_box_x, info_box_y, info_box_width, info_box_height, 8, stroke=0, fill=1)
    c.setFillAlpha(1)
    
    # Informations de l'élève (poussées vers la droite)
    x_start = 12
    y_start = height/2 + 5
    
    # Nom complet (taille encore réduite)
    c.setFillColor(colors.HexColor('#1f2937'))
    c.setFont(main_font_bold, 9)
    nom_complet = f"{eleve.prenom} {eleve.nom}".upper()
    if len(nom_complet) > 18:
        nom_complet = nom_complet[:15] + "..."
    c.drawString(x_start, y_start, nom_complet)
    
    # Ligne décorative sous le nom (ajustée à la longueur du nom)
    nom_width = c.stringWidth(nom_complet, main_font_bold, 9)
    c.setStrokeColor(colors.HexColor(primary_color))
    c.setLineWidth(2)
    c.line(x_start, y_start - 2, x_start + nom_width, y_start - 2)
    
    # Informations détaillées avec espacement
    y = y_start - 16
    c.setFillColor(colors.HexColor('#374151'))
    
    # Matricule avec espacement
    c.setFont(main_font_bold, 9)
    c.drawString(x_start, y, "N° :  ")
    c.setFont(main_font, 9)
    c.drawString(x_start + 25, y, eleve.matricule or "N/A")
    
    y -= 12
    # Classe avec espacement
    c.setFont(main_font_bold, 9)
    c.drawString(x_start, y, "Classe :  ")
    c.setFont(main_font, 9)
    c.drawString(x_start + 35, y, eleve.classe.nom)
    
    y -= 12
    # Zone avec espacement
    zone_text = abonnement.zone or "Non spécifiée"
    if len(zone_text) > 16:
        zone_text = zone_text[:13] + "..."
    c.setFont(main_font_bold, 9)
    c.drawString(x_start, y, "Zone :  ")
    c.setFont(main_font, 8)
    c.drawString(x_start + 28, y, zone_text)
    
    y -= 12
    # Point d'arrêt avec espacement
    point_arret = abonnement.point_arret or "Non spécifié"
    if len(point_arret) > 16:
        point_arret = point_arret[:13] + "..."
    c.setFont(main_font_bold, 9)
    c.drawString(x_start, y, "Arrêt :  ")
    c.setFont(main_font, 7)
    c.drawString(x_start + 28, y, point_arret)
    
    y -= 12
    # Période de validité avec espacement
    validite = f"{abonnement.date_debut.strftime('%d/%m')} - {abonnement.date_expiration.strftime('%d/%m/%Y')}"
    c.setFont(main_font_bold, 9)
    c.drawString(x_start, y, "Validité :  ")
    c.setFont(main_font, 6)
    c.drawString(x_start + 35, y, validite)
    
    # Pied de page moderne
    c.setFillColor(colors.HexColor(primary_color))
    c.setFillAlpha(0.05)
    c.roundRect(6, 4, width-12, 9, 4, stroke=0, fill=1)
    c.setFillAlpha(1)
    
    c.setFont(main_font, 6)
    c.setFillColor(colors.HexColor('#6b7280'))
    c.drawCentredString(width/2, 6, f"Généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')}")
    
    c.showPage()
    c.save()
    
    # Log de l'action
    try:
        JournalActivite.objects.create(
            user=request.user,
            action='GENERATION_PDF',
            type_objet='TICKET_BUS',
            description=f"Ticket bus généré pour {eleve.prenom} {eleve.nom} ({eleve.matricule})",
            adresse_ip=request.META.get('REMOTE_ADDR', ''),
            user_agent=request.META.get('HTTP_USER_AGENT', '')[:200]
        )
    except:
        pass
    
    return response


@login_required
def generer_tickets_retrait_classe_pdf(request, classe_id):
    """Génère tous les tickets de retrait pour une classe (primaire/maternelle) en un seul PDF"""
    classe = get_object_or_404(Classe, id=classe_id)
    
    # Vérifier les permissions
    if not user_is_admin(request.user):
        user_school_obj = user_school(request.user)
        if not user_school_obj or classe.ecole != user_school_obj:
            messages.error(request, "Vous n'avez pas accès à cette classe.")
            return redirect('eleves:liste_eleves')
    
    # Vérifier que c'est une classe du primaire/maternelle
    niveau = classe.niveau.upper() if classe.niveau else ''
    if not any(x in niveau for x in ['PRIMAIRE', 'PN', 'MATERNELLE', 'GARDERIE']):
        messages.warning(request, "Les tickets de retrait sont réservés aux classes du primaire et de la maternelle.")
        return redirect('eleves:liste_eleves')
    
    # Récupérer tous les élèves actifs de la classe
    eleves = Eleve.objects.filter(
        classe=classe,
        statut='ACTIF'
    ).select_related('classe', 'classe__ecole', 'responsable_principal').order_by('nom', 'prenom')
    
    if not eleves.exists():
        messages.warning(request, "Aucun élève actif trouvé dans cette classe.")
        return redirect('eleves:liste_eleves')
    
    # Créer le PDF avec tous les tickets (2 par page)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="tickets_retrait_{classe.nom}.pdf"'
    
    from reportlab.lib.units import mm
    from reportlab.lib.pagesizes import A4
    width_page, height_page = A4
    
    c = canvas.Canvas(response, pagesize=A4)
    
    # Enregistrer les polices
    try:
        pdfmetrics.registerFont(TTFont('Arial', 'C:/Windows/Fonts/arial.ttf'))
        pdfmetrics.registerFont(TTFont('Arial-Bold', 'C:/Windows/Fonts/arialbd.ttf'))
        main_font = 'Arial'
        main_font_bold = 'Arial-Bold'
    except:
        main_font = 'Helvetica'
        main_font_bold = 'Helvetica-Bold'
    
    # Dimensions d'un ticket
    ticket_width = 86*mm
    ticket_height = 54*mm
    margin = 10*mm
    
    # Position pour 2 tickets par page (un en haut, un en bas)
    positions = [
        (margin, height_page - margin - ticket_height),  # Haut
        (margin, height_page - margin - 2*ticket_height - 15*mm),  # Bas
    ]
    
    ticket_count = 0
    
    for eleve in eleves:
        # Calculer la position du ticket
        pos_index = ticket_count % 2
        x_offset, y_offset = positions[pos_index]
        
        # Dessiner un ticket
        _dessiner_ticket_retrait(c, eleve, x_offset, y_offset, ticket_width, ticket_height, main_font, main_font_bold)
        
        ticket_count += 1
        
        # Nouvelle page tous les 2 tickets
        if ticket_count % 2 == 0 and ticket_count < eleves.count():
            c.showPage()
    
    c.showPage()
    c.save()
    
    return response


@login_required
def generer_tickets_bus_classe_pdf(request, classe_id):
    """Génère tous les tickets bus pour une classe en un seul PDF"""
    from bus.models import AbonnementBus
    
    classe = get_object_or_404(Classe, id=classe_id)
    
    # Vérifier les permissions
    if not user_is_admin(request.user):
        user_school_obj = user_school(request.user)
        if not user_school_obj or classe.ecole != user_school_obj:
            messages.error(request, "Vous n'avez pas accès à cette classe.")
            return redirect('eleves:liste_eleves')
    
    # Récupérer les élèves avec abonnement bus actif
    eleves_ids = AbonnementBus.objects.filter(
        eleve__classe=classe,
        statut='ACTIF'
    ).values_list('eleve_id', flat=True)
    
    eleves = Eleve.objects.filter(
        id__in=eleves_ids,
        statut='ACTIF'
    ).select_related('classe', 'classe__ecole', 'responsable_principal').order_by('nom', 'prenom')
    
    if not eleves.exists():
        messages.warning(request, "Aucun élève avec abonnement bus actif trouvé dans cette classe.")
        return redirect('eleves:liste_eleves')
    
    # Créer le PDF avec tous les tickets (2 par page)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="tickets_bus_{classe.nom}.pdf"'
    
    from reportlab.lib.units import mm
    from reportlab.lib.pagesizes import A4
    width_page, height_page = A4
    
    c = canvas.Canvas(response, pagesize=A4)
    
    # Enregistrer les polices
    try:
        pdfmetrics.registerFont(TTFont('Arial', 'C:/Windows/Fonts/arial.ttf'))
        pdfmetrics.registerFont(TTFont('Arial-Bold', 'C:/Windows/Fonts/arialbd.ttf'))
        main_font = 'Arial'
        main_font_bold = 'Arial-Bold'
    except:
        main_font = 'Helvetica'
        main_font_bold = 'Helvetica-Bold'
    
    # Dimensions d'un ticket
    ticket_width = 86*mm
    ticket_height = 54*mm
    margin = 10*mm
    
    # Position pour 2 tickets par page
    positions = [
        (margin, height_page - margin - ticket_height),  # Haut
        (margin, height_page - margin - 2*ticket_height - 15*mm),  # Bas
    ]
    
    ticket_count = 0
    
    for eleve in eleves:
        # Récupérer l'abonnement
        abonnement = AbonnementBus.objects.filter(
            eleve=eleve,
            statut='ACTIF'
        ).order_by('-date_debut').first()
        
        if not abonnement:
            continue
        
        # Calculer la position du ticket
        pos_index = ticket_count % 2
        x_offset, y_offset = positions[pos_index]
        
        # Dessiner un ticket
        _dessiner_ticket_bus(c, eleve, abonnement, x_offset, y_offset, ticket_width, ticket_height, main_font, main_font_bold)
        
        ticket_count += 1
        
        # Nouvelle page tous les 2 tickets
        if ticket_count % 2 == 0 and ticket_count < eleves.count():
            c.showPage()
    
    c.showPage()
    c.save()
    
    return response


def _extraire_couleurs_logo(logo_path):
    """Extrait les couleurs dominantes du logo de l'école"""
    try:
        from PIL import Image
        import colorsys
        
        img = Image.open(logo_path)
        img = img.convert('RGB')
        img.thumbnail((100, 100))
        
        # Obtenir les pixels
        pixels = list(img.getdata())
        
        # Filtrer les pixels trop clairs ou trop sombres
        filtered_pixels = []
        for r, g, b in pixels:
            brightness = (r + g + b) / 3
            if 30 < brightness < 220:  # Éviter blanc et noir
                filtered_pixels.append((r, g, b))
        
        if not filtered_pixels:
            filtered_pixels = pixels
        
        # Compter les couleurs
        from collections import Counter
        color_counts = Counter(filtered_pixels)
        most_common = color_counts.most_common(5)
        
        # Trouver la couleur la plus saturée (plus vive)
        best_color = None
        best_saturation = 0
        
        for color, count in most_common:
            r, g, b = color
            h, s, v = colorsys.rgb_to_hsv(r/255, g/255, b/255)
            if s > best_saturation and v > 0.3:  # Saturation et luminosité minimales
                best_saturation = s
                best_color = color
        
        if best_color:
            r, g, b = best_color
            # Créer une version plus claire pour l'arrière-plan
            light_r = min(255, r + 80)
            light_g = min(255, g + 80)
            light_b = min(255, b + 80)
            
            primary = f'#{r:02x}{g:02x}{b:02x}'
            light = f'#{light_r:02x}{light_g:02x}{light_b:02x}'
            
            return primary, light
    except:
        pass
    
    # Couleurs par défaut si extraction échoue
    return '#6366f1', '#e0e7ff'


def _ticket_safe_text(value, default='-'):
    if value is None:
        return default
    value = str(value).strip()
    return value if value else default


def _ticket_fit_text(c, text, x, y, max_width, font_name, max_size, min_size=5, color=None):
    text = _ticket_safe_text(text)
    size = max_size
    while size > min_size and pdfmetrics.stringWidth(text, font_name, size) > max_width:
        size -= 0.5

    if pdfmetrics.stringWidth(text, font_name, size) > max_width:
        while text and pdfmetrics.stringWidth(text + '...', font_name, size) > max_width:
            text = text[:-1]
        text = text + '...' if text else '...'

    if color:
        c.setFillColor(colors.HexColor(color))
    c.setFont(font_name, size)
    c.drawString(x, y, text)


def _ticket_draw_photo(c, eleve, x, y, size, accent_color, main_font_bold):
    c.setFillColor(colors.white)
    c.roundRect(x, y, size, size, 6, stroke=0, fill=1)
    c.setStrokeColor(colors.HexColor(accent_color))
    c.setLineWidth(1.5)
    c.roundRect(x, y, size, size, 6, stroke=1, fill=0)

    try:
        if eleve.photo and hasattr(eleve.photo, 'path') and os.path.exists(eleve.photo.path):
            c.drawImage(eleve.photo.path, x + 2, y + 2, size - 4, size - 4, preserveAspectRatio=True, anchor='c', mask='auto')
            return
    except Exception:
        pass

    c.setFillColor(colors.HexColor('#f1f5f9'))
    c.roundRect(x + 2, y + 2, size - 4, size - 4, 5, stroke=0, fill=1)
    c.setFillColor(colors.HexColor(accent_color))
    c.setFont(main_font_bold, 6)
    c.drawCentredString(x + size / 2, y + size / 2 - 2, 'PHOTO')


def _ticket_draw_logo(c, ecole, x, y, size):
    try:
        if ecole.logo and hasattr(ecole.logo, 'path') and os.path.exists(ecole.logo.path):
            c.setFillColor(colors.white)
            c.roundRect(x, y, size, size, 5, stroke=0, fill=1)
            c.drawImage(ecole.logo.path, x + 2, y + 2, size - 4, size - 4, preserveAspectRatio=True, anchor='c', mask='auto')
            return True
    except Exception:
        pass
    return False


def _ticket_draw_row(c, label, value, x, y, value_width, accent_color, main_font, main_font_bold):
    c.setFillColor(colors.HexColor('#64748b'))
    c.setFont(main_font_bold, 5.8)
    c.drawString(x, y + 5, label.upper())
    _ticket_fit_text(c, value, x, y - 3, value_width, main_font_bold, 8.2, 5.2, '#0f172a')


def _dessiner_ticket_carte(c, eleve, x, y, width, height, main_font, main_font_bold, title, accent_color, light_color, rows, serial_label):
    c.saveState()
    from reportlab.lib.units import mm

    ecole = eleve.classe.ecole
    primary = '#1746a2'
    accent = accent_color or '#0f766e'
    dark = '#0f172a'
    muted = '#64748b'
    line = '#dbe3ef'
    soft = '#f5f8fc'
    footer_soft = '#eef4fb'

    margin = 2.2 * mm
    header_h = 10.5 * mm
    footer_h = 6.2 * mm
    radius = 4.5

    school_name = _ticket_safe_text(getattr(ecole, 'nom', '')).upper()
    full_name = f"{_ticket_safe_text(eleve.prenom, '')} {_ticket_safe_text(eleve.nom, '')}".strip().upper()

    c.setFillColor(colors.white)
    c.setStrokeColor(colors.HexColor(line))
    c.setLineWidth(0.7)
    c.roundRect(x, y, width, height, radius, stroke=1, fill=1)

    c.setFillColor(colors.HexColor(primary))
    c.roundRect(x + 0.8, y + height - header_h - 0.8, width - 1.6, header_h, radius, stroke=0, fill=1)
    c.rect(x + 0.8, y + height - header_h - 0.8, width - 1.6, header_h / 2, stroke=0, fill=1)

    logo_size = 7.2 * mm
    logo_x = x + margin
    logo_y = y + height - header_h + 1.1 * mm
    c.setFillColor(colors.white)
    c.circle(logo_x + logo_size / 2, logo_y + logo_size / 2, logo_size / 2, stroke=0, fill=1)
    try:
        if ecole.logo and hasattr(ecole.logo, 'path') and os.path.exists(ecole.logo.path):
            c.drawImage(
                ecole.logo.path,
                logo_x + 0.6,
                logo_y + 0.6,
                width=logo_size - 1.2,
                height=logo_size - 1.2,
                preserveAspectRatio=True,
                mask='auto',
            )
        else:
            raise ValueError('No logo')
    except Exception:
        c.setFillColor(colors.HexColor(primary))
        c.setFont(main_font_bold, 7)
        c.drawCentredString(logo_x + logo_size / 2, logo_y + logo_size / 2 - 2, school_name[:2] or 'EC')

    title_x = logo_x + logo_size + 2 * mm
    title_w = width - (title_x - x) - margin
    _ticket_fit_text(c, school_name, title_x, y + height - 5.1 * mm, title_w, main_font_bold, 7.6, 4.8, '#ffffff')
    _ticket_fit_text(c, title, title_x, y + height - 8.2 * mm, title_w, main_font, 5.2, 4.2, '#dbeafe')

    try:
        c.saveState()
        c.setFillAlpha(0.06)
        if ecole.logo and hasattr(ecole.logo, 'path') and os.path.exists(ecole.logo.path):
            mark = 30 * mm
            c.drawImage(
                ecole.logo.path,
                x + width - mark - 4 * mm,
                y + footer_h + 5 * mm,
                width=mark,
                height=mark,
                preserveAspectRatio=True,
                mask='auto',
            )
        else:
            c.setFillColor(colors.HexColor(primary))
            c.setFont(main_font_bold, 28)
            c.drawCentredString(x + width * 0.70, y + height * 0.46, school_name[:3])
        c.restoreState()
    except Exception:
        try:
            c.restoreState()
        except Exception:
            pass

    photo_w = 22.5 * mm
    photo_h = 27.5 * mm
    photo_x = x + margin
    photo_y = y + footer_h + 3.2 * mm
    c.setFillColor(colors.HexColor(soft))
    c.setStrokeColor(colors.HexColor(line))
    c.setLineWidth(0.7)
    c.roundRect(photo_x, photo_y, photo_w, photo_h, 3.2, stroke=1, fill=1)

    photo_drawn = False
    try:
        if eleve.photo and hasattr(eleve.photo, 'path') and eleve.photo.name and os.path.exists(eleve.photo.path):
            c.drawImage(eleve.photo.path, photo_x + 1, photo_y + 1, photo_w - 2, photo_h - 2, preserveAspectRatio=True, anchor='c', mask='auto')
            photo_drawn = True
    except Exception:
        photo_drawn = False

    if not photo_drawn:
        initials = (_ticket_safe_text(getattr(eleve, 'prenom', 'E'), 'E')[:1] + _ticket_safe_text(getattr(eleve, 'nom', 'L'), 'L')[:1]).upper()
        c.setFillColor(colors.HexColor('#e8eef8'))
        c.roundRect(photo_x + 1, photo_y + 1, photo_w - 2, photo_h - 2, 2.5, stroke=0, fill=1)
        c.setFillColor(colors.HexColor(primary))
        c.setFont(main_font_bold, 17)
        c.drawCentredString(photo_x + photo_w / 2, photo_y + photo_h / 2 - 4, initials or 'EL')

    info_x = photo_x + photo_w + 3 * mm
    info_w = x + width - margin - info_x
    name_y = y + height - header_h - 4.4 * mm
    _ticket_fit_text(c, full_name, info_x, name_y, info_w, main_font_bold, 8.7, 5.8, dark)

    c.setStrokeColor(colors.HexColor(accent))
    c.setLineWidth(1.0)
    c.line(info_x, name_y - 1.7 * mm, info_x + info_w, name_y - 1.7 * mm)

    row_y = name_y - 5.3 * mm
    label_w = 14 * mm
    value_w = info_w - label_w
    for label, value in rows[:5]:
        c.setFillColor(colors.HexColor(muted))
        c.setFont(main_font_bold, 5.6)
        c.drawString(info_x, row_y, _ticket_safe_text(label).upper())
        _ticket_fit_text(c, value, info_x + label_w, row_y, value_w, main_font, 6.6, 4.7, dark)
        row_y -= 4.2 * mm

    c.setFillColor(colors.HexColor(footer_soft))
    c.rect(x + 0.8, y + 0.8, width - 1.6, footer_h, stroke=0, fill=1)
    c.setStrokeColor(colors.HexColor(line))
    c.setLineWidth(0.5)
    c.line(x + 0.8, y + footer_h + 0.8, x + width - 0.8, y + footer_h + 0.8)

    annee = _ticket_safe_text(getattr(eleve.classe, 'annee_scolaire', ''))
    _ticket_fit_text(c, f'ANNEE SCOLAIRE {annee}', x + margin, y + 2.4 * mm, width * 0.55, main_font_bold, 5.8, 4.2, primary)
    c.setFillColor(colors.HexColor(muted))
    c.setFont(main_font, 4.6)
    c.drawRightString(x + width - margin, y + 2.4 * mm, f'{serial_label} #{getattr(eleve, "id", 0):06d}')

    c.setStrokeColor(colors.HexColor(primary))
    c.setLineWidth(0.9)
    c.roundRect(x, y, width, height, radius, stroke=1, fill=0)
    c.restoreState()


def _dessiner_ticket_retrait(c, eleve, x, y, width, height, main_font, main_font_bold):
    """Fonction helper pour dessiner un ticket de retrait avec design moderne"""
    responsable = getattr(eleve, 'responsable_principal', None)
    responsable_nom = 'Non renseigne'
    responsable_tel = '-'
    if responsable:
        responsable_nom = _ticket_safe_text(
            getattr(responsable, 'nom_complet', None) or f"{getattr(responsable, 'prenom', '')} {getattr(responsable, 'nom', '')}",
            'Non renseigne'
        )
        responsable_tel = _ticket_safe_text(getattr(responsable, 'telephone', None))

    rows = [
        ('Matricule', getattr(eleve, 'matricule', None)),
        ('Classe', getattr(eleve.classe, 'nom', None)),
        ('Parent', responsable_nom),
        ('Telephone', responsable_tel),
    ]
    return _dessiner_ticket_carte(
        c, eleve, x, y, width, height, main_font, main_font_bold,
        'CARTE DE RETRAIT', '#0f766e', '#dbeafe', rows, 'RETRAIT'
    )

    c.saveState()
    
    # Extraire les couleurs du logo
    primary_color = '#6366f1'
    light_color = '#e0e7ff'
    
    try:
        if eleve.classe.ecole.logo and hasattr(eleve.classe.ecole.logo, 'path'):
            logo_path = eleve.classe.ecole.logo.path
            if os.path.exists(logo_path):
                primary_color, light_color = _extraire_couleurs_logo(logo_path)
    except:
        pass
    
    # Fond avec dégradé simulé
    c.setFillColor(colors.white)
    c.rect(x, y, width, height, stroke=0, fill=1)
    
    # Forme géométrique décorative en arrière-plan (cercles)
    c.setFillColor(colors.HexColor(light_color))
    c.setFillAlpha(0.15)
    c.circle(x + width - 20, y + height - 20, 40, stroke=0, fill=1)
    c.circle(x + 15, y + 15, 30, stroke=0, fill=1)
    c.setFillAlpha(1)
    
    # Bordure avec coins arrondis et ombre
    c.setStrokeColor(colors.HexColor(primary_color))
    c.setLineWidth(2.5)
    c.roundRect(x+2, y+2, width-4, height-4, 10, stroke=1, fill=0)
    
    # Bande diagonale décorative en haut à gauche
    c.saveState()
    c.translate(x, y + height)
    c.rotate(-25)
    c.setFillColor(colors.HexColor(primary_color))
    c.setFillAlpha(0.1)
    c.rect(-10, -15, 60, 25, stroke=0, fill=1)
    c.setFillAlpha(1)
    c.restoreState()
    
    # En-tête moderne avec forme ondulée
    c.setFillColor(colors.HexColor(primary_color))
    path = c.beginPath()
    path.moveTo(x+5, y+height-5)
    path.lineTo(x+width-5, y+height-5)
    path.lineTo(x+width-5, y+height-42)
    # Courbe ondulée
    path.curveTo(x+width*0.75, y+height-38, x+width*0.5, y+height-46, x+width*0.25, y+height-42)
    path.curveTo(x+width*0.15, y+height-40, x+5, y+height-38, x+5, y+height-42)
    path.close()
    c.drawPath(path, fill=1, stroke=0)
    
    # Titre avec effet
    c.setFillColor(colors.white)
    c.setFont(main_font_bold, 14)
    c.drawCentredString(x+width/2, y+height-20, "TICKET DE RETRAIT")
    
    # Sous-titre école
    c.setFont(main_font, 7)
    c.setFillAlpha(0.9)
    c.drawCentredString(x+width/2, y+height-32, eleve.classe.ecole.nom[:50])
    c.setFillAlpha(1)
    
    # Photo avec cadre moderne hexagonal simulé
    photo_x = x + width - 28
    photo_y = y + 26
    photo_radius = 18
    
    # Ombre de la photo
    c.setFillColor(colors.HexColor('#000000'))
    c.setFillAlpha(0.1)
    c.circle(photo_x + 1, photo_y - 1, photo_radius + 2, stroke=0, fill=1)
    c.setFillAlpha(1)
    
    # Cercle de la photo avec double bordure
    c.setFillColor(colors.white)
    c.circle(photo_x, photo_y, photo_radius + 2, stroke=0, fill=1)
    c.setStrokeColor(colors.HexColor(primary_color))
    c.setLineWidth(3)
    c.circle(photo_x, photo_y, photo_radius, stroke=1, fill=1)
    
    # Placeholder photo
    c.setFillColor(colors.HexColor(light_color))
    c.circle(photo_x, photo_y, photo_radius - 2, stroke=0, fill=1)
    c.setFillColor(colors.HexColor(primary_color))
    c.setFont(main_font, 7)
    c.drawCentredString(photo_x, photo_y - 2, "PHOTO")
    
    # Carte d'information avec fond coloré
    info_box_x = x + 8
    info_box_y = y + 8
    info_box_width = width - 50
    info_box_height = height - 52
    
    c.setFillColor(colors.HexColor(light_color))
    c.setFillAlpha(0.08)
    c.roundRect(info_box_x, info_box_y, info_box_width, info_box_height, 8, stroke=0, fill=1)
    c.setFillAlpha(1)
    
    # Nom de l'élève avec style moderne
    x_start = x + 12
    y_start = y + height - 50
    
    c.setFillColor(colors.HexColor('#1f2937'))
    c.setFont(main_font_bold, 12)
    nom_complet = f"{eleve.prenom} {eleve.nom}".upper()
    if len(nom_complet) > 24:
        nom_complet = nom_complet[:21] + "..."
    c.drawString(x_start, y_start, nom_complet)
    
    # Ligne décorative sous le nom
    c.setStrokeColor(colors.HexColor(primary_color))
    c.setLineWidth(2)
    c.line(x_start, y_start - 2, x_start + 55, y_start - 2)
    
    # Informations avec icônes et style moderne
    y_info = y_start - 16
    c.setFillColor(colors.HexColor('#374151'))
    
    # Matricule
    c.setFont(main_font_bold, 9)
    c.drawString(x_start, y_info, "N°")
    c.setFont(main_font, 9)
    c.drawString(x_start + 12, y_info, eleve.matricule or "N/A")
    
    y_info -= 12
    # Classe
    c.setFont(main_font_bold, 9)
    c.drawString(x_start, y_info, "Classe")
    c.setFont(main_font, 9)
    c.drawString(x_start + 25, y_info, eleve.classe.nom)
    
    y_info -= 12
    # Responsable
    if eleve.responsable_principal:
        c.setFont(main_font_bold, 9)
        c.drawString(x_start, y_info, "Parent")
        resp_nom = f"{eleve.responsable_principal.prenom} {eleve.responsable_principal.nom}"
        if len(resp_nom) > 24:
            resp_nom = resp_nom[:21] + "..."
        c.setFont(main_font, 8)
        c.drawString(x_start + 25, y_info, resp_nom)
    
    # Pied de page moderne
    c.setFillColor(colors.HexColor(primary_color))
    c.setFillAlpha(0.05)
    c.roundRect(x+6, y+4, width-12, 9, 4, stroke=0, fill=1)
    c.setFillAlpha(1)
    
    c.setFont(main_font, 6)
    c.setFillColor(colors.HexColor('#6b7280'))
    c.drawCentredString(x+width/2, y+6, f"Généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')}")
    
    c.restoreState()


def _dessiner_ticket_bus(c, eleve, abonnement, x, y, width, height, main_font, main_font_bold):
    """Fonction helper pour dessiner un ticket bus avec design moderne"""
    validite = '-'
    if getattr(abonnement, 'date_debut', None) and getattr(abonnement, 'date_expiration', None):
        validite = f"{abonnement.date_debut.strftime('%d/%m')} - {abonnement.date_expiration.strftime('%d/%m/%Y')}"

    rows = [
        ('Matricule', getattr(eleve, 'matricule', None)),
        ('Classe', getattr(eleve.classe, 'nom', None)),
        ('Zone', getattr(abonnement, 'zone', None) or 'Non specifiee'),
        ('Arret', getattr(abonnement, 'point_arret', None) or 'Non specifie'),
        ('Validite', validite),
    ]
    return _dessiner_ticket_carte(
        c, eleve, x, y, width, height, main_font, main_font_bold,
        'CARTE BUS', '#0f766e', '#fef3c7', rows, 'BUS'
    )

    c.saveState()
    
    # Extraire les couleurs du logo (version orange pour bus)
    primary_color = '#f59e0b'
    light_color = '#fef3c7'
    
    try:
        if eleve.classe.ecole.logo and hasattr(eleve.classe.ecole.logo, 'path'):
            logo_path = eleve.classe.ecole.logo.path
            if os.path.exists(logo_path):
                extracted_primary, extracted_light = _extraire_couleurs_logo(logo_path)
                # Garder une teinte orange/jaune pour le bus
                primary_color = extracted_primary
                light_color = extracted_light
    except:
        pass
    
    # Fond avec motif géométrique
    c.setFillColor(colors.white)
    c.rect(x, y, width, height, stroke=0, fill=1)
    
    # Motif de lignes diagonales en arrière-plan
    c.setStrokeColor(colors.HexColor(light_color))
    c.setLineWidth(15)
    c.setStrokeAlpha(0.1)
    for i in range(0, int(width + height), 25):
        c.line(x + i, y, x + i - height, y + height)
    c.setStrokeAlpha(1)
    
    # Bordure moderne avec double ligne
    c.setStrokeColor(colors.HexColor(primary_color))
    c.setLineWidth(2.5)
    c.roundRect(x+2, y+2, width-4, height-4, 10, stroke=1, fill=0)
    
    # Accent coins
    corner_size = 8
    c.setFillColor(colors.HexColor(primary_color))
    # Coin haut gauche
    c.rect(x+2, y+height-corner_size-2, corner_size, corner_size, stroke=0, fill=1)
    # Coin bas droit
    c.rect(x+width-corner_size-2, y+2, corner_size, corner_size, stroke=0, fill=1)
    
    # En-tête avec forme moderne
    c.setFillColor(colors.HexColor(primary_color))
    path = c.beginPath()
    path.moveTo(x+5, y+height-5)
    path.lineTo(x+width-5, y+height-5)
    path.lineTo(x+width-5, y+height-44)
    # Forme en vague
    path.curveTo(x+width*0.8, y+height-40, x+width*0.6, y+height-48, x+width*0.4, y+height-44)
    path.curveTo(x+width*0.2, y+height-40, x+5, y+height-42, x+5, y+height-44)
    path.close()
    c.drawPath(path, fill=1, stroke=0)
    
    # Icône bus stylisée
    c.setFillColor(colors.white)
    c.setFillAlpha(0.3)
    c.roundRect(x+width-45, y+height-32, 18, 22, 3, stroke=0, fill=1)
    c.setFillAlpha(1)
    
    # Titre
    c.setFillColor(colors.white)
    c.setFont(main_font_bold, 13)
    c.drawCentredString(x+width/2, y+height-20, "ABONNEMENT BUS")
    
    # Sous-titre
    c.setFont(main_font, 7)
    c.setFillAlpha(0.9)
    c.drawCentredString(x+width/2, y+height-33, eleve.classe.ecole.nom[:50])
    c.setFillAlpha(1)
    
    # Photo avec style moderne
    photo_x = x + width - 28
    photo_y = y + 26
    photo_radius = 18
    
    # Ombre
    c.setFillColor(colors.HexColor('#000000'))
    c.setFillAlpha(0.15)
    c.circle(photo_x + 1, photo_y - 1, photo_radius + 2, stroke=0, fill=1)
    c.setFillAlpha(1)
    
    # Photo avec double bordure
    c.setFillColor(colors.white)
    c.circle(photo_x, photo_y, photo_radius + 2, stroke=0, fill=1)
    c.setStrokeColor(colors.HexColor(primary_color))
    c.setLineWidth(3)
    c.circle(photo_x, photo_y, photo_radius, stroke=1, fill=1)
    
    # Placeholder
    c.setFillColor(colors.HexColor(light_color))
    c.circle(photo_x, photo_y, photo_radius - 2, stroke=0, fill=1)
    c.setFillColor(colors.HexColor(primary_color))
    c.setFont(main_font, 7)
    c.drawCentredString(photo_x, photo_y - 2, "PHOTO")
    
    # Zone d'information avec fond subtil
    info_box_x = x + 8
    info_box_y = y + 8
    info_box_width = width - 50
    info_box_height = height - 52
    
    c.setFillColor(colors.HexColor(light_color))
    c.setFillAlpha(0.1)
    c.roundRect(info_box_x, info_box_y, info_box_width, info_box_height, 8, stroke=0, fill=1)
    c.setFillAlpha(1)
    
    # Nom élève
    x_start = x + 12
    y_start = y + height - 52
    
    c.setFillColor(colors.HexColor('#1f2937'))
    c.setFont(main_font_bold, 11)
    nom_complet = f"{eleve.prenom} {eleve.nom}".upper()
    if len(nom_complet) > 22:
        nom_complet = nom_complet[:19] + "..."
    c.drawString(x_start, y_start, nom_complet)
    
    # Ligne décorative
    c.setStrokeColor(colors.HexColor(primary_color))
    c.setLineWidth(2)
    c.line(x_start, y_start - 2, x_start + 50, y_start - 2)
    
    # Informations détaillées
    y_info = y_start - 15
    c.setFillColor(colors.HexColor('#374151'))
    
    # Matricule
    c.setFont(main_font_bold, 8)
    c.drawString(x_start, y_info, "N°")
    c.setFont(main_font, 8)
    c.drawString(x_start + 10, y_info, eleve.matricule or "N/A")
    
    y_info -= 10
    # Classe
    c.setFont(main_font_bold, 8)
    c.drawString(x_start, y_info, "Classe")
    c.setFont(main_font, 8)
    c.drawString(x_start + 22, y_info, eleve.classe.nom)
    
    y_info -= 10
    # Zone
    zone_text = abonnement.zone or "Non spécifiée"
    if len(zone_text) > 22:
        zone_text = zone_text[:19] + "..."
    c.setFont(main_font_bold, 8)
    c.drawString(x_start, y_info, "Zone")
    c.setFont(main_font, 8)
    c.drawString(x_start + 18, y_info, zone_text)
    
    y_info -= 10
    # Arrêt
    point_arret = abonnement.point_arret or "Non spécifié"
    if len(point_arret) > 22:
        point_arret = point_arret[:19] + "..."
    c.setFont(main_font_bold, 8)
    c.drawString(x_start, y_info, "Arrêt")
    c.setFont(main_font, 7)
    c.drawString(x_start + 18, y_info, point_arret)
    
    y_info -= 10
    # Validité
    validite = f"{abonnement.date_debut.strftime('%d/%m')} - {abonnement.date_expiration.strftime('%d/%m/%Y')}"
    c.setFont(main_font_bold, 8)
    c.drawString(x_start, y_info, "Validité")
    c.setFont(main_font, 6)
    c.drawString(x_start + 25, y_info, validite)
    
    # Pied de page moderne
    c.setFillColor(colors.HexColor(primary_color))
    c.setFillAlpha(0.05)
    c.roundRect(x+6, y+4, width-12, 9, 4, stroke=0, fill=1)
    c.setFillAlpha(1)
    
    c.setFont(main_font, 6)
    c.setFillColor(colors.HexColor('#6b7280'))
    c.drawCentredString(x+width/2, y+6, f"Généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')}")
    
    c.restoreState()


@login_required
def carte_scolaire_preview(request, eleve_id):
    """Affiche un aperçu HTML de la carte scolaire"""
    eleve = get_object_or_404(
        Eleve.objects.select_related('classe', 'classe__ecole', 'responsable_principal'),
        id=eleve_id
    )
    
    # Vérifier permissions
    if not user_is_admin(request.user):
        user_school_obj = user_school(request.user)
        if not user_school_obj or eleve.classe.ecole != user_school_obj:
            messages.error(request, "Vous n'avez pas accès à cet élève.")
            return redirect('eleves:liste_eleves')
    
    context = {
        'eleve': eleve,
        'titre_page': f'Carte Scolaire - {eleve.nom_complet}'
    }
    return render(request, 'eleves/carte_scolaire_preview.html', context)


@login_required
def generer_carte_scolaire_pdf(request, eleve_id):
    """Génère une carte scolaire moderne pour un élève"""
    from .carte_scolaire_generator import generer_carte_scolaire_moderne, generer_carte_pvc_haute_qualite
    
    eleve = get_object_or_404(
        Eleve.objects.select_related('classe', 'classe__ecole', 'responsable_principal'),
        id=eleve_id
    )
    
    # Vérifier permissions
    if not user_is_admin(request.user):
        user_school_obj = user_school(request.user)
        if not user_school_obj or eleve.classe.ecole != user_school_obj:
            messages.error(request, "Vous n'avez pas accès à cet élève.")
            return redirect('eleves:liste_eleves')
    
    # Par défaut, utiliser le format PVC pour les cartes individuelles
    # Le format standard n'est utilisé que si explicitement demandé
    format_standard = request.GET.get('format') == 'standard'
    
    # Créer le PDF avec le nouveau design
    response = HttpResponse(content_type='application/pdf')
    
    if format_standard:
        # Format standard (si explicitement demandé)
        response['Content-Disposition'] = f'attachment; filename="carte_scolaire_{eleve.matricule}.pdf"'
        return generer_carte_scolaire_moderne(eleve, response)
    else:
        # Par défaut : format PVC optimisé pour impression directe
        response['Content-Disposition'] = f'attachment; filename="carte_pvc_{eleve.matricule}.pdf"'
        # Utiliser le générateur moderne qui est déjà au format carte bancaire (86mm x 54mm)
        return generer_carte_scolaire_moderne(eleve, response)
    
    # Polices
    try:
        pdfmetrics.registerFont(TTFont('Arial', 'C:/Windows/Fonts/arial.ttf'))
        pdfmetrics.registerFont(TTFont('Arial-Bold', 'C:/Windows/Fonts/arialbd.ttf'))
        main_font = 'Arial'
        main_font_bold = 'Arial-Bold'
    except:
        main_font = 'Helvetica'
        main_font_bold = 'Helvetica-Bold'
    
    primary_color = '#2563eb'
    light_color = '#dbeafe'
    
    # Fond et bordure
    c.setFillColor(colors.white)
    c.rect(0, 0, width, height, stroke=0, fill=1)
    c.setStrokeColor(colors.HexColor(primary_color))
    c.setLineWidth(2.5)
    c.roundRect(2, 2, width-4, height-4, 10, stroke=1, fill=0)
    
    # Bande supérieure (plus grande)
    c.setFillColor(colors.HexColor(primary_color))
    c.roundRect(5, height-42, width-10, 37, 8, stroke=0, fill=1)
    
    # Logo école (plus grand)
    try:
        if eleve.classe.ecole.logo and hasattr(eleve.classe.ecole.logo, 'path'):
            if os.path.exists(eleve.classe.ecole.logo.path):
                logo_size = 30
                c.setFillColor(colors.white)
                c.circle(10 + logo_size/2, height - 35 + logo_size/2, logo_size/2 + 1, stroke=0, fill=1)
                c.drawImage(eleve.classe.ecole.logo.path, 10, height - 35, 
                          width=logo_size, height=logo_size, preserveAspectRatio=True, mask='auto')
    except:
        pass
    
    # Nom école (texte plus grand)
    c.setFillColor(colors.white)
    c.setFont(main_font_bold, 11)
    c.drawString(45, height-15, eleve.classe.ecole.nom[:40])
    c.setFont(main_font, 9)
    c.drawString(45, height-30, f"Année Scolaire: {eleve.classe.annee_scolaire}")
    
    # Photo élève (plus grande)
    photo_x = width - 48
    photo_y = height/2 + 3
    photo_size = 42
    
    c.setFillColor(colors.white)
    c.roundRect(photo_x, photo_y, photo_size, photo_size, 6, stroke=0, fill=1)
    c.setStrokeColor(colors.HexColor(primary_color))
    c.setLineWidth(2.5)
    c.roundRect(photo_x, photo_y, photo_size, photo_size, 6, stroke=1, fill=0)
    
    if eleve.photo:
        try:
            from PIL import Image
            if hasattr(eleve.photo, 'path') and os.path.exists(eleve.photo.path):
                img = Image.open(eleve.photo.path)
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                img.thumbnail((photo_size - 5, photo_size - 5), Image.Resampling.LANCZOS)
                temp_buffer = io.BytesIO()
                img.save(temp_buffer, format='JPEG')
                temp_buffer.seek(0)
                c.drawImage(temp_buffer, photo_x + 2.5, photo_y + 2.5, 
                          width=photo_size - 5, height=photo_size - 5, preserveAspectRatio=True)
        except:
            c.setFillColor(colors.HexColor(primary_color))
            c.setFont(main_font_bold, 18)
            c.drawCentredString(photo_x + photo_size/2, photo_y + photo_size/2 - 5, 
                              f"{eleve.prenom[0]}{eleve.nom[0]}")
    else:
        c.setFillColor(colors.HexColor(primary_color))
        c.setFont(main_font_bold, 18)
        c.drawCentredString(photo_x + photo_size/2, photo_y + photo_size/2 - 5, 
                          f"{eleve.prenom[0]}{eleve.nom[0]}")
    
    # Section informations élève (plus d'espace, plus d'infos)
    y_pos = height - 58
    x_margin = 10
    
    # Nom et prénom (plus grand)
    c.setFillColor(colors.HexColor('#1f2937'))
    c.setFont(main_font_bold, 13)
    nom_complet = f"{eleve.prenom} {eleve.nom}".upper()
    if len(nom_complet) > 25:
        nom_complet = nom_complet[:25] + "."
    c.drawString(x_margin, y_pos, nom_complet)
    
    # Ligne de séparation
    y_pos -= 8
    c.setStrokeColor(colors.HexColor(light_color))
    c.setLineWidth(1)
    c.line(x_margin, y_pos, width - photo_size - 15, y_pos)
    
    # Informations principales (2 colonnes)
    y_pos -= 10
    c.setFont(main_font_bold, 9)
    c.setFillColor(colors.HexColor('#374151'))
    
    # Colonne 1
    c.drawString(x_margin, y_pos, "Matricule:")
    c.setFont(main_font, 9)
    c.setFillColor(colors.HexColor('#6b7280'))
    c.drawString(x_margin + 32, y_pos, eleve.matricule)
    
    # Colonne 2
    c.setFont(main_font_bold, 9)
    c.setFillColor(colors.HexColor('#374151'))
    sexe_display = "Masculin" if eleve.sexe == 'M' else "Féminin"
    c.drawString(x_margin, y_pos - 9, "Sexe:")
    c.setFont(main_font, 9)
    c.setFillColor(colors.HexColor('#6b7280'))
    c.drawString(x_margin + 32, y_pos - 9, sexe_display)
    
    # Ligne 2
    c.setFont(main_font_bold, 9)
    c.setFillColor(colors.HexColor('#374151'))
    c.drawString(x_margin, y_pos - 18, "Classe:")
    c.setFont(main_font, 9)
    c.setFillColor(colors.HexColor('#6b7280'))
    c.drawString(x_margin + 32, y_pos - 18, eleve.classe.nom)
    
    c.setFont(main_font_bold, 9)
    c.setFillColor(colors.HexColor('#374151'))
    c.drawString(x_margin, y_pos - 27, "Né(e) le:")
    c.setFont(main_font, 9)
    c.setFillColor(colors.HexColor('#6b7280'))
    c.drawString(x_margin + 32, y_pos - 27, eleve.date_naissance.strftime('%d/%m/%Y'))
    
    # Contact d'urgence
    if eleve.responsable_principal:
        c.setFont(main_font_bold, 8)
        c.setFillColor(colors.HexColor('#374151'))
        c.drawString(x_margin, y_pos - 37, "Contact urgence:")
        c.setFont(main_font, 8)
        c.setFillColor(colors.HexColor('#6b7280'))
        tel = eleve.responsable_principal.telephone or "Non renseigné"
        c.drawString(x_margin + 50, y_pos - 37, tel[:20])
    
    # Pied de page avec adresse école
    c.setFont(main_font, 7)
    c.setFillColor(colors.HexColor('#9ca3af'))
    if eleve.classe.ecole.adresse:
        adresse_courte = eleve.classe.ecole.adresse[:45]
        c.drawString(x_margin, 10, adresse_courte)
    
    if eleve.classe.ecole.telephone:
        c.drawString(x_margin, 4, f"Tél: {eleve.classe.ecole.tous_telephones}")
    
    c.setFont(main_font, 6)
    c.drawRightString(width - 5, 6, f"Généré le {timezone.now().strftime('%d/%m/%Y')}")
    
    c.showPage()
    c.save()
    return response


@login_required
def generer_cartes_classe_pdf(request, classe_id):
    """Génère toutes les cartes d'une classe (4 cartes par page A4)"""
    from .carte_scolaire_generator import generer_cartes_classe_moderne
    
    classe = get_object_or_404(Classe, id=classe_id)
    
    if not user_is_admin(request.user):
        user_school_obj = user_school(request.user)
        if not user_school_obj or classe.ecole != user_school_obj:
            messages.error(request, "Vous n'avez pas accès à cette classe.")
            return redirect('eleves:liste_eleves')
    
    eleves = Eleve.objects.filter(classe=classe, statut='ACTIF').select_related(
        'classe', 'classe__ecole', 'responsable_principal').order_by('nom', 'prenom')
    
    if not eleves.exists():
        messages.warning(request, "Aucun élève actif dans cette classe.")
        return redirect('eleves:liste_eleves')
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="cartes_scolaires_{classe.nom}.pdf"'
    
    # Utiliser le nouveau générateur
    return generer_cartes_classe_moderne(classe, eleves, response)
    
    from reportlab.lib.units import mm
    from reportlab.lib.pagesizes import A4
    
    # Nouvelle taille : 105mm x 74mm
    card_width, card_height = 105*mm, 74*mm
    margin = 10*mm
    spacing = 8*mm
    
    c = canvas.Canvas(response, pagesize=A4)
    page_width, page_height = A4
    
    try:
        pdfmetrics.registerFont(TTFont('Arial', 'C:/Windows/Fonts/arial.ttf'))
        pdfmetrics.registerFont(TTFont('Arial-Bold', 'C:/Windows/Fonts/arialbd.ttf'))
        main_font = 'Arial'
        main_font_bold = 'Arial-Bold'
    except:
        main_font = 'Helvetica'
        main_font_bold = 'Helvetica-Bold'
    
    primary_color = '#2563eb'
    positions = [
        (margin, page_height - margin - card_height),
        (margin, page_height - margin - (2 * card_height) - spacing),
    ]
    
    card_count = 0
    
    for eleve in eleves:
        pos_index = card_count % 2
        x, y = positions[pos_index]
        
        # Bordure
        c.setStrokeColor(colors.HexColor(primary_color))
        c.setLineWidth(1.5)
        c.roundRect(x, y, card_width, card_height, 8, stroke=1, fill=0)
        
        # Bande supérieure
        c.setFillColor(colors.HexColor(primary_color))
        c.roundRect(x+2, y+card_height-30, card_width-4, 28, 6, stroke=0, fill=1)
        
        # Nom école
        c.setFillColor(colors.white)
        c.setFont(main_font_bold, 8)
        c.drawString(x+5, y+card_height-12, classe.ecole.nom[:40])
        c.setFont(main_font, 6)
        c.drawString(x+5, y+card_height-22, f"Année: {classe.annee_scolaire}")
        
        # Nom élève
        c.setFillColor(colors.HexColor('#1f2937'))
        c.setFont(main_font_bold, 10)
        c.drawString(x+5, y+card_height-42, f"{eleve.prenom} {eleve.nom}".upper()[:25])
        
        # Infos
        c.setFont(main_font, 7)
        c.setFillColor(colors.HexColor('#4b5563'))
        c.drawString(x+5, y+card_height-52, f"Mat: {eleve.matricule}")
        c.drawString(x+5, y+card_height-60, f"Classe: {classe.nom}")
        
        # Photo
        photo_size = 28
        photo_x = x + card_width - photo_size - 5
        photo_y = y + card_height/2 - photo_size/2
        
        c.setFillColor(colors.white)
        c.roundRect(photo_x, photo_y, photo_size, photo_size, 4, stroke=0, fill=1)
        c.setStrokeColor(colors.HexColor(primary_color))
        c.setLineWidth(1)
        c.roundRect(photo_x, photo_y, photo_size, photo_size, 4, stroke=1, fill=0)
        
        if eleve.photo:
            try:
                from PIL import Image
                if hasattr(eleve.photo, 'path') and os.path.exists(eleve.photo.path):
                    img = Image.open(eleve.photo.path)
                    if img.mode != 'RGB':
                        img = img.convert('RGB')
                    img.thumbnail((photo_size-2, photo_size-2), Image.Resampling.LANCZOS)
                    temp_buffer = io.BytesIO()
                    img.save(temp_buffer, format='JPEG')
                    temp_buffer.seek(0)
                    c.drawImage(temp_buffer, photo_x+1, photo_y+1, 
                              width=photo_size-2, height=photo_size-2, preserveAspectRatio=True)
            except:
                c.setFillColor(colors.HexColor(primary_color))
                c.setFont(main_font_bold, 12)
                c.drawCentredString(photo_x+photo_size/2, photo_y+photo_size/2-3, 
                                  f"{eleve.prenom[0]}{eleve.nom[0]}")
        else:
            c.setFillColor(colors.HexColor(primary_color))
            c.setFont(main_font_bold, 12)
            c.drawCentredString(photo_x+photo_size/2, photo_y+photo_size/2-3, 
                              f"{eleve.prenom[0]}{eleve.nom[0]}")
        
        card_count += 1
        if card_count % 2 == 0 and card_count < eleves.count():
            c.showPage()
    
    c.showPage()
    c.save()
    return response


@login_required
@require_http_methods(["POST"])
def ajax_modifier_telephone_responsable(request):
    """Modifie le numéro de téléphone du responsable principal d'un élève (AJAX)"""
    import json
    import re

    try:
        data = json.loads(request.body.decode('utf-8'))
        eleve_id = data.get('eleve_id')
        nouveau_telephone = data.get('telephone', '').strip()
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Données invalides'}, status=400)

    if not eleve_id or not nouveau_telephone:
        return JsonResponse({'success': False, 'error': 'Élève et téléphone requis'}, status=400)

    # Validation du format téléphone guinéen
    if not re.match(r'^\+224\d{8,9}$', nouveau_telephone):
        return JsonResponse({'success': False, 'error': 'Format invalide. Utilisez +224XXXXXXXXX'}, status=400)

    # Récupérer l'élève avec vérification des permissions
    try:
        eleve = Eleve.objects.select_related('responsable_principal', 'classe', 'classe__ecole').get(pk=eleve_id)
    except Eleve.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Élève introuvable'}, status=404)

    # Vérifier que l'utilisateur a accès à cet élève
    if not user_is_admin(request.user):
        user_school_obj = user_school(request.user)
        if not user_school_obj or eleve.classe.ecole != user_school_obj:
            return JsonResponse({'success': False, 'error': 'Accès non autorisé'}, status=403)

    if not eleve.responsable_principal:
        return JsonResponse({'success': False, 'error': 'Aucun responsable principal pour cet élève'}, status=400)

    ancien_telephone = eleve.responsable_principal.telephone
    eleve.responsable_principal.telephone = nouveau_telephone
    eleve.responsable_principal.save(update_fields=['telephone'])

    # Journaliser la modification
    try:
        JournalActivite.objects.create(
            user=request.user,
            action='MODIFICATION',
            type_objet='RESPONSABLE',
            objet_id=eleve.responsable_principal.id,
            description=f"Téléphone modifié pour {eleve.responsable_principal.nom_complet} "
                        f"(parent de {eleve.prenom} {eleve.nom}): {ancien_telephone} → {nouveau_telephone}",
            adresse_ip=request.META.get('REMOTE_ADDR', ''),
            user_agent=request.META.get('HTTP_USER_AGENT', '')[:200]
        )
    except Exception:
        pass

    return JsonResponse({
        'success': True,
        'telephone': nouveau_telephone,
        'message': f'Téléphone mis à jour: {nouveau_telephone}'
    })
