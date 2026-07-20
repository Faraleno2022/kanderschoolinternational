"""
Module Pointage & Présence : pointage journalier des élèves par classe,
rapport de présence avec alertes de taux d'absence élevé, et historique.
"""
from datetime import timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Count, Q
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone

from eleves.models import Classe, Eleve
from eleves.utils_annee import get_annee_active
from utilisateurs.utils import user_is_superadmin, user_school, filter_by_user_school

from .models import Presence

# Seuils d'alerte (proportion 0..1)
SEUIL_ABSENCE_CLASSE_JOUR = 0.20      # > 20 % d'absents dans une classe un jour
SEUIL_ABSENCE_ELEVE_PERIODE = 0.20    # > 20 % d'absences pour un élève sur la période

# Statuts comptés comme "absence réelle" dans les alertes
STATUTS_ABSENCE = [Presence.STATUT_ABSENT]


def _classes_accessibles(request):
    """Classes de l'école de l'utilisateur, limitées à l'année active."""
    qs = Classe.objects.select_related('ecole').all()
    if not user_is_superadmin(request.user):
        qs = filter_by_user_school(qs, request.user, 'ecole')
    ecole = user_school(request.user)
    annee = get_annee_active(request, ecole) if ecole else None
    if annee:
        qs = qs.filter(annee_scolaire=annee)
    return qs.order_by('niveau', 'nom')


def _eleves_actifs_classe(classe):
    return Eleve.objects.filter(classe=classe, statut='ACTIF').order_by('nom', 'prenom')


def _presences_accessibles(request):
    qs = Presence.objects.select_related('eleve', 'classe', 'classe__ecole', 'saisi_par')
    if not user_is_superadmin(request.user):
        qs = filter_by_user_school(qs, request.user, 'classe__ecole')
    return qs


def _parse_date(value, defaut=None):
    from datetime import datetime
    if value:
        try:
            return datetime.strptime(value, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            pass
    return defaut if defaut is not None else timezone.localdate()


@login_required
def presence_dashboard(request):
    """Tableau de bord du pointage : stats du jour, alertes, accès rapide au pointage."""
    classes = list(_classes_accessibles(request))
    aujourd_hui = _parse_date(request.GET.get('date'))

    presences_jour = _presences_accessibles(request).filter(date=aujourd_hui)

    # Stats globales du jour
    total_pointes = presences_jour.count()
    total_absents = presences_jour.filter(statut=Presence.STATUT_ABSENT).count()
    total_retards = presences_jour.filter(statut=Presence.STATUT_RETARD).count()
    total_excuses = presences_jour.filter(statut=Presence.STATUT_EXCUSE).count()
    total_presents = presences_jour.filter(statut=Presence.STATUT_PRESENT).count()

    # Détail + alertes par classe
    lignes_classes = []
    alertes = []
    for classe in classes:
        eff = _eleves_actifs_classe(classe).count()
        p_classe = presences_jour.filter(classe=classe)
        pointes = p_classe.count()
        absents = p_classe.filter(statut=Presence.STATUT_ABSENT).count()
        taux = (absents / pointes) if pointes else 0
        alerte = taux > SEUIL_ABSENCE_CLASSE_JOUR
        ligne = {
            'classe': classe,
            'effectif': eff,
            'pointes': pointes,
            'absents': absents,
            'retards': p_classe.filter(statut=Presence.STATUT_RETARD).count(),
            'taux_absence': round(taux * 100, 1),
            'fait': pointes > 0,
            'alerte': alerte,
        }
        lignes_classes.append(ligne)
        if alerte:
            alertes.append(ligne)

    context = {
        'titre_page': 'Pointage & Présence',
        'date_selection': aujourd_hui,
        'stats': {
            'total_pointes': total_pointes,
            'total_presents': total_presents,
            'total_absents': total_absents,
            'total_retards': total_retards,
            'total_excuses': total_excuses,
        },
        'lignes_classes': lignes_classes,
        'alertes': alertes,
        'seuil_pourcent': int(SEUIL_ABSENCE_CLASSE_JOUR * 100),
    }
    return render(request, 'vie_scolaire/presence/dashboard.html', context)


@login_required
def pointage_classe(request, classe_id):
    """Saisie/modification du pointage d'une classe pour une date donnée."""
    classe = get_object_or_404(_classes_accessibles(request), id=classe_id)
    date_pointage = _parse_date(request.POST.get('date') if request.method == 'POST' else request.GET.get('date'))
    eleves = list(_eleves_actifs_classe(classe))

    if request.method == 'POST':
        existantes = {p.eleve_id: p for p in Presence.objects.filter(classe=classe, date=date_pointage)}
        valides = dict(Presence.STATUT_CHOICES)
        with transaction.atomic():
            for eleve in eleves:
                statut = request.POST.get(f'statut_{eleve.id}', Presence.STATUT_PRESENT)
                if statut not in valides:
                    statut = Presence.STATUT_PRESENT
                motif = (request.POST.get(f'motif_{eleve.id}', '') or '').strip()[:200]
                presence = existantes.get(eleve.id)
                if presence:
                    presence.statut = statut
                    presence.motif = motif
                    presence.saisi_par = request.user
                    presence.save(update_fields=['statut', 'motif', 'saisi_par', 'date_saisie'])
                else:
                    Presence.objects.create(
                        eleve=eleve, classe=classe, date=date_pointage,
                        statut=statut, motif=motif, saisi_par=request.user,
                    )
        messages.success(
            request,
            f"Pointage enregistré pour {classe.nom} le {date_pointage:%d/%m/%Y} "
            f"({len(eleves)} élève(s))."
        )
        return redirect(f"{request.path}?date={date_pointage:%Y-%m-%d}")

    # GET : pré-remplir avec les statuts déjà saisis (PRÉSENT par défaut)
    existantes = {p.eleve_id: p for p in Presence.objects.filter(classe=classe, date=date_pointage)}
    lignes = []
    for eleve in eleves:
        p = existantes.get(eleve.id)
        lignes.append({
            'eleve': eleve,
            'statut': p.statut if p else Presence.STATUT_PRESENT,
            'motif': p.motif if p else '',
        })

    deja_fait = bool(existantes)
    context = {
        'titre_page': f"Pointage — {classe.nom}",
        'classe': classe,
        'date_pointage': date_pointage,
        'lignes': lignes,
        'statuts': Presence.STATUT_CHOICES,
        'deja_fait': deja_fait,
    }
    return render(request, 'vie_scolaire/presence/pointage.html', context)


def _calcul_rapport(request, classe, date_debut, date_fin):
    """Construit les lignes du rapport de présence par élève pour une classe/période."""
    presences = _presences_accessibles(request).filter(
        classe=classe, date__gte=date_debut, date__lte=date_fin
    )
    # Jours effectivement pointés dans la classe (base du taux)
    jours_pointes = presences.values_list('date', flat=True).distinct().count()

    par_eleve = {}
    for p in presences:
        d = par_eleve.setdefault(p.eleve_id, {
            'eleve': p.eleve, 'present': 0, 'absent': 0, 'retard': 0, 'excuse': 0, 'total': 0,
        })
        d['total'] += 1
        if p.statut == Presence.STATUT_PRESENT:
            d['present'] += 1
        elif p.statut == Presence.STATUT_ABSENT:
            d['absent'] += 1
        elif p.statut == Presence.STATUT_RETARD:
            d['retard'] += 1
        elif p.statut == Presence.STATUT_EXCUSE:
            d['excuse'] += 1

    lignes = []
    for d in par_eleve.values():
        base = d['total'] or 1
        taux_abs = d['absent'] / base
        d['taux_absence'] = round(taux_abs * 100, 1)
        d['taux_presence'] = round(d['present'] / base * 100, 1)
        d['a_risque'] = taux_abs > SEUIL_ABSENCE_ELEVE_PERIODE
        lignes.append(d)
    lignes.sort(key=lambda x: (-x['taux_absence'], x['eleve'].nom, x['eleve'].prenom))
    return {
        'lignes': lignes,
        'jours_pointes': jours_pointes,
        'nb_a_risque': sum(1 for l in lignes if l['a_risque']),
    }


@login_required
def rapport_presence(request):
    """Rapport de présence par classe et période, avec alertes d'absentéisme."""
    classes = list(_classes_accessibles(request))
    classe_id = request.GET.get('classe')
    aujourd_hui = timezone.localdate()
    date_debut = _parse_date(request.GET.get('date_debut'), aujourd_hui - timedelta(days=30))
    date_fin = _parse_date(request.GET.get('date_fin'), aujourd_hui)

    classe = None
    rapport = None
    if classe_id:
        classe = get_object_or_404(_classes_accessibles(request), id=classe_id)
        rapport = _calcul_rapport(request, classe, date_debut, date_fin)

    context = {
        'titre_page': 'Rapport de présence',
        'classes': classes,
        'classe_selection': classe,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'rapport': rapport,
        'seuil_pourcent': int(SEUIL_ABSENCE_ELEVE_PERIODE * 100),
    }
    return render(request, 'vie_scolaire/presence/rapport.html', context)


@login_required
def historique_presence_eleve(request, eleve_id):
    """Journal complet des présences d'un élève."""
    eleves = Eleve.objects.select_related('classe', 'classe__ecole')
    if not user_is_superadmin(request.user):
        eleves = filter_by_user_school(eleves, request.user, 'classe__ecole')
    eleve = get_object_or_404(eleves, id=eleve_id)

    presences = eleve.presences.select_related('classe', 'saisi_par').all()
    total = presences.count()
    absents = presences.filter(statut=Presence.STATUT_ABSENT).count()
    context = {
        'titre_page': f"Présences — {eleve.nom_complet}",
        'eleve': eleve,
        'presences': presences[:365],
        'total': total,
        'absents': absents,
        'taux_absence': round(absents / total * 100, 1) if total else 0,
    }
    return render(request, 'vie_scolaire/presence/historique_eleve.html', context)


@login_required
def export_rapport_presence_excel(request, classe_id):
    """Export Excel du rapport de présence d'une classe/période."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse

    classe = get_object_or_404(_classes_accessibles(request), id=classe_id)
    aujourd_hui = timezone.localdate()
    date_debut = _parse_date(request.GET.get('date_debut'), aujourd_hui - timedelta(days=30))
    date_fin = _parse_date(request.GET.get('date_fin'), aujourd_hui)
    rapport = _calcul_rapport(request, classe, date_debut, date_fin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Présence"

    ws['A1'] = f"Rapport de présence — {classe.nom}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Période : {date_debut:%d/%m/%Y} au {date_fin:%d/%m/%Y}"
    ws['A3'] = f"Jours pointés : {rapport['jours_pointes']}"

    entetes = ['Matricule', 'Nom', 'Prénom', 'Présences', 'Absences', 'Retards',
               'Excusées', 'Taux présence (%)', 'Taux absence (%)', 'Alerte']
    ligne_entete = 5
    header_fill = PatternFill(start_color='2E86C1', end_color='2E86C1', fill_type='solid')
    for col, titre in enumerate(entetes, start=1):
        c = ws.cell(row=ligne_entete, column=col, value=titre)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    alerte_fill = PatternFill(start_color='F5B7B1', end_color='F5B7B1', fill_type='solid')
    r = ligne_entete + 1
    for l in rapport['lignes']:
        e = l['eleve']
        valeurs = [
            e.matricule, e.nom, e.prenom, l['present'], l['absent'], l['retard'],
            l['excuse'], l['taux_presence'], l['taux_absence'],
            'OUI' if l['a_risque'] else '',
        ]
        for col, val in enumerate(valeurs, start=1):
            c = ws.cell(row=r, column=col, value=val)
            if l['a_risque']:
                c.fill = alerte_fill
        r += 1

    for col in range(1, len(entetes) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nom_fichier = f"presence_{classe.nom}_{date_debut:%Y%m%d}_{date_fin:%Y%m%d}.xlsx".replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{nom_fichier}"'
    wb.save(response)
    return response


@login_required
def export_rapport_presence_pdf(request, classe_id):
    """Export PDF du rapport de présence d'une classe/période."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from django.http import HttpResponse
    import io

    classe = get_object_or_404(_classes_accessibles(request), id=classe_id)
    aujourd_hui = timezone.localdate()
    date_debut = _parse_date(request.GET.get('date_debut'), aujourd_hui - timedelta(days=30))
    date_fin = _parse_date(request.GET.get('date_fin'), aujourd_hui)
    rapport = _calcul_rapport(request, classe, date_debut, date_fin)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"Rapport de présence — {classe.nom}", styles['Title']),
        Paragraph(
            f"Période : {date_debut:%d/%m/%Y} au {date_fin:%d/%m/%Y} — "
            f"Jours pointés : {rapport['jours_pointes']} — "
            f"Élèves à risque : {rapport['nb_a_risque']}",
            styles['Normal']
        ),
        Spacer(1, 0.5 * cm),
    ]

    data = [['Matricule', 'Nom & Prénom', 'Prés.', 'Abs.', 'Ret.', 'Exc.', '% Abs.']]
    for l in rapport['lignes']:
        e = l['eleve']
        data.append([
            e.matricule, f"{e.nom} {e.prenom}", l['present'], l['absent'],
            l['retard'], l['excuse'], f"{l['taux_absence']}%",
        ])

    table = Table(data, repeatRows=1)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86C1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ])
    # Surligner les élèves à risque
    for idx, l in enumerate(rapport['lignes'], start=1):
        if l['a_risque']:
            style.add('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#F5B7B1'))
    table.setStyle(style)
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    nom_fichier = f"presence_{classe.nom}_{date_debut:%Y%m%d}_{date_fin:%Y%m%d}.pdf".replace(' ', '_')
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nom_fichier}"'
    return response
